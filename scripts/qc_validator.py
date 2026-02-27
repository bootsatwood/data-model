#!/usr/bin/env python3
"""
QC Validator for Eventus Combined Database

Three modes:
  snapshot  — Capture full baseline before an edit session
  compare   — Diff current state against last snapshot
  validate  — Check against known-good expected values (PASS/FAIL)

Usage:
  python qc_validator.py snapshot
  python qc_validator.py compare
  python qc_validator.py validate
"""

import hashlib
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "qc_config.json"
SNAPSHOT_DIR = SCRIPT_DIR / ".snapshots"
VAULT_ROOT = Path.home() / "OneDrive - Eventus WholeHealth" / "Vault"
QC_REPORT_DIR = VAULT_ROOT / "new_02_Data_Model" / "QC"


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config():
    """Load qc_config.json and resolve file paths."""
    with CONFIG_PATH.open(encoding="utf-8") as f:
        cfg = json.load(f)
    cfg["_vault_dir"] = VAULT_ROOT / cfg["vault_data_model"]
    return cfg


# ---------------------------------------------------------------------------
# Database reader
# ---------------------------------------------------------------------------

def read_database(filepath):
    """Read an Excel file into a list of dicts (one per row).

    Uses openpyxl in read-only/data-only mode for speed.
    Returns (headers, rows) where rows is a list of dicts.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    row_iter = ws.iter_rows(values_only=True)
    headers = list(next(row_iter))

    rows = []
    for i, values in enumerate(row_iter, start=2):
        row = dict(zip(headers, values))
        row["_excel_row"] = i
        rows.append(row)

    wb.close()
    return headers, rows


# ---------------------------------------------------------------------------
# Dimension computation
# ---------------------------------------------------------------------------

def _safe(val):
    """Normalize a cell value for counting — treat None and '' as empty."""
    if val is None:
        return ""
    return str(val).strip()


def _has_barrier(val):
    """True if the Barrier field contains a real barrier value."""
    s = _safe(val)
    return s != "" and s != "None"


def _is_deactivated_barrier(val):
    """True if the Barrier field is a DUPLICATE-DEACTIVATED marker."""
    return _safe(val).startswith("DUPLICATE-DEACTIVATED")


def compute_row_key(row, key_fields):
    """Build a composite key string from key fields."""
    parts = [_safe(row.get(f, "")) for f in key_fields]
    return "|".join(parts)


def compute_row_fingerprint(row, check_fields):
    """Hash the check fields of a row for change detection."""
    parts = [_safe(row.get(f, "")) for f in check_fields]
    return hashlib.sha256("|".join(parts).encode()).hexdigest()[:16]


def compute_dimensions(rows, cfg):
    """Compute every countable dimension from the database rows.

    Returns a dict suitable for JSON serialization.
    """
    key_fields = cfg["database_columns"]["key_fields"]
    check_fields = cfg["database_columns"]["check_fields"]

    dims = {
        "totals": {},
        "per_state": {},
        "barriers_by_type": {},
        "geo_tiers": {},
        "contract_status": {},
        "metro_assignments": {},
        "fingerprints": {},
    }

    # Accumulators
    total = 0
    snf = alf = 0
    corporate = independent = 0
    served = 0
    barriers = 0
    deactivated_barriers = 0
    integrated = pcp_only = mh_only = 0

    state_data = {}   # state -> sub-dict of counts
    barrier_types = {}
    geo_tiers = {}
    contract_statuses = {}
    metros = {}

    for row in rows:
        total += 1

        st = _safe(row.get("State", ""))
        src = _safe(row.get("Source_Type", ""))
        own = _safe(row.get("Ownership_Type", ""))
        serve = _safe(row.get("Do_We_Serve", ""))
        integ = _safe(row.get("Integrated_Flag", ""))
        pcp = _safe(row.get("PCP_Flag", ""))
        mh = _safe(row.get("MH_Flag", ""))
        barrier = _safe(row.get("Barrier", ""))
        geo = _safe(row.get("Geographic_Tier", ""))
        contract = _safe(row.get("Contract_Status", ""))
        metro = _safe(row.get("Metro_Assignment", ""))

        # --- Totals ---
        if src == "SNF":
            snf += 1
        elif src == "ALF":
            alf += 1

        if own == "Corporate":
            corporate += 1
        elif own == "Independent":
            independent += 1

        if serve == "Yes":
            served += 1

        if _has_barrier(barrier):
            barriers += 1
            if _is_deactivated_barrier(barrier):
                deactivated_barriers += 1

        if integ == "Yes":
            integrated += 1
        elif pcp == "Yes" and mh != "Yes":
            pcp_only += 1
        elif mh == "Yes" and pcp != "Yes":
            mh_only += 1

        # --- Per-state ---
        if st:
            sd = state_data.setdefault(st, {
                "facilities": 0, "snf": 0, "alf": 0,
                "corporate": 0, "independent": 0,
                "served": 0, "barriers": 0,
                "integrated": 0, "pcp_only": 0, "mh_only": 0,
            })
            sd["facilities"] += 1
            if src == "SNF":
                sd["snf"] += 1
            elif src == "ALF":
                sd["alf"] += 1
            if own == "Corporate":
                sd["corporate"] += 1
            elif own == "Independent":
                sd["independent"] += 1
            if serve == "Yes":
                sd["served"] += 1
            if _has_barrier(barrier):
                sd["barriers"] += 1
            if integ == "Yes":
                sd["integrated"] += 1
            elif pcp == "Yes" and mh != "Yes":
                sd["pcp_only"] += 1
            elif mh == "Yes" and pcp != "Yes":
                sd["mh_only"] += 1

        # --- Barrier types ---
        if _has_barrier(barrier):
            barrier_types[barrier] = barrier_types.get(barrier, 0) + 1

        # --- Geo tiers ---
        if geo:
            geo_tiers[geo] = geo_tiers.get(geo, 0) + 1

        # --- Contract status ---
        if contract:
            contract_statuses[contract] = contract_statuses.get(contract, 0) + 1

        # --- Metro ---
        if metro:
            metros[metro] = metros.get(metro, 0) + 1

        # --- Row fingerprint ---
        key = compute_row_key(row, key_fields)
        fp = compute_row_fingerprint(row, check_fields)
        dims["fingerprints"][key] = {
            "fingerprint": fp,
            "excel_row": row["_excel_row"],
            "facility": _safe(row.get("Facility_Name", "")),
            "state": st,
        }

    dims["totals"] = {
        "total_facilities": total,
        "snf_count": snf,
        "alf_count": alf,
        "corporate_count": corporate,
        "independent_count": independent,
        "served_count": served,
        "barrier_count": barriers,
        "deactivated_barrier_count": deactivated_barriers,
        "integrated_count": integrated,
        "pcp_only_count": pcp_only,
        "mh_only_count": mh_only,
        "states_count": len(state_data),
        "metros_count": len(metros),
    }
    dims["per_state"] = dict(sorted(state_data.items()))
    dims["barriers_by_type"] = dict(sorted(barrier_types.items()))
    dims["geo_tiers"] = dict(sorted(geo_tiers.items()))
    dims["contract_status"] = dict(sorted(contract_statuses.items()))
    dims["metro_assignments"] = dict(sorted(metros.items()))

    return dims


# ---------------------------------------------------------------------------
# Markdown report writers
# ---------------------------------------------------------------------------

def write_validate_report(cfg, dims, results, now):
    """Write a human-readable markdown validation report to the Vault."""
    QC_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"QC_Validate_{cfg['version']}_{now.strftime('%Y%m%d')}.md"
    path = QC_REPORT_DIR / filename

    t = dims["totals"]
    lines = [
        f"# QC Validation Report — {cfg['version']}",
        "",
        f"**Date:** {now.strftime('%Y-%m-%d %H:%M')}",
        f"**Database:** `{cfg['files']['database']}`",
        f"**Result:** {results['pass']} PASS / {results['fail']} FAIL / {results['warn']} WARN",
        "",
        "---",
        "",
        "## Summary Counts",
        "",
        "| Metric | Actual | Expected | Status |",
        "|--------|--------|----------|--------|",
    ]

    expected = cfg["expected"]
    count_labels = [
        ("total_facilities", "Total Facilities"),
        ("snf_count", "SNF"),
        ("alf_count", "ALF"),
        ("corporate_count", "Corporate"),
        ("independent_count", "Independent"),
        ("served_count", "Served"),
        ("barrier_count", "Barriers"),
        ("integrated_count", "Integrated"),
        ("pcp_only_count", "PCP-only"),
        ("mh_only_count", "MH-only"),
        ("states_count", "States"),
    ]
    for key, label in count_labels:
        actual = t.get(key, 0)
        exp = expected.get(key, "—")
        status = "PASS" if actual == exp else "FAIL"
        lines.append(f"| {label} | {actual:,} | {exp:,} | {status} |")

    lines += [
        "",
        "## Geographic Tiers",
        "",
        "| Tier | Count |",
        "|------|-------|",
    ]
    for tier in sorted(dims["geo_tiers"]):
        lines.append(f"| {tier} | {dims['geo_tiers'][tier]:,} |")

    lines += [
        "",
        "## Contract Status",
        "",
        "| Status | Count |",
        "|--------|-------|",
    ]
    for cs in sorted(dims["contract_status"]):
        lines.append(f"| {cs} | {dims['contract_status'][cs]:,} |")

    lines += [
        "",
        "## Rule Checks",
        "",
        "| Check | Status | Detail |",
        "|-------|--------|--------|",
    ]
    for d in results["details"]:
        if d["category"] == "rules" or d["category"] == "cross-file":
            lines.append(f"| {d['category']} | {d['status']} | {d['message']} |")

    if t.get("deactivated_barrier_count", 0) > 0:
        lines += [
            "",
            f"> [!warning] {t['deactivated_barrier_count']} DUPLICATE-DEACTIVATED barrier markers present — consider cleanup",
        ]

    lines += [
        "",
        "---",
        "",
        f"*Generated by `qc_validator.py` on {now.strftime('%Y-%m-%d %H:%M')}*",
        "",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Report saved to: {path}")
    return path


def write_compare_report(cfg, snap, diff_report, current, now):
    """Write a human-readable markdown comparison report to the Vault."""
    QC_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"QC_Compare_{cfg['version']}_{now.strftime('%Y%m%d_%H%M%S')}.md"
    path = QC_REPORT_DIR / filename

    has_changes = (
        diff_report["summary_deltas"]
        or diff_report["state_changes"]
        or diff_report["row_changes"]
        or diff_report["new_rows"]
        or diff_report["removed_rows"]
    )

    t = current["totals"]
    lines = [
        f"# QC Comparison Report — {cfg['version']}",
        "",
        f"**Date:** {now.strftime('%Y-%m-%d %H:%M')}",
        f"**Snapshot baseline:** {snap['timestamp']}",
        f"**Database:** `{cfg['files']['database']}`",
        f"**Result:** {'CHANGES DETECTED' if has_changes else 'NO CHANGES — identical to snapshot'}",
        "",
        "---",
        "",
    ]

    if not has_changes:
        lines += [
            "Database is identical to the snapshot baseline. No drift detected.",
            "",
        ]
    else:
        # Summary deltas
        if diff_report["summary_deltas"]:
            lines += [
                "## Summary Deltas",
                "",
                "| Metric | Before | After | Delta |",
                "|--------|--------|-------|-------|",
            ]
            for key, vals in diff_report["summary_deltas"].items():
                sign = "+" if vals["delta"] > 0 else ""
                lines.append(f"| {key} | {vals['old']:,} | {vals['new']:,} | {sign}{vals['delta']:,} |")
            lines.append("")

        # State changes
        if diff_report["state_changes"]:
            lines += [
                "## State Changes",
                "",
                "| State | Metric | Before | After | Delta |",
                "|-------|--------|--------|-------|-------|",
            ]
            for st in sorted(diff_report["state_changes"]):
                for metric, vals in diff_report["state_changes"][st].items():
                    sign = "+" if vals["delta"] > 0 else ""
                    lines.append(f"| {st} | {metric} | {vals['old']} | {vals['new']} | {sign}{vals['delta']} |")
            lines.append("")

        # Row changes
        if diff_report["row_changes"]:
            lines += [
                f"## Row Changes ({len(diff_report['row_changes'])} rows)",
                "",
                "| Row | Facility | State |",
                "|-----|----------|-------|",
            ]
            for cr in sorted(diff_report["row_changes"], key=lambda x: x["excel_row"]):
                lines.append(f"| {cr['excel_row']} | {cr['facility']} | {cr['state']} |")
            lines.append("")

        # New rows
        if diff_report["new_rows"]:
            lines += [
                f"## New Rows ({len(diff_report['new_rows'])})",
                "",
                "| Row | Facility | State |",
                "|-----|----------|-------|",
            ]
            for nr in diff_report["new_rows"]:
                lines.append(f"| {nr['excel_row']} | {nr['facility']} | {nr['state']} |")
            lines.append("")

        # Removed rows
        if diff_report["removed_rows"]:
            lines += [
                f"## Removed Rows ({len(diff_report['removed_rows'])})",
                "",
                "| Was Row | Facility | State |",
                "|---------|----------|-------|",
            ]
            for rr in diff_report["removed_rows"]:
                lines.append(f"| {rr['excel_row']} | {rr['facility']} | {rr['state']} |")
            lines.append("")

    lines += [
        "---",
        "",
        f"*Generated by `qc_validator.py` on {now.strftime('%Y-%m-%d %H:%M')}*",
        "",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Report saved to: {path}")
    return path


# ---------------------------------------------------------------------------
# Mode 1: Snapshot
# ---------------------------------------------------------------------------

def cmd_snapshot(cfg):
    """Capture a full baseline snapshot of the database."""
    db_path = cfg["_vault_dir"] / cfg["files"]["database"]
    print(f"Reading {cfg['files']['database']}...")
    headers, rows = read_database(db_path)
    print(f"  {len(rows):,} rows loaded.")

    dims = compute_dimensions(rows, cfg)

    # Build snapshot payload
    now = datetime.now()
    snapshot = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "version": cfg["version"],
        "database_file": cfg["files"]["database"],
        "headers": headers,
        "dimensions": dims,
    }

    # Save timestamped + latest
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    ts_name = f"snapshot_{now.strftime('%Y%m%d_%H%M%S')}.json"
    ts_path = SNAPSHOT_DIR / ts_name
    latest_path = SNAPSHOT_DIR / "latest.json"

    with ts_path.open("w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)
    shutil.copy2(ts_path, latest_path)

    # Print summary
    t = dims["totals"]
    print()
    print(f"SNAPSHOT taken: {snapshot['timestamp']}")
    print(f"  Database: {cfg['files']['database']}")
    print(f"  Version:  {cfg['version']}")
    print(f"  Total facilities: {t['total_facilities']:,}")
    print(f"  SNF: {t['snf_count']:,}  |  ALF: {t['alf_count']:,}")
    print(f"  Corporate: {t['corporate_count']:,}  |  Independent: {t['independent_count']:,}")
    print(f"  Served: {t['served_count']:,}  |  Barriers: {t['barrier_count']:,}")
    if t["deactivated_barrier_count"]:
        print(f"    ({t['deactivated_barrier_count']} are DUPLICATE-DEACTIVATED markers)")
    print(f"  Integrated: {t['integrated_count']:,}  |  PCP-only: {t['pcp_only_count']:,}  |  MH-only: {t['mh_only_count']:,}")
    print(f"  States: {t['states_count']}  |  Metros: {t['metros_count']}")
    print(f"  Row fingerprints: {len(dims['fingerprints']):,} captured")
    print(f"  Saved to: {ts_path.relative_to(SCRIPT_DIR)}")


# ---------------------------------------------------------------------------
# Mode 2: Compare
# ---------------------------------------------------------------------------

def cmd_compare(cfg):
    """Compare current database state against the latest snapshot."""
    latest_path = SNAPSHOT_DIR / "latest.json"
    if not latest_path.exists():
        print("ERROR: No snapshot found. Run 'snapshot' first.")
        sys.exit(1)

    # Load snapshot
    with latest_path.open(encoding="utf-8") as f:
        snap = json.load(f)
    print(f"Loaded snapshot from {snap['timestamp']}")

    # Read current database
    db_path = cfg["_vault_dir"] / cfg["files"]["database"]
    print(f"Reading {cfg['files']['database']}...")
    headers, rows = read_database(db_path)
    print(f"  {len(rows):,} rows loaded.")

    current = compute_dimensions(rows, cfg)
    prev = snap["dimensions"]

    now = datetime.now()
    diff_report = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "snapshot_timestamp": snap["timestamp"],
        "summary_deltas": {},
        "state_changes": {},
        "row_changes": [],
        "new_rows": [],
        "removed_rows": [],
    }

    # --- Summary deltas ---
    print()
    print(f"COMPARE against snapshot from {snap['timestamp']}")
    print("=" * 60)
    print("SUMMARY")
    print("-" * 60)

    any_total_change = False
    for key in current["totals"]:
        cur_val = current["totals"][key]
        prev_val = prev["totals"].get(key, 0)
        if cur_val != prev_val:
            delta = cur_val - prev_val
            sign = "+" if delta > 0 else ""
            print(f"  {key}: {prev_val:,} -> {cur_val:,} ({sign}{delta:,})")
            diff_report["summary_deltas"][key] = {"old": prev_val, "new": cur_val, "delta": delta}
            any_total_change = True
    if not any_total_change:
        print("  No changes in totals.")

    # --- Per-state deltas ---
    all_states = sorted(set(list(current["per_state"].keys()) + list(prev["per_state"].keys())))
    state_changes = {}
    for st in all_states:
        cur_st = current["per_state"].get(st, {})
        prev_st = prev["per_state"].get(st, {})
        changes = {}
        for metric in set(list(cur_st.keys()) + list(prev_st.keys())):
            c = cur_st.get(metric, 0)
            p = prev_st.get(metric, 0)
            if c != p:
                changes[metric] = {"old": p, "new": c, "delta": c - p}
        if changes:
            state_changes[st] = changes

    if state_changes:
        print()
        print("STATE CHANGES")
        print("-" * 60)
        for st in sorted(state_changes.keys()):
            parts = []
            for metric, vals in state_changes[st].items():
                sign = "+" if vals["delta"] > 0 else ""
                parts.append(f"{metric} {vals['old']}->{vals['new']} ({sign}{vals['delta']})")
            print(f"  {st}: {', '.join(parts)}")
    diff_report["state_changes"] = state_changes

    # --- Row-level changes ---
    cur_fps = current["fingerprints"]
    prev_fps = prev["fingerprints"]

    cur_keys = set(cur_fps.keys())
    prev_keys = set(prev_fps.keys())

    new_keys = cur_keys - prev_keys
    removed_keys = prev_keys - cur_keys
    common_keys = cur_keys & prev_keys

    changed_rows = []
    for key in common_keys:
        if cur_fps[key]["fingerprint"] != prev_fps[key]["fingerprint"]:
            changed_rows.append({
                "key": key,
                "excel_row": cur_fps[key]["excel_row"],
                "facility": cur_fps[key]["facility"],
                "state": cur_fps[key]["state"],
            })

    # Enrich changed rows with field-level diffs
    if changed_rows:
        check_fields = cfg["database_columns"]["check_fields"]
        key_fields = cfg["database_columns"]["key_fields"]

        # Re-read to get actual values for changed rows (snapshot doesn't store raw values)
        # Build a lookup of current rows by key
        current_by_key = {}
        for row in rows:
            k = compute_row_key(row, key_fields)
            current_by_key[k] = row

        print()
        print(f"ROW CHANGES ({len(changed_rows)} rows)")
        print("-" * 60)
        for cr in sorted(changed_rows, key=lambda x: x["excel_row"]):
            print(f"  Row {cr['excel_row']}: {cr['facility']} ({cr['state']})")
            # We can't show old values without storing them in snapshot — show current values
            print(f"    (Fingerprint changed — re-run snapshot to capture new baseline)")
            diff_report["row_changes"].append(cr)
    else:
        if any_total_change:
            print()
            print("ROW CHANGES")
            print("-" * 60)
            print("  No individual row fingerprints changed (change may be in row count).")

    # New / removed rows
    if new_keys:
        print()
        print(f"NEW ROWS ({len(new_keys)})")
        print("-" * 60)
        for key in sorted(new_keys):
            info = cur_fps[key]
            print(f"  Row {info['excel_row']}: {info['facility']} ({info['state']})")
            diff_report["new_rows"].append({"key": key, **info})

    if removed_keys:
        print()
        print(f"REMOVED ROWS ({len(removed_keys)})")
        print("-" * 60)
        for key in sorted(removed_keys):
            info = prev_fps[key]
            print(f"  Was row {info['excel_row']}: {info['facility']} ({info['state']})")
            diff_report["removed_rows"].append({"key": key, **info})

    # No changes at all
    if not any_total_change and not state_changes and not changed_rows and not new_keys and not removed_keys:
        print()
        print("  No differences detected. Database is identical to snapshot.")

    # Save diff report (JSON to local snapshots, markdown to Vault)
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    diff_path = SNAPSHOT_DIR / f"compare_{now.strftime('%Y%m%d_%H%M%S')}.json"
    with diff_path.open("w", encoding="utf-8") as f:
        json.dump(diff_report, f, indent=2, ensure_ascii=False)
    print()
    print(f"Diff saved to: {diff_path.relative_to(SCRIPT_DIR)}")
    write_compare_report(cfg, snap, diff_report, current, now)


# ---------------------------------------------------------------------------
# Mode 3: Validate
# ---------------------------------------------------------------------------

def cmd_validate(cfg):
    """Validate database against expected values and business rules."""
    db_path = cfg["_vault_dir"] / cfg["files"]["database"]
    print(f"Reading {cfg['files']['database']}...")
    headers, rows = read_database(db_path)
    print(f"  {len(rows):,} rows loaded.")

    dims = compute_dimensions(rows, cfg)
    expected = cfg["expected"]
    expected_cols = cfg["database_columns"]["expected_columns"]

    results = {"pass": 0, "fail": 0, "warn": 0, "details": []}

    def record(status, category, message):
        tag = {"pass": "PASS", "fail": "FAIL", "warn": "WARN"}[status]
        results[status] += 1
        results["details"].append({"status": tag, "category": category, "message": message})
        print(f"  [{tag}] {message}")

    # --- Structure checks ---
    print()
    print("=" * 60)
    print("STRUCTURE")
    print("-" * 60)

    if headers == expected_cols:
        record("pass", "structure", f"{len(headers)} columns present, all match expected")
    else:
        missing = [c for c in expected_cols if c not in headers]
        extra = [c for c in headers if c not in expected_cols]
        if missing:
            record("fail", "structure", f"Missing columns: {missing}")
        if extra:
            record("warn", "structure", f"Unexpected columns: {extra}")
        if not missing and not extra:
            record("warn", "structure", f"Columns present but in different order")

    # --- Count checks ---
    print()
    print("COUNTS")
    print("-" * 60)

    count_checks = [
        ("total_facilities", "Total facilities"),
        ("snf_count", "SNF"),
        ("alf_count", "ALF"),
        ("corporate_count", "Corporate"),
        ("independent_count", "Independent"),
        ("served_count", "Served"),
        ("barrier_count", "Barriers"),
        ("integrated_count", "Integrated"),
        ("pcp_only_count", "PCP-only"),
        ("mh_only_count", "MH-only"),
        ("states_count", "States"),
    ]

    for key, label in count_checks:
        if key in expected:
            actual = dims["totals"].get(key, 0)
            exp = expected[key]
            if actual == exp:
                record("pass", "counts", f"{label}: {actual:,} (expected {exp:,})")
            else:
                delta = actual - exp
                sign = "+" if delta > 0 else ""
                record("fail", "counts", f"{label}: {actual:,} (expected {exp:,}, {sign}{delta:,})")

    # Geo tier checks
    if "geo_tiers" in expected:
        for tier, exp_count in expected["geo_tiers"].items():
            actual = dims["geo_tiers"].get(tier, 0)
            if actual == exp_count:
                record("pass", "counts", f"Geo tier {tier}: {actual:,}")
            else:
                record("fail", "counts", f"Geo tier {tier}: {actual:,} (expected {exp_count:,})")

    # Contract status checks
    if "contract_status" in expected:
        for status, exp_count in expected["contract_status"].items():
            actual = dims["contract_status"].get(status, 0)
            if actual == exp_count:
                record("pass", "counts", f"Contract {status}: {actual:,}")
            else:
                record("fail", "counts", f"Contract {status}: {actual:,} (expected {exp_count:,})")

    # Deactivated barriers
    deact = dims["totals"].get("deactivated_barrier_count", 0)
    if deact > 0:
        record("warn", "counts", f"{deact} DUPLICATE-DEACTIVATED barrier markers present — consider cleanup")

    # --- Rule checks ---
    print()
    print("RULES")
    print("-" * 60)

    # Rule: INDEPENDENT corporate name → Ownership_Type must be Independent
    indep_violations = []
    for row in rows:
        if _safe(row.get("Corporate_Name", "")) == "INDEPENDENT":
            if _safe(row.get("Ownership_Type", "")) != "Independent":
                indep_violations.append(row["_excel_row"])
    indep_total = sum(1 for r in rows if _safe(r.get("Corporate_Name", "")) == "INDEPENDENT")
    if not indep_violations:
        record("pass", "rules", f"INDEPENDENT ownership consistency: {indep_total} checked, 0 violations")
    else:
        record("fail", "rules", f"INDEPENDENT ownership violations at rows: {indep_violations[:10]}")

    # Rule: Do_We_Serve=Yes → at least one of PCP or MH = Yes
    serve_violations = []
    for row in rows:
        if _safe(row.get("Do_We_Serve", "")) == "Yes":
            pcp = _safe(row.get("PCP_Flag", ""))
            mh = _safe(row.get("MH_Flag", ""))
            if pcp != "Yes" and mh != "Yes":
                serve_violations.append(row["_excel_row"])
    served_total = dims["totals"]["served_count"]
    if not serve_violations:
        record("pass", "rules", f"Service flag consistency: {served_total:,} served, all have PCP or MH")
    else:
        record("fail", "rules", f"Served but no PCP/MH flag at rows: {serve_violations[:10]}")

    # Rule: Integrated_Flag=Yes ↔ both PCP=Yes AND MH=Yes
    integ_violations = []
    for row in rows:
        integ = _safe(row.get("Integrated_Flag", ""))
        pcp = _safe(row.get("PCP_Flag", ""))
        mh = _safe(row.get("MH_Flag", ""))
        if integ == "Yes" and (pcp != "Yes" or mh != "Yes"):
            integ_violations.append(("has_flag_missing_service", row["_excel_row"]))
        elif integ != "Yes" and pcp == "Yes" and mh == "Yes":
            integ_violations.append(("missing_flag_has_services", row["_excel_row"]))
    if not integ_violations:
        record("pass", "rules", f"Integrated flag consistency: {dims['totals']['integrated_count']:,} checked, 0 violations")
    else:
        record("fail", "rules", f"Integrated flag violations ({len(integ_violations)}): first 10 = {integ_violations[:10]}")

    # --- Cross-file checks ---
    print()
    print("CROSS-FILE")
    print("-" * 60)

    scenario_path = cfg["_vault_dir"] / cfg["files"]["scenario_1"]
    if scenario_path.exists():
        s_wb = load_workbook(scenario_path, read_only=True, data_only=True)
        s_ws = s_wb[s_wb.sheetnames[0]]
        s_row_count = sum(1 for _ in s_ws.iter_rows(min_row=2))
        s_wb.close()

        db_count = dims["totals"]["total_facilities"]
        exp_scenario = expected.get("scenario_row_count", 0)

        if s_row_count == db_count:
            record("pass", "cross-file", f"Scenario rows ({s_row_count:,}) match database ({db_count:,})")
        elif s_row_count == exp_scenario:
            record("warn", "cross-file",
                   f"Scenarios at V20.0 ({s_row_count:,} rows) vs database ({db_count:,} rows) — rebuild needed")
        else:
            record("warn", "cross-file",
                   f"Scenario rows ({s_row_count:,}) don't match database ({db_count:,}) or expected ({exp_scenario:,})")
    else:
        record("warn", "cross-file", f"Scenario file not found: {cfg['files']['scenario_1']}")

    # --- Summary ---
    print()
    print("=" * 60)
    total_checks = results["pass"] + results["fail"] + results["warn"]
    print(f"RESULT: {results['fail']} FAIL, {results['warn']} WARN, {results['pass']} PASS ({total_checks} checks)")
    print("=" * 60)

    if results["fail"] > 0:
        print("\nAction required: Fix FAIL items before delivery.")
    elif results["warn"] > 0:
        print("\nNo failures. Review WARN items and document any accepted variances.")
    else:
        print("\nAll checks passed.")

    # Save results (JSON to local snapshots, markdown to Vault)
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    result_path = SNAPSHOT_DIR / f"validate_{now.strftime('%Y%m%d_%H%M%S')}.json"
    with result_path.open("w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nResults saved to: {result_path.relative_to(SCRIPT_DIR)}")
    write_validate_report(cfg, dims, results, now)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ("snapshot", "compare", "validate"):
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]
    cfg = load_config()

    print(f"QC Validator — {cfg['version']}")
    print(f"Vault: {cfg['_vault_dir']}")
    print()

    if mode == "snapshot":
        cmd_snapshot(cfg)
    elif mode == "compare":
        cmd_compare(cfg)
    elif mode == "validate":
        cmd_validate(cfg)


if __name__ == "__main__":
    main()
