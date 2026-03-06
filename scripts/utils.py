"""
Shared utilities for Eventus database scripts.

Extracted from chain_update_v21_2.py for reuse across audit and fix scripts.
"""

import re
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

VAULT = Path.home() / "OneDrive - Eventus WholeHealth" / "Vault" / "02_Data_Model"
DB_CURRENT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_13.xlsx"
CMS_SNF_FILE = VAULT / "Reference" / "Source_SNF_Database_with_Revenue_Q4_2025.xlsx"
CMS_PROVIDER_FILE = VAULT / "Reference" / "Source_CMS_NH_ProviderInfo_Feb2026.csv"
NIC_ALF_FILE = VAULT / "Reference" / "Source_NIC_Maps_Nationwide_Inventory_Export_11.17.25.xlsx"
GLR_FILE = VAULT / "Reference" / "Source_GLR_Export_MUO_Data_2026_03.xlsx"

FOOTPRINT = {'IN', 'KY', 'NC', 'OH', 'SC', 'VA', 'MI', 'IL', 'WI', 'MN', 'FL', 'MD', 'GA', 'MO'}

REPORT_DIR = Path(__file__).resolve().parent / "audit_reports"


# ---------------------------------------------------------------------------
# Normalizers
# ---------------------------------------------------------------------------

def norm(s):
    """Normalize a string to lowercase alphanumeric only."""
    if not s:
        return ''
    return re.sub(r'[^a-z0-9]', '', s.lower())


def norm_addr(s):
    """Normalize a street address: abbreviate common words, strip punctuation."""
    if not s:
        return ''
    s = s.lower().strip()
    for word, abbr in [('street', 'st'), ('road', 'rd'), ('drive', 'dr'),
                        ('avenue', 'ave'), ('boulevard', 'blvd'), ('lane', 'ln'),
                        ('court', 'ct'), ('north', 'n'), ('south', 's'),
                        ('east', 'e'), ('west', 'w')]:
        s = re.sub(r'\b' + word + r'\b', abbr, s)
    return re.sub(r'[^a-z0-9]', '', s)


def addr_key(address, city, state):
    """Build a composite address key: norm_addr|norm_city|norm_state."""
    return norm_addr(address) + '|' + norm(city) + '|' + norm(state)


def safe(val):
    """Coerce a cell value to a stripped string (None -> '')."""
    if val is None:
        return ''
    return str(val).strip()


# ---------------------------------------------------------------------------
# Database loader
# ---------------------------------------------------------------------------

def load_db(filepath=None):
    """Load the Combined Database into a list of dicts.

    Returns (headers, rows) where each row is a dict with an '_excel_row' key.
    Uses read-only mode for speed.
    """
    filepath = filepath or DB_CURRENT
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    headers = [safe(c) for c in next(row_iter)]

    rows = []
    for i, values in enumerate(row_iter, start=2):
        row = dict(zip(headers, values))
        row['_excel_row'] = i
        rows.append(row)

    wb.close()
    return headers, rows


def load_cms_snf(filepath=None):
    """Load the CMS SNF source file into a list of dicts.

    Returns (headers, rows). Uses read-only mode.
    """
    filepath = filepath or CMS_SNF_FILE
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    headers = [safe(c) for c in next(row_iter)]

    rows = []
    for i, values in enumerate(row_iter, start=2):
        row = dict(zip(headers, values))
        row['_excel_row'] = i
        rows.append(row)

    wb.close()
    return headers, rows


def load_nic_alf(filepath=None):
    """Load the NIC Maps ALF source file into a list of dicts.

    Returns (headers, rows). Uses read-only mode.
    """
    filepath = filepath or NIC_ALF_FILE
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    headers = [safe(c) for c in next(row_iter)]

    rows = []
    for i, values in enumerate(row_iter, start=2):
        row = dict(zip(headers, values))
        row['_excel_row'] = i
        rows.append(row)

    wb.close()
    return headers, rows


# ---------------------------------------------------------------------------
# Report helper
# ---------------------------------------------------------------------------

def ensure_report_dir():
    """Create the audit_reports directory if it doesn't exist."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    return REPORT_DIR
