"""Rescore the full MUO universe from V20 to V23.

- Recomputes Enterprise Reach and Revenue Potential from V23 DB
- Scopes to 6-state operational footprint (NC, SC, VA, KY, OH, IN)
- Uses campus-collapsed (unique address) counts for ER
- Applies >=7 campus gate: under 7 = T4 Independent
- Keeps IR, SI, RS, AI from V20 (qualitative, unchanged)
- Adds 9 red-flagged entities not in V20
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import Counter, defaultdict

# --- Config ---
DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
TIERING_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V20.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

# --- V20 name -> V23 DB corporate name(s) mapping ---
# Keys = V20 workbook name (as-is), Values = set of V23 DB Corporate_Name values (UPPER)
NAME_MAP = {
    'ALG': {'ALG SENIOR'},
    'American Senior Communities': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'Cch Healthcare': {'CCH HEALTHCARE'},
    'Ciena Healthcare/Laurel Health Care': {'CIENA HEALTHCARE', 'LAUREL HEALTH CARE'},
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': {
        'CONSULATE HEALTH CARE', 'INDEPENDENCE LIVING CENTERS', 'NSPIRE HEALTHCARE', 'RAYDIANT HEALTH CARE'
    },
    'Infinity Healthcare Consulting': {'INFINITY HEALTHCARE CONSULTING'},
    'Liberty Senior Living': {'LIBERTY'},
    'Life Care Centers Of America': {'LIFE CARE CENTERS OF AMERICA'},
    'Majestic Care': {'MAJESTIC CARE'},
    'National Healthcare Corporation': {'NATIONAL HEALTHCARE CORPORATION', 'NHC'},
    'Navion': {'NAVION'},
    'Otterbein Seniorlife': {'OTTERBEIN', 'OTTERBEIN SENIORLIFE'},
    'PACS Group': {'PACS GROUP'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'Principle Long Term Care': {'PRINCIPLE LONG TERM CARE', 'PRINCIPLE'},
    'Pruitthealth': {'PRUITTCHEALTH', 'PRUITTHEALTH'},
    'Saber Healthcare Group': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'Sunrise': {'SUNRISE SENIOR LIVING', 'SUNRISE'},
    'Terra Bella': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'Tlc Management': {'TLC MANAGEMENT'},
    'Trilogy Health Services': {'TRILOGY HEALTH SERVICES', 'TRILOGY'},
    'Altercare': {'ALTERCARE'},
    'Aom Healthcare': {'AOM HEALTHCARE'},
    'Aperion Care': {'APERION CARE'},
    'Arbors At Ohio': {'ARBORS AT OHIO'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Brickyard Healthcare': {'BRICKYARD HEALTHCARE'},
    'Carespring': {'CARESPRING'},
    'Carrolton': {'CARROLTON FACILTY MANAGEMENT', 'CARROLTON NURSING HOMES', 'CARROLTON'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'Certus Healthcare': {'CERTUS HEALTHCARE'},
    'Clearview': {'CLEARVIEW'},
    'Gardant Management Solutions, Inc': {'GARDANT MANAGEMENT SOLUTIONS'},
    'HCF Management': {'HCF MANAGEMENT', 'HCF'},
    'Health Care Management Group': {'HEALTH CARE MANAGEMENT GROUP'},
    'Heritage Hall': {'HERITAGE HALL'},
    'Jag Healthcare': {'JAG HEALTHCARE'},
    'Kissito': {'KISSITO HEALTHCARE', 'KISSITO'},
    'Legacy Health Services': {'LEGACY HEALTH SERVICES'},
    'Lionstone Care': {'LIONSTONE CARE', 'LIONSTONE'},
    'Lutheran Services Carolina': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Momentous Health': {'MOMENTUS HEALTH', 'MOMENTOUS HEALTH'},
    'Morning Pointe Senior Living': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'Ohio Living Communities': {'OHIO LIVING COMMUNITIES', 'OHIO LIVING'},
    'Peak Resources, Inc.': {'PEAK RESOURCES'},
    'Phoenix Senior Living': {'PHOENIX SENIOR LIVING'},
    'Priority Life Care': {'PRIORITY LIFE CARE'},
    'Progressive Quality Care': {'PROGRESSIVE QUALITY CARE'},
    'Sanstone Health & Rehabilitation': {'SANSTONE HEALTH & REHABILITATION', 'SANSTONE'},
    'Seky Holding Co.': {'SEKY HOLDING CO.', 'SEKY HOLDING'},
    'Southern Assisted Living, LLC': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},  # Consolidated V22.10
    'Sovereign Healthcare Holdings': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},  # Same entity, consolidated V22.10
    'Sprenger Health Care Systems': {'SPRENGER HEALTH CARE SYSTEMS', 'SPRENGER'},
    'Topaz': {'TOPAZ HEALTHCARE', 'TOPAZ'},
    'White Oak Management': {'WHITE OAK MANAGEMENT'},
    'Windsor House, Inc.': {'WINDSOR HOUSE'},
    'Yad Healthcare': {'YAD HEALTHCARE'},
    'Hillstone Healthcare': {'HILLSTONE HEALTHCARE'},
    'Carecore Health': {'CARECORE HEALTH'},
    'Trio Healthcare': {'TRIO HEALTHCARE'},
    'Continuing Healthcare Solutions': {'CONTINUING HEALTHCARE SOLUTIONS'},
    "Miller's Merry Manor": {"MILLER'S MERRY MANOR", 'MILLERS MERRY MANOR'},
    'Vancrest Health Care Centers': {'VANCREST HEALTH CARE CENTERS', 'VANCREST'},
    'Optalis Health & Rehabilitation': {'OPTALIS HEALTH & REHABILITATION', 'OPTALIS'},
    'Crown Healthcare Group': {'CROWN HEALTHCARE GROUP'},
    'Divine Healthcare Management': {'DIVINE HEALTHCARE MANAGEMENT'},
    'Envive Healthcare': {'ENVIVE HEALTHCARE'},
    'Choice Health Management': {'CHOICE HEALTH MANAGEMENT'},
    'Caring Place Healthcare': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
}

# 9 red-flagged entities NOT in V20 (use default qualitative scores)
RED_FLAGGED = {
    'American Healthcare LLC': {'corp_names': {'AMERICAN HEALTHCARE, LLC'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'MCAP': {'corp_names': {'MCAP'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Greencroft': {'corp_names': {'GREENCROFT'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'LifeSpire of Virginia': {'corp_names': {'LIFESPIRE OF VIRGINIA'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Senior Lifestyle': {'corp_names': {'SENIOR LIFESTYLE'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Cedarhurst Senior Living': {'corp_names': {'CEDARHURST SENIOR LIVING'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Spring Arbor Management': {'corp_names': {'SPRING ARBOR MANAGEMENT'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'StoryPoint': {'corp_names': {'STORYPOINT'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Sonida Senior Living': {'corp_names': {'SONIDA SENIOR LIVING'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
}

# Also add entities from memory that were scored separately
ADDITIONAL = {
    'Triple Crown Senior Living': {'corp_names': {'TRIPLE CROWN SENIOR LIVING', 'CROWN SENIOR LIVING'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
    'Lutheran Life Villages': {'corp_names': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'}, 'ir': 2, 'si': 2, 'rs': 2, 'ai': 0},
}


# --- Scoring functions ---
def er_score(campuses):
    if campuses >= 40: return 5
    if campuses >= 20: return 4
    if campuses >= 10: return 3
    if campuses >= 5: return 2
    return 1

def rp_score(rev):
    if rev >= 2_000_000: return 5
    if rev >= 1_000_000: return 4
    if rev >= 500_000: return 3
    if rev >= 250_000: return 2
    return 1

def rs_score(served, total):
    if total == 0: return 1
    pct = served / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if served > 0: return 2
    return 1

def tier_label(score):
    if score >= 50: return 'T1'
    if score >= 25: return 'T2'
    return 'T3'

def safe_num(v):
    return v if isinstance(v, (int, float)) else 0

def bracket_er(n):
    if n >= 40: return '40+'
    if n >= 20: return '20-39'
    if n >= 10: return '10-19'
    if n >= 5: return '5-9'
    return '1-4'

def bracket_rp(rev):
    if rev >= 2_000_000: return '>$2M'
    if rev >= 1_000_000: return '$1.0-1.99M'
    if rev >= 500_000: return '$500K-$999K'
    if rev >= 250_000: return '$250-499K'
    return '<$250K'


# --- Load V20 scores ---
wb_t = openpyxl.load_workbook(TIERING_PATH, data_only=True)
v20_entities = {}
# Track Southern/Sovereign as same entity
southern_scores = None

for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_t[sn]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        scores = {
            'v20_tier': ws.cell(r, 2).value,
            'v20_score': ws.cell(r, 3).value,
            'v20_facs': ws.cell(r, 6).value,
            'v20_served': ws.cell(r, 7).value,
            'er': ws.cell(r, 9).value or 0,
            'ir': ws.cell(r, 10).value or 0,
            'si': ws.cell(r, 11).value or 0,
            'rp': ws.cell(r, 12).value or 0,
            'rs': ws.cell(r, 13).value or 0,
            'ai': ws.cell(r, 14).value or 0,
        }
        # Southern and Sovereign are the same entity in V23
        if name in ('Southern Assisted Living, LLC', 'Sovereign Healthcare Holdings'):
            if southern_scores is None:
                southern_scores = scores.copy()
            else:
                # Take the higher qualitative scores from either entry
                for dim in ('ir', 'si', 'rs', 'ai'):
                    southern_scores[dim] = max(southern_scores[dim], scores[dim])
            continue

        v20_entities[name] = scores

# Add consolidated Southern
if southern_scores:
    v20_entities['Southern Healthcare Mgmt'] = southern_scores
    NAME_MAP['Southern Healthcare Mgmt'] = {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'}
    # Remove the two old entries from NAME_MAP
    NAME_MAP.pop('Southern Assisted Living, LLC', None)
    NAME_MAP.pop('Sovereign Healthcare Holdings', None)


# --- Load V23 DB ---
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

all_rows = []
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    if not corp:
        continue
    all_rows.append({
        'corp': corp,
        'state': str(ws_db.cell(r, headers['State']).value or '').strip(),
        'addr': str(ws_db.cell(r, headers['Address']).value or '').strip().upper(),
        'city': str(ws_db.cell(r, headers['City']).value or '').strip().upper(),
        'served': str(ws_db.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES',
        'integrated': str(ws_db.cell(r, headers['Integrated_Flag']).value or '').strip().upper() == 'YES',
        'pcp': str(ws_db.cell(r, headers['PCP_Flag']).value or '').strip().upper() == 'YES',
        'mh': str(ws_db.cell(r, headers['MH_Flag']).value or '').strip().upper() == 'YES',
        'tot_pot': ws_db.cell(r, headers['Total_Potential_Revenue']).value or 0,
        'int_rev': ws_db.cell(r, headers['Integration_Revenue']).value or 0,
        'nb_rev': ws_db.cell(r, headers['New_Business_Revenue']).value or 0,
    })

# Build index by corp name for fast lookup
corp_rows = defaultdict(list)
for row in all_rows:
    corp_rows[row['corp']].append(row)


def compute_entity(label, corp_names, ir, si, rs_override, ai):
    """Compute V23 footprint-scoped scores for an entity."""
    # Gather all matching rows
    rows = []
    for cn in corp_names:
        rows.extend(corp_rows.get(cn, []))

    # Filter to footprint
    fp_rows = [r for r in rows if r['state'] in FOOTPRINT]

    # Campus collapse
    all_camp = set((r['addr'], r['city']) for r in rows)
    fp_camp = set((r['addr'], r['city']) for r in fp_rows)

    # Served campuses in footprint
    fp_served_camps = set()
    for r in fp_rows:
        if r['served']:
            fp_served_camps.add((r['addr'], r['city']))

    fp_served = len(fp_served_camps)
    fp_total_pot = sum(safe_num(r['tot_pot']) for r in fp_rows)
    fp_int = sum(safe_num(r['int_rev']) for r in fp_rows)
    fp_nb = sum(safe_num(r['nb_rev']) for r in fp_rows)

    # States in footprint
    fp_states = sorted(set(r['state'] for r in fp_rows))

    # Compute ER and RP
    n_camp = len(fp_camp)
    er = er_score(n_camp)
    rp = rp_score(fp_total_pot)

    # Recompute RS from footprint data
    rs = rs_score(fp_served, n_camp)

    # MUO gate
    meets_muo = n_camp >= 7

    # Total score
    score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = tier_label(score) if meets_muo else 'T4'

    return {
        'label': label,
        'all_rows': len(rows),
        'fp_rows': len(fp_rows),
        'all_camp': len(all_camp),
        'fp_camp': n_camp,
        'fp_served': fp_served,
        'fp_states': fp_states,
        'fp_tot_pot': fp_total_pot,
        'fp_int': fp_int,
        'fp_nb': fp_nb,
        'er': er,
        'ir': ir,
        'si': si,
        'rp': rp,
        'rs': rs,
        'ai': ai,
        'score': score,
        'tier': tier,
        'meets_muo': meets_muo,
    }


# --- Process all entities ---
results = []

# V20 entities
for name, scores in v20_entities.items():
    corp_names = NAME_MAP.get(name)
    if not corp_names:
        print(f"WARNING: No name mapping for '{name}' -- skipping")
        continue
    r = compute_entity(name, corp_names, scores['ir'], scores['si'], scores['rs'], scores['ai'])
    r['v20_tier'] = scores['v20_tier']
    r['v20_score'] = scores['v20_score']
    r['v20_er'] = scores['er']
    r['v20_rp'] = scores['rp']
    r['v20_rs'] = scores['rs']
    r['source'] = 'V20'
    results.append(r)

# Red-flagged entities (not in V20)
for name, info in {**RED_FLAGGED, **ADDITIONAL}.items():
    # Skip if already covered (e.g. MCAP might conflict if it was in V20 under different name)
    if any(r['label'] == name for r in results):
        continue
    r = compute_entity(name, info['corp_names'], info['ir'], info['si'], info['rs'], info['ai'])
    r['v20_tier'] = 'NEW'
    r['v20_score'] = '-'
    r['v20_er'] = '-'
    r['v20_rp'] = '-'
    r['v20_rs'] = '-'
    r['source'] = 'RED'
    results.append(r)

# Sort by V23 score descending
results.sort(key=lambda x: (-x['score'], x['label']))


# --- Output ---
print("=" * 140)
print("V23 MUO RESCORE — 6-State Footprint (NC, SC, VA, KY, OH, IN)")
print("=" * 140)
print()

# Summary table
print(f"{'Entity':<50} {'V20':>4} {'V23':>4} {'Scr20':>5} {'Scr23':>5} {'Chg':>4}  {'Camp':>5} {'Srv':>4} {'ER':>3} {'IR':>3} {'SI':>3} {'RP':>3} {'RS':>3} {'AI':>3}  {'FP TotPot':>13}  {'Notes'}")
print("-" * 140)

tier_changes = []
for r in results:
    v20s = r['v20_score'] if isinstance(r['v20_score'], int) else 0
    chg = r['score'] - v20s if isinstance(r['v20_score'], int) else 0
    chg_str = f"{chg:+d}" if isinstance(r['v20_score'], int) else 'NEW'

    notes = []
    if not r['meets_muo']:
        notes.append(f"GATE->T4 ({r['fp_camp']} camp)")
    if r['source'] == 'RED':
        notes.append('New entity (was red-flagged)')
    if isinstance(r['v20_tier'], str) and r['v20_tier'] != 'NEW' and r['tier'] != r['v20_tier']:
        notes.append(f"TIER {r['v20_tier']}->{r['tier']}")
        tier_changes.append(r)

    v20t = r['v20_tier'] if r['v20_tier'] else '?'
    v20sc = str(r['v20_score']) if r['v20_score'] != '-' else '-'

    print(f"{r['label']:<50} {v20t:>4} {r['tier']:>4} {v20sc:>5} {r['score']:>5} {chg_str:>4}  {r['fp_camp']:>5} {r['fp_served']:>4} {r['er']:>3} {r['ir']:>3} {r['si']:>3} {r['rp']:>3} {r['rs']:>3} {r['ai']:>3}  ${r['fp_tot_pot']:>12,.0f}  {'; '.join(notes)}")

print("-" * 140)

# Tier distribution
tier_dist = Counter(r['tier'] for r in results)
print(f"\nTier Distribution: T1={tier_dist.get('T1',0)}  T2={tier_dist.get('T2',0)}  T3={tier_dist.get('T3',0)}  T4(gate)={tier_dist.get('T4',0)}")

# Tier changes summary
if tier_changes:
    print(f"\n{'='*60}")
    print("TIER CHANGES (V20 -> V23)")
    print(f"{'='*60}")
    for r in tier_changes:
        print(f"  {r['label']:<45} {r['v20_tier']} -> {r['tier']}  (score {r['v20_score']} -> {r['score']})")

# Entities that fail MUO gate
gated = [r for r in results if not r['meets_muo']]
if gated:
    print(f"\n{'='*60}")
    print("FAILED MUO GATE (< 7 campuses in footprint -> T4)")
    print(f"{'='*60}")
    for r in gated:
        print(f"  {r['label']:<45} {r['fp_camp']} campuses in {', '.join(r['fp_states'])}")

# No DB match
no_match = [r for r in results if r['all_rows'] == 0]
if no_match:
    print(f"\n{'='*60}")
    print("NO V23 DB MATCH")
    print(f"{'='*60}")
    for r in no_match:
        print(f"  {r['label']}")
