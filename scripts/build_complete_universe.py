"""Enumerate the complete MUO universe across all sources:
1. Finance 60 (MUO_Scoring_Workbook_V23_v8.xlsx)
2. V20 universe (Final_MUO_Tiering_V20.xlsx)
3. BD workbook output (Final_MUO_Tiering_V23.xlsx)
Shows what's in each source, what's missing, and the combined total."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

V20_PATH = "C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx"
V23_FIN_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_MUO_Scoring_Workbook_V23_v8.xlsx"
V23_BD_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_Final_MUO_Tiering_V23.xlsx"

# ============================================================
# READ V20
# ============================================================
print("Reading V20...")
wb_v20 = openpyxl.load_workbook(V20_PATH, data_only=True)

v20_scored = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_v20[sn]
    tier = sn.split()[0]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name and str(name).strip().upper() != 'TOTAL':
            name = str(name).strip()
            score = ws.cell(r, 3).value or 0
            v20_scored[name] = {'tier': tier, 'score': int(score) if score else 0}

v20_t5 = set()
for sn in wb_v20.sheetnames:
    if 'T5' in sn or 'Barrier' in sn:
        ws = wb_v20[sn]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip().upper() != 'TOTAL':
                v20_t5.add(str(name).strip())

print(f"  V20: {len(v20_scored)} scored + {len(v20_t5)} T5 = {len(v20_scored) + len(v20_t5)}")

# ============================================================
# READ FINANCE V23 WORKBOOK
# ============================================================
print("Reading Finance V23 workbook...")
wb_fin = openpyxl.load_workbook(V23_FIN_PATH, data_only=True)

fin_scored = {}
ws_sum = wb_fin['Summary']
for r in range(2, ws_sum.max_row + 1):
    name = ws_sum.cell(r, 1).value
    if not name or str(name).strip() in ('', 'TOTAL'):
        continue
    name = str(name).strip()
    score = ws_sum.cell(r, 15).value
    tier = ws_sum.cell(r, 16).value
    er = ws_sum.cell(r, 3).value or 0
    ir = ws_sum.cell(r, 5).value or 0
    si = ws_sum.cell(r, 7).value or 0
    rp = ws_sum.cell(r, 9).value or 0
    rs = ws_sum.cell(r, 11).value or 0
    ai = ws_sum.cell(r, 13).value or 0
    fin_scored[name] = {
        'tier': tier, 'score': int(score) if score else 0,
        'ER': int(er) if er else 0, 'IR': int(ir) if ir else 0,
        'SI': int(si) if si else 0, 'RP': int(rp) if rp else 0,
        'RS': int(rs) if rs else 0, 'AI': int(ai) if ai else 0,
    }

fin_t4 = {}
ws_t4 = wb_fin['T4 - Independents']
for r in range(2, ws_t4.max_row + 1):
    name = ws_t4.cell(r, 1).value
    if not name or str(name).strip() in ('', 'TOTAL', 'FACILITY DETAIL'):
        continue
    name = str(name).strip()
    # Check if it looks like a header row or facility detail
    col2 = ws_t4.cell(r, 2).value
    if col2 and str(col2).strip() in ('Finance Tier', 'Qualifying Campuses'):
        continue
    camps = ws_t4.cell(r, 3).value or ws_t4.cell(r, 2).value or 0
    fin_t4[name] = {'tier': 'T4', 'campuses': int(camps) if isinstance(camps, (int, float)) else 0}

# T5 from Bracket Reference or known list
fin_t5 = set()
for sn in wb_fin.sheetnames:
    if 'T5' in sn or 'Barrier' in sn:
        ws = wb_fin[sn]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL'):
                fin_t5.add(str(name).strip())

# If no T5 tab, use known list
if not fin_t5:
    fin_t5 = {'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
               'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare'}

fin_total = len(fin_scored) + len(fin_t4) + len(fin_t5)
print(f"  Finance: {len(fin_scored)} scored + {len(fin_t4)} T4 + {len(fin_t5)} T5 = {fin_total}")

# ============================================================
# READ BD V23 WORKBOOK
# ============================================================
print("Reading BD V23 workbook...")
wb_bd = openpyxl.load_workbook(V23_BD_PATH, data_only=True)

bd_entities = {}
for sn in wb_bd.sheetnames:
    ws = wb_bd[sn]
    if any(t in sn for t in ['T1', 'T2', 'T3']):
        tier = sn.split()[0]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL'):
                name = str(name).strip()
                score = ws.cell(r, 3).value or 0
                bd_entities[name] = {'tier': tier, 'score': int(score) if score else 0, 'source_tab': sn}
    elif 'T4' in sn:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL', 'FACILITY DETAIL'):
                name = str(name).strip()
                camps = ws.cell(r, 2).value
                if isinstance(camps, (int, float)) and camps > 0:
                    bd_entities[name] = {'tier': 'T4', 'campuses': int(camps), 'source_tab': sn}
    elif 'T5' in sn or 'Barrier' in sn:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL'):
                bd_entities[str(name).strip()] = {'tier': 'T5', 'source_tab': sn}

print(f"  BD workbook: {len(bd_entities)} total entities")

# ============================================================
# BUILD NAME MATCHING
# ============================================================
# Known name mappings (V20 name -> canonical display name)
# and (Finance name -> canonical display name)
CANONICAL = {}

# Finance names are already canonical (they're the V23 standard)
for name in fin_scored:
    CANONICAL[name.upper()] = name
for name in fin_t4:
    CANONICAL[name.upper()] = name
for name in fin_t5:
    CANONICAL[name.upper()] = name

# BD names
for name in bd_entities:
    if name.upper() not in CANONICAL:
        CANONICAL[name.upper()] = name

# V20 names (may differ)
V20_ALIASES = {
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': 'Avardis',
    'Liberty Senior Living': 'Liberty',
    'Life Care Centers Of America': 'Lifecare',
    'Principle Long Term Care': 'Principle',
    'Pruitthealth': 'Pruitt Health',
    'Terra Bella': 'TerraBella Senior Living',
    'Tlc Management': 'TLC Management',
    'Trilogy Health Services': 'Trilogy',
    'Arbors At Ohio': 'Arbors',
    'Heritage Hall': 'American Healthcare LLC',
    'Jag Healthcare': 'JAG',
    'Kissito': 'Kissito Healthcare',
    'Momentous Health': 'Momentus Health',
    'Peak Resources, Inc.': 'Peak Resources',
    'Priority Life Care': 'Priority',
    'Sanstone Health & Rehabilitation': 'Sanstone',
    'Topaz': 'Topaz Healthcare',
    'Yad Healthcare': 'YAD',
    'Cch Healthcare': 'CCH Healthcare',
    'Otterbein Seniorlife': 'Otterbein Senior Life',
    'Lutheran Services Carolina': 'Lutheran Services Carolinas',
    'Southern Assisted Living, LLC': '(ABSORBED → Brookdale)',
    'Sovereign Healthcare Holdings': '(ABSORBED → Southern Healthcare Mgmt)',
    'Gardant Management Solutions, Inc': 'Gardant Management Solutions',
    'Seky Holding Co.': 'SEKY Holding Co.',
    'Sunrise': 'Sunrise Senior Living',
    'Windsor House, Inc.': 'Windsor House',
    'Aom Healthcare': 'AOM Healthcare',
    'Fundamental Ltc': 'Fundamental LTC',
    'Southern Healthcare Management, LLC': 'Southern Healthcare Mgmt',
}

for v20_name in v20_scored:
    canonical = V20_ALIASES.get(v20_name, v20_name)
    if canonical.startswith('(ABSORBED'):
        continue
    if canonical.upper() not in CANONICAL:
        CANONICAL[canonical.upper()] = canonical

for v20_name in v20_t5:
    if v20_name.upper() not in CANONICAL:
        CANONICAL[v20_name.upper()] = v20_name

# ============================================================
# MERGE INTO COMPLETE UNIVERSE
# ============================================================
universe = {}  # canonical_name -> {sources, tiers}

def add_entity(canonical_name, source, tier, score=None, extra=None):
    if canonical_name not in universe:
        universe[canonical_name] = {'sources': set(), 'tiers': {}, 'extra': {}}
    universe[canonical_name]['sources'].add(source)
    universe[canonical_name]['tiers'][source] = tier
    if score is not None:
        universe[canonical_name]['extra'][f'{source}_score'] = score
    if extra:
        universe[canonical_name]['extra'].update(extra)

# Finance scored
for name, data in fin_scored.items():
    add_entity(name, 'Finance', data['tier'], data['score'])

# Finance T4
for name, data in fin_t4.items():
    add_entity(name, 'Finance', 'T4', extra={'fin_campuses': data.get('campuses', 0)})

# Finance T5
for name in fin_t5:
    add_entity(name, 'Finance', 'T5')

# V20 scored
for v20_name, data in v20_scored.items():
    canonical = V20_ALIASES.get(v20_name, v20_name)
    if canonical.startswith('(ABSORBED'):
        continue
    add_entity(canonical, 'V20', data['tier'], data['score'])

# V20 T5
for v20_name in v20_t5:
    canonical = V20_ALIASES.get(v20_name, v20_name)
    add_entity(canonical, 'V20', 'T5')

# BD workbook
for name, data in bd_entities.items():
    add_entity(name, 'BD', data['tier'], data.get('score'))

# ============================================================
# REPORT
# ============================================================
print(f"\n{'='*130}")
print(f"COMPLETE MUO UNIVERSE — ALL SOURCES COMBINED")
print(f"{'='*130}")
print(f"Total unique entities: {len(universe)}")

# Categorize
in_all_three = []
in_finance_and_v20 = []
in_finance_only = []
in_v20_only = []
in_bd_only = []

for name, info in sorted(universe.items()):
    sources = info['sources']
    if 'Finance' in sources and 'V20' in sources and 'BD' in sources:
        in_all_three.append((name, info))
    elif 'Finance' in sources and 'V20' in sources:
        in_finance_and_v20.append((name, info))
    elif 'Finance' in sources and 'BD' not in sources:
        in_finance_only.append((name, info))
    elif 'V20' in sources and 'Finance' not in sources:
        in_v20_only.append((name, info))
    elif 'BD' in sources and 'Finance' not in sources and 'V20' not in sources:
        in_bd_only.append((name, info))

print(f"\n  In all three (Finance + V20 + BD): {len(in_all_three)}")
print(f"  In Finance + V20 (not in BD):       {len(in_finance_and_v20)}")
print(f"  Finance only (not in BD or V20):     {len(in_finance_only)}")
print(f"  V20 only (not in Finance):           {len(in_v20_only)}")
print(f"  BD only:                             {len(in_bd_only)}")

# Finance entities NOT in BD workbook
fin_not_in_bd = []
for name, info in sorted(universe.items()):
    if 'Finance' in info['sources'] and 'BD' not in info['sources']:
        fin_tier = info['tiers'].get('Finance', '?')
        fin_score = info['extra'].get('Finance_score', '')
        fin_camps = info['extra'].get('fin_campuses', '')
        fin_not_in_bd.append((name, fin_tier, fin_score, fin_camps))

if fin_not_in_bd:
    print(f"\n{'='*130}")
    print(f"FINANCE ENTITIES MISSING FROM BD WORKBOOK ({len(fin_not_in_bd)})")
    print(f"{'='*130}")
    print(f"{'Entity':<45} {'Fin Tier':>8} {'Fin Score':>9} {'Campuses':>9}")
    print('-' * 75)
    for name, tier, score, camps in sorted(fin_not_in_bd, key=lambda x: (x[1], x[0])):
        s = str(score) if score else '-'
        c = str(camps) if camps else '-'
        print(f"  {name:<43} {tier:>8} {s:>9} {c:>9}")

# Complete universe by V23 tier
print(f"\n{'='*130}")
print(f"COMPLETE UNIVERSE BY V23 TIER (what the combined workbook should look like)")
print(f"{'='*130}")

# Determine each entity's "best" V23 tier
final_list = []
for name, info in universe.items():
    # Priority: Finance tier > BD tier > V20 tier
    if 'Finance' in info['tiers']:
        tier = info['tiers']['Finance']
        score = info['extra'].get('Finance_score', info['extra'].get('BD_score', ''))
    elif 'BD' in info['tiers']:
        tier = info['tiers']['BD']
        score = info['extra'].get('BD_score', '')
    else:
        tier = info['tiers'].get('V20', '?')
        score = info['extra'].get('V20_score', '')

    sources = '+'.join(sorted(info['sources']))
    in_bd = 'Y' if 'BD' in info['sources'] else 'N'
    in_fin = 'Y' if 'Finance' in info['sources'] else 'N'
    in_v20 = 'Y' if 'V20' in info['sources'] else 'N'
    final_list.append((name, tier, score, in_fin, in_v20, in_bd))

# Sort by tier then name
tier_order = {'T1': 1, 'T2': 2, 'T3': 3, 'T4': 4, 'T5': 5}
final_list.sort(key=lambda x: (tier_order.get(x[1], 99), x[0]))

print(f"\n{'Entity':<48} {'Tier':>4} {'Score':>5}  {'Fin':>3} {'V20':>3} {'BD':>3}  Source Coverage")
print('-' * 100)

current_tier = None
tier_counts = {}
for name, tier, score, in_fin, in_v20, in_bd, in final_list:
    if tier != current_tier:
        if current_tier:
            print()
        current_tier = tier
        tier_counts[tier] = 0

    tier_counts[tier] = tier_counts.get(tier, 0) + 1
    s = str(score) if score else '-'
    coverage = []
    if in_fin == 'Y': coverage.append('Finance')
    if in_v20 == 'Y': coverage.append('V20')
    if in_bd == 'Y': coverage.append('BD')
    gap = ' ** MISSING FROM BD **' if in_bd == 'N' else ''
    print(f"  {name:<46} {tier:>4} {s:>5}  {in_fin:>3} {in_v20:>3} {in_bd:>3}  {", ".join(coverage)}{gap}")

print(f"\n{'='*130}")
print(f"FINAL COUNTS")
print(f"{'='*130}")
for tier in ['T1', 'T2', 'T3', 'T4', 'T5']:
    count = tier_counts.get(tier, 0)
    print(f"  {tier}: {count}")
print(f"  Total: {sum(tier_counts.values())}")

# One-off profile check
print(f"\n{'='*130}")
print(f"ONE-OFF PROFILE STATUS (Brooke's 4 requests)")
print(f"{'='*130}")
oneoffs = ['Liberty', 'Southern Healthcare Mgmt', 'Triple Crown Senior Living', 'Lutheran Life Villages']
for name in oneoffs:
    if name in universe:
        info = universe[name]
        sources = ', '.join(sorted(info['sources']))
        tier = info['tiers'].get('Finance', info['tiers'].get('BD', '?'))
        in_bd = 'YES' if 'BD' in info['sources'] else 'NO'
        print(f"  {name:<40} Tier: {tier}  In BD: {in_bd}  Sources: {sources}")
    else:
        # Try case insensitive
        found = False
        for uname, uinfo in universe.items():
            if uname.upper() == name.upper():
                sources = ', '.join(sorted(uinfo['sources']))
                tier = uinfo['tiers'].get('Finance', uinfo['tiers'].get('BD', '?'))
                in_bd = 'YES' if 'BD' in uinfo['sources'] else 'NO'
                print(f"  {name:<40} Tier: {tier}  In BD: {in_bd}  Sources: {sources}")
                found = True
                break
        if not found:
            print(f"  {name:<40} ** NOT FOUND IN ANY SOURCE **")
