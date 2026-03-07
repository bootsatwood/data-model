"""Test bracket thresholds against the 60 Finance-tracked MUO entities.
Source: MUO Data (1).xlsx Analysis sheet."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

# --- Load Finance entity list ---
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']

finance_entities = []
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    tier = ws_a.cell(r, 18).value
    glr_facs = ws_a.cell(r, 4).value
    core_rev = ws_a.cell(r, 14).value
    if name and str(name).strip() != 'TOTAL TOP 60':
        finance_entities.append({
            'name': str(name).strip(),
            'tier': tier,
            'glr_facs': int(glr_facs) if glr_facs else 0,
            'core_rev': core_rev or 0,
        })

print(f"Finance-tracked entities: {len(finance_entities)}")

# --- Map Finance names to V23 DB corporate names ---
FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'Bluegrass/Encore': {'BLUEGRASS/ENCORE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'SIGNATURE HEALTH': {'SIGNATURE HEALTH'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'MFA': {'MFA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'COMMUNICARE': {'COMMUNICARE'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES'},
    'Hill Valley': {'HILL VALLEY'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'CLEARVIEW': {'CLEARVIEW'},
    'SINGH': {'SINGH'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'Brighton': {'BRIGHTON'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE'},
    'Momentus Health': {'MOMENTUS HEALTH'},
    'CARDON & ASSOCIATES': {'CARDON & ASSOCIATES'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
}

# --- Load V23 DB footprint data ---
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb_db.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0, 'int_rev': 0, 'nb_rev': 0, 'states': set()})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0
    intg = ws.cell(r, headers['Integration_Revenue']).value or 0
    nb = ws.cell(r, headers['New_Business_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    d['states'].add(state)
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if isinstance(intg, (int, float)):
        d['int_rev'] += intg
    if isinstance(nb, (int, float)):
        d['nb_rev'] += nb
    if served:
        d['served_camps'].add((addr, city))

# --- Match Finance entities to DB ---
def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def rs_score(srv, total):
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

results = []
for fe in finance_entities:
    db_names = FINANCE_TO_DB.get(fe['name'])
    if not db_names:
        results.append({
            'name': fe['name'], 'fin_tier': fe['tier'], 'glr_facs': fe['glr_facs'],
            'core_rev': fe['core_rev'],
            'camp': 0, 'srv': 0, 'rev': 0, 'states': [],
            'matched': False,
        })
        continue

    # Aggregate across all matching DB names
    total_camp = set()
    total_srv = set()
    total_rev = 0
    total_states = set()
    for dn in db_names:
        d = corp_fp.get(dn)
        if d:
            total_camp |= d['campuses']
            total_srv |= d['served_camps']
            total_rev += d['tot_pot']
            total_states |= d['states']

    results.append({
        'name': fe['name'], 'fin_tier': fe['tier'], 'glr_facs': fe['glr_facs'],
        'core_rev': fe['core_rev'],
        'camp': len(total_camp), 'srv': len(total_srv), 'rev': total_rev,
        'states': sorted(total_states),
        'matched': len(total_camp) > 0,
    })

# --- Show the 60 entities with FP data ---
print()
print(f"{'#':>3} {'Entity':<45} {'Fin T':>5} {'GLR':>4} {'FP Camp':>7} {'FP Srv':>6} {'FP S2 Rev':>14} {'Core Rev':>14} {'States'}")
print("-" * 130)
for i, r in enumerate(results, 1):
    tier_str = str(r['fin_tier'] or '?').strip()
    core = f"${r['core_rev']:>13,.0f}" if isinstance(r['core_rev'], (int, float)) else f"{'N/A':>14}"
    states = ', '.join(r['states']) if r['states'] else 'NO MATCH'
    muo_flag = '' if r['camp'] >= 7 else ' *'
    print(f"{i:>3} {r['name']:<45} {tier_str:>5} {r['glr_facs']:>4} {r['camp']:>7} {r['srv']:>6} ${r['rev']:>13,.0f} {core} {states}{muo_flag}")

# Stats
matched = [r for r in results if r['matched']]
revs = sorted([r['rev'] for r in matched if r['rev'] > 0])
camps = sorted([r['camp'] for r in matched])
n = len(revs)

print()
print(f"Matched: {len(matched)}/{len(results)}  |  >=7 campuses: {sum(1 for r in matched if r['camp'] >= 7)}  |  <7: {sum(1 for r in matched if r['camp'] < 7)}")
print(f"* = fails MUO gate (<7 campuses in footprint)")
print()
print("--- Revenue Distribution (matched, rev > 0) ---")
print(f"Count:  {n}")
if n > 0:
    print(f"Min:    ${revs[0]:>13,.0f}")
    print(f"P20:    ${revs[int(n*0.2)]:>13,.0f}")
    print(f"P40:    ${revs[int(n*0.4)]:>13,.0f}")
    print(f"Median: ${revs[n//2]:>13,.0f}")
    print(f"P60:    ${revs[int(n*0.6)]:>13,.0f}")
    print(f"P80:    ${revs[int(n*0.8)]:>13,.0f}")
    print(f"Max:    ${revs[-1]:>13,.0f}")

# --- Current tier distribution from Finance ---
fin_tiers = Counter()
for r in results:
    t = str(r['fin_tier'] or '').strip()
    if t in ('1', '2', '3', '4', '5'):
        fin_tiers[f'T{t}'] += 1
    elif 'not on list' in t.lower():
        fin_tiers['Not listed'] += 1
    elif '4' in t:
        fin_tiers['T4'] += 1
    else:
        fin_tiers['Other'] += 1

print()
print("--- Finance Current Tier Distribution ---")
for t in ['T1', 'T2', 'T3', 'T4', 'T5', 'Not listed', 'Other']:
    if fin_tiers.get(t, 0) > 0:
        print(f"  {t}: {fin_tiers[t]}")

# --- Test ER/RP brackets on the Finance 60 ---
er_options = {
    'V20 (5/10/20/40)': [5, 10, 20, 40],
    'A (10/15/25/50)': [10, 15, 25, 50],
    'C (9/13/20/40)': [9, 13, 20, 40],
}

rp_options = {
    'V20 ($250K/500K/1M/2M)': [250_000, 500_000, 1_000_000, 2_000_000],
    'B ($1M/2.5M/5M/10M)': [1_000_000, 2_500_000, 5_000_000, 10_000_000],
    'C ($2M/4M/7M/12M)': [2_000_000, 4_000_000, 7_000_000, 12_000_000],
    'E ($2M/5M/10M/20M)': [2_000_000, 5_000_000, 10_000_000, 20_000_000],
}

# Only use matched entities with >=7 campuses for bracket testing
scoreable = [r for r in matched if r['camp'] >= 7]
print(f"\nScoreable (>=7 FP campuses): {len(scoreable)}")

print()
print("=" * 80)
print("ER DISTRIBUTIONS (Finance 60, scoreable)")
print("=" * 80)
for name, t in er_options.items():
    dist = Counter(score_dim(e['camp'], t) for e in scoreable)
    print(f"\n{name}:")
    for s in range(1, 6):
        bar = '#' * dist.get(s, 0)
        print(f"  ER={s}: {dist.get(s,0):>3}  {bar}")

print()
print("=" * 80)
print("RP DISTRIBUTIONS (Finance 60, scoreable)")
print("=" * 80)
for name, t in rp_options.items():
    dist = Counter(score_dim(e['rev'], t) for e in scoreable)
    print(f"\n{name}:")
    for s in range(1, 6):
        bar = '#' * dist.get(s, 0)
        print(f"  RP={s}: {dist.get(s,0):>3}  {bar}")
