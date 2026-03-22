#!/usr/bin/env python3
"""
NW Indiana Metro Report -- Blitz Prep for Ian

Three-tier geography (no overlap, no IL):
  1. Core Calumet: East Chicago, Hammond, Gary, Portage (Ian's scoped cities)
  2. Lake County Adjacency: Crown Point, Merrillville, Dyer, Hobart, Munster,
     Schererville, Saint John, Lake Station, Whiting, Lowell
  3. Val-MC Corridor: Valparaiso, Chesterton, Michigan City, La Porte + surrounding

Original report (V20.0) had 92 Calumet (39 IL + 53 IN) and 59 Val-MC with
massive city-level overlap. This version deduplicates geography, drops IL,
and frames the Lake County ring as adjacency density.

Data source: V24.1 (25,583 rows)
Revenue model: Scenario 2 (PCP Enhanced)
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import safe, load_db, VAULT

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

V24_1 = VAULT / "Current" / "1_Combined_Database_FINAL_V24_1.xlsx"
OUTPUT = VAULT / "Current" / "Metro_Reports" / "NW_Indiana_Metro_Report_V24_1_2026-03-18.xlsx"

# ---------------------------------------------------------------------------
# S2 Fee Structure (from build_scenario_2.py)
# ---------------------------------------------------------------------------

FEES = {
    'SNF': {
        'PCP': 3078.00, 'MH_adj': 605.50, 'CCM_adj': 108.00,
        'SS_adj': 792.00, 'TOTAL': 4583.50,
    },
    'ALF': {
        'PCP': 2084.00, 'MH_adj': 715.50, 'CCM_adj': 108.00,
        'SS_adj': 792.00, 'TOTAL': 3699.50,
    },
}
FEES['ILF'] = FEES['ALF']


def calc_s2(row):
    """Calculate S2 revenue columns for a facility row."""
    source_type = safe(row.get('Source_Type', ''))
    fee = FEES.get(source_type)
    if not fee:
        return 0, 0, 0, 0

    try:
        census = float(row.get('Census') or 0)
    except (TypeError, ValueError):
        census = 0

    served = safe(row.get('Do_We_Serve', '')).upper() == 'YES'
    has_barrier = bool(safe(row.get('Barrier', '')))
    integrated = safe(row.get('Integrated_Flag', '')).upper() == 'YES'
    pcp_only = safe(row.get('PCP_Flag', '')).upper() == 'YES'
    mh_only = safe(row.get('MH_Flag', '')).upper() == 'YES'

    current_rev = integration_rev = new_biz_rev = 0

    if served:
        if integrated:
            current_rev = census * fee['TOTAL']
        elif pcp_only:
            current_rev = census * (fee['PCP'] + fee['CCM_adj'] + fee['SS_adj'])
        elif mh_only:
            current_rev = census * fee['MH_adj']
        if not has_barrier:
            if pcp_only:
                integration_rev = census * fee['MH_adj']
            elif mh_only:
                integration_rev = census * (fee['PCP'] + fee['CCM_adj'] + fee['SS_adj'])
    else:
        if not has_barrier:
            new_biz_rev = census * fee['TOTAL']

    total_potential = integration_rev + new_biz_rev
    return current_rev, integration_rev, new_biz_rev, total_potential


# ---------------------------------------------------------------------------
# Three-Tier Geography (no overlap, IN only)
# ---------------------------------------------------------------------------

CORE_CALUMET_CITIES = {'east chicago', 'hammond', 'gary', 'portage'}

# Lake County cities NOT in Core Calumet (adjacency ring)
LAKE_ADJACENCY_CITIES = {
    'crown point', 'merrillville', 'dyer', 'hobart', 'munster',
    'schererville', 'saint john', 'st. john', 'lake station', 'whiting',
    'lowell', 'griffith', 'highland', 'cedar lake',
}

# Val-MC corridor: Porter County (minus Portage) + LaPorte County
VALMC_COUNTIES = {'porter', 'laporte', 'la porte'}

TIER_LABELS = {
    'Core Calumet':       'Core Calumet (East Chicago, Hammond, Gary, Portage)',
    'Lake Co. Adjacency': 'Lake County Adjacency (Crown Point, Merrillville, Dyer, Hobart, Munster +)',
    'Val-MC Corridor':    'Valparaiso-Michigan City Corridor (Porter & LaPorte Counties)',
}

TIER_DESCRIPTIONS = {
    'Core Calumet': (
        "Ian's scoped blitz geography. The industrial lakeshore corridor of "
        "East Chicago, Gary, Hammond, and Portage along I-80/I-94 at the "
        "southern tip of Lake Michigan."
    ),
    'Lake Co. Adjacency': (
        "Adjacent Lake County density surrounding the Core Calumet cities. "
        "Crown Point, Merrillville, Dyer, and Hobart anchor this ring, which "
        "was part of the original Calumet market's 15-mile radius. These "
        "facilities are not in Ian's scoped blitz cities but represent "
        "meaningful nearby density."
    ),
    'Val-MC Corridor': (
        "The eastern NW Indiana corridor extending through Porter and LaPorte "
        "counties. Valparaiso (Porter County seat) and Michigan City anchor "
        "the two ends, connected by US-20 and I-94. Includes Chesterton, "
        "La Porte, and surrounding communities. Entirely within Indiana."
    ),
}


def classify_tier(row):
    """Return 'Core Calumet', 'Lake Co. Adjacency', 'Val-MC Corridor', or None."""
    if safe(row.get('State')).upper() != 'IN':
        return None
    city = safe(row.get('City')).lower()
    county = safe(row.get('County')).lower()

    # Tier 1: Core Calumet (Ian's 4 cities)
    if city in CORE_CALUMET_CITIES:
        return 'Core Calumet'

    # Tier 2: Lake County adjacency
    if city in LAKE_ADJACENCY_CITIES or (county == 'lake' and city not in CORE_CALUMET_CITIES):
        return 'Lake Co. Adjacency'

    # Tier 3: Val-MC corridor (Porter excl. Portage + LaPorte)
    if county in VALMC_COUNTIES:
        return 'Val-MC Corridor'

    # Also catch NW Indiana metro-assigned facilities not covered above
    metro = safe(row.get('Metro_Assignment')).lower()
    if metro == 'northwest indiana':
        return 'Lake Co. Adjacency'

    return None


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="004B87", end_color="004B87", fill_type="solid")
TIER1_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
TIER2_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
TIER3_FILL = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
CURRENCY_FMT = '$#,##0'
PERCENT_FMT = '0.0%'

TIER_FILLS = {
    'Core Calumet': TIER1_FILL,
    'Lake Co. Adjacency': TIER2_FILL,
    'Val-MC Corridor': TIER3_FILL,
}


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = best + 2


# ---------------------------------------------------------------------------
# Summary Sheet
# ---------------------------------------------------------------------------

def build_summary_block(ws, tier_name, facilities, start_row=None):
    """Write a summary statistics block. Returns next available row."""
    if start_row:
        # Position cursor
        while ws.max_row < start_row - 1:
            ws.append([])

    label = TIER_LABELS.get(tier_name, tier_name)
    ws.append([label])
    title_row = ws.max_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=4)
    ws[f'A{title_row}'].font = Font(bold=True, size=13, color="FFFFFF")
    fill = TIER_FILLS.get(tier_name, HEADER_FILL)
    for col in range(1, 5):
        ws.cell(row=title_row, column=col).fill = fill

    # Description
    desc = TIER_DESCRIPTIONS.get(tier_name, '')
    if desc:
        ws.append([desc])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        ws[f'A{ws.max_row}'].font = Font(italic=True, size=9, color="555555")
        ws[f'A{ws.max_row}'].alignment = Alignment(wrap_text=True)
    ws.append([])

    total = len(facilities)
    if total == 0:
        ws.append(["No facilities in this tier."])
        ws.append([])
        return ws.max_row

    snf = [f for f in facilities if safe(f.get('Source_Type')) == 'SNF']
    alf = [f for f in facilities if safe(f.get('Source_Type')) == 'ALF']
    ilf = [f for f in facilities if safe(f.get('Source_Type')) == 'ILF']
    served = [f for f in facilities if safe(f.get('Do_We_Serve', '')).upper() == 'YES']
    barrier = [f for f in facilities if safe(f.get('Barrier', ''))]
    integrated = [f for f in served if safe(f.get('Integrated_Flag', '')).upper() == 'YES']
    pcp_only = [f for f in served if safe(f.get('PCP_Flag', '')).upper() == 'YES']
    mh_only = [f for f in served if safe(f.get('MH_Flag', '')).upper() == 'YES']

    def beds_sum(lst):
        return sum(float(f.get('Total_Beds') or 0) for f in lst)
    def census_sum(lst):
        return sum(float(f.get('Census') or 0) for f in lst)

    total_beds = beds_sum(facilities)
    total_census = census_sum(facilities)
    total_current = sum(f['_current_rev'] for f in facilities)
    total_integration = sum(f['_integration_rev'] for f in facilities)
    total_newbiz = sum(f['_newbiz_rev'] for f in facilities)
    total_potential = sum(f['_total_potential'] for f in facilities)

    # Cities breakdown
    city_counts = Counter(safe(f.get('City')) for f in facilities)
    cities_str = ', '.join(f"{c} ({n})" for c, n in sorted(city_counts.items(), key=lambda x: -x[1]))

    stats = [
        ("Total Facilities", f"{total} ({len(snf)} SNF / {len(alf)} ALF" + (f" / {len(ilf)} ILF" if ilf else "") + ")"),
        ("Cities", cities_str),
        ("Total Licensed Beds", f"{int(total_beds):,}"),
        ("Average Daily Census", f"{total_census:,.1f}"),
        ("Avg Occupancy", f"{total_census/total_beds:.1%}" if total_beds else "N/A"),
        ("", ""),
        ("Currently Served", f"{len(served)} ({len(served)/total:.1%} penetration)"),
        ("  Integrated", str(len(integrated))),
        ("  PCP Only", str(len(pcp_only))),
        ("  MH Only", str(len(mh_only))),
        ("Barrier Facilities", str(len(barrier))),
        ("Growth Eligible", str(total - len(barrier))),
        ("", ""),
        ("Current Revenue", f"${total_current:,.0f}"),
        ("Potential from Integration", f"${total_integration:,.0f}"),
        ("New Business Revenue", f"${total_newbiz:,.0f}"),
        ("Total Potential Revenue", f"${total_potential:,.0f}"),
        ("Grand Total (Current + Potential)", f"${total_current + total_potential:,.0f}"),
    ]

    for label_text, value in stats:
        ws.append([label_text, value])
        if label_text and not label_text.startswith("  "):
            ws[f'A{ws.max_row}'].font = Font(bold=True)

    ws.append([])
    return ws.max_row


# ---------------------------------------------------------------------------
# Facility Listing Sheet
# ---------------------------------------------------------------------------

FACILITY_COLS = [
    ('Facility Name',    'Facility_Name',    35, None),
    ('Type',             'Source_Type',        6, None),
    ('Corporate',        'Corporate_Name',   30, None),
    ('City',             'City',             18, None),
    ('County',           'County',           14, None),
    ('ZIP',              'ZIP',               8, None),
    ('Beds',             'Total_Beds',        7, '#,##0'),
    ('Census',           'Census',            8, '#,##0'),
    ('Served',           'Do_We_Serve',       8, None),
    ('INT',              'Integrated_Flag',   5, None),
    ('PCP',              'PCP_Flag',          5, None),
    ('MH',               'MH_Flag',           5, None),
    ('Barrier',          'Barrier',          20, None),
    ('Current Rev',      '_current_rev',     14, CURRENCY_FMT),
    ('Integration Rev',  '_integration_rev', 14, CURRENCY_FMT),
    ('New Biz Rev',      '_newbiz_rev',      14, CURRENCY_FMT),
    ('Total Potential',  '_total_potential',  14, CURRENCY_FMT),
    ('Tier',             '_tier',            20, None),
]


def build_facility_sheet(ws, sheet_title, facilities):
    """Build a facility listing sheet sorted by Total Potential Revenue desc."""
    ws.title = sheet_title

    # Header row
    for i, (label, _, width, _) in enumerate(FACILITY_COLS, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Sort by Total Potential desc, then Current Rev desc
    sorted_facs = sorted(facilities,
                         key=lambda f: (f['_total_potential'], f['_current_rev']),
                         reverse=True)

    for row_num, fac in enumerate(sorted_facs, start=2):
        for col_num, (_, key, _, fmt) in enumerate(FACILITY_COLS, 1):
            val = fac.get(key, '')
            if val is None:
                val = ''
            if key in ('Total_Beds', 'Census') and val:
                try:
                    val = int(float(val))
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=row_num, column=col_num, value=val)
            if fmt:
                cell.number_format = fmt

    # Totals row
    total_row = len(sorted_facs) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_num, (_, key, _, fmt) in enumerate(FACILITY_COLS, 1):
        if key in ('Total_Beds', 'Census', '_current_rev', '_integration_rev', '_newbiz_rev', '_total_potential'):
            total_val = sum(float(f.get(key) or 0) for f in sorted_facs)
            if '_rev' in key or '_potential' in key:
                cell_val = round(total_val, 2)
            else:
                cell_val = int(total_val)
            cell = ws.cell(row=total_row, column=col_num, value=cell_val)
            cell.font = Font(bold=True)
            if fmt:
                cell.number_format = fmt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(FACILITY_COLS))}{len(sorted_facs) + 1}"
    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Corporate Operators Sheet
# ---------------------------------------------------------------------------

def build_corporate_sheet(ws, all_facs):
    """Build corporate operator analysis across all tiers."""
    ws.title = "Corporate Operators"

    corp_groups = defaultdict(lambda: {
        'count': 0, 'beds': 0, 'census': 0, 'served': 0,
        'current_rev': 0, 'total_potential': 0, 'tiers': set(),
        'types': set(),
    })

    for fac in all_facs:
        corp = safe(fac.get('Corporate_Name')) or '(Independent)'
        g = corp_groups[corp]
        g['count'] += 1
        g['beds'] += float(fac.get('Total_Beds') or 0)
        g['census'] += float(fac.get('Census') or 0)
        if safe(fac.get('Do_We_Serve', '')).upper() == 'YES':
            g['served'] += 1
        g['current_rev'] += fac['_current_rev']
        g['total_potential'] += fac['_total_potential']
        g['tiers'].add(fac['_tier'])
        g['types'].add(safe(fac.get('Source_Type')))

    headers = ['Corporate Name', 'Facilities', 'Type', 'Beds', 'Census',
               'Served', 'Penetration', 'Tiers Present', 'Current Rev',
               'Total Potential']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Sort by total potential desc
    sorted_corps = sorted(corp_groups.items(),
                          key=lambda x: x[1]['total_potential'] + x[1]['current_rev'],
                          reverse=True)

    for row_num, (corp, g) in enumerate(sorted_corps, start=2):
        types_str = '/'.join(sorted(g['types']))
        tiers_str = ', '.join(sorted(g['tiers']))
        pct = g['served'] / g['count'] if g['count'] else 0

        ws.cell(row=row_num, column=1, value=corp)
        ws.cell(row=row_num, column=2, value=g['count'])
        ws.cell(row=row_num, column=3, value=types_str)
        ws.cell(row=row_num, column=4, value=int(g['beds']))
        ws.cell(row=row_num, column=5, value=int(g['census']))
        ws.cell(row=row_num, column=6, value=g['served'])
        cell = ws.cell(row=row_num, column=7, value=pct)
        cell.number_format = PERCENT_FMT
        ws.cell(row=row_num, column=8, value=tiers_str)
        ws.cell(row=row_num, column=9, value=round(g['current_rev'], 2))
        ws.cell(row=row_num, column=9).number_format = CURRENCY_FMT
        ws.cell(row=row_num, column=10, value=round(g['total_potential'], 2))
        ws.cell(row=row_num, column=10).number_format = CURRENCY_FMT

    widths = [35, 10, 10, 8, 8, 8, 12, 30, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:J{len(sorted_corps) + 1}"
    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Comparison Sheet (V20.0 original vs V24.1 regen)
# ---------------------------------------------------------------------------

def build_comparison_sheet(ws, tier_data):
    """Side-by-side comparison: original report vs this regen."""
    ws.title = "V20 vs V24.1 Comparison"

    # Original report numbers
    orig = {
        'Calumet (V20.0)': {
            'facilities': 92, 'il_facilities': 39, 'in_facilities': 53,
            'beds': 13543, 'census': 10172, 'served': 7,
            'current_rev': 657298, 'total_potential': 43120000,
            'barriers': 1, 'note': 'Included 39 IL facilities (south suburban Chicago)',
        },
        'Val-MC (V20.0)': {
            'facilities': 59, 'il_facilities': 0, 'in_facilities': 59,
            'beds': 6122, 'census': 4831, 'served': 12,
            'current_rev': 1980000, 'total_potential': 17830000,
            'barriers': 0, 'note': 'Had significant overlap with Calumet in Crown Point, Merrillville, Hobart, Portage',
        },
    }

    headers = ['Metric', 'Calumet (V20.0)', 'Val-MC (V20.0)',
               'Core Calumet', 'Lake Co. Adjacency', 'Val-MC Corridor', 'NW IN Combined']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    core = tier_data.get('Core Calumet', [])
    adj = tier_data.get('Lake Co. Adjacency', [])
    valmc = tier_data.get('Val-MC Corridor', [])
    combined = core + adj + valmc

    def stats(facs):
        t = len(facs)
        b = sum(float(f.get('Total_Beds') or 0) for f in facs)
        c = sum(float(f.get('Census') or 0) for f in facs)
        s = sum(1 for f in facs if safe(f.get('Do_We_Serve', '')).upper() == 'YES')
        cr = sum(f['_current_rev'] for f in facs)
        tp = sum(f['_total_potential'] for f in facs)
        br = sum(1 for f in facs if safe(f.get('Barrier', '')))
        return t, b, c, s, cr, tp, br

    c1, b1, cs1, s1, cr1, tp1, br1 = stats(core)
    c2, b2, cs2, s2, cr2, tp2, br2 = stats(adj)
    c3, b3, cs3, s3, cr3, tp3, br3 = stats(valmc)
    ct, bt, cst, st, crt, tpt, brt = stats(combined)

    rows_data = [
        ('Facilities',       92,         59,         c1, c2, c3, ct),
        ('  IL Facilities',  39,         0,          0,  0,  0,  0),
        ('  IN Facilities',  53,         59,         c1, c2, c3, ct),
        ('Licensed Beds',    '13,543',   '6,122',    f"{int(b1):,}", f"{int(b2):,}", f"{int(b3):,}", f"{int(bt):,}"),
        ('Avg Daily Census', '10,172',   '4,831',    f"{cs1:,.0f}", f"{cs2:,.0f}", f"{cs3:,.0f}", f"{cst:,.0f}"),
        ('Served',           7,          12,         s1, s2, s3, st),
        ('Barriers',         1,          0,          br1, br2, br3, brt),
        ('Current Rev',      '$657K',    '$1.98M',   f"${cr1:,.0f}", f"${cr2:,.0f}", f"${cr3:,.0f}", f"${crt:,.0f}"),
        ('Total Potential',  '$43.1M',   '$17.8M',   f"${tp1:,.0f}", f"${tp2:,.0f}", f"${tp3:,.0f}", f"${tpt:,.0f}"),
    ]

    for row_num, row_vals in enumerate(rows_data, start=2):
        for col_num, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            if col_num == 1:
                cell.font = Font(bold=not str(val).startswith('  '))

    # Notes
    note_row = len(rows_data) + 3
    ws.cell(row=note_row, column=1, value="Notes on original report:").font = Font(bold=True)
    ws.cell(row=note_row + 1, column=1, value="- Calumet (V20.0) included 39 IL facilities (south suburban Chicago) -- removed per Ian's scope")
    ws.cell(row=note_row + 2, column=1, value="- 6 cities appeared in BOTH markets (Crown Point, Merrillville, Hobart, Portage, Gary, Lake Station) -- deduplicated")
    ws.cell(row=note_row + 3, column=1, value="- V24.1 data reflects 27 row deletions and 72 field updates vs V20.0 source")
    ws.cell(row=note_row + 4, column=1, value="- Lake County Adjacency captures density that was in the original Calumet 15-mile radius but outside Ian's 4 blitz cities")

    widths = [22, 18, 18, 16, 20, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("NW Indiana Metro Report -- V24.1 (Three-Tier)")
    print("=" * 60)
    print(f"Source: {V24_1.name}")
    print(f"Output: {OUTPUT.name}")
    print()

    headers, rows = load_db(V24_1)
    print(f"Loaded {len(rows):,} rows from V24.1")

    # Classify into tiers
    tier_data = {
        'Core Calumet': [],
        'Lake Co. Adjacency': [],
        'Val-MC Corridor': [],
    }

    for r in rows:
        tier = classify_tier(r)
        if not tier:
            continue

        cur, integ, newbiz, total_pot = calc_s2(r)
        r['_current_rev'] = round(cur, 2)
        r['_integration_rev'] = round(integ, 2)
        r['_newbiz_rev'] = round(newbiz, 2)
        r['_total_potential'] = round(total_pot, 2)
        r['_tier'] = tier

        tier_data[tier].append(r)

    # Console output
    for tier_name, facs in tier_data.items():
        print(f"\n{tier_name}: {len(facs)} facilities")
        city_counts = Counter(safe(f.get('City')) for f in facs)
        for city, count in sorted(city_counts.items(), key=lambda x: -x[1]):
            print(f"  {city}: {count}")

    all_facs = []
    for facs in tier_data.values():
        all_facs.extend(facs)

    # Build workbook
    wb = Workbook()

    # Sheet 1: Summary (all three tiers + combined)
    ws_sum = wb.active
    ws_sum.title = "Summary"

    for tier_name in ['Core Calumet', 'Lake Co. Adjacency', 'Val-MC Corridor']:
        build_summary_block(ws_sum, tier_name, tier_data[tier_name])
        ws_sum.append([])

    # Combined
    ws_sum.append(["NW Indiana Combined (All Tiers)"])
    ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=4)
    ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, size=13, color="FFFFFF")
    for col in range(1, 5):
        ws_sum.cell(row=ws_sum.max_row, column=col).fill = HEADER_FILL
    ws_sum.append([])

    total_facs = len(all_facs)
    total_beds = sum(float(f.get('Total_Beds') or 0) for f in all_facs)
    total_census = sum(float(f.get('Census') or 0) for f in all_facs)
    total_served = sum(1 for f in all_facs if safe(f.get('Do_We_Serve', '')).upper() == 'YES')
    total_current = sum(f['_current_rev'] for f in all_facs)
    total_potential = sum(f['_total_potential'] for f in all_facs)

    combined_stats = [
        ("Total Facilities", total_facs),
        ("Total Licensed Beds", f"{int(total_beds):,}"),
        ("Average Daily Census", f"{total_census:,.1f}"),
        ("Currently Served", f"{total_served} ({total_served/total_facs:.1%} penetration)" if total_facs else "0"),
        ("Current Revenue", f"${total_current:,.0f}"),
        ("Total Potential Revenue", f"${total_potential:,.0f}"),
        ("Grand Total", f"${total_current + total_potential:,.0f}"),
    ]
    for label_text, value in combined_stats:
        ws_sum.append([label_text, value])
        ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True)

    auto_width(ws_sum, min_width=12, max_width=55)
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 55

    # Sheet 2: Core Calumet facilities
    ws_t1 = wb.create_sheet()
    build_facility_sheet(ws_t1, "Core Calumet", tier_data['Core Calumet'])

    # Sheet 3: Lake County Adjacency
    ws_t2 = wb.create_sheet()
    build_facility_sheet(ws_t2, "Lake Co Adjacency", tier_data['Lake Co. Adjacency'])

    # Sheet 4: Val-MC Corridor
    ws_t3 = wb.create_sheet()
    build_facility_sheet(ws_t3, "Val-MC Corridor", tier_data['Val-MC Corridor'])

    # Sheet 5: All facilities (combined, with tier column)
    ws_all = wb.create_sheet()
    build_facility_sheet(ws_all, "All NW Indiana", all_facs)

    # Sheet 6: Corporate operators
    ws_corp = wb.create_sheet()
    build_corporate_sheet(ws_corp, all_facs)

    # Sheet 7: V20 vs V24.1 comparison
    ws_comp = wb.create_sheet()
    build_comparison_sheet(ws_comp, tier_data)

    # Sheet 8: Barriers
    ws_barrier = wb.create_sheet("Barriers")
    barrier_facs = [f for f in all_facs if safe(f.get('Barrier', ''))]
    barrier_headers = ['Facility Name', 'Type', 'City', 'Tier', 'Barrier', 'Beds', 'Census']
    for i, h in enumerate(barrier_headers, 1):
        cell = ws_barrier.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    if barrier_facs:
        for row_num, f in enumerate(sorted(barrier_facs, key=lambda x: x['_tier']), start=2):
            ws_barrier.cell(row=row_num, column=1, value=safe(f.get('Facility_Name')))
            ws_barrier.cell(row=row_num, column=2, value=safe(f.get('Source_Type')))
            ws_barrier.cell(row=row_num, column=3, value=safe(f.get('City')))
            ws_barrier.cell(row=row_num, column=4, value=f['_tier'])
            ws_barrier.cell(row=row_num, column=5, value=safe(f.get('Barrier')))
            try:
                ws_barrier.cell(row=row_num, column=6, value=int(float(f.get('Total_Beds') or 0)))
            except (TypeError, ValueError):
                ws_barrier.cell(row=row_num, column=6, value=0)
            try:
                ws_barrier.cell(row=row_num, column=7, value=int(float(f.get('Census') or 0)))
            except (TypeError, ValueError):
                ws_barrier.cell(row=row_num, column=7, value=0)
    else:
        ws_barrier.cell(row=2, column=1, value="No barrier facilities in NW Indiana market.")
    auto_width(ws_barrier, min_width=10, max_width=40)
    ws_barrier.freeze_panes = 'A2'

    # Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")

    # Console summary
    print(f"\n{'=' * 60}")
    print("REPORT SUMMARY")
    print(f"{'=' * 60}")

    for tier_name in ['Core Calumet', 'Lake Co. Adjacency', 'Val-MC Corridor']:
        facs = tier_data[tier_name]
        total = len(facs)
        served = sum(1 for f in facs if safe(f.get('Do_We_Serve', '')).upper() == 'YES')
        beds = sum(float(f.get('Total_Beds') or 0) for f in facs)
        census = sum(float(f.get('Census') or 0) for f in facs)
        current = sum(f['_current_rev'] for f in facs)
        potential = sum(f['_total_potential'] for f in facs)
        barriers = sum(1 for f in facs if safe(f.get('Barrier', '')))

        print(f"\n  {tier_name}:")
        print(f"    Facilities:  {total}")
        print(f"    Beds:        {int(beds):,}")
        print(f"    Census:      {int(census):,}")
        if total:
            print(f"    Served:      {served} ({served/total:.0%})")
        else:
            print(f"    Served:      0")
        print(f"    Barriers:    {barriers}")
        print(f"    Current Rev: ${current:,.0f}")
        print(f"    Potential:   ${potential:,.0f}")
        print(f"    Grand Total: ${current + potential:,.0f}")

    print(f"\n  {'- ' * 30}")
    print(f"  NW Indiana Combined:")
    print(f"    Facilities:  {len(all_facs)}")
    print(f"    Beds:        {int(total_beds):,}")
    print(f"    Census:      {int(total_census):,}")
    print(f"    Served:      {total_served} ({total_served/len(all_facs):.0%})")
    print(f"    Current Rev: ${total_current:,.0f}")
    print(f"    Potential:   ${total_potential:,.0f}")
    print(f"    Grand Total: ${total_current + total_potential:,.0f}")

    print(f"\n{'=' * 60}")
    print("Three-Tier Geography (IN only, no overlap):")
    print("  1. Core Calumet: East Chicago, Hammond, Gary, Portage")
    print("  2. Lake Co. Adjacency: Crown Point, Merrillville, Dyer, Hobart, Munster +")
    print("  3. Val-MC Corridor: Valparaiso, Chesterton, Michigan City, La Porte +")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
