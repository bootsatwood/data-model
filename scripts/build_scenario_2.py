#!/usr/bin/env python3
"""
Build Scenario 2 (PCP Enhanced) Economic Model from V23 database.

Applies S2 service package assumptions to every facility and calculates:
  - Current Revenue (served facilities)
  - Integration Revenue (upsell to served)
  - New Business Revenue (unserved, no barrier)
  - Total Potential Revenue (Integration + New Business)

S2 service packages:
  PCP Only  -> PCP + CCM(adj) + SS(adj)
  MH Only   -> MH(adj)
  Integrated -> TOTAL (PCP + MH(adj) + CCM(adj) + SS(adj))

Usage:
  python build_scenario_2.py           # builds the workbook
  python build_scenario_2.py --dry-run # calculates and prints summary only
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, VAULT, DB_CURRENT

DB_OUTPUT = VAULT / "Current" / "4_Economic_Model_Scenario_2_Combined_V23.xlsx"

# ---------------------------------------------------------------------------
# Fee Structure (from Core Rulebook V20.0 Section 3)
# ---------------------------------------------------------------------------

FEES = {
    'SNF': {
        'PCP':     3078.00,
        'MH_adj':   605.50,
        'CCM_adj':  108.00,
        'SS_adj':   792.00,
        'TOTAL':   4583.50,
    },
    'ALF': {
        'PCP':     2084.00,
        'MH_adj':   715.50,
        'CCM_adj':  108.00,
        'SS_adj':   792.00,
        'TOTAL':   3699.50,
    },
}

# ILF uses ALF fees (ILF was carved from ALF in V22.11)
FEES['ILF'] = FEES['ALF']


# ---------------------------------------------------------------------------
# S2 Revenue Calculation
# ---------------------------------------------------------------------------

def calc_s2(row):
    """Calculate Scenario 2 revenue columns for a single facility row.

    Returns (current_rev, integration_rev, new_biz_rev, total_potential).
    """
    source_type = safe(row.get('Source_Type', ''))
    fee = FEES.get(source_type)
    if not fee:
        return 0, 0, 0, 0

    census_val = row.get('Census')
    try:
        census = float(census_val)
    except (TypeError, ValueError):
        census = 0

    served = safe(row.get('Do_We_Serve', '')).upper() == 'YES'
    has_barrier = bool(safe(row.get('Barrier', '')))
    integrated = safe(row.get('Integrated_Flag', '')).upper() == 'YES'
    pcp_only = safe(row.get('PCP_Flag', '')).upper() == 'YES'
    mh_only = safe(row.get('MH_Flag', '')).upper() == 'YES'

    current_rev = 0
    integration_rev = 0
    new_biz_rev = 0

    if served:
        # --- Current Revenue (never affected by barriers) ---
        if integrated:
            current_rev = census * fee['TOTAL']
        elif pcp_only:
            # S2: PCP Only = PCP + CCM(adj) + SS(adj)
            current_rev = census * (fee['PCP'] + fee['CCM_adj'] + fee['SS_adj'])
        elif mh_only:
            # S2: MH Only = MH(adj)
            current_rev = census * fee['MH_adj']

        # --- Integration Revenue (remaining services) ---
        if not has_barrier:
            if integrated:
                integration_rev = 0  # already has everything
            elif pcp_only:
                # S2: remaining = MH(adj)
                integration_rev = census * fee['MH_adj']
            elif mh_only:
                # S2: remaining = PCP + CCM(adj) + SS(adj)
                integration_rev = census * (fee['PCP'] + fee['CCM_adj'] + fee['SS_adj'])
        # barrier -> integration = $0
    else:
        # --- New Business Revenue ---
        if not has_barrier:
            new_biz_rev = census * fee['TOTAL']
        # barrier -> new biz = $0

    total_potential = integration_rev + new_biz_rev
    return current_rev, integration_rev, new_biz_rev, total_potential


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    dry_run = '--dry-run' in sys.argv

    print("Economic Model — Scenario 2 (PCP Enhanced) — V23")
    print("=" * 60)
    print(f"Source: {DB_CURRENT.name}")
    if not dry_run:
        print(f"Output: {DB_OUTPUT.name}")
    print()

    headers, rows = load_db()
    print(f"Loaded {len(rows):,} rows")

    # --- Calculate revenue for every row ---
    rev_cols = ['Current_Revenue', 'Integration_Revenue',
                'New_Business_Revenue', 'Total_Potential_Revenue']

    totals = {c: 0.0 for c in rev_cols}
    type_totals = {}  # by Source_Type
    served_count = Counter()
    flag_counts = Counter()

    for r in rows:
        cur, integ, newbiz, total_pot = calc_s2(r)
        r['Current_Revenue'] = round(cur, 2)
        r['Integration_Revenue'] = round(integ, 2)
        r['New_Business_Revenue'] = round(newbiz, 2)
        r['Total_Potential_Revenue'] = round(total_pot, 2)

        totals['Current_Revenue'] += cur
        totals['Integration_Revenue'] += integ
        totals['New_Business_Revenue'] += newbiz
        totals['Total_Potential_Revenue'] += total_pot

        st = safe(r.get('Source_Type', ''))
        if st not in type_totals:
            type_totals[st] = {c: 0.0 for c in rev_cols}
        type_totals[st]['Current_Revenue'] += cur
        type_totals[st]['Integration_Revenue'] += integ
        type_totals[st]['New_Business_Revenue'] += newbiz
        type_totals[st]['Total_Potential_Revenue'] += total_pot

        if safe(r.get('Do_We_Serve', '')).upper() == 'YES':
            if safe(r.get('Integrated_Flag', '')).upper() == 'YES':
                flag_counts['Integrated'] += 1
            elif safe(r.get('PCP_Flag', '')).upper() == 'YES':
                flag_counts['PCP Only'] += 1
            elif safe(r.get('MH_Flag', '')).upper() == 'YES':
                flag_counts['MH Only'] += 1
            served_count[st] += 1

    # --- Summary ---
    print(f"\n{'=' * 60}")
    print("SCENARIO 2 (PCP Enhanced) REVENUE SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Current Revenue:        ${totals['Current_Revenue']:>14,.2f}")
    print(f"  Integration Revenue:    ${totals['Integration_Revenue']:>14,.2f}")
    print(f"  New Business Revenue:   ${totals['New_Business_Revenue']:>14,.2f}")
    print(f"  Total Potential:        ${totals['Total_Potential_Revenue']:>14,.2f}")

    print(f"\n  Grand Total (Current + Potential): ${totals['Current_Revenue'] + totals['Total_Potential_Revenue']:>14,.2f}")

    print(f"\n{'-' * 60}")
    print("BY SOURCE TYPE")
    for st in sorted(type_totals.keys()):
        tt = type_totals[st]
        print(f"\n  {st}:")
        print(f"    Current:      ${tt['Current_Revenue']:>14,.2f}")
        print(f"    Integration:  ${tt['Integration_Revenue']:>14,.2f}")
        print(f"    New Business: ${tt['New_Business_Revenue']:>14,.2f}")
        print(f"    Total Pot:    ${tt['Total_Potential_Revenue']:>14,.2f}")

    print(f"\n{'-' * 60}")
    print("SERVED FACILITIES")
    total_served = sum(served_count.values())
    print(f"  Total served: {total_served}")
    for st in sorted(served_count.keys()):
        print(f"    {st}: {served_count[st]}")
    print(f"  By service flag:")
    for flag in ['PCP Only', 'MH Only', 'Integrated']:
        print(f"    {flag}: {flag_counts[flag]}")

    barrier_count = sum(1 for r in rows if safe(r.get('Barrier', '')))
    print(f"\n  Barriers: {barrier_count}")

    if dry_run:
        print(f"\n  ** DRY RUN -- no file written **")
        return

    # --- Write workbook ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_CURRENT)
    ws = wb.active

    # Read existing header row to find column count
    existing_headers = [safe(ws.cell(row=1, column=c).value)
                        for c in range(1, ws.max_column + 1)]
    start_col = len(existing_headers) + 1

    # Check if revenue columns already exist (shouldn't, but be safe)
    for i, col_name in enumerate(rev_cols):
        if col_name in existing_headers:
            print(f"  WARNING: {col_name} already exists at column {existing_headers.index(col_name) + 1}")
            start_col = existing_headers.index(rev_cols[0]) + 1
            break

    # Write revenue column headers
    for i, col_name in enumerate(rev_cols):
        ws.cell(row=1, column=start_col + i, value=col_name)

    # Write revenue values for each data row
    for r in rows:
        excel_row = r['_excel_row']
        ws.cell(row=excel_row, column=start_col, value=r['Current_Revenue'])
        ws.cell(row=excel_row, column=start_col + 1, value=r['Integration_Revenue'])
        ws.cell(row=excel_row, column=start_col + 2, value=r['New_Business_Revenue'])
        ws.cell(row=excel_row, column=start_col + 3, value=r['Total_Potential_Revenue'])

    wb.save(DB_OUTPUT)
    print(f"  Saved: {DB_OUTPUT}")
    print(f"  {len(rows):,} rows × {len(rev_cols)} revenue columns written")


if __name__ == '__main__':
    main()
