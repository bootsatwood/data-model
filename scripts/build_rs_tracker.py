"""Build RS Score Audit Trail workbook.

Three touchpoints:
  Touch 1: Brooke V20 Scoring (Dec 2025)
  Touch 2: Tom SS Scoring for V23 (Mar 2026)
  Touch 3: Email Thread (Mar 11, 2026) — both Brooke and Tom on 7 gap entities
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/RS_Score_Audit_Trail.xlsx"

# --- Styling ---
NAVY = "193241"
PURPLE = "877BE9"
PEWTER = "507E8E"
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL_ENTITY = PatternFill("solid", fgColor=NAVY)
HEADER_FILL_T1 = PatternFill("solid", fgColor="2E75B6")
HEADER_FILL_T2 = PatternFill("solid", fgColor=PURPLE)
HEADER_FILL_T3 = PatternFill("solid", fgColor=PEWTER)
SUBHEAD_FONT = Font(bold=True, size=9)
DATA_FONT = Font(size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow for score changes
CONFIRMED_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for confirmed
NO_DATA_FILL = PatternFill("solid", fgColor="F2F2F2")  # gray for no input

# --- Data ---
# Each entity: (name, tier, brooke_rs, tom_rs, tom_notes,
#               email_tom_notes, email_brooke_notes)
# None = no input for that touchpoint

ENTITIES = [
    # T1
    ("American Senior Communities", "T1", 5, 5, "Have it on our paper, they are happy", None, None),
    ("Brookdale Senior Living", "T1", 5, 5, "Have on our paper", None, None),
    ("Otterbein Senior Life", "T1", 5, 1, "Do not really know well", None, None),
    ("Saber Healthcare Group", "T1", 5, 2, "Not very focused, not working", None, None),
    ("Majestic Care", "T1", 5, 3, "Expressed interest but has not gone anywhere", None, None),
    ("Trilogy", "T1", 4, 4, "Already covered, speaking with Dr. McNamara", None, None),
    ("Pruitt Health", "T1", 4, 3, "Getting more PC, focused on own ISNP", None, None),
    ("Pavilion Healthcare", "T1", 4, 1, "Do not really know well", None, None),
    ("Liberty", "T1", 3, 4, "Gave them proposal in RFP, Liberty Advantage ISNP", None, None),
    ("ALG", "T1", 3, 2, "Presented SS, Charlie not interested", None, None),
    ("Avardis", "T1", 3, 2, "Do not really know well, do not want SS", None, None),
    ("CCH Healthcare", "T1", 3, 2, "Talk to them defensively", None, None),
    ("Navion", "T1", 3, 3, "Good to start talking, growing", None, None),
    ("Principle", "T1", 3, 3, "Meeting scheduled on the 17th", None, None),
    ("Sunrise Senior Living", "T1", 3, 1, None,
     "Established corporate relationship. Current opp in NJ/WI (outside footprint). VA awarded to Curana for entire state. Limited opp in our states.",
     "Need to stay close, never know when incumbent could fail. Need to find terms of Curana partnership, auto renewal?"),
    ("TerraBella Senior Living", "T1", 3, 2, "Would have to drive growth", None, None),
    ("TLC Management", "T1", 3, 3, "Part of Hoosier Alliance, not happy with ISNP, active contact", None, None),
    ("Infinity Healthcare Consulting", "T1", 1, 2, "Getting nowhere", None, None),
    ("Ciena Healthcare", "T1", 1, 1, None, None, None),
    ("Lifecare", "T1", 1, 1, "Do not really know well", None, None),
    ("National Healthcare Corp", "T1", 1, 1, None, None, None),
    ("PACS Group", "T1", 1, 1, None, None, None),
    ("Kisco Senior Living", "T1", None, 1, "Do not really know well", None, None),
    ("Lutheran Life Villages", "T1", None, 4, "Cathy worked with Mike to put them in SS", None, None),
    ("Sonida Senior Living", "T1", None, 1, "Growth opp, not doing much", None, None),
    # T2
    ("Lutheran Services Carolinas", "T2", 5, 4, "Verbal commitment for SS, part of how we won business", None, None),
    ("Arbors", "T2", 4, 1, "Do not really know well", None, None),
    ("AOM Healthcare", "T2", 3, 1, None, None, None),
    ("Aperion Care", "T2", 3, 1, None, None, None),
    ("American Healthcare LLC", "T2", 3, 3, "Probably need to do something, Tom to talk to Colvin", None, None),
    ("Castle Healthcare", "T2", 3, 3, "Brooke and Ian have been more involved", None, None),
    ("Lionstone Care", "T2", 3, 3, "Tom needs to talk to Kim B (action pending)", None, None),
    ("Morning Pointe Senior Living", "T2", 3, 2, "Not a great fit, do not recommend doing anything", None, None),
    ("Ohio Living Communities", "T2", 3, 1, None, None, None),
    ("Peak Resources", "T2", 3, 2, "Not responsive or interested in SS", None, None),
    ("Priority", "T2", 3, 1, "Do not really know well", None, None),
    ("Sanstone", "T2", 3, 2, "Presented SS, canceled meetings, not ready to commit", None, None),
    ("JAG", "T2", 3, 2, "Not much growth opp", None, None),
    ("Clearview", "T2", 1, 3, "In process of partnering with Telos", None, None),
    ("Kissito Healthcare", "T2", 1, 2, "COO not interested in SS, just stay close", None, None),
    ("YAD", "T2", 1, 2, "Tried but never gotten an audience, defensive", None, None),
    ("BHI Senior Living", "T2", 1, 1, "All psych", None, None),
    ("Brickyard Healthcare", "T2", 1, 1, None,
     "Existing relationship. Years of fits and starts. Hoosier Alliance member, viable for IN.",
     "Lost business to Rounding Providers (emotional decision by prior COO). Jami Patterson has great relationship with new COO, recovery meeting mid-end April. New AE Gunner Grider churned up new opp via relationship with Regionals."),
    ("Carespring", "T2", 1, 1, None, None, None),
    ("HCF Management", "T2", 1, 1, None,
     "No HQ relationship I am aware of", None),
    ("Phoenix Senior Living", "T2", 1, 1, None,
     "Need input from Brooke and Ian",
     "Two facilities in Charlotte NC. Going on EMR, currently do not want to give us access. Could be more to come."),
    ("Southern Healthcare Mgmt", "T2", 1, 1, "Not in MUO Data notes", None, None),
    ("Topaz Healthcare", "T2", 1, 1, "Do not really know well", None, None),
    ("Eldercare Partners", "T2", None, 3, "Good relationship with them, out of KY", None, None),
    ("Storypoint", "T2", None, 3, "Part of a growth plan", None, None),
    ("Runk & Pratt", "T2", None, 1, "Do not really know well", None, None),
    ("MCAP", "T2", None, 1, "Do not really know well", None, None),
    ("Greencroft", "T2", None, 1, "Do not really know well", None, None),
    ("LifeSpire of Virginia", "T2", None, 1, "Do not really know well", None, None),
    ("Senior Lifestyle", "T2", None, 1, "Do not really know well", None, None),
    ("Cedarhurst Senior Living", "T2", None, 1, "Do not really know well", None, None),
    ("Spring Arbor Management", "T2", None, 1, "Do not really know well", None, None),
    ("Triple Crown", "T2", None, 1, "Not in MUO Data notes", None, None),
    # T3
    ("Caring Place Healthcare", "T3", 1, 2, "Psych only, will not do PC, have their own", None, None),
    ("Carecore Health", "T3", 1, 1, None, None, None),
    ("Envive Healthcare", "T3", 1, 1, None,
     "Need input from Brooke and Ian", "Unknown"),
    ("Miller's Merry Manor", "T3", 1, 1, None,
     "Post-covid reduced dramatically. Hoosier Alliance member. Struggled to have meaningful dialogue.",
     "True, but we do have facilities, many even an employee clinic at one or two. Will double check."),
    ("Trio Healthcare", "T3", 1, 1, None,
     "Need input from Brooke and Ian",
     "No relationship. Termed in a couple of VA facilities. VA-based only. Emailed corp but could not get a call back."),
    ("Fundamental LTC", "T3", None, 1, "Psych only", None, None),
]

def score_effect(before, after):
    """Describe the effect on production RS."""
    if before is None:
        return f"Set to {after}"
    if before == after:
        return "No change"
    delta = after - before
    direction = "+" if delta > 0 else ""
    return f"{before} -> {after} ({direction}{delta})"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RS Audit Trail"

# --- Column layout ---
# A: MUO Name
# B: Tier
# C-E: Touch 1 (Brooke V20): Score, RS Effect, Notes
# F-H: Touch 2 (Tom V23): Score, RS Effect, Notes
# I-K: Touch 3 (Email Mar 11): Input, RS Effect, Notes
# L: Current Production RS

# Row 1: Touchpoint group headers
merge_ranges = [
    ("A1:B1", "Entity", HEADER_FILL_ENTITY),
    ("C1:E1", "Touch 1: Brooke V20 Scoring (Dec 2025)", HEADER_FILL_T1),
    ("F1:H1", "Touch 2: Tom SS Scoring (Mar 2026)", HEADER_FILL_T2),
    ("I1:K1", "Touch 3: Email Thread (Mar 11, 2026)", HEADER_FILL_T3),
    ("L1:L1", "Current", HEADER_FILL_ENTITY),
]

for rng, title, fill in merge_ranges:
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = title
    cell.font = HEADER_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Row 2: Column sub-headers
sub_headers = [
    ("A2", "MUO Name"),
    ("B2", "Tier"),
    ("C2", "Brooke RS"),
    ("D2", "Effect on Production RS"),
    ("E2", "Score Logic"),
    ("F2", "Tom RS"),
    ("G2", "Effect on Production RS"),
    ("H2", "Tom's Notes"),
    ("I2", "Input (Tom=red, Brooke=blue)"),
    ("J2", "Effect on Production RS"),
    ("K2", "Score Logic"),
    ("L2", "Production RS"),
]

for cell_ref, title in sub_headers:
    cell = ws[cell_ref]
    cell.value = title
    cell.font = SUBHEAD_FONT
    cell.alignment = WRAP
    cell.border = THIN_BORDER

# --- Data rows ---
row = 3
for name, tier, brooke_rs, tom_rs, tom_notes, email_tom, email_brooke in ENTITIES:
    # Touch 1: Brooke V20
    if brooke_rs is not None:
        t1_score = brooke_rs
        t1_effect = f"Set to {brooke_rs}"
        t1_logic = "Brooke's BD behavioral assessment"
        prod_after_t1 = brooke_rs
    else:
        t1_score = None
        t1_effect = None
        t1_logic = None
        prod_after_t1 = None

    # Touch 2: Tom V23
    t2_score = tom_rs
    if prod_after_t1 is not None:
        t2_effect = score_effect(prod_after_t1, tom_rs)
        if tom_rs != prod_after_t1:
            t2_logic = f"Tom's SS score replaced Brooke's (was {prod_after_t1})"
        else:
            t2_logic = "Tom confirmed Brooke's score"
    else:
        t2_effect = f"Set to {tom_rs}"
        t2_logic = "First score for this entity (Brooke did not assess)"
    prod_after_t2 = tom_rs

    # Touch 3: Email (Mar 11)
    if email_tom or email_brooke:
        parts = []
        if email_tom:
            parts.append(f"[Tom] {email_tom}")
        if email_brooke:
            parts.append(f"[Brooke] {email_brooke}")
        t3_input = "\n".join(parts)
        # Email was qualitative, did not change production scores
        t3_effect = "No change (qualitative input, score not updated)"
        t3_logic = "Captured as notes. Score review pending."
        prod_final = prod_after_t2
    else:
        t3_input = None
        t3_effect = None
        t3_logic = None
        prod_final = prod_after_t2

    # Write row
    data = [
        name, tier,
        t1_score, t1_effect, t1_logic,
        t2_score, t2_effect, tom_notes,
        t3_input, t3_effect, t3_logic,
        prod_final,
    ]

    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

        # Conditional formatting
        if col_idx in (4, 7, 10):  # Effect columns
            if val and "No change" in str(val):
                cell.fill = CONFIRMED_FILL
            elif val and "->" in str(val):
                cell.fill = CHANGED_FILL
            elif val is None:
                cell.fill = NO_DATA_FILL
        elif val is None:
            cell.fill = NO_DATA_FILL

    row += 1

# --- Column widths ---
widths = {
    "A": 30, "B": 6,
    "C": 10, "D": 22, "E": 30,
    "F": 10, "G": 22, "H": 45,
    "I": 55, "J": 22, "K": 30,
    "L": 12,
}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w

# Freeze panes
ws.freeze_panes = "C3"

# Row height for header
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 35

wb.save(OUT_PATH)
print(f"Saved to {OUT_PATH}")
print(f"Entities: {len(ENTITIES)}")
