"""
Q1 Sale Type Verification — 9 Remaining Facilities
Pull V25.9 DB + NIC MAP data for each facility to support the 11-step verification.
"""
import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

DB_PATH = r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_9.xlsx"
NIC_PATH = r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Reference\Source_NIC_Maps_Nationwide_Inventory_Export_11.17.25.xlsx"

# ---- The 9 facilities to verify ----
# (display_name, city, state, is_est, claimed_corp, claimed_class, search_terms)
FACILITIES = [
    ("Elderwood Rehab & Nursing", "GEORGETOWN", "SC", False, "OAK HOLLOW HEALTHCARE MANAGEMENT", "New Door",
     ["elderwood", "oak hollow", "georgetown"]),
    ("PruittHealth - Raleigh", "RALEIGH", "NC", True, "PRUITT HEALTH", "Integration",
     ["pruitt", "raleigh"]),
    ("Gabriel Manor", "FAYETTEVILLE", "NC", True, "SABER HEALTHCARE GROUP", "Integration",
     ["gabriel"]),
    ("Brookdale Harrisonburg", "HARRISONBURG", "VA", False, "BROOKDALE SENIOR LIVING", "New Door",
     ["brookdale", "harrisonburg"]),
    ("TerraBella Little Avenue", "CHARLOTTE", "NC", False, "TERRABELLA SENIOR LIVING", "New Door",
     ["terrabella", "little"]),
    ("Marquette Manor Memory Care", "INDIANAPOLIS", "IN", True, "LIFECARE", "Integration",
     ["marquette"]),
    ("Franklin Manor", "FRANKLIN", "NC", True, "SABER HEALTHCARE GROUP", "Integration",
     ["franklin manor"]),
    ("Brookdale Lakeview Crossing", "COLUMBUS", "OH", False, "BROOKDALE SENIOR LIVING", "New Door",
     ["brookdale", "lakeview"]),
    ("Cardinal Landing Memory Care", "FORT MITCHELL", "KY", True, "TRIPLE CROWN SENIOR LIVING", "Integration",
     ["cardinal landing"]),
]

# ---- Load V25.9 DB ----
print("Loading V25.9 database...")
wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
header = None
db_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        header = [str(c or "").strip() for c in row]
        continue
    db_rows.append(row)
wb.close()

# Print header for reference
print(f"DB Columns: {header[:15]}")
print(f"DB Rows: {len(db_rows)}")

def fc(name, hdr):
    for i, h in enumerate(hdr):
        if name.lower() == h.lower():
            return i
    for i, h in enumerate(hdr):
        if name.lower() in h.lower():
            return i
    return None

c_fac = fc("Facility_Name", header)
c_corp = fc("Corporate_Name", header)
c_serve = fc("Do_We_Serve", header)
c_city = fc("City", header)
c_state = fc("State", header)
c_own = fc("Ownership_Type", header)
c_beds = fc("Beds", header)
c_type = fc("Source_Type", header) or fc("Facility_Type", header)
c_ccn = fc("CCN", header)
c_npi = fc("NPI", header)

# Build corporate portfolio
corp_port = defaultdict(lambda: {"total": 0, "served": 0, "facs": []})
for row in db_rows:
    corp = (str(row[c_corp] or "")).strip().upper()
    serve = str(row[c_serve] or "").strip()
    fac = str(row[c_fac] or "").strip()
    city = str(row[c_city] or "").strip()
    state = str(row[c_state] or "").strip()
    if corp:
        corp_port[corp]["total"] += 1
        if serve == "Yes":
            corp_port[corp]["served"] += 1
            corp_port[corp]["facs"].append(f"{fac} ({city}, {state})")

# Search function
def find_in_db(terms, target_city, target_state):
    """Find facility in DB by terms + city + state."""
    candidates = []
    for idx, row in enumerate(db_rows):
        row_state = str(row[c_state] or "").strip().upper()
        if row_state != target_state.upper():
            continue
        row_fac = str(row[c_fac] or "").strip().lower()
        row_city = str(row[c_city] or "").strip().upper()

        for term in terms:
            if term.lower() in row_fac:
                city_match = target_city.upper() in row_city or row_city in target_city.upper()
                candidates.append((idx, row, city_match))
                break

    # Prefer city match
    city_hits = [(i, r) for i, r, cm in candidates if cm]
    if city_hits:
        return city_hits[0]
    if candidates:
        return (candidates[0][0], candidates[0][1])
    return None

# ---- Load NIC MAP ----
print("\nLoading NIC MAP...")
try:
    nic_wb = openpyxl.load_workbook(NIC_PATH, read_only=True, data_only=True)
    # Find Buildings sheet
    nic_sheet = None
    for name in nic_wb.sheetnames:
        if "building" in name.lower():
            nic_sheet = nic_wb[name]
            break
    if not nic_sheet:
        nic_sheet = nic_wb[nic_wb.sheetnames[0]]

    nic_header = None
    nic_rows = []
    for i, row in enumerate(nic_sheet.iter_rows(values_only=True)):
        if i == 0:
            nic_header = [str(c or "").strip() for c in row]
            continue
        nic_rows.append(row)
    nic_wb.close()

    print(f"NIC MAP Columns (first 50): {nic_header[:50]}")
    print(f"NIC MAP Rows: {len(nic_rows)}")

    # Find NIC MAP columns
    nc_name = fc("Property Name", nic_header) or fc("Building Name", nic_header) or fc("Name", nic_header)
    nc_city = fc("City", nic_header)
    nc_state = fc("State", nic_header)
    nc_owner = None
    nc_operator = None

    # Find Owner and Operator columns by index or name
    for i, h in enumerate(nic_header or []):
        h_lower = h.lower()
        if "owner" in h_lower and "name" in h_lower and nc_owner is None:
            nc_owner = i
        if "operator" in h_lower and "name" in h_lower and nc_operator is None:
            nc_operator = i

    # Try by column index if not found by name
    if nc_owner is None and len(nic_header) > 18:
        nc_owner = 17  # col 18 (0-indexed = 17)
    if nc_operator is None and len(nic_header) > 41:
        nc_operator = 40  # col 41 (0-indexed = 40)

    print(f"NIC MAP Key Cols: Name={nc_name}, City={nc_city}, State={nc_state}, Owner={nc_owner} ({nic_header[nc_owner] if nc_owner and nc_owner < len(nic_header) else '?'}), Operator={nc_operator} ({nic_header[nc_operator] if nc_operator and nc_operator < len(nic_header) else '?'})")

    def find_in_nic(terms, target_city, target_state):
        matches = []
        for row in nic_rows:
            if nc_state is not None:
                row_state = str(row[nc_state] or "").strip().upper()
                if row_state != target_state.upper():
                    continue
            if nc_name is not None:
                row_name = str(row[nc_name] or "").strip().lower()
            else:
                continue

            for term in terms:
                if term.lower() in row_name:
                    row_city = str(row[nc_city] or "").strip() if nc_city is not None else ""
                    row_owner = str(row[nc_owner] or "").strip() if nc_owner is not None else ""
                    row_operator = str(row[nc_operator] or "").strip() if nc_operator is not None else ""
                    matches.append({
                        "name": str(row[nc_name] or "").strip(),
                        "city": row_city,
                        "state": target_state,
                        "owner": row_owner,
                        "operator": row_operator,
                    })
                    break
        return matches

    NIC_LOADED = True
except Exception as e:
    print(f"NIC MAP load failed: {e}")
    NIC_LOADED = False

# ---- Process each facility ----
print("\n" + "=" * 140)
print("VERIFICATION DATA PULL — 9 REMAINING FACILITIES")
print("=" * 140)

for display, city, state, is_est, claimed_corp, claimed_class, terms in FACILITIES:
    print(f"\n{'='*140}")
    print(f"FACILITY: {display} ({city}, {state})")
    print(f"Claimed: {claimed_corp} | {claimed_class} | EST={is_est}")
    print(f"{'='*140}")

    # Step 1: DB match
    result = find_in_db(terms, city, state)
    if result:
        idx, row = result
        db_fac = str(row[c_fac] or "").strip()
        db_corp = str(row[c_corp] or "").strip()
        db_serve = str(row[c_serve] or "").strip()
        db_city = str(row[c_city] or "").strip()
        db_state = str(row[c_state] or "").strip()
        db_own = str(row[c_own] or "").strip()
        db_beds = str(row[c_beds] or "").strip() if c_beds else "?"
        db_type = str(row[c_type] or "").strip() if c_type else "?"
        db_ccn = str(row[c_ccn] or "").strip() if c_ccn else ""
        db_npi_val = str(row[c_npi] or "").strip() if c_npi else ""

        print(f"\n  [DB V25.9] Row {idx+2}")
        print(f"    Facility:  {db_fac}")
        print(f"    City:      {db_city}, {db_state}")
        print(f"    Corporate: {db_corp}")
        print(f"    Ownership: {db_own}")
        print(f"    Beds:      {db_beds}")
        print(f"    Type:      {db_type}")
        print(f"    Serve:     {db_serve}")
        print(f"    CCN:       {db_ccn}")
        print(f"    NPI:       {db_npi_val}")
    else:
        print(f"\n  [DB V25.9] NOT FOUND in {state}")
        db_corp = ""
        db_ccn = ""
        db_npi_val = ""
        db_serve = ""
        db_fac = ""

    # Step 2: Corporate portfolio
    corp_key = (db_corp or claimed_corp).upper()
    port = corp_port.get(corp_key, {"total": 0, "served": 0, "facs": []})
    print(f"\n  [PORTFOLIO] {corp_key}")
    print(f"    Total facilities: {port['total']}")
    print(f"    Served:           {port['served']}")
    if port['served'] > 0 and port['served'] <= 10:
        print(f"    Served list:")
        for f in port['facs']:
            print(f"      - {f}")

    # Step 3: NIC MAP
    if NIC_LOADED:
        nic_matches = find_in_nic(terms, city, state)
        if nic_matches:
            print(f"\n  [NIC MAP] {len(nic_matches)} match(es)")
            for m in nic_matches[:3]:
                print(f"    Name:     {m['name']}")
                print(f"    City:     {m['city']}, {m['state']}")
                print(f"    Owner:    {m['owner']}  (PropCo — never code as operator)")
                print(f"    Operator: {m['operator']}")
        else:
            print(f"\n  [NIC MAP] No match found")

    # Step 4: Classification check
    print(f"\n  [CLASSIFICATION CHECK]")
    if is_est:
        print(f"    EST item → Integration (by definition)")
        calc_class = "Integration"
    elif db_serve == "Yes":
        print(f"    Do_We_Serve=Yes → Integration")
        print(f"    NOTE: If this was marked Yes AFTER the Q1 win, classify at time of consent")
        calc_class = "Integration (check timing)"
    elif port["served"] > 0:
        print(f"    {port['served']} served under {corp_key} → New Door")
        calc_class = "New Door"
    else:
        print(f"    0 served under {corp_key} → New Logo")
        calc_class = "New Logo"

    print(f"    Claimed: {claimed_class}  |  Calculated: {calc_class}")
    if calc_class != claimed_class and "check timing" not in calc_class:
        print(f"    *** DISCREPANCY ***")

    # CMS/ProPublica lookup hints
    if db_ccn:
        print(f"\n  [EXTERNAL VERIFICATION URLS]")
        print(f"    CMS Provider Info: data.cms.gov (search CCN {db_ccn})")
        print(f"    ProPublica:        projects.propublica.org/nursing-homes/homes/h-{db_ccn}")
    if db_npi_val:
        print(f"    NPI Registry:      npiprofile.com/npi/{db_npi_val}")
