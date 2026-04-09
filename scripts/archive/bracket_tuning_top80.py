"""Test bracket thresholds on TOP 80 MUO entities by FP revenue.
Matches V20 scope (70 scored entities)."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0, 'states': set()})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    d['states'].add(state)
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served:
        d['served_camps'].add((addr, city))

# All MUO-qualifying, exclude INDEPENDENT
all_muos = []
for corp, d in corp_fp.items():
    n = len(d['campuses'])
    if n >= 7 and corp != 'INDEPENDENT':
        all_muos.append({
            'corp': corp,
            'camp': n,
            'srv': len(d['served_camps']),
            'rev': d['tot_pot'],
        })

# Take top 80 by revenue
all_muos.sort(key=lambda x: -x['rev'])
top = all_muos[:80]

print(f"Full MUO universe: {len(all_muos)} entities")
print(f"Analyzing top 80 by FP revenue")
print()

# Show the top 80
print(f"{'#':>3} {'Entity':<50} {'Camp':>5} {'Srv':>4} {'FP Rev':>14}")
print("-" * 82)
for i, e in enumerate(top, 1):
    print(f"{i:>3} {e['corp']:<50} {e['camp']:>5} {e['srv']:>4} ${e['rev']:>13,.0f}")

# Revenue stats for top 80
revs = sorted([e['rev'] for e in top])
camps = sorted([e['camp'] for e in top])
n = len(revs)

print()
print("--- Top 80 Revenue Distribution ---")
print(f"Min:    ${revs[0]:>13,.0f}")
print(f"P20:    ${revs[int(n*0.2)]:>13,.0f}")
print(f"P40:    ${revs[int(n*0.4)]:>13,.0f}")
print(f"Median: ${revs[n//2]:>13,.0f}")
print(f"P60:    ${revs[int(n*0.6)]:>13,.0f}")
print(f"P80:    ${revs[int(n*0.8)]:>13,.0f}")
print(f"Max:    ${revs[-1]:>13,.0f}")

print()
print("--- Top 80 Campus Distribution ---")
print(f"Min: {camps[0]}, P20: {camps[int(n*0.2)]}, Median: {camps[n//2]}, P80: {camps[int(n*0.8)]}, Max: {camps[-1]}")

# ER options
er_options = {
    'V20 Original': [5, 10, 20, 40],
    'Option A (10/15/25/50)': [10, 15, 25, 50],
    'Option C (9/13/20/40)': [9, 13, 20, 40],
}

# RP options
rp_options = {
    'V20 Original': [250_000, 500_000, 1_000_000, 2_000_000],
    'Option B ($1M/2.5M/5M/10M)': [1_000_000, 2_500_000, 5_000_000, 10_000_000],
    'Option C ($2M/4M/7M/12M)': [2_000_000, 4_000_000, 7_000_000, 12_000_000],
    'Option D ($1.5M/3.5M/7M/15M)': [1_500_000, 3_500_000, 7_000_000, 15_000_000],
    'Option E ($2M/5M/10M/20M)': [2_000_000, 5_000_000, 10_000_000, 20_000_000],
}

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def score_rp(rev, thresholds):
    for i, t in enumerate(thresholds):
        if rev < t:
            return i + 1
    return 5

def rs_score(srv, total):
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

print()
print("=" * 80)
print("ER DISTRIBUTIONS (top 80)")
print("=" * 80)
for name, t in er_options.items():
    dist = Counter(score_dim(e['camp'], t) for e in top)
    print(f"\n{name}:")
    for s in range(1, 6):
        bar = '#' * dist.get(s, 0)
        print(f"  ER={s}: {dist.get(s,0):>3}  {bar}")

print()
print("=" * 80)
print("RP DISTRIBUTIONS (top 80)")
print("=" * 80)
print(f"\nV20 target: RP=1:14  RP=2:16  RP=3:14  RP=4:8  RP=5:18")
for name, t in rp_options.items():
    dist = Counter(score_rp(e['rev'], t) for e in top)
    labels = [f"<${t[0]/1e6:.1f}M", f"${t[0]/1e6:.1f}-{t[1]/1e6:.1f}M",
              f"${t[1]/1e6:.1f}-{t[2]/1e6:.1f}M", f"${t[2]/1e6:.1f}-{t[3]/1e6:.1f}M",
              f">${t[3]/1e6:.1f}M"]
    print(f"\n{name}:")
    for s in range(1, 6):
        bar = '#' * dist.get(s, 0)
        print(f"  RP={s} ({labels[s-1]:>15}): {dist.get(s,0):>3}  {bar}")

# Tier simulations with real RS, placeholder IR/SI/AI
DEFAULT_IR = 2
DEFAULT_SI = 2
DEFAULT_AI = 0

print()
print("=" * 80)
print("TIER DISTRIBUTION (top 80, placeholder qualitative scores)")
print("=" * 80)
print(f"V20: T1=22 (31%), T2=36 (51%), T3=12 (17%)")
print()

combos = [
    ('V20 Original', 'V20 Original'),
    ('Option C (9/13/20/40)', 'Option B ($1M/2.5M/5M/10M)'),
    ('Option C (9/13/20/40)', 'Option C ($2M/4M/7M/12M)'),
    ('Option C (9/13/20/40)', 'Option D ($1.5M/3.5M/7M/15M)'),
    ('Option C (9/13/20/40)', 'Option E ($2M/5M/10M/20M)'),
    ('Option A (10/15/25/50)', 'Option C ($2M/4M/7M/12M)'),
    ('Option A (10/15/25/50)', 'Option E ($2M/5M/10M/20M)'),
]

for er_name, rp_name in combos:
    er_t = er_options[er_name]
    rp_t = rp_options[rp_name]
    tier_dist = Counter()
    for e in top:
        er = score_dim(e['camp'], er_t)
        rp = score_rp(e['rev'], rp_t)
        rs = rs_score(e['srv'], e['camp'])
        score = (er * 4) + (DEFAULT_IR * 3) + (DEFAULT_SI * 3) + (rp * 4) + (rs * 3) + (DEFAULT_AI * 3)
        if score >= 50:
            tier_dist['T1'] += 1
        elif score >= 25:
            tier_dist['T2'] += 1
        else:
            tier_dist['T3'] += 1

    t1, t2, t3 = tier_dist.get('T1',0), tier_dist.get('T2',0), tier_dist.get('T3',0)
    print(f"ER={er_name:<25} RP={rp_name:<30} T1={t1:>3} ({t1/80*100:4.0f}%)  T2={t2:>3} ({t2/80*100:4.0f}%)  T3={t3:>3} ({t3/80*100:4.0f}%)")
