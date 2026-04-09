"""Test aggressive street-number matching: collapse addresses that share
the same street number + city for the same entity.
Compare against the standard abbreviation normalization to see what's new."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15

T5_EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
}

FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Triple Crown': {'TRIPLE CROWN'},
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE', 'WARM HEARTH'},
    'Brighton': {'BRIGHTON'},
    'Momentus Health': {'MOMENTUS HEALTH'},
}

ER_THRESHOLDS = [9, 13, 20, 40]

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def normalize_address(addr):
    """Standard abbreviation normalization."""
    s = addr.upper().strip().rstrip('.')
    s = re.sub(r'\bSTREET\b', 'ST', s)
    s = re.sub(r'\bDRIVE\b', 'DR', s)
    s = re.sub(r'\bROAD\b', 'RD', s)
    s = re.sub(r'\bAVENUE\b', 'AVE', s)
    s = re.sub(r'\bBOULEVARD\b', 'BLVD', s)
    s = re.sub(r'\bLANE\b', 'LN', s)
    s = re.sub(r'\bCOURT\b', 'CT', s)
    s = re.sub(r'\bCIRCLE\b', 'CIR', s)
    s = re.sub(r'\bPLACE\b', 'PL', s)
    s = re.sub(r'\bTERRACE\b', 'TER', s)
    s = re.sub(r'\bPARKWAY\b', 'PKWY', s)
    s = re.sub(r'\bHIGHWAY\b', 'HWY', s)
    s = re.sub(r'\bNORTH\b', 'N', s)
    s = re.sub(r'\bSOUTH\b', 'S', s)
    s = re.sub(r'\bEAST\b', 'E', s)
    s = re.sub(r'\bWEST\b', 'W', s)
    s = re.sub(r'\.', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def extract_street_number(addr):
    """Pull leading street number from address."""
    m = re.match(r'^(\d+)', addr.strip())
    return m.group(1) if m else None

# Load
print("Loading DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

wb_muo = openpyxl.load_workbook("C:/Users/ratwood/Downloads/MUO Data (1).xlsx", data_only=True)
ws_a = wb_muo['Analysis']
finance_names = set()
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in T5_EXCLUDE:
        finance_names.add(str(name).strip())

all_db_names = set()
db_to_finance = {}
for fname, db_names in FINANCE_TO_DB.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_finance[dn] = fname

# Collect
entity_rows = defaultdict(list)
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue
    fname = db_to_finance[corp]
    if fname not in finance_names:
        continue

    beds = ws_db.cell(r, headers['Total_Beds']).value or 0
    beds = beds if isinstance(beds, (int, float)) else 0
    addr_raw = str(ws_db.cell(r, headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, headers['City']).value or '').strip().upper()
    fac_name = str(ws_db.cell(r, headers['Facility_Name']).value or '').strip()
    tot_rev = ws_db.cell(r, headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0

    norm_addr = normalize_address(addr_raw)
    street_num = extract_street_number(addr_raw)

    entity_rows[fname].append({
        'fac_name': fac_name,
        'addr_raw': addr_raw,
        'norm_addr': norm_addr,
        'street_num': street_num,
        'city': city,
        'beds': beds,
        'tot_rev': tot_rev,
        'norm_key': (norm_addr, city),
        'num_key': (street_num, city) if street_num else (norm_addr, city),
    })

# For street-number matching: within each entity, group by (street_num, city)
# and collapse. If two addresses share the same number+city for the same entity,
# treat them as one campus.
def get_num_campuses(rows):
    """Collapse by street number + city within an entity."""
    campuses = set()
    for r in rows:
        campuses.add(r['num_key'])
    return campuses

def get_norm_campuses(rows):
    """Collapse by normalized address + city."""
    return set(r['norm_key'] for r in rows)

# Compare the two approaches
print(f"\n{'='*140}")
print(f"OPTION 2: STREET-NUMBER MATCHING vs STANDARD NORMALIZATION (both with 15-bed rule)")
print(f"{'='*140}")
print(f"\nShowing only entities where street-number matching produces DIFFERENT results than standard normalization:\n")
print(f"{'Entity':<35} {'Norm':>4} {'NumM':>4} {'Diff':>4}  {'ER_n':>4} {'ER_nm':>5}  {'Gate_n':>6} {'Gate_nm':>7}  Additional collapses")
print(f"{'-'*140}")

extra_collapses_all = []

for fname in sorted(entity_rows.keys()):
    rows = entity_rows[fname]
    kept = [r for r in rows if r['beds'] > BED_MIN]
    if not kept:
        continue

    norm_camps = get_norm_campuses(kept)
    num_camps = get_num_campuses(kept)
    n_norm = len(norm_camps)
    n_num = len(num_camps)

    if n_norm == n_num:
        continue

    er_norm = score_dim(n_norm, ER_THRESHOLDS)
    er_num = score_dim(n_num, ER_THRESHOLDS)
    gate_norm = 'PASS' if n_norm >= 7 else 'T4'
    gate_num = 'PASS' if n_num >= 7 else 'T4'

    # Find which addresses collapsed additionally
    # Group kept rows by num_key and find groups with multiple different norm_keys
    num_groups = defaultdict(set)
    num_group_addrs = defaultdict(list)
    for r in kept:
        num_groups[r['num_key']].add(r['norm_key'])
        num_group_addrs[r['num_key']].append((r['norm_addr'], r['fac_name'], r['beds']))

    extra = []
    for nk, norm_keys in num_groups.items():
        if len(norm_keys) > 1:
            addrs_in_group = num_group_addrs[nk]
            unique_addrs = sorted(set(a[0] for a in addrs_in_group))
            extra.append(f"  #{nk[0]} {nk[1]}: {' | '.join(unique_addrs)}")
            extra_collapses_all.append((fname, nk, unique_addrs, addrs_in_group))

    gate_marker = ' ***' if gate_norm != gate_num else ''
    er_marker = ' *' if er_norm != er_num and gate_norm == gate_num else ''

    print(f"{fname:<35} {n_norm:>4} {n_num:>4} {n_norm-n_num:>+4}  {er_norm:>4} {er_num:>5}  {gate_norm:>6} {gate_num:>7}{gate_marker}{er_marker}")
    for e in extra:
        print(f"    {e}")

# Full detail of all extra collapses
print(f"\n{'='*140}")
print(f"ALL ADDITIONAL COLLAPSES (street-number matching catches these, standard normalization does not)")
print(f"{'='*140}")
for fname, nk, unique_addrs, addr_details in sorted(extra_collapses_all):
    print(f"\n  {fname} — #{nk[0]} in {nk[1]}:")
    for addr, fac, beds in sorted(addr_details, key=lambda x: x[0]):
        print(f"    {addr:<45} {fac:<50} {beds:>4} beds")

# Gate comparison summary
print(f"\n{'='*140}")
print(f"GATE COMPARISON SUMMARY")
print(f"{'='*140}")

for label, get_camps in [("Standard normalization + bed rule", lambda rows: get_norm_campuses([r for r in rows if r['beds'] > BED_MIN])),
                          ("Street-number matching + bed rule", lambda rows: get_num_campuses([r for r in rows if r['beds'] > BED_MIN]))]:
    t4_list = []
    scored = 0
    for fname in sorted(entity_rows.keys()):
        rows = entity_rows[fname]
        camps = get_camps(rows)
        if len(camps) >= 7:
            scored += 1
        else:
            t4_list.append((fname, len(camps)))
    print(f"\n{label}:")
    print(f"  Scored (>=7 campuses): {scored}")
    print(f"  T4 (<7 campuses): {len(t4_list)}")
    for fname, nc in t4_list:
        print(f"    {fname:<35} {nc} campuses")
