"""
Q1 Sale Type Re-Verification against V25.9
Checks the CORPORATE ENTITY health for each operator involved in the 23-facility reconciliation.
Compares V25.6 portfolio vs V25.9 portfolio to see what DB cleanup fixed.
"""
import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

# ---- Load V25.6 ----
V26_PATH = r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_6.xlsx"
V29_PATH = r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_9.xlsx"

def load_db(path, label):
    print(f"Loading {label}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c or "").strip() for c in row]
            continue
        rows.append(row)
    wb.close()

    def fc(name):
        for i, h in enumerate(header):
            if name.lower() == h.lower():
                return i
        for i, h in enumerate(header):
            if name.lower() in h.lower():
                return i
        return None

    return rows, fc("Facility_Name"), fc("Corporate_Name"), fc("Do_We_Serve"), fc("City"), fc("State"), fc("Ownership_Type"), len(rows)

v26_rows, c_fac6, c_corp6, c_serve6, c_city6, c_state6, c_own6, n26 = load_db(V26_PATH, "V25.6")
v29_rows, c_fac9, c_corp9, c_serve9, c_city9, c_state9, c_own9, n29 = load_db(V29_PATH, "V25.9")

def build_portfolio(rows, c_corp, c_serve, c_fac, c_city, c_state):
    portfolio = defaultdict(lambda: {"total": 0, "served": 0, "served_facs": [], "all_names": set()})
    for row in rows:
        corp = (str(row[c_corp] or "")).strip().upper()
        serve = str(row[c_serve] or "").strip()
        fac = str(row[c_fac] or "").strip()
        city = str(row[c_city] or "").strip()
        state = str(row[c_state] or "").strip()
        if corp:
            portfolio[corp]["total"] += 1
            if serve == "Yes":
                portfolio[corp]["served"] += 1
                portfolio[corp]["served_facs"].append(f"{fac} ({city}, {state})")
    return portfolio

print("Building portfolios...")
p26 = build_portfolio(v26_rows, c_corp6, c_serve6, c_fac6, c_city6, c_state6)
p29 = build_portfolio(v29_rows, c_corp9, c_serve9, c_fac9, c_city9, c_state9)

# Also search for specific facility matches in V25.9
def find_facility_v27(terms, city, state):
    """Find a specific facility in V25.9 by name terms + city + state."""
    candidates = []
    for row in v29_rows:
        row_state = str(row[c_state9] or "").strip().upper()
        if row_state != state.upper():
            continue
        row_fac = str(row[c_fac9] or "").strip().lower()
        row_city = str(row[c_city9] or "").strip().upper()
        row_corp = str(row[c_corp9] or "").strip()
        row_serve = str(row[c_serve9] or "").strip()

        for term in terms:
            if term.lower() in row_fac:
                candidates.append({
                    "fac": str(row[c_fac9] or "").strip(),
                    "city": row_city,
                    "corp": row_corp,
                    "serve": row_serve,
                    "city_match": city.upper() in row_city or row_city in city.upper()
                })
                break

    # Prefer city match
    city_hits = [c for c in candidates if c["city_match"]]
    if city_hits:
        return city_hits[0]
    if candidates:
        return candidates[0]
    return None


# ---- The 23 facilities and their corporate entities ----
# Structure: (display_name, city, state, is_est, v26_verified_corp, v26_class, search_terms_for_db)
facilities = [
    ("Richland Bean-Blossom HC Center", "SPENCER", "IN", False, "PUTNAM COUNTY HOSPITAL", "New Door",
     ["richland", "bean-blossom", "bean blossom"]),
    ("Pinewood Nursing Care", "SUMTER", "SC", False, "OAK HOLLOW HEALTHCARE MANAGEMENT", "New Logo",
     ["pinewood"]),
    ("McCoy Memorial Nursing Center", "BISHOPVILLE", "SC", False, "CARLYLE SENIOR CARE", "New Logo",
     ["mccoy"]),
    ("Seven Acres Senior Living", "HAMILTON", "OH", False, "INDEPENDENT", "New Logo",
     ["seven acres"]),
    ("Signature HC of Terre Haute", "TERRE HAUTE", "IN", False, "SIGNATURE HEALTH", "New Door",
     ["signature", "terre haute"]),
    ("Rehabilitation Center of Cheraw", "CHERAW", "SC", False, "FUNDAMENTAL LTC", "New Door",
     ["cheraw"]),
    ("Elderwood Rehab & Nursing", "GEORGETOWN", "SC", False, "OAK HOLLOW HEALTHCARE MANAGEMENT", "New Door",
     ["elderwood"]),
    ("Green Meadows Health & Rehab", "FLORENCE", "KY", False, "LYON HEALTHCARE", "New Door",
     ["green meadow"]),
    ("Trinity Grove (Addl Consents)", "WILMINGTON", "NC", True, "SABER HEALTHCARE GROUP", "Integration",
     ["trinity grove"]),
    ("PruittHealth - Raleigh", "RALEIGH", "NC", True, "PRUITT HEALTH", "Integration",
     ["pruitt", "raleigh"]),
    ("Elkhorn Health & Rehab", "FRANKFORT", "KY", False, "LYON HEALTHCARE", "New Door",
     ["elkhorn"]),
    ("Gabriel Manor", "FAYETTEVILLE", "NC", True, "SABER HEALTHCARE GROUP", "Integration",
     ["gabriel"]),
    ("Belmont Terrace", "FLORENCE", "KY", True, "BLUEGRASS/ENCORE", "Integration",
     ["belmont"]),
    ("JAG Healthcare Mansfield", "MANSFIELD", "OH", False, "JAG", "New Door",
     ["jag", "mansfield"]),
    ("Trinity Ridge PCP", "HICKORY", "NC", False, "LUTHERAN SERVICES CAROLINAS", "New Door",
     ["trinity ridge"]),
    ("Lutheran Life Villages Kendallville", "KENDALLVILLE", "IN", False, "LUTHERAN LIFE VILLAGES", "New Door",
     ["lutheran life", "kendallville"]),
    ("Brookdale Harrisonburg", "HARRISONBURG", "VA", False, "BROOKDALE SENIOR LIVING", "New Door",
     ["brookdale", "harrisonburg"]),
    ("TerraBella Little Avenue", "CHARLOTTE", "NC", False, "TERRABELLA SENIOR LIVING", "New Door",
     ["terrabella", "little"]),
    ("Marquette Manor Memory Care", "INDIANAPOLIS", "IN", True, "LIFECARE", "Integration",
     ["marquette"]),
    ("Franklin Manor", "FRANKLIN", "NC", True, "SABER HEALTHCARE GROUP", "Integration",
     ["franklin manor"]),
    ("Brookdale Lakeview Crossing", "COLUMBUS", "OH", False, "BROOKDALE SENIOR LIVING", "New Door",
     ["brookdale", "lakeview"]),
    ("Elliott Nursing & Rehab", "SANDY HOOK", "KY", True, "MAJESTIC CARE", "Integration",
     ["elliott"]),
    ("Cardinal Landing Memory Care", "FORT MITCHELL", "KY", True, "TRIPLE CROWN SENIOR LIVING", "Integration",
     ["cardinal landing"]),
]

print("\n" + "=" * 140)
print("Q1 2026 SALE TYPE RE-VERIFICATION — CORPORATE ENTITY ANALYSIS")
print("V25.6 (reconciliation baseline) vs V25.9 (current DB)")
print("=" * 140)

# Collect unique corporate entities
corp_entities = set()
for f in facilities:
    if f[4] not in ("INDEPENDENT",):
        corp_entities.add(f[4])

# ---- SECTION 1: Corporate Entity Portfolio Comparison ----
print(f"\n{'='*140}")
print("SECTION 1: CORPORATE ENTITY PORTFOLIO HEALTH")
print(f"{'='*140}")

for corp in sorted(corp_entities):
    e26 = p26.get(corp, {"total": 0, "served": 0})
    e27 = p29.get(corp, {"total": 0, "served": 0})

    # Check for variant names in V25.9
    # (entity may have been renamed/consolidated)
    variant_note = ""
    if e27["total"] == 0 and e26["total"] > 0:
        # Entity disappeared — likely renamed
        variant_note = " *** ENTITY NOT FOUND IN V25.9 — likely renamed/consolidated ***"

    total_changed = e26["total"] != e27["total"]
    served_changed = e26["served"] != e27["served"]

    if total_changed or served_changed or variant_note:
        marker = " ***"
    else:
        marker = ""

    print(f"\n  {corp}{marker}")
    print(f"    V25.6: {e26['total']:>4} total, {e26['served']:>4} served")
    print(f"    V25.9: {e27['total']:>4} total, {e27['served']:>4} served")
    if total_changed:
        delta = e27["total"] - e26["total"]
        print(f"    Total delta: {'+' if delta > 0 else ''}{delta}")
    if served_changed:
        delta = e27["served"] - e26["served"]
        print(f"    Served delta: {'+' if delta > 0 else ''}{delta}")
    if variant_note:
        print(f"    {variant_note}")

    # Also check for similar names in V25.9 if entity vanished
    if e27["total"] == 0 and e26["total"] > 0:
        key_word = corp.split()[0] if corp.split() else corp
        similar = [(k, v) for k, v in p29.items() if key_word in k and v["total"] > 0]
        if similar:
            print(f"    Possible V25.9 variants:")
            for name, data in sorted(similar, key=lambda x: -x[1]["total"])[:5]:
                print(f"      {name}: {data['total']} total, {data['served']} served")

# ---- SECTION 2: Facility-Level Lookup ----
print(f"\n{'='*140}")
print("SECTION 2: FACILITY-LEVEL DB MATCH (V25.9)")
print(f"{'='*140}")

oak_hollow_seen = False
results = []

for display, city, state, is_est, v26_corp, v26_class, terms in facilities:
    match = find_facility_v27(terms, city, state)

    # Determine V25.9 classification
    if is_est:
        # EST items are Integration regardless
        v29_class = "Integration"
        if match:
            v29_corp = match["corp"].upper()
        else:
            v29_corp = v26_corp
    elif match:
        v29_corp = match["corp"].upper() if match["corp"] else ""
        v29_serve = match["serve"]

        if v29_serve == "Yes":
            v29_class = "Integration"
        elif not v29_corp or v29_corp in ("INDEPENDENT", ""):
            v29_class = "New Logo"
        elif v29_corp == "OAK HOLLOW HEALTHCARE MANAGEMENT":
            if not oak_hollow_seen:
                v29_class = "New Logo"
                oak_hollow_seen = True
            else:
                v29_class = "New Door"
        else:
            port = p29.get(v29_corp, {"total": 0, "served": 0})
            v29_class = "New Door" if port["served"] > 0 else "New Logo"
    else:
        # Not found in DB — use corporate portfolio check
        v29_corp = v26_corp
        if v26_corp == "INDEPENDENT":
            v29_class = "New Logo"
        else:
            port = p29.get(v26_corp, {"total": 0, "served": 0})
            v29_class = "New Door" if port["served"] > 0 else "New Logo"

    port = p29.get(v29_corp, {"total": 0, "served": 0}) if v29_corp and v29_corp != "INDEPENDENT" else {"total": 0, "served": 0}

    corp_changed = v29_corp != v26_corp.upper()
    class_changed = v29_class != v26_class

    row_data = {
        "display": display,
        "city": city,
        "state": state,
        "is_est": is_est,
        "v26_corp": v26_corp,
        "v29_corp": v29_corp or "(Independent)",
        "v26_class": v26_class,
        "v29_class": v29_class,
        "corp_changed": corp_changed,
        "class_changed": class_changed,
        "match_fac": match["fac"] if match else "NOT FOUND",
        "match_city": match["city"] if match else "",
        "match_serve": match["serve"] if match else "",
        "port_total": port["total"],
        "port_served": port["served"],
    }
    results.append(row_data)

    status = ""
    if corp_changed:
        status += " [CORP CHANGED]"
    if class_changed:
        status += " [CLASS CHANGED]"
    if not match:
        status += " [NO DB MATCH]"

    print(f"\n  {display} ({city}, {state}){status}")
    print(f"    DB Match: {row_data['match_fac']} ({row_data['match_city']})" if match else f"    DB Match: NOT FOUND IN {state}")
    print(f"    V25.6 Corp: {v26_corp}")
    if corp_changed:
        print(f"    V25.9 Corp: {row_data['v29_corp']}  *** CHANGED ***")
    else:
        print(f"    V25.9 Corp: {row_data['v29_corp']}")
    print(f"    Portfolio:  {port['total']} total, {port['served']} served | Do_We_Serve: {row_data['match_serve'] or 'N/A'}")
    print(f"    V25.6 Class: {v26_class}  |  V25.9 Class: {v29_class}", end="")
    if class_changed:
        print("  *** CHANGED ***")
    else:
        print()

# ---- SECTION 3: Summary ----
print(f"\n{'='*140}")
print("SECTION 3: CLASSIFICATION SUMMARY")
print(f"{'='*140}")

print(f"\n  {'Facility':<45} {'V25.6 Corp':<35} {'V25.9 Corp':<35} {'V26 Class':<12} {'V27 Class':<12} {'Delta'}")
print(f"  {'-'*45} {'-'*35} {'-'*35} {'-'*12} {'-'*12} {'-'*8}")
for r in results:
    delta = "CHANGED" if r["class_changed"] else ""
    v27c = r["v29_corp"][:33] if len(r["v29_corp"]) > 33 else r["v29_corp"]
    v26c = r["v26_corp"][:33] if len(r["v26_corp"]) > 33 else r["v26_corp"]
    print(f"  {r['display']:<45} {v26c:<35} {v27c:<35} {r['v26_class']:<12} {r['v29_class']:<12} {delta}")

print(f"\n  Totals:")
print(f"  {'Sale Type':<15} {'V25.6':>6} {'V25.9':>6} {'Delta':>6}")
for st in ["New Logo", "New Door", "Integration"]:
    c26 = sum(1 for r in results if r["v26_class"] == st)
    c27 = sum(1 for r in results if r["v29_class"] == st)
    delta = c27 - c26
    print(f"  {st:<15} {c26:>6} {c27:>6} {'+' + str(delta) if delta > 0 else str(delta):>6}")

print(f"\n  Classification changes: {sum(1 for r in results if r['class_changed'])} of {len(results)}")
print(f"  Corporate name changes: {sum(1 for r in results if r['corp_changed'])} of {len(results)}")
