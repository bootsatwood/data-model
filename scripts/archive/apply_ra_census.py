#!/usr/bin/env python3
"""Apply RA census corrections from Brookdale_Dedup_Review_RA.xlsx to V22.1."""

import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
sys.path.insert(0, ".")
from utils import safe, load_db, VAULT
from openpyxl import load_workbook

DB = VAULT / "Current" / "1_Combined_Database_FINAL_V22_1.xlsx"

# Build the update map from RA file
ra_wb = load_workbook("audit_reports/Brookdale_Dedup_Review_RA.xlsx",
                      read_only=True, data_only=True)
ra_ws = ra_wb.active
ra_headers = [safe(c.value) for c in next(ra_ws.iter_rows(min_row=1, max_row=1))]
ci_action = ra_headers.index("Action")
ci_fac = ra_headers.index("Facility_Name")
ci_city = ra_headers.index("City")
ci_state = ra_headers.index("State")
ci_census = ra_headers.index("Census")

ra_keep = []
for row in ra_ws.iter_rows(min_row=2, values_only=True):
    if safe(row[ci_action]) == "KEEP":
        ra_keep.append({
            "facility": safe(row[ci_fac]),
            "city": safe(row[ci_city]),
            "state": safe(row[ci_state]),
            "census": row[ci_census],
        })
ra_wb.close()
print(f"RA KEEP rows loaded: {len(ra_keep)}")

# Load V22.1 read-only to find matching rows
_, v22_rows = load_db(DB)
match_map = {}
for ra in ra_keep:
    for r in v22_rows:
        if (safe(r.get("Facility_Name", "")) == ra["facility"]
                and safe(r.get("City", "")) == ra["city"]
                and safe(r.get("State", "")) == ra["state"]
                and safe(r.get("Corporate_Name", "")).upper().strip()
                    == "BROOKDALE SENIOR LIVING"):
            old_cens = r.get("Census", "")
            if str(old_cens) != str(ra["census"]):
                match_map[r["_excel_row"]] = ra["census"]
            break

print(f"Census values that differ: {len(match_map)}")

# Open V22.1 in write mode and update
wb = load_workbook(DB)
ws = wb.active
ws_headers = [safe(c.value) for c in ws[1]]
ci_census_ws = ws_headers.index("Census") + 1
ci_fac_ws = ws_headers.index("Facility_Name") + 1

updated = 0
for excel_row, new_census in sorted(match_map.items()):
    old_val = ws.cell(row=excel_row, column=ci_census_ws).value
    ws.cell(row=excel_row, column=ci_census_ws).value = new_census
    fac = safe(ws.cell(row=excel_row, column=ci_fac_ws).value)
    print(f"  Row {excel_row}: {fac} | Census: {old_val} -> {new_census}")
    updated += 1

wb.save(DB)
wb.close()
print(f"\nApplied {updated} census corrections to V22.1.")
