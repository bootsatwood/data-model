"""
GLR Export Coverage Analysis
Compare GLR Export (from MUO Data.xlsx) against V22.2 Combined Database.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, norm, norm_addr

# ── State name -> abbreviation mapping ──
STATE_ABBREV = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
    'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
    'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
    'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
    'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
    'wisconsin': 'WI', 'wyoming': 'WY', 'district of columbia': 'DC',
}


def glr_state_to_abbrev(s):
    """Convert GLR full state name to 2-letter abbreviation."""
    s = safe(s).strip().lower()
    if s in STATE_ABBREV:
        return STATE_ABBREV[s]
    if len(s) == 2:
        return s.upper()
    return s.upper()


def db_key(r):
    addr = safe(r.get('Address', ''))
    city = safe(r.get('City', ''))
    state = safe(r.get('State', ''))
    return norm_addr(addr) + '|' + norm(city) + '|' + norm(state)


# ── Load GLR Export ──
print("Loading GLR Export...")
glr_path = Path.home() / 'Downloads' / 'MUO Data.xlsx'
wb = load_workbook(glr_path, read_only=True, data_only=True)
ws = wb['GLR Export']
glr_iter = ws.iter_rows(values_only=True)
glr_headers = [safe(c) for c in next(glr_iter)]
glr_rows = []
for vals in glr_iter:
    row = dict(zip(glr_headers, vals))
    glr_rows.append(row)
wb.close()
print(f"GLR Export rows: {len(glr_rows)}")

# ── Build GLR address index ──
glr_by_addr = {}
glr_state_raw = set()
for r in glr_rows:
    addr = safe(r.get('AddressLine1', ''))
    city = safe(r.get('City', ''))
    state_raw = safe(r.get('StateCode', ''))
    glr_state_raw.add(state_raw)
    state_abbrev = glr_state_to_abbrev(state_raw)
    key = norm_addr(addr) + '|' + norm(city) + '|' + norm(state_abbrev)
    if key not in glr_by_addr:
        glr_by_addr[key] = []
    glr_by_addr[key].append(r)

print(f"Unique GLR address keys: {len(glr_by_addr)}")
print(f"GLR raw state values: {sorted(glr_state_raw)}")

# ── Load DB ──
print("\nLoading V22.2 database...")
headers, db_rows = load_db()
print(f"DB total rows: {len(db_rows)}")

served = [r for r in db_rows if safe(r.get('Do_We_Serve', '')).lower() == 'yes']
not_served = [r for r in db_rows if safe(r.get('Do_We_Serve', '')).lower() != 'yes']
print(f"DB served (Do_We_Serve=Yes): {len(served)}")
print(f"DB not-served: {len(not_served)}")

# ── Match served DB -> GLR ──
served_matched = []
served_unmatched = []
glr_keys_hit = set()

for r in served:
    k = db_key(r)
    if k in glr_by_addr:
        served_matched.append((r, glr_by_addr[k]))
        glr_keys_hit.add(k)
    else:
        served_unmatched.append(r)

# ── Match not-served DB -> GLR (to check if GLR covers non-served) ──
notserved_matched = 0
notserved_matched_examples = []
for r in not_served:
    k = db_key(r)
    if k in glr_by_addr:
        notserved_matched += 1
        glr_keys_hit.add(k)
        if len(notserved_matched_examples) < 5:
            notserved_matched_examples.append(r)

# GLR rows that did not match ANY DB row
glr_unmatched_keys = set(glr_by_addr.keys()) - glr_keys_hit
glr_unmatched_count = sum(len(glr_by_addr[k]) for k in glr_unmatched_keys)

print("\n" + "=" * 70)
print("RESULTS")
print("=" * 70)
print(f"1. Served DB facilities matched in GLR:     {len(served_matched)} / {len(served)}")
print(f"2. GLR rows NOT matching ANY DB row:         {glr_unmatched_count}")
print(f"3. Served DB rows NOT in GLR:                {len(served_unmatched)}")
print(f"4. Non-served DB rows matched in GLR:        {notserved_matched}")
if notserved_matched > 0:
    print(f"   => GLR DOES cover some non-served facilities")
else:
    print(f"   => GLR is strictly served facilities")

# ── Corporate name comparison ──
corp_match = 0
corp_mismatch = 0
corp_mismatch_examples = []
both_blank = 0

for db_row, glr_matches in served_matched:
    db_corp = safe(db_row.get('Corporate_Name', ''))
    glr_corp = safe(glr_matches[0].get('Parent Company', ''))
    if norm(db_corp) == norm(glr_corp):
        if not db_corp and not glr_corp:
            both_blank += 1
        corp_match += 1
    else:
        corp_mismatch += 1
        if len(corp_mismatch_examples) < 20:
            corp_mismatch_examples.append({
                'facility': safe(db_row.get('Facility_Name', '')),
                'db_corp': db_corp,
                'glr_parent': glr_corp,
                'state': safe(db_row.get('State', '')),
            })

print(f"\n5. Matched served: GLR Parent == DB Corp:    {corp_match}  (of which {both_blank} are both blank)")
print(f"6. Matched served: GLR Parent != DB Corp:    {corp_mismatch}")

print(f"\n--- Sample mismatches (GLR Parent != DB Corporate_Name) ---")
for ex in corp_mismatch_examples:
    print(f"  {ex['facility']} ({ex['state']})")
    print(f"    DB Corporate_Name:  {ex['db_corp']}")
    print(f"    GLR Parent Company: {ex['glr_parent']}")

# ── Sample unmatched GLR rows ──
print(f"\n--- Sample GLR rows with NO DB match ({min(15, glr_unmatched_count)} of {glr_unmatched_count}) ---")
count = 0
for k in sorted(glr_unmatched_keys):
    for r in glr_by_addr[k]:
        if count >= 15:
            break
        print(f"  {safe(r.get('Facility Name', ''))} | {safe(r.get('AddressLine1', ''))} | "
              f"{safe(r.get('City', ''))} | {safe(r.get('StateCode', ''))} | "
              f"Status={safe(r.get('Status', ''))}")
        count += 1
    if count >= 15:
        break

# ── Sample unmatched served DB rows ──
print(f"\n--- Sample served DB rows NOT in GLR ({min(15, len(served_unmatched))} of {len(served_unmatched)}) ---")
for r in served_unmatched[:15]:
    print(f"  {safe(r.get('Facility_Name', ''))} | {safe(r.get('Address', ''))} | "
          f"{safe(r.get('City', ''))} | {safe(r.get('State', ''))} | "
          f"Corp={safe(r.get('Corporate_Name', ''))}")

# ── Non-served DB rows that ARE in GLR ──
if notserved_matched_examples:
    print(f"\n--- Sample non-served DB rows that ARE in GLR ({len(notserved_matched_examples)} shown) ---")
    for r in notserved_matched_examples:
        print(f"  {safe(r.get('Facility_Name', ''))} | {safe(r.get('Address', ''))} | "
              f"{safe(r.get('City', ''))} | {safe(r.get('State', ''))} | "
              f"Do_We_Serve={safe(r.get('Do_We_Serve', ''))}")

# ── GLR Status distribution ──
glr_statuses = Counter(safe(r.get('Status', '')) for r in glr_rows)
print(f"\n--- GLR Status distribution ---")
for s, c in glr_statuses.most_common():
    print(f"  {s}: {c}")

# ── GLR Client Status distribution ──
glr_client_statuses = Counter(safe(r.get('Client Status', '')) for r in glr_rows)
print(f"\n--- GLR Client Status distribution ---")
for s, c in glr_client_statuses.most_common():
    print(f"  {s}: {c}")

# ── State breakdown of served-but-not-in-GLR ──
unmatched_by_state = Counter(safe(r.get('State', '')) for r in served_unmatched)
print(f"\n--- Served DB NOT in GLR, by state ---")
for s, c in unmatched_by_state.most_common():
    print(f"  {s}: {c}")
