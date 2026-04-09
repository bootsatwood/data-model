"""
V25.6 — BD Slides Sale Type Reconciliation Fixes
Applied: 2026-03-31
Source: BD Slides review + CMS verification + web research

CONFIRMED CHANGES (no verification needed):
  #19: Brothers Healthcare / CJM Advisors → OAK HOLLOW HEALTHCARE MANAGEMENT (2 rows)
       + Do_We_Serve → Yes (both won Q1 2026)
  #20: BHP/Encore → BLUEGRASS/ENCORE (1 row)
       + Do_We_Serve → Yes (EST, won Feb 2026)

PENDING VERIFICATION (not applied here):
  #21: 9 Brookdale/INDEPENDENT OH facilities — needs operator confirmation before recoding
       Brookdale Lakeview Crossing Do_We_Serve → Yes is confirmed but corp recode is not.

Discovered during: BD Slides sale type reconciliation, 2026-03-31
Logged in: data-model/reference/MUO_Corporate_History.md
"""

import openpyxl
import sys
import os
from datetime import datetime
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

INPUT = r'C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_5.xlsx'
OUTPUT = r'C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_6.xlsx'

# Column indices (1-based)
COL_FACILITY_NAME = 2
COL_CORPORATE_NAME = 3
COL_DO_WE_SERVE = 12
COL_CORP_ATTRIBUTION = 26

# --- Define changes ---

CORPORATE_RENAMES = {
    'Brothers Healthcare / CJM Advisors': 'OAK HOLLOW HEALTHCARE MANAGEMENT',
    'BHP/Encore': 'BLUEGRASS/ENCORE',
}

# Facilities to set Do_We_Serve = Yes (by facility name match)
SERVE_YES_FACILITIES = [
    'OAK HOLLOW OF SUMTER REHABILITATION CENTER',
    'OAK HOLLOW OF GEORGETOWN REHABILITATION CENTER LLC',
    'BELMONT TERRACE NURSING AND REHABILITATION CENTER',
    'BROOKDALE LAKEVIEW CROSSING',  # Won Mar 2026 — corp recode is separate/pending
]

def main():
    if not os.path.exists(INPUT):
        print(f'ERROR: Input file not found: {INPUT}')
        sys.exit(1)

    print(f'Loading {INPUT}...')
    wb = openpyxl.load_workbook(INPUT)
    ws = wb[wb.sheetnames[0]]

    changes = []
    row_count = 0

    for row in ws.iter_rows(min_row=2):
        row_count += 1
        row_num = row[0].row
        fac_name = (row[COL_FACILITY_NAME - 1].value or '').strip()
        corp_name = (row[COL_CORPORATE_NAME - 1].value or '').strip()
        serve = (row[COL_DO_WE_SERVE - 1].value or '').strip()

        changed = False

        # Corporate name renames
        if corp_name in CORPORATE_RENAMES:
            new_corp = CORPORATE_RENAMES[corp_name]
            changes.append(f'Row {row_num}: Corp rename "{corp_name}" → "{new_corp}" ({fac_name})')
            row[COL_CORPORATE_NAME - 1].value = new_corp
            changed = True

        # Do_We_Serve updates
        fac_upper = fac_name.upper()
        for target in SERVE_YES_FACILITIES:
            if fac_upper == target.upper() and serve != 'Yes':
                changes.append(f'Row {row_num}: Do_We_Serve "{serve}" → "Yes" ({fac_name})')
                row[COL_DO_WE_SERVE - 1].value = 'Yes'
                changed = True
                break

    print(f'\nScanned {row_count} rows.')
    print(f'Changes to apply: {len(changes)}')
    print()
    for c in changes:
        print(f'  {c}')

    if not changes:
        print('\nNo changes found. Exiting without saving.')
        sys.exit(0)

    print(f'\nSaving to {OUTPUT}...')
    wb.save(OUTPUT)
    print('Done.')

    # Summary
    print(f'\n=== V25.6 SUMMARY ===')
    print(f'Input:  V25.5 ({INPUT})')
    print(f'Output: V25.6 ({OUTPUT})')
    print(f'Changes: {len(changes)}')
    print(f'  - Corporate renames: {sum(1 for c in changes if "Corp rename" in c)}')
    print(f'  - Do_We_Serve updates: {sum(1 for c in changes if "Do_We_Serve" in c)}')
    print(f'\nPENDING (not applied):')
    print(f'  - #21: 9 Brookdale/INDEPENDENT OH facilities — awaiting operator verification')

if __name__ == '__main__':
    main()
