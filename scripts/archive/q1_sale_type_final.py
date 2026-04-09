"""
Q1 2026 Sale Type Classification — FINAL
Source: Monday.com CRM Pipeline, "Consented On (Expected)" date, "AR with New Assumptions" column
DB: V25.6 (corrected Oak Hollow, BHP/Encore, Brookdale)
"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

# Load V25.6 DB
wb = openpyxl.load_workbook(
    r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_6.xlsx",
    read_only=True, data_only=True
)
ws = wb[wb.sheetnames[0]]

corps_we_serve = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    corp = (row[2] or "").strip()
    serve = row[11]
    if serve == "Yes" and corp:
        corps_we_serve.add(corp.upper())
wb.close()

# Q1 items: (name, ae, month, ar, is_est)
items = [
    # JANUARY (14 items, $1,770,972)
    ("Signature HC of Terre Haute", "Adam Sabie", "Jan", 405307.2, False),
    ("EST - PruittHealth - Raleigh", "Tonya Hoffman", "Jan", 307756.0, True),
    ("Elkhorn Health & Rehabilitation", "Adam Sabie", "Jan", 291876.1, False),
    ("Richland Bean-Blossom HC Center", "Adam Sabie", "Jan", 189878.0, False),
    ("EST - Belmont Terrace", "Adam Sabie", "Jan", 109020.6, True),
    ("Rehabilitation Center of Cheraw", "Chad Vukelich", "Jan", 89639.16, False),
    ("McCoy Memorial Nursing Center", "Chad Vukelich", "Jan", 84793.8, False),
    ("Green Meadows Health and Rehab", "Adam Sabie", "Jan", 54510.3, False),
    ("Pinewood Nursing Care", "Chad Vukelich", "Jan", 54510.3, False),
    ("EST - Signature HC of Terre Haute", "Adam Sabie", "Jan", 42396.9, True),
    ("Elderwood Rehab and Nursing Center", "Chad Vukelich", "Jan", 42396.9, False),
    ("Seven Acres Senior Living", "Adam Sabie", "Jan", 36340.2, False),
    ("ADDITIONAL CONSENTS- Trinity Grove", "Tonya Hoffman", "Jan", 33917.52, True),
    ("EST Gabriel Manor", "Tonya Hoffman", "Jan", 28629.2, True),
    # FEBRUARY (10 items, $1,410,370)
    ("JAG Healthcare Mansfield", "Adam Sabie", "Feb", 199110.68, False),
    ("Trinity Ridge PCP / NO MD", "Chad Vukelich", "Feb", 320066.24, False),
    ("EST - Elliott Nursing and Rehab", "Adam Sabie", "Feb", 236041.4, True),
    ("Lutheran Life Villages at Kendallville", "Jami Patterson", "Feb", 303497.4, False),
    ("Terra Bella- Little Ave", "Chad Vukelich", "Feb", 92586.06, False),
    ("Brookdale Lakeview Crossing", "Adam Sabie", "Feb", 93788.1, False),
    ("EST-Cardinal Landing Memory Care", "Samantha Roark", "Feb", 52104.5, True),
    ("Brookdale Harrisonburg", "Melissa Been", "Feb", 35786.5, False),
    ("EST-Franklin Manor", "Tonya Hoffman", "Feb", 35786.5, True),
    ("EST- Marquette Manor- Memory Care", "Gunner Grider", "Feb", 41683.6, True),
    # MARCH (16 items, $1,096,746)
    ("EST-PruittHealth - Durham", "Tonya Hoffman", "Mar", 246204.8, True),
    ("OAKVIEW NURSING & REHAB CENTER", "Adam Sabie", "Mar", 153878.0, False),
    ("The Cardinal at North Hills SNF", "Tonya Hoffman", "Mar", 90355.46, False),
    ("EST- Waters of Wakarusa", "Jami Patterson", "Mar", 72680.4, True),
    ("Arbors At Fairlawn", "Adam Sabie", "Mar", 72680.4, False),
    ("Chandler Park Assisted Living", "Samantha Roark", "Mar", 70312.8, False),
    ("EST -Grey Stone Health & Rehab", "Jami Patterson", "Mar", 61551.2, True),
    ("EST-Green Leaf Care Center", "Tonya Hoffman", "Mar", 50101.1, True),
    ("EST - JAG Healthcare Shelby", "Jami Patterson", "Mar", 42396.9, True),
    ("Magnolia Springs Florence", "Adam Sabie", "Mar", 41683.6, False),
    ("Legacy Heights Senior Living Cross Sell", "Chad Vukelich", "Mar", 40080.88, True),
    ("ALG Mount Pleasant House Cross Sell", "Chad Vukelich", "Mar", 35786.5, True),
    ("Bickford of Carmel", "Gunner Grider", "Mar", 31492.12, False),
    ("Willowbrook Healthcare Center", "Adam Sabie", "Mar", 30283.5, False),
    ("1019 Belles Place of Crawfordsville", "Gunner Grider", "Mar", 28629.2, False),
    ("Forest Ridge ALF", "Chad Vukelich", "Mar", 28629.2, False),
]

# Corporate owner mapping (resolved during reconciliation)
corp_map = {
    "Signature HC of Terre Haute": "SIGNATURE HEALTH",
    "EST - PruittHealth - Raleigh": "PRUITT HEALTH",
    "Elkhorn Health & Rehabilitation": "A&M HEALTHCARE INVESTMENT LLC",
    "Richland Bean-Blossom HC Center": None,  # Independent
    "EST - Belmont Terrace": "BLUEGRASS/ENCORE",
    "Rehabilitation Center of Cheraw": "FUNDAMENTAL LTC",
    "McCoy Memorial Nursing Center": "CONCIERGE HEALTHCARE",
    "Green Meadows Health and Rehab": "LYON HEALTHCARE",
    "Pinewood Nursing Care": "OAK HOLLOW HEALTHCARE MANAGEMENT",
    "EST - Signature HC of Terre Haute": "SIGNATURE HEALTH",
    "Elderwood Rehab and Nursing Center": "OAK HOLLOW HEALTHCARE MANAGEMENT",
    "Seven Acres Senior Living": None,  # Independent
    "ADDITIONAL CONSENTS- Trinity Grove": "SABER HEALTHCARE GROUP",
    "EST Gabriel Manor": "SABER HEALTHCARE GROUP",
    "JAG Healthcare Mansfield": "JAG",
    "Trinity Ridge PCP / NO MD": "LUTHERAN SERVICES CAROLINAS",
    "EST - Elliott Nursing and Rehab": "MAJESTIC CARE",
    "Lutheran Life Villages at Kendallville": "LUTHERAN LIFE VILLAGES",
    "Terra Bella- Little Ave": "TERRABELLA SENIOR LIVING",
    "Brookdale Lakeview Crossing": "BROOKDALE SENIOR LIVING",
    "EST-Cardinal Landing Memory Care": "TRIPLE CROWN SENIOR LIVING",
    "Brookdale Harrisonburg": "BROOKDALE SENIOR LIVING",
    "EST-Franklin Manor": "SABER HEALTHCARE GROUP",
    "EST- Marquette Manor- Memory Care": "LIFECARE",
    "EST-PruittHealth - Durham": "PRUITT HEALTH",
    "OAKVIEW NURSING & REHAB CENTER": "OAKDALE SENIOR ALLIANCE",
    "The Cardinal at North Hills SNF": "CARDINAL",
    "EST- Waters of Wakarusa": "GREENCROFT",
    "Arbors At Fairlawn": "ARBORS AT OHIO",
    "Chandler Park Assisted Living": "CARDON & ASSOCIATES",
    "EST -Grey Stone Health & Rehab": "MAJESTIC CARE",
    "EST-Green Leaf Care Center": "SABER HEALTHCARE GROUP",
    "EST - JAG Healthcare Shelby": "JAG",
    "Magnolia Springs Florence": "MAGNOLIA SPRINGS HEALTH SYSTEM",
    "Legacy Heights Senior Living Cross Sell": "ALG SENIOR",
    "ALG Mount Pleasant House Cross Sell": "ALG SENIOR",
    "Bickford of Carmel": "BICKFORD",
    "Willowbrook Healthcare Center": "CONTINUING HEALTHCARE SOLUTIONS",
    "1019 Belles Place of Crawfordsville": "1019 HEALTHCARE",
    "Forest Ridge ALF": "NAVION SENIOR SOLUTIONS",
}

# Track Oak Hollow first win
oak_hollow_first = True

results = []
for name, ae, month, ar, is_est in items:
    corp = corp_map.get(name)

    if is_est:
        sale_type = "Integration"
    elif corp is None:
        sale_type = "New Logo"
    elif corp == "OAK HOLLOW HEALTHCARE MANAGEMENT":
        if oak_hollow_first:
            sale_type = "New Logo"
            oak_hollow_first = False
        else:
            sale_type = "New Door"
    elif corp.upper() in corps_we_serve:
        sale_type = "New Door"
    else:
        sale_type = "New Logo"

    results.append((name, ae, month, ar, is_est, sale_type, corp or "Independent"))

# Print by sale type
print("=" * 90)
print("Q1 2026 SALE TYPE CLASSIFICATION — FINAL")
print("Source: Monday.com CRM Pipeline | AR: New Assumptions | DB: V25.6")
print("=" * 90)

for st in ["New Logo", "New Door", "Integration"]:
    group = [r for r in results if r[5] == st]
    group_ar = sum(r[3] for r in group)
    print(f"\n--- {st} ({len(group)} facilities, ${group_ar:,.0f}) ---")
    print(f"  {'Facility':<45} {'AE':<20} {'Month':<5} {'AR':>12} {'Corp'}")
    for r in group:
        print(f"  {r[0]:<45} {r[1]:<20} {r[2]:<5} ${r[3]:>11,.0f} {r[6]}")

# Summary
total_ar = sum(r[3] for r in results)
print("\n" + "=" * 90)
print("SUMMARY")
print("=" * 90)
print(f"\n  {'Sale Type':<20} {'Count':>6} {'AR':>14} {'% of Total':>10}")
print(f"  {'-'*20} {'-'*6} {'-'*14} {'-'*10}")
for st in ["New Logo", "New Door", "Integration"]:
    group = [r for r in results if r[5] == st]
    group_ar = sum(r[3] for r in group)
    pct = group_ar / total_ar * 100
    print(f"  {st:<20} {len(group):>6} ${group_ar:>13,.0f} {pct:>9.1f}%")
print(f"  {'-'*20} {'-'*6} {'-'*14} {'-'*10}")
print(f"  {'TOTAL':<20} {len(results):>6} ${total_ar:>13,.0f} {'100.0%':>10}")

# Monthly breakdown
print(f"\n  {'Month':<10}", end="")
for st in ["New Logo", "New Door", "Integration"]:
    print(f" {st:>14}", end="")
print(f" {'Total':>14}")
print(f"  {'-'*10}", end="")
for _ in range(4):
    print(f" {'-'*14}", end="")
print()
for m in ["Jan", "Feb", "Mar"]:
    print(f"  {m:<10}", end="")
    for st in ["New Logo", "New Door", "Integration"]:
        val = sum(r[3] for r in results if r[2] == m and r[5] == st)
        print(f" ${val:>13,.0f}", end="")
    mtotal = sum(r[3] for r in results if r[2] == m)
    print(f" ${mtotal:>13,.0f}")
