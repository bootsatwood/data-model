"""Test different RP and ER bracket thresholds against MUO universe.
Goal: find brackets that produce a V20-like tier distribution."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0, 'states': set()})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    d['states'].add(state)
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served:
        d['served_camps'].add((addr, city))

# Filter: >=7 campuses, exclude INDEPENDENT
muos = []
for corp, d in corp_fp.items():
    n = len(d['campuses'])
    if n >= 7 and corp != 'INDEPENDENT':
        muos.append({
            'corp': corp,
            'camp': n,
            'srv': len(d['served_camps']),
            'rev': d['tot_pot'],
        })

print(f"MUO universe: {len(muos)} entities (excl INDEPENDENT)")
print()

# --- Define bracket options ---

# ER options
er_options = {
    'V20 Original': [5, 10, 20, 40],           # 1-4, 5-9, 10-19, 20-39, 40+
    'Option A': [10, 15, 25, 50],               # 7-9, 10-14, 15-24, 25-49, 50+
    'Option B': [10, 20, 35, 60],               # 7-9, 10-19, 20-34, 35-59, 60+
    'Option C': [9, 13, 20, 40],                # 7-8, 9-12, 13-19, 20-39, 40+
}

# RP options ($ thresholds between scores 1/2, 2/3, 3/4, 4/5)
rp_options = {
    'V20 Original': [250_000, 500_000, 1_000_000, 2_000_000],
    'Option A (quintile)': [1_500_000, 3_000_000, 4_500_000, 7_500_000],
    'Option B (wider)': [1_000_000, 2_500_000, 5_000_000, 10_000_000],
    'Option C (aggressive)': [2_000_000, 4_000_000, 7_000_000, 12_000_000],
    'Option D (steep)': [1_500_000, 3_500_000, 7_000_000, 15_000_000],
}

def score_er(camp, thresholds):
    # thresholds = [t1, t2, t3, t4] where score = 1 if < t1, 2 if < t2, etc.
    for i, t in enumerate(thresholds):
        if camp < t:
            return i + 1
    return 5

def score_rp(rev, thresholds):
    for i, t in enumerate(thresholds):
        if rev < t:
            return i + 1
    return 5

def rs_score(srv, total):
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

# Default qualitative scores for simulation (use 2 for everything as placeholder)
DEFAULT_IR = 2
DEFAULT_SI = 2
DEFAULT_AI = 0

print("=" * 100)
print("ER BRACKET OPTIONS")
print("=" * 100)
for name, thresholds in er_options.items():
    dist = Counter()
    for e in muos:
        s = score_er(e['camp'], thresholds)
        dist[s] += 1
    labels = [
        f"<{thresholds[0]}",
        f"{thresholds[0]}-{thresholds[1]-1}",
        f"{thresholds[1]}-{thresholds[2]-1}",
        f"{thresholds[2]}-{thresholds[3]-1}",
        f"{thresholds[3]}+",
    ]
    print(f"\n{name}: thresholds {thresholds}")
    for i in range(1, 6):
        bar = '#' * dist.get(i, 0)
        print(f"  ER={i} ({labels[i-1]:>8}): {dist.get(i,0):>4}  {bar}")

print()
print("=" * 100)
print("RP BRACKET OPTIONS")
print("=" * 100)
for name, thresholds in rp_options.items():
    dist = Counter()
    for e in muos:
        s = score_rp(e['rev'], thresholds)
        dist[s] += 1
    labels = [
        f"<${thresholds[0]/1e6:.1f}M",
        f"${thresholds[0]/1e6:.1f}-{thresholds[1]/1e6:.1f}M",
        f"${thresholds[1]/1e6:.1f}-{thresholds[2]/1e6:.1f}M",
        f"${thresholds[2]/1e6:.1f}-{thresholds[3]/1e6:.1f}M",
        f">${thresholds[3]/1e6:.1f}M",
    ]
    print(f"\n{name}: thresholds ${[t/1e6 for t in thresholds]}M")
    for i in range(1, 6):
        bar = '#' * dist.get(i, 0)
        print(f"  RP={i} ({labels[i-1]:>15}): {dist.get(i,0):>4}  {bar}")

# Now test combinations for tier distribution
print()
print("=" * 100)
print("TIER DISTRIBUTION BY BRACKET COMBINATION")
print("=" * 100)
print(f"V20 target: T1=22 (31%), T2=36 (51%), T3=12 (17%)")
print(f"V23 universe: {len(muos)} entities")
print(f"Proportional target: T1~{int(len(muos)*0.31)}, T2~{int(len(muos)*0.51)}, T3~{int(len(muos)*0.17)}")
print()

combos = [
    ('V20 Original', 'V20 Original'),
    ('Option A', 'Option A (quintile)'),
    ('Option A', 'Option B (wider)'),
    ('Option B', 'Option B (wider)'),
    ('Option C', 'Option C (aggressive)'),
    ('Option A', 'Option C (aggressive)'),
    ('Option A', 'Option D (steep)'),
    ('Option C', 'Option D (steep)'),
]

for er_name, rp_name in combos:
    er_t = er_options[er_name]
    rp_t = rp_options[rp_name]
    tier_dist = Counter()
    for e in muos:
        er = score_er(e['camp'], er_t)
        rp = score_rp(e['rev'], rp_t)
        rs = rs_score(e['srv'], e['camp'])
        score = (er * 4) + (DEFAULT_IR * 3) + (DEFAULT_SI * 3) + (rp * 4) + (rs * 3) + (DEFAULT_AI * 3)
        if score >= 50:
            tier_dist['T1'] += 1
        elif score >= 25:
            tier_dist['T2'] += 1
        else:
            tier_dist['T3'] += 1

    t1 = tier_dist.get('T1', 0)
    t2 = tier_dist.get('T2', 0)
    t3 = tier_dist.get('T3', 0)
    total = t1 + t2 + t3
    p1 = t1/total*100 if total else 0
    p2 = t2/total*100 if total else 0
    p3 = t3/total*100 if total else 0

    print(f"ER={er_name:<12} + RP={rp_name:<22}  T1={t1:>3} ({p1:4.0f}%)  T2={t2:>3} ({p2:4.0f}%)  T3={t3:>3} ({p3:4.0f}%)")
