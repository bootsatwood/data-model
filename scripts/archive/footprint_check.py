"""Check MUO entities against 6-state operational footprint.
Compares all-state vs footprint-only campus counts, revenue, and scoring impact."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

# Exact corp name sets for each entity
ENTITIES = {
    'American Healthcare LLC': {'AMERICAN HEALTHCARE, LLC'},
    'MCAP': {'MCAP'},
    'Greencroft': {'GREENCROFT'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'Senior Lifestyle': {'SENIOR LIFESTYLE'},
    'Cedarhurst Senior Living': {'CEDARHURST SENIOR LIVING'},
    'Spring Arbor Management': {'SPRING ARBOR MANAGEMENT'},
    'StoryPoint': {'STORYPOINT'},
    'Sonida Senior Living': {'SONIDA SENIOR LIVING'},
    'Triple Crown Senior Living': {'TRIPLE CROWN SENIOR LIVING', 'CROWN SENIOR LIVING'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Liberty': {'LIBERTY'},
}


def er_score(f):
    if f >= 40: return 5
    if f >= 20: return 4
    if f >= 10: return 3
    if f >= 5: return 2
    return 1

def rp_score(rev):
    if rev >= 2_000_000: return 5
    if rev >= 1_000_000: return 4
    if rev >= 500_000: return 3
    if rev >= 250_000: return 2
    return 1

def bracket(f):
    if f >= 40: return '40+'
    if f >= 20: return '20-39'
    if f >= 10: return '10-19'
    if f >= 5: return '5-9'
    return '1-4'

def rev_bracket(rev):
    if rev >= 2_000_000: return '>$2M'
    if rev >= 1_000_000: return '$1.0-1.99M'
    if rev >= 500_000: return '$500K-$999K'
    if rev >= 250_000: return '$250-499K'
    return '<$250K'

def safe_num(v):
    return v if isinstance(v, (int, float)) else 0


wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

# Read all rows with a corporate name
all_rows = []
for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    if not corp:
        continue
    all_rows.append({
        'corp': corp,
        'state': str(ws.cell(r, headers['State']).value or '').strip(),
        'addr': str(ws.cell(r, headers['Address']).value or '').strip().upper(),
        'city': str(ws.cell(r, headers['City']).value or '').strip().upper(),
        'type': str(ws.cell(r, headers['Source_Type']).value or '').strip(),
        'served': str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES',
        'integrated': str(ws.cell(r, headers['Integrated_Flag']).value or '').strip().upper() == 'YES',
        'pcp': str(ws.cell(r, headers['PCP_Flag']).value or '').strip().upper() == 'YES',
        'mh': str(ws.cell(r, headers['MH_Flag']).value or '').strip().upper() == 'YES',
        'cur_rev': ws.cell(r, headers['Current_Revenue']).value or 0,
        'int_rev': ws.cell(r, headers['Integration_Revenue']).value or 0,
        'nb_rev': ws.cell(r, headers['New_Business_Revenue']).value or 0,
        'tot_pot': ws.cell(r, headers['Total_Potential_Revenue']).value or 0,
    })

# --- Summary table ---
print("=" * 120)
hdr = f"{'Entity':<30} {'Rows':>5} {'FP':>4} {'Camp':>5} {'FP C':>5} {'Srv':>4} {'FP S':>4}  {'All TotPot':>13} {'FP TotPot':>13}  {'FP States'}"
print(hdr)
print("=" * 120)

for label, corp_names in ENTITIES.items():
    rows = [r for r in all_rows if r['corp'] in corp_names]
    fp_rows = [r for r in rows if r['state'] in FOOTPRINT]

    all_camp = set((r['addr'], r['city']) for r in rows)
    fp_camp = set((r['addr'], r['city']) for r in fp_rows)

    all_srv = sum(1 for r in rows if r['served'])
    fp_srv = sum(1 for r in fp_rows if r['served'])

    all_rev = sum(safe_num(r['tot_pot']) for r in rows)
    fp_rev = sum(safe_num(r['tot_pot']) for r in fp_rows)

    sc = Counter(r['state'] for r in rows)
    fp_sc = sorted((s, c) for s, c in sc.items() if s in FOOTPRINT)
    fp_str = ", ".join(f"{s}({c})" for s, c in fp_sc)

    print(f"{label:<30} {len(rows):>5} {len(fp_rows):>4} {len(all_camp):>5} {len(fp_camp):>5} {all_srv:>4} {fp_srv:>4}  ${all_rev:>12,.0f} ${fp_rev:>12,.0f}  {fp_str}")

print("=" * 120)
print()

# --- Detailed per-entity ---
for label, corp_names in ENTITIES.items():
    rows = [r for r in all_rows if r['corp'] in corp_names]
    fp_rows = [r for r in rows if r['state'] in FOOTPRINT]

    all_camp = set((r['addr'], r['city']) for r in rows)
    fp_camp = set((r['addr'], r['city']) for r in fp_rows)

    all_srv = sum(1 for r in rows if r['served'])
    fp_srv = sum(1 for r in fp_rows if r['served'])

    all_rev = sum(safe_num(r['tot_pot']) for r in rows)
    fp_rev = sum(safe_num(r['tot_pot']) for r in fp_rows)

    sc = Counter(r['state'] for r in rows)
    fp_states = {s: c for s, c in sc.items() if s in FOOTPRINT}
    out_states = {s: c for s, c in sc.items() if s not in FOOTPRINT}

    pct_c = len(fp_camp) / len(all_camp) * 100 if all_camp else 0
    pct_r = fp_rev / all_rev * 100 if all_rev else 0

    er_a = er_score(len(all_camp))
    er_f = er_score(len(fp_camp))
    rp_a = rp_score(all_rev)
    rp_f = rp_score(fp_rev)

    er_delta = f"  ** ER CHANGES: {er_a} -> {er_f} (x4 = {(er_a-er_f)*4:+d} pts) **" if er_a != er_f else ""
    rp_delta = f"  ** RP CHANGES: {rp_a} -> {rp_f} (x4 = {(rp_a-rp_f)*4:+d} pts) **" if rp_a != rp_f else ""
    meets_muo = len(fp_camp) >= 7

    print(f"--- {label} ---")
    print(f"  Total: {len(rows)} rows, {len(all_camp)} campuses  |  Footprint: {len(fp_rows)} rows, {len(fp_camp)} campuses ({pct_c:.0f}%)")
    print(f"  Served: {all_srv} total, {fp_srv} in footprint")
    print(f"  Revenue: All ${all_rev:,.0f}  |  FP ${fp_rev:,.0f} ({pct_r:.0f}%)")
    print(f"  FP states: {fp_states}")
    if out_states:
        print(f"  Out-of-FP: {out_states}")
    print(f"  ER: {er_a} [{bracket(len(all_camp))}] -> {er_f} [{bracket(len(fp_camp))}]{er_delta}")
    print(f"  RP: {rp_a} [{rev_bracket(all_rev)}] -> {rp_f} [{rev_bracket(fp_rev)}]{rp_delta}")
    if not meets_muo:
        print(f"  *** FAILS MUO 7-CAMPUS THRESHOLD ({len(fp_camp)} campuses) ***")
    print()
