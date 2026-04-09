"""Check how address normalization (St/Street, Rd/Road, etc.) affects
campus counts across ALL entities, not just Atrium/Brighton."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15

T5_EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
}

FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Triple Crown': {'TRIPLE CROWN'},
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE', 'WARM HEARTH'},
    'Brighton': {'BRIGHTON'},
    'Momentus Health': {'MOMENTUS HEALTH'},
}

ER_THRESHOLDS = [9, 13, 20, 40]

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def normalize_address(addr):
    """Normalize common abbreviations so co-located SNF+ALF collapse properly."""
    s = addr.upper().strip().rstrip('.')
    # Street suffix normalization
    s = re.sub(r'\bSTREET\b', 'ST', s)
    s = re.sub(r'\bDRIVE\b', 'DR', s)
    s = re.sub(r'\bROAD\b', 'RD', s)
    s = re.sub(r'\bAVENUE\b', 'AVE', s)
    s = re.sub(r'\bBOULEVARD\b', 'BLVD', s)
    s = re.sub(r'\bLANE\b', 'LN', s)
    s = re.sub(r'\bCOURT\b', 'CT', s)
    s = re.sub(r'\bCIRCLE\b', 'CIR', s)
    s = re.sub(r'\bPLACE\b', 'PL', s)
    s = re.sub(r'\bTERRACE\b', 'TER', s)
    s = re.sub(r'\bPARKWAY\b', 'PKWY', s)
    s = re.sub(r'\bHIGHWAY\b', 'HWY', s)
    # Direction normalization
    s = re.sub(r'\bNORTH\b', 'N', s)
    s = re.sub(r'\bSOUTH\b', 'S', s)
    s = re.sub(r'\bEAST\b', 'E', s)
    s = re.sub(r'\bWEST\b', 'W', s)
    # Remove trailing periods, extra spaces
    s = re.sub(r'\.', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# Load
print("Loading DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

wb_muo = openpyxl.load_workbook("C:/Users/ratwood/Downloads/MUO Data (1).xlsx", data_only=True)
ws_a = wb_muo['Analysis']
finance_names = set()
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in T5_EXCLUDE:
        finance_names.add(str(name).strip())

all_db_names = set()
db_to_finance = {}
for fname, db_names in FINANCE_TO_DB.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_finance[dn] = fname

# Collect
entity_rows = defaultdict(list)
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue
    fname = db_to_finance[corp]
    if fname not in finance_names:
        continue

    beds = ws_db.cell(r, headers['Total_Beds']).value or 0
    beds = beds if isinstance(beds, (int, float)) else 0
    addr = str(ws_db.cell(r, headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, headers['City']).value or '').strip()
    tot_rev = ws_db.cell(r, headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0

    entity_rows[fname].append({
        'raw_key': (addr.upper(), city.upper()),
        'norm_key': (normalize_address(addr), city.upper()),
        'beds': beds,
        'tot_rev': tot_rev,
    })

# Compare
print(f"\n{'='*120}")
print(f"ADDRESS NORMALIZATION + 15-BED RULE IMPACT")
print(f"{'='*120}")
print(f"{'Entity':<35} {'Raw':>4} {'Norm':>4} {'+Bed':>4}  {'ER_raw':>6} {'ER_nrm':>6} {'ER_both':>7}  {'Gate_raw':>8} {'Gate_both':>9}  Notes")
print(f"{'-'*120}")

changes = []
for fname in sorted(entity_rows.keys()):
    rows = entity_rows[fname]

    # Raw (current V7 behavior)
    raw_camps = set(r['raw_key'] for r in rows)
    n_raw = len(raw_camps)

    # Normalized addresses only
    norm_camps = set(r['norm_key'] for r in rows)
    n_norm = len(norm_camps)

    # Normalized + bed minimum
    kept = [r for r in rows if r['beds'] > BED_MIN]
    both_camps = set(r['norm_key'] for r in kept)
    n_both = len(both_camps)

    er_raw = score_dim(n_raw, ER_THRESHOLDS)
    er_norm = score_dim(n_norm, ER_THRESHOLDS)
    er_both = score_dim(n_both, ER_THRESHOLDS) if n_both > 0 else 0

    gate_raw = 'PASS' if n_raw >= 7 else 'T4'
    gate_both = 'PASS' if n_both >= 7 else 'T4'

    notes = []
    if n_raw != n_norm:
        notes.append(f'addr collapse: {n_raw}->{n_norm} (-{n_raw-n_norm})')
    if n_norm != n_both:
        notes.append(f'bed rule: {n_norm}->{n_both} (-{n_norm-n_both})')
    if er_raw != er_both:
        notes.append(f'ER: {er_raw}->{er_both}')
    if gate_raw != gate_both:
        notes.append(f'GATE: {gate_raw}->{gate_both}')

    has_change = n_raw != n_both
    marker = ' ***' if gate_raw != gate_both else (' *' if er_raw != er_both else '')

    if has_change:
        changes.append(fname)
        print(f"{fname:<35} {n_raw:>4} {n_norm:>4} {n_both:>4}  {er_raw:>6} {er_norm:>6} {er_both:>7}  {gate_raw:>8} {gate_both:>9}  {'; '.join(notes)}{marker}")

# Unchanged
unchanged = [f for f in sorted(entity_rows.keys()) if f not in changes]
print(f"\n{len(unchanged)} entities unchanged (raw campus count = normalized + bed rule count)")

# Summary
print(f"\n{'='*120}")
print(f"SUMMARY")
print(f"{'='*120}")

gate_movers = []
er_movers = []
for fname in sorted(entity_rows.keys()):
    rows = entity_rows[fname]
    raw_camps = set(r['raw_key'] for r in rows)
    kept = [r for r in rows if r['beds'] > BED_MIN]
    both_camps = set(r['norm_key'] for r in kept)
    n_raw = len(raw_camps)
    n_both = len(both_camps)
    gate_raw = 'PASS' if n_raw >= 7 else 'T4'
    gate_both = 'PASS' if n_both >= 7 else 'T4'
    er_raw = score_dim(n_raw, ER_THRESHOLDS)
    er_both = score_dim(n_both, ER_THRESHOLDS) if n_both > 0 else 0
    if gate_raw != gate_both:
        gate_movers.append((fname, n_raw, n_both))
    elif er_raw != er_both:
        er_movers.append((fname, n_raw, n_both, er_raw, er_both))

print(f"\nGATE CHANGES (PASS -> T4):")
for fname, n_raw, n_both in gate_movers:
    print(f"  {fname:<35} {n_raw} raw -> {n_both} qualifying campuses")

print(f"\nER SCORE CHANGES (no gate change):")
for fname, n_raw, n_both, er_raw, er_both in er_movers:
    print(f"  {fname:<35} {n_raw} raw -> {n_both} qualifying  ER: {er_raw} -> {er_both}")

print(f"\nTotal entities affected: {len(changes)}")
print(f"Gate movers: {len(gate_movers)}")
print(f"ER-only movers: {len(er_movers)}")
