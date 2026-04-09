#!/usr/bin/env python3
"""
Phantom Fix — Wave 1 (V21.1 -> V22.0)

Implements all confirmed phantom triage decisions from Facility DB cross-reference.
Three action types applied in dependency order:

  1. UPDATE  — Migrate service flags from ALF->SNF before removal
  2. REMOVE  — Delete confirmed phantom ALF rows
  3. RECLASS — Change Source_Type (ALF->SNF or ALF->ILF)

Two modes:
  preview  — Dry run: show all proposed changes, no file modifications
  apply    — Write changes to new V22.0 database file

Usage:
  python phantom_fix.py preview
  python phantom_fix.py apply
"""

import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
from pathlib import Path
from openpyxl import load_workbook
from utils import safe, VAULT

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V21_1.xlsx"
DB_TARGET = VAULT / "Current" / "1_Combined_Database_FINAL_V22_0.xlsx"

# ---------------------------------------------------------------------------
# Change definitions — from Facility DB cross-reference (2026-03-03)
# ---------------------------------------------------------------------------

# Action 1: UPDATE SNF rows (migrate service flags before removing ALF pair)
SNF_UPDATES = {
    # excel_row: {column: new_value, ...}
    18779: {
        'Corporate_Name': 'CHI LIVING',
        'Do_We_Serve': 'Yes',
        'MH_Flag': 'Yes',
    },
    5200: {
        'Do_We_Serve': 'Yes',
        'MH_Flag': 'Yes',
    },
    20792: {
        'Do_We_Serve': 'Yes',
        'Integrated_Flag': 'Yes',
    },
    17209: {
        'Do_We_Serve': 'Yes',
        'PCP_Flag': 'Yes',
        'Facility_Name': 'GLENWOOD HEALTHCARE CENTER SNF/NF',
    },
}

# Action 2: REMOVE phantom ALF rows
# Served phantoms — Cat A (Pattern 1: SNF in ALF name)
REMOVE_SERVED_PHANTOM = [
    4760,   # Four Seasons Retirement Center - SNF, Columbus IN
    4879,   # Heritage Park SNF, Fort Wayne IN
    5368,   # Wesley Manor SNF, Frankfort IN
]

# Served phantoms — Cat B (service mismatch, flags migrated to SNF above)
REMOVE_SERVED_MIGRATED = [
    4895,   # Hoosier Health & Living, Brownstown IN (flags -> row 18779)
    17212,  # Glenwood Healthcare Center SNF/NF, Princeton WV (flags -> row 17209)
]

# Non-served Group 1 — have matching SNF at same address, redundant
REMOVE_NONSERVED_GROUP1 = [
    5104,   # River Terrace Retirement Community, Bluffton IN
    19748,  # River Terrace Health Care Center, Bluffton IN
    18422,  # Four Seasons Retirement Center, Columbus IN
    4519,   # Adams Woodcrest of Decatur, Decatur IN
    4877,   # Heritage Park, Fort Wayne IN
    20566,  # Towne House Retirement Center, Fort Wayne IN
    5350,   # Waterford Crossing, Goshen IN
    5180,   # St. Charles Health Campus, Jasper IN
    5351,   # Waterford Place Health Campus, Kokomo IN
    4701,   # Creasy Springs Health Campus, Lafayette IN
    5103,   # River Terrace Health Campus, Madison IN
    5096,   # Rawlins House & Fall Creek Retirement Village, Pendleton IN
    4611,   # Blair Ridge Health Campus, Peru IN
    5038,   # Oakwood Health Campus, Tell City IN
    4690,   # Cobblestone Crossings Health Campus, Terre Haute IN
    4882,   # Heritage Pointe of Warren, Warren IN
    5226,   # The Villages at Oak Ridge, Washington IN
    20814,  # Westminster Village, West Lafayette IN
    18528,  # Grace Village, Winona Lake IN
    18496,  # Givens Highland Farms, Black Mountain NC
    19854,  # Scotia Village Retirement Community, Laurinburg NC
    20312,  # The Hillsman House Independent Living, Tryon NC
    18231,  # Deupree House, Cincinnati OH
    18524,  # Grace Brethren Village, Englewood OH
    17821,  # Buffalo Valley Lutheran Village, Lewisburg PA
    20491,  # The Village at Penn State, State College PA
    20454,  # The Seabrook of Hilton Head, Hilton Head Island SC
]

# Non-served Group 2 duplicate — Lima Estates appears twice, remove one
REMOVE_DUPLICATE = [
    19013,  # Lima Estates (duplicate of 19012), Media PA
]

ALL_REMOVALS = (
    REMOVE_SERVED_PHANTOM
    + REMOVE_SERVED_MIGRATED
    + REMOVE_NONSERVED_GROUP1
    + REMOVE_DUPLICATE
)

# Action 3: RECLASSIFY Source_Type
# Group 2 with CMS CCN — ALF -> SNF (these are real nursing facilities)
RECLASS_TO_SNF = {
    # excel_row: {column: new_value}
    17484: {
        'Source_Type': 'SNF',
        # Arrowood at Southwestern, Pittsburgh PA — CMS: Southwestern Nursing & Rehab
    },
    18434: {
        'Source_Type': 'SNF',
        # Frederick Living, Frederick PA — CMS: Frederick Living - Cedarwood
    },
    19012: {
        'Source_Type': 'SNF',
        # Lima Estates, Media PA — CMS: Willowbrooke Court SKD Care Center
        # Note: CMS city is "Lima" not "Media" — address fix deferred to Wave 2
    },
}

# Group 2 without CCN — ALF -> ILF (NIC says NC units but no CMS certification)
RECLASS_TO_ILF = {
    17803: {'Source_Type': 'ILF'},   # Brooks-Howell Home, Asheville NC
    19226: {'Source_Type': 'ILF'},   # Moravian Village of Bethlehem, PA
    20020: {'Source_Type': 'ILF'},   # Springmoor Life Care Retirement, Raleigh NC
    20223: {'Source_Type': 'ILF'},   # The Cedars of Chapel Hill, NC
    20251: {'Source_Type': 'ILF'},   # The Cypress of Raleigh, NC
    20494: {'Source_Type': 'ILF'},   # The Village at St Barnabas, Gibsonia PA
    20880: {'Source_Type': 'ILF'},   # Windsor Run, Matthews NC
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col_index(headers, name):
    """Return 0-based column index for a header name."""
    try:
        return headers.index(name)
    except ValueError:
        raise ValueError(f"Column '{name}' not found. Available: {headers}")


def verify_row(ws, excel_row, headers, expected_facility=None):
    """Read a row and return a dict of its values for verification."""
    row_data = {}
    for ci, hdr in enumerate(headers):
        val = safe(ws.cell(row=excel_row, column=ci + 1).value)
        row_data[hdr] = val
    return row_data


# ---------------------------------------------------------------------------
# Preview mode
# ---------------------------------------------------------------------------

def do_preview():
    print("PHANTOM FIX — PREVIEW MODE (no changes written)")
    print("=" * 70)
    print()
    print(f"Source: {DB_SOURCE.name}")
    print(f"Target: {DB_TARGET.name} (will be created on apply)")
    print()

    # Load all rows into dict by excel_row for random access
    from utils import load_db
    print("Loading database...")
    headers, db_rows = load_db(DB_SOURCE)
    rows_by_excel = {row['_excel_row']: row for row in db_rows}
    print(f"  {len(db_rows):,} rows loaded.")
    print()

    def get_row(excel_row):
        return rows_by_excel.get(excel_row, {})

    def row_label(r):
        return f"{safe(r.get('Facility_Name', ''))} ({safe(r.get('City', ''))}, {safe(r.get('State', ''))})"

    # --- Action 1: SNF Updates ---
    print("-" * 70)
    print(f"ACTION 1: UPDATE {len(SNF_UPDATES)} SNF rows (migrate service flags)")
    print("-" * 70)
    for excel_row, updates in sorted(SNF_UPDATES.items()):
        r = get_row(excel_row)
        print(f"\n  Row {excel_row}: {row_label(r)}")
        for col, new_val in updates.items():
            old_val = safe(r.get(col, ''))
            marker = " <<<" if old_val != new_val else ""
            print(f"    {col}: '{old_val}' -> '{new_val}'{marker}")

    # --- Action 2: Removals ---
    print()
    print("-" * 70)
    print(f"ACTION 2: REMOVE {len(ALL_REMOVALS)} ALF rows")
    print("-" * 70)

    print(f"\n  Served phantoms (Pattern 1 — SNF in name): {len(REMOVE_SERVED_PHANTOM)}")
    for excel_row in REMOVE_SERVED_PHANTOM:
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))} | Serve={safe(r.get('Do_We_Serve', ''))}")

    print(f"\n  Served phantoms (flags migrated to SNF): {len(REMOVE_SERVED_MIGRATED)}")
    for excel_row in REMOVE_SERVED_MIGRATED:
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))} | Serve={safe(r.get('Do_We_Serve', ''))}")

    print(f"\n  Non-served Group 1 (SNF exists at address): {len(REMOVE_NONSERVED_GROUP1)}")
    for excel_row in REMOVE_NONSERVED_GROUP1:
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))}")

    print(f"\n  Duplicate removal: {len(REMOVE_DUPLICATE)}")
    for excel_row in REMOVE_DUPLICATE:
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))}")

    # --- Action 3: Reclassifications ---
    print()
    print("-" * 70)
    print(f"ACTION 3: RECLASSIFY {len(RECLASS_TO_SNF)} ALF->SNF + {len(RECLASS_TO_ILF)} ALF->ILF")
    print("-" * 70)

    print(f"\n  ALF -> SNF (have CMS CCN): {len(RECLASS_TO_SNF)}")
    for excel_row in sorted(RECLASS_TO_SNF.keys()):
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))} | Source_Type: ALF -> SNF")

    print(f"\n  ALF -> ILF (no CMS CCN, NC+IL campus): {len(RECLASS_TO_ILF)}")
    for excel_row in sorted(RECLASS_TO_ILF.keys()):
        r = get_row(excel_row)
        print(f"    Row {excel_row}: {safe(r.get('Facility_Name', ''))} | {safe(r.get('City', ''))}, {safe(r.get('State', ''))} | Source_Type: ALF -> ILF")

    # --- Summary ---
    print()
    print("=" * 70)
    print("PREVIEW SUMMARY")
    print("=" * 70)
    print(f"  SNF updates (flag migration):  {len(SNF_UPDATES)}")
    print(f"  Row removals:                  {len(ALL_REMOVALS)}")
    print(f"    Served phantoms:               {len(REMOVE_SERVED_PHANTOM)}")
    print(f"    Served (flags migrated):       {len(REMOVE_SERVED_MIGRATED)}")
    print(f"    Non-served Group 1:            {len(REMOVE_NONSERVED_GROUP1)}")
    print(f"    Duplicate:                     {len(REMOVE_DUPLICATE)}")
    print(f"  Reclassify ALF->SNF:            {len(RECLASS_TO_SNF)}")
    print(f"  Reclassify ALF->ILF:            {len(RECLASS_TO_ILF)}")
    print(f"  ---")
    total_touched = len(SNF_UPDATES) + len(ALL_REMOVALS) + len(RECLASS_TO_SNF) + len(RECLASS_TO_ILF)
    print(f"  Total rows affected:           {total_touched}")
    print()
    print(f"  Net row count change:          -{len(ALL_REMOVALS)} (26,267 -> {26267 - len(ALL_REMOVALS):,})")
    print()
    print("To apply: python phantom_fix.py apply")


# ---------------------------------------------------------------------------
# Apply mode
# ---------------------------------------------------------------------------

def do_apply():
    print("PHANTOM FIX — APPLY MODE")
    print("=" * 70)
    print()
    print(f"Source: {DB_SOURCE.name}")
    print(f"Target: {DB_TARGET.name}")
    print()

    if DB_TARGET.exists():
        print(f"ERROR: {DB_TARGET.name} already exists. Remove it first to re-apply.")
        sys.exit(1)

    # Load full workbook (not read-only) for modification
    print("Loading workbook (full mode)...")
    wb = load_workbook(DB_SOURCE)
    ws = wb.active
    headers = [safe(c.value) for c in ws[1]]

    # Build column index lookup
    col_idx = {}
    for name in ['Source_Type', 'Facility_Name', 'Corporate_Name', 'Do_We_Serve',
                  'Integrated_Flag', 'PCP_Flag', 'MH_Flag', 'City', 'State']:
        col_idx[name] = headers.index(name) + 1  # openpyxl is 1-indexed

    # --- Action 1: UPDATE SNF rows ---
    print("\nAction 1: Updating SNF rows...")
    for excel_row, updates in sorted(SNF_UPDATES.items()):
        fac = safe(ws.cell(row=excel_row, column=col_idx['Facility_Name']).value)
        for col_name, new_val in updates.items():
            ci = col_idx.get(col_name)
            if ci is None:
                ci = headers.index(col_name) + 1
            old_val = safe(ws.cell(row=excel_row, column=ci).value)
            ws.cell(row=excel_row, column=ci).value = new_val
            print(f"  Row {excel_row} ({fac}): {col_name} '{old_val}' -> '{new_val}'")

    # --- Action 3: RECLASSIFY (before deletion, since row numbers shift) ---
    print("\nAction 3: Reclassifying rows...")
    all_reclass = {}
    all_reclass.update(RECLASS_TO_SNF)
    all_reclass.update(RECLASS_TO_ILF)
    for excel_row, updates in sorted(all_reclass.items()):
        fac = safe(ws.cell(row=excel_row, column=col_idx['Facility_Name']).value)
        for col_name, new_val in updates.items():
            ci = col_idx.get(col_name)
            if ci is None:
                ci = headers.index(col_name) + 1
            old_val = safe(ws.cell(row=excel_row, column=ci).value)
            ws.cell(row=excel_row, column=ci).value = new_val
            print(f"  Row {excel_row} ({fac}): {col_name} '{old_val}' -> '{new_val}'")

    # --- Action 2: REMOVE rows (delete from bottom up to preserve row numbers) ---
    print(f"\nAction 2: Removing {len(ALL_REMOVALS)} rows...")
    removals_sorted = sorted(ALL_REMOVALS, reverse=True)  # Bottom-up deletion
    for excel_row in removals_sorted:
        fac = safe(ws.cell(row=excel_row, column=col_idx['Facility_Name']).value)
        city = safe(ws.cell(row=excel_row, column=col_idx['City']).value)
        state = safe(ws.cell(row=excel_row, column=col_idx['State']).value)
        ws.delete_rows(excel_row, 1)
        print(f"  Deleted row {excel_row}: {fac} ({city}, {state})")

    # --- Save ---
    print(f"\nSaving as {DB_TARGET.name}...")
    wb.save(DB_TARGET)
    wb.close()
    print(f"  Saved to: {DB_TARGET}")

    # --- Summary ---
    print()
    print("=" * 70)
    print("APPLY COMPLETE")
    print("=" * 70)
    print(f"  SNF updates:        {len(SNF_UPDATES)}")
    print(f"  Rows removed:       {len(ALL_REMOVALS)}")
    print(f"  Reclassified:       {len(all_reclass)}")
    print(f"  New row count:      {26267 - len(ALL_REMOVALS):,}")
    print()
    print("Next steps:")
    print("  1. Run qc_validator.py snapshot on V22.0")
    print("  2. Compare against V21.1 baseline")
    print("  3. Verify: facility count = 26,267 - 33 = 26,234")
    print("  4. Verify: SNF count unchanged, ALF count decreased by 33")
    print("  5. Verify: 4 previously-unserved SNFs now show Do_We_Serve=Yes")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python phantom_fix.py [preview|apply]")
        print("  preview  — Show all proposed changes (no files modified)")
        print("  apply    — Write V22.0 database with changes applied")
        sys.exit(1)

    mode = sys.argv[1]
    if mode == 'preview':
        do_preview()
    elif mode == 'apply':
        do_apply()


if __name__ == '__main__':
    main()
