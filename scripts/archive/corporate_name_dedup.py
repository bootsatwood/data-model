#!/usr/bin/env python3
"""
Corporate Name Deduplication Analysis — V22.3

Finds Corporate_Name variants that likely refer to the same entity:
  1. norm_corp exact clusters (suffix/case/punctuation variants)
  2. Containment matches (one name is substring of another)
  3. Token overlap (shared distinctive words)
  4. Edit distance (typos/minor spelling differences)

Produces a review workbook: audit_reports/Corporate_Name_Dedup_Review.xlsx

Usage:
  python corporate_name_dedup.py
"""

import re
import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
from collections import Counter, defaultdict
from itertools import combinations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import safe, load_db, ensure_report_dir, VAULT, GLR_FILE
from corporate_fix import norm_corp, load_glr

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_3.xlsx"

# Common words to ignore in token matching (too generic to be distinctive)
STOP_WORDS = {
    'healthcare', 'health', 'care', 'senior', 'living', 'seniors',
    'management', 'group', 'services', 'of', 'the', 'and', 'at', 'in',
    'nursing', 'rehab', 'rehabilitation', 'center', 'centres', 'home',
    'homes', 'inc', 'llc', 'corp', 'company', 'co', 'ltd', 'lp',
    'incorporated', 'corporation', 'limited', 'associates', 'partners',
    'properties', 'investments', 'holdings', 'ventures', 'solutions',
    'consulting', 'operations', 'facilities', 'community', 'communities',
    'retirement', 'assisted', 'skilled', 'long', 'term',
}

# Minimum token length to consider distinctive
MIN_TOKEN_LEN = 3

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                          fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                       fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2",
                          fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4",
                          fill_type="solid")
GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9",
                         fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_headers(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(cell.value)) for cell in col if cell.value]
        if lengths:
            width = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = width + 2


# ---------------------------------------------------------------------------
# Tokenizer
# ---------------------------------------------------------------------------

def tokenize(name):
    """Split a corporate name into distinctive tokens."""
    if not name:
        return set()
    words = re.findall(r'[a-z]+', name.lower())
    return {w for w in words if len(w) >= MIN_TOKEN_LEN and w not in STOP_WORDS}


# ---------------------------------------------------------------------------
# Levenshtein distance (simple implementation)
# ---------------------------------------------------------------------------

def levenshtein(s1, s2):
    if len(s1) < len(s2):
        return levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            ins = prev[j + 1] + 1
            delete = curr[j] + 1
            sub = prev[j] + (c1 != c2)
            curr.append(min(ins, delete, sub))
        prev = curr
    return prev[-1]


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------

print("Loading V22.3 database...")
_, rows = load_db(DB_SOURCE)
print(f"  {len(rows):,} rows loaded.")

# Load GLR for canonical authority
print("Loading GLR Export for canonical authority...")
glr = load_glr()
glr_parents = set()
glr_parent_by_norm = {}  # norm_corp -> GLR spelling
for entry in glr.values():
    parent = entry['parent_company']
    if parent and parent.upper() not in ('INDEPENDENT', 'N/A', ''):
        glr_parents.add(parent)
        n = norm_corp(parent)
        # Keep the first GLR spelling per norm_corp key
        if n not in glr_parent_by_norm:
            glr_parent_by_norm[n] = parent
print(f"  {len(glr_parents)} distinct GLR Parent Companies loaded.")
print(f"  {len(glr_parent_by_norm)} distinct after normalization.")
print()


def pick_canonical(name_a, count_a, name_b, count_b):
    """Pick canonical name: GLR authority first, then highest count.

    Returns (canonical_name, variant_name, reason).
    """
    norm_a = norm_corp(name_a)
    norm_b = norm_corp(name_b)
    a_in_glr = norm_a in glr_parent_by_norm
    b_in_glr = norm_b in glr_parent_by_norm

    if a_in_glr and not b_in_glr:
        glr_spelling = glr_parent_by_norm[norm_a]
        return glr_spelling, name_b, f'GLR: "{glr_spelling}"'
    if b_in_glr and not a_in_glr:
        glr_spelling = glr_parent_by_norm[norm_b]
        return glr_spelling, name_a, f'GLR: "{glr_spelling}"'
    if a_in_glr and b_in_glr:
        # Both in GLR — use the one with more rows, note conflict
        if count_a >= count_b:
            return name_a, name_b, f'Both in GLR (conflict), using higher count'
        else:
            return name_b, name_a, f'Both in GLR (conflict), using higher count'

    # Neither in GLR — use highest count
    if count_a >= count_b:
        return name_a, name_b, 'Highest count (no GLR match)'
    else:
        return name_b, name_a, 'Highest count (no GLR match)'

# Build corporate name counts
corp_counts = Counter()
corp_served = Counter()
corp_states = defaultdict(set)
corp_types = defaultdict(set)
corp_attr_source = defaultdict(set)

for r in rows:
    corp = safe(r.get('Corporate_Name', ''))
    if not corp or corp.upper() == 'INDEPENDENT':
        continue
    corp_counts[corp] += 1
    if safe(r.get('Do_We_Serve', '')) == 'Yes':
        corp_served[corp] += 1
    corp_states[corp].add(safe(r.get('State', '')))
    corp_types[corp].add(safe(r.get('Source_Type', '')))
    src = safe(r.get('Corp_Attribution_Source', ''))
    if src:
        corp_attr_source[corp].add(src)

distinct_names = list(corp_counts.keys())
print(f"  {len(distinct_names):,} distinct Corporate_Names.")
print()

# ===================================================================
# PASS 1: norm_corp exact clusters
# ===================================================================

print("Pass 1: norm_corp exact clusters...")
norm_groups = defaultdict(list)
for name in distinct_names:
    n = norm_corp(name)
    norm_groups[n].append(name)

# Only keep clusters with 2+ variants
norm_clusters = {n: names for n, names in norm_groups.items() if len(names) > 1}
print(f"  {len(norm_clusters)} clusters with 2+ spelling variants")
total_variants = sum(len(v) for v in norm_clusters.values())
print(f"  {total_variants} total names in clusters")

# Track which names are already clustered (skip in fuzzy passes)
clustered = set()
for names in norm_clusters.values():
    clustered.update(names)

# Remaining singletons for fuzzy matching
singletons = [n for n in distinct_names if n not in clustered]
print(f"  {len(singletons)} singletons remaining for fuzzy matching")
print()

# ===================================================================
# PASS 2: Containment matching on norm_corp values
# ===================================================================

print("Pass 2: Containment matching...")

# Build norm_corp -> name mapping for singletons
singleton_norms = {}
for name in singletons:
    n = norm_corp(name)
    singleton_norms.setdefault(n, []).append(name)

# Deduplicate norm values
unique_norms = sorted(set(singleton_norms.keys()))

containment_matches = []
# Sort by length so we can check if shorter is contained in longer
by_length = sorted(unique_norms, key=len)

# For efficiency, only check pairs where the shorter is >= 5 chars
# and <= half the length of the longer (avoid "care" matching everything)
checked = 0
for i, short in enumerate(by_length):
    if len(short) < 5:
        continue
    for j in range(i + 1, len(by_length)):
        long = by_length[j]
        if len(long) < len(short) + 2:
            # Too similar in length for containment to be meaningful
            continue
        if len(short) < len(long) * 0.4:
            # Short name is less than 40% of long name -- too generic
            continue
        checked += 1
        if short in long:
            # Get actual names
            short_names = singleton_norms[short]
            long_names = singleton_norms[long]
            for sn in short_names:
                for ln in long_names:
                    containment_matches.append({
                        'name_a': sn,
                        'name_b': ln,
                        'type': 'containment',
                        'detail': f'"{norm_corp(sn)}" contained in '
                                  f'"{norm_corp(ln)}"',
                    })

print(f"  {checked:,} pairs checked")
print(f"  {len(containment_matches)} containment matches found")
print()

# ===================================================================
# PASS 3: Token overlap (Jaccard similarity)
# ===================================================================

print("Pass 3: Token overlap matching...")

# Build token index: token -> list of names
token_index = defaultdict(list)
for name in singletons:
    tokens = tokenize(name)
    for t in tokens:
        token_index[t].append(name)

# Find candidate pairs that share at least one distinctive token
candidate_pairs = set()
for token, names in token_index.items():
    if len(names) > 50:
        continue  # Skip overly common tokens
    for a, b in combinations(names, 2):
        pair = (min(a, b), max(a, b))
        candidate_pairs.add(pair)

print(f"  {len(candidate_pairs):,} candidate pairs from shared tokens")

# Compute Jaccard similarity for candidates
token_matches = []
for a, b in candidate_pairs:
    tokens_a = tokenize(a)
    tokens_b = tokenize(b)
    if not tokens_a or not tokens_b:
        continue
    intersection = tokens_a & tokens_b
    union = tokens_a | tokens_b
    jaccard = len(intersection) / len(union)

    # Require high overlap AND at least 2 shared distinctive tokens
    if jaccard >= 0.5 and len(intersection) >= 2:
        token_matches.append({
            'name_a': a,
            'name_b': b,
            'type': 'token_overlap',
            'jaccard': jaccard,
            'shared': ', '.join(sorted(intersection)),
            'detail': f'Jaccard={jaccard:.2f}, shared=[{", ".join(sorted(intersection))}]',
        })

print(f"  {len(token_matches)} token overlap matches")
print()

# ===================================================================
# PASS 4: Edit distance for similar-length names
# ===================================================================

print("Pass 4: Edit distance matching...")

# Only check names with similar norm_corp lengths (within 3 chars)
# and same first 3 characters (blocking strategy)
prefix_blocks = defaultdict(list)
for name in singletons:
    n = norm_corp(name)
    if len(n) >= 6:
        prefix = n[:4]
        prefix_blocks[prefix].append((n, name))

edit_matches = []
edit_checked = 0
for prefix, items in prefix_blocks.items():
    if len(items) > 30:
        continue  # Skip huge blocks
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            n1, name1 = items[i]
            n2, name2 = items[j]
            if abs(len(n1) - len(n2)) > 3:
                continue
            edit_checked += 1
            dist = levenshtein(n1, n2)
            max_len = max(len(n1), len(n2))
            ratio = dist / max_len if max_len else 1
            if ratio <= 0.15 and dist > 0:
                edit_matches.append({
                    'name_a': name1,
                    'name_b': name2,
                    'type': 'edit_distance',
                    'distance': dist,
                    'ratio': ratio,
                    'detail': f'edit_dist={dist}, ratio={ratio:.2f}',
                })

print(f"  {edit_checked:,} pairs checked")
print(f"  {len(edit_matches)} edit distance matches")
print()

# ===================================================================
# Deduplicate fuzzy matches (a pair might appear in multiple passes)
# ===================================================================

all_fuzzy = containment_matches + token_matches + edit_matches
fuzzy_pairs = {}
for m in all_fuzzy:
    pair = (min(m['name_a'], m['name_b']), max(m['name_a'], m['name_b']))
    if pair not in fuzzy_pairs:
        fuzzy_pairs[pair] = m
    else:
        # Merge match types
        existing = fuzzy_pairs[pair]
        existing['type'] += ' + ' + m['type']
        existing['detail'] += '; ' + m['detail']

unique_fuzzy = list(fuzzy_pairs.values())
print(f"Total unique fuzzy matches: {len(unique_fuzzy)}")
print()

# ===================================================================
# Build workbook
# ===================================================================

print("Building review workbook...")
wb = Workbook()

# --- Tab 1: norm_corp clusters ---
ws1 = wb.active
ws1.title = "Exact Norm Clusters"

headers1 = [
    "Cluster_ID", "Norm_Key", "Corporate_Name", "Row_Count",
    "Served_Count", "States", "Source_Types", "Attribution",
    "Recommended_Canonical"
]
ws1.append(headers1)
style_headers(ws1, len(headers1))

cluster_id = 0
norm_cluster_rows = 0
for n, names in sorted(norm_clusters.items(),
                       key=lambda x: -sum(corp_counts[nm] for nm in x[1])):
    cluster_id += 1
    sorted_names = sorted(names, key=lambda x: -corp_counts[x])

    # Pick canonical via GLR authority, then highest count
    # For multi-name clusters, check each against GLR
    glr_match = None
    for name in sorted_names:
        nc = norm_corp(name)
        if nc in glr_parent_by_norm:
            glr_match = glr_parent_by_norm[nc]
            break
    canonical = glr_match if glr_match else sorted_names[0]
    reason = f'GLR: "{canonical}"' if glr_match else 'Highest count'

    for i, name in enumerate(sorted_names):
        is_canonical = (norm_corp(name) == norm_corp(canonical))
        row_data = [
            cluster_id, n, name, corp_counts[name],
            corp_served.get(name, 0),
            ', '.join(sorted(corp_states.get(name, set()))),
            ', '.join(sorted(corp_types.get(name, set()))),
            ', '.join(sorted(corp_attr_source.get(name, set()))),
            f'(canonical - {reason})' if is_canonical else canonical,
        ]
        ws1.append(row_data)
        norm_cluster_rows += 1
        row_num = ws1.max_row

        if corp_served.get(name, 0) > 0:
            fill = RED_FILL if not is_canonical else ORANGE_FILL
        elif not is_canonical:
            fill = YELLOW_FILL
        else:
            fill = GREEN_FILL
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = fill

auto_width(ws1)
print(f"  Tab 1: {norm_cluster_rows} rows in {cluster_id} norm_corp clusters")

# --- Tab 2: Fuzzy matches (priority-filtered, with RA decisions) ---
ws2 = wb.create_sheet("Fuzzy Matches")

# Build top-70 set for priority filtering
top70 = set(n for n, _ in corp_counts.most_common(70))

# Load RA's existing decisions from coded workbook (if it exists)
ra_decisions = {}  # (name_a, name_b) -> decision text
ra_file = ensure_report_dir() / "Corporate_Name_Dedup_Review_RA_first 60 or so.xlsx"
if ra_file.exists():
    ra_wb = load_workbook(ra_file, read_only=True)
    ra_ws = ra_wb["Fuzzy Matches"]
    for ra_row in ra_ws.iter_rows(min_row=2, values_only=True):
        ra_a, ra_b = ra_row[1], ra_row[4]
        ra_note = str(ra_row[11]).strip() if ra_row[11] else ''
        if ra_note:
            if ra_note.upper().startswith('NOT THE SAME'):
                ra_decisions[(ra_a, ra_b)] = 'NOT THE SAME'
            elif 'check' in ra_note.lower() or 'need' in ra_note.lower():
                ra_decisions[(ra_a, ra_b)] = 'NEEDS RESEARCH'
            else:
                ra_decisions[(ra_a, ra_b)] = 'CONFIRMED'
    ra_wb.close()
    print(f"  Loaded {len(ra_decisions)} RA decisions from prior review")

headers2 = [
    "Match_Type", "Name_A", "Count_A", "Served_A",
    "Name_B", "Count_B", "Served_B",
    "Detail", "Combined_Count", "Canonical", "Reason",
    "Priority_Why", "RA_Decision"
]
ws2.append(headers2)
style_headers(ws2, len(headers2))

# Sort by combined count descending (most impactful first)
unique_fuzzy.sort(key=lambda x: -(corp_counts[x['name_a']]
                                   + corp_counts[x['name_b']]))

tab2_count = 0
tab2_precoded = 0
for m in unique_fuzzy:
    a, b = m['name_a'], m['name_b']
    ct_a = corp_counts[a]
    ct_b = corp_counts[b]
    sv_a = corp_served.get(a, 0)
    sv_b = corp_served.get(b, 0)
    combined = ct_a + ct_b

    # Priority filter: top-70 OR served
    in_top70 = a in top70 or b in top70
    has_served = sv_a > 0 or sv_b > 0
    if not in_top70 and not has_served:
        continue

    # Build priority reason
    reasons = []
    if in_top70:
        reasons.append('Top 70')
    if has_served:
        reasons.append('Served')
    priority_why = ' + '.join(reasons)

    canonical, variant, reason = pick_canonical(a, ct_a, b, ct_b)

    # Carry forward RA decision
    ra_dec = ra_decisions.get((a, b), '')
    if ra_dec:
        tab2_precoded += 1

    row_data = [
        m['type'], a, ct_a, sv_a, b, ct_b, sv_b,
        m['detail'], combined, canonical, reason,
        priority_why, ra_dec
    ]
    ws2.append(row_data)
    tab2_count += 1
    row_num = ws2.max_row

    if ra_dec == 'NOT THE SAME':
        fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # grey
    elif ra_dec == 'CONFIRMED':
        fill = GREEN_FILL
    elif ra_dec == 'NEEDS RESEARCH':
        fill = YELLOW_FILL
    elif sv_a > 0 or sv_b > 0:
        fill = ORANGE_FILL
    elif combined >= 10:
        fill = YELLOW_FILL
    else:
        fill = GREEN_FILL
    for col in range(1, len(headers2) + 1):
        ws2.cell(row=row_num, column=col).fill = fill

auto_width(ws2)
print(f"  Tab 2: {tab2_count} priority pairs (from {len(unique_fuzzy)} total)")
print(f"         {tab2_precoded} already coded by RA, {tab2_count - tab2_precoded} remaining")

# --- Tab 3: High-impact summary ---
ws3 = wb.create_sheet("High Impact Summary")

headers3 = [
    "Type", "Name_A", "Count_A", "Served_A",
    "Name_B", "Count_B", "Served_B",
    "Combined", "Canonical", "Reason"
]
ws3.append(headers3)
style_headers(ws3, len(headers3))

# Combine norm clusters + fuzzy matches, filter to high-impact
high_impact = []

# From norm clusters: where non-canonical variants have 5+ rows or any served
for n, names in norm_clusters.items():
    sorted_names = sorted(names, key=lambda x: -corp_counts[x])
    # Use GLR-aware canonical selection for first pair
    best = sorted_names[0]
    for variant in sorted_names[1:]:
        ct = corp_counts[variant]
        sv = corp_served.get(variant, 0)
        if ct >= 5 or sv > 0:
            canon, _, reason = pick_canonical(
                best, corp_counts[best], variant, ct
            )
            high_impact.append({
                'type': 'norm_corp',
                'name_a': best,
                'count_a': corp_counts[best],
                'served_a': corp_served.get(best, 0),
                'name_b': variant,
                'count_b': ct,
                'served_b': sv,
                'combined': corp_counts[best] + ct,
                'canonical': canon,
                'reason': reason,
            })

# From fuzzy: where combined count >= 10 or any served
for m in unique_fuzzy:
    a, b = m['name_a'], m['name_b']
    ct_a, ct_b = corp_counts[a], corp_counts[b]
    sv_a = corp_served.get(a, 0)
    sv_b = corp_served.get(b, 0)
    combined = ct_a + ct_b
    if combined >= 10 or sv_a > 0 or sv_b > 0:
        canon, _, reason = pick_canonical(a, ct_a, b, ct_b)
        high_impact.append({
            'type': f'fuzzy ({m["type"]})',
            'name_a': a,
            'count_a': ct_a,
            'served_a': sv_a,
            'name_b': b,
            'count_b': ct_b,
            'served_b': sv_b,
            'combined': combined,
            'canonical': canon,
            'reason': reason,
        })

high_impact.sort(key=lambda x: -x['combined'])

for item in high_impact:
    row_data = [
        item['type'], item['name_a'], item['count_a'], item['served_a'],
        item['name_b'], item['count_b'], item['served_b'],
        item['combined'], item['canonical'], item['reason']
    ]
    ws3.append(row_data)
    row_num = ws3.max_row
    if item['served_a'] > 0 or item['served_b'] > 0:
        fill = ORANGE_FILL
    else:
        fill = YELLOW_FILL
    for col in range(1, len(headers3) + 1):
        ws3.cell(row=row_num, column=col).fill = fill

auto_width(ws3)
print(f"  Tab 3: {len(high_impact)} high-impact consolidation candidates")

# ===================================================================
# Save
# ===================================================================

output_path = ensure_report_dir() / "Corporate_Name_Dedup_Review.xlsx"
wb.save(output_path)
print(f"\nSaved: {output_path}")

# ===================================================================
# Console summary
# ===================================================================

print()
print("=" * 65)
print("CORPORATE NAME DEDUP ANALYSIS SUMMARY")
print("=" * 65)
print(f"  Distinct Corporate_Names in V22.3:   {len(distinct_names):,}")
print()
print(f"  Pass 1 (norm_corp exact clusters):    {len(norm_clusters)} clusters")
print(f"    Names in clusters:                  {total_variants}")
print(f"    Consolidatable rows:                "
      f"{sum(sum(corp_counts[n] for n in names[1:]) for names in [sorted(v, key=lambda x: -corp_counts[x]) for v in norm_clusters.values()]):,}")
print()
print(f"  Pass 2 (containment):                 {len(containment_matches)} matches")
print(f"  Pass 3 (token overlap):               {len(token_matches)} matches")
print(f"  Pass 4 (edit distance):               {len(edit_matches)} matches")
print(f"  Total fuzzy (deduplicated):           {len(unique_fuzzy)} pairs")
print()
print(f"  HIGH IMPACT (Tab 3):                  {len(high_impact)} candidates")
print(f"    Affecting served facilities:         "
      f"{sum(1 for h in high_impact if h['served_a'] > 0 or h['served_b'] > 0)}")
print()

# Show top 30 high-impact
if high_impact:
    print("  TOP 30 HIGH-IMPACT CONSOLIDATION CANDIDATES:")
    print(f"  {'Name A':<35s} {'#':>5s} {'Name B':<35s} {'#':>5s} {'Comb':>5s}")
    print(f"  {'-'*35} {'-'*5} {'-'*35} {'-'*5} {'-'*5}")
    for h in high_impact[:30]:
        a = h['name_a'][:34]
        b = h['name_b'][:34]
        tag_a = '*' if h['served_a'] > 0 else ' '
        tag_b = '*' if h['served_b'] > 0 else ' '
        print(f"  {a:<35s}{h['count_a']:>5,}{tag_a}{b:<35s}{h['count_b']:>5,}{tag_b}{h['combined']:>5,}")
    if len(high_impact) > 30:
        print(f"  ... and {len(high_impact) - 30} more (see Tab 3)")
    print()
    print("  * = has served facilities")
