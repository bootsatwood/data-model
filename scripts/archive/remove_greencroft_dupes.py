#!/usr/bin/env python3
"""
Remove 3 duplicate/phantom rows from V22.6.

Rows to remove (confirmed by RA):
  - Row 4794: Greencroft Goshen ALF — phantom (no ALF in Facility DB)
  - Row 19306: Oak Grove Christian Retirement Village — duplicate of row 5022
  - Row 20674: Walnut Hills Retirement Home — duplicate of row 13013

Removes rows in-place from V22.6, preserving formatting.

Usage:
  python remove_greencroft_dupes.py preview   # show rows without modifying
  python remove_greencroft_dupes.py apply      # remove rows and overwrite V22.6
"""

import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
from openpyxl import load_workbook
from utils import safe, VAULT

DB_FILE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_6.xlsx"

ROWS_TO_REMOVE = {4794, 19306, 20674}


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python remove_greencroft_dupes.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Remove Greencroft Duplicates/Phantoms ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_FILE.name}...")
    wb = load_workbook(DB_FILE)
    ws = wb.active

    # Build header index
    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    fac_col = col_idx.get('Facility_Name')
    addr_col = col_idx.get('Address')
    city_col = col_idx.get('City')
    state_col = col_idx.get('State')
    type_col = col_idx.get('Source_Type')
    served_col = col_idx.get('Do_We_Serve')
    corp_col = col_idx.get('Corporate_Name')

    print(f"  {ws.max_row - 1:,} data rows loaded.\n")

    # Verify rows exist and show details
    print("Rows to remove:")
    print("-" * 40)
    verified = []
    for excel_row in sorted(ROWS_TO_REMOVE):
        fac = safe(ws.cell(row=excel_row, column=fac_col).value)
        addr = safe(ws.cell(row=excel_row, column=addr_col).value)
        city = safe(ws.cell(row=excel_row, column=city_col).value)
        state = safe(ws.cell(row=excel_row, column=state_col).value)
        stype = safe(ws.cell(row=excel_row, column=type_col).value)
        served = safe(ws.cell(row=excel_row, column=served_col).value)
        corp = safe(ws.cell(row=excel_row, column=corp_col).value)

        if not fac:
            print(f"  WARNING: Row {excel_row} appears empty — skipping")
            continue

        print(f"  Row {excel_row}: {fac}")
        print(f"    {addr}, {city}, {state} | {stype} | Corp: {corp} | Served: {served}")
        verified.append(excel_row)

    print(f"\n  {len(verified)} rows verified for removal")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY — no file modified **")
        print(f"  Run with 'apply' to remove rows and overwrite {DB_FILE.name}")
        return

    # Remove rows — delete from bottom up to preserve row numbers
    for excel_row in sorted(verified, reverse=True):
        ws.delete_rows(excel_row, 1)
        print(f"  Deleted row {excel_row}")

    wb.save(DB_FILE)
    print(f"\n  Saved: {DB_FILE}")
    print(f"  New row count: {ws.max_row - 1:,} data rows")


if __name__ == '__main__':
    main()
