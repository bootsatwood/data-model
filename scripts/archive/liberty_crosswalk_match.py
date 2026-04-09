#!/usr/bin/env python3
"""Match Liberty website facilities against V22.1 database."""

import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
sys.path.insert(0, ".")
from utils import load_db, norm_addr, norm, safe, ensure_report_dir, VAULT
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB = VAULT / "Current" / "1_Combined_Database_FINAL_V22_6.xlsx"
headers, rows = load_db(DB)
print(f"V22.1 loaded: {len(rows):,} rows")

# ── Styles ──
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Build DB address index ──
db_by_addr = {}
for r in rows:
    akey = norm_addr(safe(r.get("Address", ""))) + "|" + norm(safe(r.get("City", ""))) + "|" + norm(safe(r.get("State", "")))
    db_by_addr.setdefault(akey, []).append(r)

# Also index by city+state for fallback name matching
db_by_city_state = {}
for r in rows:
    key = (norm(safe(r.get("City", ""))), norm(safe(r.get("State", ""))))
    db_by_city_state.setdefault(key, []).append(r)


def find_db_matches(name, address, city, state):
    """Find DB rows matching a website facility. Returns list of DB row dicts."""
    matches = []
    city_n = norm(city)
    state_n = norm(state)
    addr_n = norm_addr(address) if address else ""

    # 1. Exact address match
    if addr_n:
        akey = addr_n + "|" + city_n + "|" + state_n
        if akey in db_by_addr:
            return db_by_addr[akey]  # address match is definitive

    # 2. Fuzzy address — try without last token (unit/suite numbers)
    if addr_n and len(addr_n) > 8:
        for db_akey, db_rows in db_by_addr.items():
            db_parts = db_akey.split("|")
            if len(db_parts) == 3 and db_parts[1] == city_n and db_parts[2] == state_n:
                # Check if addresses share a long common prefix
                if db_parts[0][:12] == addr_n[:12] and len(addr_n) > 12:
                    matches.extend(db_rows)
        if matches:
            return matches

    # 3. Name-based fallback in same city/state
    candidates = db_by_city_state.get((city_n, state_n), [])
    name_n = norm(name)
    for r in candidates:
        db_name_n = norm(safe(r.get("Facility_Name", "")))
        # Check substantial overlap
        if len(name_n) > 6 and len(db_name_n) > 6:
            if name_n[:10] == db_name_n[:10]:
                matches.append(r)
    return matches


# ── Website facility data (copied from build_liberty_crosswalk.py structure) ──
# We import the data by running the crosswalk builder's data
# Rather than duplicate, let's just define the lists here

lsl_facilities = [
    ("Brightmore of South Charlotte", "LSL - Life Plan", "10225 Old Ardrey Kell Road", "Charlotte", "NC", "28277"),
    ("Brightmore of Wilmington", "LSL - Life Plan", "2124 41st Street", "Wilmington", "NC", "28403"),
    ("Carlisle Palm Beach", "LSL - Senior Living", "450 East Ocean Avenue", "Lantana", "FL", "33462"),
    ("Carolina Bay at Autumn Hall", "LSL - Life Plan", "620 Carolina Bay Drive", "Wilmington", "NC", "28403"),
    ("The Carrollton", "LSL - Senior Living", "701 South Carrollton Avenue", "New Orleans", "LA", "70118"),
    ("Hayes Barton Place", "LSL - Life Plan", "2600 Yettington Drive", "Raleigh", "NC", "27608"),
    ("Inspire Briar Chapel", "Inspire (55+)", "152 Market Chapel Road", "Pittsboro", "NC", "27312"),
    ("Inspire Brunswick Forest", "Inspire (55+)", "6146 Liberty Hall Drive", "Leland", "NC", "28451"),
    ("Inspire Royal Park", "Inspire (55+)", "4101 Glenloch Circle", "Matthews", "NC", "28105"),
    ("Inspire Sandhill", "Inspire (55+)", "440 Town Center Place", "Columbia", "SC", "29229"),
    ("Kempton of Charleston", "LSL - Senior Living", "194 Spring Street", "Charleston", "SC", "29403"),
    ("Kempton of Jacksonville", "LSL - Senior Living", "3045 Henderson Drive Extension", "Jacksonville", "NC", "28546"),
    ("The Kempton of Hermitage", "LSL - Senior Living", "3778 Central Pike", "Hermitage", "TN", "37076"),
    ("Kempton of Rockhill", "LSL - Senior Living", "1611 Constitution Boulevard", "Rock Hill", "SC", "29732"),
    ("Oakleaf Village of Lexington", "LSL - Senior Living", "800 N Lake Drive", "Lexington", "SC", "29072"),
    ("Pisgah Valley", "LSL - Life Plan", "95 Holcombe Cove Road", "Candler", "NC", "28715"),
    ("Quail Haven Village", "LSL - Life Plan", "155 Blake Boulevard", "Pinehurst", "NC", "28374"),
    ("South Bay at Mount Pleasant", "LSL - Senior Living", "1400 Liberty Midtown Drive", "Mount Pleasant", "SC", "29464"),
    ("The Barclay at SouthPark", "LSL - Life Plan", "4801 Barclay Downs Drive", "Charlotte", "NC", "28210"),
    ("The Peninsula of Charleston", "LSL - Life Plan", "625 King Street", "Charleston", "SC", "29403"),
    ("The Preserve at Fairfield Glade", "LSL - Senior Living", "100 Samaritan Way", "Crossville", "TN", "38558"),
    ("The Templeton of Cary", "LSL - Life Plan", "125 Brightmore Drive", "Cary", "NC", "27511"),
    ("Wellington Bay", "LSL - Senior Living", "10430 Stable Lane", "Wellington", "FL", "33414"),
    # Not yet screenshotted — still have addresses from user's original list
    ("Harbour's Edge", "LSL", "", "Delray Beach", "FL", ""),
    ("The Palms of Sebring", "LSL", "", "Sebring", "FL", ""),
    ("Park Summit", "LSL", "", "Hendersonville", "NC", ""),
    ("Plantation Village", "LSL", "", "Wilmington", "NC", ""),
    ("Rosemont", "LSL", "", "Brampton", "ON", ""),
    ("The Cypress of Charlotte", "LSL", "", "Charlotte", "NC", ""),
    ("The Cypress of Hilton Head Island", "LSL", "", "Hilton Head Island", "SC", ""),
    ("The Cypress of Raleigh", "LSL", "", "Raleigh", "NC", ""),
    ("The Pines at Davidson", "LSL", "", "Davidson", "NC", ""),
    ("The Pointe at Lifespring", "LSL", "", "Knoxville", "TN", ""),
    ("The Vista at Sedge Garden", "LSL", "", "Kernersville", "NC", ""),
    ("Vista at Town Center", "LSL", "", "King", "NC", ""),
    ("Waypoint at Bridgeway Station", "LSL", "", "Morrisville", "NC", ""),
    ("Westminster Towers", "LSL", "", "Rock Hill", "SC", ""),
    ("Willow Valley Communities", "LSL", "", "Lancaster", "PA", ""),
]

lhr_facilities = [
    ("Liberty Commons N&R of Alamance County", "LHR - SNF", "791 Boone Station Drive", "Burlington", "NC", "27215"),
    ("Three Rivers Health & Rehab Center", "LHR - SNF", "1403 Connor Ave.", "Windsor", "NC", "27983"),
    ("Elizabethtown Healthcare & Rehab Center", "LHR - SNF", "208 Mercer Road", "Elizabethtown", "NC", "28337"),
    ("Southport Health & Rehab Center", "LHR - SNF", "630 N Fodale Ave", "Southport", "NC", "28461"),
    ("Pisgah Manor", "LHR - SNF", "104 Holcombe Cove Road", "Candler", "NC", "28715"),
    ("Liberty Commons N&R of Columbus County", "LHR - SNF", "1402 Pinckney Street", "Whiteville", "NC", "28472"),
    ("Shoreland Health Care & Retirement Center", "LHR - SNF", "200 Flowers-Pridgen Drive", "Whiteville", "NC", "28472"),
    ("Golden Years Nursing Home", "LHR - SNF", "7348 North West Street", "Falcon", "NC", "28342"),
    ("Woodlands Nursing & Rehab Center", "LHR - SNF", "400 Pelt Drive", "Fayetteville", "NC", "28301"),
    ("Highland House Rehab & Healthcare", "LHR - SNF", "1700 Pamalee Dr.", "Fayetteville", "NC", "28301"),
    ("Bermuda Commons N&R Center", "LHR - SNF", "316 NC Highway 801 South", "Advance", "NC", "27006"),
    ("Briar Creek at The Barclay", "LHR - SNF", "6041 Piedmont Row Drive", "Charlotte", "NC", "28208"),
    ("Oak Forest Health & Rehab Center", "LHR - SNF", "5680 Windy Hill Drive", "Winston-Salem", "NC", "27105"),
    ("The Oaks", "LHR - SNF", "901 Bethesda Road", "Winston-Salem", "NC", "27103"),
    ("Summerstone Health & Rehab Center", "LHR - SNF", "485 Veteran's Way", "Kernersville", "NC", "27284"),
    ("Louisburg Healthcare & Rehab Center", "LHR - SNF", "202 Smoketree Way", "Louisburg", "NC", "27549"),
    ("Liberty Commons N&R of Halifax County", "LHR - SNF", "101 Caroline Avenue", "Weldon", "NC", "27890"),
    ("Silver Bluff Village", "LHR - SNF", "100 Silver Bluff Dr.", "Canton", "NC", "28716"),
    ("Liberty Commons N&R of Johnston County", "LHR - SNF", "2315 Highway 242 North", "Benson", "NC", "27504"),
    ("Liberty Commons N&R of Lee County", "LHR - SNF", "310 Commerce Drive", "Sanford", "NC", "27332"),
    ("Westfield Rehab and Health Center", "LHR - SNF", "3100 Tramway Road", "Sanford", "NC", "27332"),
    ("Royal Park Rehab and Health Center", "LHR - SNF", "2700 Royal Commons Lane", "Matthews", "NC", "28105"),
    ("The Pavilion Health Center at Brightmore", "LHR - SNF", "10011 Providence Road West", "Charlotte", "NC", "28277"),
    ("Pinehurst Healthcare & Rehab Center", "LHR - SNF", "300 Blake Road", "Pinehurst", "NC", "28374"),
    ("The Inn at Quail Haven Village", "LHR - SNF", "155 Blake Road", "Pinehurst", "NC", "28374"),
    ("Bradley Creek at Carolina Bay", "LHR - SNF", "740 Diamond Shoals Road", "Wilmington", "NC", "28403"),
    ("Liberty Commons Rehab Center", "LHR - SNF", "121 Racine Drive", "Wilmington", "NC", "28403"),
    ("Parkview Health & Rehab Center", "LHR - SNF", "1718 Legion Road", "Chapel Hill", "NC", "27517"),
    ("Roxboro Healthcare & Rehab Center", "LHR - SNF", "901 Ridge Road", "Roxboro", "NC", "27573"),
    ("WoodHaven Nursing Center", "LHR - SNF", "1150 Pine Run Dr.", "Lumberton", "NC", "28358"),
    ("Liberty Commons N&R of Rowan County", "LHR - SNF", "4412 South Main Street", "Salisbury", "NC", "28147"),
    ("Mary Gran Nursing Center", "LHR - SNF", "120 Southwood Drive", "Clinton", "NC", "28329"),
    ("Southwood Nursing & Rehab Center", "LHR - SNF", "180 Southwood Drive", "Clinton", "NC", "28329"),
    ("Capital Nursing & Rehab Center", "LHR - SNF", "3000 Holston Lane", "Raleigh", "NC", "27610"),
    ("Bloomsbury at Hayes Barton Place", "LHR - SNF", "2750 Oberlin Road", "Raleigh", "NC", "27608"),
    ("Swift Creek at The Templeton", "LHR - SNF", "221 Brightmore Drive", "Cary", "NC", "27518"),
    ("Warren Hills Rehab & Nursing Center", "LHR - SNF", "864 US Hwy. 158 Business West", "Warrenton", "NC", "27589"),
    ("The Foley Center at Chestnut Ridge", "LHR - SNF", "621 Chestnut Ridge Parkway", "Blowing Rock", "NC", "28605"),
    ("Yadkin Nursing Care Center", "LHR - SNF", "903 W. Main Street", "Yadkinville", "NC", "27055"),
    ("Kempton of Charleston", "LHR - SNF", "194 Spring Street", "Charleston", "SC", "29403"),
    ("Preserve at Fairfield Glade", "LHR - SNF", "100 Samaritan Way", "Crossville", "TN", "38558"),
]

FOOTPRINT = {'IN', 'KY', 'NC', 'OH', 'SC', 'VA', 'MI', 'IL', 'WI', 'MN', 'FL', 'MD', 'GA', 'MO'}

# ── Output columns ──
out_cols = [
    "Website Facility Name", "Division", "Website Address", "City", "State", "ZIP",
    "Match Status",
    "DB Excel Row", "DB Facility_Name", "DB Corporate_Name",
    "DB Source_Type", "DB Address",
    "DB Total_Beds", "DB Census", "DB Do_We_Serve",
]


def write_out_header(ws):
    for c, col_name in enumerate(out_cols, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"


def write_out_row(ws, row_num, vals, fill=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def autofit_out(ws, max_row):
    for c in range(1, len(out_cols) + 1):
        max_len = len(str(ws.cell(row=1, column=c).value or ""))
        for r in range(2, min(max_row + 1, 100)):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 45))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max_len + 3


wb = Workbook()

# Track all matched DB row numbers so we can find orphans
all_matched_db_rows = set()


def process_tab(ws, facility_list, tab_label):
    global all_matched_db_rows
    write_out_header(ws)
    rn = 2
    matched = 0
    no_match = 0
    no_addr = 0
    out_of_footprint = 0

    for (name, division, address, city, state, zipcode) in facility_list:
        in_fp = state.upper() in FOOTPRINT if state else False

        if not address:
            vals = [name, division, "", city, state, zipcode,
                    "NO ADDRESS YET", "", "", "", "", "", "", "", ""]
            write_out_row(ws, rn, vals, gray_fill)
            rn += 1
            no_addr += 1
            continue

        if not in_fp:
            vals = [name, division, address, city, state, zipcode,
                    "OUT OF FOOTPRINT", "", "", "", "", "", "", "", ""]
            write_out_row(ws, rn, vals, gray_fill)
            rn += 1
            out_of_footprint += 1
            continue

        db_matches = find_db_matches(name, address, city, state)

        if db_matches:
            for m in db_matches:
                all_matched_db_rows.add(m["_excel_row"])
                vals = [
                    name, division, address, city, state, zipcode,
                    "MATCHED",
                    m["_excel_row"], safe(m.get("Facility_Name", "")),
                    safe(m.get("Corporate_Name", "")),
                    safe(m.get("Source_Type", "")), safe(m.get("Address", "")),
                    m.get("Total_Beds", ""), m.get("Census", ""),
                    safe(m.get("Do_We_Serve", "")),
                ]
                write_out_row(ws, rn, vals, green_fill)
                rn += 1
                matched += 1
        else:
            vals = [name, division, address, city, state, zipcode,
                    "NOT IN DB", "", "", "", "", "", "", "", ""]
            write_out_row(ws, rn, vals, yellow_fill)
            rn += 1
            no_match += 1

    autofit_out(ws, ws.max_row)
    print(f"\n{tab_label}:")
    print(f"  Matched:          {matched}")
    print(f"  Not in DB:        {no_match}")
    print(f"  No address yet:   {no_addr}")
    print(f"  Out of footprint: {out_of_footprint}")
    return matched, no_match


# Tab 1: LSL matches
ws1 = wb.active
ws1.title = "LSL Matches"
lsl_matched, lsl_missing = process_tab(ws1, lsl_facilities, "LSL - Senior Living")

# Tab 2: LHR matches
ws2 = wb.create_sheet("LHR Matches")
lhr_matched, lhr_missing = process_tab(ws2, lhr_facilities, "LHR - Healthcare & Rehab")

# Tab 3: DB Orphans — rows with "LIBERTY" in Corporate_Name that didn't match any website facility
ws3 = wb.create_sheet("DB Liberty Orphans")
write_out_header(ws3)

liberty_db_rows = [
    r for r in rows
    if "LIBERTY" in safe(r.get("Corporate_Name", "")).upper()
]
orphans = [r for r in liberty_db_rows if r["_excel_row"] not in all_matched_db_rows]
matched_liberty = [r for r in liberty_db_rows if r["_excel_row"] in all_matched_db_rows]

rn = 2
for r in sorted(orphans, key=lambda x: (safe(x.get("Corporate_Name", "")), safe(x.get("State", "")), safe(x.get("City", "")))):
    vals = [
        "", safe(r.get("Corporate_Name", "")), "", safe(r.get("City", "")),
        safe(r.get("State", "")), safe(r.get("ZIP", "")),
        "ORPHAN — no website match",
        r["_excel_row"], safe(r.get("Facility_Name", "")),
        safe(r.get("Corporate_Name", "")),
        safe(r.get("Source_Type", "")), safe(r.get("Address", "")),
        r.get("Total_Beds", ""), r.get("Census", ""),
        safe(r.get("Do_We_Serve", "")),
    ]
    write_out_row(ws3, rn, vals, red_fill)
    rn += 1

autofit_out(ws3, ws3.max_row)

print(f"\nDB Liberty Rows Summary:")
print(f"  Total DB rows with 'LIBERTY' in Corporate_Name: {len(liberty_db_rows)}")
print(f"  Matched to a website facility:                  {len(matched_liberty)}")
print(f"  Orphans (no website match):                     {len(orphans)}")

# Print corporate name breakdown for orphans
from collections import Counter
orphan_corps = Counter(safe(r.get("Corporate_Name", "")) for r in orphans)
print(f"\n  Orphan Corporate_Name breakdown:")
for corp, cnt in orphan_corps.most_common():
    print(f"    {corp}: {cnt}")

outpath = ensure_report_dir() / "Liberty_Crosswalk_DB_Match.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
