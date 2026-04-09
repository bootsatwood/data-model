"""Analyze FP-scoped revenue distribution for MUO-qualifying entities.
Used to find new RP bracket thresholds."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl, statistics
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

# Build per-corp footprint data
corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0, 'states': set()})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    d['states'].add(state)
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served:
        d['served_camps'].add((addr, city))

# Filter to >=7 campuses
muos = []
for corp, d in corp_fp.items():
    n = len(d['campuses'])
    if n >= 7:
        muos.append({
            'corp': corp,
            'camp': n,
            'srv': len(d['served_camps']),
            'rev': d['tot_pot'],
            'st': sorted(d['states']),
        })

muos.sort(key=lambda x: x['rev'])

print(f"MUO-qualifying entities (>=7 campuses in footprint): {len(muos)}")
print()
print(f"{'#':>3} {'Entity':<45} {'Camp':>5} {'Srv':>4} {'FP Tot Pot':>14}")
print("-" * 75)
for i, e in enumerate(muos, 1):
    print(f"{i:>3} {e['corp']:<45} {e['camp']:>5} {e['srv']:>4} ${e['rev']:>13,.0f}")

revs = [e['rev'] for e in muos]
sr = sorted(revs)
n = len(sr)

print()
print("--- Revenue Distribution ---")
print(f"Count:  {n}")
print(f"Min:    ${min(revs):>13,.0f}")
print(f"P20:    ${sr[int(n*0.2)]:>13,.0f}")
print(f"P40:    ${sr[int(n*0.4)]:>13,.0f}")
print(f"Median: ${statistics.median(revs):>13,.0f}")
print(f"P60:    ${sr[int(n*0.6)]:>13,.0f}")
print(f"P80:    ${sr[int(n*0.8)]:>13,.0f}")
print(f"Max:    ${max(revs):>13,.0f}")
print(f"Mean:   ${statistics.mean(revs):>13,.0f}")

# V20 tier distribution target
print()
print("--- V20 Tier Distribution (target) ---")
print("T1: 22  T2: 36  T3: 12  (total 70)")
print()

# V20 RP distribution target
print("--- V20 RP Score Distribution (target) ---")
print("RP=1: 14  RP=2: 16  RP=3: 14  RP=4: 8  RP=5: 18  (total 70)")
print()

# Try quintile-based brackets
print("--- Quintile Brackets (equal distribution) ---")
for pct, label in [(0.2, 'P20'), (0.4, 'P40'), (0.6, 'P60'), (0.8, 'P80')]:
    val = sr[int(n * pct)]
    print(f"  {label}: ${val:>13,.0f}")

# Try matching V20 proportions: 14/70=20%, 16/70=23%, 14/70=20%, 8/70=11%, 18/70=26%
# Cumulative: 20%, 43%, 63%, 74%, 100%
print()
print("--- V20-Proportional Brackets ---")
for pct, label in [(0.20, 'RP=1 ceiling (P20)'), (0.43, 'RP=2 ceiling (P43)'),
                    (0.63, 'RP=3 ceiling (P63)'), (0.74, 'RP=4 ceiling (P74)')]:
    idx = min(int(n * pct), n - 1)
    val = sr[idx]
    print(f"  {label}: ${val:>13,.0f}")

# Also show campus distribution for ER
camps = sorted([e['camp'] for e in muos])
print()
print("--- Campus Distribution ---")
print(f"Count:  {len(camps)}")
print(f"Min:    {min(camps)}")
print(f"P20:    {camps[int(n*0.2)]}")
print(f"P40:    {camps[int(n*0.4)]}")
print(f"Median: {statistics.median(camps)}")
print(f"P60:    {camps[int(n*0.6)]}")
print(f"P80:    {camps[int(n*0.8)]}")
print(f"Max:    {max(camps)}")

# V20 ER distribution
print()
print("--- V20 ER Score Distribution (target) ---")
print("ER=1: 35  ER=2: 2  ER=3: 18  ER=4: 0  ER=5: 15  (total 70)")
