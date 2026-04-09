"""
Search the V23 Combined Database for facility records at:
  31869 Chicago Trail, New Carlisle, IN

Criteria: address contains "Chicago Trail" OR city is "New Carlisle" in Indiana (state = IN)
"""

import openpyxl
import os
import sys

DB_PATH = r"C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V23.xlsx"

print(f"Opening: {DB_PATH}")
print(f"File size: {os.path.getsize(DB_PATH) / (1024*1024):.1f} MB")
print()

wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
print(f"Sheet names: {wb.sheetnames}")
print()

# We'll search every sheet, but the main facility data is likely on the first sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Read header row
    headers = []
    first_row = True
    rows_found = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if first_row:
            headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
            first_row = False
            continue

        # Build a dict for this row
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = cell
            else:
                row_dict[f"col_{i}"] = cell

        # Convert relevant fields to strings for searching (case-insensitive)
        row_str_values = {k: str(v).strip().lower() if v is not None else "" for k, v in row_dict.items()}

        # Find address-like columns and city/state columns
        address_match = False
        city_state_match = False

        for col_name, val in row_str_values.items():
            col_lower = col_name.lower()

            # Check for "Chicago Trail" in any address-like column
            if "chicago trail" in val:
                address_match = True

            # Check for "New Carlisle" in any column
            if "new carlisle" in val:
                city_state_match = True

        # Also check if "31869" appears anywhere in the row
        has_31869 = any("31869" in str(v).lower() for v in row_dict.values() if v is not None)

        if address_match or city_state_match or has_31869:
            row_dict["_row_number"] = row_idx
            row_dict["_match_reason"] = []
            if address_match:
                row_dict["_match_reason"].append("Chicago Trail in address")
            if city_state_match:
                row_dict["_match_reason"].append("New Carlisle match")
            if has_31869:
                row_dict["_match_reason"].append("31869 match")
            row_dict["_match_reason"] = " | ".join(row_dict["_match_reason"])
            rows_found.append(row_dict)

    if rows_found:
        print(f"{'='*120}")
        print(f"SHEET: '{sheet_name}' -- Found {len(rows_found)} matching row(s)")
        print(f"{'='*120}")
        print(f"\nColumn headers ({len(headers)} columns):")
        for i, h in enumerate(headers):
            print(f"  [{i:3d}] {h}")
        print()

        for match_idx, row_dict in enumerate(rows_found, 1):
            print(f"\n--- Match #{match_idx} (Excel row {row_dict['_row_number']}) | Reason: {row_dict['_match_reason']} ---")
            for col_name in headers:
                val = row_dict.get(col_name, "")
                if val is not None and str(val).strip() != "":
                    print(f"  {col_name:40s} : {val}")
            print()
    else:
        print(f"Sheet '{sheet_name}': No matches found.")

wb.close()
print("\nDone.")
