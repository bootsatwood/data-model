"""Add 'MUO Data Sequence' tab — MUO Name + V23 Tier in Finance workbook order."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/MUO_Scoring_Workbook_V23_v8.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"

# 1. MUO Data entity order (col B, rows 18+)
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_muo = wb_muo['Analysis']
muo_order = []
for r in range(18, ws_muo.max_row + 1):
    v = ws_muo.cell(r, 2).value
    if v and str(v).strip() and str(v).strip() != 'TOTAL TOP 60':
        muo_order.append(str(v).strip())

# 2. Build tier lookup from scoring workbook
wb = openpyxl.load_workbook(WB_PATH)

tiers = {}
# Summary = scored T1/T2/T3
ws_sum = wb['Summary']
for r in range(2, ws_sum.max_row + 1):
    name = str(ws_sum.cell(r, 1).value or '').strip()
    if name:
        tiers[name.upper()] = str(ws_sum.cell(r, 16).value or '')

# T4
ws_t4 = wb['T4 - Independents']
for r in range(2, ws_t4.max_row + 1):
    name = str(ws_t4.cell(r, 1).value or '').strip()
    if name and name != 'FACILITY DETAIL':
        if ws_t4.cell(r, 3).value and isinstance(ws_t4.cell(r, 3).value, (int, float)):
            tiers[name.upper()] = 'T4'

# T5 hardcoded set
T5 = {'BLUEGRASS/ENCORE', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'HILL VALLEY',
      'SINGH', 'CARDON & ASSOCIATES', 'EASTERN HEALTHCARE GROUP', 'CLEARVIEW',
      'PAVILION HEALTHCARE', 'EMBASSY', 'EXCEPTIONAL LIVING', 'PORTOPICCOLO',
      'VENZA', 'AVENTURA', 'JOURNEY'}
for t5 in T5:
    tiers[t5] = 'T5'

# 3. Build tab
if 'MUO Data Sequence' in wb.sheetnames:
    del wb['MUO Data Sequence']

ws = wb.create_sheet('MUO Data Sequence')
HDR_FONT = Font(bold=True, size=11, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='4472C4')

ws.cell(1, 1, 'MUO Name').font = HDR_FONT
ws.cell(1, 1).fill = HDR_FILL
ws.cell(1, 2, 'V23 Tier').font = HDR_FONT
ws.cell(1, 2).fill = HDR_FILL
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12

FILLS = {
    'T1': PatternFill('solid', fgColor='C6EFCE'),
    'T2': PatternFill('solid', fgColor='FFEB9C'),
    'T3': PatternFill('solid', fgColor='FFC7CE'),
    'T4': PatternFill('solid', fgColor='D6DCE4'),
    'T5': PatternFill('solid', fgColor='F2DCDB'),
}

for i, name in enumerate(muo_order, 2):
    tier = tiers.get(name.upper(), '?')
    ws.cell(i, 1, name)
    ws.cell(i, 2, tier)
    fill = FILLS.get(tier)
    if fill:
        ws.cell(i, 1).fill = fill
        ws.cell(i, 2).fill = fill
    if tier == '?':
        print(f'  NOT FOUND: {name}')

wb.save(WB_PATH)
print(f"Done. {len(muo_order)} entities, MUO Data Sequence tab saved.")
