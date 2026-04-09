"""
Sale Type AR Analysis — Q1 2026 Wins
Calculates annualized revenue by sale type (New Logo / New Door / Integration)
using corrected V25.6 corporate attributions and Monday.com consent counts.
"""

import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# AR formula rates (from Monday.com formula_mkwk98vn)
SNF_PCP = 141.58 * 12  # $1,698.96
SNF_MH = 1623.96
ALF_PCP = 1875.84
ALF_MH = 1898.40
INTEG_SNF_MH_MULT = 0.5
INTEG_ALF_MH_MULT = 0.5

# Load V25.6 DB for corporate lookup
wb = openpyxl.load_workbook(
    r'C:\Users\ratwood\OneDrive - Eventus WholeHealth\Vault\02_Data_Model\Current\1_Combined_Database_FINAL_V25_6.xlsx',
    read_only=True, data_only=True
)
ws = wb[wb.sheetnames[0]]

corps_we_serve = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    corp = (row[2] or '').strip()
    serve = row[11]
    if serve == 'Yes' and corp:
        corps_we_serve.add(corp.upper())
wb.close()

# Q1 2026 wins: (name, ae, month, is_est, snf_pcp, snf_mh, integ_snf_mh, alf_pcp, alf_mh, integ_alf_mh)
wins = [
    ("Gallatin Nursing and Rehab", "Adam Sabie", "2026-01", False, 100,0,0,0,0,0),
    ("McCoy Memorial Nursing Center", "Chad Vukelich", "2026-01", False, 0,70,0,0,0,0),
    ("Trinity Village (LUTHERAN HOME)", "Chad Vukelich", "2026-01", False, 0,52,0,0,0,0),
    ("Onslow House", "Tonya Hoffman", "2026-01", False, 0,0,0,0,20,0),
    ("EST- PruittHealth - Neuse", "Tonya Hoffman", "2026-01", True, 65,0,0,0,0,0),
    ("EST- PruittHealth - Trent", "Tonya Hoffman", "2026-01", True, 73,0,0,0,0,0),
    ("EST - JAG Healthcare Union City", "Jami Patterson", "2026-01", True, 0,20,0,0,0,0),
    ("Rehabilitation Center of Cheraw", "Chad Vukelich", "2026-01", False, 0,74,0,0,0,0),
    ("Seven Acres Senior Living", "Adam Sabie", "2026-01", False, 0,30,0,0,0,0),
    ("Green Meadows Health and Rehabilitation", "Adam Sabie", "2026-01", False, 0,45,0,0,0,0),
    ("Trinity Ridge (LUTHERAN HOME)", "Chad Vukelich", "2026-01", False, 0,40,0,0,0,0),
    ("Signature Health Care of Terre Haute", "Adam Sabie", "2026-01", False, 120,0,0,0,0,0),
    ("Richland Bean-Blossom Health Care Center", "Adam Sabie", "2026-01", False, 50,0,0,0,0,0),
    ("Pinewood Nursing Care", "Chad Vukelich", "2026-01", False, 0,45,0,0,0,0),
    ("Elderwood Rehabilitation and Nursing Center", "Chad Vukelich", "2026-02", False, 0,35,0,0,0,0),
    ("Elkhorn Health & Rehabilitation", "Adam Sabie", "2026-02", False, 85,0,50,0,0,0),
    ("EST - Belmont Terrace Nursing and Rehab", "Adam Sabie", "2026-02", True, 0,90,0,0,0,0),
    ("EST - PruittHealth - Raleigh", "Tonya Hoffman", "2026-02", True, 100,0,0,0,0,0),
    ("JAG Healthcare Mansfield", "Adam Sabie", "2026-02", False, 53,0,0,0,0,0),
    ("EST - Brookdale Wooster", "Adam Sabie", "2026-02", True, 0,0,0,0,20,0),
    ("EST Gabriel Manor", "Tonya Hoffman", "2026-02", True, 0,0,0,0,20,0),
    ("EST - Elliott Nursing and Rehabilitation", "Adam Sabie", "2026-03", True, 65,0,0,0,0,0),
    ("Trinity Ridge PCP / NO MD", "Chad Vukelich", "2026-03", False, 104,0,0,0,0,0),
    ("EST-Franklin Manor", "Tonya Hoffman", "2026-03", True, 0,0,0,0,25,0),
    ("EST- Marquette Manor- Memory Care", "Gunner Grider", "2026-03", True, 0,0,0,20,0,0),
    ("Lutheran Life Villages at Kendallville", "Jami Patterson", "2026-03", False, 75,60,0,0,0,0),
    ("Brookdale Harrisonburg", "Melissa Been", "2026-03", False, 0,0,0,0,25,0),
    ("Brookdale Lakeview Crossing", "Adam Sabie", "2026-03", False, 0,0,0,45,0,0),
    ("Terra Bella- Little Ave", "Chad Vukelich", "2026-03", False, 0,0,0,30,21,0),
    ("EST-Cardinal Landing Memory Care", "Samantha Roark", "2026-03", True, 0,0,0,25,0,0),
    ("EST-Green Leaf Care Center", "Tonya Hoffman", "2026-03", True, 0,0,0,0,35,0),
    ("EST - JAG Healthcare Shelby", "Jami Patterson", "2026-03", True, 0,35,0,0,0,0),
]

# Corporate owner mapping (manually resolved during reconciliation)
corp_map = {
    "Gallatin Nursing and Rehab": "PACS GROUP",
    "McCoy Memorial Nursing Center": "CONCIERGE HEALTHCARE",
    "Trinity Village (LUTHERAN HOME)": "LUTHERAN SERVICES CAROLINAS",
    "Onslow House": "ALG SENIOR",
    "EST- PruittHealth - Neuse": "PRUITT HEALTH",
    "EST- PruittHealth - Trent": "PRUITT HEALTH",
    "EST - JAG Healthcare Union City": "JAG",
    "Rehabilitation Center of Cheraw": "FUNDAMENTAL LTC",
    "Seven Acres Senior Living": None,  # Independent
    "Green Meadows Health and Rehabilitation": "LYON HEALTHCARE",
    "Trinity Ridge (LUTHERAN HOME)": "LUTHERAN SERVICES CAROLINAS",
    "Signature Health Care of Terre Haute": "SIGNATURE HEALTH",
    "Richland Bean-Blossom Health Care Center": None,  # Independent
    "Pinewood Nursing Care": "OAK HOLLOW HEALTHCARE MANAGEMENT",
    "Elderwood Rehabilitation and Nursing Center": "OAK HOLLOW HEALTHCARE MANAGEMENT",
    "Elkhorn Health & Rehabilitation": "A&M HEALTHCARE INVESTMENT LLC",
    "EST - Belmont Terrace Nursing and Rehab": "BLUEGRASS/ENCORE",
    "EST - PruittHealth - Raleigh": "PRUITT HEALTH",
    "JAG Healthcare Mansfield": "JAG",
    "EST - Brookdale Wooster": "BROOKDALE SENIOR LIVING",
    "EST Gabriel Manor": "SABER HEALTHCARE GROUP",
    "EST - Elliott Nursing and Rehabilitation": "MAJESTIC CARE",
    "Trinity Ridge PCP / NO MD": "LUTHERAN SERVICES CAROLINAS",
    "EST-Franklin Manor": "SABER HEALTHCARE GROUP",
    "EST- Marquette Manor- Memory Care": "LIFECARE",
    "Lutheran Life Villages at Kendallville": "NORTH WOODS VILLAGE MEMORY CARE",
    "Brookdale Harrisonburg": "BROOKDALE SENIOR LIVING",
    "Brookdale Lakeview Crossing": "BROOKDALE SENIOR LIVING",
    "Terra Bella- Little Ave": "TERRABELLA SENIOR LIVING",
    "EST-Cardinal Landing Memory Care": "TRIPLE CROWN SENIOR LIVING",
    "EST-Green Leaf Care Center": "SABER HEALTHCARE GROUP",
    "EST - JAG Healthcare Shelby": "JAG",
}

# Track Oak Hollow first-win for New Logo vs New Door
oak_hollow_first = True
results = []

for w in wins:
    name, ae, month, is_est, snf_pcp, snf_mh, integ_snf, alf_pcp, alf_mh, integ_alf = w

    ar = (snf_pcp * SNF_PCP + snf_mh * SNF_MH +
          alf_pcp * ALF_PCP + alf_mh * ALF_MH +
          (integ_snf * INTEG_SNF_MH_MULT) * SNF_MH +
          (integ_alf * INTEG_ALF_MH_MULT) * ALF_MH)

    corp = corp_map.get(name)

    if is_est:
        sale_type = "Integration"
    elif corp is None:
        sale_type = "New Logo"
    elif corp == "OAK HOLLOW HEALTHCARE MANAGEMENT":
        if oak_hollow_first:
            sale_type = "New Logo"
            oak_hollow_first = False
        else:
            sale_type = "New Door"
    elif corp.upper() in corps_we_serve:
        sale_type = "New Door"
    else:
        sale_type = "New Logo"

    results.append((name, ae, month, sale_type, ar, corp or "Independent"))

# Detail
print("=== Q1 2026 WINS — CORRECTED SALE TYPE + AR ===\n")
for st in ["New Logo", "New Door", "Integration"]:
    items = [r for r in results if r[3] == st]
    total_ar = sum(r[4] for r in items)
    print(f"--- {st} ({len(items)} facilities, ${total_ar:,.0f} AR) ---")
    for r in items:
        print(f"  {r[0]} | {r[1]} | {r[2]} | Corp: {r[5]} | AR: ${r[4]:,.0f}")
    print()

# Summary — full Q1
total_ar = sum(r[4] for r in results)
print("=== FULL Q1 SUMMARY ===")
for st in ["New Logo", "New Door", "Integration"]:
    items = [r for r in results if r[3] == st]
    type_ar = sum(r[4] for r in items)
    pct = (type_ar / total_ar * 100) if total_ar > 0 else 0
    print(f"  {st}: {len(items)} facilities | ${type_ar:,.0f} | {pct:.1f}%")
print(f"  TOTAL: {len(results)} facilities | ${total_ar:,.0f}")

# Jan+Feb only (deck reporting window)
print("\n=== JAN + FEB ONLY (deck reporting window) ===")
jf = [r for r in results if r[2] in ("2026-01", "2026-02")]
jf_total = sum(r[4] for r in jf)
for st in ["New Logo", "New Door", "Integration"]:
    items = [r for r in jf if r[3] == st]
    type_ar = sum(r[4] for r in items)
    pct = (type_ar / jf_total * 100) if jf_total > 0 else 0
    print(f"  {st}: {len(items)} facilities | ${type_ar:,.0f} | {pct:.1f}%")
print(f"  TOTAL: {len(jf)} facilities | ${jf_total:,.0f}")

# Compare to deck
print("\n=== DECK COMPARISON (Jan+Feb) ===")
print(f"  Deck says: $3.2MM total, New Logo = $1.9MM (61%)")
print(f"  Our calc:  ${jf_total:,.0f} total")
jf_nl = sum(r[4] for r in jf if r[3] == "New Logo")
jf_nd = sum(r[4] for r in jf if r[3] == "New Door")
jf_ig = sum(r[4] for r in jf if r[3] == "Integration")
print(f"  New Logo:     ${jf_nl:,.0f} ({jf_nl/jf_total*100:.1f}%)" if jf_total else "")
print(f"  New Door:     ${jf_nd:,.0f} ({jf_nd/jf_total*100:.1f}%)" if jf_total else "")
print(f"  Integration:  ${jf_ig:,.0f} ({jf_ig/jf_total*100:.1f}%)" if jf_total else "")
