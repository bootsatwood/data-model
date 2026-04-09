#!/usr/bin/env python3
"""
Corporate Dedup Fix — V22.5 -> V22.6

Applies corporate name consolidations identified through the Corporate Name
Deduplication Review (corporate_name_dedup.py + manual RA coding).

Wave 3: Greencroft consolidation.
(Wave 1 = V22.4, Wave 2 = V22.5)

Changes:
  1. Direct corporate name renames (canonical consolidation)
  2. Rerun Four-Rule ownership hierarchy after all changes

Usage:
  python corporate_dedup_fix.py preview    # show changes without writing
  python corporate_dedup_fix.py apply      # write V22.6
"""

import sys
import os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))  # archive: find utils.py
from copy import copy
from collections import Counter

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils import safe, load_db, VAULT, ensure_report_dir

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_5.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_6.xlsx"

# ---------------------------------------------------------------------------
# Corporate Name Consolidation Map
# Key = old name, Value = new canonical name
# ---------------------------------------------------------------------------

CORP_RENAMES = {
    # Greencroft — CMS uses "Greencroft Communities", GLR/served uses "Greencroft"
    "GREENCROFT COMMUNITIES": "GREENCROFT",
    "Greencroft Communities": "GREENCROFT",
}

# ---------------------------------------------------------------------------
# Facility-level reattributions
# (facility name + city + state) -> new corporate name
# These fix GLR misattributions where the facility is correct but the
# corporate parent is wrong.
# ---------------------------------------------------------------------------

FACILITY_REATTRIBUTIONS = [
    # No facility-level reattributions in this wave
]

# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts, served_set):
    """Apply Four-Rule ownership classification.

    Rule 1: 'unknown' or blank -> Independent
    Rule 2: 'INDEPENDENT' -> Independent
    Rule 3: Corp has 2+ facilities in database -> Corporate
    Rule 4: Everything else -> Independent
    """
    if not corp_name or corp_name.upper() in ('UNKNOWN', ''):
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 0) >= 2:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python corporate_dedup_fix.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Corporate Dedup Fix — V22.5 -> V22.6 ({mode} mode)")
    print("=" * 65)

    # Load database
    print("\nLoading V22.3 database...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # --- Phase 1: Corporate Name Renames ---
    print("\nPhase 1: Corporate Name Renames")
    print("-" * 40)

    rename_counts = Counter()
    rename_served = Counter()
    renamed_rows = []

    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp in CORP_RENAMES:
            new_corp = CORP_RENAMES[corp]
            served = safe(r.get('Do_We_Serve', '')) == 'Yes'
            rename_counts[f"{corp} -> {new_corp}"] += 1
            if served:
                rename_served[f"{corp} -> {new_corp}"] += 1
            renamed_rows.append(r)
            r['Corporate_Name'] = new_corp

    print(f"  {len(renamed_rows)} rows renamed across {len(CORP_RENAMES)} consolidation rules")
    print()
    for key in sorted(rename_counts.keys()):
        sv = rename_served.get(key, 0)
        sv_tag = f" ({sv} served)" if sv > 0 else ""
        print(f"    {rename_counts[key]:>4d}  {key}{sv_tag}")

    # --- Phase 2: Facility-Level Reattributions ---
    print(f"\nPhase 2: Facility-Level Reattributions")
    print("-" * 40)

    reattr_count = 0
    for r in rows:
        fac = safe(r.get('Facility_Name', '')).upper()
        city = safe(r.get('City', ''))
        state = safe(r.get('State', ''))
        for rule in FACILITY_REATTRIBUTIONS:
            if (rule["facility_name_contains"].upper() in fac
                    and city.upper() == rule["city"].upper()
                    and state.upper() == rule["state"].upper()):
                old_corp = safe(r.get('Corporate_Name', ''))
                new_corp = rule["new_corp"]
                if old_corp != new_corp:
                    served = safe(r.get('Do_We_Serve', '')) == 'Yes'
                    sv_tag = " *SERVED*" if served else ""
                    print(f"  {safe(r.get('Facility_Name',''))}, {city}, {state}")
                    print(f"    {old_corp} -> {new_corp}{sv_tag}")
                    print(f"    Reason: {rule['reason']}")
                    r['Corporate_Name'] = new_corp
                    reattr_count += 1

    print(f"\n  {reattr_count} facility reattributions applied")

    # --- Phase 3: Rerun Four-Rule Hierarchy (scoped to changed rows only) ---
    print(f"\nPhase 3: Four-Rule Ownership Reclassification")
    print("-" * 40)

    # Rebuild counts with new corporate names
    corp_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp:
            corp_counts[corp] += 1

    # Only reclassify rows whose Corporate_Name actually changed
    # (avoids touching pre-existing issues like 794 "unknown" = Corporate)
    changed_corps = set(CORP_RENAMES.keys()) | set(CORP_RENAMES.values())
    for rule in FACILITY_REATTRIBUTIONS:
        changed_corps.add(rule['new_corp'])

    reclass_count = 0
    reclass_served = 0
    reclass_detail = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp not in changed_corps:
            continue
        old_type = safe(r.get('Ownership_Type', ''))
        new_type = four_rule_classify(corp, corp_counts, None)
        if old_type != new_type:
            served = safe(r.get('Do_We_Serve', '')) == 'Yes'
            reclass_count += 1
            if served:
                reclass_served += 1
            reclass_detail[f"{old_type} -> {new_type}"] += 1
            r['Ownership_Type'] = new_type

    print(f"  {reclass_count} reclassifications ({reclass_served} served)")
    for key in sorted(reclass_detail.keys()):
        print(f"    {reclass_detail[key]:>4d}  {key}")

    # --- Summary ---
    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Corporate Name renames:        {len(renamed_rows)}")
    print(f"  Facility reattributions:       {reattr_count}")
    print(f"  Ownership reclassifications:   {reclass_count}")
    total_changes = len(renamed_rows) + reattr_count
    print(f"  Total rows modified:           {total_changes}")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY — no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.4 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    # Reopen source workbook (preserving formatting)
    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    # Build header index
    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    corp_col = col_idx.get('Corporate_Name')
    own_col = col_idx.get('Ownership_Type')

    if not corp_col or not own_col:
        print("ERROR: Could not find Corporate_Name or Ownership_Type columns")
        sys.exit(1)

    # Apply changes row by row
    changes_written = 0
    for r in rows:
        excel_row = r['_excel_row']
        # Corporate_Name
        old_corp = safe(ws.cell(row=excel_row, column=corp_col).value)
        new_corp = safe(r.get('Corporate_Name', ''))
        if old_corp != new_corp:
            ws.cell(row=excel_row, column=corp_col).value = new_corp
            changes_written += 1
        # Ownership_Type
        old_own = safe(ws.cell(row=excel_row, column=own_col).value)
        new_own = safe(r.get('Ownership_Type', ''))
        if old_own != new_own:
            ws.cell(row=excel_row, column=own_col).value = new_own

    wb.save(DB_OUTPUT)
    print(f"  {changes_written} cells written")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
