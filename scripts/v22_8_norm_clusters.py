#!/usr/bin/env python3
"""
Exact Norm Cluster Consolidation — V22.7 -> V22.8

Applies:
  1. 20 exact norm cluster consolidations (spelling/case/formatting variants)
  2. 4 obvious typo fixes from fuzzy match review
  3. Four-Rule ownership reclassification for affected groups

Usage:
  python v22_8_norm_clusters.py preview
  python v22_8_norm_clusters.py apply
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, VAULT

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_7.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_8.xlsx"

# ---------------------------------------------------------------------------
# Corporate Name Renames — Exact Norm Clusters
# Format: old_name -> canonical
# ---------------------------------------------------------------------------

NORM_CLUSTER_RENAMES = {
    # Cluster 1: GLR canonical
    "PRUITTHEALTH": "PRUITT HEALTH",
    # Cluster 2: GLR canonical
    "SUNRISE SENIOR LIVING": "Sunrise Senior Living",
    # Cluster 3: GLR canonical
    "OTTERBEIN SENIORLIFE": "OTTERBEIN SENIOR LIFE",
    # Cluster 4: GLR canonical
    "MILLER'S MERRY MANOR": "MILLERS MERRY MANOR",
    # Cluster 9: GLR canonical
    "KISSITO HEALTHCARE": "Kissito Healthcare",
    # Cluster 10: GLR canonical
    "LYON HEALTHCARE": "Lyon Healthcare",
    # Cluster 14: GLR canonical (remove trailing "?")
    "BLUE CENTRAL GROUP?": "BLUE CENTRAL GROUP",
    # Cluster 12: GLR canonical
    "CrownPointe Communities": "CROWN POINTE COMMUNITIES",
    # Cluster 20: GLR canonical
    "MP CARE II, LLC": "MPCARE II, LLC",
    # Cluster 5: no GLR, use highest count
    "PRESBYTERIAN SENIORCARE NETWORK": "Presbyterian Senior Care Network",
    # Cluster 6: no GLR, use highest count
    "BeehiveHomes": "BeeHive Homes",
    # Cluster 7: no GLR, use highest count
    "JUNIPER COMMUNITIES": "Juniper Communities",
    # Cluster 8: no GLR, use highest count
    "MERRILL GARDENS L.L.C": "Merrill Gardens",
    # Cluster 11: no GLR, use highest count
    "Citizens Memorial Healthcare": "CITIZENS MEMORIAL HEALTH CARE",
    # Cluster 15: no GLR, use highest count
    "Prelude Homes and Services": "Prelude Homes & Services",
    # Cluster 16: no GLR, use highest count
    "MILL CREEK MANOR, LLC": "MILLCREEK MANOR",
    # Cluster 17: no GLR, use highest count
    "Caring Place Healthcare Group": "CARING PLACE HEALTHCARE GROUP",
    # Cluster 18: no GLR, use highest count
    "North Port Retirement Centers, Inc.": "Northport Retirement Centers, Inc",
    # Cluster 19: no GLR, use highest count
    "Rest Haven Homes, Inc.": "Resthaven Homes, Inc",
}

# ---------------------------------------------------------------------------
# Typo Fixes — from fuzzy match review (obvious misspellings)
# ---------------------------------------------------------------------------

TYPO_FIXES = {
    # Missing 'S' in MAJESTIC
    "MAJETIC CARE": "MAJESTIC CARE",
    # Missing 'I' in SENIOR (same pattern as Cedarhurst SENOR fix in V22.7)
    "HARMONY SENOR SERVICES": "HARMONY SENIOR SERVICES",
    # Singular -> plural (14 rows vs 1)
    "EXCEPTIONAL LIVING CENTER": "EXCEPTIONAL LIVING CENTERS",
    # SENIORS -> SENIOR (3 rows 2 served vs 3 rows 0 served)
    "OAKDALE SENIORS ALLIANCE": "OAKDALE SENIOR ALLIANCE",
}

# Combined rename map
ALL_RENAMES = {**NORM_CLUSTER_RENAMES, **TYPO_FIXES}


# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    if not corp_name or corp_name.upper() in ('UNKNOWN', ''):
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 0) >= 2:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python v22_8_norm_clusters.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Exact Norm Cluster Consolidation — V22.7 -> V22.8 ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_SOURCE.name}...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # --- Phase 1: Corporate Name Renames ---
    print("\nPhase 1: Corporate Name Renames")
    print("-" * 40)

    rename_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp in ALL_RENAMES:
            new_corp = ALL_RENAMES[corp]
            rename_counts[f"{corp} -> {new_corp}"] += 1
            r['Corporate_Name'] = new_corp

    print("\n  Exact Norm Clusters:")
    norm_total = 0
    for old, new in sorted(NORM_CLUSTER_RENAMES.items()):
        key = f"{old} -> {new}"
        if rename_counts.get(key, 0) > 0:
            print(f"    {rename_counts[key]:>4d}  {old} -> {new}")
            norm_total += rename_counts[key]
    print(f"    {norm_total} norm cluster renames")

    print("\n  Typo Fixes:")
    typo_total = 0
    for old, new in sorted(TYPO_FIXES.items()):
        key = f"{old} -> {new}"
        if rename_counts.get(key, 0) > 0:
            print(f"    {rename_counts[key]:>4d}  {old} -> {new}")
            typo_total += rename_counts[key]
    print(f"    {typo_total} typo fix renames")

    total_renames = sum(rename_counts.values())
    print(f"\n  {total_renames} total renames")

    # --- Phase 2: Four-Rule Reclassification ---
    print(f"\nPhase 2: Four-Rule Ownership Reclassification")
    print("-" * 40)

    corp_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp:
            corp_counts[corp] += 1

    # Only reclassify rows with affected corporate names
    affected_corps = set(ALL_RENAMES.keys()) | set(ALL_RENAMES.values())
    reclass_count = 0
    reclass_detail = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp not in affected_corps:
            continue
        old_type = safe(r.get('Ownership_Type', ''))
        new_type = four_rule_classify(corp, corp_counts)
        if old_type != new_type:
            reclass_count += 1
            reclass_detail[f"{old_type} -> {new_type}"] += 1
            print(f"    {safe(r.get('Facility_Name',''))}, {safe(r.get('City',''))}, {safe(r.get('State',''))}")
            print(f"      Corp: {corp} | {old_type} -> {new_type}")
            r['Ownership_Type'] = new_type

    print(f"\n  {reclass_count} reclassifications")
    for key in sorted(reclass_detail.keys()):
        print(f"    {reclass_detail[key]:>4d}  {key}")

    # --- Summary ---
    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Norm cluster renames:          {norm_total}")
    print(f"  Typo fix renames:              {typo_total}")
    print(f"  Total renames:                 {total_renames}")
    print(f"  Ownership reclassifications:   {reclass_count}")
    print(f"  Rows deleted:                  0")
    print(f"  Net row delta:                 0")
    print(f"  Final row count:               {len(rows):,}")

    # --- Post-consolidation corporate name counts ---
    print(f"\nPost-consolidation counts for affected names:")
    for canonical in sorted(set(ALL_RENAMES.values())):
        cnt = corp_counts.get(canonical, 0)
        served = sum(1 for r in rows if safe(r.get('Corporate_Name', '')) == canonical
                     and safe(r.get('Do_We_Serve', '')) == 'Yes')
        print(f"    {cnt:>4d} ({served:>2d} served)  {canonical}")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY -- no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.8 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    corp_col = col_idx.get('Corporate_Name')
    own_col = col_idx.get('Ownership_Type')

    cells_updated = 0
    for r in rows:
        excel_row = r['_excel_row']

        # Corporate_Name
        old_corp = safe(ws.cell(row=excel_row, column=corp_col).value)
        new_corp = safe(r.get('Corporate_Name', ''))
        if old_corp != new_corp:
            ws.cell(row=excel_row, column=corp_col).value = new_corp
            cells_updated += 1

        # Ownership_Type
        old_own = safe(ws.cell(row=excel_row, column=own_col).value)
        new_own = safe(r.get('Ownership_Type', ''))
        if old_own != new_own:
            ws.cell(row=excel_row, column=own_col).value = new_own
            cells_updated += 1

    wb.save(DB_OUTPUT)
    print(f"  {cells_updated} cells updated")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
