#!/usr/bin/env python3
"""
IL/55+ Source_Type Reclassification — V22.10 -> V22.11

Reclassifies HIGH-confidence non-ALF facilities from Source_Type ALF to ILF.
These are Independent Living, Senior Apartments, Active Adult, and 55+ communities
that were imported from NIC Maps as ALFs but are not licensed Assisted Living Facilities.

Categories:
  - IL_ABBREVIATION: Campus IL components with "-IL" suffix (9 rows, 8 served)
  - INDEPENDENT_LIVING: Explicitly named IL facilities (8 rows, 1 served)
  - SENIOR_APARTMENTS: Age-restricted apartment buildings (25 rows, 0 served)
  - ACTIVE_ADULT/ADULT_COMMUNITY: Active adult communities (7 rows, 0 served)
  - AGE_55_PLUS: 55+ communities (2 rows, 0 served)
  - ADULT_LIVING: Adult living facility (1 row, 0 served)
  - CCRC_IL: CCRC apartment component (1 row, 1 served)

Deferred (8 rows):
  - WOODLAND HILLS INDEPENDENT LIVING, AL & MC (served, has AL & MC)
  - THE WILLOWS OF EASLEY & ASSISTED & INDEPENDENT LIVING (has "Assisted")
  - HOLLY MANOR AL IL OPERATIONS LLC (ambiguous "AL IL")
  - CCRC I-DOVER NURSING HOME (might be SNF, not IL)
  - CENTRAL CONTINUING CARE-SNF/NF (served, has SNF/NF in name)
  - TRINITY OAKS CONTINUING CARE RETIREMENT COMMUNITY x2 (ambiguous campus)
  - Collington Continuing Care Retirement Community (ambiguous campus)

Usage:
  python v22_11_il_reclassify.py preview
  python v22_11_il_reclassify.py apply
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, norm, load_db, VAULT

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_10.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_11.xlsx"

# ---------------------------------------------------------------------------
# Facilities to reclassify: (Facility_Name, State) -> ALF to ILF
# Matched by normalized name + state to handle row number shifts
# ---------------------------------------------------------------------------

RECLASSIFY_TO_ILF = [
    # --- IL_ABBREVIATION (9 rows, 8 served) ---
    ("GREENWOOD VILLAGE SOUTH IL", "IN"),
    ("MARQUETTE MANOR IL", "IN"),
    ("BENTLEY COMMONS-IL", "VA"),
    ("HARMONY ROANOKE - IL", "VA"),
    ("KENDAL IL", "VA"),
    ("RUNK & PRATT LIBERTY RIDGE-IL", "VA"),
    ("THE LODGE AT OLD TRAIL-IL", "VA"),
    ("THE SUMMIT-IL", "VA"),
    ("HARMONY COLLECTION AT COLUMBIA (IL)", "SC"),

    # --- INDEPENDENT_LIVING (8 rows, 1 served) ---
    ("GLEBE-INDEPENDENT LIVING", "VA"),
    ("CHANDLER PLACE INDEPENDENT LIVING", "SC"),
    ("PRIME INDEPENDENT LIVING GASTONIA", "NC"),
    ("SOUTHERN HERITAGE INDEPENDENT LIVING", "TN"),
    ("THE ARBORS INDEPENDENT LIVING", "VA"),
    ("THE HARMONY COLLECTION AT ROANOKE INDEPENDENT LIVING", "VA"),
    ("THE WOODLANDS INDEPENDENT LIVING AT OTTERBEIN SUNSET HOUSE", "OH"),
    ("TRADITIONS OF HERSHEY INDEPENDENT LIVING", "PA"),

    # --- SENIOR_APARTMENTS (25 rows, 0 served) ---
    ("900 ACQUA SENIOR APARTMENTS", "VA"),
    ("BROOKHAVEN AT COUNTY LINE SENIOR APARTMENTS", "IN"),
    ("CAVALIER SENIOR APARTMENTS", "VA"),
    ("CHAUNCEY POINTE SENIOR APARTMENTS", "OH"),
    ("CROMWELL HOUSE SENIOR APARTMENTS - 55+", "VA"),
    ("FAIRFIELD VILLAGE SENIOR APARTMENTS", "OH"),
    ("FIELDCHASE SENIOR APARTMENTS", "OH"),
    ("GARDENS ON GATEWAY SENIOR APARTMENTS", "IN"),
    ("GATEWAY POINTE SENIOR APARTMENTS 55+", "OH"),
    ("MANCHESTER LAKES SENIOR APARTMENT HOMES", "VA"),
    ("MERIDIAN HILLS SENIOR APARTMENTS", "KY"),
    ("MOFFETT MANOR SENIOR APARTMENTS", "VA"),
    ("NORA COMMONS ON THE MONON SENIOR APARTMENTS", "IN"),
    ("OAK HILL SENIOR APARTMENTS", "PA"),
    ("PARK PLACE SENIOR APARTMENTS", "VA"),
    ("PARSONS VILLAGE EAST SENIOR APARTMENTS", "OH"),
    ("PLAZA TOWERS SENIOR APARTMENTS", "TN"),
    ("SADDLEBROOK SENIOR APARTMENTS", "TN"),
    ("SENIOR APARTMENTS", "OH"),
    ("SYCAMORE CREEK SENIOR APARTMENTS", "OH"),
    ("THE ARBORS SENIOR APARTMENTS", "VA"),
    ("THE GABLES AT DRUID HILLS SENIOR APARTMENTS", "NC"),
    ("THE OAKS OF WELLINGTON - 55 & OLDER SENIOR APARTMENTS", "VA"),
    ("TUCKER STATION SENIOR APARTMENTS", "KY"),
    ("WESTWOOD PLACE SENIOR APARTMENTS", "OH"),

    # --- ACTIVE_ADULT / ADULT_COMMUNITY (7 rows, 0 served) ---
    ("VITALIA ACTIVE ADULT COMMUNITY AT MONTROSE", "OH"),
    ("VITALIA ACTIVE ADULT COMMUNITY AT NORTH OLMSTED", "OH"),
    ("VITALIA ACTIVE ADULT COMMUNITY AT ROCKSIDE", "OH"),
    ("VITALIA ACTIVE ADULT COMMUNITY ROCKSIDE", "OH"),
    ("VITALIA ACTIVE ADULT COMMUNITY AT NORTH ROYALTON", "OH"),
    ("VITALIA ACTIVE ADULT COMMUNITY AT ROCKSIDE (SEVEN HILLS)", "OH"),
    ("Autumn Fields Adult Community", "IL"),

    # --- AGE_55_PLUS (2 rows not already in SENIOR_APARTMENTS) ---
    ("GLENNS CREEK MANOR SENIOR LIVING 55+ APARTMENTS", "KY"),
    ("HORIZON HOMES (55+ COMMUNITY) APARTMENTS", "IN"),

    # --- ADULT_LIVING (1 row, 0 served) ---
    ("ADULT LIVING AT ROSEBROOK", "PA"),

    # --- CCRC IL component (1 row, 1 served) ---
    ("CCRC I-DOVER APARTMENTS", "OH"),
]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python v22_11_il_reclassify.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"IL/55+ Source_Type Reclassification — V22.10 -> V22.11 ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_SOURCE.name}...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # Build lookup: norm(name)|state -> list of rows
    name_state_index = {}
    for r in rows:
        fn = norm(safe(r.get('Facility_Name', '')))
        st = safe(r.get('State', ''))
        key = fn + '|' + st
        name_state_index.setdefault(key, []).append(r)

    # --- Phase 1: Source_Type Reclassification ---
    print("\nPhase 1: Source_Type ALF -> ILF Reclassification")
    print("-" * 50)

    reclass_count = 0
    not_found = []
    already_ilf = []
    wrong_type = []
    served_reclass = 0
    state_counts = Counter()

    for facility_name, state in RECLASSIFY_TO_ILF:
        key = norm(facility_name) + '|' + state
        candidates = name_state_index.get(key, [])

        if not candidates:
            not_found.append((facility_name, state))
            continue

        for r in candidates:
            old_type = safe(r.get('Source_Type', ''))
            if old_type == 'ILF':
                already_ilf.append((facility_name, state))
                continue
            if old_type != 'ALF':
                wrong_type.append((facility_name, state, old_type))
                continue

            r['Source_Type'] = 'ILF'
            reclass_count += 1
            state_counts[state] += 1
            served = safe(r.get('Do_We_Serve', ''))
            if served == 'Yes':
                served_reclass += 1
                print(f"  * Row {r['_excel_row']:>5d}  ALF -> ILF  {facility_name:<55s}  {state}  SERVED")
            else:
                print(f"    Row {r['_excel_row']:>5d}  ALF -> ILF  {facility_name:<55s}  {state}")

    print(f"\n  {reclass_count} rows reclassified (ALF -> ILF)")
    print(f"  {served_reclass} of those are served")
    print(f"  States: {dict(sorted(state_counts.items()))}")

    if not_found:
        print(f"\n  WARNING: {len(not_found)} facilities not found in DB:")
        for name, st in not_found:
            print(f"    {name} ({st})")

    if already_ilf:
        print(f"\n  NOTE: {len(already_ilf)} already ILF (no change):")
        for name, st in already_ilf:
            print(f"    {name} ({st})")

    if wrong_type:
        print(f"\n  WARNING: {len(wrong_type)} unexpected Source_Type (skipped):")
        for name, st, typ in wrong_type:
            print(f"    {name} ({st}) = {typ}")

    # --- Summary ---
    # Count Source_Type totals after changes
    type_counts = Counter()
    for r in rows:
        type_counts[safe(r.get('Source_Type', ''))] += 1

    served_ilf = sum(1 for r in rows
                     if safe(r.get('Source_Type', '')) == 'ILF'
                     and safe(r.get('Do_We_Serve', '')) == 'Yes')

    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Source_Type reclassifications:  {reclass_count} (ALF -> ILF)")
    print(f"  Served rows reclassified:      {served_reclass}")
    print(f"  Rows deleted:                  0")
    print(f"  Final row count:               {len(rows):,}")
    print(f"\n  Source_Type breakdown:")
    for t in sorted(type_counts.keys()):
        print(f"    {t}: {type_counts[t]:,}")
    print(f"\n  ILF total: {type_counts.get('ILF', 0)} ({served_ilf} served)")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY -- no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.11 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    source_type_col = col_idx.get('Source_Type')
    if not source_type_col:
        print("  ERROR: Source_Type column not found!")
        sys.exit(1)

    cells_updated = 0
    for r in rows:
        excel_row = r['_excel_row']
        old_val = safe(ws.cell(row=excel_row, column=source_type_col).value)
        new_val = safe(r.get('Source_Type', ''))
        if old_val != new_val:
            ws.cell(row=excel_row, column=source_type_col).value = new_val
            cells_updated += 1

    wb.save(DB_OUTPUT)
    print(f"  {cells_updated} cells updated")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
