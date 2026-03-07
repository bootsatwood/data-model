"""Cross-reference V20 tiering (70 scored) vs V6 tiering (45 scored).
Shows who's carried over, who's dropped, who's new, who moved to T5."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

# Load V20
wb20 = openpyxl.load_workbook('C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx', data_only=True)
v20_entities = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb20[sn]
    tier = sn.split()[0]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name and str(name).strip():
            score = ws.cell(r, 3).value
            v20_entities[str(name).strip()] = {'tier': tier, 'score': score}

v20_t5 = []
ws5 = wb20['T5- Hard Barriers ']
for r in range(2, ws5.max_row + 1):
    name = ws5.cell(r, 1).value
    barrier = ws5.cell(r, 2).value
    if name and str(name).strip() and str(name).strip() != 'TOTAL':
        v20_t5.append((str(name).strip(), str(barrier or '').strip()))

# Load V6
# Try V6, fall back to V5 if locked
try:
    wb6 = openpyxl.load_workbook('C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/MUO_Scoring_Workbook_V23_v6.xlsx', data_only=True)
except PermissionError:
    wb6 = openpyxl.load_workbook('C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/MUO_Scoring_Workbook_V23_v5.xlsx', data_only=True)
    print("NOTE: V6 locked, using V5 (same entities, old thresholds)")
ws6 = wb6['Summary']
v6_entities = {}
for r in range(2, ws6.max_row + 1):
    name = ws6.cell(r, 1).value
    if not name: break
    v6_entities[name] = {'tier': ws6.cell(r, 16).value, 'score': ws6.cell(r, 15).value}

# V20 name -> V6 name crosswalk
V20_TO_V6 = {
    'ALG': 'ALG',
    'American Senior Communities': 'AMERICAN SENIOR COMMUNITIES',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'Cch Healthcare': 'CCH HEALTHCARE',
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': 'Avardis',
    'Infinity Healthcare Consulting': 'INFINITY HEALTHCARE CONSULTING',
    'Liberty Senior Living': 'Liberty',
    'Life Care Centers Of America': 'Lifecare',
    'Majestic Care': 'Majestic Care',
    'Navion': 'NAVION',
    'Otterbein Seniorlife': 'OTTERBEIN SENIOR LIFE',
    'Principle Long Term Care': 'PRINCIPLE',
    'Pruitthealth': 'PRUITT HEALTH',
    'Saber Healthcare Group': 'SABER HEALTHCARE GROUP',
    'Terra Bella': 'TERRABELLA SENIOR LIVING',
    'Tlc Management': 'TLC Management',
    'Trilogy Health Services': 'TRILOGY',
    'Arbors At Ohio': 'ARBORS',
    'BHI Senior Living': 'BHI Senior Living',
    'Castle Healthcare': 'Castle Healthcare',
    'Jag Healthcare': 'JAG',
    'Kissito': 'Kissito Healthcare',
    'Lionstone Care': 'LIONSTONE CARE',
    'Lutheran Services Carolina': 'Lutheran Services Carolinas',
    'Morning Pointe Senior Living': 'MORNING POINTE SENIOR LIVING',
    'Peak Resources, Inc.': 'PEAK RESOURCES',
    'Priority Life Care': 'PRIORITY',
    'Sanstone Health & Rehabilitation': 'SANSTONE',
    'Topaz': 'TOPAZ HEALTHCARE',
    'Yad Healthcare': 'YAD',
    'Caring Place Healthcare': 'CARING PLACE HEALTHCARE',
}

# V20 entities that moved to T5 in V6
MOVED_TO_T5 = {'Pavilion Healthcare', 'Clearview'}

# Build reverse lookup
v6_from_v20 = set(V20_TO_V6.values())

print("=" * 110)
print("INVENTORY: V20 (70 scored) vs V6 (45 scored)")
print("=" * 110)

# 1. In both
print()
print("--- RESCORED: In both V20 and V6 ---")
print(f"{'V20 Name':<55} {'V20T':>4} {'V20s':>5}  {'V6T':>4} {'V6s':>5}  Tier Change")
print("-" * 100)
both = 0
for v20_name in sorted(v20_entities.keys()):
    v6_name = V20_TO_V6.get(v20_name)
    if v6_name and v6_name in v6_entities:
        v20d = v20_entities[v20_name]
        v6d = v6_entities[v6_name]
        change = ''
        if v20d['tier'] != v6d['tier']:
            change = f"{v20d['tier']} -> {v6d['tier']}"
        print(f"{v20_name:<55} {v20d['tier']:>4} {v20d['score']:>5}  {v6d['tier']:>4} {v6d['score']:>5}  {change}")
        both += 1
print(f"Count: {both}")

# 2. Moved to T5
print()
print("--- MOVED TO T5 IN V6 (were scored in V20) ---")
for v20_name in sorted(MOVED_TO_T5):
    if v20_name in v20_entities:
        v20d = v20_entities[v20_name]
        print(f"  {v20_name:<55} was {v20d['tier']} (score {v20d['score']}) -> T5 barrier in V6")

# 3. V20 scored but not in Finance 60
print()
print("--- DROPPED: In V20 but NOT in Finance 60 universe ---")
dropped = []
for v20_name in sorted(v20_entities.keys()):
    v6_name = V20_TO_V6.get(v20_name)
    if v6_name is None and v20_name not in MOVED_TO_T5:
        v20d = v20_entities[v20_name]
        dropped.append((v20_name, v20d['tier'], v20d['score']))
        print(f"  {v20_name:<55} was {v20d['tier']} (score {v20d['score']})")
print(f"Count: {len(dropped)}")

# 4. New in V6
print()
print("--- NEW: In V6 but NOT in V20 ---")
new = []
for v6_name in sorted(v6_entities.keys()):
    if v6_name not in v6_from_v20:
        v6d = v6_entities[v6_name]
        new.append((v6_name, v6d['tier'], v6d['score']))
        print(f"  {v6_name:<55} {v6d['tier']:>4} (score {v6d['score']})")
print(f"Count: {len(new)}")

# 5. V20 T5 barriers
print()
print("--- V20 T5 BARRIERS ---")
v6_t5 = {'Hill Valley', 'Cardon & Associates', 'Signature Health', 'CommuniCare',
          'Bluegrass/Encore', 'MFA', 'Singh', 'Eastern Healthcare Group',
          'Clearview', 'Pavilion Healthcare'}
for name, barrier in v20_t5:
    status = "(still T5)" if any(name.lower() in t5.lower() for t5 in v6_t5) else "(NOT in V6 T5)"
    print(f"  {name:<30} {barrier:<25} {status}")

# Summary
print()
print("=" * 110)
print("SUMMARY")
print("=" * 110)
t1 = sum(1 for e in v6_entities.values() if e['tier'] == 'T1')
t2 = sum(1 for e in v6_entities.values() if e['tier'] == 'T2')
t3 = sum(1 for e in v6_entities.values() if e['tier'] == 'T3')
print(f"V20 scored:  70 entities (T1=22, T2=36, T3=12)")
print(f"V6 scored:   {len(v6_entities)} entities (T1={t1}, T2={t2}, T3={t3})")
print(f"  Rescored:  {both} entities (in both V20 and V6)")
print(f"  Dropped:   {len(dropped)} entities (V20 scored, not in Finance 60)")
print(f"  To T5:     2 entities (Pavilion, Clearview — barrier flags)")
print(f"  New:       {len(new)} entities (in Finance 60, not in V20)")
print()
print("COVERAGE GAP: {0} V20 entities not yet rescored in V6".format(len(dropped)))
print("These were scored in V20 but are NOT in the Finance 60 universe.")
print("Decision needed: are they still relevant, or is Finance 60 the correct universe?")
