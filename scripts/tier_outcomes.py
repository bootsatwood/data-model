"""Show FINAL TIER outcomes (T1/T2/T3) for Finance 60, excluding T4/T5.
Tests different ER+RP bracket combinations."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

# Entities to EXCLUDE from scoring (T4/T5/gate failures)
EXCLUDE = {
    # T5 - Barriers
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    # T4 - Independents
    'Sunnyside Communities', 'ATRIUM HEALTH', 'Warm Hearth Village', 'Brighton',
    # Gate failures (<7 FP campuses)
    'Momentus Health',
}

# Finance entity -> DB corporate names
FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'CLEARVIEW': {'CLEARVIEW'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'CARDON & ASSOCIATES': {'CARDON & ASSOCIATES'},
}

# --- Load Finance entity list ---
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']

finance_entities = []
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    tier = ws_a.cell(r, 18).value
    glr_facs = ws_a.cell(r, 4).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in EXCLUDE:
        finance_entities.append({
            'name': str(name).strip(),
            'fin_tier': tier,
            'glr_facs': int(glr_facs) if glr_facs else 0,
        })

# --- Load V23 DB ---
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb_db.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served:
        d['served_camps'].add((addr, city))

# --- Match ---
entities = []
for fe in finance_entities:
    db_names = FINANCE_TO_DB.get(fe['name'])
    if not db_names:
        continue
    camp = set()
    srv = set()
    rev = 0
    for dn in db_names:
        d = corp_fp.get(dn)
        if d:
            camp |= d['campuses']
            srv |= d['served_camps']
            rev += d['tot_pot']
    if len(camp) < 7:
        continue  # MUO gate
    entities.append({
        'name': fe['name'],
        'fin_tier': fe['fin_tier'],
        'camp': len(camp),
        'srv': len(srv),
        'rev': rev,
    })

print(f"Scoreable entities (excl T4/T5/barriers, >=7 campuses): {len(entities)}")
print()

# --- Finance tier distribution for comparison ---
fin_t1 = sum(1 for e in entities if str(e['fin_tier']).strip() == '1')
fin_t2 = sum(1 for e in entities if str(e['fin_tier']).strip() == '2')
fin_t3 = sum(1 for e in entities if str(e['fin_tier']).strip() == '3')
fin_new = sum(1 for e in entities if 'not on list' in str(e['fin_tier']).lower())
fin_other = len(entities) - fin_t1 - fin_t2 - fin_t3 - fin_new
print(f"Finance current: T1={fin_t1}  T2={fin_t2}  T3={fin_t3}  Not listed={fin_new}  Other={fin_other}")
print()

# --- Scoring functions ---
def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def rs_score(srv, total):
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

# Using placeholder qualitative scores
DEFAULT_IR = 2
DEFAULT_SI = 2
DEFAULT_AI = 0

er_options = {
    'V20 (5/10/20/40)': [5, 10, 20, 40],
    'A (10/15/25/50)': [10, 15, 25, 50],
    'C (9/13/20/40)': [9, 13, 20, 40],
}

rp_options = {
    'V20 ($250K/$500K/$1M/$2M)': [250_000, 500_000, 1_000_000, 2_000_000],
    'B ($1M/$2.5M/$5M/$10M)': [1_000_000, 2_500_000, 5_000_000, 10_000_000],
    'C ($2M/$4M/$7M/$12M)': [2_000_000, 4_000_000, 7_000_000, 12_000_000],
    'E ($2M/$5M/$10M/$20M)': [2_000_000, 5_000_000, 10_000_000, 20_000_000],
}

# --- Test all combos ---
print("=" * 95)
print(f"{'ER Brackets':<22} {'RP Brackets':<30} {'T1':>5} {'T2':>5} {'T3':>5}   Shape")
print("=" * 95)

for er_name, er_t in er_options.items():
    for rp_name, rp_t in rp_options.items():
        tiers = Counter()
        for e in entities:
            er = score_dim(e['camp'], er_t)
            rp = score_dim(e['rev'], rp_t)
            rs = rs_score(e['srv'], e['camp'])
            score = (er * 4) + (DEFAULT_IR * 3) + (DEFAULT_SI * 3) + (rp * 4) + (rs * 3) + (DEFAULT_AI * 3)
            if score >= 50: tiers['T1'] += 1
            elif score >= 25: tiers['T2'] += 1
            else: tiers['T3'] += 1

        t1, t2, t3 = tiers.get('T1', 0), tiers.get('T2', 0), tiers.get('T3', 0)
        total = t1 + t2 + t3

        # Rate how close to Finance distribution (T1=18, T2=15, T3=6 -> ~46% T1, 38% T2, 15% T3)
        shape = ""
        if t3 == 0:
            shape = "no T3"
        elif t1 > t2:
            shape = "top-heavy"
        elif t2 > t1 * 3:
            shape = "middle-heavy"
        else:
            shape = "balanced"

        print(f"{er_name:<22} {rp_name:<30} {t1:>5} {t2:>5} {t3:>5}   {shape}")

# --- Show best combo in detail ---
print()
print("=" * 95)
print("DETAILED VIEW: ER-C (9/13/20/40) + RP-B ($1M/$2.5M/$5M/$10M)")
print("=" * 95)

er_t = [9, 13, 20, 40]
rp_t = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

detail = []
for e in entities:
    er = score_dim(e['camp'], er_t)
    rp = score_dim(e['rev'], rp_t)
    rs = rs_score(e['srv'], e['camp'])
    score = (er * 4) + (DEFAULT_IR * 3) + (DEFAULT_SI * 3) + (rp * 4) + (rs * 3) + (DEFAULT_AI * 3)
    tier = 'T1' if score >= 50 else 'T2' if score >= 25 else 'T3'
    detail.append({**e, 'er': er, 'rp': rp, 'rs': rs, 'score': score, 'tier': tier})

detail.sort(key=lambda x: (-x['score'], x['name']))

print(f"\n{'Entity':<45} {'Fin':>4} {'V23':>4} {'Scr':>4}  {'Camp':>5} {'Srv':>4} {'ER':>3} {'RP':>3} {'RS':>3}  {'FP Revenue':>13}")
print("-" * 110)
for d in detail:
    ft = str(d['fin_tier'] or '?').strip()
    if 'not on list' in ft.lower():
        ft = 'NEW'
    elif ft in ('1','2','3'):
        ft = f'T{ft}'
    print(f"{d['name']:<45} {ft:>4} {d['tier']:>4} {d['score']:>4}  {d['camp']:>5} {d['srv']:>4} {d['er']:>3} {d['rp']:>3} {d['rs']:>3}  ${d['rev']:>12,.0f}")

tiers = Counter(d['tier'] for d in detail)
print("-" * 110)
print(f"TOTALS: T1={tiers.get('T1',0)}  T2={tiers.get('T2',0)}  T3={tiers.get('T3',0)}")
