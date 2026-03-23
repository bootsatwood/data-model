"""
V25.1 Migration — Data Quality Cleanup (Tiering Analysis + Audit Report Fixes)

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25.xlsx
Output: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_1.xlsx

Content-based matching (name_frag + city + state + corp) for safety.
Already-applied fixes will report NOT FOUND and be skipped — safe to re-run.

Parts:
  A: Avardis legacy name consolidation (~40 rows)
  B: Spring Harbor typo fix (1 row)
  C: Scoring board duplicate consolidations (13 pairs, ~130 rows)
  D: Cedarhurst facility naming normalization
  E: Greencroft/Oak Grove NC (delete ALF dupe + update SNF corp)
  F: Atrium swing bed removal (1 row)
  G: Gardant fixes (Regency reattrib, name consolidation, entity merge, bed count)
  H: Vitality/Triple Crown fixes (reattribs, deletes, entity merge, missing facs)
  I: RUI fixes (phantom delete, dup deletes, reattribs, entity merge, county fills)
  J: Barrier flag fills (16 T5 entities)
  K: Address normalization (street suffix standardization)
  L: Campus/Site ID column (computed from address + GPS)
  M: Miscellaneous field fixes (city typos, bed counts, county fills)
"""
import openpyxl
import re
import os
import math
from collections import defaultdict

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
VAULT = os.path.expanduser(
    "~/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current"
)
V25_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25.xlsx")
V25_1_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_1.xlsx")

# ---------------------------------------------------------------------------
# Column map (1-based for openpyxl) — will be auto-detected from headers
# ---------------------------------------------------------------------------
COL = {}

def detect_columns(ws):
    """Auto-detect column indices from header row."""
    global COL
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            COL[val] = c
    # Convenience aliases
    aliases = {
        'type': 'Source_Type', 'name': 'Facility_Name',
        'corp': 'Corporate_Name', 'addr': 'Address',
        'city': 'City', 'state': 'State', 'zip': 'ZIP',
        'county': 'County', 'beds': 'Total_Beds', 'census': 'Census',
        'served': 'Do_We_Serve', 'int': 'Integrated_Flag',
        'pcp': 'PCP_Flag', 'mh': 'MH_Flag', 'barrier': 'Barrier',
        'lat': 'Latitude', 'lon': 'Longitude',
        'dq': 'Data_Quality_Flag', 'corp_source': 'Corp_Attribution_Source',
    }
    for alias, real in aliases.items():
        if real in COL:
            COL[alias] = COL[real]

def norm(s):
    """Normalize string for matching."""
    if not s:
        return ''
    s = str(s).upper().strip()
    s = re.sub(r'\.(?=\s|$)', '', s)
    s = re.sub(r'(?<=\b[A-Z])\.', '', s)
    s = ' '.join(s.split())
    return s

def find_row(ws, name_frag, state, city=None, corp=None):
    """Find row by content match. Returns row index, None, or list of matches."""
    matches = []
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['name']).value)
        rs = norm(ws.cell(r, COL['state']).value)
        if name_frag.upper() not in rn:
            continue
        if state.upper() != rs:
            continue
        if city and city.upper() != norm(ws.cell(r, COL['city']).value):
            continue
        if corp and corp.upper() != norm(ws.cell(r, COL['corp']).value):
            continue
        matches.append(r)
    if len(matches) == 1:
        return matches[0]
    elif len(matches) == 0:
        return None
    else:
        return matches

def find_all_rows(ws, corp_exact, state=None):
    """Find all rows with exact corporate name match."""
    matches = []
    for r in range(2, ws.max_row + 1):
        rc = norm(ws.cell(r, COL['corp']).value)
        if rc != corp_exact.upper():
            continue
        if state and norm(ws.cell(r, COL['state']).value) != state.upper():
            continue
        matches.append(r)
    return matches

def find_all_rows_like(ws, corp_pattern):
    """Find all rows where corporate name contains pattern."""
    matches = []
    pat = corp_pattern.upper()
    for r in range(2, ws.max_row + 1):
        rc = norm(ws.cell(r, COL['corp']).value)
        if pat in rc:
            matches.append(r)
    return matches

# ---------------------------------------------------------------------------
# CHANGE DEFINITIONS
# ---------------------------------------------------------------------------

# === PART A: Avardis legacy name consolidation ===
AVARDIS_RENAMES = {
    "CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE": "AVARDIS",
    "Consulate Health Care": "AVARDIS",
    "CONSULATE OR LAVIE OR NSPIRE": "AVARDIS",
    "INSPIRED HEALTHCARE MANAGEMENT": "AVARDIS",
}

# === PART B: Spring Harbor typo ===
# 1 GA row: "Spring Harbor" -> "Spring Arbor Management"
SPRING_HARBOR_FIX = ("Savannah Court Of Camilla", "GA", "Spring Harbor", "Spring Arbor Management")

# === PART C: Scoring board duplicate consolidations ===
# (old_corp_exact, new_corp) — rename old -> new
SCORING_BOARD_RENAMES = {
    "ALG": "ALG Senior",
    "TRILOGY": "TRILOGY HEALTH SERVICES",
    "SUNRISE": "Sunrise Senior Living",
    "SANSTONE": "SANSTONE HEALTH & REHABILITATION",
    "BRICKYARD": "BRICKYARD HEALTHCARE",
    "YAD": "YAD HEALTHCARE",
    "ENVIVE": "ENVIVE HEALTHCARE",
    "CARECORE": "CARECORE HEALTH",
    "CARESPRING HEALTH CARE MANAGEMENT": "CARESPRING",
    "AOM HEALTHCARE SOLUTIONS": "AOM HEALTHCARE",
    "Optalis Healthcare": "OPTALIS HEALTH & REHABILITATION",
    "PACs": "PACS GROUP",
    "GARDANT MANAGEMENT SOLUTIONS": "GARDANT MANAGEMENT SOLUTIONS, INC",
    "APERION": "APERION CARE",
    "TOPAZ": "TOPAZ HEALTHCARE",
    "TerraBella / Discovery Senior Living": "TerraBella Senior Living",
    "Infinity Health Care Management": "INFINITY HEALTHCARE CONSULTING",
    "KISSITO": "Kissito Healthcare",
    "CARROLTON NURSING HOMES": "CARROLTON FACILTY MANAGEMENT",
    "MFA MARYVILLE RE LLC": "MFA",
    "Encore Senior Living": "BLUEGRASS/ENCORE",
    "FUNDAMENTAL HEALTHCARE": "FUNDAMENTAL LTC",
    "AVENTURA HEALTH GROUP": "AVENTURA",
}

# === PART D: Cedarhurst facility name normalization ===
# Facility names under CEDARHURST SENIOR LIVING that need standardizing
CEDARHURST_FACILITY_FIXES = [
    # (name_frag, city, state, new_facility_name)
    ("CEDAR HURST LIVING OF MARION", "Marion", "IN", "CEDARHURST OF MARION"),
    ("CEDAR CREEK ASSISTED LIVING", None, None, None),  # flag only — multiple possible
]

# === PART E: Greencroft/Oak Grove NC ===
# Delete: ALF row "OAK GROVE HEALTHCARE CENTER" under GREENCROFT in Rutherfordton NC
# Update: SNF row "OAK GROVE HEALTHCARE" under Consulate legacy -> SNF CARE CENTERS
OAK_GROVE_DELETE = ("OAK GROVE HEALTHCARE CENTER", "Rutherfordton", "NC", "GREENCROFT")
OAK_GROVE_UPDATE = ("OAK GROVE HEALTHCARE", "Rutherfordton", "NC",
                     "CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE",
                     "SNF CARE CENTERS")

# === PART F: Atrium swing bed removal ===
ATRIUM_DELETE = ("WILKES REGIONAL MEDICAL CTR SN", "NORTH WILKESBORO", "NC", "ATRIUM HEALTH")

# === PART G: Gardant fixes ===
GARDANT_REGENCY = ("REGENCY AT AUGUSTA", "Fishersville", "VA", "GATEWAY SENIOR LLC",
                   "GARDANT MANAGEMENT SOLUTIONS, INC")
GARDANT_ENTITY_MERGE = [
    # (name_frag, city, state) — rows on entity "GARDANT" to rename to "GARDANT MANAGEMENT SOLUTIONS, INC"
    ("GREEN OAKS OF GOSHEN", "Goshen", "IN"),
    ("BELLAIRE", "Harrisonburg", "VA"),
]
GARDANT_BED_FIX = ("LANDING OF LONG COVE", "Mason", "OH", 82, 66)  # beds: 17->82, census estimate
GARDANT_NAME_FIX = ("CARRIAGE COURT GROVE OF CITY", "Grove City", "OH", "CARRIAGE COURT OF GROVE CITY")
# Deferred V24.2 deletes
GARDANT_DEFERRED_DELETES = [
    ("SWEET GALILEE AT THE WIGWAM", "Anderson", "IN", "SWEET GALILEE AT THE WIGWAM LLC",
     "Propco LLC duplicate — Gardant is actual operator"),
]

# === PART H: Vitality / Triple Crown fixes ===
VITALITY_REATTRIBS = [
    # (name_frag, city, state, old_corp, new_corp)
    ("SUNRISE OF FIVE FORKS", "Lilburn", "GA", "VITALITY LIVING", "Sunrise Senior Living"),
    ("VITALITY LIVING-COLUMBIA", "Columbia", "SC", "Watercrest Senior Living Group", "VITALITY LIVING"),
    ("ACCLAIM AT BELMONT BAY", "Woodbridge", "VA", "Senior Lifestyle", "VITALITY LIVING"),
]
VITALITY_DELETES = [
    # (name_frag, city, state, corp, reason) — deferred from V24.2
    ("VITALITY ELIZABETHTOWN", "Elizabethtown", "KY", "VITALITY LIVING",
     "Duplicate of VITALITY LIVING ELIZABETHTOWN (smaller, surrogate)"),
    ("SUMMIT OF EDGEWOOD", "Edgewood", "KY", "TRIPLE CROWN SENIOR LIVING",
     "Duplicate of Summit at Ft Mitchell (smaller, more quality flags)"),
    ("WATERCREST COLUMBIA", "Columbia", "SC", None,
     "Duplicate of VITALITY LIVING-COLUMBIA (Watercrest-branded copy)"),
]
# Triple Crown entity merge: VITALITY SENIOR SERVICES -> Triple Crown Senior Living
TRIPLE_CROWN_RENAME = ("VITALITY SENIOR SERVICES", "Triple Crown Senior Living")
# City typo
VITALITY_CITY_FIX = ("VITALITY LIVING HENDERSONVILLE", "Hendersonvlle", "TN", "Hendersonville")

# Missing Vitality/Triple Crown facilities — DEFERRED TO V25.2
# Addresses need confirmation before insert:
#   - Aspen Alcove at Bardstown, KY (41 units)
#   - Aspen Alcove at Elizabethtown, KY (42 units)
#   - Vitality Living West End Richmond, VA

# === PART I: RUI fixes ===
RUI_REATTRIBS = [
    # (name_frag, city, state, old_corp, new_corp)
    ("Watercrest Palm Beach Gardens", "Palm Beach Gardens", "FL",
     "Retirement Unlimited, Inc.", "Watercrest Senior Living Group"),
    ("Capital Square At Tallahassee", "Tallahassee", "FL",
     "Retirement Unlimited, Inc.", "YourLife Senior Living"),
    ("YOURLife of West Melbourne", "West Melbourne", "FL",
     "Retirement Unlimited, Inc.", "YourLife Senior Living"),
]
RUI_DELETES = [
    # (name_frag, city, state, corp, reason) — deferred from V24.2
    ("MORGAN AT FORD", "Williamsburg", "VA", "Retirement Unlimited, Inc.",
     "Phantom facility — RUI withdrew Jan 2025, never built"),
    ("WELLINGTON AT LAKE MANASSAS COMMUNITY", "Gainesville", "VA", "INDEPENDENT",
     "Duplicate of Wellington (LEGACY surrogate, misattributed)"),
    ("ASHLEIGH AT LANSDOWNE COMMUNITY", "Leesburg", "VA", "INDEPENDENT",
     "Duplicate of Ashleigh (LEGACY surrogate, misattributed)"),
    ("WESTMONT AT SHORT PUMP", "Glen Allen", "VA", "Retirement Unlimited, Inc.",
     "Duplicate of Westmont (surrogate beds/census, missing county)"),
    ("WOODLAND HILLS COMMUNITY", "Roanoke", "VA", "RUI",
     "Duplicate of Woodland Hills IL, AL & MC (surrogate, missing county)"),
]
# Entity merge: RUI -> Retirement Unlimited, Inc.
RUI_ENTITY_RENAME = ("RUI", "Retirement Unlimited, Inc.")
# County fills for RUI VA facilities
RUI_COUNTY_FILLS = [
    ("PAUL SPRING", "Alexandria", "VA", "Fairfax"),
    ("HEATHERWOOD", "Burke", "VA", "Fairfax"),
    ("HIDENWOOD", "Newport News", "VA", "Newport News City"),
    ("BARRINGTON AT HIOAKS", "Richmond", "VA", "Richmond City"),
    ("AARONDALE", "Springfield", "VA", "Fairfax"),
    ("BAY LAKE", "Virginia Beach", "VA", "Virginia Beach City"),
]

# === PART J: Barrier flag fills (16 T5 entities) ===
# Fill rows where barrier is None/empty with the correct barrier value
BARRIER_FILLS = [
    # (corp_name_exact_or_pattern, barrier_value, match_type)
    # match_type: 'exact' = corporate_name_raw must match exactly
    #             'like' = corporate_name_raw must contain pattern
    ("AVENTURA HEALTH GROUP", "Termination Risk", "exact"),
    ("CARDON & ASSOCIATES", "Permanent - Cardon", "exact"),
    ("CLEARVIEW", "Alliance, Own Provider Group", "exact"),
    ("COMMUNICARE", "Own Provider Group", "exact"),
    ("EMBASSY HEALTHCARE", "Reputation", "exact"),
    ("EXCEPTIONAL LIVING CENTERS", "Alliance", "exact"),
    ("Hill Valley", "MH Only Opportunity", "exact"),
    ("JOURNEY HEALTHCARE", "Alliance", "exact"),
    ("MFA", "Alliance", "exact"),
    ("PAVILION HEALTHCARE", "Alliance, Own Provider Group", "exact"),
    ("ACCORDIUS HEALTH LLC/PORTOPICCOLO GROUP", "Alliance", "exact"),
    ("SIGNATURE HEALTH", "MH Only Opportunity", "exact"),
    ("SINGH", "Own Provider Group", "exact"),
    ("VENZA", "Alliance", "exact"),
    ("VENZA CARE MANAGEMENT", "Alliance", "exact"),
]

# === PART K: Address normalization ===
# Standardize street suffixes to abbreviations (most common form in DB)
ADDR_SUFFIX_MAP = [
    # (pattern, replacement) — regex patterns with word boundaries
    (r'\bDrive\b', 'Dr'),
    (r'\bStreet\b', 'St'),
    (r'\bBoulevard\b', 'Blvd'),
    (r'\bRoad\b', 'Rd'),
    (r'\bAvenue\b', 'Ave'),
    (r'\bLane\b', 'Ln'),
    (r'\bCircle\b', 'Cir'),
    (r'\bCourt\b', 'Ct'),
    (r'\bPlace\b', 'Pl'),
    (r'\bParkway\b', 'Pkwy'),
    (r'\bHighway\b', 'Hwy'),
    (r'\bTrail\b', 'Trl'),
    (r'\bTerrace\b', 'Ter'),
    (r'\bSquare\b', 'Sq'),
]

# === PART L: Campus/Site ID ===
def extract_street_num(addr):
    """Extract leading street number from address."""
    if not addr:
        return None
    m = re.match(r'(\d+)', str(addr).strip())
    return m.group(1) if m else None

def haversine_miles(lat1, lon1, lat2, lon2):
    """Distance in miles between two GPS points."""
    R = 3958.8  # Earth radius in miles
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))

# === PART M: Miscellaneous field fixes ===
CITY_FIXES = [
    # (name_frag, old_city, state, new_city)
    ("VITALITY LIVING HENDERSONVILLE", "Hendersonvlle", "TN", "Hendersonville"),
]
BED_FIXES = [
    # (name_frag, city, state, new_beds, new_census)
    ("LANDING OF LONG COVE", "Mason", "OH", 82, 66),
]
COUNTY_FIXES = [
    # (name_frag, city, state, county)
    ("GREEN OAKS OF GOSHEN", "Goshen", "IN", "Elkhart"),
    ("CHANCELLOR", "Fredericksburg", "VA", "Spotsylvania"),
]

# ============================================================
# PROCESS
# ============================================================

print("Loading V25...")
wb = openpyxl.load_workbook(V25_PATH)
ws = wb.active
detect_columns(ws)
total_rows = ws.max_row - 1
print(f"V25 rows: {total_rows}")
print(f"Detected columns: {len(COL)}")

errors = []
changes = defaultdict(list)
rows_to_delete = []

# --- PART A: Avardis legacy name consolidation ---
print("\n--- Part A: Avardis Legacy Name Consolidation ---")
for old_name, new_name in AVARDIS_RENAMES.items():
    rows = find_all_rows(ws, old_name)
    if rows:
        for r in rows:
            ws.cell(r, COL['corp']).value = new_name
            fac = ws.cell(r, COL['name']).value
            changes['avardis'].append(f"Row {r}: {fac} | '{old_name}' -> '{new_name}'")
    else:
        errors.append(f"AVARDIS: No rows found for '{old_name}' (may already be fixed)")

# --- PART B: Spring Harbor typo ---
print("--- Part B: Spring Harbor Typo ---")
name_frag, st, old_corp, new_corp = SPRING_HARBOR_FIX
row = find_row(ws, name_frag, st, corp=old_corp)
if row is None:
    errors.append(f"SPRING HARBOR: NOT FOUND -- '{name_frag}' in {st} under '{old_corp}'")
elif isinstance(row, list):
    errors.append(f"SPRING HARBOR: MULTIPLE MATCHES ({len(row)})")
else:
    ws.cell(row, COL['corp']).value = new_corp
    fac = ws.cell(row, COL['name']).value
    changes['spring_harbor'].append(f"Row {row}: {fac} | '{old_corp}' -> '{new_corp}'")

# --- PART C: Scoring board duplicate consolidations ---
print("--- Part C: Scoring Board Duplicate Consolidations ---")
for old_corp, new_corp in SCORING_BOARD_RENAMES.items():
    rows = find_all_rows(ws, old_corp)
    if rows:
        for r in rows:
            ws.cell(r, COL['corp']).value = new_corp
            fac = ws.cell(r, COL['name']).value
            changes['scoring_board'].append(f"Row {r}: {fac} | '{old_corp}' -> '{new_corp}'")
    else:
        errors.append(f"SCORING BOARD: No rows for '{old_corp}' (may already be fixed)")

# --- PART D: Cedarhurst facility name normalization ---
print("--- Part D: Cedarhurst Facility Naming ---")
for r in range(2, ws.max_row + 1):
    corp = norm(ws.cell(r, COL['corp']).value)
    if corp == "CEDARHURST SENIOR LIVING":
        fac = ws.cell(r, COL['name']).value
        if fac and "CEDAR HURST" in str(fac).upper():
            old = fac
            new = str(fac).replace("CEDAR HURST", "CEDARHURST").replace("Cedar Hurst", "Cedarhurst")
            ws.cell(r, COL['name']).value = new
            changes['cedarhurst'].append(f"Row {r}: '{old}' -> '{new}'")

# --- PART E: Greencroft/Oak Grove NC ---
print("--- Part E: Greencroft/Oak Grove NC ---")
# Delete ALF dupe
nf, city, st, corp = OAK_GROVE_DELETE
row = find_row(ws, nf, st, city=city, corp=corp)
if row is None:
    errors.append(f"OAK GROVE DELETE: NOT FOUND -- '{nf}' under {corp}")
elif isinstance(row, list):
    errors.append(f"OAK GROVE DELETE: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    changes['oak_grove'].append(f"DELETE Row {row}: {fac} | {city}, {st} | beds={beds} | Greencroft misattribution")
# Update SNF row — corp may have been renamed by Part A (Avardis), so try multiple searches
nf, city, st, old_corp, new_corp = OAK_GROVE_UPDATE
row = find_row(ws, nf, st, city=city, corp=old_corp)
if row is None:
    # Try under AVARDIS (Part A may have renamed it)
    row = find_row(ws, nf, st, city=city, corp="AVARDIS")
if row is None:
    # Last resort: find by name without corp, but exclude the GREENCROFT row we're deleting
    candidates = find_row(ws, nf, st, city=city)
    if isinstance(candidates, list):
        # Pick the one that's NOT in rows_to_delete
        candidates = [r for r in candidates if r not in rows_to_delete]
        if len(candidates) == 1:
            row = candidates[0]
        else:
            errors.append(f"OAK GROVE UPDATE: MULTIPLE MATCHES after filtering ({len(candidates)})")
    else:
        row = candidates
if row is None:
    errors.append(f"OAK GROVE UPDATE: NOT FOUND -- '{nf}' in {city}, {st}")
elif not isinstance(row, list):
    old = ws.cell(row, COL['corp']).value
    ws.cell(row, COL['corp']).value = new_corp
    changes['oak_grove'].append(f"Row {row}: OAK GROVE HEALTHCARE | '{old}' -> '{new_corp}'")

# --- PART F: Atrium swing bed removal ---
print("--- Part F: Atrium Swing Bed Removal ---")
nf, city, st, corp = ATRIUM_DELETE
row = find_row(ws, nf, st, city=city, corp=corp)
if row is None:
    errors.append(f"ATRIUM DELETE: NOT FOUND -- '{nf}' in {city}, {st}")
elif isinstance(row, list):
    errors.append(f"ATRIUM DELETE: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    changes['atrium'].append(f"DELETE Row {row}: {fac} | {city}, {st} | beds={beds} | 10-bed hospital swing bed")

# --- PART G: Gardant fixes ---
print("--- Part G: Gardant Fixes ---")
# Regency at Augusta reattribution
nf, city, st, old_corp, new_corp = GARDANT_REGENCY
row = find_row(ws, nf, st, city=city, corp=old_corp)
if row is None:
    errors.append(f"GARDANT REGENCY: NOT FOUND -- '{nf}' under '{old_corp}'")
elif isinstance(row, list):
    errors.append(f"GARDANT REGENCY: MULTIPLE MATCHES ({len(row)})")
else:
    ws.cell(row, COL['corp']).value = new_corp
    changes['gardant'].append(f"Row {row}: REGENCY AT AUGUSTA | '{old_corp}' -> '{new_corp}'")

# Entity merge: "GARDANT" -> "GARDANT MANAGEMENT SOLUTIONS, INC"
for nf, city, st in GARDANT_ENTITY_MERGE:
    row = find_row(ws, nf, st, city=city, corp="GARDANT")
    if row is None:
        # May already be fixed or under different corp
        row = find_row(ws, nf, st, city=city, corp="GARDANT MANAGEMENT SOLUTIONS")
    if row is None:
        errors.append(f"GARDANT ENTITY MERGE: NOT FOUND -- '{nf}' in {city}, {st}")
    elif isinstance(row, list):
        errors.append(f"GARDANT ENTITY MERGE: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = "GARDANT MANAGEMENT SOLUTIONS, INC"
        changes['gardant'].append(f"Row {row}: {nf} | '{old}' -> 'GARDANT MANAGEMENT SOLUTIONS, INC'")

# Facility name fix
nf, city, st, new_name = GARDANT_NAME_FIX
row = find_row(ws, nf, st, city=city)
if row:
    if not isinstance(row, list):
        old = ws.cell(row, COL['name']).value
        ws.cell(row, COL['name']).value = new_name
        changes['gardant'].append(f"Row {row}: name '{old}' -> '{new_name}'")

# Deferred deletes
for nf, city, st, corp, reason in GARDANT_DEFERRED_DELETES:
    row = find_row(ws, nf, st, city=city, corp=corp)
    if row is None:
        errors.append(f"GARDANT DELETE: NOT FOUND -- '{nf}' under '{corp}' (may already be deleted)")
    elif isinstance(row, list):
        errors.append(f"GARDANT DELETE: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        fac = ws.cell(row, COL['name']).value
        beds = ws.cell(row, COL['beds']).value
        rows_to_delete.append(row)
        changes['gardant'].append(f"DELETE Row {row}: {fac} | {city}, {st} | beds={beds} | {reason}")

# --- PART H: Vitality / Triple Crown fixes ---
print("--- Part H: Vitality / Triple Crown Fixes ---")
# Reattributions
for nf, city, st, old_corp, new_corp in VITALITY_REATTRIBS:
    row = find_row(ws, nf, st, city=city, corp=old_corp)
    if row is None:
        errors.append(f"VITALITY REATTRIB: NOT FOUND -- '{nf}' under '{old_corp}' (may already be fixed)")
    elif isinstance(row, list):
        errors.append(f"VITALITY REATTRIB: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        ws.cell(row, COL['corp']).value = new_corp
        fac = ws.cell(row, COL['name']).value
        changes['vitality'].append(f"Row {row}: {fac} | '{old_corp}' -> '{new_corp}'")

# Deletes
for nf, city, st, corp, reason in VITALITY_DELETES:
    if corp:
        row = find_row(ws, nf, st, city=city, corp=corp)
    else:
        row = find_row(ws, nf, st, city=city)
    if row is None:
        errors.append(f"VITALITY DELETE: NOT FOUND -- '{nf}' in {city}, {st} (may already be deleted)")
    elif isinstance(row, list):
        errors.append(f"VITALITY DELETE: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        fac = ws.cell(row, COL['name']).value
        beds = ws.cell(row, COL['beds']).value
        rows_to_delete.append(row)
        changes['vitality'].append(f"DELETE Row {row}: {fac} | {city}, {st} | beds={beds} | {reason}")

# Triple Crown entity rename
old, new = TRIPLE_CROWN_RENAME
rows = find_all_rows(ws, old)
for r in rows:
    ws.cell(r, COL['corp']).value = new
    fac = ws.cell(r, COL['name']).value
    changes['vitality'].append(f"Row {r}: {fac} | '{old}' -> '{new}'")

# City typo fix
nf, old_city, st, new_city = VITALITY_CITY_FIX
row = find_row(ws, nf, st, city=old_city)
if row:
    if not isinstance(row, list):
        ws.cell(row, COL['city']).value = new_city
        changes['vitality'].append(f"Row {row}: city '{old_city}' -> '{new_city}'")

# --- PART I: RUI fixes ---
print("--- Part I: RUI Fixes ---")
for nf, city, st, old_corp, new_corp in RUI_REATTRIBS:
    row = find_row(ws, nf, st, city=city, corp=old_corp)
    if row is None:
        errors.append(f"RUI REATTRIB: NOT FOUND -- '{nf}' under '{old_corp}' (may already be fixed)")
    elif isinstance(row, list):
        errors.append(f"RUI REATTRIB: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        ws.cell(row, COL['corp']).value = new_corp
        fac = ws.cell(row, COL['name']).value
        changes['rui'].append(f"Row {row}: {fac} | '{old_corp}' -> '{new_corp}'")

for nf, city, st, corp, reason in RUI_DELETES:
    row = find_row(ws, nf, st, city=city, corp=corp)
    if row is None:
        errors.append(f"RUI DELETE: NOT FOUND -- '{nf}' under '{corp}' (may already be deleted)")
    elif isinstance(row, list):
        errors.append(f"RUI DELETE: MULTIPLE ({len(row)}) -- '{nf}'")
    else:
        fac = ws.cell(row, COL['name']).value
        beds = ws.cell(row, COL['beds']).value
        rows_to_delete.append(row)
        changes['rui'].append(f"DELETE Row {row}: {fac} | {city}, {st} | beds={beds} | {reason}")

# Entity rename: RUI -> Retirement Unlimited, Inc.
old, new = RUI_ENTITY_RENAME
rows = find_all_rows(ws, old)
for r in rows:
    ws.cell(r, COL['corp']).value = new
    fac = ws.cell(r, COL['name']).value
    changes['rui'].append(f"Row {r}: {fac} | '{old}' -> '{new}'")

# County fills
for nf, city, st, county in RUI_COUNTY_FILLS:
    row = find_row(ws, nf, st, city=city)
    if row is None:
        errors.append(f"RUI COUNTY: NOT FOUND -- '{nf}' in {city}, {st}")
    elif isinstance(row, list):
        # Take first match
        row = row[0]
    if row and not isinstance(row, list):
        old_county = ws.cell(row, COL['county']).value
        if not old_county or str(old_county).strip() in ('', 'None'):
            ws.cell(row, COL['county']).value = county
            fac = ws.cell(row, COL['name']).value
            changes['rui'].append(f"Row {row}: {fac} | county '' -> '{county}'")

# --- PART J: Barrier flag fills ---
print("--- Part J: Barrier Flag Fills ---")
for corp_name, barrier_val, match_type in BARRIER_FILLS:
    if match_type == 'exact':
        rows = find_all_rows(ws, corp_name)
    else:
        rows = find_all_rows_like(ws, corp_name)
    filled = 0
    for r in rows:
        current = ws.cell(r, COL['barrier']).value
        if not current or str(current).strip() in ('', 'None'):
            ws.cell(r, COL['barrier']).value = barrier_val
            filled += 1
    if filled > 0:
        changes['barrier'].append(f"{corp_name}: filled {filled} rows with '{barrier_val}'")

# --- PART K: Address normalization ---
print("--- Part K: Address Normalization ---")
addr_changes = 0
for r in range(2, ws.max_row + 1):
    addr = ws.cell(r, COL['addr']).value
    if not addr:
        continue
    new_addr = str(addr)
    for pattern, replacement in ADDR_SUFFIX_MAP:
        new_addr = re.sub(pattern, replacement, new_addr)
    if new_addr != str(addr):
        ws.cell(r, COL['addr']).value = new_addr
        addr_changes += 1
changes['address'].append(f"Normalized {addr_changes} addresses (street suffix abbreviation)")

# --- PART L: Campus/Site ID ---
print("--- Part L: Campus/Site ID ---")
# Add new column header
campus_col = ws.max_column + 1
ws.cell(1, campus_col).value = "Campus_ID"

# Build campus groups by (street_number, city_upper, state)
campus_groups = defaultdict(list)
for r in range(2, ws.max_row + 1):
    addr = ws.cell(r, COL['addr']).value
    city = norm(ws.cell(r, COL['city']).value)
    state = norm(ws.cell(r, COL['state']).value)
    snum = extract_street_num(addr)
    if snum:
        key = f"{snum}|{city}|{state}"
        campus_groups[key].append(r)

# Assign campus IDs
campus_id_counter = 1
row_to_campus = {}
for key in sorted(campus_groups.keys()):
    rows_in_campus = campus_groups[key]
    if len(rows_in_campus) > 0:
        for r in rows_in_campus:
            row_to_campus[r] = campus_id_counter
        campus_id_counter += 1

# GPS proximity pass for unmatched rows
gps_threshold = 0.1  # miles
unmatched = []
for r in range(2, ws.max_row + 1):
    if r not in row_to_campus:
        lat = ws.cell(r, COL.get('lat', COL.get('Latitude', 999))).value
        lon = ws.cell(r, COL.get('lon', COL.get('Longitude', 999))).value
        if lat and lon:
            unmatched.append((r, float(lat), float(lon)))

# Try to match unmatched rows to existing campuses via GPS
matched_by_gps = 0
for r, lat, lon in unmatched:
    # Check against existing campus centroids
    best_campus = None
    best_dist = gps_threshold
    for key, rows_in_campus in campus_groups.items():
        for cr in rows_in_campus[:1]:  # Check first row of each campus
            clat = ws.cell(cr, COL.get('lat', COL.get('Latitude', 999))).value
            clon = ws.cell(cr, COL.get('lon', COL.get('Longitude', 999))).value
            if clat and clon:
                try:
                    dist = haversine_miles(lat, lon, float(clat), float(clon))
                    if dist < best_dist:
                        best_dist = dist
                        best_campus = row_to_campus.get(cr)
                except (ValueError, TypeError):
                    continue
    if best_campus:
        row_to_campus[r] = best_campus
        matched_by_gps += 1
    else:
        row_to_campus[r] = campus_id_counter
        campus_id_counter += 1

# Write campus IDs
for r in range(2, ws.max_row + 1):
    ws.cell(r, campus_col).value = row_to_campus.get(r, campus_id_counter)

multi_row_campuses = sum(1 for k, v in campus_groups.items() if len(v) > 1)
changes['campus_id'].append(
    f"Added Campus_ID column. {campus_id_counter - 1} unique campuses. "
    f"{multi_row_campuses} multi-facility campuses. {matched_by_gps} GPS proximity matches."
)

# --- PART M: Miscellaneous field fixes ---
print("--- Part M: Miscellaneous Field Fixes ---")
for nf, city, st, county in COUNTY_FIXES:
    row = find_row(ws, nf, st, city=city)
    if row and not isinstance(row, list):
        old = ws.cell(row, COL['county']).value
        if not old or str(old).strip() in ('', 'None'):
            ws.cell(row, COL['county']).value = county
            changes['misc'].append(f"Row {row}: {nf} | county '' -> '{county}'")

for nf, city, st, new_beds, new_census in BED_FIXES:
    row = find_row(ws, nf, st, city=city)
    if row and not isinstance(row, list):
        old_beds = ws.cell(row, COL['beds']).value
        old_census = ws.cell(row, COL['census']).value
        ws.cell(row, COL['beds']).value = new_beds
        ws.cell(row, COL['census']).value = new_census
        changes['misc'].append(
            f"Row {row}: {nf} | beds {old_beds}->{new_beds}, census {old_census}->{new_census}"
        )

# --- APPLY DELETES (bottom-up to preserve row indices) ---
print(f"\n--- Applying {len(rows_to_delete)} deletes ---")
for row_idx in sorted(set(rows_to_delete), reverse=True):
    ws.delete_rows(row_idx)

# ============================================================
# REPORT
# ============================================================
print(f"\n{'=' * 70}")
print(f"V25.1 Migration Report")
print(f"{'=' * 70}")

category_labels = {
    'avardis': 'PART A: Avardis Legacy Consolidation',
    'spring_harbor': 'PART B: Spring Harbor Typo',
    'scoring_board': 'PART C: Scoring Board Consolidations',
    'cedarhurst': 'PART D: Cedarhurst Facility Naming',
    'oak_grove': 'PART E: Greencroft/Oak Grove NC',
    'atrium': 'PART F: Atrium Swing Bed',
    'gardant': 'PART G: Gardant Fixes',
    'vitality': 'PART H: Vitality/Triple Crown',
    'rui': 'PART I: RUI Fixes',
    'barrier': 'PART J: Barrier Flag Fills',
    'address': 'PART K: Address Normalization',
    'campus_id': 'PART L: Campus/Site ID',
    'misc': 'PART M: Miscellaneous Fixes',
}

total_changes = 0
for key, label in category_labels.items():
    items = changes.get(key, [])
    count = len(items)
    total_changes += count
    print(f"{label}: {count}")

print(f"{'=' * 70}")
print(f"TOTAL CHANGES: {total_changes}")
print(f"DELETES: {len(rows_to_delete)}")
print(f"Final row count: {ws.max_row - 1}")
print()

for key, label in category_labels.items():
    items = changes.get(key, [])
    if items:
        print(f"\n{label}:")
        for item in items:
            print(f"  {item}")

# ============================================================
# VALIDATION
# ============================================================
expected_final = total_rows - len(set(rows_to_delete))
actual_final = ws.max_row - 1
if actual_final != expected_final:
    errors.append(f"Row count mismatch: expected {expected_final}, got {actual_final}")

if errors:
    print(f"\n{'=' * 70}")
    print(f"!! {len(errors)} NOTES/WARNINGS !!")
    for e in errors:
        print(f"  {e}")
    print()
    critical = [e for e in errors if "Row count mismatch" in e]
    if critical:
        print("CRITICAL ERRORS -- NOT SAVING.")
    else:
        print(f"{len(errors)} non-critical notes (NOT FOUND = already applied). Saving...")
        print(f"Saving to {V25_1_PATH}...")
        wb.save(V25_1_PATH)
        print("Done.")
else:
    print("\nAll operations completed.")
    print(f"Saving to {V25_1_PATH}...")
    wb.save(V25_1_PATH)
    print("Done.")

wb.close()
