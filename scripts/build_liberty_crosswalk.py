#!/usr/bin/env python3
"""Build Liberty Facility Crosswalk — independent dataset from website."""

import sys
sys.path.insert(0, ".")
from utils import ensure_report_dir
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Styles ──
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

cols = [
    "Facility Name", "Division", "Address", "City", "State", "ZIP",
    "Care Levels", "Unit/Bed Details", "Campus Notes",
    "Phone", "Website",
]


def write_header(ws):
    for c, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"


def write_row(ws, row_num, vals, fill=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def autofit(ws, max_row):
    for c in range(1, len(cols) + 1):
        max_len = len(str(ws.cell(row=1, column=c).value or ""))
        for r in range(2, min(max_row + 1, 60)):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max_len + 3


# ── LSL Website Data (all from screenshots) ──
lsl_facilities = [
    # --- Batch 1 (8 facilities) ---
    {
        "name": "Brightmore of South Charlotte",
        "division": "LSL - Life Plan",
        "address": "10225 Old Ardrey Kell Road",
        "city": "Charlotte", "state": "NC", "zip": "28277",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "148 IL units, 14 AL units, 16 memory support units. 255,000 sq ft, 9 acres. Opened fall 2014.",
        "campus": "Campus includes Pavilion Heath Center at Brightmore — 120-bed SNF.",
        "phone": "704-557-0511", "website": "www.brightmoreofsouthcharlotte.com",
    },
    {
        "name": "Brightmore of Wilmington",
        "division": "LSL - Life Plan",
        "address": "2124 41st Street",
        "city": "Wilmington", "state": "NC", "zip": "28403",
        "care": "IL, AL, Living Smart",
        "units": "137 IL apartments. 191,200 sq ft. Also on campus: The Brightmore at 2110 (4-story, 82 AL + skilled living). 201 healthcare beds.",
        "campus": "Part of community since 1990.",
        "phone": "910-796-8150", "website": "www.brighmorewilmington.com",
    },
    {
        "name": "Carlisle Palm Beach",
        "division": "LSL - Senior Living",
        "address": "450 East Ocean Avenue",
        "city": "Lantana", "state": "FL", "zip": "33462",
        "care": "IL, AL, Memory Care",
        "units": "Luxury oceanfront. Steps from beach.",
        "campus": "",
        "phone": "561-533-6960", "website": "www.carlislepalmbeach.com",
    },
    {
        "name": "Carolina Bay at Autumn Hall",
        "division": "LSL - Life Plan",
        "address": "620 Carolina Bay Drive",
        "city": "Wilmington", "state": "NC", "zip": "28403",
        "care": "IL, AL, Memory Care, Skilled Nursing",
        "units": "Main campus: 12 buildings, 21 acres, 130,000+ sq ft.",
        "campus": "Life Plan Community. Multiple buildings totaling 214k+ sq ft.",
        "phone": "866-803-6099", "website": "www.carolinabayatautumnhall.com",
    },
    {
        "name": "The Carrollton",
        "division": "LSL - Senior Living",
        "address": "701 South Carrollton Avenue",
        "city": "New Orleans", "state": "LA", "zip": "70118",
        "care": "AL, Memory Care",
        "units": "73 AL + 20 memory care private units.",
        "campus": "Historic Greek revival courthouse (1855). Renovated and expanded.",
        "phone": "504-380-0067", "website": "www.thecarrollton.com",
    },
    {
        "name": "Hayes Barton Place",
        "division": "LSL - Life Plan",
        "address": "2600 Yettington Drive",
        "city": "Raleigh", "state": "NC", "zip": "27608",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "214 IL residences. On-site Bloomsbury Health Center: 85 private rooms (SNF/rehab). 7+ acres.",
        "campus": "Life Plan Community.",
        "phone": "919-803-6734", "website": "www.hayesbartonplace.com",
    },
    {
        "name": "Inspire Briar Chapel",
        "division": "Inspire by Liberty (55+)",
        "address": "152 Market Chapel Road",
        "city": "Pittsboro", "state": "NC", "zip": "27312",
        "care": "Senior 55+ Rental Apartments",
        "units": "1-2 bedroom floor plans.",
        "campus": "",
        "phone": "910-705-3925", "website": "www.inspirebriarchapel.com",
    },
    {
        "name": "Inspire Brunswick Forest",
        "division": "Inspire by Liberty (55+)",
        "address": "6146 Liberty Hall Drive",
        "city": "Leland", "state": "NC", "zip": "28451",
        "care": "Senior 55+ Rental Apartments",
        "units": "1-2 bedroom. Now Open.",
        "campus": "",
        "phone": "803-828-3322", "website": "www.inspirebrunswickforest.com",
    },
    # --- Batch 2 (15 facilities) ---
    {
        "name": "Inspire Royal Park",
        "division": "Inspire by Liberty (55+)",
        "address": "4101 Glenloch Circle",
        "city": "Matthews", "state": "NC", "zip": "28105",
        "care": "55+ Rental Apartments",
        "units": "1-3 bedroom + cottage homes.",
        "campus": "",
        "phone": "980-766-1590", "website": "www.inspireroyalpark.com",
    },
    {
        "name": "Inspire Sandhill",
        "division": "Inspire by Liberty (55+)",
        "address": "440 Town Center Place",
        "city": "Columbia", "state": "SC", "zip": "29229",
        "care": "Active Senior 55+ Rental Apartments",
        "units": "1-3 bedroom.",
        "campus": "",
        "phone": "803-828-3322", "website": "www.inspiresandhill.com",
    },
    {
        "name": "Kempton of Charleston",
        "division": "LSL - Senior Living",
        "address": "194 Spring Street",
        "city": "Charleston", "state": "SC", "zip": "29403",
        "care": "AL, Memory Support, Skilled Nursing, Rehab",
        "units": "94,000 sq ft. 71 AL + 21 memory support units.",
        "campus": "Adjacent to Medical District of Charleston. High barriers-to-entry area.",
        "phone": "843-754-6940", "website": "www.kemptonofcharleston.com",
    },
    {
        "name": "Kempton of Jacksonville",
        "division": "LSL - Senior Living",
        "address": "3045 Henderson Drive Extension",
        "city": "Jacksonville", "state": "NC", "zip": "28546",
        "care": "AL, Memory Support",
        "units": "",
        "campus": "",
        "phone": "910-355-1996", "website": "www.kemptonofjacksonville.com",
    },
    {
        "name": "The Kempton of Hermitage",
        "division": "LSL - Senior Living",
        "address": "3778 Central Pike",
        "city": "Hermitage", "state": "TN", "zip": "37076",
        "care": "AL, Memory Support",
        "units": "77 AL + 22 memory care units. 3.65 acres. Class A.",
        "campus": "Liberty's first TN venture. Operated by Liberty since Oct 2019.",
        "phone": "615-928-1400", "website": "www.kemptonofhermitage.com",
    },
    {
        "name": "Kempton of Rockhill",
        "division": "LSL - Senior Living",
        "address": "1611 Constitution Boulevard",
        "city": "Rock Hill", "state": "SC", "zip": "29732",
        "care": "AL, Memory Support",
        "units": "",
        "campus": "",
        "phone": "803-900-4125", "website": "www.kemptonofrockhill.com",
    },
    {
        "name": "Oakleaf Village of Lexington",
        "division": "LSL - Senior Living",
        "address": "800 N Lake Drive",
        "city": "Lexington", "state": "SC", "zip": "29072",
        "care": "AL, Memory Support",
        "units": "66 AL + 24 memory care apartments. 100% DHEC survey for 2 years.",
        "campus": "More recent addition to Liberty Senior Living portfolio.",
        "phone": "803-808-3477", "website": "www.oakleafvillagelexington.com",
    },
    {
        "name": "Pisgah Valley",
        "division": "LSL - Life Plan",
        "address": "95 Holcombe Cove Road",
        "city": "Candler", "state": "NC", "zip": "28715",
        "care": "IL, AL, Rehab, Skilled Care",
        "units": "72 IL patio homes, 24 AL units, 118-bed SNF. 40 acres.",
        "campus": "Life Plan Community. Operated by Liberty since Feb 2018. Near Asheville.",
        "phone": "828-418-2333", "website": "www.pisgahvalleyretirement.com",
    },
    {
        "name": "Quail Haven Village",
        "division": "LSL - Life Plan",
        "address": "155 Blake Boulevard",
        "city": "Pinehurst", "state": "NC", "zip": "28374",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "94 IL apartments + 60 skilled nursing beds.",
        "campus": "Life Plan Community. Operated by Liberty Healthcare since 2013. Golfing community.",
        "phone": "910-295-2294", "website": "www.quailhavenvillage.com",
    },
    {
        "name": "South Bay at Mount Pleasant",
        "division": "LSL - Senior Living",
        "address": "1400 Liberty Midtown Drive",
        "city": "Mount Pleasant", "state": "SC", "zip": "29464",
        "care": "IL, AL, Memory Support",
        "units": "143 IL apartments + 78 AL/memory support units. 27 acres.",
        "campus": "Continuum of care in place. Charleston metro area.",
        "phone": "843-936-2800", "website": "www.southbayatmountpleasant.com",
    },
    {
        "name": "The Barclay at SouthPark",
        "division": "LSL - Life Plan",
        "address": "4801 Barclay Downs Drive",
        "city": "Charlotte", "state": "NC", "zip": "28210",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "10.5 acres. 700-1550 sq ft units.",
        "campus": "Life Plan Community. SouthPark area of Charlotte.",
        "phone": "980-224-8540", "website": "www.barclayatsouthpark.com",
    },
    {
        "name": "The Peninsula of Charleston",
        "division": "LSL - Life Plan",
        "address": "625 King Street",
        "city": "Charleston", "state": "SC", "zip": "29403",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "",
        "campus": "Only Life Plan Community in downtown Charleston. Home Ownership model.",
        "phone": "", "website": "www.thepeninsulaofcharleston.com",
    },
    {
        "name": "The Preserve at Fairfield Glade",
        "division": "LSL - Senior Living",
        "address": "100 Samaritan Way",
        "city": "Crossville", "state": "TN", "zip": "38558",
        "care": "IL, AL, Rehab, Skilled Care",
        "units": "",
        "campus": "Mountain views.",
        "phone": "", "website": "www.thepreserveatfairfieldglade.com",
    },
    {
        "name": "The Templeton of Cary",
        "division": "LSL - Life Plan",
        "address": "125 Brightmore Drive",
        "city": "Cary", "state": "NC", "zip": "27511",
        "care": "IL, AL, Memory Support, Rehab, Skilled Care",
        "units": "196 luxury suites. 20 wooded acres across from WakeMed Cary Hospital.",
        "campus": "Life Plan. Rehab/Skilled Care on site.",
        "phone": "919-651-3699", "website": "www.thetempletonofcary.com",
    },
    {
        "name": "Wellington Bay",
        "division": "LSL - Senior Living",
        "address": "10430 Stable Lane",
        "city": "Wellington", "state": "FL", "zip": "33414",
        "care": "IL, AL, Memory Care",
        "units": "46-acre, 283-unit Class AA community.",
        "campus": "Country club experience. Equine therapy (horse paddock). Village of Wellington.",
        "phone": "561-423-9001", "website": "www.wellingtonbayfl.com",
    },
    # --- From original user list but NOT in screenshots yet ---
    {"name": "Harbour's Edge", "division": "LSL", "address": "", "city": "Delray Beach", "state": "FL", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Palms of Sebring", "division": "LSL", "address": "", "city": "Sebring", "state": "FL", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Park Summit", "division": "LSL", "address": "", "city": "Hendersonville", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Plantation Village", "division": "LSL", "address": "", "city": "Wilmington", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Rosemont", "division": "LSL", "address": "", "city": "Brampton", "state": "ON", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Cypress of Charlotte", "division": "LSL", "address": "", "city": "Charlotte", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Cypress of Hilton Head Island", "division": "LSL", "address": "", "city": "Hilton Head Island", "state": "SC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Cypress of Raleigh", "division": "LSL", "address": "", "city": "Raleigh", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Pines at Davidson", "division": "LSL", "address": "", "city": "Davidson", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Pointe at Lifespring", "division": "LSL", "address": "", "city": "Knoxville", "state": "TN", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "The Vista at Sedge Garden", "division": "LSL", "address": "", "city": "Kernersville", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Vista at Town Center", "division": "LSL", "address": "", "city": "King", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Waypoint at Bridgeway Station", "division": "LSL", "address": "", "city": "Morrisville", "state": "NC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Westminster Towers", "division": "LSL", "address": "", "city": "Rock Hill", "state": "SC", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
    {"name": "Willow Valley Communities", "division": "LSL", "address": "", "city": "Lancaster", "state": "PA", "zip": "", "care": "", "units": "", "campus": "", "phone": "", "website": ""},
]

wb = Workbook()

# ── Tab 1: LSL - Senior Living ──
ws1 = wb.active
ws1.title = "LSL - Senior Living"
write_header(ws1)

rn = 2
for fac in lsl_facilities:
    vals = [
        fac["name"], fac["division"], fac["address"], fac["city"],
        fac["state"], fac["zip"], fac["care"], fac["units"],
        fac["campus"], fac["phone"], fac["website"],
    ]
    fill = blue_fill if fac["address"] else None
    write_row(ws1, rn, vals, fill)
    rn += 1

filled = sum(1 for f in lsl_facilities if f["address"])
print(f"LSL Tab: {len(lsl_facilities)} facilities ({filled} with full data, {len(lsl_facilities) - filled} awaiting screenshots)")

# ── LHR Website Data (from LHRS-Directory_Jan2026.pdf) ──
# 41 locations across NC, SC, TN — "3 States, 28 Counties, 41 Locations"
lhr_facilities = [
    # --- NORTH CAROLINA ---
    # Alamance County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Alamance County", "division": "LHR - SNF", "address": "791 Boone Station Drive", "city": "Burlington", "state": "NC", "zip": "27215", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-586-9850", "website": ""},
    # Bertie County
    {"name": "Three Rivers Health & Rehabilitation Center", "division": "LHR - SNF", "address": "1403 Connor Ave.", "city": "Windsor", "state": "NC", "zip": "27983", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "252-794-4441", "website": ""},
    # Bladen County
    {"name": "Elizabethtown Healthcare & Rehabilitation Center", "division": "LHR - SNF", "address": "208 Mercer Road", "city": "Elizabethtown", "state": "NC", "zip": "28337", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-862-8181", "website": ""},
    # Brunswick County
    {"name": "Southport Health & Rehabilitation Center", "division": "LHR - SNF", "address": "630 N Fodale Ave", "city": "Southport", "state": "NC", "zip": "28461", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-457-9581", "website": ""},
    # Buncombe County
    {"name": "Pisgah Manor", "division": "LHR - SNF", "address": "104 Holcombe Cove Road", "city": "Candler", "state": "NC", "zip": "28715", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On Pisgah Valley Life Plan campus.", "phone": "828-667-9851", "website": ""},
    # Columbus County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Columbus County", "division": "LHR - SNF", "address": "1402 Pinckney Street", "city": "Whiteville", "state": "NC", "zip": "28472", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-642-4245", "website": ""},
    {"name": "Shoreland Health Care & Retirement Center", "division": "LHR - SNF", "address": "200 Flowers-Pridgen Drive", "city": "Whiteville", "state": "NC", "zip": "28472", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-642-4300", "website": ""},
    # Cumberland County
    {"name": "Golden Years Nursing Home", "division": "LHR - SNF", "address": "7348 North West Street", "city": "Falcon", "state": "NC", "zip": "28342", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-980-1271", "website": ""},
    {"name": "Woodlands Nursing & Rehabilitation Center", "division": "LHR - SNF", "address": "400 Pelt Drive", "city": "Fayetteville", "state": "NC", "zip": "28301", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-822-0515", "website": ""},
    {"name": "Highland House Rehabilitation & Healthcare", "division": "LHR - SNF", "address": "1700 Pamalee Dr.", "city": "Fayetteville", "state": "NC", "zip": "28301", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-488-2295", "website": ""},
    # Davie County
    {"name": "Bermuda Commons Nursing & Rehabilitation Center", "division": "LHR - SNF", "address": "316 NC Highway 801 South", "city": "Advance", "state": "NC", "zip": "27006", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-998-0240", "website": ""},
    # Mecklenburg County
    {"name": "Briar Creek at The Barclay", "division": "LHR - SNF", "address": "6041 Piedmont Row Drive", "city": "Charlotte", "state": "NC", "zip": "28208", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On The Barclay at SouthPark campus.", "phone": "980-443-6760", "website": ""},
    # Forsyth County
    {"name": "Oak Forest Health & Rehabilitation Center", "division": "LHR - SNF", "address": "5680 Windy Hill Drive", "city": "Winston-Salem", "state": "NC", "zip": "27105", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-776-5000", "website": ""},
    {"name": "The Oaks", "division": "LHR - SNF", "address": "901 Bethesda Road", "city": "Winston-Salem", "state": "NC", "zip": "27103", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-768-2211", "website": ""},
    {"name": "Summerstone Health & Rehabilitation Center", "division": "LHR - SNF", "address": "485 Veteran's Way", "city": "Kernersville", "state": "NC", "zip": "27284", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-515-3000", "website": ""},
    # Franklin County
    {"name": "Louisburg Healthcare & Rehabilitation Center and Louisburg Manor", "division": "LHR - SNF", "address": "202 Smoketree Way", "city": "Louisburg", "state": "NC", "zip": "27549", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "919-496-2188", "website": ""},
    # Halifax County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Halifax County", "division": "LHR - SNF", "address": "101 Caroline Avenue", "city": "Weldon", "state": "NC", "zip": "27890", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "252-536-4817", "website": ""},
    # Haywood County
    {"name": "Silver Bluff Village", "division": "LHR - SNF", "address": "100 Silver Bluff Dr.", "city": "Canton", "state": "NC", "zip": "28716", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "828-648-2044", "website": ""},
    # Johnston County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Johnston County", "division": "LHR - SNF", "address": "2315 Highway 242 North", "city": "Benson", "state": "NC", "zip": "27504", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "919-207-1717", "website": ""},
    # Lee County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Lee County", "division": "LHR - SNF", "address": "310 Commerce Drive", "city": "Sanford", "state": "NC", "zip": "27332", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "919-499-2206", "website": ""},
    {"name": "Westfield Rehabilitation and Health Center", "division": "LHR - SNF", "address": "3100 Tramway Road", "city": "Sanford", "state": "NC", "zip": "27332", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "919-775-5404", "website": ""},
    # Mecklenburg County (continued)
    {"name": "Royal Park Rehabilitation and Health Center", "division": "LHR - SNF", "address": "2700 Royal Commons Lane", "city": "Matthews", "state": "NC", "zip": "28105", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "704-849-6990", "website": ""},
    {"name": "The Pavilion Health Center at Brightmore", "division": "LHR - SNF", "address": "10011 Providence Road West", "city": "Charlotte", "state": "NC", "zip": "28277", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On Brightmore of South Charlotte campus.", "phone": "980-245-8500", "website": ""},
    # Moore County
    {"name": "Pinehurst Healthcare & Rehabilitation Center", "division": "LHR - SNF", "address": "300 Blake Road", "city": "Pinehurst", "state": "NC", "zip": "28374", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-295-6158", "website": ""},
    {"name": "The Inn at Quail Haven Village", "division": "LHR - SNF", "address": "155 Blake Road", "city": "Pinehurst", "state": "NC", "zip": "28374", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On Quail Haven Village Life Plan campus.", "phone": "910-295-2294", "website": ""},
    # New Hanover County
    {"name": "Bradley Creek at Carolina Bay", "division": "LHR - SNF", "address": "740 Diamond Shoals Road", "city": "Wilmington", "state": "NC", "zip": "28403", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On Carolina Bay at Autumn Hall campus.", "phone": "910-769-7550", "website": ""},
    {"name": "Liberty Commons Rehabilitation Center", "division": "LHR - SNF", "address": "121 Racine Drive", "city": "Wilmington", "state": "NC", "zip": "28403", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-452-4070", "website": ""},
    # Orange County
    {"name": "Parkview Health & Rehabilitation Center", "division": "LHR - SNF", "address": "1718 Legion Road", "city": "Chapel Hill", "state": "NC", "zip": "27517", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "984-234-3600", "website": ""},
    # Person County
    {"name": "Roxboro Healthcare & Rehabilitation Center", "division": "LHR - SNF", "address": "901 Ridge Road", "city": "Roxboro", "state": "NC", "zip": "27573", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-599-0106", "website": ""},
    # Robeson County
    {"name": "WoodHaven Nursing Center", "division": "LHR - SNF", "address": "1150 Pine Run Dr.", "city": "Lumberton", "state": "NC", "zip": "28358", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-671-5703", "website": ""},
    # Rowan County
    {"name": "Liberty Commons Nursing & Rehabilitation Center of Rowan County", "division": "LHR - SNF", "address": "4412 South Main Street", "city": "Salisbury", "state": "NC", "zip": "28147", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "704-637-3040", "website": ""},
    # Sampson County
    {"name": "Mary Gran Nursing Center", "division": "LHR - SNF", "address": "120 Southwood Drive", "city": "Clinton", "state": "NC", "zip": "28329", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-592-7981", "website": ""},
    {"name": "Southwood Nursing & Rehabilitation Center", "division": "LHR - SNF", "address": "180 Southwood Drive", "city": "Clinton", "state": "NC", "zip": "28329", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "910-592-8165", "website": ""},
    # Wake County
    {"name": "Capital Nursing & Rehabilitation Center", "division": "LHR - SNF", "address": "3000 Holston Lane", "city": "Raleigh", "state": "NC", "zip": "27610", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "919-231-6045", "website": ""},
    {"name": "Bloomsbury at Hayes Barton Place", "division": "LHR - SNF", "address": "2750 Oberlin Road", "city": "Raleigh", "state": "NC", "zip": "27608", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On Hayes Barton Place Life Plan campus.", "phone": "919-703-6365", "website": ""},
    {"name": "Swift Creek at The Templeton", "division": "LHR - SNF", "address": "221 Brightmore Drive", "city": "Cary", "state": "NC", "zip": "27518", "care": "Skilled Nursing, Rehab", "units": "", "campus": "On The Templeton of Cary campus.", "phone": "984-465-4800", "website": ""},
    # Warren County
    {"name": "Warren Hills Rehabilitation & Nursing Center", "division": "LHR - SNF", "address": "864 US Hwy. 158 Business West", "city": "Warrenton", "state": "NC", "zip": "27589", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "252-257-2011", "website": ""},
    # Watauga County
    {"name": "The Foley Center at Chestnut Ridge", "division": "LHR - SNF", "address": "621 Chestnut Ridge Parkway", "city": "Blowing Rock", "state": "NC", "zip": "28605", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "828-386-3300", "website": ""},
    # Yadkin County
    {"name": "Yadkin Nursing Care Center & Magnolias Over Yadkin", "division": "LHR - SNF", "address": "903 W. Main Street", "city": "Yadkinville", "state": "NC", "zip": "27055", "care": "Skilled Nursing, Rehab", "units": "", "campus": "", "phone": "336-679-8863", "website": ""},
    # --- SOUTH CAROLINA ---
    # Charleston County
    {"name": "Kempton of Charleston", "division": "LHR - SNF", "address": "194 Spring Street", "city": "Charleston", "state": "SC", "zip": "29403", "care": "Skilled Nursing, Rehab", "units": "", "campus": "Also listed under LSL (AL/Memory). Dual-division campus.", "phone": "854-500-7778", "website": ""},
    # --- TENNESSEE ---
    # Cumberland County
    {"name": "Preserve at Fairfield Glade", "division": "LHR - SNF", "address": "100 Samaritan Way", "city": "Crossville", "state": "TN", "zip": "38558", "care": "Skilled Nursing, Rehab", "units": "", "campus": "Also listed under LSL (IL/AL). Dual-division campus.", "phone": "931-429-0110", "website": ""},
]

ws2 = wb.create_sheet("LHR - Healthcare & Rehab")
write_header(ws2)

rn = 2
for fac in lhr_facilities:
    vals = [
        fac["name"], fac["division"], fac["address"], fac["city"],
        fac["state"], fac["zip"], fac["care"], fac["units"],
        fac["campus"], fac["phone"], fac["website"],
    ]
    write_row(ws2, rn, vals, blue_fill)
    rn += 1

nc_count = sum(1 for f in lhr_facilities if f["state"] == "NC")
sc_count = sum(1 for f in lhr_facilities if f["state"] == "SC")
tn_count = sum(1 for f in lhr_facilities if f["state"] == "TN")
print(f"LHR Tab: {len(lhr_facilities)} facilities (NC={nc_count}, SC={sc_count}, TN={tn_count})")

# ── Tab 3: Other Divisions (blank) ──
ws3 = wb.create_sheet("Other Divisions")
write_header(ws3)
print("Other Divisions Tab: blank")

for ws in [ws1, ws2, ws3]:
    autofit(ws, ws.max_row)

outpath = ensure_report_dir() / "Liberty_Facility_Crosswalk.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
