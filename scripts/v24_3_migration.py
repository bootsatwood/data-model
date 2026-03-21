"""
V24.3 Migration -- Resolve 5 deferred items from V24.2.
Content-based matching with verification.

Deferred items:
  1. Sweet Galilee (Gardant duplicate) -- city was wrong
  2. Morgan at Ford's Village (RUI phantom) -- corrupted apostrophe
  3. Woodland Hills (RUI duplicate) -- 2 matches, needed differentiator
  4. Westmont at Short Pump (RUI duplicate) -- 2 matches, needed city
  5. Sunrise of Five Forks (Vitality duplicate) -- wrong state (GA not SC)
"""
import openpyxl
import re

V24_2 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_2.xlsx"
V24_3 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_3.xlsx"

COL = {
    'type': 1, 'name': 2, 'corp': 3, 'addr': 4, 'city': 5,
    'state': 6, 'zip': 7, 'county': 8, 'beds': 10, 'census': 11,
    'served': 12, 'int': 13, 'pcp': 14, 'mh': 15, 'dq': 19,
    'corp_source': 26
}

def norm(s):
    if not s: return ''
    s = str(s).upper().strip()
    s = re.sub(r'[^\x00-\x7F]', '', s)  # strip non-ASCII (handles mojibake)
    s = re.sub(r'\.(?=\s|$)', '', s)
    s = re.sub(r'(?<=\b[A-Z])\.', '', s)
    s = ' '.join(s.split())
    return s

def find_row(ws, name_frag, state, city=None, corp=None):
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        row_name = norm(ws.cell(row_idx, COL['name']).value)
        row_state = norm(ws.cell(row_idx, COL['state']).value)
        if name_frag.upper() not in row_name:
            continue
        if state.upper() != row_state:
            continue
        if city and city.upper() != norm(ws.cell(row_idx, COL['city']).value):
            continue
        if corp and corp.upper() != norm(ws.cell(row_idx, COL['corp']).value):
            continue
        matches.append(row_idx)
    if len(matches) == 1:
        return matches[0]
    elif len(matches) == 0:
        return None
    else:
        return matches

# ============================================================
# LOAD
# ============================================================
print("Loading V24.2...")
wb = openpyxl.load_workbook(V24_2)
ws = wb.active
total_rows = ws.max_row - 1
print(f"V24.2 rows: {total_rows}")

errors = []
deletes = []
rows_to_delete = []

# ============================================================
# 1. Sweet Galilee -- DELETE the duplicate (SWEET GALILEE AT THE WIGWAM)
#    Keep id 5168 (GARDANT MANAGEMENT SOLUTIONS)
#    Delete the one under "SWEET GALILEE AT THE WIGWAM LLC" in Anderson, IN
# ============================================================
print("\n1. Sweet Galilee...")
row = find_row(ws, "SWEET GALILEE AT THE WIGWAM", "IN", city="Anderson")
if row is None:
    errors.append("Sweet Galilee: NOT FOUND")
elif isinstance(row, list):
    errors.append(f"Sweet Galilee: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    corp = ws.cell(row, COL['corp']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    deletes.append(f"Row {row}: {fac} | {corp} | Anderson, IN | beds={beds} | Gardant duplicate")

# ============================================================
# 2. Morgan at Ford's Village -- DELETE phantom (never built)
#    Corrupted apostrophe in name, use MORGAN + Williamsburg + VA
# ============================================================
print("2. Morgan at Ford's Village...")
row = find_row(ws, "MORGAN", "VA", city="Williamsburg")
if row is None:
    errors.append("Morgan at Ford's Village: NOT FOUND")
elif isinstance(row, list):
    errors.append(f"Morgan at Ford's Village: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    corp = ws.cell(row, COL['corp']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    deletes.append(f"Row {row}: {fac} | {corp} | Williamsburg, VA | beds={beds} | RUI phantom -- never built")

# ============================================================
# 3. Woodland Hills -- DELETE the surrogate (WOODLAND HILLS COMMUNITY)
#    Keep WOODLAND HILLS INDEPENDENT LIVING, AL & MC (served, has county)
#    Both were renamed from RUI to Retirement Unlimited, Inc. in V24.2
# ============================================================
print("3. Woodland Hills...")
row = find_row(ws, "WOODLAND HILLS COMMUNITY", "VA", city="Roanoke")
if row is None:
    # Try without city
    row = find_row(ws, "WOODLAND HILLS COMMUNITY", "VA")
if row is None:
    errors.append("Woodland Hills Community: NOT FOUND")
elif isinstance(row, list):
    errors.append(f"Woodland Hills Community: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    corp = ws.cell(row, COL['corp']).value
    beds = ws.cell(row, COL['beds']).value
    county = ws.cell(row, COL['county']).value
    rows_to_delete.append(row)
    deletes.append(f"Row {row}: {fac} | {corp} | Roanoke, VA | beds={beds} | county={county} | RUI duplicate (surrogate, missing county)")

# ============================================================
# 4. Westmont at Short Pump -- DELETE the surrogate (Glen Allen city)
#    Keep the Richmond one (real beds, has county)
# ============================================================
print("4. Westmont at Short Pump...")
row = find_row(ws, "WESTMONT AT SHORT PUMP", "VA", city="Glen Allen")
if row is None:
    errors.append("Westmont at Short Pump (Glen Allen): NOT FOUND")
elif isinstance(row, list):
    errors.append(f"Westmont at Short Pump (Glen Allen): MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    corp = ws.cell(row, COL['corp']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    deletes.append(f"Row {row}: {fac} | {corp} | Glen Allen, VA | beds={beds} | RUI duplicate (surrogate)")

# ============================================================
# 5. Sunrise of Five Forks -- DELETE the Vitality misattribution
#    Correct Sunrise record (id 22213) already exists under Sunrise Senior Living
#    Delete the VITALITY LIVING duplicate in Lilburn, GA
# ============================================================
print("5. Sunrise of Five Forks...")
row = find_row(ws, "SUNRISE OF FIVE FORKS", "GA", city="Lilburn")
if row is None:
    # Try broader match
    row = find_row(ws, "SUNRISE OF FIVE FORKS", "GA")
if row is None:
    # Try without state filter but with Vitality corp
    row = find_row(ws, "SUNRISE OF FIVE FORKS", "GA", corp="VITALITY LIVING")
if row is None:
    errors.append("Sunrise of Five Forks: NOT FOUND in GA")
elif isinstance(row, list):
    errors.append(f"Sunrise of Five Forks: MULTIPLE MATCHES ({len(row)})")
else:
    fac = ws.cell(row, COL['name']).value
    corp = ws.cell(row, COL['corp']).value
    beds = ws.cell(row, COL['beds']).value
    rows_to_delete.append(row)
    deletes.append(f"Row {row}: {fac} | {corp} | Lilburn, GA | beds={beds} | Vitality misattribution (Sunrise duplicate)")

# ============================================================
# APPLY DELETES (bottom-up)
# ============================================================
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx)

# ============================================================
# REPORT
# ============================================================
print(f"\n{'='*60}")
print(f"V24.3 Migration Report -- Deferred V24.2 Items")
print(f"{'='*60}")
print(f"Deletes: {len(deletes)}")
print(f"Final row count: {ws.max_row - 1}")
print()

if deletes:
    print("DELETES:")
    for d in deletes:
        print(f"  {d}")
    print()

# ============================================================
# VALIDATION
# ============================================================
expected_final = total_rows - len(deletes)
actual_final = ws.max_row - 1
if actual_final != expected_final:
    errors.append(f"Row count mismatch: expected {expected_final}, got {actual_final}")

if errors:
    print(f"!! {len(errors)} ERRORS !!")
    for e in errors:
        print(f"  {e}")
    critical = [e for e in errors if "Row count mismatch" in e or "MULTIPLE MATCHES" in e]
    if critical:
        print("\nCRITICAL ERRORS -- NOT SAVING.")
    else:
        print(f"\nSaving with {len(errors)} non-critical warnings...")
        print(f"Saving to {V24_3}...")
        wb.save(V24_3)
        print("Done.")
else:
    print("All validations passed.")
    print(f"Saving to {V24_3}...")
    wb.save(V24_3)
    print("Done.")

wb.close()
