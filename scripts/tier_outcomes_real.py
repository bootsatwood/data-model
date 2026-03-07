"""Final tier outcomes using Brooke's real V20 qualitative scores.
New ER/RP brackets: ER-C (9/13/20/40), RP-B ($1M/$2.5M/$5M/$10M).
Excludes T4/T5. New entities get IR=2, SI=2, RS=computed, AI=0."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
TIERING_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V20.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

# T4/T5 exclusions
EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'Sunnyside Communities', 'ATRIUM HEALTH', 'Warm Hearth Village', 'Brighton',
    'Momentus Health',
}

# Finance name -> DB corp names
FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'CLEARVIEW': {'CLEARVIEW'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'CARDON & ASSOCIATES': {'CARDON & ASSOCIATES'},
}

# Finance name -> V20 workbook name (for qualitative score lookup)
FINANCE_TO_V20 = {
    'ALG': 'ALG',
    'AMERICAN SENIOR COMMUNITIES': 'American Senior Communities',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'SABER HEALTHCARE GROUP': 'Saber Healthcare Group',
    'INFINITY HEALTHCARE CONSULTING': 'Infinity Healthcare Consulting',
    'NAVION': 'Navion',
    'Majestic Care': 'Majestic Care',
    'PRUITT HEALTH': 'Pruitthealth',
    'Kisco Senior Living': None,  # Not in V20
    'TRILOGY': 'Trilogy Health Services',
    'Pavilion Healthcare': 'Pavilion Healthcare',
    'TOPAZ HEALTHCARE': 'Topaz',
    'MORNING POINTE SENIOR LIVING': 'Morning Pointe Senior Living',
    'PRINCIPLE': 'Principle Long Term Care',
    'Liberty': 'Liberty Senior Living',
    'TERRABELLA SENIOR LIVING': 'Terra Bella',
    'SANSTONE': 'Sanstone Health & Rehabilitation',
    'LIONSTONE CARE': 'Lionstone Care',
    'ELDERCARE PARTNERS': None,
    'OTTERBEIN SENIOR LIFE': 'Otterbein Seniorlife',
    'TLC Management': 'Tlc Management',
    'BHI Senior Living': 'BHI Senior Living',
    'Avardis': 'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care',
    'PEAK RESOURCES': 'Peak Resources, Inc.',
    'ARBORS': 'Arbors At Ohio',
    'AMERICAN HEALTHCARE LLC': None,
    'Runk & Pratt': None,
    'MCAP': None,
    'Lutheran Services Carolinas': 'Lutheran Services Carolina',
    'Lutheran Life Villages': None,
    'Greencroft': None,
    'CCH HEALTHCARE': 'Cch Healthcare',
    'LifeSpire of Virginia': None,
    'SENIOR LIFESTYLE': None,
    'CLEARVIEW': 'Clearview',
    'PRIORITY': 'Priority Life Care',
    'CEDARHURST SENOR LIVING': None,
    'Lifecare': 'Life Care Centers Of America',
    'Kissito Healthcare': 'Kissito',
    'SPRING ARBOR MANAGEMENT': None,
    'YAD': 'Yad Healthcare',
    'STORYPOINT': None,
    'CARING PLACE HEALTHCARE': 'Caring Place Healthcare',
    'Eastern Healthcare Group': None,
    'SONIDA SENIOR LIVING': None,
    'Castle Healthcare': 'Castle Healthcare',
    'JAG': 'Jag Healthcare',
    'FUNDAMENTAL LTC': None,
    'CARDON & ASSOCIATES': None,
}

# --- Load V20 qualitative scores ---
wb_t = openpyxl.load_workbook(TIERING_PATH, data_only=True)
v20_scores = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws_t = wb_t[sn]
    for r in range(2, ws_t.max_row + 1):
        name = ws_t.cell(r, 1).value
        if not name:
            continue
        v20_scores[name] = {
            'ir': ws_t.cell(r, 10).value or 0,
            'si': ws_t.cell(r, 11).value or 0,
            'rs': ws_t.cell(r, 13).value or 0,
            'ai': ws_t.cell(r, 14).value or 0,
        }

# --- Load Finance entity list ---
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']

finance_entities = []
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    tier = ws_a.cell(r, 18).value
    glr_facs = ws_a.cell(r, 4).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in EXCLUDE:
        finance_entities.append({
            'name': str(name).strip(),
            'fin_tier': tier,
            'glr_facs': int(glr_facs) if glr_facs else 0,
        })

# --- Load V23 DB ---
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb_db.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0})

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if not corp or state not in FOOTPRINT:
        continue
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0

    d = corp_fp[corp]
    d['campuses'].add((addr, city))
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served:
        d['served_camps'].add((addr, city))

# --- Scoring ---
def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def rs_score(srv, total):
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

ER_T = [9, 13, 20, 40]
RP_T = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

# --- Score each entity ---
results = []
for fe in finance_entities:
    db_names = FINANCE_TO_DB.get(fe['name'])
    if not db_names:
        continue

    camp = set()
    srv = set()
    rev = 0
    for dn in db_names:
        d = corp_fp.get(dn)
        if d:
            camp |= d['campuses']
            srv |= d['served_camps']
            rev += d['tot_pot']

    if len(camp) < 7:
        continue

    # Get qualitative scores
    v20_name = FINANCE_TO_V20.get(fe['name'])
    if v20_name and v20_name in v20_scores:
        qs = v20_scores[v20_name]
        ir = qs['ir']
        si = qs['si']
        ai = qs['ai']
        qual_source = 'V20'
    else:
        ir = 2
        si = 2
        ai = 0
        qual_source = 'Default'

    er = score_dim(len(camp), ER_T)
    rp = score_dim(rev, RP_T)
    rs = rs_score(len(srv), len(camp))

    score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = 'T1' if score >= 50 else 'T2' if score >= 25 else 'T3'

    results.append({
        'name': fe['name'],
        'fin_tier': fe['fin_tier'],
        'camp': len(camp),
        'srv': len(srv),
        'rev': rev,
        'er': er, 'ir': ir, 'si': si, 'rp': rp, 'rs': rs, 'ai': ai,
        'score': score, 'tier': tier, 'qual_source': qual_source,
    })

results.sort(key=lambda x: (-x['score'], x['name']))

# --- Output ---
print(f"Scoreable: {len(results)} entities")
print(f"Brackets: ER (9/13/20/40)  RP ($1M/$2.5M/$5M/$10M)")
print()

print(f"{'Entity':<45} {'Fin':>4} {'V23':>4} {'Scr':>4}  {'Camp':>5} {'Srv':>4}  {'ER':>3} {'IR':>3} {'SI':>3} {'RP':>3} {'RS':>3} {'AI':>3}  {'Qual':>7}")
print("-" * 115)
for d in results:
    ft = str(d['fin_tier'] or '?').strip()
    if 'not on list' in ft.lower():
        ft = 'NEW'
    elif 'pulled' in ft.lower():
        ft = 'T1*'
    elif ft in ('1', '2', '3'):
        ft = f'T{ft}'
    print(f"{d['name']:<45} {ft:>4} {d['tier']:>4} {d['score']:>4}  {d['camp']:>5} {d['srv']:>4}  {d['er']:>3} {d['ir']:>3} {d['si']:>3} {d['rp']:>3} {d['rs']:>3} {d['ai']:>3}  {d['qual_source']:>7}")

tiers = Counter(d['tier'] for d in results)
print("-" * 115)
print(f"TOTALS:  T1={tiers.get('T1',0)}  T2={tiers.get('T2',0)}  T3={tiers.get('T3',0)}")

# Finance comparison
fin_t1 = sum(1 for r in results if str(r['fin_tier']).strip() == '1' or 'pulled' in str(r['fin_tier']).lower())
fin_t2 = sum(1 for r in results if str(r['fin_tier']).strip() == '2')
fin_t3 = sum(1 for r in results if str(r['fin_tier']).strip() == '3')
print(f"Finance: T1={fin_t1}  T2={fin_t2}  T3={fin_t3}")

# Show tier changes
print()
print("--- Tier Changes vs Finance ---")
for d in results:
    ft = str(d['fin_tier'] or '').strip()
    if 'not on list' in ft.lower():
        print(f"  NEW  -> {d['tier']}  {d['name']}")
    elif 'pulled' in ft.lower():
        if d['tier'] != 'T1':
            print(f"  T1*  -> {d['tier']}  {d['name']}")
    elif ft in ('1', '2', '3'):
        old = f'T{ft}'
        if old != d['tier']:
            print(f"  {old}   -> {d['tier']}  {d['name']}")
