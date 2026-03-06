#!/usr/bin/env python3
"""
Southern Healthcare Management Consolidation — V22.9 -> V22.10

Resolves DCR #1 (Southern rename) and DCR #2 (missing facilities):
  - SOVEREIGN HEALTHCARE HOLDINGS (43 rows, CMS chain name) is the same entity
    as SOUTHERN HEALTHCARE MANAGEMENT, LLC (GLR operating company name)
  - Evidence: identical state distribution (FL 31, GA 2, NC 10), NC facilities
    at same addresses as GLR-confirmed Southern Healthcare facilities
  - Also normalizes "SOUTHERN HEALTHCARE" (1 LEGACY row) to the GLR canonical

Usage:
  python v22_10_southern_fix.py preview
  python v22_10_southern_fix.py apply
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, VAULT

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_9.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_10.xlsx"

# ---------------------------------------------------------------------------
# Corporate Name Renames
# ---------------------------------------------------------------------------

RENAMES = {
    # CMS chain name -> GLR operating company name
    "SOVEREIGN HEALTHCARE HOLDINGS": "SOUTHERN HEALTHCARE MANAGEMENT, LLC",
    # LEGACY variant -> GLR canonical
    "SOUTHERN HEALTHCARE": "SOUTHERN HEALTHCARE MANAGEMENT, LLC",
}


# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    if not corp_name or corp_name.upper() in ('UNKNOWN', ''):
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 0) >= 2:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python v22_10_southern_fix.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Southern Healthcare Consolidation — V22.9 -> V22.10 ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_SOURCE.name}...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # --- Phase 1: Corporate Name Renames ---
    print("\nPhase 1: Corporate Name Renames")
    print("-" * 40)

    rename_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp in RENAMES:
            new_corp = RENAMES[corp]
            rename_counts[f"{corp} -> {new_corp}"] += 1
            r['Corporate_Name'] = new_corp

    for key in sorted(rename_counts.keys()):
        print(f"  {rename_counts[key]:>4d}  {key}")

    total_renames = sum(rename_counts.values())
    print(f"\n  {total_renames} total renames")

    # --- Phase 2: Four-Rule Reclassification ---
    print(f"\nPhase 2: Four-Rule Ownership Reclassification")
    print("-" * 40)

    corp_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp:
            corp_counts[corp] += 1

    affected_corps = set(RENAMES.keys()) | set(RENAMES.values())
    reclass_count = 0
    reclass_detail = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp not in affected_corps:
            continue
        old_type = safe(r.get('Ownership_Type', ''))
        new_type = four_rule_classify(corp, corp_counts)
        if old_type != new_type:
            reclass_count += 1
            reclass_detail[f"{old_type} -> {new_type}"] += 1
            r['Ownership_Type'] = new_type

    print(f"  {reclass_count} reclassifications")
    for key in sorted(reclass_detail.keys()):
        print(f"    {reclass_detail[key]:>4d}  {key}")

    # --- Summary ---
    canonical = "SOUTHERN HEALTHCARE MANAGEMENT, LLC"
    cnt = corp_counts.get(canonical, 0)
    served = sum(1 for r in rows if safe(r.get('Corporate_Name', '')) == canonical
                 and safe(r.get('Do_We_Serve', '')) == 'Yes')
    states = Counter()
    for r in rows:
        if safe(r.get('Corporate_Name', '')) == canonical:
            states[safe(r.get('State', ''))] += 1

    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Total renames:                 {total_renames}")
    print(f"  Ownership reclassifications:   {reclass_count}")
    print(f"  Rows deleted:                  0")
    print(f"  Final row count:               {len(rows):,}")
    print(f"\n  SOUTHERN HEALTHCARE MANAGEMENT, LLC:")
    print(f"    {cnt} total rows, {served} served")
    print(f"    States: {dict(sorted(states.items()))}")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY -- no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.10 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    corp_col = col_idx.get('Corporate_Name')
    own_col = col_idx.get('Ownership_Type')

    cells_updated = 0
    for r in rows:
        excel_row = r['_excel_row']

        old_corp = safe(ws.cell(row=excel_row, column=corp_col).value)
        new_corp = safe(r.get('Corporate_Name', ''))
        if old_corp != new_corp:
            ws.cell(row=excel_row, column=corp_col).value = new_corp
            cells_updated += 1

        old_own = safe(ws.cell(row=excel_row, column=own_col).value)
        new_own = safe(r.get('Ownership_Type', ''))
        if old_own != new_own:
            ws.cell(row=excel_row, column=own_col).value = new_own
            cells_updated += 1

    wb.save(DB_OUTPUT)
    print(f"  {cells_updated} cells updated")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
