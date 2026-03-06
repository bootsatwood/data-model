#!/usr/bin/env python3
"""Build Liberty Phases 1b/1c/1d QC review workbook."""

import sys
sys.path.insert(0, '.')
from utils import load_db, norm_addr, norm, safe, ensure_report_dir, VAULT
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB = VAULT / "Current" / "1_Combined_Database_FINAL_V22_1.xlsx"
headers, rows = load_db(DB)
print(f"V22.1 loaded: {len(rows):,} rows")

# Styles
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

cols = [
    "Phase", "Action", "Old Corporate_Name", "New Corporate_Name", "Excel Row",
    "Facility_Name", "Address", "City", "State", "ZIP", "County",
    "Source_Type", "Total_Beds", "Census", "Do_We_Serve",
    "Integrated_Flag", "PCP_Flag", "MH_Flag", "Barrier",
    "Contract_Status", "Ownership_Type", "Geographic_Tier", "Metro_Assignment",
    "Latitude", "Longitude",
]


def write_header(ws):
    for c, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"


def write_row(ws, row_num, vals, fill):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.fill = fill
        cell.border = thin_border


def row_vals(phase, action, old_corp, new_corp, r):
    return [
        phase, action, old_corp, new_corp, r["_excel_row"],
        safe(r.get("Facility_Name", "")), safe(r.get("Address", "")),
        safe(r.get("City", "")), safe(r.get("State", "")),
        safe(r.get("ZIP", "")), safe(r.get("County", "")),
        safe(r.get("Source_Type", "")),
        r.get("Total_Beds", ""), r.get("Census", ""),
        safe(r.get("Do_We_Serve", "")), safe(r.get("Integrated_Flag", "")),
        safe(r.get("PCP_Flag", "")), safe(r.get("MH_Flag", "")),
        safe(r.get("Barrier", "")), safe(r.get("Contract_Status", "")),
        safe(r.get("Ownership_Type", "")), safe(r.get("Geographic_Tier", "")),
        safe(r.get("Metro_Assignment", "")),
        r.get("Latitude", ""), r.get("Longitude", ""),
    ]


def autofit(ws, max_row):
    for c in range(1, len(cols) + 1):
        max_len = len(str(ws.cell(row=1, column=c).value))
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max_len + 3


wb = Workbook()

# ── SHEET 1: Phase 1b ──
ws1 = wb.active
ws1.title = "1b - Liberty Short Name"
write_header(ws1)

p1b = [r for r in rows if safe(r.get("Corporate_Name", "")) == "LIBERTY"]
rn = 2
for r in sorted(p1b, key=lambda x: (safe(x.get("State", "")), safe(x.get("City", "")))):
    write_row(ws1, rn, row_vals("1b", "RENAME", "LIBERTY", "LIBERTY SENIOR LIVING", r), blue_fill)
    rn += 1
print(f"  Sheet 1b: {len(p1b)} rename rows")

# ── SHEET 2: Phase 1c ──
ws2 = wb.create_sheet("1c - Liberty Village")
write_header(ws2)

LIBERTY_VILLAGE_CORPS = {
    "Liberty Village of Freeport", "Liberty Village of Geneseo",
    "Liberty Village of Jerseyville", "Liberty Village of Peoria",
    "Liberty Village of Peru", "Liberty Village of Rochelle",
    "Liberty Village of Streator",
}
p1c = [r for r in rows if safe(r.get("Corporate_Name", "")) in LIBERTY_VILLAGE_CORPS]
rn = 2
for r in sorted(p1c, key=lambda x: safe(x.get("Corporate_Name", ""))):
    old_corp = safe(r.get("Corporate_Name", ""))
    write_row(ws2, rn, row_vals("1c", "RENAME", old_corp, "LIBERTY VILLAGE", r), blue_fill)
    rn += 1
print(f"  Sheet 1c: {len(p1c)} rename rows")

# ── SHEET 3: Phase 1d ──
ws3 = wb.create_sheet("1d - Liberty PROPCO Dupes")
write_header(ws3)

LIBERTY_PROPCO_SNF_PATTERN = "LIBERTY HEALTHCARE PROPERTIES"
rows_by_addr = defaultdict(list)
for r in rows:
    key = (norm_addr(safe(r.get("Address", "")))
           + "|" + norm(safe(r.get("City", "")))
           + "|" + norm(safe(r.get("State", ""))))
    r["_addr_key"] = key
    rows_by_addr[key].append(r)

pairs = []
for r in rows:
    corp_upper = safe(r.get("Corporate_Name", "")).upper()
    stype = safe(r.get("Source_Type", ""))
    if stype == "SNF" and LIBERTY_PROPCO_SNF_PATTERN in corp_upper:
        key = r["_addr_key"]
        siblings = rows_by_addr.get(key, [])
        real_snf = [
            s for s in siblings
            if s is not r
            and safe(s.get("Source_Type", "")) == "SNF"
            and safe(s.get("Corporate_Name", "")).upper() == "LIBERTY SENIOR LIVING"
        ]
        if not real_snf:
            city = safe(r.get("City", "")).upper()
            state = safe(r.get("State", "")).upper()
            fac_prefix = norm(safe(r.get("Facility_Name", "")))[:15]
            real_snf = [
                s for s in rows
                if s is not r
                and safe(s.get("Source_Type", "")) == "SNF"
                and safe(s.get("City", "")).upper() == city
                and safe(s.get("State", "")).upper() == state
                and safe(s.get("Corporate_Name", "")).upper() == "LIBERTY SENIOR LIVING"
                and norm(safe(s.get("Facility_Name", "")))[:15] == fac_prefix
            ]
        if real_snf:
            pairs.append((r, real_snf[0]))

rn = 2
for rem, keep in sorted(pairs, key=lambda x: safe(x[0].get("City", ""))):
    # KEEP row first (green)
    write_row(ws3, rn,
              row_vals("1d", "KEEP",
                       safe(keep.get("Corporate_Name", "")), "", keep),
              green_fill)
    rn += 1
    # REMOVE row (red)
    write_row(ws3, rn,
              row_vals("1d", "REMOVE",
                       safe(rem.get("Corporate_Name", "")), "", rem),
              red_fill)
    rn += 1

print(f"  Sheet 1d: {len(pairs)} pairs ({len(pairs)} KEEP + {len(pairs)} REMOVE)")

# Auto-fit all sheets
for ws in [ws1, ws2, ws3]:
    autofit(ws, ws.max_row)

outpath = ensure_report_dir() / "Liberty_Phases_1b_1c_1d_Review.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
