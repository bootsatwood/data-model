#!/usr/bin/env python3
"""
ALF Different-Name Pair Classification — V23 → V24 Prep

Finds all addresses with 2+ ALF rows that have DIFFERENT facility names,
then classifies each pair as likely-duplicate, likely-legitimate, or
needs-manual-review using NIC Maps unit data, name pattern analysis,
and attribution source signals.

Outputs a 3-tab Excel review workbook:
  Tab 1: Auto-Classified Duplicates (high confidence)
  Tab 2: Auto-Classified Legitimate (MC vs AL campus pairs)
  Tab 3: Manual Review (everything else)

Usage:
  python alf_diffname_classify.py
"""

import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils import safe, norm, addr_key, load_db, load_nic_alf, ensure_report_dir


# ---------------------------------------------------------------------------
# NIC index (reused from nic_enrichment.py)
# ---------------------------------------------------------------------------

def _int(val):
    try:
        return int(val or 0)
    except (ValueError, TypeError):
        return 0


def build_nic_index(nic_rows):
    """Index NIC buildings by address key. Keep largest building per address."""
    index = {}
    for row in nic_rows:
        key = addr_key(
            safe(row.get('Address')),
            safe(row.get('City')),
            safe(row.get('State')),
        )
        if not key or key == '||':
            continue
        total = _int(row.get('Total Units'))
        if key in index:
            if total <= _int(index[key].get('Total Units')):
                continue
        index[key] = row
    return index


# ---------------------------------------------------------------------------
# Name analysis helpers
# ---------------------------------------------------------------------------

# Suffixes that indicate care type
MC_INDICATORS = re.compile(
    r'\b(memory\s*care|mc|alzheimer|dementia|memory)\b', re.I
)
AL_INDICATORS = re.compile(
    r'\b(assisted\s*living|assisted|al\b|personal\s*care)', re.I
)
IL_INDICATORS = re.compile(
    r'\b(independent\s*living|independent|retirement\s*(home|community|village|center)|'
    r'senior\s*(apartments?|living)|55\+|active\s*adult)\b', re.I
)
TYPE_SUFFIXES = re.compile(
    r'\s*[-–—]\s*(assisted\s*living|memory\s*care|personal\s*care|'
    r'skilled\s*nursing|rehabilitation|rehab|snf|alf|mc|al|il)\s*$', re.I
)


def strip_type_suffix(name):
    """Remove trailing care-type suffix from a facility name."""
    return TYPE_SUFFIXES.sub('', name).strip()


def name_is_substring(a, b):
    """Check if one name substantially contains the other."""
    na, nb = norm(a), norm(b)
    if not na or not nb:
        return False
    if len(na) < 5 or len(nb) < 5:
        return False
    return na in nb or nb in na


def names_match_after_strip(a, b):
    """Check if names match after removing type suffixes."""
    sa = norm(strip_type_suffix(a))
    sb = norm(strip_type_suffix(b))
    if not sa or not sb:
        return False
    return sa == sb


def one_mc_one_al(name_a, name_b):
    """Check if one name signals MC and the other signals AL."""
    a_mc = bool(MC_INDICATORS.search(name_a))
    a_al = bool(AL_INDICATORS.search(name_a))
    b_mc = bool(MC_INDICATORS.search(name_b))
    b_al = bool(AL_INDICATORS.search(name_b))
    return (a_mc and b_al) or (a_al and b_mc)


# ---------------------------------------------------------------------------
# Row scoring (from v22_13)
# ---------------------------------------------------------------------------

def score_row(r):
    """Score a row by data completeness. Higher = keep."""
    s = 0
    if safe(r.get('Do_We_Serve', '')) == 'Yes':
        s += 1000
    corp = safe(r.get('Corporate_Name', ''))
    if corp and corp != 'INDEPENDENT':
        s += 10
    src = safe(r.get('Corp_Attribution_Source', ''))
    if src == 'GLR':
        s += 8
    elif src == 'CMS':
        s += 6
    elif src.startswith('NIC'):
        s += 4
    elif src == 'LEGACY':
        s += 2
    for h, v in r.items():
        if h.startswith('_'):
            continue
        if safe(v):
            s += 1
    return s


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify_pair(rows, nic_index):
    """Classify a group of 2+ ALF rows at the same address with different names.

    Returns a dict with classification, reasoning, and row details.
    """
    ak = addr_key(safe(rows[0].get('Address', '')),
                   safe(rows[0].get('City', '')),
                   safe(rows[0].get('State', '')))
    nic = nic_index.get(ak)

    al_units = _int(nic.get('AL Open Units')) if nic else 0
    mc_units = _int(nic.get('MC Open Units')) if nic else 0
    il_units = _int(nic.get('IL Open Units')) if nic else 0
    nc_units = _int(nic.get('NC Open Units')) if nic else 0
    total_units = _int(nic.get('Total Units')) if nic else 0
    nic_matched = nic is not None

    names = [safe(r.get('Facility_Name', '')) for r in rows]
    corps = [safe(r.get('Corporate_Name', '')) or '(blank)' for r in rows]
    sources = [safe(r.get('Corp_Attribution_Source', '')) or '-' for r in rows]
    served = [safe(r.get('Do_We_Serve', '')) for r in rows]
    any_served = any(s == 'Yes' for s in served)
    both_served = all(s == 'Yes' for s in served)

    reasons = []
    classification = 'REVIEW'  # default

    # --- Test 1: Names match after stripping type suffixes ---
    if len(rows) == 2 and names_match_after_strip(names[0], names[1]):
        reasons.append('Names match after removing type suffix')
        classification = 'LIKELY_DUPLICATE'

    # --- Test 2: One name is substring of the other ---
    if len(rows) == 2 and name_is_substring(names[0], names[1]):
        if classification != 'LIKELY_DUPLICATE':
            reasons.append('One name contains the other')
            classification = 'LIKELY_DUPLICATE'

    # --- Test 3: MC vs AL name indicators ---
    if len(rows) == 2 and one_mc_one_al(names[0], names[1]):
        reasons.append('One name signals MC, other signals AL')
        if nic_matched and mc_units > 0 and al_units > 0:
            reasons.append(f'NIC confirms AL={al_units} + MC={mc_units} units')
            classification = 'LIKELY_LEGITIMATE'
        elif nic_matched and mc_units == 0:
            reasons.append(f'NIC shows MC=0 — only AL={al_units} units')
            classification = 'LIKELY_DUPLICATE'
        else:
            classification = 'REVIEW'

    # --- Test 4: NIC has only 1 type of non-NC unit ---
    if nic_matched and classification == 'REVIEW':
        unit_types_present = sum([al_units > 0, mc_units > 0, il_units > 0])
        if unit_types_present <= 1 and total_units > 0:
            reasons.append(f'NIC shows single unit type (AL={al_units}, MC={mc_units}, IL={il_units})')
            classification = 'LIKELY_DUPLICATE'

    # --- Test 5: NIC has AL + MC → likely two legitimate programs ---
    if nic_matched and al_units > 0 and mc_units > 0 and classification != 'LIKELY_DUPLICATE':
        reasons.append(f'NIC shows AL={al_units} + MC={mc_units} — possible two-program campus')
        if classification == 'REVIEW':
            classification = 'LIKELY_LEGITIMATE'

    # --- Test 6: No NIC match at all ---
    if not nic_matched:
        reasons.append('No NIC match — cannot verify unit types')

    # --- Test 7: IL indicator in one name ---
    if len(rows) == 2:
        il_flags = [bool(IL_INDICATORS.search(n)) for n in names]
        if any(il_flags) and not all(il_flags):
            reasons.append('One name suggests IL/retirement — may be Source_Type issue')

    # --- Determine keep/remove for duplicates ---
    scored = [(score_row(r), i, r) for i, r in enumerate(rows)]
    scored.sort(key=lambda x: x[0], reverse=True)

    return {
        'address_key': ak,
        'address': safe(rows[0].get('Address', '')),
        'city': safe(rows[0].get('City', '')),
        'state': safe(rows[0].get('State', '')),
        'classification': classification,
        'reasons': ' | '.join(reasons) if reasons else 'No automated signal',
        'any_served': any_served,
        'both_served': both_served,
        'nic_matched': nic_matched,
        'al_units': al_units,
        'mc_units': mc_units,
        'il_units': il_units,
        'nc_units': nc_units,
        'total_units': total_units,
        'rows': rows,
        'names': names,
        'corps': corps,
        'sources': sources,
        'served': served,
        'scored': scored,
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
SERVED_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

COLUMNS = [
    ('Address', 30),
    ('City', 18),
    ('State', 6),
    ('Row A — Name', 40),
    ('Row A — Corp', 30),
    ('Row A — Source', 8),
    ('Row A — Served', 8),
    ('Row A — Beds', 8),
    ('Row A — Score', 8),
    ('Row A — Excel Row', 10),
    ('Row B — Name', 40),
    ('Row B — Corp', 30),
    ('Row B — Source', 8),
    ('Row B — Served', 8),
    ('Row B — Beds', 8),
    ('Row B — Score', 8),
    ('Row B — Excel Row', 10),
    ('NIC AL', 7),
    ('NIC MC', 7),
    ('NIC IL', 7),
    ('NIC NC', 7),
    ('NIC Total', 8),
    ('Classification', 18),
    ('Reasoning', 60),
]


def write_pair_row(ws, row_num, pair):
    """Write one pair to the worksheet."""
    rows = pair['rows']
    scored = pair['scored']

    # Row A = highest scored, Row B = second
    ra = scored[0][2]
    rb = scored[1][2] if len(scored) > 1 else {}
    sa = scored[0][0]
    sb = scored[1][0] if len(scored) > 1 else 0

    values = [
        pair['address'], pair['city'], pair['state'],
        safe(ra.get('Facility_Name', '')),
        safe(ra.get('Corporate_Name', '')) or '(blank)',
        safe(ra.get('Corp_Attribution_Source', '')) or '-',
        safe(ra.get('Do_We_Serve', '')),
        safe(ra.get('Total_Beds', '')),
        sa,
        ra.get('_excel_row', ''),
        safe(rb.get('Facility_Name', '')),
        safe(rb.get('Corporate_Name', '')) or '(blank)',
        safe(rb.get('Corp_Attribution_Source', '')) or '-',
        safe(rb.get('Do_We_Serve', '')),
        safe(rb.get('Total_Beds', '')),
        sb,
        rb.get('_excel_row', ''),
        pair['al_units'], pair['mc_units'], pair['il_units'],
        pair['nc_units'], pair['total_units'],
        pair['classification'],
        pair['reasons'],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        if pair['any_served']:
            cell.fill = SERVED_FILL


def setup_sheet(ws, title):
    """Apply headers and column widths to a worksheet."""
    ws.title = title
    for col, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{chr(64 + len(COLUMNS))}1"


def write_workbook(duplicates, legitimate, review, outpath):
    """Write the 3-tab review workbook."""
    wb = Workbook()

    # Tab 1: Likely Duplicates
    ws1 = wb.active
    setup_sheet(ws1, f'Likely Duplicate ({len(duplicates)})')
    for i, pair in enumerate(duplicates, 2):
        write_pair_row(ws1, i, pair)

    # Tab 2: Likely Legitimate
    ws2 = wb.create_sheet()
    setup_sheet(ws2, f'Likely Legitimate ({len(legitimate)})')
    for i, pair in enumerate(legitimate, 2):
        write_pair_row(ws2, i, pair)

    # Tab 3: Manual Review
    ws3 = wb.create_sheet()
    setup_sheet(ws3, f'Manual Review ({len(review)})')
    for i, pair in enumerate(review, 2):
        write_pair_row(ws3, i, pair)

    wb.save(outpath)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("ALF Different-Name Pair Classification")
    print("=" * 55)

    print("\nLoading Combined Database (V23)...")
    headers, rows = load_db()
    print(f"  {len(rows):,} rows loaded.")

    print("Loading NIC Maps source...")
    _, nic_rows = load_nic_alf()
    print(f"  {len(nic_rows):,} NIC buildings loaded.")

    print("Building NIC index...")
    nic_index = build_nic_index(nic_rows)
    print(f"  {len(nic_index):,} NIC address keys.")

    # --- Find addresses with 2+ ALFs with different names ---
    print("\nFinding different-name ALF pairs...")
    alf_by_addr = defaultdict(list)
    for r in rows:
        if safe(r.get('Source_Type', '')) != 'ALF':
            continue
        ak = addr_key(safe(r.get('Address', '')),
                       safe(r.get('City', '')),
                       safe(r.get('State', '')))
        if not ak or ak == '||':
            continue
        alf_by_addr[ak].append(r)

    diff_name_groups = []
    for ak, rlist in alf_by_addr.items():
        if len(rlist) < 2:
            continue
        names = set(norm(safe(r.get('Facility_Name', ''))) for r in rlist)
        if len(names) >= 2:
            diff_name_groups.append(rlist)

    print(f"  {len(diff_name_groups)} addresses with 2+ differently-named ALFs")

    # --- Classify each group ---
    print("\nClassifying pairs...")
    duplicates = []
    legitimate = []
    review = []

    for group in diff_name_groups:
        result = classify_pair(group, nic_index)
        if result['classification'] == 'LIKELY_DUPLICATE':
            duplicates.append(result)
        elif result['classification'] == 'LIKELY_LEGITIMATE':
            legitimate.append(result)
        else:
            review.append(result)

    # Sort: served first, then by state
    for lst in [duplicates, legitimate, review]:
        lst.sort(key=lambda x: (not x['any_served'], x['state'], x['address']))

    # --- Summary ---
    print(f"\n{'=' * 55}")
    print("CLASSIFICATION SUMMARY")
    print(f"{'=' * 55}")
    print(f"  Likely Duplicate:   {len(duplicates):>4}  "
          f"({sum(1 for d in duplicates if d['any_served'])} served)")
    print(f"  Likely Legitimate:  {len(legitimate):>4}  "
          f"({sum(1 for d in legitimate if d['any_served'])} served)")
    print(f"  Manual Review:      {len(review):>4}  "
          f"({sum(1 for d in review if d['any_served'])} served)")
    print(f"  Total:              {len(diff_name_groups):>4}")

    # State breakdown
    print(f"\n  By State:")
    state_counts = defaultdict(int)
    for lst in [duplicates, legitimate, review]:
        for p in lst:
            state_counts[p['state']] += 1
    for state, count in sorted(state_counts.items(), key=lambda x: -x[1]):
        print(f"    {state}: {count}")

    # --- Write workbook ---
    report_dir = ensure_report_dir()
    outpath = report_dir / "ALF_DiffName_Classification.xlsx"
    print(f"\nWriting review workbook: {outpath.name}")
    write_workbook(duplicates, legitimate, review, outpath)
    print(f"  Done. {outpath}")


if __name__ == '__main__':
    main()
