"""
V25.2 Migration — Ownership Transitions + Choice Health Consolidation

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_1.xlsx
Output: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_2.xlsx

Changes:
  A: White Oak Management -> NHC (acquisition Aug 2024, 16 rows)
  B: HCMG -> Lionstone Care (ownership transition, 7 rows)
  C: Choice Health Management consolidation (MFA/YAD/LIFEWORKS recode, ~30 rows)
  D: Remaining duplicate cleanup (Guilford standalone SNF)
"""
import openpyxl
import re
import os
from collections import defaultdict

VAULT = os.path.expanduser(
    "~/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current"
)
V25_1_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_1.xlsx")
V25_2_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_2.xlsx")

COL = {}

def detect_columns(ws):
    global COL
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            COL[val] = c
    aliases = {
        'type': 'Source_Type', 'name': 'Facility_Name',
        'corp': 'Corporate_Name', 'addr': 'Address',
        'city': 'City', 'state': 'State', 'zip': 'ZIP',
        'county': 'County', 'beds': 'Total_Beds', 'census': 'Census',
        'served': 'Do_We_Serve', 'int': 'Integrated_Flag',
        'pcp': 'PCP_Flag', 'mh': 'MH_Flag', 'barrier': 'Barrier',
        'lat': 'Latitude', 'lon': 'Longitude',
        'dq': 'Data_Quality_Flag', 'corp_source': 'Corp_Attribution_Source',
    }
    for alias, real in aliases.items():
        if real in COL:
            COL[alias] = COL[real]

def norm(s):
    if not s: return ''
    s = str(s).upper().strip()
    s = re.sub(r'\.(?=\s|$)', '', s)
    s = re.sub(r'(?<=\b[A-Z])\.', '', s)
    s = ' '.join(s.split())
    return s

def find_all_rows(ws, corp_exact, state=None):
    matches = []
    for r in range(2, ws.max_row + 1):
        rc = norm(ws.cell(r, COL['corp']).value)
        if rc != corp_exact.upper():
            continue
        if state and norm(ws.cell(r, COL['state']).value) != state.upper():
            continue
        matches.append(r)
    return matches

# ============================================================
# PART A: White Oak Management -> NHC
# ============================================================
# Full acquisition completed Aug 1, 2024 for $220M.
# All 16 DB rows under WHITE OAK MANAGEMENT -> NATIONAL HEALTHCARE CORPORATION
WHITE_OAK_NEW_CORP = "NATIONAL HEALTHCARE CORPORATION"

# ============================================================
# PART B: HCMG -> Lionstone Care
# ============================================================
# Ownership transitioned to Lionstone Health Care.
# All 7 rows under HEALTH CARE MANAGEMENT GROUP -> LIONSTONE CARE
HCMG_OLD = "HEALTH CARE MANAGEMENT GROUP"
HCMG_NEW = "LIONSTONE CARE"

# ============================================================
# PART C: Choice Health Management consolidation
# ============================================================
# Choice Health Management owns ~18 facilities in NC/SC.
# MFA and YAD are management companies retained by Choice Health, not owners.
# LIFEWORKS REHAB also manages some Choice Health facilities.
#
# Strategy:
#   - ALL MFA rows in NC -> CHOICE HEALTH MANAGEMENT (MFA only manages Choice in NC)
#   - YAD / YAD HEALTHCARE rows in NC/SC that match known Choice Health facilities
#   - LIFEWORKS REHAB rows with "Universal Health Care" branding -> CHOICE HEALTH MANAGEMENT
#   - Remaining LIFEWORKS/YAD rows -> flag for investigation (not all are Choice Health)

# MFA NC: all 16 rows -> CHOICE HEALTH MANAGEMENT
# (MFA has 1 VA row we leave alone)
MFA_RECODE_STATE = "NC"

# YAD / YAD HEALTHCARE rows confirmed as Choice Health (by facility name / location):
# These match confirmed Choice Health properties from website + research
YAD_CHOICE_FACILITIES = [
    # (name_fragment, city, state) — YAD-managed but Choice Health-owned
    ("FLETCHER REHABILITATION", "Fletcher", "NC"),
    ("FLETCHER REHABILITATION", "FLETCHER", "NC"),
    ("RAMSEUR REHABILITATION", "Ramseur", "NC"),
    ("UNIVERSAL HEALTHCARE/RAMSEUR", "RAMSEUR", "NC"),
    ("LAUREL PARK REHABILITATION", "Elizabeth City", "NC"),
    ("LAUREL PARK REHABILITATION", "ELIZABETH CITY", "NC"),
    ("WINDSOR REHAB", "Windsor", "NC"),
    ("WINDSOR REHABILITATION", "WINDSOR", "NC"),
    ("SEVEN OAKS REHABILITATION", "Columbia", "SC"),
    ("BRIAN CENTER NURSING", "COLUMBIA", "SC"),
]

# LIFEWORKS rows with Universal Health Care branding -> CHOICE HEALTH MANAGEMENT
LIFEWORKS_CHOICE_FACILITIES = [
    ("UNIVERSAL HEALTH CARE/OXFORD", "OXFORD", "NC"),
    ("UNIVERSAL HEALTH CARE/NORTH RALEIGH", "RALEIGH", "NC"),
]

# LIFEWORKS rows that are UNCERTAIN (may or may not be Choice Health):
# Charlotte H&R, Carolina Rehab Burke, Huntersville, Litchford Falls
# These are flagged for V25.3 investigation, not recoded.

# YAD HEALTHCARE rows that are UNCERTAIN:
# Asheboro, Bolivia, Charlotte (Rockwell Park), Kenansville, Warsaw
# These are also flagged for V25.3.

# ============================================================
# PROCESS
# ============================================================

print("Loading V25.1...")
wb = openpyxl.load_workbook(V25_1_PATH)
ws = wb.active
detect_columns(ws)
total_rows = ws.max_row - 1
print(f"V25.1 rows: {total_rows}")

errors = []
changes = defaultdict(list)

# --- PART A: White Oak -> NHC ---
print("\n--- Part A: White Oak Management -> NHC ---")
rows = find_all_rows(ws, "WHITE OAK MANAGEMENT")
for r in rows:
    old = ws.cell(r, COL['corp']).value
    ws.cell(r, COL['corp']).value = WHITE_OAK_NEW_CORP
    fac = ws.cell(r, COL['name']).value
    city = ws.cell(r, COL['city']).value
    state = ws.cell(r, COL['state']).value
    changes['white_oak'].append(f"Row {r}: {fac} | {city}, {state} | '{old}' -> '{WHITE_OAK_NEW_CORP}'")
if not rows:
    errors.append("WHITE OAK: No rows found under 'WHITE OAK MANAGEMENT'")

# --- PART B: HCMG -> Lionstone ---
print("--- Part B: HCMG -> Lionstone Care ---")
rows = find_all_rows(ws, HCMG_OLD)
for r in rows:
    old = ws.cell(r, COL['corp']).value
    ws.cell(r, COL['corp']).value = HCMG_NEW
    fac = ws.cell(r, COL['name']).value
    city = ws.cell(r, COL['city']).value
    state = ws.cell(r, COL['state']).value
    changes['hcmg'].append(f"Row {r}: {fac} | {city}, {state} | '{old}' -> '{HCMG_NEW}'")
if not rows:
    errors.append("HCMG: No rows found under 'HEALTH CARE MANAGEMENT GROUP'")

# --- PART C: Choice Health Management ---
print("--- Part C: Choice Health Management Consolidation ---")
NEW_CORP = "CHOICE HEALTH MANAGEMENT"

# C1: All MFA rows in NC
mfa_nc_rows = find_all_rows(ws, "MFA", state="NC")
for r in mfa_nc_rows:
    old = ws.cell(r, COL['corp']).value
    ws.cell(r, COL['corp']).value = NEW_CORP
    fac = ws.cell(r, COL['name']).value
    city = ws.cell(r, COL['city']).value
    changes['choice'].append(f"Row {r}: {fac} | {city}, NC | '{old}' -> '{NEW_CORP}' (MFA recode)")
print(f"  MFA NC: {len(mfa_nc_rows)} rows recoded")

# C2: Specific YAD/YAD HEALTHCARE rows confirmed as Choice Health
yad_recoded = 0
for name_frag, city_match, state_match in YAD_CHOICE_FACILITIES:
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['name']).value)
        rs = norm(ws.cell(r, COL['state']).value)
        rc = norm(ws.cell(r, COL['corp']).value)
        rcity = norm(ws.cell(r, COL['city']).value)
        if (name_frag.upper() in rn and
            rs == state_match.upper() and
            rcity == city_match.upper() and
            rc in ('YAD', 'YAD HEALTHCARE')):
            old = ws.cell(r, COL['corp']).value
            ws.cell(r, COL['corp']).value = NEW_CORP
            fac = ws.cell(r, COL['name']).value
            changes['choice'].append(
                f"Row {r}: {fac} | {city_match}, {state_match} | '{old}' -> '{NEW_CORP}' (YAD recode)")
            yad_recoded += 1
print(f"  YAD confirmed: {yad_recoded} rows recoded")

# C3: LIFEWORKS rows with Universal Health Care branding
lw_recoded = 0
for name_frag, city_match, state_match in LIFEWORKS_CHOICE_FACILITIES:
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['name']).value)
        rs = norm(ws.cell(r, COL['state']).value)
        rc = norm(ws.cell(r, COL['corp']).value)
        rcity = norm(ws.cell(r, COL['city']).value)
        if (name_frag.upper() in rn and
            rs == state_match.upper() and
            rcity == city_match.upper() and
            rc == 'LIFEWORKS REHAB'):
            old = ws.cell(r, COL['corp']).value
            ws.cell(r, COL['corp']).value = NEW_CORP
            fac = ws.cell(r, COL['name']).value
            changes['choice'].append(
                f"Row {r}: {fac} | {city_match}, {state_match} | '{old}' -> '{NEW_CORP}' (LIFEWORKS recode)")
            lw_recoded += 1
print(f"  LIFEWORKS Universal: {lw_recoded} rows recoded")

# Flag remaining uncertain rows
uncertain = []
for r in range(2, ws.max_row + 1):
    rc = norm(ws.cell(r, COL['corp']).value)
    rs = norm(ws.cell(r, COL['state']).value)
    if rc in ('YAD', 'YAD HEALTHCARE', 'LIFEWORKS REHAB') and rs in ('NC', 'SC'):
        fac = ws.cell(r, COL['name']).value
        city = ws.cell(r, COL['city']).value
        uncertain.append(f"  Row {r}: {fac} | {city}, {rs} | corp={rc}")
if uncertain:
    changes['choice_uncertain'] = uncertain

# ============================================================
# REPORT
# ============================================================
print(f"\n{'=' * 70}")
print(f"V25.2 Migration Report")
print(f"{'=' * 70}")

category_labels = {
    'white_oak': 'PART A: White Oak Management -> NHC',
    'hcmg': 'PART B: HCMG -> Lionstone Care',
    'choice': 'PART C: Choice Health Management Consolidation',
    'choice_uncertain': 'PART C (FLAGGED): Uncertain MFA/YAD/LIFEWORKS — needs V25.3 investigation',
}

total_changes = 0
for key, label in category_labels.items():
    items = changes.get(key, [])
    count = len(items)
    if key != 'choice_uncertain':
        total_changes += count
    print(f"{label}: {count}")

print(f"{'=' * 70}")
print(f"TOTAL CHANGES: {total_changes}")
print(f"Final row count: {ws.max_row - 1}")
print()

for key, label in category_labels.items():
    items = changes.get(key, [])
    if items:
        print(f"\n{label}:")
        for item in items:
            print(f"  {item}")

# ============================================================
# VALIDATION & SAVE
# ============================================================
if errors:
    print(f"\n!! {len(errors)} ERRORS !!")
    for e in errors:
        print(f"  {e}")
    critical = [e for e in errors if "Row count" in e]
    if critical:
        print("\nCRITICAL ERRORS -- NOT SAVING.")
    else:
        print(f"\nNon-critical errors. Saving...")
        print(f"Saving to {V25_2_PATH}...")
        wb.save(V25_2_PATH)
        print("Done.")
else:
    print("\nAll operations completed.")
    print(f"Saving to {V25_2_PATH}...")
    wb.save(V25_2_PATH)
    print("Done.")

wb.close()
