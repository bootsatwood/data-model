"""
Load tier and scoring data into bd.market_intel_corporate_entities.

Sources:
  - T1/T2/T3 scored entities from 2026-03-07_Final_MUO_Tiering_V23.xlsx
  - T4 independents (bd_tier='T4', no scores)
  - T5 hard barriers (bd_tier='T5', barrier_type set)
  - V25 additions (Heritage Hall, RUI, Triple Crown, Vitality)

Matching strategy:
  Uses the ENTITY_UNIVERSE db_names mapping from build_final_muo_tiering.py
  to match display names to PostgreSQL corporate_name values (case-insensitive).

Usage:
  python scripts/load_scoring_to_pg.py [--dry-run]
"""

import os
import sys
import psycopg2
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
VAULT = os.path.expanduser("~/OneDrive - Eventus WholeHealth/Vault")
TIERING_PATH = os.path.join(
    VAULT, "03_Corporate_Accounts", "Tiering",
    "2026-03-07_Final_MUO_Tiering_V23.xlsx",
)

# PostgreSQL connection
PG_HOST = "keystone-platform-postgres.postgres.database.azure.com"
PG_PORT = 5432
PG_DATABASE = "postgres"
PG_USER = "ratwood"
PG_PASSWORD = "Ch0EeU3yz7ep5bEWTKT1UoC^NMPCZXzN"


def get_connection():
    return psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        database=PG_DATABASE,
        user=PG_USER,
        password=PG_PASSWORD,
        sslmode="require",
    )


# ---------------------------------------------------------------------------
# ENTITY_UNIVERSE: display_name -> set of DB-level corporate_name values
# (mirrors build_final_muo_tiering.py exactly)
# ---------------------------------------------------------------------------
ENTITY_UNIVERSE = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'American Senior Communities': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'Saber Healthcare Group': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'Infinity Healthcare Consulting': {'INFINITY HEALTHCARE CONSULTING'},
    'Navion': {'NAVION', 'NAVION SENIOR SOLUTIONS', 'NAVION SENIOR LIVING', 'Navion Senior Solutions'},
    'Majestic Care': {'MAJESTIC CARE', 'MAJESTIC / MAJESTIC CARE', 'MAJESTIC / BLUEGRASS CONSULTING GROUP'},
    'Pruitt Health': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Trilogy': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Topaz Healthcare': {'TOPAZ HEALTHCARE'},
    'Morning Pointe Senior Living': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'Principle': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TerraBella Senior Living': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'Sanstone': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'Lionstone Care': {'LIONSTONE CARE', 'LIONSTONE'},
    'Eldercare Partners': {'ELDERCARE PARTNERS'},
    'Otterbein Senior Life': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'Peak Resources': {'PEAK RESOURCES'},
    'Arbors': {'ARBORS', 'ARBORS AT OHIO'},
    'Heritage Hall': {'HERITAGE HALL', 'AMERICAN HEALTHCARE, LLC', 'AHC'},
    'CCH Healthcare': {'CCH HEALTHCARE'},
    'Priority': {'PRIORITY LIFE CARE'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'Caring Place Healthcare': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'Fundamental LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Momentus Health': {'MOMENTUS HEALTH'},
    'Altercare': {'ALTERCARE'},
    'AOM Healthcare': {'AOM HEALTHCARE'},
    'Aperion Care': {'APERION CARE'},
    'Brickyard Healthcare': {'BRICKYARD HEALTHCARE'},
    'Carecore Health': {'CARECORE HEALTH'},
    'Carespring': {'CARESPRING'},
    'Carrolton': {'CARROLTON FACILTY MANAGEMENT', 'CARROLTON NURSING HOMES'},
    'Certus Healthcare': {'CERTUS HEALTHCARE'},
    'Choice Health Management': {'CHOICE HEALTH MANAGEMENT'},
    'Ciena Healthcare/Laurel Health Care': {'CIENA HEALTHCARE/LAUREL HEALTH CARE'},
    'Continuing Healthcare Solutions': {'CONTINUING HEALTHCARE SOLUTIONS'},
    'Crown Healthcare Group': {'CROWN HEALTHCARE GROUP'},
    'Divine Healthcare Management': {'DIVINE HEALTHCARE MANAGEMENT'},
    'Envive Healthcare': {'ENVIVE HEALTHCARE'},
    'Gardant Management Solutions': {'GARDANT MANAGEMENT SOLUTIONS, INC', 'GARDANT MANAGEMENT SOLUTIONS', 'GARDANT'},
    'HCF Management': {'HCF MANAGEMENT'},
    'Health Care Management Group': {'HEALTH CARE MANAGEMENT GROUP'},
    'Hillstone Healthcare': {'HILLSTONE HEALTHCARE'},
    'Legacy Health Services': {'LEGACY HEALTH SERVICES'},
    "Miller's Merry Manor": {'MILLERS MERRY MANOR', 'MILLERS MERRY MANOR INDIANAPOLIS EAST LLC'},
    'National Healthcare Corporation': {'NATIONAL HEALTHCARE CORPORATION'},
    'Ohio Living Communities': {'OHIO LIVING COMMUNITIES'},
    'Optalis Health & Rehabilitation': {'OPTALIS HEALTH & REHABILITATION'},
    'PACS Group': {'PACS GROUP'},
    'Phoenix Senior Living': {'PHOENIX SENIOR LIVING'},
    'Progressive Quality Care': {'PROGRESSIVE QUALITY CARE'},
    'SEKY Holding Co.': {'SEKY HOLDING CO.'},
    'Sprenger Health Care Systems': {'SPRENGER HEALTH CARE SYSTEMS'},
    'Sunrise Senior Living': {'SUNRISE SENIOR LIVING', 'SUNRISE'},
    'Trio Healthcare': {'TRIO HEALTHCARE'},
    'Vancrest Health Care Centers': {'VANCREST HEALTH CARE CENTERS'},
    'White Oak Management': {'WHITE OAK MANAGEMENT'},
    'Windsor House': {'WINDSOR HOUSE, INC.'},
    'Retirement Unlimited (RUI)': {'RETIREMENT UNLIMITED, INC.', 'RETIREMENT UNLIMITED, INC', 'RUI'},
    'Vitality Living': {'VITALITY LIVING'},
    'Triple Crown Senior Living': {'TRIPLE CROWN SENIOR LIVING', 'VITALITY SENIOR SERVICES'},
    # --- Reconciliation additions (2026-04-07) ---
    'Sonida Senior Living': {'SONIDA SENIOR LIVING', 'SONIDA'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS'},
    'StoryPoint': {'STORYPOINT', 'STORYPOINT SENIOR LIVING'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING', 'KISCO'},
    'Senior Lifestyle': {'SENIOR LIFESTYLE', 'SENIOR LIFESTYLE CORPORATION'},
    'Harmony Senior Services': {'HARMONY SENIOR SERVICES'},
    'Spring Arbor Management': {'SPRING ARBOR MANAGEMENT', 'SPRING ARBOR'},
    'Greencroft': {'GREENCROFT', 'GREENCROFT COMMUNITIES'},
    'Cedarhurst Senior Living': {'CEDARHURST SENIOR LIVING', 'CEDARHURST'},
    'MCAP': {'MCAP', 'MCAP HEALTHCARE'},
    'Mainstay Senior Living': {'MAINSTAY SENIOR LIVING', 'MAINSTAY'},
    'Traditions': {'TRADITIONS', 'TRADITION SENIOR LIVING'},
    'Commonwealth Senior Living': {'COMMONWEALTH SENIOR LIVING'},
    'Cogir Senior Living': {'COGIR USA', 'COGIR SENIOR LIVING', 'COGIR'},
    'National Church Residences': {'NATIONAL CHURCH RESIDENCES'},
    'Carlyle Senior Care': {'CARLYLE SENIOR CARE', 'CARLYLE'},
}

# T5 barriers: display_name -> (set of DB names, barrier_type)
T5_BARRIERS = {
    'Bluegrass/Encore': ({'BLUEGRASS/ENCORE', 'BLUEGRASS', 'ENCORE SENIOR LIVING'}, 'Alliance'),
    'SIGNATURE HEALTH': ({'SIGNATURE HEALTHCARE', 'SIGNATURE HEALTH'}, 'Own provider group'),
    'MFA': ({'MFA', 'MFA MARYVILLE RE LLC'}, 'Alliance (VA)'),
    'COMMUNICARE': ({'COMMUNICARE'}, 'Own provider group'),
    'Hill Valley': ({'HILL VALLEY'}, 'Own provider group'),
    'SINGH': ({'SINGH'}, 'Own provider group'),
    'Embassy': ({'EMBASSY', 'EMBASSY HEALTHCARE'}, 'Own provider group'),
    'Exceptional Living': ({'EXCEPTIONAL LIVING CENTERS'}, 'Own provider group'),
    'Portopiccolo': ({'PORTOPICCOLO', 'ACCORDIUS HEALTH LLC/PORTOPICCOLO GROUP'}, 'Alliance (Telos)'),
    'Venza': ({'VENZA CARE MANAGEMENT'}, 'Alliance'),
    'Aventura': ({'AVENTURA'}, 'Recent LOB'),
    'Journey': ({'JOURNEY', 'JOURNEY HEALTHCARE'}, 'Explicit rejection'),
    'CARDON & ASSOCIATES': ({'CARDON & ASSOCIATES'}, 'Own provider group (91%)'),
    'Eastern Healthcare Group': ({'EASTERN HEALTHCARE GROUP'}, 'Termination Risk (100%)'),
    'CLEARVIEW': ({'CLEARVIEW'}, 'Alliance + Own Provider Group (38%)'),
    'Pavilion Healthcare': ({'PAVILION HEALTHCARE'}, 'Alliance + Own Provider Group (29%)'),
    # --- Reconciliation additions (2026-04-07) ---
    'Genesis Healthcare': ({'GENESIS'}, 'Own Provider Group (AlignMed Partners)'),
    'Alliance Health Group': ({'ALLIANCE HEALTH GROUP', 'ALLIANCE'}, 'Own Provider Group'),
}

# V25 manual overrides: entities whose scores come from the MUO Candidate
# Evaluation rather than the workbook. These take precedence over workbook rows
# if present. (Heritage Hall is scored differently in V25 vs V23 workbook.)
V25_OVERRIDES = {
    'Heritage Hall': {
        'bd_tier': 'T1', 'total_score': 58,
        'enterprise_reach': 3, 'integration_ready': 3,
        'strategic_influence': 2, 'revenue_potential_score': 4,
        'relationship_strength': 5, 'ai_tech_adoption': 0,
    },
    'Retirement Unlimited (RUI)': {
        'bd_tier': 'T2', 'total_score': 40,
        'enterprise_reach': 2, 'integration_ready': 2,
        'strategic_influence': 1, 'revenue_potential_score': 5,
        'relationship_strength': 1, 'ai_tech_adoption': 0,
    },
    'Triple Crown Senior Living': {
        'bd_tier': 'T2', 'total_score': 40,
        'enterprise_reach': 2, 'integration_ready': 2,
        'strategic_influence': 1, 'revenue_potential_score': 5,
        'relationship_strength': 1, 'ai_tech_adoption': 0,
    },
    'Vitality Living': {
        'bd_tier': 'T3', 'total_score': 29,
        'enterprise_reach': 1, 'integration_ready': 1,
        'strategic_influence': 1, 'revenue_potential_score': 4,
        'relationship_strength': 1, 'ai_tech_adoption': 0,
    },
    # --- Reconciliation additions (2026-04-07) ---
    # V23 computed scores (from Finance workbook / HTML universe build)
    'Sonida Senior Living': {
        'bd_tier': 'T1', 'total_score': 59,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Lutheran Services Carolinas': {
        'bd_tier': 'T1', 'total_score': 68,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'StoryPoint': {
        'bd_tier': 'T1', 'total_score': 60,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Kisco Senior Living': {
        'bd_tier': 'T2', 'total_score': 54,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Senior Lifestyle': {
        'bd_tier': 'T2', 'total_score': 49,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Spring Arbor Management': {
        'bd_tier': 'T2', 'total_score': 41,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Greencroft': {
        'bd_tier': 'T2', 'total_score': 41,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Cedarhurst Senior Living': {
        'bd_tier': 'T2', 'total_score': 37,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'MCAP': {
        'bd_tier': 'T3', 'total_score': 32,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    # Brooke-assigned tiers (3/27 call) — no computed scores yet
    'Harmony Senior Services': {
        'bd_tier': 'T2', 'total_score': None,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Mainstay Senior Living': {
        'bd_tier': 'T3', 'total_score': None,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
    'Traditions': {
        'bd_tier': 'T3', 'total_score': None,
        'enterprise_reach': None, 'integration_ready': None,
        'strategic_influence': None, 'revenue_potential_score': None,
        'relationship_strength': None, 'ai_tech_adoption': None,
    },
}


# ---------------------------------------------------------------------------
# Step 1: Read scored entities from the V23 tiering workbook
# ---------------------------------------------------------------------------
def read_tiering_workbook():
    """Read T1/T2/T3 tabs and return {display_name: score_dict}."""
    wb = openpyxl.load_workbook(TIERING_PATH, read_only=True, data_only=True)
    scored = {}

    for sheet_name in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name:
                continue
            name = str(name).strip()
            scored[name] = {
                'bd_tier': str(row[1]).strip(),
                'total_score': int(row[2]) if row[2] is not None else None,
                'enterprise_reach': int(row[8]) if row[8] is not None else None,
                'integration_ready': int(row[9]) if row[9] is not None else None,
                'strategic_influence': int(row[10]) if row[10] is not None else None,
                'revenue_potential_score': int(row[11]) if row[11] is not None else None,
                'relationship_strength': int(row[12]) if row[12] is not None else None,
                'ai_tech_adoption': int(row[13]) if row[13] is not None else None,
            }

    # T4 independents: tier only, no dimension scores
    ws4 = wb['T4 - Independents']
    t4_names = []
    for row in ws4.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        t4_names.append(name)
        scored[name] = {
            'bd_tier': 'T4',
            'total_score': None,
            'enterprise_reach': None,
            'integration_ready': None,
            'strategic_influence': None,
            'revenue_potential_score': None,
            'relationship_strength': None,
            'ai_tech_adoption': None,
        }

    wb.close()
    print(f"  Read {len(scored)} entities from workbook")
    print(f"    T1: {sum(1 for v in scored.values() if v['bd_tier']=='T1')}")
    print(f"    T2: {sum(1 for v in scored.values() if v['bd_tier']=='T2')}")
    print(f"    T3: {sum(1 for v in scored.values() if v['bd_tier']=='T3')}")
    print(f"    T4: {len(t4_names)}")
    return scored


# ---------------------------------------------------------------------------
# Step 2: Apply V25 overrides
# ---------------------------------------------------------------------------
def apply_v25_overrides(scored):
    """Override or add V25 candidate evaluation scores."""
    for display_name, override in V25_OVERRIDES.items():
        if display_name in scored:
            print(f"  V25 override (replacing workbook): {display_name}")
        else:
            print(f"  V25 addition (new entity): {display_name}")
        scored[display_name] = override
    return scored


# ---------------------------------------------------------------------------
# Step 3: Build the update plan
# ---------------------------------------------------------------------------
def build_update_plan(scored, pg_names):
    """
    Match display_name -> PG corporate_name(s) using ENTITY_UNIVERSE db_names.

    pg_names: dict of {UPPER_NAME: (id, original_name)} from PostgreSQL.

    Returns a list of (pg_id, pg_name, display_name, score_dict) tuples,
    plus a list of unmatched display_names.
    """
    updates = []
    unmatched = []

    # Scored entities (T1-T4)
    for display_name, score_dict in scored.items():
        db_names = ENTITY_UNIVERSE.get(display_name, set())
        matched = False
        for dn in db_names:
            key = dn.upper()
            if key in pg_names:
                pg_id, pg_original = pg_names[key]
                updates.append((pg_id, pg_original, display_name, score_dict))
                matched = True
        if not matched:
            unmatched.append(('SCORED', display_name))

    # T5 barriers
    for display_name, (db_names_set, barrier_type) in T5_BARRIERS.items():
        matched = False
        for dn in db_names_set:
            key = dn.upper()
            if key in pg_names:
                pg_id, pg_original = pg_names[key]
                t5_dict = {
                    'bd_tier': 'T5',
                    'total_score': None,
                    'enterprise_reach': None,
                    'integration_ready': None,
                    'strategic_influence': None,
                    'revenue_potential_score': None,
                    'relationship_strength': None,
                    'ai_tech_adoption': None,
                    'barrier_type': barrier_type,
                }
                updates.append((pg_id, pg_original, display_name, t5_dict))
                matched = True
        if not matched:
            unmatched.append(('T5', display_name))

    return updates, unmatched


# ---------------------------------------------------------------------------
# Step 4: Execute updates
# ---------------------------------------------------------------------------
def execute_updates(cursor, updates, dry_run=False):
    """Run parameterized UPDATE statements."""
    sql = """
        UPDATE bd.market_intel_corporate_entities
        SET bd_tier = %s,
            total_score = %s,
            enterprise_reach = %s,
            integration_ready = %s,
            strategic_influence = %s,
            revenue_potential_score = %s,
            relationship_strength = %s,
            ai_tech_adoption = %s,
            barrier_type = %s
        WHERE id = %s
    """
    count = 0
    for pg_id, pg_name, display_name, score_dict in updates:
        params = (
            score_dict.get('bd_tier'),
            score_dict.get('total_score'),
            score_dict.get('enterprise_reach'),
            score_dict.get('integration_ready'),
            score_dict.get('strategic_influence'),
            score_dict.get('revenue_potential_score'),
            score_dict.get('relationship_strength'),
            score_dict.get('ai_tech_adoption'),
            score_dict.get('barrier_type'),
            pg_id,
        )
        if not dry_run:
            cursor.execute(sql, params)
        count += 1
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 70)
    print(f"Load Scoring Data to PostgreSQL {'(DRY RUN)' if dry_run else ''}")
    print("=" * 70)

    # Verify workbook exists
    if not os.path.exists(TIERING_PATH):
        print(f"ERROR: Tiering workbook not found: {TIERING_PATH}")
        sys.exit(1)
    print(f"  Source: {TIERING_PATH}")

    # Step 1: Read workbook
    print()
    print("Step 1: Reading V23 tiering workbook...")
    scored = read_tiering_workbook()

    # Step 2: Apply V25 overrides
    print()
    print("Step 2: Applying V25 overrides...")
    scored = apply_v25_overrides(scored)
    print(f"  Total scored entities: {len(scored)}")

    # Step 3: Connect to PostgreSQL and get current names
    print()
    print("Step 3: Connecting to PostgreSQL...")
    conn = get_connection()
    cursor = conn.cursor()
    print("  Connected.")

    cursor.execute("SELECT id, corporate_name FROM bd.market_intel_corporate_entities")
    rows = cursor.fetchall()
    pg_names = {}
    for row_id, name in rows:
        pg_names[name.upper()] = (row_id, name)
    print(f"  Loaded {len(pg_names)} corporate entities from PostgreSQL")

    # Step 4: Build update plan
    print()
    print("Step 4: Building update plan...")
    updates, unmatched = build_update_plan(scored, pg_names)

    # Deduplicate by pg_id (one entity can match multiple db_names to the same PG row)
    seen_ids = set()
    deduped = []
    for item in updates:
        pg_id = item[0]
        if pg_id not in seen_ids:
            seen_ids.add(pg_id)
            deduped.append(item)
    updates = deduped

    print(f"  Updates to execute: {len(updates)}")
    if unmatched:
        print(f"  WARNING: {len(unmatched)} entities could not be matched to PostgreSQL:")
        for kind, name in unmatched:
            print(f"    [{kind}] {name}")

    # Tier breakdown
    tier_counts = {}
    for _, _, _, sd in updates:
        t = sd.get('bd_tier', '?')
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"  Tier breakdown of updates:")
    for t in sorted(tier_counts):
        print(f"    {t}: {tier_counts[t]}")

    if dry_run:
        print()
        print("DRY RUN complete. No changes made.")
        print()
        print("Planned updates:")
        for pg_id, pg_name, display_name, sd in sorted(updates, key=lambda x: x[3].get('bd_tier', 'Z')):
            score_str = f"score={sd.get('total_score')}" if sd.get('total_score') else "no score"
            barrier_str = f", barrier={sd.get('barrier_type')}" if sd.get('barrier_type') else ""
            print(f"  {sd.get('bd_tier'):>3} | {pg_name:<55} <- {display_name} ({score_str}{barrier_str})")
        cursor.close()
        conn.close()
        return

    # Step 5: Execute updates
    print()
    print("Step 5: Executing updates...")
    count = execute_updates(cursor, updates, dry_run=False)
    print(f"  Executed {count} UPDATE statements")

    # Step 6: Commit
    print()
    print("Step 6: Committing transaction...")
    conn.commit()
    print("  Committed.")

    # Step 7: Verification
    print()
    print("Step 7: Verification...")
    cursor.execute("SELECT bd_tier, COUNT(*) FROM bd.market_intel_corporate_entities WHERE bd_tier IS NOT NULL GROUP BY bd_tier ORDER BY bd_tier")
    tier_rows = cursor.fetchall()
    print("  Rows by tier:")
    total_set = 0
    for tier, cnt in tier_rows:
        print(f"    {tier}: {cnt}")
        total_set += cnt
    print(f"  Total entities with bd_tier set: {total_set}")

    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_corporate_entities WHERE total_score IS NOT NULL")
    score_count = cursor.fetchone()[0]
    print(f"  Total entities with total_score set: {score_count}")

    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_corporate_entities WHERE barrier_type IS NOT NULL")
    barrier_count = cursor.fetchone()[0]
    print(f"  Total entities with barrier_type set: {barrier_count}")

    # Show all updated rows
    print()
    print("  Updated entities:")
    cursor.execute("""
        SELECT bd_tier, corporate_name, total_score, enterprise_reach,
               integration_ready, strategic_influence, revenue_potential_score,
               relationship_strength, ai_tech_adoption, barrier_type
        FROM bd.market_intel_corporate_entities
        WHERE bd_tier IS NOT NULL
        ORDER BY bd_tier, total_score DESC NULLS LAST, corporate_name
    """)
    for row in cursor.fetchall():
        tier, name, score, er, ir, si, rp, rs, ai, barrier = row
        dims = f"ER={er} IR={ir} SI={si} RP={rp} RS={rs} AI={ai}" if score else ""
        barrier_str = f" | barrier={barrier}" if barrier else ""
        score_str = f"score={score}" if score else "no score"
        print(f"    {tier} | {name:<55} {score_str:>10} {dims}{barrier_str}")

    cursor.close()
    conn.close()
    print()
    print("Done.")


if __name__ == "__main__":
    main()
