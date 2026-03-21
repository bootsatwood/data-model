"""
V24.2 Migration -- MUO Candidate Evaluation corrections.
Content-based matching with verification for safety.

Source: MUO Candidate Evaluation (board review, 2026-03-20/21)
Audit reports: data-model/scripts/audit_reports/
Corporate history: data-model/reference/MUO_Corporate_History.md

Changes:
  - Heritage Hall reattribution (American Healthcare/AHC -> Heritage Hall)
  - Heritage Senior Living split (4 separate operators)
  - Pinnacle corrections (WI -> Pennant Group, IL -> Vantage, VA normalize)
  - Navion entity consolidation + Goldsboro misattribution fix
  - Gardant misattributions + entity merge
  - RUI misattributions + phantom delete + duplicate deletes + entity merge
  - Vitality/Triple Crown corrections
  - Facility name, county, bed, city fixes
"""
import openpyxl
import re

V24_1 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_1.xlsx"
V24_2 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_2.xlsx"

def norm(s):
    if not s: return ''
    s = str(s).upper().strip()
    s = re.sub(r'\.(?=\s|$)', '', s)
    s = re.sub(r'(?<=\b[A-Z])\.', '', s)
    s = ' '.join(s.split())
    return s

# Column indices (1-based for openpyxl)
COL = {
    'type': 1, 'name': 2, 'corp': 3, 'addr': 4, 'city': 5,
    'state': 6, 'zip': 7, 'county': 8, 'beds': 10, 'census': 11,
    'served': 12, 'int': 13, 'pcp': 14, 'mh': 15, 'dq': 19,
    'corp_source': 26
}

# ============================================================
# HELPER: Find row by content match
# ============================================================
def find_row(ws, name_frag, state, city=None, corp=None):
    """Find a row matching name fragment + state + optional city/corp. Returns row index or None."""
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        row_name = norm(ws.cell(row_idx, COL['name']).value)
        row_state = norm(ws.cell(row_idx, COL['state']).value)
        if name_frag.upper() not in row_name:
            continue
        if state.upper() != row_state:
            continue
        if city and city.upper() != norm(ws.cell(row_idx, COL['city']).value):
            continue
        if corp and corp.upper() != norm(ws.cell(row_idx, COL['corp']).value):
            continue
        matches.append(row_idx)
    if len(matches) == 1:
        return matches[0]
    elif len(matches) == 0:
        return None
    else:
        return matches  # multiple matches -- caller must handle


def find_all_rows(ws, corp_exact, state=None):
    """Find all rows with exact corporate name match, optional state filter."""
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        row_corp = norm(ws.cell(row_idx, COL['corp']).value)
        if row_corp != corp_exact.upper():
            continue
        if state and norm(ws.cell(row_idx, COL['state']).value) != state.upper():
            continue
        matches.append(row_idx)
    return matches

# ============================================================
# CHANGE DEFINITIONS
# ============================================================

# --- PART A: Heritage Hall Reattribution ---
# 11 rows: AMERICAN HEALTHCARE, LLC -> HERITAGE HALL
# 2 rows: AHC -> HERITAGE HALL
HERITAGE_HALL_REATTRIB = [
    # (name_fragment, city, state, old_corp)
    ("814 HERITAGE HALL GRUNDY", "Grundy", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("818 HERITAGE HALL LEXINGTON", "Lexington", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("819 HERITAGE HALL LAUREL MEADOWS", "Laurel Fork", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("931 HERITAGE HALL BLACKSBURG", "Blacksburg", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("933 HERITAGE HALL TAZEWELL", "Tazewell", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL - BIG STONE GAP", "Big Stone Gap", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL BIG STONE GAP", "BIG STONE GAP", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL BLACKSBURG", "BLACKSBURG", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL GRUNDY", "GRUNDY", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL - LAUREL MEADOWS", "LAUREL FORK", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("HERITAGE HALL TAZEWELL", "TAZEWELL", "VA", "AMERICAN HEALTHCARE, LLC"),
    ("822 SOUTH ROANOKE NURSING HOME", "Roanoke", "VA", "AHC"),
    ("SOUTH ROANOKE NURSING AND REHABILITATION", "ROANOKE", "VA", "AHC"),
]

# --- PART B: Heritage Senior Living Split ---
# PA rows -> Heritage Senior Living LLC
HSL_PA_NAMES = [
    ("TRADITIONS OF HANOVER", "Bethlehem"),
    ("KEYSTONE VILLA AT FLEETWOOD", "Blandon"),
    ("CHESTNUT KNOLL", "Boyertown"),
    ("KEYSTONE VILLA AT DOUGLASSVILLE", "Douglassville"),
    ("THE BIRCHES OF THE LEHIGH VALLEY", "Easton"),
    ("THE BIRCHES AT HARLEYSVILLE", "Harleysville"),
    ("TRADITIONS OF LANSDALE", "Lansdale"),
    ("BELLE REVE", "Milford"),
    ("THE BIRCHES AT NEW HOPE", "New Hope"),
    ("BIRCHES OF NEWTOWN", "Newtown"),
    ("TRADITIONS OF HERSHEY", "Palmyra"),
    ("THE MANOR AT MARKET SQUARE", "Reading"),
    ("HERITAGE HILL SENIOR COMMUNITY", "Weatherly"),
    ("SENIOR COMMONS AT POWDER MILL", "York"),
]

# VA rows -> Heritage Senior Living LLC
HSL_VA_NAMES = [
    ("CHANCELLOR'S VILLAGE", "Fredericksburg"),
    ("HERITAGE GREEN", "Lynchburg"),
    ("THE CROSSROADS AT BON AIR", "Richmond"),
]

# --- PART C: Pinnacle Corrections ---
PINNACLE_TO_PENNANT = [
    # WI facilities -> The Pennant Group: (name_frag, city, state, old_corp)
    ("Kenosha Senior Living", "Kenosha", "WI", "Pinnacle Living"),
    ("North Point Senior Living", "Kenosha", "WI", "Pinnacle Living"),
    ("Madison Pointe Senior Living", "Madison", "WI", "PINNACLE"),
    ("Pleasant Point Senior Living", "Racine", "WI", "Pinnacle Living"),
    ("Meadow View Assisted Living", "Two Rivers", "WI", "Pinnacle Senior Living"),
]

PINNACLE_TO_VANTAGE = [
    ("New City Supportive Living", "Chicago", "IL"),
]

PINNACLE_VA_NORMALIZE = [
    # Roanoke rows: PINNACLE -> Pinnacle Living
    ("HERMITAGE ROANOKE - ALF", "Roanoke", "VA"),
    ("HERMITAGE ROANOKE -HC", "Roanoke", "VA"),
    ("HERMITAGE ROANOKE", "Roanoke", "VA"),
]

# --- PART D: Navion Consolidation ---
# Entity 12 "NAVION" -> "Navion Senior Solutions"
# Entity 569 "NAVION SENIOR LIVING" -> "Navion Senior Solutions"
# ID 9309 Brookdale -> "Navion Senior Solutions"

# --- PART E: Gardant Corrections ---
GARDANT_REATTRIB = [
    ("CARRIAGE COURT GROVE OF CITY", "Grove City", "OH"),
    ("CARRIAGE COURT OF WASH", "Washington Court House", "OH"),
    ("REFLECTIONS RETIREMENT", "Lancaster", "OH"),
    ("THE VILLAGE AT WESTERVILLE", "Westerville", "OH"),
    ("KEEPSAKE VILLAGE OF COLUMBUS", "Columbus", "IN"),
]

# --- PART F: RUI Corrections ---
RUI_FL_REATTRIB = [
    # (name_frag, city, state, new_corp)
    ("Watercrest Palm Beach Gardens", "Palm Beach Gardens", "FL", "Watercrest Senior Living Group"),
    ("Capital Square At Tallahassee", "Tallahassee", "FL", "YourLife Senior Living"),
    ("YOURLife of West Melbourne", "West Melbourne", "FL", "YourLife Senior Living"),
]

# --- PART G: Vitality / Triple Crown ---
VITALITY_FIXES = [
    # (name_frag, city, state, old_corp, new_corp)
    # Note: Sunrise of Five Forks deferred to V24.3 -- city needs investigation
]

VITALITY_STALE_CORP = [
    ("VITALITY LIVING-COLUMBIA", "Columbia", "SC", "Watercrest Senior Living Group", "Vitality Living"),
    ("ACCLAIM AT BELMONT BAY", "Woodbridge", "VA", "Senior Lifestyle", "Vitality Living"),
]

# --- PART H: Deletes ---
# DELETE_TARGETS now uses (name_fragment, city, state, old_corp, reason)
# old_corp is used to disambiguate duplicates -- we delete the row with the WRONG corp
DELETE_TARGETS = [
    ("HERITAGE POINT ASSISTED LIVING", "New Boston", "OH", None, "Duplicate of Heritage Square at same address"),
    ("GLASSWATER CREEK", "Lafayette", "IN", None, "Gardant duplicate"),
    ("VITALITY LIVING ELIZABETHTOWN", "Elizabethtown", "KY", None, "Vitality duplicate (smaller, surrogate)"),
    ("SUMMIT OF EDGEWOOD", "Edgewood", "KY", "TRIPLE CROWN SENIOR LIVING", "Vitality duplicate (smaller)"),
    ("WATERCREST", "Columbia", "SC", None, "Vitality/Watercrest duplicate"),
]

# These deletes need corp-based disambiguation -- handled separately
DELETE_BY_CORP = [
    # (name_frag, city, state, target_corp_to_delete, reason)
    ("WELLINGTON AT LAKE MANASSAS", "Gainesville", "VA", "INDEPENDENT", "RUI duplicate (surrogate)"),
    ("ASHLEIGH AT LANSDOWNE", "Leesburg", "VA", "INDEPENDENT", "RUI duplicate (surrogate)"),
]

# These need investigation -- skip for V24.2, defer to V24.3
# SWEET GALILEE -- city might be wrong
# THE MORGAN AT FORD'S VILLAGE -- apostrophe matching issue
# WOODLAND HILLS -- city might not be Richmond
# THE WESTMONT AT SHORT PUMP -- city might not be Richmond
# Sunrise Of Five Forks -- city might not be Simpsonville

# --- PART I: Facility Name Fixes ---
NAME_FIXES = [
    # (old_name_frag, city, state, new_name)
    ("HERITAGE SQUARE NEW BOSTON", "New Boston", "OH", "Heritage Pointe Assisted Living"),
    ("CARRIAGE COURT GROVE OF CITY", "Grove City", "OH", "CARRIAGE COURT OF GROVE CITY"),
]

# --- PART J: County Fills ---
COUNTY_FILLS = [
    # (name_frag, city, state, county)
    ("Heritage Pointe", "New Boston", "OH", "Scioto"),  # after name fix
    ("Green Oaks", "Goshen", "IN", "Elkhart"),
]

# --- PART K: Bed Count Fixes ---
BED_FIXES = [
    # (name_frag, city, state, new_beds, new_census)
    ("Heritage Pointe", "New Boston", "OH", 116, 93),  # after name fix; 93 from original Heritage Square census
]

# --- PART L: City Fixes ---
CITY_FIXES = [
    # (name_frag, old_city, state, new_city, old_corp)
    ("North Point Senior Living", "Kenosha", "WI", "Somers", "The Pennant Group"),  # corp already changed by Part C
    ("Pleasant Point Senior Living", "Racine", "WI", "Mt. Pleasant", "The Pennant Group"),  # corp already changed
]

# ============================================================
# PROCESS
# ============================================================

print("Loading V24.1...")
wb = openpyxl.load_workbook(V24_1)
ws = wb.active
total_rows = ws.max_row - 1
print(f"V24.1 rows: {total_rows}")

errors = []
changes = {
    'corp_reattrib': [],
    'corp_split': [],
    'corp_normalize': [],
    'name_fix': [],
    'county_fill': [],
    'bed_fix': [],
    'city_fix': [],
    'delete': [],
}

# --- PART A: Heritage Hall Reattribution ---
print("\n--- Part A: Heritage Hall Reattribution ---")
for name_frag, city, state, old_corp in HERITAGE_HALL_REATTRIB:
    row = find_row(ws, name_frag, state, city=city, corp=old_corp)
    if row is None:
        errors.append(f"HERITAGE HALL: NOT FOUND -- '{name_frag}' in {city}, {state} under '{old_corp}'")
    elif isinstance(row, list):
        errors.append(f"HERITAGE HALL: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}' in {city}, {state}")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = "HERITAGE HALL"
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> 'HERITAGE HALL'")

# --- PART B: Heritage Senior Living Split ---
print("--- Part B: Heritage Senior Living Split ---")

# PA -> Heritage Senior Living LLC
for name_frag, city in HSL_PA_NAMES:
    row = find_row(ws, name_frag, "PA", city=city, corp="HERITAGE SENIOR LIVING")
    if row is None:
        errors.append(f"HSL SPLIT PA: NOT FOUND -- '{name_frag}' in {city}, PA")
    elif isinstance(row, list):
        errors.append(f"HSL SPLIT PA: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}' in {city}, PA")
    else:
        ws.cell(row, COL['corp']).value = "Heritage Senior Living LLC"
        fac = ws.cell(row, COL['name']).value
        changes['corp_split'].append(f"Row {row}: {fac} | PA -> 'Heritage Senior Living LLC'")

# VA -> Heritage Senior Living LLC
for name_frag, city in HSL_VA_NAMES:
    row = find_row(ws, name_frag, "VA", city=city, corp="HERITAGE SENIOR LIVING")
    if row is None:
        errors.append(f"HSL SPLIT VA: NOT FOUND -- '{name_frag}' in {city}, VA")
    elif isinstance(row, list):
        errors.append(f"HSL SPLIT VA: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}' in {city}, VA")
    else:
        ws.cell(row, COL['corp']).value = "Heritage Senior Living LLC"
        fac = ws.cell(row, COL['name']).value
        changes['corp_split'].append(f"Row {row}: {fac} | VA -> 'Heritage Senior Living LLC'")

# OH Marysville -> Heritage Senior Living of Marysville
row = find_row(ws, "HERITAGE SENIOR LIVING OF MARYSVILLE", "OH", city="Marysville", corp="HERITAGE SENIOR LIVING")
if row is None:
    errors.append("HSL SPLIT OH: NOT FOUND -- Marysville")
elif isinstance(row, list):
    errors.append(f"HSL SPLIT OH: MULTIPLE MATCHES ({len(row)}) -- Marysville")
else:
    ws.cell(row, COL['corp']).value = "Heritage Senior Living of Marysville"
    changes['corp_split'].append(f"Row {row}: Heritage Senior Living of Marysville")

# OH New Boston -> Heritage Legacy Health Services (after name fix in Part I)
row = find_row(ws, "HERITAGE SQUARE NEW BOSTON", "OH", city="New Boston", corp="HERITAGE SENIOR LIVING")
if row is None:
    errors.append("HSL SPLIT OH: NOT FOUND -- Heritage Square New Boston")
elif isinstance(row, list):
    errors.append(f"HSL SPLIT OH: MULTIPLE MATCHES ({len(row)}) -- Heritage Square New Boston")
else:
    ws.cell(row, COL['corp']).value = "Heritage Legacy Health Services"
    changes['corp_split'].append(f"Row {row}: Heritage Square New Boston -> 'Heritage Legacy Health Services'")

# WI stays as HERITAGE SENIOR LIVING -- no change needed

# --- PART C: Pinnacle Corrections ---
print("--- Part C: Pinnacle Corrections ---")

for name_frag, city, state, old_corp in PINNACLE_TO_PENNANT:
    row = find_row(ws, name_frag, state, city=city, corp=old_corp)
    if row is None:
        errors.append(f"PINNACLE->PENNANT: NOT FOUND -- '{name_frag}' in {city}, {state} under '{old_corp}'")
    elif isinstance(row, list):
        errors.append(f"PINNACLE->PENNANT: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = "The Pennant Group"
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> 'The Pennant Group'")

for name_frag, city, state in PINNACLE_TO_VANTAGE:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"PINNACLE->VANTAGE: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"PINNACLE->VANTAGE: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = "Vantage Senior Care"
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> 'Vantage Senior Care'")

for name_frag, city, state in PINNACLE_VA_NORMALIZE:
    row = find_row(ws, name_frag, state, city=city, corp="PINNACLE")
    if row is None:
        errors.append(f"PINNACLE NORMALIZE: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"PINNACLE NORMALIZE: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        ws.cell(row, COL['corp']).value = "Pinnacle Living"
        fac = ws.cell(row, COL['name']).value
        changes['corp_normalize'].append(f"Row {row}: {fac} | 'PINNACLE' -> 'Pinnacle Living'")

# --- PART D: Navion Consolidation ---
print("--- Part D: Navion Consolidation ---")

# NAVION -> Navion Senior Solutions
navion_rows = find_all_rows(ws, "NAVION")
for row in navion_rows:
    ws.cell(row, COL['corp']).value = "Navion Senior Solutions"
    fac = ws.cell(row, COL['name']).value
    changes['corp_normalize'].append(f"Row {row}: {fac} | 'NAVION' -> 'Navion Senior Solutions'")

# NAVION SENIOR LIVING -> Navion Senior Solutions
navion_sl_rows = find_all_rows(ws, "NAVION SENIOR LIVING")
for row in navion_sl_rows:
    ws.cell(row, COL['corp']).value = "Navion Senior Solutions"
    fac = ws.cell(row, COL['name']).value
    changes['corp_normalize'].append(f"Row {row}: {fac} | 'NAVION SENIOR LIVING' -> 'Navion Senior Solutions'")

# Goldsboro misattribution: Brookdale -> Navion
row = find_row(ws, "NAVION OF GOLDSBORO", "NC", city="Goldsboro", corp="BROOKDALE SENIOR LIVING")
if row is None:
    # Try without corp filter -- might be under different Brookdale variant
    row = find_row(ws, "NAVION OF GOLDSBORO", "NC", city="Goldsboro")
if row is None:
    errors.append("NAVION GOLDSBORO: NOT FOUND")
elif isinstance(row, list):
    errors.append(f"NAVION GOLDSBORO: MULTIPLE MATCHES ({len(row)})")
else:
    old = ws.cell(row, COL['corp']).value
    ws.cell(row, COL['corp']).value = "Navion Senior Solutions"
    changes['corp_reattrib'].append(f"Row {row}: NAVION OF GOLDSBORO | '{old}' -> 'Navion Senior Solutions'")

# --- PART E: Gardant Corrections ---
print("--- Part E: Gardant Corrections ---")

for name_frag, city, state in GARDANT_REATTRIB:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"GARDANT REATTRIB: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"GARDANT REATTRIB: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = "GARDANT MANAGEMENT SOLUTIONS"
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> 'GARDANT MANAGEMENT SOLUTIONS'")

# Gardant entity merge: GARDANT -> GARDANT MANAGEMENT SOLUTIONS
gardant_rows = find_all_rows(ws, "GARDANT")
for row in gardant_rows:
    ws.cell(row, COL['corp']).value = "GARDANT MANAGEMENT SOLUTIONS"
    fac = ws.cell(row, COL['name']).value
    changes['corp_normalize'].append(f"Row {row}: {fac} | 'GARDANT' -> 'GARDANT MANAGEMENT SOLUTIONS'")

# --- PART F: RUI Corrections ---
print("--- Part F: RUI Corrections ---")

for name_frag, city, state, new_corp in RUI_FL_REATTRIB:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"RUI FL REATTRIB: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"RUI FL REATTRIB: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = new_corp
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> '{new_corp}'")

# RUI entity merge: RUI -> Retirement Unlimited, Inc.
rui_rows = find_all_rows(ws, "RUI")
for row in rui_rows:
    ws.cell(row, COL['corp']).value = "Retirement Unlimited, Inc."
    fac = ws.cell(row, COL['name']).value
    changes['corp_normalize'].append(f"Row {row}: {fac} | 'RUI' -> 'Retirement Unlimited, Inc.'")

# --- PART G: Vitality / Triple Crown ---
print("--- Part G: Vitality / Triple Crown ---")

# Sunrise misattribution
for name_frag, city, state, old_corp, new_corp in VITALITY_FIXES:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"VITALITY FIX: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"VITALITY FIX: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = new_corp
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> '{new_corp}'")

# Stale corp updates
for name_frag, city, state, old_corp, new_corp in VITALITY_STALE_CORP:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"VITALITY STALE: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"VITALITY STALE: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['corp']).value
        ws.cell(row, COL['corp']).value = new_corp
        fac = ws.cell(row, COL['name']).value
        changes['corp_reattrib'].append(f"Row {row}: {fac} | '{old}' -> '{new_corp}'")

# Triple Crown rename: VITALITY SENIOR SERVICES -> Triple Crown Senior Living
vss_rows = find_all_rows(ws, "VITALITY SENIOR SERVICES")
for row in vss_rows:
    ws.cell(row, COL['corp']).value = "Triple Crown Senior Living"
    fac = ws.cell(row, COL['name']).value
    changes['corp_normalize'].append(f"Row {row}: {fac} | 'VITALITY SENIOR SERVICES' -> 'Triple Crown Senior Living'")

# --- PART H: Deletes ---
print("--- Part H: Deletes ---")
rows_to_delete = []

# Standard deletes (unique matches)
for name_frag, city, state, old_corp, reason in DELETE_TARGETS:
    if old_corp:
        row = find_row(ws, name_frag, state, city=city, corp=old_corp)
    elif city:
        row = find_row(ws, name_frag, state, city=city)
    else:
        row = find_row(ws, name_frag, state)
    if row is None:
        errors.append(f"DELETE: NOT FOUND -- '{name_frag}' in {city or '?'}, {state} ({reason})")
    elif isinstance(row, list):
        errors.append(f"DELETE: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}' in {city or '?'}, {state}")
    else:
        fac = ws.cell(row, COL['name']).value
        beds = ws.cell(row, COL['beds']).value
        rows_to_delete.append(row)
        changes['delete'].append(f"Row {row}: {fac} | {city}, {state} | beds={beds} | {reason}")

# Corp-disambiguated deletes (delete the row with the target corp, keep the other)
for name_frag, city, state, target_corp, reason in DELETE_BY_CORP:
    row = find_row(ws, name_frag, state, city=city, corp=target_corp)
    if row is None:
        errors.append(f"DELETE BY CORP: NOT FOUND -- '{name_frag}' in {city}, {state} under '{target_corp}' ({reason})")
    elif isinstance(row, list):
        errors.append(f"DELETE BY CORP: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}' under '{target_corp}'")
    else:
        fac = ws.cell(row, COL['name']).value
        beds = ws.cell(row, COL['beds']).value
        rows_to_delete.append(row)
        changes['delete'].append(f"Row {row}: {fac} | {city}, {state} | corp='{target_corp}' | beds={beds} | {reason}")

# --- PART I: Facility Name Fixes ---
print("--- Part I: Facility Name Fixes ---")
for old_frag, city, state, new_name in NAME_FIXES:
    row = find_row(ws, old_frag, state, city=city)
    if row is None:
        errors.append(f"NAME FIX: NOT FOUND -- '{old_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"NAME FIX: MULTIPLE MATCHES ({len(row)}) -- '{old_frag}'")
    else:
        old = ws.cell(row, COL['name']).value
        ws.cell(row, COL['name']).value = new_name
        changes['name_fix'].append(f"Row {row}: '{old}' -> '{new_name}'")

# --- PART J: County Fills ---
print("--- Part J: County Fills ---")
for name_frag, city, state, county in COUNTY_FILLS:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"COUNTY FILL: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"COUNTY FILL: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old = ws.cell(row, COL['county']).value
        ws.cell(row, COL['county']).value = county
        fac = ws.cell(row, COL['name']).value
        changes['county_fill'].append(f"Row {row}: {fac} | county '{old}' -> '{county}'")

# --- PART K: Bed Count Fixes ---
print("--- Part K: Bed Count Fixes ---")
for name_frag, city, state, new_beds, new_census in BED_FIXES:
    row = find_row(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"BED FIX: NOT FOUND -- '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        errors.append(f"BED FIX: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        old_beds = ws.cell(row, COL['beds']).value
        old_census = ws.cell(row, COL['census']).value
        ws.cell(row, COL['beds']).value = new_beds
        ws.cell(row, COL['census']).value = new_census
        fac = ws.cell(row, COL['name']).value
        changes['bed_fix'].append(f"Row {row}: {fac} | beds {old_beds}->{new_beds}, census {old_census}->{new_census}")

# --- PART L: City Fixes ---
print("--- Part L: City Fixes ---")
for name_frag, old_city, state, new_city, old_corp in CITY_FIXES:
    row = find_row(ws, name_frag, state, city=old_city, corp=old_corp)
    if row is None:
        # Try without corp filter as fallback
        row = find_row(ws, name_frag, state, city=old_city)
    if row is None:
        errors.append(f"CITY FIX: NOT FOUND -- '{name_frag}' in {old_city}, {state}")
    elif isinstance(row, list):
        errors.append(f"CITY FIX: MULTIPLE MATCHES ({len(row)}) -- '{name_frag}'")
    else:
        ws.cell(row, COL['city']).value = new_city
        fac = ws.cell(row, COL['name']).value
        changes['city_fix'].append(f"Row {row}: {fac} | city '{old_city}' -> '{new_city}'")

# --- APPLY DELETES (bottom-up) ---
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx)

# ============================================================
# REPORT
# ============================================================
print(f"\n{'='*60}")
print(f"V24.2 Migration Report -- MUO Candidate Evaluation")
print(f"{'='*60}")
print(f"Corporate reattributions:  {len(changes['corp_reattrib'])}")
print(f"Corporate splits:          {len(changes['corp_split'])}")
print(f"Corporate normalizations:  {len(changes['corp_normalize'])}")
print(f"Facility name fixes:       {len(changes['name_fix'])}")
print(f"County fills:              {len(changes['county_fill'])}")
print(f"Bed count fixes:           {len(changes['bed_fix'])}")
print(f"City fixes:                {len(changes['city_fix'])}")
print(f"Deletes:                   {len(changes['delete'])}")
total_changes = sum(len(v) for v in changes.values())
print(f"{'='*60}")
print(f"TOTAL CHANGES: {total_changes}")
print(f"Final row count: {ws.max_row - 1}")
print()

for category, label in [
    ('corp_reattrib', 'CORPORATE REATTRIBUTIONS'),
    ('corp_split', 'CORPORATE SPLITS'),
    ('corp_normalize', 'CORPORATE NORMALIZATIONS'),
    ('name_fix', 'FACILITY NAME FIXES'),
    ('county_fill', 'COUNTY FILLS'),
    ('bed_fix', 'BED COUNT FIXES'),
    ('city_fix', 'CITY FIXES'),
    ('delete', 'DELETES'),
]:
    items = changes[category]
    if items:
        print(f"{label}:")
        for item in items:
            print(f"  {item}")
        print()

# ============================================================
# VALIDATION
# ============================================================
expected_deletes = len(DELETE_TARGETS)
actual_deletes = len(changes['delete'])
expected_final = total_rows - actual_deletes
actual_final = ws.max_row - 1

if actual_final != expected_final:
    errors.append(f"Row count mismatch: expected {expected_final}, got {actual_final}")

if errors:
    print(f"\n!! {len(errors)} VALIDATION ERRORS !!")
    for e in errors:
        print(f"  {e}")
    print("\nReview errors. NOT FOUND items may need manual investigation.")
    print("Script will still save if only NOT FOUND errors (non-destructive).")

    # Only block save on critical errors (row count mismatch, etc.)
    critical = [e for e in errors if "Row count mismatch" in e or "MULTIPLE MATCHES" in e]
    if critical:
        print("\nCRITICAL ERRORS -- NOT SAVING.")
    else:
        print(f"\n{len(errors)} non-critical errors (NOT FOUND). Saving with available changes...")
        print(f"Saving to {V24_2}...")
        wb.save(V24_2)
        print("Done. Review NOT FOUND items for next patch.")
else:
    print("All validations passed.")
    print(f"Saving to {V24_2}...")
    wb.save(V24_2)
    print("Done.")

wb.close()
