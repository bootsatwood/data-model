"""
V25.4 Migration — Facility-Level Dedup (T1-T5 Confirmed Duplicates)

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_3.xlsx
Output: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_4.xlsx

All classifications verified by authority sources:
  - CMS CCN (SNFs): 1 CCN at address = confirmed duplicate
  - GLR Extract: 1 facility at address = confirmed duplicate
  - State registries: NC DHSR, VA DSS, KY, OH, IN licensing

47 deletes, 13 bed updates, 4 field fixes. Reviewed one-by-one with Roian.
"""
import openpyxl
import re
import os

VAULT = os.path.expanduser(
    "~/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current"
)
V25_3_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_3.xlsx")
V25_4_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_4.xlsx")

COL = {}

def detect_columns(ws):
    global COL
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            COL[val] = c

def norm(s):
    if not s: return ''
    return ' '.join(str(s).upper().strip().split())

def find_row_by_name(ws, name_frag, state, city=None):
    """Find row by facility name fragment + state + optional city."""
    matches = []
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['Facility_Name']).value)
        rs = norm(ws.cell(r, COL['State']).value)
        if name_frag.upper() not in rn or state.upper() != rs:
            continue
        if city and city.upper() != norm(ws.cell(r, COL['City']).value):
            continue
        matches.append(r)
    if len(matches) == 1: return matches[0]
    elif len(matches) == 0: return None
    else: return matches

# ============================================================
# DELETES — 47 rows confirmed as duplicates
# ============================================================
# (name_fragment, city, state) — enough to uniquely identify each row
DELETES = [
    # ALG Senior (20)
    ("MEADOWVIEW TERRACE OF WADESBORO", "Wadesboro", "NC", 60, "NIC-A"),  # beds=60, not the served 51
    ("THE GARDENS OF HENDERSONVILLE", "Hendersonville", "NC", 60, None),
    ("THE LANDINGS OF OAK ISLAND", "Oak Island", "NC", 80, None),  # beds=80, not served 55
    # Littleton triple — delete both non-served
    ("THE LANDINGS OF LAKE GASTON", "Littleton", "NC", 60, "Surrogate"),  # first 60-bed
    ("WOODRIDGE ASSISTED LIVING FACILITY", "Monroe", "NC", None, None),
    ("THE GARDENS OF PAMLICO", "Grantsboro", "NC", None, None),
    ("THE LANDINGS OF MILLS RIVER", "Mills River", "NC", None, None),
    ("CRANBERRY HOUSE", "Newland", "NC", 60, None),  # beds=60, not served 21
    ("CLEVELAND HOUSE", "Shelby", "NC", 72, None),  # beds=72, not served 45
    ("AHOSKIE HOUSE", "Ahoskie", "NC", 60, None),  # beds=60, not served 26
    ("THE GARDENS OF ROSEBORO", "Roseboro", "NC", None, None),
    ("WINDSOR HOUSE", "Windsor", "NC", 60, None),  # beds=60 at 336 Rhodes
    ("MOCKSVILLE SENIOR LIVING AND MEMORY CARE", "Mocksville", "NC", 69, None),
    ("THE LANDINGS OF ALBEMARLE", "Hertford", "NC", 50, None),  # beds=50, not served 34
    ("THE MEADOWS OF ROCKWELL RETIREMENT CENTER", "Rockwell", "NC", None, None),
    ("GREENBRIER OF FAIRMONT", "Fairmont", "NC", None, None),
    ("EAST TOWNE", "Charlotte", "NC", 120, None),  # beds=120, the non-served one
    ("THE BLAIR MEMORY CARE CENTER OF CARY", "Cary", "NC", None, None),
    ("ZEBULON HOUSE", "Zebulon", "NC", 60, None),  # beds=60, not served 57
    # ALTERCARE (2)
    ("SUMMIT ACRES NURSING HOME", "Caldwell", "OH", 95, None),
    ("ALTER CARE NEWARK SOUTH", "Newark", "OH", None, None),
    # AVARDIS (1)
    ("COTTAGES AT LOCUST GROVE", "Mifflin", "PA", None, None),
    # BROOKDALE (11)
    ("BROOKDALE CONCORD PARKWAY", "Concord", "NC", 69, None),  # NIC-A 69, not LEGACY 112
    ("Brookdale Meridian MC (MI)", "Haslett", "MI", None, None),
    ("Brookdale Delta AL (MI)", "Lansing", "MI", None, None),
    ("BROOKDALE CARRIAGE CLUB PROVIDENCE I", "Charlotte", "NC", 77, None),  # PROPCO 77
    ("BROOKDALE CARRIAGE CLUB PROVIDENCE II", "Charlotte", "NC", 34, None),  # PROPCO 34
    ("BROOKDALE HIGH POINT NORTH", "High Point", "NC", 65, None),  # not the MC row
    ("BROOKDALE UNION PARK", "Monroe", "NC", 87, None),  # LEGACY 87
    ("BROOKDALE REIDSVILLE", "Reidsville", "NC", 46, None),  # NIC-A 46
    ("BROOKDALE BRISTOL", "Bristol", "VA", 125, None),  # the not-served 125
    ("BROOKDALE WAKE FOREST", "Wake Forest", "NC", 70, None),  # LEGACY 70
    ("BROOKDALE WOOSTER", "Wooster", "OH", 79, None),  # the not-served 79
    # CCH (1)
    ("RIVERWOOD ASSISTED LIVING FACILITY", "Dobson", "NC", None, None),
    # ELDERCARE (3)
    ("REGENCY MANOR", "Independence", "KY", 59, None),  # the not-served 59
    ("CHRISTIAN COUNTY MANOR I LLC", "Hopkinsville", "KY", None, None),
    ("CORNERSTONE MANOR I LLC", "Scottsville", "KY", None, None),
    # ENVIVE (2)
    ("ENVIVE HEALTHCARE OF BEECH GROVE", "Beech Grove", "IN", None, None),
    ("ENVIVE OF BEECH GROVE", "BEECH GROVE", "IN", 52, None),  # CMS SNF 52, not GLR SNF 46
    # GARDANT (2)
    ("Heritage Woods Of Bolingbrook", "Bolingbrook", "IL", None, None),  # "Of" variant
    ("Green Oaks Supportive Living Of River Oaks", "Calumet City", "IL", None, None),
    # LIBERTY (1)
    ("WARREN HILLS NURSING CENTER SNF", "Warrenton", "NC", None, None),
    # LIONSTONE (1)
    ("PAULDING CARE CENTER", "Paulding", "OH", None, None),
    # MORNING POINTE (1)
    ("MORNING POINTE OF LOUISVILLE", "Louisville", "KY", 73, None),  # LEGACY 73
    # OTTERBEIN (1)
    ("OTTERBEIN LEBANON RETIREMENT COMMUNITY", "Lebanon", "OH", None, None),
    # TRILOGY (1)
    ("SILVER OAKS HEALTH CAMPUS", "Columbus", "IN", 49, None),  # LEGACY ALF 49
]

# ============================================================
# BED UPDATES — 13 rows where kept row has wrong bed count
# ============================================================
# (name_fragment, city, state, new_beds)
BED_UPDATES = [
    ("EAST TOWNE ASSISTED LIVING", "Charlotte", "NC", 120),
    ("BROOKDALE CARRIAGE CLUB PROVIDENCE I ALF", "Charlotte", "NC", 77),
    ("BROOKDALE CARRIAGE CLUB II ALZ", "Charlotte", "NC", 34),
    ("BROOKDALE BRISTOL", "Bristol", "VA", 125),  # keep row is the 33-bed served one
    ("BROOKDALE WAKE FOREST", "Wake Forest", "NC", 70),  # keep row is 36-bed served
    ("BROOKDALE WOOSTER AL", "Wooster", "OH", 117),
    ("RIVERWOOD", "Dobson", "NC", 65),  # keep row is 40-bed served
    ("REGENCY MANOR - INDEPENDENCE", "Independence", "KY", 59),
    ("CHRISTIAN COUNTY MANOR", "Hopkinsville", "KY", 78),
    ("CORNERSTONE MANOR", "Scottsville", "KY", 36),
    ("ENVIVE OF BEECH GROVE - ILF/ALF", "Beech Grove", "IN", 57),
    ("ENVIVE OF BEECH GROVE - SNF", "Beech Grove", "IN", 52),
    ("OTTERBEIN LEBANON SNF", "Lebanon", "OH", 61),
]

# ============================================================
# FIELD FIXES — 4 rows
# ============================================================
FIELD_FIXES = [
    # (name_frag, city, state, field, new_value)
    ("KINGS DAUGHTERS", "Staunton", "VA", "Source_Type", "SNF"),
    ("THE LEGACY AT NORTH AUGUSTA", "Staunton", "VA", "Total_Beds", 71),
    ("THE LEGACY AT NORTH AUGUSTA", "Staunton", "VA", "Do_We_Serve", "Yes"),
    ("SOLISTA HIGH POINT", "High Point", "NC", "Source_Type", "ILF"),
]

# ============================================================
# PROCESS
# ============================================================
print("Loading V25.3...")
wb = openpyxl.load_workbook(V25_3_PATH)
ws = wb.active
detect_columns(ws)
total_rows = ws.max_row - 1
print(f"V25.3 rows: {total_rows}")

errors = []
deletes = []
updates = []

# --- DELETES ---
print("\n--- Processing deletes ---")
rows_to_delete = []
for name_frag, city, state, beds_hint, extra_hint in DELETES:
    candidates = []
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['Facility_Name']).value)
        rs = norm(ws.cell(r, COL['State']).value)
        rc = norm(ws.cell(r, COL['City']).value)
        if name_frag.upper() not in rn or state.upper() != rs:
            continue
        if city and city.upper() != rc:
            continue
        # If beds hint provided, use it to disambiguate
        if beds_hint:
            rb = ws.cell(r, COL['Total_Beds']).value
            if rb and abs(int(rb) - beds_hint) > 5:
                continue
        candidates.append(r)

    if len(candidates) == 1:
        r = candidates[0]
        fac = ws.cell(r, COL['Facility_Name']).value
        beds = ws.cell(r, COL['Total_Beds']).value
        rows_to_delete.append(r)
        deletes.append(f"Row {r}: {fac} | {city}, {state} | beds={beds}")
    elif len(candidates) == 0:
        errors.append(f"DELETE NOT FOUND: '{name_frag}' in {city}, {state} (beds~{beds_hint})")
    else:
        # Multiple matches — try to pick the not-served one
        not_served = [r for r in candidates
                      if not ws.cell(r, COL['Do_We_Serve']).value
                      or str(ws.cell(r, COL['Do_We_Serve']).value).upper() not in ('YES','TRUE')]
        if len(not_served) == 1:
            r = not_served[0]
            fac = ws.cell(r, COL['Facility_Name']).value
            beds = ws.cell(r, COL['Total_Beds']).value
            rows_to_delete.append(r)
            deletes.append(f"Row {r}: {fac} | {city}, {state} | beds={beds} (disambiguated by served)")
        else:
            errors.append(f"DELETE MULTIPLE ({len(candidates)}): '{name_frag}' in {city}, {state} (beds~{beds_hint})")

# Handle Littleton triple — need to delete 2 rows
# The triple has 3 rows: served 48-bed (keep), and two 60-bed not-served (delete both)
littleton_extra = []
for r in range(2, ws.max_row + 1):
    rn = norm(ws.cell(r, COL['Facility_Name']).value)
    rs = norm(ws.cell(r, COL['State']).value)
    rc = norm(ws.cell(r, COL['City']).value)
    if 'LANDINGS' in rn and 'LAKE GASTON' in rn and rs == 'NC' and rc == 'LITTLETON':
        beds = ws.cell(r, COL['Total_Beds']).value
        served = str(ws.cell(r, COL['Do_We_Serve']).value or '').upper()
        if beds and int(beds) == 60 and served not in ('YES', 'TRUE') and r not in rows_to_delete:
            littleton_extra.append(r)

for r in littleton_extra:
    fac = ws.cell(r, COL['Facility_Name']).value
    rows_to_delete.append(r)
    deletes.append(f"Row {r}: {fac} | Littleton, NC | beds=60 (Littleton triple extra)")

print(f"  Deletes found: {len(rows_to_delete)}")

# --- BED UPDATES ---
print("--- Processing bed updates ---")
for name_frag, city, state, new_beds in BED_UPDATES:
    row = find_row_by_name(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"BED UPDATE NOT FOUND: '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        # Pick the served one
        served = [r for r in row
                  if str(ws.cell(r, COL['Do_We_Serve']).value or '').upper() in ('YES','TRUE')]
        if len(served) == 1:
            row = served[0]
        else:
            errors.append(f"BED UPDATE MULTIPLE ({len(row)}): '{name_frag}' in {city}, {state}")
            continue
    if row and not isinstance(row, list):
        old = ws.cell(row, COL['Total_Beds']).value
        ws.cell(row, COL['Total_Beds']).value = new_beds
        fac = ws.cell(row, COL['Facility_Name']).value
        updates.append(f"Row {row}: {fac} | beds {old} -> {new_beds}")

# --- FIELD FIXES ---
print("--- Processing field fixes ---")
for name_frag, city, state, field, new_value in FIELD_FIXES:
    row = find_row_by_name(ws, name_frag, state, city=city)
    if row is None:
        errors.append(f"FIELD FIX NOT FOUND: '{name_frag}' in {city}, {state}")
    elif isinstance(row, list):
        row = row[0]  # take first match
    if row and not isinstance(row, list):
        old = ws.cell(row, COL[field]).value
        ws.cell(row, COL[field]).value = new_value
        fac = ws.cell(row, COL['Facility_Name']).value
        updates.append(f"Row {row}: {fac} | {field} '{old}' -> '{new_value}'")

# --- APPLY DELETES ---
print(f"\n--- Applying {len(set(rows_to_delete))} deletes ---")
for r in sorted(set(rows_to_delete), reverse=True):
    ws.delete_rows(r)

# ============================================================
# REPORT
# ============================================================
final_rows = ws.max_row - 1
print(f"\n{'=' * 70}")
print(f"V25.4 Dedup Migration Report")
print(f"{'=' * 70}")
print(f"Deletes: {len(set(rows_to_delete))}")
print(f"Bed updates: {len([u for u in updates if 'beds' in u])}")
print(f"Field fixes: {len([u for u in updates if 'beds' not in u])}")
print(f"Final row count: {final_rows}")
print()

if deletes:
    print("DELETES:")
    for d in deletes:
        print(f"  {d}")

if updates:
    print("\nUPDATES:")
    for u in updates:
        print(f"  {u}")

if errors:
    print(f"\n!! {len(errors)} WARNINGS !!")
    for e in errors:
        print(f"  {e}")

# ============================================================
# SAVE
# ============================================================
print(f"\nSaving to {V25_4_PATH}...")
wb.save(V25_4_PATH)
print("Done.")
wb.close()
