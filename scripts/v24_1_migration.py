"""
V24.1 Migration -- Apply accumulated IN fixes to V24 database.
Row-number based with content verification for safety.
"""
import openpyxl
import re

V24 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24.xlsx"
V24_1 = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_1.xlsx"

def norm(s):
    if not s: return ''
    s = str(s).upper().strip()
    s = re.sub(r'\.(?=\s|$)', '', s)
    s = re.sub(r'(?<=\b[A-Z])\.', '', s)
    s = ' '.join(s.split())
    return s

# Column indices (1-based for openpyxl)
COL = {
    'type': 1, 'name': 2, 'corp': 3, 'addr': 4, 'city': 5,
    'state': 6, 'zip': 7, 'county': 8, 'beds': 10, 'census': 11,
    'served': 12, 'int': 13, 'pcp': 14, 'mh': 15, 'dq': 19,
    'corp_source': 26
}

# ============================================================
# CHANGE DEFINITIONS (all V24 row numbers)
# ============================================================

# -- DELETES: 26 rows --
# (row_number, name_fragment_for_verification)
DELETE_ROWS = [
    # Cluster review (22)
    (19991, "TOWNE PARK ASSISTED LIVING"),
    (18104, "GLASSWATER CREEK"),
    (20108, "VITA SENIOR LIVING OF GREENFIELD"),
    (5044,  "RESTORACY OF CARMEL"),
    (18093, "GENTRY PARK"),
    (17729, "COMPASS PARK"),
    (18171, "GREENWOOD VILLAGE SOUTH"),
    (4850,  "HELLENIC SENIOR LIVING OF ELKHART"),
    (18267, "HELLENIC SENIOR LIVING OF ELKHART"),
    (4852,  "HELLENIC SENIOR LIVING OF MISHAWAKA"),
    (17947, "ELMS"),
    (18620, "LUTHERAN LIFE VILLAGES"),
    (18740, "MENNONITE MEMORIAL HOME"),
    (5182,  "TERRACE AT SOLARBRON"),
    (19594, "SWEET GALILEE"),
    (19117, "SACRED HEART"),
    (19595, "SWISS VILLAGE"),
    (18366, "HOOSIER CHRISTIAN VILLAGE"),
    (17536, "CARMEL CARE CENTER"),
    (20270, "WOODCREST OF DECATUR"),
    (18374, "HUBBARD HILL"),
    (18765, "MILLER"),
    (18764, "MILLER"),
    # IL contamination (4)
    (17478, "BROOKHAVEN"),
    (18081, "GARDENS ON GATEWAY"),
    (18370, "HORIZON HOMES"),
    (18870, "NORA COMMONS"),
]

# -- TYPE FIXES: 11 ALF -> SNF --
TYPE_FIX_ROWS = [
    (4732,  "ENVIVE OF BEECH GROVE - SNF"),
    (4774,  "GOLDEN YEARS - FORT WAYNE - SNF"),
    (4846,  "HEARTHSTONE HEALTH SNF"),
    (5035,  "PAOLI HEALTH AND LIVING COMMUNITY SNF"),
    (5075,  "ROBIN RUN SNF"),
    (5143,  "STONEBRIDGE SNF"),
    (5145,  "STONECROFT CAMPUS SNF"),
    (5189,  "THE WATERS OF GEORGETOWN SNF"),
    (5202,  "TIMBERCREST SNF"),
    (5206,  "TOWNE HOUSE RETIREMENT COMMUNITY SNF"),
    (5327,  "WILDWOOD HEALTHCARE CENTER SNF"),
]

# -- CITY FIX: 1 row --
CITY_FIX = (5202, "TIMBERCREST SNF", "North Manchester")

# -- SERVICE FLAG TRANSFER: 1 row --
SVC_TRANSFER = (4830, "HOOSIER HEALTH")  # served=Yes, MH=Yes (actual name: HOOSIER HEALTH & LIVING COMMUNITY)

# -- CORP NAME FILLS: 3 rows --
CORP_FILLS = [
    (4547, "BELLTOWER", "FUNDAMENTAL HEALTHCARE", "CMS"),
    (4765, "GREENWOOD VILLAGE SOUTH", "LIFE CARE SERVICES", "CMS"),
    (5300, "WOODLAND MANOR", "IDE MANAGEMENT GROUP", "CMS"),
]

# -- CORP NAME FIX: 1 row --
CORP_FIX = (5320, "WEST LAFAYETTE", "WICKSHIRE SENIOR LIVING")

# -- COUNTY FILLS: 55 rows --
ZIP_TO_COUNTY = {
    "46060": "Hamilton", "46072": "Tipton", "46077": "Boone", "46112": "Hendricks",
    "46131": "Johnson", "46140": "Hancock", "46143": "Johnson", "46176": "Shelby",
    "46182": "Shelby", "46202": "Marion", "46218": "Marion", "46219": "Marion",
    "46220": "Marion", "46237": "Marion", "46268": "Marion", "46304": "Porter",
    "46307": "Lake", "46514": "Elkhart", "46528": "Elkhart", "46530": "St. Joseph",
    "46545": "St. Joseph", "46552": "St. Joseph", "46703": "Steuben",
    "46750": "Huntington", "46804": "Allen", "46815": "Allen", "46825": "Allen",
    "46835": "Allen", "46962": "Wabash", "46975": "Fulton", "47006": "Ripley",
    "47012": "Franklin", "47025": "Dearborn", "47043": "Switzerland",
    "47201": "Bartholomew", "47265": "Jennings", "47353": "Union", "47356": "Henry",
    "47401": "Monroe", "47421": "Lawrence", "47454": "Orange", "47546": "Dubois",
    "47803": "Vigo", "47804": "Vigo", "47904": "Tippecanoe",
}

# ============================================================
# PROCESS
# ============================================================

print("Loading V24...")
wb = openpyxl.load_workbook(V24)
ws = wb.active
total_rows = ws.max_row - 1
print(f"V24 rows: {total_rows}")

errors = []
deleted = []
type_fixed = []
city_fixed = []
svc_transferred = []
corp_filled = []
corp_fixed_list = []
county_filled = []

def verify(row_idx, name_frag):
    """Verify row contains expected name fragment."""
    val = str(ws.cell(row_idx, COL['name']).value or '')
    if name_frag.upper() not in val.upper():
        errors.append(f"VERIFY FAIL row {row_idx}: expected '{name_frag}', got '{val}'")
        return False
    return True

# -- DELETES --
rows_to_delete = []
for row_idx, name_frag in DELETE_ROWS:
    if verify(row_idx, name_frag):
        val = ws.cell(row_idx, COL['name']).value
        beds = ws.cell(row_idx, COL['beds']).value
        city = ws.cell(row_idx, COL['city']).value
        rows_to_delete.append(row_idx)
        deleted.append(f"Row {row_idx}: {val} | {city} | beds={beds}")

# -- TYPE FIXES --
for row_idx, name_frag in TYPE_FIX_ROWS:
    if verify(row_idx, name_frag):
        old_type = ws.cell(row_idx, COL['type']).value
        if old_type != 'ALF':
            errors.append(f"TYPE FIX row {row_idx}: expected ALF, got '{old_type}'")
        else:
            ws.cell(row_idx, COL['type']).value = 'SNF'
            val = ws.cell(row_idx, COL['name']).value
            type_fixed.append(f"Row {row_idx}: {val} ALF->SNF")

# -- CITY FIX --
row_idx, name_frag, new_city = CITY_FIX
if verify(row_idx, name_frag):
    old_city = ws.cell(row_idx, COL['city']).value
    ws.cell(row_idx, COL['city']).value = new_city
    val = ws.cell(row_idx, COL['name']).value
    city_fixed.append(f"Row {row_idx}: {val} city '{old_city}' -> '{new_city}'")

# -- SERVICE FLAG TRANSFER --
row_idx, name_frag = SVC_TRANSFER
if verify(row_idx, name_frag):
    old_served = ws.cell(row_idx, COL['served']).value
    old_mh = ws.cell(row_idx, COL['mh']).value
    ws.cell(row_idx, COL['served']).value = 'Yes'
    ws.cell(row_idx, COL['mh']).value = 'Yes'
    val = ws.cell(row_idx, COL['name']).value
    svc_transferred.append(f"Row {row_idx}: {val} served={old_served}->Yes, MH={old_mh}->Yes")

# -- CORP NAME FILLS --
for row_idx, name_frag, new_corp, source in CORP_FILLS:
    if verify(row_idx, name_frag):
        old_corp = ws.cell(row_idx, COL['corp']).value
        ws.cell(row_idx, COL['corp']).value = new_corp
        ws.cell(row_idx, COL['corp_source']).value = source
        val = ws.cell(row_idx, COL['name']).value
        corp_filled.append(f"Row {row_idx}: {val} corp='{old_corp}' -> '{new_corp}'")

# -- CORP NAME FIX --
row_idx, name_frag, new_corp = CORP_FIX
if verify(row_idx, name_frag):
    old_corp = ws.cell(row_idx, COL['corp']).value
    ws.cell(row_idx, COL['corp']).value = new_corp
    val = ws.cell(row_idx, COL['name']).value
    corp_fixed_list.append(f"Row {row_idx}: {val} corp='{old_corp}' -> '{new_corp}'")

# -- COUNTY FILLS (content-based: IN rows with blank county) --
for row_idx in range(2, ws.max_row + 1):
    state = str(ws.cell(row_idx, COL['state']).value or '').strip()
    county = str(ws.cell(row_idx, COL['county']).value or '').strip()
    if state == 'IN' and (not county or county == 'None'):
        zip_val = str(ws.cell(row_idx, COL['zip']).value or '')[:5]
        mapped = ZIP_TO_COUNTY.get(zip_val)
        if mapped:
            ws.cell(row_idx, COL['county']).value = mapped
            val = ws.cell(row_idx, COL['name']).value
            county_filled.append(f"Row {row_idx}: {val} -> county={mapped}")

# -- APPLY DELETES (bottom-up) --
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx)

# ============================================================
# REPORT
# ============================================================
print(f"\n{'='*60}")
print(f"V24.1 Migration Report")
print(f"{'='*60}")
print(f"Deletes:              {len(deleted)} rows")
print(f"Type fixes (ALF->SNF): {len(type_fixed)} rows")
print(f"City fixes:           {len(city_fixed)} rows")
print(f"Service transfers:    {len(svc_transferred)} rows")
print(f"Corp name fills:      {len(corp_filled)} rows")
print(f"Corp name fixes:      {len(corp_fixed_list)} rows")
print(f"County fills:         {len(county_filled)} rows")
print(f"{'='*60}")
print(f"Final row count: {ws.max_row - 1}")
print()

if deleted:
    print("DELETED:")
    for d in deleted:
        print(f"  {d}")
    print()

if type_fixed:
    print("TYPE FIXED (ALF->SNF):")
    for t in type_fixed:
        print(f"  {t}")
    print()

if city_fixed:
    print("CITY FIXED:")
    for c in city_fixed:
        print(f"  {c}")
    print()

if svc_transferred:
    print("SERVICE FLAG TRANSFER:")
    for s in svc_transferred:
        print(f"  {s}")
    print()

if corp_filled:
    print("CORP NAME FILLS:")
    for c in corp_filled:
        print(f"  {c}")
    print()

if corp_fixed_list:
    print("CORP NAME FIXES:")
    for c in corp_fixed_list:
        print(f"  {c}")
    print()

print(f"COUNTY FILLS: {len(county_filled)} rows")
if county_filled:
    for c in county_filled[:5]:
        print(f"  {c}")
    if len(county_filled) > 5:
        print(f"  ... and {len(county_filled) - 5} more")
    print()

# ============================================================
# VALIDATION
# ============================================================

if len(deleted) != 27:
    errors.append(f"Expected 27 deletes, got {len(deleted)}")
if len(type_fixed) != 11:
    errors.append(f"Expected 11 type fixes, got {len(type_fixed)}")
if len(city_fixed) != 1:
    errors.append(f"Expected 1 city fix, got {len(city_fixed)}")
if len(svc_transferred) != 1:
    errors.append(f"Expected 1 service transfer, got {len(svc_transferred)}")
if len(corp_filled) != 3:
    errors.append(f"Expected 3 corp fills, got {len(corp_filled)}")
if len(corp_fixed_list) != 1:
    errors.append(f"Expected 1 corp fix, got {len(corp_fixed_list)}")

expected_final = total_rows - 27
actual_final = ws.max_row - 1
if actual_final != expected_final:
    errors.append(f"Expected final count {expected_final}, got {actual_final}")

if errors:
    print("!! VALIDATION ERRORS !!")
    for e in errors:
        print(f"  {e}")
    print("\nNOT SAVING -- fix errors first.")
else:
    print("All validations passed.")
    print(f"Saving to {V24_1}...")
    wb.save(V24_1)
    print("Done.")

wb.close()
