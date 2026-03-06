#!/usr/bin/env python3
"""Apply Liberty corrections to V22.1 -> V22.2.

Changes:
  - 61 rows: Corporate_Name -> "LIBERTY" (confirmed Liberty Health facilities)
  - 9 rows: Corporate_Name -> "LIBERTY VILLAGE" (separate IL company)
  - 2 rows: Source_Type ALF -> SNF (8774, 9983)
"""

import sys
sys.path.insert(0, ".")
from utils import safe, load_db, norm_addr, norm, VAULT
from openpyxl import load_workbook

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_1.xlsx"
DB_TARGET = VAULT / "Current" / "1_Combined_Database_FINAL_V22_2.xlsx"

# ── Build change map from same logic as QC review ──
headers, rows = load_db(DB_SOURCE)
print(f"V22.1 loaded: {len(rows):,} rows")

db_by_addr = {}
for r in rows:
    akey = norm_addr(safe(r.get("Address", ""))) + "|" + norm(safe(r.get("City", ""))) + "|" + norm(safe(r.get("State", "")))
    db_by_addr.setdefault(akey, []).append(r)


def addr_lookup(address, city, state):
    akey = norm_addr(address) + "|" + norm(city) + "|" + norm(state)
    return db_by_addr.get(akey, [])


# ── Confirmed LHR addresses ──
lhr_addresses = [
    ("791 Boone Station Drive", "Burlington", "NC"),
    ("1403 Connor Ave.", "Windsor", "NC"),
    ("208 Mercer Road", "Elizabethtown", "NC"),
    ("630 N Fodale Ave", "Southport", "NC"),
    ("104 Holcombe Cove Road", "Candler", "NC"),
    ("1402 Pinckney Street", "Whiteville", "NC"),
    ("200 Flowers-Pridgen Drive", "Whiteville", "NC"),
    ("7348 North West Street", "Falcon", "NC"),
    ("400 Pelt Drive", "Fayetteville", "NC"),
    ("1700 Pamalee Dr.", "Fayetteville", "NC"),
    ("316 NC Highway 801 South", "Advance", "NC"),
    ("6041 Piedmont Row Drive", "Charlotte", "NC"),
    ("5680 Windy Hill Drive", "Winston-Salem", "NC"),
    ("901 Bethesda Road", "Winston-Salem", "NC"),
    ("485 Veteran's Way", "Kernersville", "NC"),
    ("202 Smoketree Way", "Louisburg", "NC"),
    ("101 Caroline Avenue", "Weldon", "NC"),
    ("100 Silver Bluff Dr.", "Canton", "NC"),
    ("2315 Highway 242 North", "Benson", "NC"),
    ("310 Commerce Drive", "Sanford", "NC"),
    ("3100 Tramway Road", "Sanford", "NC"),
    ("2700 Royal Commons Lane", "Matthews", "NC"),
    ("10011 Providence Road West", "Charlotte", "NC"),
    ("300 Blake Road", "Pinehurst", "NC"),
    ("155 Blake Road", "Pinehurst", "NC"),
    ("740 Diamond Shoals Road", "Wilmington", "NC"),
    ("121 Racine Drive", "Wilmington", "NC"),
    ("1718 Legion Road", "Chapel Hill", "NC"),
    ("901 Ridge Road", "Roxboro", "NC"),
    ("1150 Pine Run Dr.", "Lumberton", "NC"),
    ("4412 South Main Street", "Salisbury", "NC"),
    ("120 Southwood Drive", "Clinton", "NC"),
    ("180 Southwood Drive", "Clinton", "NC"),
    ("3000 Holston Lane", "Raleigh", "NC"),
    ("2750 Oberlin Road", "Raleigh", "NC"),
    ("221 Brightmore Drive", "Cary", "NC"),
    ("864 US Hwy. 158 Business West", "Warrenton", "NC"),
    ("621 Chestnut Ridge Parkway", "Blowing Rock", "NC"),
    ("903 W. Main Street", "Yadkinville", "NC"),
    ("194 Spring Street", "Charleston", "SC"),
    ("100 Samaritan Way", "Crossville", "TN"),
]

# ── Confirmed LSL addresses ──
lsl_addresses = [
    ("10225 Old Ardrey Kell Road", "Charlotte", "NC"),
    ("2124 41st Street", "Wilmington", "NC"),
    ("450 East Ocean Avenue", "Lantana", "FL"),
    ("620 Carolina Bay Drive", "Wilmington", "NC"),
    ("2600 Yettington Drive", "Raleigh", "NC"),
    ("152 Market Chapel Road", "Pittsboro", "NC"),
    ("6146 Liberty Hall Drive", "Leland", "NC"),
    ("4101 Glenloch Circle", "Matthews", "NC"),
    ("440 Town Center Place", "Columbia", "SC"),
    ("194 Spring Street", "Charleston", "SC"),
    ("3045 Henderson Drive Extension", "Jacksonville", "NC"),
    ("1611 Constitution Boulevard", "Rock Hill", "SC"),
    ("800 N Lake Drive", "Lexington", "SC"),
    ("95 Holcombe Cove Road", "Candler", "NC"),
    ("155 Blake Boulevard", "Pinehurst", "NC"),
    ("1400 Liberty Midtown Drive", "Mount Pleasant", "SC"),
    ("4801 Barclay Downs Drive", "Charlotte", "NC"),
    ("625 King Street", "Charleston", "SC"),
    ("125 Brightmore Drive", "Cary", "NC"),
    ("10430 Stable Lane", "Wellington", "FL"),
]

# ── Not Liberty Health — exclude from renames ──
NOT_LIBERTY_HEALTH = {
    "SEACOAST AT LIBERTY RIDGE",
    "LIBERTY RIDGE SENIOR LIVING INC",
    "LIBERTY RETIREMENT PROPERTIES OF LIMA LTD",
    "WEST LIBERTY GARDEN APARTMENTS",
}

LIBERTY_VILLAGE_CORPS = {
    "LIBERTY VILLAGE",
    "Liberty Village of Freeport", "Liberty Village of Geneseo",
    "Liberty Village of Jerseyville", "Liberty Village of Peoria",
    "Liberty Village of Peru", "Liberty Village of Rochelle",
    "Liberty Village of Streator",
}

BARE_LIBERTY_CONFIRMED = {8773, 8774, 9291, 9487, 9641, 9982, 9983}

SOURCE_TYPE_FIXES = {
    8774: "SNF",
    9983: "SNF",
}

# ── Collect rows to rename to "LIBERTY" ──
rename_to_liberty = set()

for (addr, city, state) in lhr_addresses:
    for m in addr_lookup(addr, city, state):
        rename_to_liberty.add(m["_excel_row"])

for (addr, city, state) in lsl_addresses:
    for m in addr_lookup(addr, city, state):
        rename_to_liberty.add(m["_excel_row"])

for row_num in BARE_LIBERTY_CONFIRMED:
    rename_to_liberty.add(row_num)

for r in rows:
    if safe(r.get("Corporate_Name", "")).upper() == "LIBERTY SENIOR LIVING":
        rename_to_liberty.add(r["_excel_row"])

# Remove rows that are already "LIBERTY"
already_liberty = set()
for r in rows:
    if r["_excel_row"] in rename_to_liberty and safe(r.get("Corporate_Name", "")) == "LIBERTY":
        already_liberty.add(r["_excel_row"])
rename_to_liberty -= already_liberty

# Remove not-Liberty-Health and Liberty Village from the set
for r in rows:
    corp = safe(r.get("Corporate_Name", ""))
    if corp in NOT_LIBERTY_HEALTH or corp in LIBERTY_VILLAGE_CORPS:
        rename_to_liberty.discard(r["_excel_row"])

# ── Collect Liberty Village renames ──
rename_to_lv = set()
for r in rows:
    if safe(r.get("Corporate_Name", "")) in LIBERTY_VILLAGE_CORPS:
        rename_to_lv.add(r["_excel_row"])

print(f"\nChanges to apply:")
print(f"  Corporate_Name -> 'LIBERTY':        {len(rename_to_liberty)} rows")
print(f"  Corporate_Name -> 'LIBERTY VILLAGE': {len(rename_to_lv)} rows")
print(f"  Source_Type ALF -> SNF:              {len(SOURCE_TYPE_FIXES)} rows")

# ── Open V22.1 in write mode ──
wb = load_workbook(DB_SOURCE)
ws = wb.active
ws_headers = [safe(c.value) for c in ws[1]]
ci_corp = ws_headers.index("Corporate_Name") + 1
ci_stype = ws_headers.index("Source_Type") + 1
ci_fac = ws_headers.index("Facility_Name") + 1

# Apply Corporate_Name -> "LIBERTY"
corp_updated = 0
for excel_row in sorted(rename_to_liberty):
    old_val = safe(ws.cell(row=excel_row, column=ci_corp).value)
    ws.cell(row=excel_row, column=ci_corp).value = "LIBERTY"
    fac = safe(ws.cell(row=excel_row, column=ci_fac).value)
    print(f"  Row {excel_row}: {fac} | Corp: '{old_val}' -> 'LIBERTY'")
    corp_updated += 1

# Apply Corporate_Name -> "LIBERTY VILLAGE"
lv_updated = 0
for excel_row in sorted(rename_to_lv):
    old_val = safe(ws.cell(row=excel_row, column=ci_corp).value)
    ws.cell(row=excel_row, column=ci_corp).value = "LIBERTY VILLAGE"
    fac = safe(ws.cell(row=excel_row, column=ci_fac).value)
    print(f"  Row {excel_row}: {fac} | Corp: '{old_val}' -> 'LIBERTY VILLAGE'")
    lv_updated += 1

# Apply Source_Type fixes
stype_updated = 0
for excel_row, new_stype in SOURCE_TYPE_FIXES.items():
    old_val = safe(ws.cell(row=excel_row, column=ci_stype).value)
    ws.cell(row=excel_row, column=ci_stype).value = new_stype
    fac = safe(ws.cell(row=excel_row, column=ci_fac).value)
    print(f"  Row {excel_row}: {fac} | Source_Type: '{old_val}' -> '{new_stype}'")
    stype_updated += 1

wb.save(DB_TARGET)
wb.close()
print(f"\n{'='*60}")
print(f"Applied to V22.2:")
print(f"  Corporate_Name -> LIBERTY:        {corp_updated}")
print(f"  Corporate_Name -> LIBERTY VILLAGE: {lv_updated}")
print(f"  Source_Type ALF -> SNF:            {stype_updated}")
print(f"  Total changes:                     {corp_updated + lv_updated + stype_updated}")
print(f"\nSaved: {DB_TARGET}")

# ── Quick QC counts ──
print(f"\n{'='*60}")
print("QC VERIFICATION")
print(f"{'='*60}")
h2, r2 = load_db(DB_TARGET)
total = len(r2)
snf = sum(1 for r in r2 if safe(r.get("Source_Type", "")) == "SNF")
alf = sum(1 for r in r2 if safe(r.get("Source_Type", "")) == "ALF")
ilf = sum(1 for r in r2 if safe(r.get("Source_Type", "")) == "ILF")
served = sum(1 for r in r2 if safe(r.get("Do_We_Serve", "")).upper() in ("YES", "Y", "TRUE", "1"))
liberty_count = sum(1 for r in r2 if safe(r.get("Corporate_Name", "")) == "LIBERTY")
lv_count = sum(1 for r in r2 if safe(r.get("Corporate_Name", "")) == "LIBERTY VILLAGE")
lsl_count = sum(1 for r in r2 if safe(r.get("Corporate_Name", "")).upper() == "LIBERTY SENIOR LIVING")

print(f"  Total facilities: {total:,}")
print(f"  SNF:              {snf:,}")
print(f"  ALF:              {alf:,}")
print(f"  ILF:              {ilf:,}")
print(f"  Served:           {served:,}")
print(f"  Corp='LIBERTY':       {liberty_count}")
print(f"  Corp='LIBERTY VILLAGE': {lv_count}")
print(f"  Corp='LIBERTY SENIOR LIVING': {lsl_count} (should be 0)")
