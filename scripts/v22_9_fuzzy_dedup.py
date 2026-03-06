#!/usr/bin/env python3
"""
High-Confidence Fuzzy Match Consolidation — V22.8 -> V22.9

Applies:
  1. 14 fuzzy match consolidations (confirmed same-entity via GLR cross-ref,
     geographic analysis, and corporate research)
  2. Four-Rule ownership reclassification for affected groups

All pairs were identified in Corporate_Name_Dedup_Review_ra_v2.xlsx (Fuzzy
Matches sheet) and verified against the V22.8 database. Only high-confidence
consolidations are included — ambiguous pairs are deferred.

Usage:
  python v22_9_fuzzy_dedup.py preview
  python v22_9_fuzzy_dedup.py apply
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, VAULT

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_8.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_9.xlsx"

# ---------------------------------------------------------------------------
# Corporate Name Renames — High-Confidence Fuzzy Matches
# ---------------------------------------------------------------------------

FUZZY_RENAMES = {
    # GLR canonical consolidations
    "GENESIS HEALTHCARE": "GENESIS",
    "StoryPoint Group": "STORYPOINT",
    "FOUNDATIONS HEALTH SOLUTIONS": "FOUNDATIONS HEALTH",
    "PRINCIPLE LONG TERM CARE": "PRINCIPLE",
    "Principle LTC": "PRINCIPLE",
    "The Arbor Company": "ARBOR COMPANY",
    "BICKFORD SENIOR LIVING": "BICKFORD",
    "A PHOENIX SENIOR LIVING COMMUNITY": "PHOENIX SENIOR LIVING",
    # GLR — canonical is the form with most served facilities
    "TERRA BELLA": "TerraBella Senior Living",
    "LIONSTONE": "LIONSTONE CARE",
    # CMS state-variant of GLR name
    "HCF MANAGEMENT INDIANA": "HCF MANAGEMENT",
    # Same company, highest count
    "SRI MANAGEMENT GROUP": "SRI Management",
    # Full legal name → CMS short form (0 served either side, data cleanup)
    "Evangelical Lutheran Good Samaritan Society/ Sandford": "GOOD SAMARITAN SOCIETY",
    "THE EVANGELICAL LUTHERAN GOOD SAMARITAN SOCIETY": "GOOD SAMARITAN SOCIETY",
}

# ---------------------------------------------------------------------------
# NOT SAME ENTITY — documented here for auditability
# ---------------------------------------------------------------------------
# NATIONAL HEALTHCARE CORPORATION (86 rows, SE US: AL,GA,KY,MO,SC,TN,VA)
#   vs NATIONAL HEALTH CARE ASSOCIATES (42 rows, New England: CT,MA,ME,NH,NY,VT)
#   -> DIFFERENT companies, completely separate geographies
#
# LIFECARE (8 rows, IN only, GLR)
#   vs LIFE CARE CENTERS OF AMERICA (194 rows, national, CMS)
#   vs LIFE CARE SERVICES (75 rows, national management co, CMS)
#   vs LIFE CARE CENTERS (2 rows, OH, GLR)
#   -> 4 SEPARATE entities per GLR authority
#
# CASCADIA HEALTHCARE (44 rows, Pacific NW) vs CASCADES HEALTHCARE (20 rows, ID/TX/UT)
#   -> DEFERRED — geographic overlap in ID/WA but uncertain affiliation
#
# LAUREL HEALTH CARE COMPANY (2 rows, GLR) vs CIENA HEALTHCARE/LAUREL HEALTH CARE (86 rows, CMS)
#   -> DEFERRED — acquisition/naming transition, needs more research


# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    if not corp_name or corp_name.upper() in ('UNKNOWN', ''):
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 0) >= 2:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python v22_9_fuzzy_dedup.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"High-Confidence Fuzzy Match Consolidation — V22.8 -> V22.9 ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_SOURCE.name}...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # --- Phase 1: Corporate Name Renames ---
    print("\nPhase 1: Corporate Name Renames")
    print("-" * 40)

    rename_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp in FUZZY_RENAMES:
            new_corp = FUZZY_RENAMES[corp]
            rename_counts[f"{corp} -> {new_corp}"] += 1
            r['Corporate_Name'] = new_corp

    for key in sorted(rename_counts.keys()):
        print(f"  {rename_counts[key]:>4d}  {key}")

    total_renames = sum(rename_counts.values())
    print(f"\n  {total_renames} total renames")

    # --- Phase 2: Four-Rule Reclassification ---
    print(f"\nPhase 2: Four-Rule Ownership Reclassification")
    print("-" * 40)

    corp_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp:
            corp_counts[corp] += 1

    affected_corps = set(FUZZY_RENAMES.keys()) | set(FUZZY_RENAMES.values())
    reclass_count = 0
    reclass_detail = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp not in affected_corps:
            continue
        old_type = safe(r.get('Ownership_Type', ''))
        new_type = four_rule_classify(corp, corp_counts)
        if old_type != new_type:
            reclass_count += 1
            reclass_detail[f"{old_type} -> {new_type}"] += 1
            print(f"    {safe(r.get('Facility_Name',''))}, {safe(r.get('City',''))}, {safe(r.get('State',''))}")
            print(f"      Corp: {corp} | {old_type} -> {new_type}")
            r['Ownership_Type'] = new_type

    print(f"\n  {reclass_count} reclassifications")
    for key in sorted(reclass_detail.keys()):
        print(f"    {reclass_detail[key]:>4d}  {key}")

    # --- Summary ---
    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Total renames:                 {total_renames}")
    print(f"  Ownership reclassifications:   {reclass_count}")
    print(f"  Rows deleted:                  0")
    print(f"  Net row delta:                 0")
    print(f"  Final row count:               {len(rows):,}")

    # Post-consolidation counts
    print(f"\nPost-consolidation counts for affected names:")
    for canonical in sorted(set(FUZZY_RENAMES.values())):
        cnt = corp_counts.get(canonical, 0)
        served = sum(1 for r in rows if safe(r.get('Corporate_Name', '')) == canonical
                     and safe(r.get('Do_We_Serve', '')) == 'Yes')
        print(f"    {cnt:>4d} ({served:>2d} served)  {canonical}")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY -- no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.9 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    corp_col = col_idx.get('Corporate_Name')
    own_col = col_idx.get('Ownership_Type')

    cells_updated = 0
    for r in rows:
        excel_row = r['_excel_row']

        old_corp = safe(ws.cell(row=excel_row, column=corp_col).value)
        new_corp = safe(r.get('Corporate_Name', ''))
        if old_corp != new_corp:
            ws.cell(row=excel_row, column=corp_col).value = new_corp
            cells_updated += 1

        old_own = safe(ws.cell(row=excel_row, column=own_col).value)
        new_own = safe(r.get('Ownership_Type', ''))
        if old_own != new_own:
            ws.cell(row=excel_row, column=own_col).value = new_own
            cells_updated += 1

    wb.save(DB_OUTPUT)
    print(f"  {cells_updated} cells updated")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
