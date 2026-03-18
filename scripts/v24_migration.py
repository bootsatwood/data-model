"""
V24 Migration Script
Applies all confirmed fixes from data_remediation_progress.md to V23, saves as V24.
Matches every fix by content (name + address + city + state), NOT by row ID.
"""

import openpyxl
from copy import copy
import sys

V23_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V23.xlsx"
V24_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24.xlsx"

# Column indices (0-based)
COL = {
    'Source_Type': 0,
    'Facility_Name': 1,
    'Corporate_Name': 2,
    'Address': 3,
    'City': 4,
    'State': 5,
    'ZIP': 6,
    'County': 7,
    'Ownership_Type': 8,
    'Total_Beds': 9,
    'Census': 10,
    'Do_We_Serve': 11,
    'Integrated_Flag': 12,
    'PCP_Flag': 13,
    'MH_Flag': 14,
    'Barrier': 15,
    'Latitude': 16,
    'Longitude': 17,
    'Data_Quality_Flag': 18,
}

def norm(s):
    """Normalize string for matching: uppercase, strip, collapse whitespace, standardize abbreviations."""
    if s is None:
        return ''
    result = ' '.join(str(s).upper().strip().split())
    # Remove trailing periods
    result = result.rstrip('.')
    # Remove periods after single letters (N. -> N, S. -> S, etc.)
    import re
    result = re.sub(r'\b([A-Z])\.\s', r'\1 ', result)
    result = re.sub(r'\b([A-Z])\.$', r'\1', result)
    return result


def norm_addr(s):
    """Normalize address for matching: norm() plus street suffix standardization."""
    result = norm(s)
    # Standardize common street suffixes
    replacements = [
        (' ROAD', ' RD'), (' STREET', ' ST'), (' AVENUE', ' AVE'),
        (' BOULEVARD', ' BLVD'), (' DRIVE', ' DR'), (' LANE', ' LN'),
        (' CIRCLE', ' CIR'), (' COURT', ' CT'), (' PLACE', ' PL'),
        (' TRAIL', ' TRL'), (' PIKE', ' PIKE'),  # Pike stays
    ]
    for old, new in replacements:
        if result.endswith(old):
            result = result[:-len(old)] + new
    return result


# ============================================================
# DEFINE ALL FIXES
# ============================================================

# DELETES: (name, address, city, state, beds) — beds used for disambiguation
DELETES = [
    ("OTTERBEIN FRANKLIN SENIORLIFE COMM RES & COM CARE", "1070 W JEFFERSON ST", "FRANKLIN", "IN", 208),
    ("ST ANDREWS HEALTH CAMPUS", "1400 LAMMERS PIKE", "BATESVILLE", "IN", 66),
    ("ST. ANDREWS HEALTH CAMPUS", "1400 LAMMERS PIKE", "BATESVILLE", "IN", 40),
    ("APERION CARE HANOVER", "410 W LAGRANGE RD", "HANOVER", "IN", 125),
    ("APERION CARE HANOVER", "410 W LAGRANGE RD", "HANOVER", "IN", 12),
    ("RIDGEWOOD HEALTH CAMPUS", "181 CAMPUS DR", "LAWRENCEBURG", "IN", 71),
    ("RIDGEWOOD HEALTH CAMPUS", "181 CAMPUS DR", "LAWRENCEBURG", "IN", 61),
    ("HAMILTON POINTE HEALTH AND REHABILITATION CENTER", "3800 ELI PL", "NEWBURGH", "IN", 86),
    ("OSSIAN HEALTH CARE AND REHABILITATION CENTER", "215 DAVIS RD", "OSSIAN", "IN", 100),
    ("OSSIAN HEALTH AND REHABILITATION CENTER", "215 DAVIS RD", "OSSIAN", "IN", 62),
    ("RICHMOND HILL", "95 RICHMOND HILL RD", "ASHEVILLE", "NC", 46),
    ("AUTUMN VILLAGE 2C", "746 MCDONOUGH ROAD", "JACKSON", "GA", 15),
    ("ASPEN PLACE HEALTH CAMPUS", "2320 N MONTGOMERY RD", "GREENSBURG", "IN", 39),
    ("CENTURY VILLA HEALTH CARE", "705 N MERIDIAN ST", "GREENTOWN", "IN", 84),  # disambiguate: ALF, unserved
    ("BROOKSIDE VILLAGE", "1111 CHURCH AVE", "JASPER", "IN", 78),
    ("THE GLEBE", "200 THE GLEBE BLVD", "DALEVILLE", "VA", 72),
    ("WOODLAND TERRACE", "300 KILDAIRE WOODS DR", "CARY", "NC", 84),
    ("HAMILTON GROVE", "31869 CHICAGO TRAIL", "NEW CARLISLE", "IN", 85),
    ("FIVE STAR RESIDENCES OF NOBLESVILLE", "7235 RIVERWALK WAY N", "NOBLESVILLE", "IN", 151),
    ("CROSSING AT NOBLESVILLE", "7235 RIVERWALK WAY N", "NOBLESVILLE", "IN", 75),
    ("WATERS OF WAKARUSA SKILLED NURSING FACILITY, THE", "300 N WASHINGTON ST", "WAKARUSA", "IN", 133),
    ("WATERS OF WAKARUSA", "300 N WASHINGTON ST", "WAKARUSA", "IN", 61),
    ("THE VILLAGE OF LEBANON", "105 VILLAGE WAY", "LEBANON", "KY", 29),
    ("MAPLE GROVE SENIOR LIVING LLC", "711 FRANKFORT ROAD", "SHELBYVILLE", "KY", 117),
    ("PILLARS ASSISTED LIVING COMMUNITY", "711 FRANKFORT RD", "SHELBYVILLE", "KY", 21),
    ("ABSOLUTE CARE ASSISTED LIVING", "431 JUNNY RD", "ANGIER", "NC", 12),
    ("ABSOLUTE CARE ASSISTED LIVING II", "431 JUNNY RD", "ANGIER", "NC", 12),
    # Cluster 27: delete the WILHAM RIDGE with wrong ZIP (28805)
    # Disambiguated by ZIP since two rows share same name/addr/beds
    ("WILHAM RIDGE", "30 DALEA DR", "ASHEVILLE", "NC", 54),  # ZIP=28805 version
    # Punchlist #3
    ("CHERRY POINT BAY NURSING AND REHAB CENTER", "110 MCCOTTER BLVD", "HAVELOCK", "NC", 47),
    ("CHERRY POINT BAY NURSING AND REHABILITATION CENTER", "110 MCCOTTER BOULEVARD", "HAVELOCK", "NC", 80),
    ("PENDER MEMORIAL HOSP SNF", "507 E FREMONT STREET", "BURGAW", "NC", 53),
    ("RICHMOND PINES HEALTHCARE AND REHABILITATION CENTE", "HIGHWAY 177 S BOX 1489", "HAMLET", "NC", 125),
    ("MOUNTAIN VIEW CARE AND REHABILITATION CENTER", "2309 STAFFORD AVENUE", "SCRANTON", "PA", 180),
    ("MOUNTAIN VIEW CARE AND REHABILITATION CENTER", "2309 STAFFORD AVENUE", "SCRANTON", "PA", 204),
    ("OAK GLEN HEALTHCARE AND REHABILITATION CENTER", "3201 RIVER ROAD", "LEWISBURG", "PA", 251),
    # Clusters 28-56
    ("ARBOR TERRACE OF ASHEVILLE", "3199 SWEETEN CREEK RD", "ASHEVILLE", "NC", 70),
    ("MARJORIE MCCUNE MEMORIAL CENTER", "101 LIONS WAY", "BLACK MOUNTAIN", "NC", 64),
    ("ALAMANCE HOUSE", "2766 GRAND OAKS BLVD", "BURLINGTON", "NC", 94),
    ("CAROL WOODS RETIREMENT COMMUNITY", "750 WEAVER DAIRY RD", "CHAPEL HILL", "NC", 465),
    ("ALDERSGATE RETIREMENT COMMUNITY", "3800 SHAMROCK DR", "CHARLOTTE", "NC", 577),
    ("BROOKDALE CHARLOTTE EAST", "6053 WILORA LAKE RD", "CHARLOTTE", "NC", 50),
    ("CAROLINA RESERVE OF DURHAM", "4523 HOPE VALLEY RD", "DURHAM", "NC", 60),
    ("BECKY'S REST HOME #1", "316 LOWER BRUSH CREEK RD", "FLETCHER", "NC", 15),
    ("BECKY'S REST HOME #2", "316 LOWER BRUSH CREEK RD", "FLETCHER", "NC", 15),
    ("CADENCE GARNER", "200 MINGLEWOOD DR", "GARNER", "NC", 84),
    ("WHITESTONE MASONIC AND EASTERN STAR COMMUNITY", "700 S HOLDEN RD", "GREENSBORO", "NC", 335),
    ("THE ARBORETUM AT HERITAGE GREENS", "709 MEADOWOOD ST", "GREENSBORO", "NC", 66),
    ("VERRA SPRINGS AT HERITAGE GREENS", "803 MEADOWOOD ST", "GREENSBORO", "NC", 45),
    ("GRACELAND LIVING CENTER I", "1290 DENNY RD", "KING", "NC", 12),
    ("GRACELAND LIVING CENTER II", "1290 DENNY RD", "KING", "NC", 11),
    ("TERRABELLA KNIGHTDALE", "2408 HODGE RD", "KNIGHTDALE", "NC", 96),
    ("TERRABELLA KNIGHTDALE", "2408 HODGE RD", "KNIGHTDALE", "NC", 51),
    ("MONROE MANOR ASSISTED LIVING BUILDING I", "1101 BAUCOM RD", "MONROE", "NC", 12),
    ("MONROE MANOR ASSISTED LIVING BUILDING II", "1101 BAUCOM RD", "MONROE", "NC", 12),
    ("SCARLET OAKS RETIREMENT COMMUNITY", "440 LAFAYETTE AVE", "CINCINNATI", "OH", 112),
    ("BUCKEYE FOREST AT FAIRFIELD", "3801 WOODRIDGE BLVD", "FAIRFIELD", "OH", 43),
    ("FOX RUN OF FINDLAY ASSISTED LIVING", "600 FOX RUN RD", "FINDLAY", "OH", 132),
    ("WOODLANDS OF FINDLAY", "600 FOX RUN RD", "FINDLAY", "OH", 133),
    ("AYDEN HEALTHCARE OF JACKSON", "8668 STATE ROUTE 93", "JACKSON", "OH", 82),
    ("OTTERBEIN LEBANON SENIORLIFE COMMUNITY", "585 OH-741", "LEBANON", "OH", 550),
    ("OAK GROVE MANOR", "1670 CRIDER RD", "MANSFIELD", "OH", 75),
    ("THE DISTRICT HOME", "8605 CENTREVILLE RD", "MANASSAS", "VA", 72),
]

# Special disambiguation for WILHAM RIDGE delete (two identical rows except ZIP)
WILHAM_RIDGE_DELETE_ZIP = "28805"  # Delete the one with wrong ZIP

# TYPE UPDATES: ALF -> SNF
# (name, address, city, state)
TYPE_ALF_TO_SNF = [
    ("OTTERBEIN FRANKLIN - SNF", "1070 W JEFFERSON ST", "FRANKLIN", "IN"),
    ("ST ANDREWS HEALTH CAMPUS - BATESVILLE SNF", "1400 LAMMERS PIKE", "BATESVILLE", "IN"),
    ("APERION CARE HANOVER SNF", "410 W LAGRANGE RD", "HANOVER", "IN"),
    ("RIDGEWOOD HEALTH SNF", "181 CAMPUS DR", "LAWRENCEBURG", "IN"),
    ("HAMILTON POINTE HEALTH AND REHAB - SNF", "3800 ELI PL", "NEWBURGH", "IN"),
    ("OSSIAN HEALTHCARE SNF", "215 DAVIS RD", "OSSIAN", "IN"),
    ("ASPEN PLACE HEALTH CAMPUS SNF", "2320 N MONTGOMERY RD", "GREENSBURG", "IN"),
    ("LLV - SOUTH ANTHONY SNF", "6723 S ANTHONY BLVD", "FORT WAYNE", "IN"),
    ("ELWOOD SNF", "2300 PARKVIEW LN", "ELWOOD", "IN"),
    ("GREENWOOD VILLAGE SOUTH - PAVILION", "295 VILLAGE LN", "GREENWOOD", "IN"),
    ("WINDSOR POINT SNF", "1221 BROAD ST", "FUQUAY-VARINA", "NC"),
    ("GLEBE-SNF/NF", "200 THE GLEBE BLVD", "DALEVILLE", "VA"),
    ("WARM HEARTH-THE COVE-SNF/NF", "2387 WARM HEARTH DR", "BLACKSBURG", "VA"),
    ("HAMILTON GROVE SNF", "31869 CHICAGO TRAIL", "NEW CARLISLE", "IN"),
    ("THE WATERS OF WAKARUSA SNF", "300 N WASHINGTON ST", "WAKARUSA", "IN"),
    ("MAPLE GROVE SENIOR LIVING", "711 FRANKFORT RD", "SHELBYVILLE", "KY"),
    ("WHITESTONE SNF", "700 S HOLDEN RD", "GREENSBORO", "NC"),
    ("SCARLET OAKS NURSING AND REHAB SNF", "440 LAFAYETTE AVE", "CINCINNATI", "OH"),
    ("OTTERBEIN LEBANON SNF", "585 OH-741", "LEBANON", "OH"),
    ("BRISTOL VILLAGE SNF", "444 CHERRY ST", "WAVERLY", "OH"),
    ("BRANDON OAKS NURSING AND REHABILITATION CENTER SNF", "3837 BRANDON AVE SW", "ROANOKE", "VA"),
]

# TYPE UPDATES: ALF -> ILF
TYPE_ALF_TO_ILF = [
    ("WOODLAND TERRACE IN", "300 KILDAIRE WOODS DR", "CARY", "NC"),
    ("WINDSOR POINT IN", "1221 BROAD ST", "FUQUAY-VARINA", "NC"),
    ("MARJORIE MCCUNE MEMORIAL IN", "101 LIONS WAY", "BLACK MOUNTAIN", "NC"),
    ("WALTONWOOD CARY IN", "750 SE CARY PKWY", "CARY", "NC"),
    ("WALTONWOOD PROVIDENCE IN", "11945 PROVIDENCE RD.", "CHARLOTTE", "NC"),
    ("BROOKDALE CHARLOTTE EAST IN", "6053 WILORA LAKE RD", "CHARLOTTE", "NC"),
    ("HERITAGE GREENS IN", "709 MEADOWOOD ST", "GREENSBORO", "NC"),
    ("THE CARDINAL AT NORTH HILLS IN", "4030 CARDINAL AT N HILLS ST", "RALEIGH", "NC"),
    ("SUNNYSIDE ILF", "3935 SUNNYSIDE DR", "HARRISONBURG", "VA"),
]

# BED COUNT UPDATES: (name, address, city, state, old_beds, new_beds)
BED_UPDATES = [
    ("MAPLE GROVE SENIOR LIVING", "711 FRANKFORT RD", "SHELBYVILLE", "KY", 83, 117),
]

# NAME UPDATES: (old_name, new_name, address, city, state)
# Using address/city/state to disambiguate
NAME_UPDATES = [
    # Cluster 27: rename the 54-bed WILHAM RIDGE (ZIP=28801) to Chase Samaritan
    ("WILHAM RIDGE", "CHASE SAMARITAN ASSISTED LIVING", "30 DALEA DR", "ASHEVILLE", "NC"),
    # Backslash cleanups (Cluster 9)
    ("SHULER HEALTH CARE/CRANE VILLA", "SHULER HEALTH CARE CRANE VILLA", None, "KERNERSVILLE", "NC"),
    ("SHULER HEALTH CARE/PHILLIPS VILLA", "SHULER HEALTH CARE PHILLIPS VILLA", None, "KERNERSVILLE", "NC"),
    ("SHULER HEALTH CARE/PIERCE VILLA", "SHULER HEALTH CARE PIERCE VILLA", None, "KERNERSVILLE", "NC"),
    ("SHULER HEALTH CARE/RECORD VILLA", "SHULER HEALTH CARE RECORD VILLA", None, "KERNERSVILLE", "NC"),
    ("SHULER HEALTH CARE/STOREY VILLA", "SHULER HEALTH CARE STOREY VILLA", None, "KERNERSVILLE", "NC"),
]

# FIELD UPDATES (Punchlist #3): (name, address, city, state, field, old_value, new_value)
FIELD_UPDATES = [
    ("CHERRY POINT BAY NURSING AND REHABILITATION CENTER", "110 MCCOTTER BOULEVARD", "HAVELOCK", "NC",
     "MH_Flag", None, "Yes"),
    ("CHERRY POINT BAY NURSING AND REHABILITATION CENTER", "110 MCCOTTER BOULEVARD", "HAVELOCK", "NC",
     "Integrated_Flag", "Yes", None),
    ("OAK GLEN HEALTHCARE AND REHABILITATION CENTER", "3201 RIVER ROAD", "LEWISBURG", "PA",
     "Address", "3201 RIVER ROAD", "15 Ridgecrest Cir"),
    ("OAK GLEN HEALTHCARE AND REHABILITATION CENTER", None, None, "PA",
     "Corporate_Name", None, "OAK GLEN HEALTHCARE AND REHABILITATION CENTER LLC"),
]

# ADDRESS/NAME CORRECTIONS
# (facility_name_or_partial, address_match, city, state, field, old_value, new_value)
ADDR_NAME_CORRECTIONS = [
    ("AUTUMN VILLAGE 2A", "746 MCDONOUGH ROAD", "JACKSON", "GA", "Address", "746 MCDONOUGH ROAD", "753 Covington Street"),
    ("AUTUMN VILLAGE 2A", "746 MCDONOUGH ROAD", "JACKSON", "GA", "Facility_Name", "AUTUMN VILLAGE 2A", "Autumn Village 1"),
    ("AUTUMN VILLAGE 2B", "746 MCDONOUGH ROAD", "JACKSON", "GA", "Facility_Name", "AUTUMN VILLAGE 2B", "Autumn Village 2"),
]

# COUNTY UPDATES: (name, address, city, state, county)
COUNTY_UPDATES = [
    ("HAMILTON GROVE ALF", "31869 CHICAGO TRAIL", "NEW CARLISLE", "IN", "St. Joseph"),
    ("THE WATERS OF WAKARUSA AL", "300 N. WASHINGTON ST", "WAKARUSA", "IN", "Elkhart"),
]


def find_row(all_data, name, address, city, state, beds=None, zip_code=None,
             source_type=None, exclude_rows=None):
    """Find a row by content matching. Returns list of (excel_row, row_data) tuples."""
    matches = []
    n_name = norm(name)
    n_addr = norm_addr(address) if address else None
    n_city = norm(city) if city else None
    n_state = norm(state) if state else None

    for excel_row, row_data in all_data.items():
        if exclude_rows and excel_row in exclude_rows:
            continue

        r_name = norm(row_data[COL['Facility_Name']])
        r_addr = norm_addr(row_data[COL['Address']])
        r_city = norm(row_data[COL['City']])
        r_state = norm(row_data[COL['State']])
        r_beds = row_data[COL['Total_Beds']]
        r_zip = str(row_data[COL['ZIP']] or '').strip()
        r_type = norm(row_data[COL['Source_Type']])

        # Name must match
        if r_name != n_name:
            continue
        # Address match (if specified)
        if n_addr and r_addr != n_addr:
            continue
        # City match (if specified)
        if n_city and r_city != n_city:
            continue
        # State match (if specified)
        if n_state and r_state != n_state:
            continue
        # Beds match (if specified)
        if beds is not None and r_beds != beds:
            continue
        # ZIP match (if specified)
        if zip_code is not None and r_zip != str(zip_code):
            continue
        # Source type match (if specified)
        if source_type is not None and r_type != norm(source_type):
            continue

        matches.append((excel_row, row_data))

    return matches


def main():
    print("=" * 70)
    print("V24 MIGRATION — Content-Based Matching")
    print("=" * 70)

    # Phase 1: Load all data
    print("\n[Phase 1] Loading V23...")
    wb = openpyxl.load_workbook(V23_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    all_data = {}  # excel_row -> tuple of values
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        all_data[i] = list(row)
    wb.close()
    print(f"  Loaded {len(all_data)} rows, {len(headers)} columns")

    # Track all changes
    rows_to_delete = set()
    updates = {}  # excel_row -> dict of {col_index: new_value}
    errors = []
    matched = []

    # Phase 2: Match all fixes
    print(f"\n[Phase 2] Matching fixes by content...")

    # --- DELETES ---
    print(f"\n  DELETES ({len(DELETES)} entries):")
    for name, address, city, state, beds in DELETES:
        # Special case: WILHAM RIDGE at 30 Dalea Dr — disambiguate by ZIP
        if name == "WILHAM RIDGE" and "DALEA" in address.upper():
            matches = find_row(all_data, name, address, city, state, beds=beds,
                               zip_code=WILHAM_RIDGE_DELETE_ZIP, exclude_rows=rows_to_delete)
        # Special case: CENTURY VILLA — disambiguate by source_type=ALF (delete the ALF, keep the SNF)
        elif name == "CENTURY VILLA HEALTH CARE" and city == "GREENTOWN":
            matches = find_row(all_data, name, address, city, state, beds=beds,
                               source_type="ALF", exclude_rows=rows_to_delete)
        else:
            matches = find_row(all_data, name, address, city, state, beds=beds,
                               exclude_rows=rows_to_delete)

        if len(matches) == 1:
            excel_row = matches[0][0]
            rows_to_delete.add(excel_row)
            matched.append(f"  DEL row {excel_row}: {name} ({city}, {state}) [{beds} beds]")
        elif len(matches) == 0:
            errors.append(f"  DEL NO MATCH: {name} | {address} | {city} | {state} | {beds} beds")
        else:
            # Multiple matches — try to pick one not already marked for deletion
            remaining = [(r, d) for r, d in matches if r not in rows_to_delete]
            if len(remaining) == 1:
                excel_row = remaining[0][0]
                rows_to_delete.add(excel_row)
                matched.append(f"  DEL row {excel_row}: {name} ({city}, {state}) [{beds} beds] (disambiguated)")
            else:
                errors.append(f"  DEL AMBIGUOUS ({len(matches)} matches): {name} | {address} | {city} | {state} | {beds} beds")
                for r, d in matches:
                    errors.append(f"    -> row {r}: ZIP={d[COL['ZIP']]} served={d[COL['Do_We_Serve']]}")

    print(f"    Matched: {len(rows_to_delete)}, Errors: {len(errors)}")

    # --- TYPE UPDATES: ALF -> SNF ---
    print(f"\n  TYPE ALF->SNF ({len(TYPE_ALF_TO_SNF)} entries):")
    type_snf_count = 0
    for name, address, city, state in TYPE_ALF_TO_SNF:
        matches = find_row(all_data, name, address, city, state, exclude_rows=rows_to_delete)
        if len(matches) == 1:
            excel_row = matches[0][0]
            current_type = matches[0][1][COL['Source_Type']]
            if norm(current_type) != 'ALF':
                errors.append(f"  TYPE SNF WRONG CURRENT: row {excel_row} {name} is '{current_type}', expected ALF")
            else:
                updates.setdefault(excel_row, {})[COL['Source_Type']] = 'SNF'
                type_snf_count += 1
                matched.append(f"  TYPE row {excel_row}: {name} ALF->SNF")
        elif len(matches) == 0:
            errors.append(f"  TYPE SNF NO MATCH: {name} | {address} | {city} | {state}")
        else:
            errors.append(f"  TYPE SNF AMBIGUOUS ({len(matches)}): {name} | {address} | {city} | {state}")
    print(f"    Matched: {type_snf_count}, Errors: {len(errors) - len(rows_to_delete) + type_snf_count if False else 'see below'}")

    # --- TYPE UPDATES: ALF -> ILF ---
    print(f"\n  TYPE ALF->ILF ({len(TYPE_ALF_TO_ILF)} entries):")
    type_ilf_count = 0
    for name, address, city, state in TYPE_ALF_TO_ILF:
        matches = find_row(all_data, name, address, city, state, exclude_rows=rows_to_delete)
        if len(matches) == 1:
            excel_row = matches[0][0]
            current_type = matches[0][1][COL['Source_Type']]
            if norm(current_type) != 'ALF':
                errors.append(f"  TYPE ILF WRONG CURRENT: row {excel_row} {name} is '{current_type}', expected ALF")
            else:
                updates.setdefault(excel_row, {})[COL['Source_Type']] = 'ILF'
                type_ilf_count += 1
                matched.append(f"  TYPE row {excel_row}: {name} ALF->ILF")
        elif len(matches) == 0:
            errors.append(f"  TYPE ILF NO MATCH: {name} | {address} | {city} | {state}")
        else:
            errors.append(f"  TYPE ILF AMBIGUOUS ({len(matches)}): {name} | {address} | {city} | {state}")

    # --- BED COUNT UPDATES ---
    print(f"\n  BED UPDATES ({len(BED_UPDATES)} entries):")
    bed_count = 0
    for name, address, city, state, old_beds, new_beds in BED_UPDATES:
        matches = find_row(all_data, name, address, city, state, beds=old_beds, exclude_rows=rows_to_delete)
        if len(matches) == 1:
            excel_row = matches[0][0]
            updates.setdefault(excel_row, {})[COL['Total_Beds']] = new_beds
            bed_count += 1
            matched.append(f"  BEDS row {excel_row}: {name} {old_beds}->{new_beds}")
        elif len(matches) == 0:
            errors.append(f"  BEDS NO MATCH: {name} | {address} | {city} | {state} | old={old_beds}")
        else:
            errors.append(f"  BEDS AMBIGUOUS ({len(matches)}): {name}")

    # --- NAME UPDATES ---
    print(f"\n  NAME UPDATES ({len(NAME_UPDATES)} entries):")
    name_count = 0
    for old_name, new_name, address, city, state in NAME_UPDATES:
        # For WILHAM RIDGE rename: need to exclude the row being deleted AND the
        # row being kept (WILHAM RIDGE ASSISTED LIVING, 11 beds)
        # The rename target is the remaining WILHAM RIDGE with 54 beds and ZIP=28801
        if old_name == "WILHAM RIDGE" and address and "DALEA" in address.upper():
            matches = find_row(all_data, old_name, address, city, state, beds=54,
                               zip_code="28801", exclude_rows=rows_to_delete)
        else:
            matches = find_row(all_data, old_name, address, city, state, exclude_rows=rows_to_delete)

        if len(matches) == 1:
            excel_row = matches[0][0]
            updates.setdefault(excel_row, {})[COL['Facility_Name']] = new_name
            name_count += 1
            matched.append(f"  NAME row {excel_row}: '{old_name}' -> '{new_name}'")
        elif len(matches) == 0:
            errors.append(f"  NAME NO MATCH: '{old_name}' | {address} | {city} | {state}")
        else:
            errors.append(f"  NAME AMBIGUOUS ({len(matches)}): '{old_name}' | {address} | {city} | {state}")
            for r, d in matches:
                errors.append(f"    -> row {r}: beds={d[COL['Total_Beds']]} ZIP={d[COL['ZIP']]}")

    # --- FIELD UPDATES (Punchlist #3) ---
    print(f"\n  FIELD UPDATES ({len(FIELD_UPDATES)} entries):")
    field_count = 0
    for name, address, city, state, field, old_val, new_val in FIELD_UPDATES:
        # For OAK GLEN corporate update, the address may have already been changed
        # by a prior field update in same batch. Match by name+state only if address is None.
        if address:
            matches = find_row(all_data, name, address, city, state, exclude_rows=rows_to_delete)
        else:
            matches = find_row(all_data, name, None, None, state, exclude_rows=rows_to_delete)

        if len(matches) == 1:
            excel_row = matches[0][0]
            col_idx = COL[field]
            current_val = matches[0][1][col_idx]

            # Verify old value matches (loose comparison)
            if old_val is not None and norm(current_val) != norm(old_val):
                errors.append(f"  FIELD OLD MISMATCH: row {excel_row} {name}.{field} = '{current_val}', expected '{old_val}'")
            else:
                updates.setdefault(excel_row, {})[col_idx] = new_val
                field_count += 1
                matched.append(f"  FIELD row {excel_row}: {name}.{field} = '{old_val}' -> '{new_val}'")
        elif len(matches) == 0:
            errors.append(f"  FIELD NO MATCH: {name} | {address} | {city} | {state} | {field}")
        else:
            # For OAK GLEN with no address filter, try disambiguating by address
            if address is None:
                # Try with the original address from the DB
                for r, d in matches:
                    addr_val = norm(d[COL['Address']])
                    if '3201' in addr_val or 'RIVER' in addr_val:
                        updates.setdefault(r, {})[COL[field]] = new_val
                        field_count += 1
                        matched.append(f"  FIELD row {r}: {name}.{field} -> '{new_val}' (disambiguated by addr)")
                        break
                else:
                    errors.append(f"  FIELD AMBIGUOUS ({len(matches)}): {name} | {field}")

    # --- ADDRESS/NAME CORRECTIONS ---
    print(f"\n  ADDRESS/NAME CORRECTIONS ({len(ADDR_NAME_CORRECTIONS)} entries):")
    corr_count = 0
    for name, address, city, state, field, old_val, new_val in ADDR_NAME_CORRECTIONS:
        # Match by old facility name and old address
        matches = find_row(all_data, name, address, city, state, exclude_rows=rows_to_delete)
        if len(matches) == 0:
            # Try case-insensitive partial match
            matches = find_row(all_data, name, None, city, state, exclude_rows=rows_to_delete)

        if len(matches) == 1:
            excel_row = matches[0][0]
            col_idx = COL[field]
            updates.setdefault(excel_row, {})[col_idx] = new_val
            corr_count += 1
            matched.append(f"  CORR row {excel_row}: {name}.{field} -> '{new_val}'")
        elif len(matches) == 0:
            errors.append(f"  CORR NO MATCH: {name} | {address} | {city} | {state}")
        else:
            errors.append(f"  CORR AMBIGUOUS ({len(matches)}): {name} | {city} | {state}")

    # --- COUNTY UPDATES ---
    print(f"\n  COUNTY UPDATES ({len(COUNTY_UPDATES)} entries):")
    county_count = 0
    for name, address, city, state, county in COUNTY_UPDATES:
        matches = find_row(all_data, name, address, city, state, exclude_rows=rows_to_delete)
        if len(matches) == 1:
            excel_row = matches[0][0]
            current_county = matches[0][1][COL['County']]
            if current_county is not None and str(current_county).strip() != '':
                errors.append(f"  COUNTY NOT BLANK: row {excel_row} {name} county='{current_county}', expected blank")
            else:
                updates.setdefault(excel_row, {})[COL['County']] = county
                county_count += 1
                matched.append(f"  COUNTY row {excel_row}: {name} -> '{county}'")
        elif len(matches) == 0:
            errors.append(f"  COUNTY NO MATCH: {name} | {address} | {city} | {state}")
        else:
            errors.append(f"  COUNTY AMBIGUOUS ({len(matches)}): {name}")

    # Phase 3: Report
    print("\n" + "=" * 70)
    print("MATCHING RESULTS")
    print("=" * 70)
    print(f"\n  Deletes matched: {len(rows_to_delete)}")
    print(f"  Type ALF->SNF matched: {type_snf_count}")
    print(f"  Type ALF->ILF matched: {type_ilf_count}")
    print(f"  Bed updates matched: {bed_count}")
    print(f"  Name updates matched: {name_count}")
    print(f"  Field updates matched: {field_count}")
    print(f"  Address/Name corrections matched: {corr_count}")
    print(f"  County updates matched: {county_count}")
    print(f"\n  Total rows to update: {len(updates)}")
    print(f"  Total rows to delete: {len(rows_to_delete)}")

    if errors:
        print(f"\n  *** ERRORS ({len(errors)}) ***")
        for e in errors:
            print(f"  {e}")
        print("\nAborting — fix errors before proceeding.")
        sys.exit(1)

    # Print all matched operations for verification
    print("\n  ALL MATCHED OPERATIONS:")
    for m in sorted(matched):
        print(f"    {m}")

    # Phase 4: Apply fixes
    print(f"\n[Phase 4] Applying fixes...")

    # Apply updates to in-memory data
    for excel_row, col_updates in updates.items():
        for col_idx, new_val in col_updates.items():
            all_data[excel_row][col_idx] = new_val

    # Phase 5: Write V24
    print(f"\n[Phase 5] Writing V24...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws_out.cell(row=1, column=col_idx, value=header)

    # Write data rows, skipping deleted rows
    out_row = 2
    deleted_count = 0
    kept_count = 0
    for excel_row in sorted(all_data.keys()):
        if excel_row in rows_to_delete:
            deleted_count += 1
            continue
        row_data = all_data[excel_row]
        for col_idx, value in enumerate(row_data, start=1):
            ws_out.cell(row=out_row, column=col_idx, value=value)
        out_row += 1
        kept_count += 1

    wb_out.save(V24_PATH)
    wb_out.close()

    print(f"\n  V24 saved to: {V24_PATH}")
    print(f"  Rows deleted: {deleted_count}")
    print(f"  Rows kept: {kept_count}")
    print(f"  Expected: {len(all_data) - len(rows_to_delete)} rows")

    # Sanity check
    expected = 25672 - len(rows_to_delete)
    if kept_count != expected:
        print(f"  *** WARNING: Expected {expected} rows but wrote {kept_count} ***")
    else:
        print(f"  Row count verified: {kept_count}")

    print("\nDone.")


if __name__ == "__main__":
    main()
