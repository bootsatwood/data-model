#!/usr/bin/env python3
"""
Combined Fix — V22.6 -> V22.7

Applies:
  1. Liberty QC corrections (9 deletions + 1 service flag fix)
  2. Singh/Waltonwood corporate consolidation (renames + 11 dupe removals)
  3. Cedarhurst typo fix (SENOR -> SENIOR + 3 dupe removals)
  4. Spring Arbor consolidation (renames + 5 dupe removals)

Usage:
  python v22_7_fix.py preview
  python v22_7_fix.py apply
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from utils import safe, load_db, VAULT

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_6.xlsx"
DB_OUTPUT = VAULT / "Current" / "1_Combined_Database_FINAL_V22_7.xlsx"

# ---------------------------------------------------------------------------
# Corporate Name Renames
# ---------------------------------------------------------------------------

CORP_RENAMES = {
    # Singh/Waltonwood → SINGH (GLR canonical)
    "Singh Developments (Waltonwood)": "SINGH",
    "Waltonwood": "SINGH",
    "Waltonwood Senior Living": "SINGH",
    "WALTONWOOD AT CARY, LLC": "SINGH",
    # Cedarhurst typo
    "CEDARHURST SENOR LIVING": "CEDARHURST SENIOR LIVING",
    # Spring Arbor → SPRING ARBOR MANAGEMENT (GLR canonical)
    "Spring Arbor Senior Living": "SPRING ARBOR MANAGEMENT",
    "SPRING ARBOR KILL DEVIL HILLS NC TENANT, LLC": "SPRING ARBOR MANAGEMENT",
    "SPRING ARBOR COTTAGE OF FREDERICKSB": "SPRING ARBOR MANAGEMENT",
}

# Row 25975 has Corporate_Name="unknown" but is Waltonwood At Lakeside — being deleted

# ---------------------------------------------------------------------------
# Rows to Delete (by current V22.6 row numbers)
# ---------------------------------------------------------------------------

ROWS_TO_DELETE = {
    # --- Liberty CMS SNF duplicates (6) ---
    # Keep lower row (Row A), delete higher row (Row B)
    20080,  # Swift Creek Health Center dupe (of 9612)
    18953,  # Liberty HC Svcs of Golden Years dupe (of 9261)
    20844,  # Woodlands N&R Center dupe (of 9964)
    20836,  # Woodhaven Nursing Center dupe (of 9963)
    20203,  # The Foley Center at Chestnut Ridge dupe (of 9717)
    17642,  # Bradley Creek Health Center dupe (of 8747)

    # --- Liberty Pisgah Manor duplicate (1) ---
    19513,  # Pisgah Manor 150 beds dupe (keep 9448 with 103/118 beds)

    # --- Liberty phantoms (2) ---
    9287,   # Liberty Commons Alamance ALF (phantom, SNF row 9259 retained)
    9637,   # Silver Bluff Village-SNF (phantom ALF, SNF 9595 + ALF 9636 retained)

    # --- Singh/Waltonwood duplicates (11) ---
    9974,   # Waltonwood Cary Parkway (LEGACY dupe of GLR 9971-9973)
    9980,   # Waltonwood At Providence (GLR dupe of 9977-9979)
    20679,  # Waltonwood Ashburn (NIC-B dupe of 20678)
    20680,  # Waltonwood At Cotswold (NIC-B dupe of 9975)
    20681,  # Waltonwood Lake Boone (NIC-B dupe of 9976)
    25971,  # Waltonwood at Carriage Park (LEGACY dupe, wrong address)
    25974,  # Waltonwood At Cherry Hill II (LEGACY dupe)
    25975,  # Waltonwood At Lakeside (LEGACY dupe, unknown corp)
    25977,  # Waltonwood At Main (LEGACY dupe)
    25979,  # Waltonwood At Twelve Oaks II (LEGACY dupe)
    25981,  # Waltonwood At University II (LEGACY dupe)

    # --- Cedarhurst duplicates (3) ---
    5761,   # Cedarhurst Of Beaumont (dupe of 5760, Lexington KY)
    17853,  # Cedar Creek Of Bloomington (dupe of 4671, Bloomington IN)
    17856,  # Cedar Creek Of Marion (dupe of 4676, Marion IN)

    # --- Spring Arbor duplicates (5) ---
    9660,   # Spring Arbor Of Albemarle (dupe of 9658)
    9662,   # Spring Arbor Of Apex (dupe of 9661)
    9663,   # Spring Arbor Of Cary (dupe of GLR 9664-9665)
    9666,   # Spring Arbor Of Greensboro (dupe of 9659)
    19934,  # Spring Arbor Of Cary (GLR dupe of 9664-9665)
}

# ---------------------------------------------------------------------------
# Service Flag Fixes
# ---------------------------------------------------------------------------

# Woodlands N&R Center, Fayetteville NC — incorrectly marked served
WOODLANDS_FIX = {
    "facility": "WOODLANDS",
    "city": "FAYETTEVILLE",
    "state": "NC",
    "fixes": {
        "Do_We_Serve": "No",
    },
}

# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    if not corp_name or corp_name.upper() in ('UNKNOWN', ''):
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 0) >= 2:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print("Usage: python v22_7_fix.py [preview|apply]")
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Combined Fix — V22.6 -> V22.7 ({mode} mode)")
    print("=" * 65)

    print(f"\nLoading {DB_SOURCE.name}...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    # --- Phase 1: Corporate Name Renames ---
    print("\nPhase 1: Corporate Name Renames")
    print("-" * 40)

    rename_counts = Counter()
    for r in rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp in CORP_RENAMES:
            new_corp = CORP_RENAMES[corp]
            rename_counts[f"{corp} -> {new_corp}"] += 1
            r['Corporate_Name'] = new_corp

    for key in sorted(rename_counts.keys()):
        print(f"  {rename_counts[key]:>4d}  {key}")
    print(f"  {sum(rename_counts.values())} total renames")

    # --- Phase 2: Service Flag Fixes ---
    print(f"\nPhase 2: Service Flag Fixes")
    print("-" * 40)

    flag_fixes = 0
    for r in rows:
        if r['_excel_row'] in ROWS_TO_DELETE:
            continue
        fac = safe(r.get('Facility_Name', '')).upper()
        city = safe(r.get('City', '')).upper()
        state = safe(r.get('State', ''))
        if (WOODLANDS_FIX["facility"] in fac
                and city == WOODLANDS_FIX["city"]
                and state == WOODLANDS_FIX["state"]):
            old_served = safe(r.get('Do_We_Serve', ''))
            if old_served == 'Yes':
                for col, new_val in WOODLANDS_FIX["fixes"].items():
                    old_val = safe(r.get(col, ''))
                    print(f"  Row {r['_excel_row']}: {safe(r.get('Facility_Name',''))}")
                    print(f"    {col}: {old_val} -> {new_val}")
                    r[col] = new_val
                flag_fixes += 1

    # Also clear MH flag columns
    for r in rows:
        if r['_excel_row'] in ROWS_TO_DELETE:
            continue
        fac = safe(r.get('Facility_Name', '')).upper()
        city = safe(r.get('City', '')).upper()
        state = safe(r.get('State', ''))
        if (WOODLANDS_FIX["facility"] in fac
                and city == WOODLANDS_FIX["city"]
                and state == WOODLANDS_FIX["state"]
                and safe(r.get('Do_We_Serve', '')) == 'No'):
            # Find and clear MH flag
            for h in headers:
                if h.upper() in ('MH', 'MH_FLAG'):
                    old_val = safe(r.get(h, ''))
                    if old_val and old_val != 'No':
                        print(f"    {h}: {old_val} -> No")
                        r[h] = 'No'

    print(f"  {flag_fixes} service flag corrections")

    # --- Phase 3: Verify Deletions ---
    print(f"\nPhase 3: Row Deletions ({len(ROWS_TO_DELETE)} rows)")
    print("-" * 40)

    verified = []
    for excel_row in sorted(ROWS_TO_DELETE):
        found = False
        for r in rows:
            if r['_excel_row'] == excel_row:
                fac = safe(r.get('Facility_Name', ''))
                city = safe(r.get('City', ''))
                state = safe(r.get('State', ''))
                served = safe(r.get('Do_We_Serve', ''))
                print(f"  Row {excel_row}: {fac}, {city}, {state} (Served={served})")
                verified.append(excel_row)
                found = True
                break
        if not found:
            print(f"  WARNING: Row {excel_row} not found!")

    print(f"\n  {len(verified)} rows verified for deletion")

    # Count served rows being deleted
    served_deletes = 0
    for r in rows:
        if r['_excel_row'] in ROWS_TO_DELETE:
            if safe(r.get('Do_We_Serve', '')) == 'Yes':
                served_deletes += 1
    print(f"  {served_deletes} of those are served (service captured on retained rows)")

    # --- Phase 4: Four-Rule Reclassification (scoped) ---
    print(f"\nPhase 4: Four-Rule Ownership Reclassification")
    print("-" * 40)

    # Rebuild counts excluding deleted rows
    active_rows = [r for r in rows if r['_excel_row'] not in ROWS_TO_DELETE]
    corp_counts = Counter()
    for r in active_rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp:
            corp_counts[corp] += 1

    changed_corps = set(CORP_RENAMES.keys()) | set(CORP_RENAMES.values())
    reclass_count = 0
    reclass_detail = Counter()
    for r in active_rows:
        corp = safe(r.get('Corporate_Name', ''))
        if corp not in changed_corps:
            continue
        old_type = safe(r.get('Ownership_Type', ''))
        new_type = four_rule_classify(corp, corp_counts)
        if old_type != new_type:
            reclass_count += 1
            reclass_detail[f"{old_type} -> {new_type}"] += 1
            r['Ownership_Type'] = new_type

    print(f"  {reclass_count} reclassifications")
    for key in sorted(reclass_detail.keys()):
        print(f"    {reclass_detail[key]:>4d}  {key}")

    # --- Summary ---
    print(f"\n{'=' * 65}")
    print("SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Corporate Name renames:        {sum(rename_counts.values())}")
    print(f"  Service flag corrections:      {flag_fixes}")
    print(f"  Rows deleted:                  {len(verified)}")
    print(f"  Ownership reclassifications:   {reclass_count}")
    print(f"  Net row delta:                 -{len(verified)}")
    print(f"  Expected final row count:      {len(rows) - len(verified):,}")

    if mode == 'preview':
        print(f"\n  ** PREVIEW ONLY — no file written **")
        print(f"  Run with 'apply' to write {DB_OUTPUT.name}")
        return

    # --- Write V22.7 ---
    print(f"\nWriting {DB_OUTPUT.name}...")

    wb = load_workbook(DB_SOURCE)
    ws = wb.active

    header_row = [safe(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(header_row)}

    corp_col = col_idx.get('Corporate_Name')
    own_col = col_idx.get('Ownership_Type')
    served_col = col_idx.get('Do_We_Serve')

    # Find MH column
    mh_col = None
    for h, idx in col_idx.items():
        if h.upper() in ('MH', 'MH_FLAG'):
            mh_col = idx
            break

    # Apply renames and flag fixes
    changes_written = 0
    for r in rows:
        excel_row = r['_excel_row']
        if excel_row in ROWS_TO_DELETE:
            continue

        # Corporate_Name
        old_corp = safe(ws.cell(row=excel_row, column=corp_col).value)
        new_corp = safe(r.get('Corporate_Name', ''))
        if old_corp != new_corp:
            ws.cell(row=excel_row, column=corp_col).value = new_corp
            changes_written += 1

        # Ownership_Type
        old_own = safe(ws.cell(row=excel_row, column=own_col).value)
        new_own = safe(r.get('Ownership_Type', ''))
        if old_own != new_own:
            ws.cell(row=excel_row, column=own_col).value = new_own
            changes_written += 1

        # Do_We_Serve
        old_served = safe(ws.cell(row=excel_row, column=served_col).value)
        new_served = safe(r.get('Do_We_Serve', ''))
        if old_served != new_served:
            ws.cell(row=excel_row, column=served_col).value = new_served
            changes_written += 1

        # MH flag
        if mh_col:
            for h in headers:
                if h.upper() in ('MH', 'MH_FLAG'):
                    old_mh = safe(ws.cell(row=excel_row, column=col_idx[h]).value)
                    new_mh = safe(r.get(h, ''))
                    if old_mh != new_mh:
                        ws.cell(row=excel_row, column=col_idx[h]).value = new_mh
                        changes_written += 1

    # Delete rows (bottom up)
    for excel_row in sorted(verified, reverse=True):
        ws.delete_rows(excel_row, 1)

    wb.save(DB_OUTPUT)
    print(f"  {changes_written} cells updated")
    print(f"  {len(verified)} rows deleted")
    print(f"  Saved: {DB_OUTPUT}")


if __name__ == '__main__':
    main()
