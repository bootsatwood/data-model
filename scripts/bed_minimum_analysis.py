"""Analyze impact of a 15-bed minimum rule on ER and RP scoring.
Excludes facility rows with Total_Beds <= 15 from campus counting and revenue.
Shows before/after for all scored + T4 entities."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
TIERING_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V20.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15

T5_EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
}

FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Triple Crown': {'TRIPLE CROWN'},
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE', 'WARM HEARTH'},
    'Brighton': {'BRIGHTON'},
    'Momentus Health': {'MOMENTUS HEALTH'},
}

ER_THRESHOLDS = [9, 13, 20, 40]
RP_THRESHOLDS = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

# Load data
print("Loading DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

print("Loading Finance entity list...")
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']
finance_names = set()
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in T5_EXCLUDE:
        finance_names.add(str(name).strip())

# Build lookups
all_db_names = set()
db_to_finance = {}
for fname, db_names in FINANCE_TO_DB.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_finance[dn] = fname

# Collect facility rows
print("Scanning DB...")
entity_rows = defaultdict(list)
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue
    fname = db_to_finance[corp]
    if fname not in finance_names:
        continue

    beds = ws_db.cell(r, headers['Total_Beds']).value or 0
    beds = beds if isinstance(beds, (int, float)) else 0
    addr = str(ws_db.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws_db.cell(r, headers['City']).value or '').strip().upper()
    tot_rev = ws_db.cell(r, headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0
    fac_name = str(ws_db.cell(r, headers['Facility_Name']).value or '').strip()
    source = str(ws_db.cell(r, headers['Source_Type']).value or '').strip()

    entity_rows[fname].append({
        'fac_name': fac_name,
        'campus_key': (addr, city),
        'beds': beds,
        'tot_rev': tot_rev,
        'state': state,
        'source': source,
    })

# Compute before/after for each entity
print(f"\nBED MINIMUM RULE: Exclude rows with Total_Beds <= {BED_MIN}")
print(f"{'='*140}")
print(f"{'Entity':<35} {'Rows':>4} {'Drop':>4}  {'Camp':>5}{'->':>3}{'New':>5} {'ER':>3}{'->':>3}{'New':>3}  {'Revenue':>14}{'->':>3}{'New Revenue':>14} {'RP':>3}{'->':>3}{'New':>3}  {'Gate':>6} {'Notes'}")
print(f"{'-'*140}")

changes = []
gate_changes = []
dropped_rows_detail = []

for fname in sorted(entity_rows.keys()):
    rows = entity_rows[fname]

    # BEFORE (all rows)
    camps_before = set(r['campus_key'] for r in rows)
    rev_before = sum(r['tot_rev'] for r in rows)
    n_camp_before = len(camps_before)
    er_before = score_dim(n_camp_before, ER_THRESHOLDS)
    rp_before = score_dim(rev_before, RP_THRESHOLDS)

    # AFTER (exclude rows with beds <= BED_MIN)
    kept = [r for r in rows if r['beds'] > BED_MIN]
    dropped = [r for r in rows if r['beds'] <= BED_MIN]
    camps_after = set(r['campus_key'] for r in kept)
    rev_after = sum(r['tot_rev'] for r in kept)
    n_camp_after = len(camps_after)
    er_after = score_dim(n_camp_after, ER_THRESHOLDS) if n_camp_after > 0 else 0
    rp_after = score_dim(rev_after, RP_THRESHOLDS) if n_camp_after > 0 else 0

    # Gate status
    gate_before = 'PASS' if n_camp_before >= 7 else 'T4'
    gate_after = 'PASS' if n_camp_after >= 7 else 'T4'
    gate_flag = ''
    if gate_before != gate_after:
        gate_flag = f'{gate_before}->{gate_after}'
        gate_changes.append(fname)
    else:
        gate_flag = gate_after

    notes = []
    if er_before != er_after:
        notes.append(f'ER {er_before}->{er_after}')
    if rp_before != rp_after:
        notes.append(f'RP {rp_before}->{rp_after}')
    if len(dropped) > 0:
        notes.append(f'{len(dropped)} rows dropped')

    er_arrow = '->' if er_before != er_after else '  '
    rp_arrow = '->' if rp_before != rp_after else '  '

    has_change = er_before != er_after or rp_before != rp_after or gate_before != gate_after

    print(f"{fname:<35} {len(rows):>4} {len(dropped):>4}  {n_camp_before:>5}{er_arrow:>3}{n_camp_after:>5} {er_before:>3}{er_arrow:>3}{er_after:>3}  ${rev_before:>13,.0f}{rp_arrow:>3}${rev_after:>13,.0f} {rp_before:>3}{rp_arrow:>3}{rp_after:>3}  {gate_flag:>6} {'; '.join(notes) if has_change else ''}")

    if has_change:
        changes.append((fname, n_camp_before, n_camp_after, er_before, er_after, rev_before, rev_after, rp_before, rp_after, gate_before, gate_after))

    if dropped:
        for d in dropped:
            dropped_rows_detail.append((fname, d['fac_name'], d['beds'], d['tot_rev'], d['source'], d['campus_key']))

# Summary
print(f"\n{'='*140}")
print(f"SUMMARY")
print(f"{'='*140}")
print(f"Total entities analyzed: {len(entity_rows)}")
print(f"Entities with scoring changes: {len(changes)}")
print(f"Entities with gate change (PASS->T4): {len(gate_changes)}")
if gate_changes:
    for g in gate_changes:
        print(f"  - {g}")

# Show all dropped rows
print(f"\n{'='*140}")
print(f"ALL DROPPED ROWS (Total_Beds <= {BED_MIN})")
print(f"{'='*140}")
print(f"{'Entity':<35} {'Facility':<50} {'Beds':>5} {'Revenue':>12} {'Source':<6} {'Campus Key'}")
print(f"{'-'*140}")
for fname, fac, beds, rev, source, ck in sorted(dropped_rows_detail, key=lambda x: (x[0], -x[2])):
    print(f"{fname:<35} {fac:<50} {beds:>5} ${rev:>11,.0f} {source:<6} {ck[0]}, {ck[1]}")
print(f"\nTotal dropped rows: {len(dropped_rows_detail)}")

# ER and RP distribution shift
print(f"\n{'='*140}")
print(f"ER DISTRIBUTION SHIFT")
print(f"{'='*140}")
for label, get_vals in [('BEFORE', lambda r: score_dim(len(set(x['campus_key'] for x in r)), ER_THRESHOLDS)),
                         ('AFTER', lambda r: score_dim(len(set(x['campus_key'] for x in r if x['beds'] > BED_MIN)), ER_THRESHOLDS) if any(x['beds'] > BED_MIN for x in r) else 0)]:
    dist = defaultdict(int)
    for fname in entity_rows:
        rows = entity_rows[fname]
        if label == 'BEFORE':
            camps = set(x['campus_key'] for x in rows)
            if len(camps) >= 7:
                dist[score_dim(len(camps), ER_THRESHOLDS)] += 1
        else:
            kept = [x for x in rows if x['beds'] > BED_MIN]
            camps = set(x['campus_key'] for x in kept)
            if len(camps) >= 7:
                dist[score_dim(len(camps), ER_THRESHOLDS)] += 1
    print(f"  {label}: ER1={dist.get(1,0)}  ER2={dist.get(2,0)}  ER3={dist.get(3,0)}  ER4={dist.get(4,0)}  ER5={dist.get(5,0)}  (scored={sum(dist.values())})")

print(f"\nRP DISTRIBUTION SHIFT")
for label in ['BEFORE', 'AFTER']:
    dist = defaultdict(int)
    for fname in entity_rows:
        rows = entity_rows[fname]
        if label == 'BEFORE':
            camps = set(x['campus_key'] for x in rows)
            if len(camps) >= 7:
                rev = sum(x['tot_rev'] for x in rows)
                dist[score_dim(rev, RP_THRESHOLDS)] += 1
        else:
            kept = [x for x in rows if x['beds'] > BED_MIN]
            camps = set(x['campus_key'] for x in kept)
            if len(camps) >= 7:
                rev = sum(x['tot_rev'] for x in kept)
                dist[score_dim(rev, RP_THRESHOLDS)] += 1
    print(f"  {label}: RP1={dist.get(1,0)}  RP2={dist.get(2,0)}  RP3={dist.get(3,0)}  RP4={dist.get(4,0)}  RP5={dist.get(5,0)}  (scored={sum(dist.values())})")
