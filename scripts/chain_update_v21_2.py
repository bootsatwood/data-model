#!/usr/bin/env python3
"""
CMS Chain ID Corporate Name Enhancement — V21.1 -> V21.2

Facilities we classify as Independent but CMS assigns to a chain get their
Corporate_Name updated. The Four-Rule hierarchy is then rerun to reclassify
Ownership_Type across the entire database.

Three modes:
  preview   — Dry run: show all proposed Corporate_Name changes
  apply     — Write changes to 1_Combined_Database_FINAL_V21_2.xlsx
  four-rule — Recompute Ownership_Type on the V21.2 file using Four-Rule hierarchy

Usage (from repo root):
  python scripts/chain_update_v21_2.py preview
  python scripts/chain_update_v21_2.py apply
  python scripts/chain_update_v21_2.py four-rule
"""

import csv
import re
import sys
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

VAULT = Path.home() / "OneDrive - Eventus WholeHealth" / "Vault" / "02_Data_Model"
CMS_FILE = VAULT / "Reference" / "Source_CMS_NH_ProviderInfo_Feb2026.csv"
DB_V21_1 = VAULT / "Current" / "1_Combined_Database_FINAL_V21_1.xlsx"
DB_V21_2 = VAULT / "Current" / "1_Combined_Database_FINAL_V21_2.xlsx"

FOOTPRINT = {'IN', 'KY', 'NC', 'OH', 'SC', 'VA', 'MI', 'IL', 'WI', 'MN', 'FL', 'MD', 'GA', 'MO'}


# ---------------------------------------------------------------------------
# Normalizers (shared with chain_crossref.py)
# ---------------------------------------------------------------------------

def norm(s):
    if not s:
        return ''
    return re.sub(r'[^a-z0-9]', '', s.lower())


def norm_addr(s):
    if not s:
        return ''
    s = s.lower().strip()
    for word, abbr in [('street', 'st'), ('road', 'rd'), ('drive', 'dr'), ('avenue', 'ave'),
                        ('boulevard', 'blvd'), ('lane', 'ln'), ('court', 'ct'),
                        ('north', 'n'), ('south', 's'), ('east', 'e'), ('west', 'w')]:
        s = re.sub(r'\b' + word + r'\b', abbr, s)
    return re.sub(r'[^a-z0-9]', '', s)


# ---------------------------------------------------------------------------
# CMS data loader
# ---------------------------------------------------------------------------

def load_cms():
    """Load CMS provider data keyed by normalized address."""
    cms = {}
    with open(CMS_FILE, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            addr = row.get('Provider Address', '')
            city = row.get('City/Town', '')
            state = row.get('State', '')
            key = norm_addr(addr) + '|' + norm(city) + '|' + norm(state)
            cms[key] = {
                'chain': row['Chain Name'].strip(),
                'chain_id': row['Chain ID'].strip(),
                'fac_in_chain': row['Number of Facilities in Chain'].strip(),
            }
    return cms


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def col_index(headers, name):
    return headers.index(name)


def safe(val):
    if val is None:
        return ''
    return str(val).strip()


# ---------------------------------------------------------------------------
# Build canonical Corporate_Name universe
# ---------------------------------------------------------------------------

def build_canonical_names(db_path):
    """Read the database and return a dict: norm(corp_name) -> canonical string.

    For each normalized name, the canonical form is the spelling used most often
    among Corporate-classified rows.
    """
    wb = load_workbook(db_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci_corp = col_index(headers, 'Corporate_Name')
    ci_own = col_index(headers, 'Ownership_Type')

    counts = {}  # norm -> Counter of original spellings
    for row in ws.iter_rows(min_row=2, values_only=True):
        corp = safe(row[ci_corp])
        own = safe(row[ci_own])
        if corp and own == 'Corporate':
            n = norm(corp)
            if n not in counts:
                counts[n] = Counter()
            counts[n][corp] += 1

    wb.close()

    # Pick the most common spelling for each normalized name
    canonical = {}
    for n, ctr in counts.items():
        canonical[n] = ctr.most_common(1)[0][0]
    return canonical


# ---------------------------------------------------------------------------
# Identify proposed changes
# ---------------------------------------------------------------------------

def find_changes(db_path, cms, canonical):
    """Find all Independent SNFs in footprint where CMS assigns a chain.

    Returns a list of dicts with change details.
    """
    wb = load_workbook(db_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    ci = {name: col_index(headers, name) for name in [
        'Source_Type', 'Facility_Name', 'Corporate_Name', 'Address',
        'City', 'State', 'Ownership_Type', 'Do_We_Serve',
    ]}

    changes = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        source = safe(row[ci['Source_Type']])
        state = safe(row[ci['State']])
        ownership = safe(row[ci['Ownership_Type']])

        if source != 'SNF' or state not in FOOTPRINT or ownership != 'Independent':
            continue

        addr = safe(row[ci['Address']])
        city = safe(row[ci['City']])
        key = norm_addr(addr) + '|' + norm(city) + '|' + norm(state)

        if key not in cms:
            continue
        c = cms[key]
        if not c['chain']:
            continue

        old_corp = safe(row[ci['Corporate_Name']])
        cms_chain = c['chain']
        cms_norm = norm(cms_chain)
        served = safe(row[ci['Do_We_Serve']]) == 'Yes'

        # Determine new Corporate_Name and section
        if cms_norm in canonical:
            new_corp = canonical[cms_norm]
            section = 'A'
        else:
            new_corp = cms_chain
            section = 'B'

        changes.append({
            'excel_row': excel_row,
            'facility': safe(row[ci['Facility_Name']]),
            'city': city,
            'state': state,
            'old_corp': old_corp,
            'new_corp': new_corp,
            'cms_chain': cms_chain,
            'chain_id': c['chain_id'],
            'chain_size': c['fac_in_chain'],
            'served': served,
            'section': section,
        })

    wb.close()
    return changes


# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    """Apply the Four-Rule Count-Based Hierarchy.

    1. Blank Corporate Name -> Independent
    2. "INDEPENDENT" placeholder -> Independent
    3. Multi-facility chains (count > 1) -> Corporate
    4. Single-facility operators (count = 1) -> Independent
    """
    if not corp_name:
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    count = corp_counts.get(corp_name, 1)
    if count > 1:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Mode: preview
# ---------------------------------------------------------------------------

def cmd_preview():
    print("Loading CMS data...")
    cms = load_cms()
    print(f"  {len(cms):,} CMS records loaded.")

    print("Building canonical Corporate_Name universe...")
    canonical = build_canonical_names(DB_V21_1)
    print(f"  {len(canonical):,} distinct corporate names.")

    print("Scanning for Independent facilities with CMS chain assignment...")
    changes = find_changes(DB_V21_1, cms, canonical)

    section_a = [c for c in changes if c['section'] == 'A']
    section_b = [c for c in changes if c['section'] == 'B']
    served = [c for c in changes if c['served']]

    print()
    print("=" * 70)
    print("PREVIEW: CMS Chain ID Corporate Name Enhancement (V21.2)")
    print("=" * 70)
    print()
    print(f"Total proposed changes: {len(changes)}")
    print(f"  Section A (CMS chain matches our existing name): {len(section_a)}")
    print(f"  Section B (CMS chain is new to us):              {len(section_b)}")
    print(f"  Served subset (affects revenue):                 {len(served)}")
    print()

    # Section A detail
    print("=" * 70)
    print(f"SECTION A: Using our canonical Corporate_Name ({len(section_a)} facilities)")
    print("=" * 70)
    print()

    # Group by new_corp
    a_by_corp = {}
    for c in sorted(section_a, key=lambda x: (x['new_corp'], x['state'], x['facility'])):
        a_by_corp.setdefault(c['new_corp'], []).append(c)

    for corp_name, facs in sorted(a_by_corp.items()):
        states = sorted(set(f['state'] for f in facs))
        served_ct = sum(1 for f in facs if f['served'])
        cms_chains = sorted(set(f['cms_chain'] for f in facs))
        chain_ids = sorted(set(f['chain_id'] for f in facs))
        print(f"  {corp_name} ({len(facs)} facilities in {', '.join(states)})")
        print(f"    CMS chain: {', '.join(cms_chains)} (ID: {', '.join(chain_ids)})")
        if served_ct:
            print(f"    *** {served_ct} SERVED ***")
        for f in facs:
            tag = " *SERVED*" if f['served'] else ""
            old_tag = f" [was: {f['old_corp']}]" if f['old_corp'] else ""
            print(f"    - {f['facility']}, {f['city']} {f['state']}{old_tag}{tag}")
        print()

    # Section B detail
    print("=" * 70)
    print(f"SECTION B: Using CMS chain name (new to us) ({len(section_b)} facilities)")
    print("=" * 70)
    print()

    b_by_chain = {}
    for c in sorted(section_b, key=lambda x: (x['cms_chain'], x['state'], x['facility'])):
        b_by_chain.setdefault(c['cms_chain'], []).append(c)

    for chain_name, facs in sorted(b_by_chain.items()):
        states = sorted(set(f['state'] for f in facs))
        served_ct = sum(1 for f in facs if f['served'])
        chain_ids = sorted(set(f['chain_id'] for f in facs))
        print(f"  {chain_name} (ID: {', '.join(chain_ids)}, {len(facs)} facilities in {', '.join(states)})")
        if served_ct:
            print(f"    *** {served_ct} SERVED ***")
        for f in facs:
            tag = " *SERVED*" if f['served'] else ""
            old_tag = f" [was: {f['old_corp']}]" if f['old_corp'] else ""
            print(f"    - {f['facility']}, {f['city']} {f['state']}{old_tag}{tag}")
        print()

    # Served summary
    if served:
        print("=" * 70)
        print(f"SERVED FACILITIES ({len(served)}) — Review these carefully")
        print("=" * 70)
        print()
        for f in sorted(served, key=lambda x: (x['section'], x['new_corp'], x['facility'])):
            print(f"  [{f['section']}] {f['facility']}, {f['city']} {f['state']}")
            print(f"      Corporate_Name: '{f['old_corp']}' -> '{f['new_corp']}'")
            print(f"      CMS chain: {f['cms_chain']} (ID: {f['chain_id']})")
            print()

    return changes


# ---------------------------------------------------------------------------
# Mode: apply
# ---------------------------------------------------------------------------

def cmd_apply():
    print("Loading CMS data...")
    cms = load_cms()

    print("Building canonical Corporate_Name universe...")
    canonical = build_canonical_names(DB_V21_1)

    print("Scanning for changes...")
    changes = find_changes(DB_V21_1, cms, canonical)
    print(f"  {len(changes)} changes to apply.")

    if not changes:
        print("No changes to apply.")
        return

    # Build lookup: excel_row -> new_corp
    row_updates = {c['excel_row']: c['new_corp'] for c in changes}

    print(f"Loading workbook (full mode) for writing...")
    wb = load_workbook(DB_V21_1)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    ci_corp = col_index(headers, 'Corporate_Name') + 1  # openpyxl is 1-indexed

    applied = 0
    for excel_row, new_corp in row_updates.items():
        cell = ws.cell(row=excel_row, column=ci_corp)
        old_val = safe(cell.value)
        cell.value = new_corp
        applied += 1

    print(f"  {applied} Corporate_Name values updated.")
    print(f"Saving as {DB_V21_2.name}...")
    wb.save(DB_V21_2)
    wb.close()
    print(f"  Saved to: {DB_V21_2}")

    # Summary
    section_a = sum(1 for c in changes if c['section'] == 'A')
    section_b = sum(1 for c in changes if c['section'] == 'B')
    served = sum(1 for c in changes if c['served'])
    print()
    print(f"APPLY COMPLETE: {applied} Corporate_Name changes written to V21.2")
    print(f"  Section A: {section_a}  |  Section B: {section_b}  |  Served: {served}")
    print()
    print("Next step: python scripts/chain_update_v21_2.py four-rule")


# ---------------------------------------------------------------------------
# Mode: four-rule
# ---------------------------------------------------------------------------

def cmd_four_rule():
    if not DB_V21_2.exists():
        print(f"ERROR: {DB_V21_2.name} not found. Run 'apply' first.")
        sys.exit(1)

    print(f"Loading {DB_V21_2.name} (full mode)...")
    wb = load_workbook(DB_V21_2)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    ci_corp = col_index(headers, 'Corporate_Name') + 1  # 1-indexed
    ci_own = col_index(headers, 'Ownership_Type') + 1

    total_rows = ws.max_row - 1  # exclude header
    print(f"  {total_rows:,} rows loaded.")

    # Pass 1: Count facilities per Corporate_Name
    print("Pass 1: Counting facilities per Corporate_Name...")
    corp_counts = Counter()
    for row_num in range(2, ws.max_row + 1):
        corp = safe(ws.cell(row=row_num, column=ci_corp).value)
        if corp and corp.upper() != 'INDEPENDENT':
            corp_counts[corp] += 1

    print(f"  {len(corp_counts):,} distinct Corporate_Name values (excluding blank/INDEPENDENT).")

    # Pass 2: Apply Four-Rule and log changes
    print("Pass 2: Applying Four-Rule hierarchy...")
    reclassifications = []
    for row_num in range(2, ws.max_row + 1):
        corp = safe(ws.cell(row=row_num, column=ci_corp).value)
        old_type = safe(ws.cell(row=row_num, column=ci_own).value)
        new_type = four_rule_classify(corp, corp_counts)

        if old_type != new_type:
            # Get facility info for logging
            ci_fac = col_index(headers, 'Facility_Name') + 1
            ci_state = col_index(headers, 'State') + 1
            ci_serve = col_index(headers, 'Do_We_Serve') + 1
            fac = safe(ws.cell(row=row_num, column=ci_fac).value)
            state = safe(ws.cell(row=row_num, column=ci_state).value)
            served = safe(ws.cell(row=row_num, column=ci_serve).value)

            reclassifications.append({
                'excel_row': row_num,
                'facility': fac,
                'state': state,
                'corp_name': corp,
                'old_type': old_type,
                'new_type': new_type,
                'served': served == 'Yes',
                'corp_count': corp_counts.get(corp, 0),
            })
            ws.cell(row=row_num, column=ci_own).value = new_type

    print(f"  {len(reclassifications)} Ownership_Type reclassifications.")

    if reclassifications:
        # Count by direction
        ind_to_corp = [r for r in reclassifications if r['old_type'] == 'Independent' and r['new_type'] == 'Corporate']
        corp_to_ind = [r for r in reclassifications if r['old_type'] == 'Corporate' and r['new_type'] == 'Independent']
        served_reclass = [r for r in reclassifications if r['served']]

        print()
        print(f"  Independent -> Corporate: {len(ind_to_corp)}")
        print(f"  Corporate -> Independent: {len(corp_to_ind)}")
        print(f"  Served facilities affected: {len(served_reclass)}")

        # Show served reclassifications in detail
        if served_reclass:
            print()
            print("  SERVED RECLASSIFICATIONS:")
            for r in sorted(served_reclass, key=lambda x: (x['facility'])):
                print(f"    Row {r['excel_row']}: {r['facility']} ({r['state']}) — "
                      f"{r['old_type']} -> {r['new_type']} [{r['corp_name']}, count={r['corp_count']}]")

        # Show all Independent->Corporate grouped by corp name
        print()
        print("  INDEPENDENT -> CORPORATE BY CHAIN:")
        by_corp = {}
        for r in ind_to_corp:
            by_corp.setdefault(r['corp_name'], []).append(r)
        for corp, facs in sorted(by_corp.items(), key=lambda x: -len(x[1])):
            states = sorted(set(f['state'] for f in facs))
            print(f"    {corp} ({len(facs)} fac, count={facs[0]['corp_count']}, {', '.join(states)})")

        # Show Corporate->Independent if any
        if corp_to_ind:
            print()
            print("  CORPORATE -> INDEPENDENT:")
            for r in sorted(corp_to_ind, key=lambda x: x['facility']):
                print(f"    Row {r['excel_row']}: {r['facility']} ({r['state']}) — "
                      f"[{r['corp_name']}, count={r['corp_count']}]")

    # Final counts
    print()
    print("Recomputing final counts...")
    corporate = 0
    independent = 0
    for row_num in range(2, ws.max_row + 1):
        own = safe(ws.cell(row=row_num, column=ci_own).value)
        if own == 'Corporate':
            corporate += 1
        elif own == 'Independent':
            independent += 1

    print(f"  Corporate:   {corporate:,}")
    print(f"  Independent: {independent:,}")
    print(f"  Total:       {corporate + independent:,}")

    print()
    print(f"Saving {DB_V21_2.name}...")
    wb.save(DB_V21_2)
    wb.close()
    print("  Saved.")

    print()
    print(f"FOUR-RULE COMPLETE: {len(reclassifications)} reclassifications")
    print()
    print("Next step: Update scripts/qc_config.json, then run:")
    print("  python scripts/qc_validator.py validate")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply', 'four-rule'):
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]
    print(f"CMS Chain ID Corporate Name Enhancement — V21.2")
    print(f"Mode: {mode}")
    print()

    if mode == 'preview':
        cmd_preview()
    elif mode == 'apply':
        cmd_apply()
    elif mode == 'four-rule':
        cmd_four_rule()


if __name__ == '__main__':
    main()
