#!/usr/bin/env python3
"""
Corporate Fix — Wave 2c (V22.2 -> V22.3)

Three-tier corporate attribution with data source tracking.

Data Authority Hierarchy:
  Tier 1 — GLR (Internal Facility DB): ground truth for all matched facilities
  Tier 2 — CMS Chain ID: authoritative for SNFs not matched by GLR
  Tier 3 — NIC Maps Operator: fallback for PROPCO-tagged ALFs, confidence-tiered
            by Operator ID clustering (Tier A = 10+, Tier B = 3-9 buildings)

Phases:
  1d: Liberty PROPCO SNF duplicate removal (DCR #7)
  2a: GLR-based corporate attribution corrections (Tier 1)
  2b: CMS-based corporate attribution for SNFs (Tier 2)
  2c: NIC-based PROPCO reattribution for ALFs (Tier 3, A/B only)
   3: Four-Rule ownership reclassification
   4: Corp_Attribution_Source QC column + validation

New column added: Corp_Attribution_Source (QC-only)
  GLR    = Internal Facility DB (ground truth)
  CMS    = Federal Chain ID
  NIC-A  = NIC Operator, high confidence (10+ buildings same Operator ID)
  NIC-B  = NIC Operator, medium confidence (3-9 buildings)
  LEGACY = Pre-V22 attribution (untouched)

Usage:
  python corporate_fix.py preview
  python corporate_fix.py apply
"""

import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from utils import (norm, norm_addr, addr_key, safe, load_db,
                   VAULT, REPORT_DIR, NIC_ALF_FILE, GLR_FILE,
                   CMS_PROVIDER_FILE)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_2.xlsx"
DB_TARGET = VAULT / "Current" / "1_Combined_Database_FINAL_V22_3.xlsx"
NIC_ENRICHMENT = REPORT_DIR / "nic_enrichment.csv"
QC_CONFIG = Path(__file__).resolve().parent / "qc_config.json"

# ---------------------------------------------------------------------------
# Phase 1d constant
# ---------------------------------------------------------------------------

LIBERTY_PROPCO_SNF_PATTERN = "LIBERTY HEALTHCARE PROPERTIES"

# ---------------------------------------------------------------------------
# State abbreviation mapping (GLR uses full state names)
# ---------------------------------------------------------------------------

STATE_ABBR = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN',
    'mississippi': 'MS', 'missouri': 'MO', 'montana': 'MT',
    'nebraska': 'NE', 'nevada': 'NV', 'new hampshire': 'NH',
    'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
    'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH',
    'oklahoma': 'OK', 'oregon': 'OR', 'pennsylvania': 'PA',
    'rhode island': 'RI', 'south carolina': 'SC', 'south dakota': 'SD',
    'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT', 'vermont': 'VT',
    'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
    'wisconsin': 'WI', 'wyoming': 'WY', 'district of columbia': 'DC',
}

# ---------------------------------------------------------------------------
# Corporate name normalization (suffix-aware)
# ---------------------------------------------------------------------------

CORP_SUFFIXES = [
    ' incorporated', ' inc.', ' inc',
    ' llc.', ' llc', ' l.l.c.', ' l.l.c',
    ' corporation', ' corp.', ' corp',
    ' company', ' co.', ' co',
    ' limited', ' ltd.', ' ltd',
    ' l.p.', ' lp',
    ' l.c.', ' lc',
]
CORP_SUFFIXES.sort(key=len, reverse=True)


def norm_corp(s):
    """Normalize a corporate name for canonical matching.

    Strips one trailing legal suffix, normalizes & -> and,
    then reduces to lowercase alphanumeric.
    """
    if not s:
        return ''
    s = s.lower().strip()
    for suffix in CORP_SUFFIXES:
        if s.endswith(suffix):
            s = s[:len(s) - len(suffix)].rstrip(' ,.')
            break
    s = s.replace('&', ' and ')
    return re.sub(r'[^a-z0-9]', '', s)


# ---------------------------------------------------------------------------
# Suspicious-operator check (NIC Tier 3 flagging)
# ---------------------------------------------------------------------------

SUSPICIOUS_PATTERNS = ['propco', 'opco', 'not avail', 'unknown',
                       'privately owned', 'private owner']


def is_suspicious_operator(name):
    """Return True if a NIC operator name looks like a holding company or
    placeholder rather than a real operating company."""
    if not name or len(name.strip()) < 3:
        return True
    low = name.lower().strip()
    for pat in SUSPICIOUS_PATTERNS:
        if pat in low:
            return True
    if low[0].isdigit() and any(p in low for p in ['llc', 'propco', 'opco']):
        return True
    return False


# ---------------------------------------------------------------------------
# Build canonical Corporate_Name map from the database
# ---------------------------------------------------------------------------

def build_canonical_names(db_path):
    """Map norm_corp(corp_name) -> most-common DB spelling.

    Includes ALL non-blank Corporate_Names so single-facility operators
    can still be matched.
    """
    wb = load_workbook(db_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci_corp = headers.index('Corporate_Name')

    counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        corp = safe(row[ci_corp])
        if corp and corp.upper() != 'INDEPENDENT':
            n = norm_corp(corp)
            if n not in counts:
                counts[n] = Counter()
            counts[n][corp] += 1

    wb.close()

    canonical = {}
    for n, ctr in counts.items():
        canonical[n] = ctr.most_common(1)[0][0]

    # Manual aliases for known variants
    canonical[norm_corp('Liberty Senior Living')] = 'LIBERTY'

    return canonical


# ---------------------------------------------------------------------------
# Four-Rule Ownership Hierarchy
# ---------------------------------------------------------------------------

def four_rule_classify(corp_name, corp_counts):
    """Apply the Four-Rule Count-Based Hierarchy.

    1. Blank Corporate Name -> Independent
    2. "INDEPENDENT" placeholder -> Independent
    3. Multi-facility chains (count > 1) -> Corporate
    4. Single-facility operators (count = 1) -> Independent
    """
    if not corp_name:
        return 'Independent'
    if corp_name.upper() == 'INDEPENDENT':
        return 'Independent'
    if corp_counts.get(corp_name, 1) > 1:
        return 'Corporate'
    return 'Independent'


# ---------------------------------------------------------------------------
# Data Loaders
# ---------------------------------------------------------------------------

def load_glr():
    """Load GLR Export keyed by addr_key.

    Returns dict: addr_key -> {parent_company, facility_name, facility_type,
                                status, client_status}
    """
    wb = load_workbook(GLR_FILE, read_only=True, data_only=True)
    ws = wb['GLR Export']
    raw_headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = safe(cell.value)
        # Strip emoji/high-Unicode from headers
        val = ''.join(c for c in val if ord(c) < 65536)
        raw_headers.append(val.strip())

    glr = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(raw_headers, row))
        addr = safe(rd.get('AddressLine1', ''))
        city = safe(rd.get('City', ''))
        state_full = safe(rd.get('StateCode', '')).lower().strip()
        state = STATE_ABBR.get(state_full, state_full.upper()[:2])

        if not addr or not city or not state:
            continue

        key = addr_key(addr, city, state)
        parent = safe(rd.get('Parent Company', ''))

        glr[key] = {
            'parent_company': parent,
            'facility_name': safe(rd.get('Facility Name', '')),
            'facility_type': safe(rd.get('Facility Type', '')),
            'status': safe(rd.get('Status', '')),
            'client_status': safe(rd.get('Client Status', '')),
        }

    wb.close()
    return glr


def load_cms():
    """Load CMS ProviderInfo keyed by addr_key.

    Returns dict: addr_key -> {chain_name, chain_id, fac_in_chain}
    Only includes rows with non-blank Chain Name.
    """
    cms = {}
    with open(CMS_PROVIDER_FILE, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            chain = row.get('Chain Name', '').strip()
            if not chain:
                continue
            addr = row.get('Provider Address', '')
            city = row.get('City/Town', '')
            state = row.get('State', '')
            key = addr_key(addr, city, state)
            cms[key] = {
                'chain_name': chain,
                'chain_id': row.get('Chain ID', '').strip(),
                'fac_in_chain': row.get('Number of Facilities in Chain', '').strip(),
            }
    return cms


def load_nic_maps():
    """Load NIC Maps keyed by addr_key.

    Returns dict: addr_key -> {operator_id, operator_name, owner_name, building_name}
    Also returns raw list of all entries for Operator ID counting.
    """
    wb = load_workbook(NIC_ALF_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [safe(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    nic_by_addr = {}
    all_entries = []  # for Operator ID counting

    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        op_id = safe(rd.get('Operator ID', ''))
        op_name = safe(rd.get('Operator Name', ''))
        owner_name = safe(rd.get('Owner Name', ''))

        entry = {
            'operator_id': op_id,
            'operator_name': op_name,
            'owner_name': owner_name,
            'building_name': safe(rd.get('Building Name', '')),
        }
        all_entries.append(entry)

        addr = safe(rd.get('Address', ''))
        city = safe(rd.get('City', ''))
        state = safe(rd.get('State', ''))
        if addr and city and state:
            key = addr_key(addr, city, state)
            # Keep first entry per address (same campus = same operator)
            if key not in nic_by_addr:
                nic_by_addr[key] = entry

    wb.close()
    return nic_by_addr, all_entries


def build_nic_tiers(all_entries):
    """Build Operator ID confidence tiers from NIC Maps data.

    Returns:
      tiers: dict op_id -> 'A'/'B'/'C'
      op_counts: Counter op_id -> building count
    """
    op_counts = Counter()
    for e in all_entries:
        op_id = e['operator_id']
        if op_id:
            op_counts[op_id] += 1

    tiers = {}
    for op_id, count in op_counts.items():
        if count >= 10:
            tiers[op_id] = 'A'
        elif count >= 3:
            tiers[op_id] = 'B'
        else:
            tiers[op_id] = 'C'

    return tiers, op_counts


def load_propco_tags():
    """Load PROPCO-tagged ALF address keys from NIC enrichment CSV.

    Returns set of addr_keys that have PROPCO_AS_CORPORATE or PROPCO_SPLIT tags.
    """
    propco_keys = set()
    with open(NIC_ENRICHMENT, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            tags = row.get('tags', '')
            if 'PROPCO_AS_CORPORATE' in tags or 'PROPCO_SPLIT' in tags:
                key = addr_key(row['address'], row['city'], row['state'])
                propco_keys.add(key)
    return propco_keys


# ---------------------------------------------------------------------------
# Resolve a source name to canonical DB spelling
# ---------------------------------------------------------------------------

def resolve_canonical(name, canonical):
    """Given a raw corporate name from any source, return the canonical
    DB spelling (or the original name if no canonical match)."""
    if not name:
        return ''
    n = norm_corp(name)
    if n in canonical:
        return canonical[n]
    return name


# ---------------------------------------------------------------------------
# Core: Compute all changes (shared by preview and apply)
# ---------------------------------------------------------------------------

def compute_all_changes():
    """Compute every change for Phases 1d, 2a-c, 3.

    Returns a dict with all change data, statistics, attribution sources,
    and the loaded rows.
    """
    print("Loading V22.2 database...")
    headers, rows = load_db(DB_SOURCE)
    print(f"  {len(rows):,} rows loaded.")

    print("Building canonical Corporate_Name map...")
    canonical = build_canonical_names(DB_SOURCE)
    print(f"  {len(canonical):,} distinct canonical names.")

    print("Loading GLR Export (Tier 1)...")
    glr = load_glr()
    print(f"  {len(glr):,} GLR facilities loaded.")

    print("Loading CMS ProviderInfo (Tier 2)...")
    cms = load_cms()
    print(f"  {len(cms):,} CMS facilities with Chain Name.")

    print("Loading NIC Maps (Tier 3)...")
    nic_by_addr, all_nic = load_nic_maps()
    print(f"  {len(nic_by_addr):,} NIC buildings indexed by address.")

    print("Building NIC Operator ID confidence tiers...")
    nic_tiers, nic_op_counts = build_nic_tiers(all_nic)
    tier_a = sum(1 for t in nic_tiers.values() if t == 'A')
    tier_b = sum(1 for t in nic_tiers.values() if t == 'B')
    tier_c = sum(1 for t in nic_tiers.values() if t == 'C')
    print(f"  Tier A (10+ buildings): {tier_a} operators")
    print(f"  Tier B (3-9 buildings): {tier_b} operators")
    print(f"  Tier C (1-2 buildings): {tier_c} operators")

    print("Loading NIC enrichment PROPCO tags...")
    propco_keys = load_propco_tags()
    print(f"  {len(propco_keys):,} PROPCO-tagged address keys.")
    print()

    # Build address-key index for V22.2 rows
    rows_by_addr = {}
    for r in rows:
        key = addr_key(safe(r.get('Address', '')),
                       safe(r.get('City', '')),
                       safe(r.get('State', '')))
        r['_addr_key'] = key
        rows_by_addr.setdefault(key, []).append(r)

    # ===================================================================
    # Phase 1d: Liberty SNF-SNF PROPCO Duplicates
    # ===================================================================

    removal_rows = set()
    phase1d = []
    for r in rows:
        corp_upper = safe(r.get('Corporate_Name', '')).upper()
        stype = safe(r.get('Source_Type', ''))
        if stype == 'SNF' and LIBERTY_PROPCO_SNF_PATTERN in corp_upper:
            key = r['_addr_key']
            siblings = rows_by_addr.get(key, [])
            real_snf = [s for s in siblings
                        if s is not r
                        and safe(s.get('Source_Type', '')) == 'SNF'
                        and safe(s.get('Corporate_Name', '')).upper() == 'LIBERTY']
            if not real_snf:
                city = safe(r.get('City', '')).upper()
                state = safe(r.get('State', '')).upper()
                fac_prefix = norm(safe(r.get('Facility_Name', '')))[:15]
                real_snf = [s for s in rows
                            if s is not r
                            and safe(s.get('Source_Type', '')) == 'SNF'
                            and safe(s.get('City', '')).upper() == city
                            and safe(s.get('State', '')).upper() == state
                            and safe(s.get('Corporate_Name', '')).upper() == 'LIBERTY'
                            and norm(safe(s.get('Facility_Name', '')))[:15]
                            == fac_prefix]
            if real_snf:
                phase1d.append({
                    'excel_row': r['_excel_row'],
                    'facility': safe(r.get('Facility_Name', '')),
                    'city': safe(r.get('City', '')),
                    'state': safe(r.get('State', '')),
                    'corp': safe(r.get('Corporate_Name', '')),
                    'beds': safe(r.get('Total_Beds', '')),
                    'served': safe(r.get('Do_We_Serve', '')),
                    'real_snf': safe(real_snf[0].get('Facility_Name', '')),
                    'real_row': real_snf[0]['_excel_row'],
                })
                removal_rows.add(r['_excel_row'])

    # ===================================================================
    # Phase 2: Three-Tier Corporate Attribution
    # ===================================================================

    # Track per-row: attribution source and corporate name changes
    attr_sources = {}       # excel_row -> source string
    corp_changes = {}       # excel_row -> {old_corp, new_corp, source, detail}

    # Phase 2a/2b/2c detail lists for preview
    phase2a_changes = []    # GLR corrections
    phase2a_confirmed = 0   # GLR matched, already correct
    phase2a_independent = 0 # GLR says INDEPENDENT
    phase2a_blank_parent = 0

    phase2b_changes = []    # CMS corrections
    phase2b_confirmed = 0   # CMS matched, already correct
    phase2b_already_glr = 0

    phase2c_auto = []       # NIC auto-applied (Tier A/B)
    phase2c_flagged = []    # NIC flagged (Tier C, suspicious, no match)
    phase2c_already_attr = 0
    phase2c_same_name = 0
    phase2c_no_nic = 0

    # --- Phase 2a: GLR corrections (Tier 1) ---
    for r in rows:
        er = r['_excel_row']
        if er in removal_rows:
            continue
        key = r['_addr_key']
        if key not in glr:
            continue

        glr_entry = glr[key]
        parent = glr_entry['parent_company']

        # Skip blank GLR Parent Company
        if not parent:
            phase2a_blank_parent += 1
            continue

        old_corp = safe(r.get('Corporate_Name', ''))

        # GLR says INDEPENDENT or N/A — clear PROPCO names
        if parent.upper() in ('INDEPENDENT', 'N/A'):
            phase2a_independent += 1
            attr_sources[er] = 'GLR'
            if old_corp and old_corp.upper() != 'INDEPENDENT':
                corp_changes[er] = {
                    'old_corp': old_corp,
                    'new_corp': '',
                    'source': 'GLR',
                    'detail': 'GLR Parent=INDEPENDENT, clearing PROPCO name',
                    'glr_parent': parent,
                    'facility': safe(r.get('Facility_Name', '')),
                    'city': safe(r.get('City', '')),
                    'state': safe(r.get('State', '')),
                    'source_type': safe(r.get('Source_Type', '')),
                    'served': safe(r.get('Do_We_Serve', '')),
                }
                phase2a_changes.append(corp_changes[er])
            continue

        # Resolve GLR parent to canonical DB spelling
        new_corp = resolve_canonical(parent, canonical)
        attr_sources[er] = 'GLR'

        # Compare normalized — skip if already correct
        if norm_corp(new_corp) == norm_corp(old_corp):
            phase2a_confirmed += 1
            continue

        corp_changes[er] = {
            'old_corp': old_corp,
            'new_corp': new_corp,
            'source': 'GLR',
            'detail': f'GLR Parent="{parent}" -> canonical="{new_corp}"',
            'glr_parent': parent,
            'facility': safe(r.get('Facility_Name', '')),
            'city': safe(r.get('City', '')),
            'state': safe(r.get('State', '')),
            'source_type': safe(r.get('Source_Type', '')),
            'served': safe(r.get('Do_We_Serve', '')),
        }
        phase2a_changes.append(corp_changes[er])

    # --- Phase 2b: CMS corrections (Tier 2 — SNFs not in GLR) ---
    for r in rows:
        er = r['_excel_row']
        if er in removal_rows:
            continue
        if er in attr_sources:
            if safe(r.get('Source_Type', '')) == 'SNF':
                phase2b_already_glr += 1
            continue
        if safe(r.get('Source_Type', '')) != 'SNF':
            continue

        key = r['_addr_key']
        if key not in cms:
            continue

        cms_entry = cms[key]
        chain = cms_entry['chain_name']

        new_corp = resolve_canonical(chain, canonical)
        old_corp = safe(r.get('Corporate_Name', ''))
        attr_sources[er] = 'CMS'

        if norm_corp(new_corp) == norm_corp(old_corp):
            phase2b_confirmed += 1
            continue

        corp_changes[er] = {
            'old_corp': old_corp,
            'new_corp': new_corp,
            'source': 'CMS',
            'detail': f'CMS Chain="{chain}" (ID={cms_entry["chain_id"]})',
            'chain_id': cms_entry['chain_id'],
            'fac_in_chain': cms_entry['fac_in_chain'],
            'facility': safe(r.get('Facility_Name', '')),
            'city': safe(r.get('City', '')),
            'state': safe(r.get('State', '')),
            'source_type': safe(r.get('Source_Type', '')),
            'served': safe(r.get('Do_We_Serve', '')),
        }
        phase2b_changes.append(corp_changes[er])

    # --- Phase 2c: NIC corrections (Tier 3 — PROPCO-tagged ALFs) ---
    for r in rows:
        er = r['_excel_row']
        if er in removal_rows:
            continue

        key = r['_addr_key']
        # Only target PROPCO-tagged ALF addresses
        if key not in propco_keys:
            continue
        # Only ALF rows at PROPCO addresses
        if safe(r.get('Source_Type', '')) != 'ALF':
            continue

        # Skip if already attributed by GLR or CMS
        if er in attr_sources:
            phase2c_already_attr += 1
            continue

        if key not in nic_by_addr:
            phase2c_no_nic += 1
            continue

        nic_entry = nic_by_addr[key]
        op_id = nic_entry['operator_id']
        op_name = nic_entry['operator_name']

        # Check for suspicious operator names
        if is_suspicious_operator(op_name):
            phase2c_flagged.append({
                'excel_row': er,
                'facility': safe(r.get('Facility_Name', '')),
                'city': safe(r.get('City', '')),
                'state': safe(r.get('State', '')),
                'old_corp': safe(r.get('Corporate_Name', '')),
                'nic_operator': op_name,
                'nic_owner': nic_entry['owner_name'],
                'reason': 'suspicious operator name',
                'served': safe(r.get('Do_We_Serve', '')),
            })
            continue

        # Check confidence tier
        tier = nic_tiers.get(op_id, 'C') if op_id else 'C'
        if tier == 'C':
            phase2c_flagged.append({
                'excel_row': er,
                'facility': safe(r.get('Facility_Name', '')),
                'city': safe(r.get('City', '')),
                'state': safe(r.get('State', '')),
                'old_corp': safe(r.get('Corporate_Name', '')),
                'nic_operator': op_name,
                'nic_owner': nic_entry['owner_name'],
                'reason': f'Tier C (Op ID {op_id}, '
                          f'{nic_op_counts.get(op_id, 0)} buildings)',
                'served': safe(r.get('Do_We_Serve', '')),
            })
            continue

        # Resolve NIC operator to canonical DB name
        new_corp = resolve_canonical(op_name, canonical)
        old_corp = safe(r.get('Corporate_Name', ''))

        if norm_corp(new_corp) == norm_corp(old_corp):
            phase2c_same_name += 1
            attr_sources[er] = f'NIC-{tier}'
            continue

        source_label = f'NIC-{tier}'
        attr_sources[er] = source_label
        corp_changes[er] = {
            'old_corp': old_corp,
            'new_corp': new_corp,
            'source': source_label,
            'detail': f'NIC Op="{op_name}" (ID={op_id}, '
                      f'{nic_op_counts.get(op_id, 0)} bldgs, Tier {tier})',
            'nic_operator': op_name,
            'nic_owner': nic_entry['owner_name'],
            'operator_id': op_id,
            'tier': tier,
            'facility': safe(r.get('Facility_Name', '')),
            'city': safe(r.get('City', '')),
            'state': safe(r.get('State', '')),
            'source_type': safe(r.get('Source_Type', '')),
            'served': safe(r.get('Do_We_Serve', '')),
        }
        phase2c_auto.append(corp_changes[er])

    # All remaining rows: LEGACY
    for r in rows:
        er = r['_excel_row']
        if er not in attr_sources and er not in removal_rows:
            attr_sources[er] = 'LEGACY'

    # ===================================================================
    # Phase 3: Four-Rule Ownership Reclassification
    # ===================================================================

    # Build the final state: all corp changes applied
    all_renames = {er: c['new_corp'] for er, c in corp_changes.items()}

    final_data = []
    for r in rows:
        er = r['_excel_row']
        if er in removal_rows:
            continue
        final_corp = all_renames.get(er, safe(r.get('Corporate_Name', '')))
        old_own = safe(r.get('Ownership_Type', ''))
        final_data.append((er, final_corp, old_own))

    corp_counts = Counter()
    for _, corp, _ in final_data:
        if corp and corp.upper() != 'INDEPENDENT':
            corp_counts[corp] += 1

    ownership_changes = []
    for er, corp, old_own in final_data:
        new_own = four_rule_classify(corp, corp_counts)
        if old_own != new_own:
            ownership_changes.append({
                'excel_row': er,
                'corp_name': corp,
                'old_own': old_own,
                'new_own': new_own,
                'corp_count': corp_counts.get(corp, 0),
            })

    ind_to_corp = sum(1 for c in ownership_changes
                      if c['old_own'] == 'Independent' and c['new_own'] == 'Corporate')
    corp_to_ind = sum(1 for c in ownership_changes
                      if c['old_own'] == 'Corporate' and c['new_own'] == 'Independent')

    # ===================================================================
    # Attribution source summary
    # ===================================================================

    source_counts = Counter(attr_sources.values())

    # ===================================================================
    # Baseline counts
    # ===================================================================

    baseline = {
        'total': len(rows),
        'snf': sum(1 for r in rows if safe(r.get('Source_Type', '')) == 'SNF'),
        'alf': sum(1 for r in rows if safe(r.get('Source_Type', '')) == 'ALF'),
        'ilf': sum(1 for r in rows if safe(r.get('Source_Type', '')) == 'ILF'),
        'served': sum(1 for r in rows if safe(r.get('Do_We_Serve', '')) == 'Yes'),
        'corporate': sum(1 for r in rows
                         if safe(r.get('Ownership_Type', '')) == 'Corporate'),
        'independent': sum(1 for r in rows
                           if safe(r.get('Ownership_Type', '')) == 'Independent'),
        'distinct_corps': len(set(
            safe(r.get('Corporate_Name', ''))
            for r in rows
            if safe(r.get('Corporate_Name', ''))
            and safe(r.get('Corporate_Name', '')).upper() != 'INDEPENDENT')),
    }

    final_total = baseline['total'] - len(removal_rows)
    final_corps = set()
    for _, corp, _ in final_data:
        if corp and corp.upper() != 'INDEPENDENT':
            final_corps.add(corp)

    projected = {
        'total': final_total,
        'corporate': baseline['corporate'] + ind_to_corp - corp_to_ind,
        'independent': baseline['independent'] - ind_to_corp + corp_to_ind,
        'distinct_corps': len(final_corps),
    }

    return {
        'headers': headers,
        'rows': rows,
        'canonical': canonical,
        'baseline': baseline,
        'projected': projected,
        'attr_sources': attr_sources,
        'source_counts': source_counts,
        # Phase 1d
        'phase1d': phase1d,
        'removal_rows': removal_rows,
        # Phase 2a — GLR
        'phase2a_changes': phase2a_changes,
        'phase2a_confirmed': phase2a_confirmed,
        'phase2a_independent': phase2a_independent,
        'phase2a_blank_parent': phase2a_blank_parent,
        # Phase 2b — CMS
        'phase2b_changes': phase2b_changes,
        'phase2b_confirmed': phase2b_confirmed,
        'phase2b_already_glr': phase2b_already_glr,
        # Phase 2c — NIC
        'phase2c_auto': phase2c_auto,
        'phase2c_flagged': phase2c_flagged,
        'phase2c_already_attr': phase2c_already_attr,
        'phase2c_same_name': phase2c_same_name,
        'phase2c_no_nic': phase2c_no_nic,
        # Phase 3
        'ownership_changes': ownership_changes,
        'corp_counts': corp_counts,
        'ind_to_corp': ind_to_corp,
        'corp_to_ind': corp_to_ind,
        # Combined
        'corp_changes': corp_changes,
        'all_renames': {er: c['new_corp'] for er, c in corp_changes.items()},
    }


# ---------------------------------------------------------------------------
# Preview mode
# ---------------------------------------------------------------------------

def do_preview():
    print("CORPORATE FIX — PREVIEW MODE (no changes written)")
    print("Three-tier hierarchy: GLR -> CMS -> NIC")
    print("=" * 70)
    print()
    print(f"Source: {DB_SOURCE.name}")
    print(f"Target: {DB_TARGET.name} (will be created on apply)")
    print()

    data = compute_all_changes()

    # ===================================================================
    print("=" * 70)
    print("PHASE 1d: Liberty PROPCO SNF Duplicate Removal (DCR #7)")
    print("=" * 70)
    print()
    if data['phase1d']:
        for d in data['phase1d']:
            print(f"  Row {d['excel_row']:5d}: {d['facility']}, "
                  f"{d['city']}, {d['state']}")
            print(f"           Corp: {d['corp']} | Beds: {d['beds']} | "
                  f"Served: {d['served']}")
            print(f"           Real SNF (row {d['real_row']:5d}): {d['real_snf']}")
            if d['served'] == 'Yes':
                print(f"           *** WARNING: SERVED FACILITY ***")
            print()
    else:
        print("  None found.")
    print(f"  Total: {len(data['phase1d'])} rows to remove")
    print()

    # ===================================================================
    print("=" * 70)
    print("PHASE 2a: GLR Corporate Attribution (Tier 1 — Ground Truth)")
    print("=" * 70)
    print()
    glr_total_matched = (len(data['phase2a_changes']) + data['phase2a_confirmed']
                         + data['phase2a_independent'] + data['phase2a_blank_parent'])
    print(f"  GLR-matched DB rows:          {glr_total_matched:,}")
    print(f"    Already correct:            {data['phase2a_confirmed']:,}")
    print(f"    GLR says INDEPENDENT:       {data['phase2a_independent']:,}")
    print(f"    Blank GLR Parent (skipped): {data['phase2a_blank_parent']:,}")
    print(f"    CORPORATE NAME CHANGES:     {len(data['phase2a_changes']):,}")
    print()

    if data['phase2a_changes']:
        # Split by served/not served
        served_glr = [c for c in data['phase2a_changes']
                      if c.get('served') == 'Yes']
        unserved_glr = [c for c in data['phase2a_changes']
                        if c.get('served') != 'Yes']

        if served_glr:
            print(f"  SERVED GLR CHANGES ({len(served_glr)}):")
            for c in served_glr[:20]:
                print(f"    {c['facility']}, {c['city']}, {c['state']}")
                print(f"      \"{c['old_corp']}\" -> \"{c['new_corp']}\"")
                print(f"      {c['detail']}")
            if len(served_glr) > 20:
                print(f"    ... and {len(served_glr) - 20} more")
            print()

        if unserved_glr:
            by_new = Counter(c['new_corp'] for c in unserved_glr)
            print(f"  UNSERVED GLR CHANGES ({len(unserved_glr)}) — top new names:")
            for corp, ct in by_new.most_common(20):
                print(f"    {ct:4d}  {corp}")
            if len(by_new) > 20:
                print(f"    ... and {len(by_new) - 20} more names")
            print()

    # ===================================================================
    print("=" * 70)
    print("PHASE 2b: CMS Corporate Attribution (Tier 2 — SNFs)")
    print("=" * 70)
    print()
    cms_total = (len(data['phase2b_changes']) + data['phase2b_confirmed']
                 + data['phase2b_already_glr'])
    print(f"  SNFs processed by CMS:        "
          f"{len(data['phase2b_changes']) + data['phase2b_confirmed']:,}")
    print(f"    Already GLR-attributed:     {data['phase2b_already_glr']:,}")
    print(f"    Already correct:            {data['phase2b_confirmed']:,}")
    print(f"    CORPORATE NAME CHANGES:     {len(data['phase2b_changes']):,}")
    print()

    if data['phase2b_changes']:
        served_cms = [c for c in data['phase2b_changes']
                      if c.get('served') == 'Yes']
        unserved_cms = [c for c in data['phase2b_changes']
                        if c.get('served') != 'Yes']

        if served_cms:
            print(f"  SERVED CMS CHANGES ({len(served_cms)}):")
            for c in served_cms[:20]:
                print(f"    {c['facility']}, {c['city']}, {c['state']}")
                print(f"      \"{c['old_corp']}\" -> \"{c['new_corp']}\"")
                print(f"      {c['detail']}")
            if len(served_cms) > 20:
                print(f"    ... and {len(served_cms) - 20} more")
            print()

        if unserved_cms:
            by_new = Counter(c['new_corp'] for c in unserved_cms)
            print(f"  UNSERVED CMS CHANGES ({len(unserved_cms)}) — top new names:")
            for corp, ct in by_new.most_common(20):
                print(f"    {ct:4d}  {corp}")
            if len(by_new) > 20:
                print(f"    ... and {len(by_new) - 20} more names")
            print()

    # ===================================================================
    print("=" * 70)
    print("PHASE 2c: NIC Corporate Attribution (Tier 3 — PROPCO ALFs)")
    print("=" * 70)
    print()
    nic_total = (len(data['phase2c_auto']) + len(data['phase2c_flagged'])
                 + data['phase2c_already_attr'] + data['phase2c_same_name']
                 + data['phase2c_no_nic'])
    print(f"  PROPCO-tagged ALFs evaluated: {nic_total:,}")
    print(f"    Already GLR/CMS-attributed: {data['phase2c_already_attr']:,}")
    print(f"    Already correct (same name):{data['phase2c_same_name']:,}")
    print(f"    No NIC match:               {data['phase2c_no_nic']:,}")
    print(f"    AUTO-APPLIED (Tier A/B):    {len(data['phase2c_auto']):,}")
    print(f"    FLAGGED (Tier C/suspicious):{len(data['phase2c_flagged']):,}")
    print()

    if data['phase2c_auto']:
        tier_a_ct = sum(1 for c in data['phase2c_auto']
                        if c.get('tier') == 'A')
        tier_b_ct = sum(1 for c in data['phase2c_auto']
                        if c.get('tier') == 'B')
        served_nic = sum(1 for c in data['phase2c_auto']
                         if c.get('served') == 'Yes')
        print(f"    Tier A auto-applied:        {tier_a_ct:,}")
        print(f"    Tier B auto-applied:        {tier_b_ct:,}")
        print(f"    Served in auto-applied:     {served_nic:,}")
        print()

        by_new = Counter(c['new_corp'] for c in data['phase2c_auto'])
        print(f"  TOP NIC OPERATOR CHANGES (by new corporate name):")
        for corp, ct in by_new.most_common(25):
            print(f"    {ct:4d}  {corp}")
        if len(by_new) > 25:
            print(f"    ... and {len(by_new) - 25} more")
        print()

    if data['phase2c_flagged']:
        served_flag = sum(1 for f in data['phase2c_flagged']
                          if f.get('served') == 'Yes')
        print(f"  FLAGGED ({len(data['phase2c_flagged'])} rows, "
              f"{served_flag} served):")
        for f in data['phase2c_flagged'][:15]:
            tag = " *SERVED*" if f.get('served') == 'Yes' else ""
            print(f"    {f['facility']}, {f['city']}, {f['state']}{tag}")
            print(f"      Corp=\"{f['old_corp']}\" | NIC Op=\"{f['nic_operator']}\"")
            print(f"      Reason: {f['reason']}")
        if len(data['phase2c_flagged']) > 15:
            print(f"    ... and {len(data['phase2c_flagged']) - 15} more")
        print()

    # ===================================================================
    print("=" * 70)
    print("ATTRIBUTION SOURCE SUMMARY (Corp_Attribution_Source column)")
    print("=" * 70)
    print()
    sc = data['source_counts']
    total_attr = sum(sc.values())
    for src in ['GLR', 'CMS', 'NIC-A', 'NIC-B', 'LEGACY']:
        ct = sc.get(src, 0)
        pct = ct / total_attr * 100 if total_attr else 0
        print(f"  {src:<8s} {ct:>7,} rows  ({pct:5.1f}%)")
    print(f"  {'TOTAL':<8s} {total_attr:>7,} rows")
    print()

    # ===================================================================
    print("=" * 70)
    print("PHASE 3: Four-Rule Ownership Reclassification")
    print("=" * 70)
    print()
    own_changes = data['ownership_changes']
    print(f"  Total reclassifications:    {len(own_changes):,}")
    print(f"    Independent -> Corporate:  {data['ind_to_corp']:,}")
    print(f"    Corporate -> Independent:  {data['corp_to_ind']:,}")
    print()

    if own_changes:
        i2c = [c for c in own_changes
               if c['old_own'] == 'Independent' and c['new_own'] == 'Corporate']
        c2i = [c for c in own_changes
               if c['old_own'] == 'Corporate' and c['new_own'] == 'Independent']

        if i2c:
            by_corp = {}
            for c in i2c:
                by_corp.setdefault(c['corp_name'], []).append(c)
            print("  INDEPENDENT -> CORPORATE (top 20 by count):")
            for corp, facs in sorted(by_corp.items(),
                                     key=lambda x: -len(x[1]))[:20]:
                count = data['corp_counts'].get(corp, 0)
                print(f"    {len(facs):4d}  {corp} (total={count})")
            print()

        if c2i:
            print(f"  CORPORATE -> INDEPENDENT ({len(c2i)} rows, first 20):")
            for c in c2i[:20]:
                print(f"    Row {c['excel_row']:5d}: {c['corp_name']} "
                      f"(count={c['corp_count']})")
            if len(c2i) > 20:
                print(f"    ... and {len(c2i) - 20} more")
            print()

    # ===================================================================
    print("=" * 70)
    print("QC COMPARISON (V22.2 -> V22.3)")
    print("=" * 70)
    print()

    b = data['baseline']
    p = data['projected']

    print(f"  {'Metric':<30s} {'V22.2':>10s} {'V22.3':>10s} {'Delta':>10s}")
    print(f"  {'-'*30} {'-'*10} {'-'*10} {'-'*10}")
    print(f"  {'Total facilities':<30s} {b['total']:>10,} {p['total']:>10,} "
          f"{p['total']-b['total']:>+10,}")
    print(f"  {'Corporate':<30s} {b['corporate']:>10,} {p['corporate']:>10,} "
          f"{p['corporate']-b['corporate']:>+10,}")
    print(f"  {'Independent':<30s} {b['independent']:>10,} {p['independent']:>10,} "
          f"{p['independent']-b['independent']:>+10,}")
    print(f"  {'Distinct Corporate_Names':<30s} {b['distinct_corps']:>10,} "
          f"{p['distinct_corps']:>10,} "
          f"{p['distinct_corps']-b['distinct_corps']:>+10,}")
    print()
    print(f"  SNF/ALF/ILF counts: unchanged (no type reclassification)")
    print(f"  Served count: unchanged (no service flag changes)")
    print()

    # ===================================================================
    total_corp_changes = len(data['corp_changes'])
    total_changes = (total_corp_changes + len(data['removal_rows'])
                     + len(data['ownership_changes']))
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print()
    print(f"  Phase 1d (Liberty SNF dupe):   {len(data['phase1d']):,} removed")
    print(f"  Phase 2a (GLR corrections):    {len(data['phase2a_changes']):,} renames")
    print(f"  Phase 2b (CMS corrections):    {len(data['phase2b_changes']):,} renames")
    print(f"  Phase 2c (NIC Tier A/B):       {len(data['phase2c_auto']):,} renames")
    print(f"  Phase 2c (NIC flagged):        {len(data['phase2c_flagged']):,} (no change)")
    print(f"  Phase 3  (Four-Rule):          {len(data['ownership_changes']):,} reclassified")
    print(f"  ---")
    print(f"  Total Corporate_Name changes:  {total_corp_changes:,}")
    print(f"  Total row removals:            {len(data['removal_rows']):,}")
    print(f"  Total Ownership_Type changes:  {len(data['ownership_changes']):,}")
    print(f"  New column: Corp_Attribution_Source ({sum(data['source_counts'].values()):,} rows)")
    print(f"  Net row count: {b['total']:,} -> {p['total']:,}")
    print()
    print("To apply: python corporate_fix.py apply")


# ---------------------------------------------------------------------------
# Apply mode
# ---------------------------------------------------------------------------

def do_apply():
    print("CORPORATE FIX — APPLY MODE")
    print("Three-tier hierarchy: GLR -> CMS -> NIC")
    print("=" * 70)
    print()
    print(f"Source: {DB_SOURCE.name}")
    print(f"Target: {DB_TARGET.name}")
    print()

    if DB_TARGET.exists():
        print(f"ERROR: {DB_TARGET.name} already exists. Remove it first to re-apply.")
        sys.exit(1)

    data = compute_all_changes()
    corp_changes = data['corp_changes']
    all_renames = data['all_renames']
    removal_rows = data['removal_rows']
    attr_sources = data['attr_sources']

    total_changes = (len(all_renames) + len(removal_rows)
                     + len(data['ownership_changes']))
    if total_changes == 0 and not attr_sources:
        print("No changes to apply.")
        return

    # Load workbook in full mode for writing
    print("Loading workbook (full mode) for writing...")
    wb = load_workbook(DB_SOURCE)
    ws = wb.active
    ws_headers = [safe(c.value) for c in ws[1]]

    ci_corp = ws_headers.index('Corporate_Name') + 1
    ci_own = ws_headers.index('Ownership_Type') + 1
    ci_fac = ws_headers.index('Facility_Name') + 1

    # --- Add Corp_Attribution_Source column ---
    new_col = len(ws_headers) + 1
    ws.cell(row=1, column=new_col).value = 'Corp_Attribution_Source'

    # Set attribution source for every data row
    for row_num in range(2, ws.max_row + 1):
        source = attr_sources.get(row_num, 'LEGACY')
        ws.cell(row=row_num, column=new_col).value = source

    print(f"  Added Corp_Attribution_Source column (col {new_col}).")

    # --- Apply Corporate_Name renames ---
    print(f"\nApplying {len(all_renames):,} Corporate_Name changes...")
    rename_count = 0
    for excel_row, new_corp in sorted(all_renames.items()):
        ws.cell(row=excel_row, column=ci_corp).value = new_corp
        rename_count += 1
    print(f"  {rename_count:,} Corporate_Name values updated.")

    # --- Delete Phase 1d rows (bottom-up) ---
    if removal_rows:
        print(f"\nRemoving {len(removal_rows)} Liberty PROPCO duplicate rows...")
        for excel_row in sorted(removal_rows, reverse=True):
            fac = safe(ws.cell(row=excel_row, column=ci_fac).value)
            ws.delete_rows(excel_row, 1)
            print(f"  Deleted row {excel_row}: {fac}")

    # --- Apply Four-Rule (Phase 3) ---
    print(f"\nPhase 3: Applying Four-Rule ownership reclassification...")

    # Pass 1: Count facilities per Corporate_Name (post-changes, post-deletions)
    corp_counts = Counter()
    for row_num in range(2, ws.max_row + 1):
        corp = safe(ws.cell(row=row_num, column=ci_corp).value)
        if corp and corp.upper() != 'INDEPENDENT':
            corp_counts[corp] += 1

    print(f"  {len(corp_counts):,} distinct Corporate_Name values.")

    # Pass 2: Apply Four-Rule
    reclass_count = 0
    for row_num in range(2, ws.max_row + 1):
        corp = safe(ws.cell(row=row_num, column=ci_corp).value)
        old_own = safe(ws.cell(row=row_num, column=ci_own).value)
        new_own = four_rule_classify(corp, corp_counts)
        if old_own != new_own:
            ws.cell(row=row_num, column=ci_own).value = new_own
            reclass_count += 1

    print(f"  {reclass_count:,} Ownership_Type reclassifications applied.")

    # --- Save ---
    print(f"\nSaving as {DB_TARGET.name}...")
    wb.save(DB_TARGET)
    wb.close()
    print(f"  Saved to: {DB_TARGET}")

    # --- Final counts ---
    print()
    print("Verifying final counts...")
    _, final_rows = load_db(DB_TARGET)
    final_total = len(final_rows)
    final_snf = sum(1 for r in final_rows
                    if safe(r.get('Source_Type', '')) == 'SNF')
    final_alf = sum(1 for r in final_rows
                    if safe(r.get('Source_Type', '')) == 'ALF')
    final_ilf = sum(1 for r in final_rows
                    if safe(r.get('Source_Type', '')) == 'ILF')
    final_served = sum(1 for r in final_rows
                       if safe(r.get('Do_We_Serve', '')) == 'Yes')
    final_corp_ct = sum(1 for r in final_rows
                        if safe(r.get('Ownership_Type', '')) == 'Corporate')
    final_ind_ct = sum(1 for r in final_rows
                       if safe(r.get('Ownership_Type', '')) == 'Independent')
    final_distinct = len(set(
        safe(r.get('Corporate_Name', ''))
        for r in final_rows
        if safe(r.get('Corporate_Name', ''))
        and safe(r.get('Corporate_Name', '')).upper() != 'INDEPENDENT'))

    # Attribution source distribution from final file
    final_src = Counter()
    for r in final_rows:
        src = safe(r.get('Corp_Attribution_Source', ''))
        if src:
            final_src[src] += 1

    b = data['baseline']
    print()
    print(f"  {'Metric':<30s} {'V22.2':>10s} {'V22.3':>10s} {'Delta':>10s}")
    print(f"  {'-'*30} {'-'*10} {'-'*10} {'-'*10}")
    print(f"  {'Total facilities':<30s} {b['total']:>10,} {final_total:>10,} "
          f"{final_total-b['total']:>+10,}")
    print(f"  {'SNF':<30s} {b['snf']:>10,} {final_snf:>10,} "
          f"{final_snf-b['snf']:>+10,}")
    print(f"  {'ALF':<30s} {b['alf']:>10,} {final_alf:>10,} "
          f"{final_alf-b['alf']:>+10,}")
    print(f"  {'ILF':<30s} {b['ilf']:>10,} {final_ilf:>10,} "
          f"{final_ilf-b['ilf']:>+10,}")
    print(f"  {'Served':<30s} {b['served']:>10,} {final_served:>10,} "
          f"{final_served-b['served']:>+10,}")
    print(f"  {'Corporate':<30s} {b['corporate']:>10,} {final_corp_ct:>10,} "
          f"{final_corp_ct-b['corporate']:>+10,}")
    print(f"  {'Independent':<30s} {b['independent']:>10,} {final_ind_ct:>10,} "
          f"{final_ind_ct-b['independent']:>+10,}")
    print(f"  {'Distinct Corporate_Names':<30s} {b['distinct_corps']:>10,} "
          f"{final_distinct:>10,} {final_distinct-b['distinct_corps']:>+10,}")
    print()
    print("  Attribution source distribution:")
    for src in ['GLR', 'CMS', 'NIC-A', 'NIC-B', 'LEGACY']:
        ct = final_src.get(src, 0)
        print(f"    {src:<8s} {ct:>7,}")
    print()

    # --- Update qc_config.json ---
    print("Updating qc_config.json...")
    with open(QC_CONFIG, 'r') as f:
        qc = json.load(f)

    today = __import__('datetime').date.today().isoformat()
    qc['v22_3_actuals'] = {
        '_comment': f'V22.3 actuals after Wave 2c three-tier corporate fix ({today})',
        'total_facilities': final_total,
        'snf_count': final_snf,
        'alf_count': final_alf,
        'ilf_count': final_ilf,
        'served_count': final_served,
        'corporate_count': final_corp_ct,
        'independent_count': final_ind_ct,
        'distinct_corporate_names': final_distinct,
        'corp_name_changes': len(all_renames),
        'rows_removed': len(removal_rows),
        'ownership_reclassified': reclass_count,
        'attribution_sources': dict(final_src),
        'phase2a_glr_changes': len(data['phase2a_changes']),
        'phase2b_cms_changes': len(data['phase2b_changes']),
        'phase2c_nic_changes': len(data['phase2c_auto']),
        'phase2c_flagged': len(data['phase2c_flagged']),
    }

    qc['wave_log'].append({
        'wave': '2c',
        'label': 'Three-tier Corporate Attribution (GLR -> CMS -> NIC) + Four-Rule',
        'status': 'complete',
        'date_started': today,
        'date_completed': today,
        'script': 'corporate_fix.py',
        'output': '1_Combined_Database_FINAL_V22_3.xlsx',
        'notes': (
            f'{len(all_renames):,} Corporate_Name changes: '
            f'{len(data["phase2a_changes"])} GLR, '
            f'{len(data["phase2b_changes"])} CMS, '
            f'{len(data["phase2c_auto"])} NIC. '
            f'{len(removal_rows)} row(s) removed (Liberty PROPCO dupe). '
            f'{reclass_count:,} Ownership_Type reclassifications. '
            f'{len(data["phase2c_flagged"])} NIC rows flagged (Tier C/suspicious). '
            f'New column: Corp_Attribution_Source.'
        ),
    })

    qc['expected']['total_facilities'] = final_total

    with open(QC_CONFIG, 'w') as f:
        json.dump(qc, f, indent=2)
        f.write('\n')
    print(f"  Updated: {QC_CONFIG}")

    # --- Summary ---
    print()
    print("=" * 70)
    print("APPLY COMPLETE")
    print("=" * 70)
    print(f"  Corporate_Name changes:    {len(all_renames):,}")
    print(f"    GLR (Tier 1):            {len(data['phase2a_changes']):,}")
    print(f"    CMS (Tier 2):            {len(data['phase2b_changes']):,}")
    print(f"    NIC (Tier 3):            {len(data['phase2c_auto']):,}")
    print(f"  Rows removed:              {len(removal_rows)}")
    print(f"  Ownership reclassified:    {reclass_count:,}")
    print(f"  New column added:          Corp_Attribution_Source")
    print(f"  Final row count:           {final_total:,}")
    print()
    print("Next steps:")
    print("  1. Review flagged NIC rows (Tier C/suspicious)")
    print("  2. Spot-check GLR and CMS corrections for served facilities")
    print("  3. Update Database_Change_Request_V22.md DCR status")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('preview', 'apply'):
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]
    print(f"Corporate Fix — Wave 2c (V22.2 -> V22.3)")
    print(f"Mode: {mode}")
    print()

    if mode == 'preview':
        do_preview()
    elif mode == 'apply':
        do_apply()


if __name__ == '__main__':
    main()
