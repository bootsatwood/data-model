"""Complete MUO universe — case-insensitive merge across Finance, V20, and BD."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

V20_PATH = "C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx"
V23_FIN_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_MUO_Scoring_Workbook_V23_v8.xlsx"
V23_BD_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_Final_MUO_Tiering_V23.xlsx"

# Canonical display names — resolves all aliases
ALIASES = {
    # V20 -> canonical
    'consulate health care/independence living centers/nspire healthcare/raydiant health care': 'Avardis',
    'liberty senior living': 'Liberty',
    'life care centers of america': 'Lifecare',
    'principle long term care': 'Principle',
    'pruitthealth': 'Pruitt Health',
    'terra bella': 'TerraBella Senior Living',
    'tlc management': 'TLC Management',
    'trilogy health services': 'Trilogy',
    'arbors at ohio': 'Arbors',
    'heritage hall': 'American Healthcare LLC',
    'jag healthcare': 'JAG',
    'kissito': 'Kissito Healthcare',
    'momentous health': 'Momentus Health',
    'peak resources, inc.': 'Peak Resources',
    'priority life care': 'Priority',
    'sanstone health & rehabilitation': 'Sanstone',
    'topaz': 'Topaz Healthcare',
    'yad healthcare': 'YAD',
    'cch healthcare': 'CCH Healthcare',
    'otterbein seniorlife': 'Otterbein Senior Life',
    'lutheran services carolina': 'Lutheran Services Carolinas',
    'gardant management solutions, inc': 'Gardant Management Solutions',
    'seky holding co.': 'SEKY Holding Co.',
    'sunrise': 'Sunrise Senior Living',
    'windsor house, inc.': 'Windsor House',
    'aom healthcare': 'AOM Healthcare',
    'fundamental ltc': 'Fundamental LTC',
    'southern healthcare management, llc': 'Southern Healthcare Mgmt',
    'southern assisted living, llc': '(ABSORBED)',
    'sovereign healthcare holdings': '(ABSORBED)',
    # Finance UPPERCASE -> canonical
    'american senior communities': 'American Senior Communities',
    'infinity healthcare consulting': 'Infinity Healthcare Consulting',
    'lionstone care': 'Lionstone Care',
    'navion': 'Navion',
    'otterbein senior life': 'Otterbein Senior Life',
    'principle': 'Principle',
    'pruitt health': 'Pruitt Health',
    'saber healthcare group': 'Saber Healthcare Group',
    'trilogy': 'Trilogy',
    'arbors': 'Arbors',
    'eldercare partners': 'Eldercare Partners',
    'fundamental ltc': 'Fundamental LTC',
    'morning pointe senior living': 'Morning Pointe Senior Living',
    'peak resources': 'Peak Resources',
    'priority': 'Priority',
    'sanstone': 'Sanstone',
    'terrabella senior living': 'TerraBella Senior Living',
    'topaz healthcare': 'Topaz Healthcare',
    'american healthcare llc': 'American Healthcare LLC',
    'atrium health': 'Atrium Health',
    'caring place healthcare': 'Caring Place Healthcare',
    'clearview': 'Clearview',
    'communicare': 'CommuniCare',
    'cardon & associates': 'Cardon & Associates',
    'signature health': 'Signature Health',
    'singh': 'Singh',
    'sonida senior living': 'Sonida Senior Living',
    'storypoint': 'StoryPoint',
    'cedarhurst senor living': 'Cedarhurst Senior Living',
    'cedarhurst senior living': 'Cedarhurst Senior Living',
    'spring arbor management': 'Spring Arbor Management',
    'senior lifestyle': 'Senior Lifestyle',
    'mcap': 'MCAP',
    # V20 T5 short names -> canonical
    'bluegrass': 'Bluegrass/Encore',
    'encore': 'Bluegrass/Encore',
    'bluegrass/encore': 'Bluegrass/Encore',
    'signature': 'Signature Health',
    'mfa': 'MFA',
    'communicare': 'CommuniCare',
    'hill valley': 'Hill Valley',
    'cardon': 'Cardon & Associates',
    'eastern': 'Eastern Healthcare Group',
    'eastern healthcare group': 'Eastern Healthcare Group',
    'embassy': 'Embassy',
    'exceptional living': 'Exceptional Living',
    'portopiccolo': 'Portopiccolo',
    'venza': 'Venza',
    'aventura': 'Aventura',
    'journey': 'Journey',
    'pavilion healthcare': 'Pavilion Healthcare',
}

def canonicalize(name):
    name = name.strip()
    low = name.lower()
    if low in ALIASES:
        return ALIASES[low]
    return name

# ============================================================
# READ ALL SOURCES
# ============================================================
print("Reading V20...")
wb_v20 = openpyxl.load_workbook(V20_PATH, data_only=True)

v20_entities = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_v20[sn]
    tier = sn.split()[0]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip().upper() == 'TOTAL':
            continue
        canon = canonicalize(str(name))
        if canon == '(ABSORBED)':
            continue
        score = ws.cell(r, 3).value or 0
        v20_entities[canon] = {'tier': tier, 'score': int(score) if score else 0}

v20_t5 = {}
for sn in wb_v20.sheetnames:
    if 'T5' in sn or 'Barrier' in sn:
        ws = wb_v20[sn]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip().upper() != 'TOTAL':
                canon = canonicalize(str(name))
                v20_t5[canon] = {'tier': 'T5'}

print(f"  V20: {len(v20_entities)} scored + {len(v20_t5)} T5 = {len(v20_entities) + len(v20_t5)}")

print("\nReading Finance V23...")
wb_fin = openpyxl.load_workbook(V23_FIN_PATH, data_only=True)

fin_entities = {}
ws_sum = wb_fin['Summary']
for r in range(2, ws_sum.max_row + 1):
    name = ws_sum.cell(r, 1).value
    if not name or str(name).strip() in ('', 'TOTAL'):
        continue
    canon = canonicalize(str(name))
    score = ws_sum.cell(r, 15).value
    tier = ws_sum.cell(r, 16).value
    campuses = ws_sum.cell(r, 17).value
    fin_entities[canon] = {
        'tier': str(tier).strip() if tier else '?',
        'score': int(score) if score else 0,
        'campuses': int(campuses) if campuses else 0,
    }

ws_t4 = wb_fin['T4 - Independents']
for r in range(2, ws_t4.max_row + 1):
    name = ws_t4.cell(r, 1).value
    if not name:
        continue
    name_str = str(name).strip()
    if name_str in ('', 'TOTAL', 'FACILITY DETAIL', 'Finance Tier'):
        continue
    # Skip non-entity rows
    col2 = ws_t4.cell(r, 2).value
    if col2 and str(col2).strip() in ('Finance Tier', 'Qualifying Campuses'):
        continue
    canon = canonicalize(name_str)
    if canon not in fin_entities:
        camps = ws_t4.cell(r, 3).value
        if camps is None:
            camps = ws_t4.cell(r, 2).value
        fin_entities[canon] = {
            'tier': 'T4',
            'score': None,
            'campuses': int(camps) if isinstance(camps, (int, float)) else 0,
        }

fin_t5 = {'Bluegrass/Encore', 'Signature Health', 'MFA', 'CommuniCare', 'Hill Valley', 'Singh',
           'Cardon & Associates', 'Eastern Healthcare Group', 'Clearview', 'Pavilion Healthcare'}
for name in fin_t5:
    if name not in fin_entities:
        fin_entities[name] = {'tier': 'T5', 'score': None, 'campuses': 0}

print(f"  Finance: {len(fin_entities)} entities")

print("\nReading BD V23...")
wb_bd = openpyxl.load_workbook(V23_BD_PATH, data_only=True)

bd_entities = {}
for sn in wb_bd.sheetnames:
    ws = wb_bd[sn]
    if any(t in sn for t in ['T1', 'T2', 'T3']) and 'T4' not in sn and 'T5' not in sn:
        tier = sn.split()[0]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL'):
                canon = canonicalize(str(name))
                score = ws.cell(r, 3).value or 0
                bd_entities[canon] = {'tier': tier, 'score': int(score) if score else 0}
    elif 'T4' in sn:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL', 'FACILITY DETAIL'):
                canon = canonicalize(str(name))
                camps = ws.cell(r, 2).value
                if isinstance(camps, (int, float)) and camps >= 0:
                    bd_entities[canon] = {'tier': 'T4', 'campuses': int(camps)}
    elif 'T5' in sn or 'Barrier' in sn:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() not in ('', 'TOTAL'):
                canon = canonicalize(str(name))
                bd_entities[canon] = {'tier': 'T5'}

print(f"  BD: {len(bd_entities)} entities")

# ============================================================
# MERGE
# ============================================================
all_names = set()
all_names.update(v20_entities.keys())
all_names.update(v20_t5.keys())
all_names.update(fin_entities.keys())
all_names.update(bd_entities.keys())

print(f"\n{'='*130}")
print(f"COMPLETE MUO UNIVERSE: {len(all_names)} unique entities")
print(f"{'='*130}")

# Build combined records
records = []
for name in sorted(all_names):
    in_v20 = name in v20_entities or name in v20_t5
    in_fin = name in fin_entities
    in_bd = name in bd_entities

    # Determine V23 tier (Finance is authoritative, then BD, then V20)
    if in_fin:
        tier = fin_entities[name]['tier']
        score = fin_entities[name].get('score')
        campuses = fin_entities[name].get('campuses', 0)
    elif in_bd:
        tier = bd_entities[name]['tier']
        score = bd_entities[name].get('score')
        campuses = bd_entities[name].get('campuses', 0)
    else:
        # V20 only — need to determine what V23 would assign
        if name in v20_t5:
            tier = 'T5'
            score = None
            campuses = 0
        else:
            tier = v20_entities[name]['tier'] + '?'
            score = v20_entities[name].get('score')
            campuses = 0

    records.append({
        'name': name,
        'tier': tier,
        'score': score,
        'campuses': campuses,
        'in_fin': in_fin,
        'in_v20': in_v20,
        'in_bd': in_bd,
    })

# Sort by tier then score then name
tier_order = {'T1': 1, 'T2': 2, 'T3': 3, 'T4': 4, 'T5': 5}
records.sort(key=lambda x: (tier_order.get(x['tier'], 6), -(x['score'] or 0), x['name']))

# Print
current_tier = None
tier_counts = {}
missing_from_bd = []

print(f"\n{'Entity':<48} {'Tier':>4} {'Score':>5} {'Camps':>5}  {'Fin':>3} {'V20':>3} {'BD':>3}  Status")
print('-' * 115)

for rec in records:
    if rec['tier'] != current_tier:
        if current_tier:
            print()
        current_tier = rec['tier']

    tier_counts[rec['tier']] = tier_counts.get(rec['tier'], 0) + 1

    s = str(rec['score']) if rec['score'] else '-'
    c = str(rec['campuses']) if rec['campuses'] else '-'
    f = 'Y' if rec['in_fin'] else '.'
    v = 'Y' if rec['in_v20'] else '.'
    b = 'Y' if rec['in_bd'] else '.'

    status = ''
    if rec['in_fin'] and not rec['in_bd']:
        status = '<< MISSING FROM BD'
        missing_from_bd.append(rec)
    elif not rec['in_fin'] and rec['in_bd']:
        status = '(V20/BD only)'
    elif rec['in_fin'] and rec['in_v20'] and rec['in_bd']:
        status = ''

    print(f"  {rec['name']:<46} {rec['tier']:>4} {s:>5} {c:>5}  {f:>3} {v:>3} {b:>3}  {status}")

print(f"\n{'='*130}")
print(f"TIER TOTALS (Combined Universe)")
print(f"{'='*130}")
for tier in ['T1', 'T2', 'T3', 'T4', 'T5']:
    print(f"  {tier}: {tier_counts.get(tier, 0)}")
print(f"  Total: {sum(tier_counts.values())}")

print(f"\n{'='*130}")
print(f"COVERAGE ANALYSIS")
print(f"{'='*130}")
all_in_fin = sum(1 for r in records if r['in_fin'])
all_in_v20 = sum(1 for r in records if r['in_v20'])
all_in_bd = sum(1 for r in records if r['in_bd'])
in_both_fin_bd = sum(1 for r in records if r['in_fin'] and r['in_bd'])
print(f"  In Finance workbook:  {all_in_fin}")
print(f"  In V20 workbook:      {all_in_v20}")
print(f"  In BD workbook:       {all_in_bd}")
print(f"  In Finance AND BD:    {in_both_fin_bd}")
print(f"  In Finance NOT BD:    {all_in_fin - in_both_fin_bd}  << GAP TO CLOSE")
print(f"  In BD NOT Finance:    {all_in_bd - in_both_fin_bd}  (V20-only gap entities)")

print(f"\n{'='*130}")
print(f"FINANCE ENTITIES MISSING FROM BD ({len(missing_from_bd)})")
print(f"{'='*130}")
scored_missing = [r for r in missing_from_bd if r['tier'] in ('T1', 'T2', 'T3')]
t4_missing = [r for r in missing_from_bd if r['tier'] == 'T4']
t5_missing = [r for r in missing_from_bd if r['tier'] == 'T5']

if scored_missing:
    print(f"\n  SCORED ({len(scored_missing)}) — these need full 6-dimension rows added to T1/T2/T3 tabs:")
    for r in scored_missing:
        s = str(r['score']) if r['score'] else '-'
        print(f"    {r['name']:<46} {r['tier']}  score {s}")

if t4_missing:
    print(f"\n  T4 INDEPENDENTS ({len(t4_missing)}) — add to T4 tab:")
    for r in t4_missing:
        c = str(r['campuses']) if r['campuses'] else '?'
        print(f"    {r['name']:<46} {c} campuses")

if t5_missing:
    print(f"\n  T5 BARRIERS ({len(t5_missing)}) — add to T5 tab:")
    for r in t5_missing:
        print(f"    {r['name']}")

# One-off check
print(f"\n{'='*130}")
print(f"BROOKE'S 4 ONE-OFF PROFILES")
print(f"{'='*130}")
for name in ['Liberty', 'Southern Healthcare Mgmt', 'Triple Crown Senior Living', 'Lutheran Life Villages']:
    if name in {r['name'] for r in records}:
        rec = next(r for r in records if r['name'] == name)
        f = 'Y' if rec['in_fin'] else 'N'
        b = 'Y' if rec['in_bd'] else 'N'
        print(f"  {name:<40} Tier: {rec['tier']}  Finance: {f}  BD: {b}")
    else:
        # Try partial match
        matches = [r for r in records if name.lower() in r['name'].lower()]
        if matches:
            for m in matches:
                f = 'Y' if m['in_fin'] else 'N'
                b = 'Y' if m['in_bd'] else 'N'
                print(f"  {name:<40} -> {m['name']}: Tier {m['tier']}  Finance: {f}  BD: {b}")
        else:
            print(f"  {name:<40} ** NOT FOUND — may need to be added to DB / Finance universe **")
