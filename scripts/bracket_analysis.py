"""Show current ER/RP distributions and test tighter ceilings."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'Sunnyside Communities', 'ATRIUM HEALTH', 'Warm Hearth Village', 'Brighton',
    'Momentus Health', 'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
}

FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
}

# RS scores from Tom's notes
TOM_RS = {
    'ALG': 2, 'AMERICAN SENIOR COMMUNITIES': 5, 'Brookdale Senior Living': 5,
    'SABER HEALTHCARE GROUP': 2, 'INFINITY HEALTHCARE CONSULTING': 2, 'NAVION': 3,
    'Majestic Care': 3, 'PRUITT HEALTH': 3, 'Kisco Senior Living': 1, 'TRILOGY': 4,
    'TOPAZ HEALTHCARE': 1, 'MORNING POINTE SENIOR LIVING': 2, 'PRINCIPLE': 3,
    'Liberty': 4, 'TERRABELLA SENIOR LIVING': 2, 'SANSTONE': 2, 'LIONSTONE CARE': 3,
    'ELDERCARE PARTNERS': 3, 'OTTERBEIN SENIOR LIFE': 1, 'TLC Management': 3,
    'BHI Senior Living': 1, 'Avardis': 2, 'PEAK RESOURCES': 2, 'ARBORS': 1,
    'AMERICAN HEALTHCARE LLC': 3, 'Runk & Pratt': 1, 'MCAP': 1,
    'Lutheran Services Carolinas': 4, 'Lutheran Life Villages': 4, 'Greencroft': 1,
    'CCH HEALTHCARE': 2, 'LifeSpire of Virginia': 1, 'SENIOR LIFESTYLE': 1,
    'PRIORITY': 1, 'CEDARHURST SENOR LIVING': 1, 'Lifecare': 1,
    'Kissito Healthcare': 2, 'SPRING ARBOR MANAGEMENT': 1, 'YAD': 2,
    'STORYPOINT': 3, 'CARING PLACE HEALTHCARE': 2, 'SONIDA SENIOR LIVING': 1,
    'Castle Healthcare': 3, 'JAG': 2, 'FUNDAMENTAL LTC': 1,
}

# SI scores from research
DATA_SI = {
    'ALG': 5, 'Brookdale Senior Living': 5, 'SABER HEALTHCARE GROUP': 5,
    'TRILOGY': 5, 'Liberty': 5, 'PRUITT HEALTH': 5, 'Avardis': 5,
    'SONIDA SENIOR LIVING': 5, 'Lifecare': 5,
    'AMERICAN SENIOR COMMUNITIES': 4, 'INFINITY HEALTHCARE CONSULTING': 4,
    'Majestic Care': 4, 'NAVION': 4, 'Kisco Senior Living': 4,
    'TLC Management': 4, 'Lutheran Life Villages': 4, 'Lutheran Services Carolinas': 4,
    'SANSTONE': 3, 'MORNING POINTE SENIOR LIVING': 3, 'OTTERBEIN SENIOR LIFE': 3,
    'PRINCIPLE': 3, 'LIONSTONE CARE': 3, 'Kissito Healthcare': 3,
    'CCH HEALTHCARE': 3, 'TERRABELLA SENIOR LIVING': 3, 'PEAK RESOURCES': 3,
    'BHI Senior Living': 3,
}

# Load Finance entities
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']
finance_names = set()
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in EXCLUDE:
        finance_names.add(str(name).strip())

all_db = set()
db_to_fin = {}
for fn, dns in FINANCE_TO_DB.items():
    if fn not in finance_names:
        continue
    for dn in dns:
        all_db.add(dn)
        db_to_fin[dn] = fn

wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

corp_fp = defaultdict(lambda: {'campuses': set(), 'served_camps': set(), 'tot_pot': 0,
                                'mh_camps': set(), 'pcp_camps': set(), 'integ_camps': set()})
for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if corp not in all_db or state not in FOOTPRINT:
        continue
    fn = db_to_fin[corp]
    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    tot = ws.cell(r, headers['Total_Potential_Revenue']).value or 0
    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws.cell(r, headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws.cell(r, headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws.cell(r, headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')

    d = corp_fp[fn]
    campus = (addr, city)
    d['campuses'].add(campus)
    if isinstance(tot, (int, float)):
        d['tot_pot'] += tot
    if served: d['served_camps'].add(campus)
    if mh: d['mh_camps'].add(campus)
    if pcp: d['pcp_camps'].add(campus)
    if integ: d['integ_camps'].add(campus)

# Build entities with all V5 scores except ER/RP (which we'll vary)
entities = []
for fn in finance_names:
    d = corp_fp.get(fn)
    if not d or len(d['campuses']) < 7:
        continue

    n_camp = len(d['campuses'])
    n_srv = len(d['served_camps'])

    # IR inference
    n_integ = len(d['integ_camps'])
    n_mh = len(d['mh_camps'])
    n_pcp = len(d['pcp_camps'])
    dual = len(d['mh_camps'] & d['pcp_camps'])
    int_pct = n_integ / n_camp if n_camp > 0 else 0

    if n_srv == 0: ir = 1
    elif int_pct >= 0.5: ir = 5
    elif int_pct >= 0.25: ir = 4
    elif dual > 0: ir = 4
    elif n_mh > 0 and n_pcp > 0: ir = 3
    elif n_mh > 0 or n_pcp > 0: ir = 2
    else: ir = 1

    si = DATA_SI.get(fn, 2)
    rs = TOM_RS.get(fn, 1)
    ai = 0  # simplified — most are 0

    # Fixed portion of score (IR + SI + RS + AI, weighted)
    fixed = (ir * 3) + (si * 3) + (rs * 3) + (ai * 3)

    entities.append({
        'name': fn, 'camp': n_camp, 'rev': d['tot_pot'],
        'ir': ir, 'si': si, 'rs': rs, 'ai': ai, 'fixed': fixed,
    })

print(f"Scoreable: {len(entities)} entities")
print()

# Current brackets
def sc(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t: return i + 1
    return 5

# Show current ER/RP distributions
camps = sorted([e['camp'] for e in entities])
revs = sorted([e['rev'] for e in entities])
n = len(entities)
print("=== CAMPUS DISTRIBUTION ===")
print(f"Min={camps[0]}  P25={camps[n//4]}  Med={camps[n//2]}  P75={camps[3*n//4]}  Max={camps[-1]}")

er_cur = Counter(sc(e['camp'], [9,13,20,40]) for e in entities)
print(f"\nCurrent ER (9/13/20/40):")
for s in range(1, 6):
    print(f"  ER={s}: {er_cur.get(s,0):>3}  {'#' * er_cur.get(s,0)}")

print(f"\n=== REVENUE DISTRIBUTION ===")
print(f"Min=${revs[0]:,.0f}  P25=${revs[n//4]:,.0f}  Med=${revs[n//2]:,.0f}  P75=${revs[3*n//4]:,.0f}  Max=${revs[-1]:,.0f}")

rp_cur = Counter(sc(e['rev'], [1e6,2.5e6,5e6,10e6]) for e in entities)
print(f"\nCurrent RP ($1M/$2.5M/$5M/$10M):")
for s in range(1, 6):
    print(f"  RP={s}: {rp_cur.get(s,0):>3}  {'#' * rp_cur.get(s,0)}")

# Test tighter bracket options
print("\n" + "=" * 95)
print("BRACKET OPTIONS — Tier outcomes (V20 target: T1~22, T2~36, T3~12)")
print("=" * 95)

er_opts = {
    'V5 current (9/13/20/40)':      [9, 13, 20, 40],
    'Tighter-A  (10/15/25/50)':     [10, 15, 25, 50],
    'Tighter-B  (12/18/30/60)':     [12, 18, 30, 60],
    'Tighter-C  (10/20/35/60)':     [10, 20, 35, 60],
    'Cap at 4   (10/15/25/999)':    [10, 15, 25, 999],
}

rp_opts = {
    'V5 current ($1M/2.5/5/10)':    [1e6, 2.5e6, 5e6, 10e6],
    'Tighter-A  ($1.5/3.5/7/15)':   [1.5e6, 3.5e6, 7e6, 15e6],
    'Tighter-B  ($2/5/10/20)':      [2e6, 5e6, 10e6, 20e6],
    'Tighter-C  ($2/4/8/15)':       [2e6, 4e6, 8e6, 15e6],
    'Cap at 4   ($1.5/3/6/999M)':   [1.5e6, 3e6, 6e6, 999e6],
}

print(f"\n{'ER Brackets':<28} {'RP Brackets':<30} {'T1':>4} {'T2':>4} {'T3':>4}  Shape")
print("-" * 90)

for er_name, er_t in er_opts.items():
    for rp_name, rp_t in rp_opts.items():
        tiers = Counter()
        for e in entities:
            er = sc(e['camp'], er_t)
            rp = sc(e['rev'], rp_t)
            score = (er * 4) + (rp * 4) + e['fixed']
            tier = 'T1' if score >= 50 else 'T2' if score >= 25 else 'T3'
            tiers[tier] += 1

        t1, t2, t3 = tiers.get('T1',0), tiers.get('T2',0), tiers.get('T3',0)
        if t3 == 0:
            shape = "no T3"
        elif t1 > 30:
            shape = "top-heavy"
        elif t3 > 10:
            shape = "GOOD spread"
        elif t3 >= 5:
            shape = "decent"
        else:
            shape = f"thin T3"

        # Only show combos that produce T3
        if t3 > 0:
            print(f"{er_name:<28} {rp_name:<30} {t1:>4} {t2:>4} {t3:>4}  {shape}")

# Highlight best combos
print()
print("=" * 95)
print("BEST CANDIDATES (closest to V20 shape: ~30% T1, ~50% T2, ~20% T3)")
print("=" * 95)
best = []
for er_name, er_t in er_opts.items():
    for rp_name, rp_t in rp_opts.items():
        tiers = Counter()
        for e in entities:
            er = sc(e['camp'], er_t)
            rp = sc(e['rev'], rp_t)
            score = (er * 4) + (rp * 4) + e['fixed']
            tier = 'T1' if score >= 50 else 'T2' if score >= 25 else 'T3'
            tiers[tier] += 1
        t1, t2, t3 = tiers.get('T1',0), tiers.get('T2',0), tiers.get('T3',0)
        # Score how close to V20 proportions (49% T1, 51% T2, 17% T3 of 70)
        # For our 45: ~13 T1, ~23 T2, ~9 T3 would be ideal
        dist = abs(t1/45 - 0.30) + abs(t2/45 - 0.50) + abs(t3/45 - 0.20)
        best.append((dist, er_name, rp_name, t1, t2, t3))

best.sort()
print(f"\n{'ER Brackets':<28} {'RP Brackets':<30} {'T1':>4} {'T2':>4} {'T3':>4}  {'Fit':>6}")
print("-" * 90)
for dist, er_name, rp_name, t1, t2, t3 in best[:8]:
    print(f"{er_name:<28} {rp_name:<30} {t1:>4} {t2:>4} {t3:>4}  {dist:.3f}")
