"""What happens to V23 tiers if we swap Tom's RS back to Brooke's V20 RS?

RS weight = x3, so each RS point = 3 total-score points.
T1 >= 55, T2 = 35-54, T3 < 35.

We don't need to rerun the full pipeline -- just adjust the total score
by (brooke_rs - tom_rs) * 3 and check for tier boundary crossings."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

RS_WEIGHT = 3

# V23 current scores and tiers (from build output / workbook)
# Need: entity, v23_total_score, v23_tier, tom_rs
# I'll read these from the Final MUO Tiering V23 workbook (BD universe)
# But if locked, use the scoring workbook data.

# Actually let's just read the V20 tiering for Brooke's RS
# and the build_final_muo_tiering.py for V23 scores.
# The BD tiering workbook has the scores in it.

import openpyxl

# Try reading the Final MUO Tiering V23
BD_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_Final_MUO_Tiering_V23.xlsx"

try:
    wb = openpyxl.load_workbook(BD_PATH, read_only=True, data_only=True)
except PermissionError:
    print("BD workbook is locked (open in Excel). Close it and retry.")
    sys.exit(1)

# Read all scored entities from T1, T2, T3 tabs
v23_entities = {}
for sn in wb.sheetnames:
    if sn.startswith('T1') or sn.startswith('T2') or sn.startswith('T3'):
        ws = wb[sn]
        # Check headers
        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                headers[str(val).strip()] = c

        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            name = str(name).strip()

            # Fixed column positions: col2=tier, col3=total_score, col13=RS
            tier = ws.cell(r, 2).value
            total = ws.cell(r, 3).value
            rs = ws.cell(r, 13).value

            if total and rs:
                v23_entities[name] = {
                    'tier': tier or sn[:2],
                    'score': total,
                    'rs': rs,
                }

wb.close()

# Brooke V20 RS
V20_RS = {
    'ALG': 3,
    'American Senior Communities': 5,
    'Brookdale Senior Living': 5,
    'CCH Healthcare': 3, # was 'Cch Healthcare'
    'Ciena Healthcare': 1, # was 'Ciena Healthcare/Laurel Health Care'
    'Avardis': 3, # was Consulate/Nspire/etc
    'Infinity Healthcare Consulting': 1,
    'Liberty': 3, # was 'Liberty Senior Living'
    'Lifecare': 1, # was 'Life Care Centers Of America'
    'Majestic Care': 5,
    'National Healthcare Corp': 1, # was 'National Healthcare Corporation'
    'Navion': 3,
    'Otterbein Senior Life': 5, # was 'Otterbein Seniorlife'
    'PACS Group': 1,
    'Pavilion Healthcare': 4,
    'Principle': 3, # was 'Principle Long Term Care'
    'Pruitt Health': 4, # was 'Pruitthealth'
    'Saber Healthcare Group': 5,
    'Sunrise Senior Living': 3,
    'TerraBella Senior Living': 3, # was 'Terra Bella'
    'TLC Management': 3, # was 'Tlc Management'
    'Trilogy': 4, # was 'Trilogy Health Services'
    'AOM Healthcare': 3, # was 'Aom Healthcare'
    'Aperion Care': 3,
    'Arbors': 4, # was 'Arbors At Ohio'
    'BHI Senior Living': 1,
    'Brickyard Healthcare': 1,
    'Carespring': 1,
    'Castle Healthcare': 3,
    'Clearview': 1,
    'HCF Management': 1,
    'American Healthcare LLC': 3, # was 'Heritage Hall'
    'JAG': 3, # was 'Jag Healthcare'
    'Kissito Healthcare': 1, # was 'Kissito'
    'Lionstone Care': 3,
    'Lutheran Services Carolinas': 5, # was 'Lutheran Services Carolina'
    'Morning Pointe Senior Living': 3,
    'Ohio Living Communities': 3,
    'Peak Resources': 3, # was 'Peak Resources, Inc.'
    'Phoenix Senior Living': 1,
    'Priority': 3, # was 'Priority Life Care'
    'Sanstone': 3, # was 'Sanstone Health & Rehabilitation'
    'Southern Healthcare Mgmt': 1,
    'Topaz Healthcare': 1, # was 'Topaz'
    'YAD': 1, # was 'Yad Healthcare'
    'Carecore Health': 1,
    'Trio Healthcare': 1,
    'Caring Place Healthcare': 1,
    'Envive Healthcare': 1,
    "Miller's Merry Manor": 1,
}

# Match V23 entities to V20 RS and compute impact
# Need to fuzzy-match names between V23 workbook and our V20 dict
def normalize(s):
    return s.upper().strip()

# Build lookup
v20_lookup = {normalize(k): (k, v) for k, v in V20_RS.items()}

print(f"V23 scored entities found: {len(v23_entities)}")
print()

results = []
for v23_name, data in v23_entities.items():
    v23_score = data['score']
    v23_tier = data['tier']
    tom_rs = data['rs']

    # Try to match to V20
    norm = normalize(v23_name)
    match = v20_lookup.get(norm)

    if not match:
        # Try partial matching
        for v20_norm, (v20_name, brooke_rs) in v20_lookup.items():
            if v20_norm in norm or norm in v20_norm:
                match = (v20_name, brooke_rs)
                break

    if match:
        v20_name, brooke_rs = match
        rs_delta = brooke_rs - tom_rs
        score_impact = rs_delta * RS_WEIGHT
        new_score = v23_score + score_impact
        new_tier = 'T1' if new_score >= 55 else 'T2' if new_score >= 35 else 'T3'
        tier_changed = new_tier != v23_tier
        results.append((v23_name, v23_tier, v23_score, tom_rs, brooke_rs, rs_delta, score_impact, new_score, new_tier, tier_changed))
    else:
        # No V20 match -- Brooke didn't score them
        results.append((v23_name, v23_tier, v23_score, tom_rs, None, 0, 0, v23_score, v23_tier, False))

# Sort: tier changers first, then by score impact
results.sort(key=lambda x: (not x[9], -abs(x[6]), x[0]))

print(f"{'Entity':<40} {'V23Tier':>7} {'V23Scr':>6} {'TomRS':>5} {'BrkRS':>5} {'Delta':>5} {'Impact':>6} {'NewScr':>6} {'NewTier':>7} {'SHIFT':>6}")
print("=" * 110)

for r in results:
    entity, v23_tier, v23_score, tom_rs, brooke_rs, rs_delta, score_impact, new_score, new_tier, tier_changed = r
    brk_str = str(brooke_rs) if brooke_rs is not None else 'n/a'
    delta_str = f"{rs_delta:+d}" if rs_delta != 0 else ' 0'
    impact_str = f"{score_impact:+d}" if score_impact != 0 else ' 0'
    shift = f">> {new_tier}" if tier_changed else ''
    print(f"{entity:<40} {v23_tier:>7} {v23_score:>6} {tom_rs:>5} {brk_str:>5} {delta_str:>5} {impact_str:>6} {new_score:>6} {new_tier:>7} {shift:>6}")

# Summary
changers = [r for r in results if r[9]]
print()
if changers:
    print(f"TIER SHIFTS: {len(changers)} entities would change tier")
    for r in changers:
        entity, v23_tier, v23_score, tom_rs, brooke_rs, rs_delta, score_impact, new_score, new_tier, _ = r
        print(f"  {entity:<38} {v23_tier} (score {v23_score}) -> {new_tier} (score {new_score})  [RS: Tom={tom_rs} -> Brooke={brooke_rs}, impact={score_impact:+d}]")
else:
    print("NO TIER SHIFTS -- all entities stay in same tier")
