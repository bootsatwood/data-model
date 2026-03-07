"""Full tier spread for both normalization options + 15-bed rule."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict, Counter

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
TIERING_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V20.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15

T5_EXCLUDE = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
}

FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'}, 'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS'}, 'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'}, 'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'}, 'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'}, 'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'}, 'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'}, 'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'}, 'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'}, 'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'}, 'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'}, 'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'}, 'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'PRIORITY': {'PRIORITY LIFE CARE'}, 'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'}, 'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'}, 'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'}, 'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'}, 'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Triple Crown': {'TRIPLE CROWN'}, 'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'}, 'Warm Hearth Village': {'WARM HEARTH VILLAGE', 'WARM HEARTH'},
    'Brighton': {'BRIGHTON'}, 'Momentus Health': {'MOMENTUS HEALTH'},
}

TOM_RS_SCORES = {
    'ALG': 2, 'AMERICAN SENIOR COMMUNITIES': 5, 'Brookdale Senior Living': 5,
    'SABER HEALTHCARE GROUP': 2, 'INFINITY HEALTHCARE CONSULTING': 2, 'NAVION': 3,
    'Majestic Care': 3, 'PRUITT HEALTH': 3, 'Kisco Senior Living': 1, 'TRILOGY': 4,
    'TOPAZ HEALTHCARE': 1, 'MORNING POINTE SENIOR LIVING': 2, 'PRINCIPLE': 3,
    'Liberty': 4, 'TERRABELLA SENIOR LIVING': 2, 'SANSTONE': 2, 'LIONSTONE CARE': 3,
    'ELDERCARE PARTNERS': 3, 'OTTERBEIN SENIOR LIFE': 1, 'TLC Management': 3,
    'BHI Senior Living': 1, 'Avardis': 2, 'PEAK RESOURCES': 2, 'ARBORS': 1,
    'AMERICAN HEALTHCARE LLC': 3, 'Runk & Pratt': 1, 'MCAP': 1,
    'Lutheran Services Carolinas': 4, 'Lutheran Life Villages': 4, 'Greencroft': 1,
    'CCH HEALTHCARE': 2, 'LifeSpire of Virginia': 1, 'SENIOR LIFESTYLE': 1,
    'PRIORITY': 1, 'CEDARHURST SENOR LIVING': 1, 'Lifecare': 1,
    'Kissito Healthcare': 2, 'SPRING ARBOR MANAGEMENT': 1, 'YAD': 2, 'STORYPOINT': 3,
    'CARING PLACE HEALTHCARE': 2, 'SONIDA SENIOR LIVING': 1, 'Castle Healthcare': 3,
    'JAG': 2, 'FUNDAMENTAL LTC': 1, 'Southern Healthcare Mgmt': 1, 'Triple Crown': 1,
}
DATA_SI_SCORES = {
    'ALG': 5, 'Brookdale Senior Living': 5, 'SABER HEALTHCARE GROUP': 5, 'TRILOGY': 5,
    'Liberty': 5, 'PRUITT HEALTH': 5, 'Avardis': 5, 'SONIDA SENIOR LIVING': 5, 'Lifecare': 5,
    'AMERICAN SENIOR COMMUNITIES': 4, 'INFINITY HEALTHCARE CONSULTING': 4, 'Majestic Care': 4,
    'NAVION': 4, 'Kisco Senior Living': 4, 'TLC Management': 4, 'Lutheran Life Villages': 4,
    'Lutheran Services Carolinas': 4,
    'SANSTONE': 3, 'MORNING POINTE SENIOR LIVING': 3, 'OTTERBEIN SENIOR LIFE': 3,
    'PRINCIPLE': 3, 'LIONSTONE CARE': 3, 'Kissito Healthcare': 3, 'CCH HEALTHCARE': 3,
    'TERRABELLA SENIOR LIVING': 3, 'PEAK RESOURCES': 3, 'BHI Senior Living': 3,
}

FINANCE_TO_V20 = {
    'ALG': 'ALG', 'AMERICAN SENIOR COMMUNITIES': 'American Senior Communities',
    'Brookdale Senior Living': 'Brookdale Senior Living', 'SABER HEALTHCARE GROUP': 'Saber Healthcare Group',
    'INFINITY HEALTHCARE CONSULTING': 'Infinity Healthcare Consulting', 'NAVION': 'Navion',
    'Majestic Care': 'Majestic Care', 'PRUITT HEALTH': 'Pruitthealth', 'TRILOGY': 'Trilogy Health Services',
    'TOPAZ HEALTHCARE': 'Topaz', 'MORNING POINTE SENIOR LIVING': 'Morning Pointe Senior Living',
    'PRINCIPLE': 'Principle Long Term Care', 'Liberty': 'Liberty Senior Living',
    'TERRABELLA SENIOR LIVING': 'Terra Bella', 'SANSTONE': 'Sanstone Health & Rehabilitation',
    'LIONSTONE CARE': 'Lionstone Care', 'OTTERBEIN SENIOR LIFE': 'Otterbein Seniorlife',
    'TLC Management': 'Tlc Management', 'BHI Senior Living': 'BHI Senior Living',
    'Avardis': 'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care',
    'PEAK RESOURCES': 'Peak Resources, Inc.', 'ARBORS': 'Arbors At Ohio',
    'Lutheran Services Carolinas': 'Lutheran Services Carolina', 'CCH HEALTHCARE': 'Cch Healthcare',
    'PRIORITY': 'Priority Life Care', 'Lifecare': 'Life Care Centers Of America',
    'Kissito Healthcare': 'Kissito', 'YAD': 'Yad Healthcare',
    'CARING PLACE HEALTHCARE': 'Caring Place Healthcare', 'Castle Healthcare': 'Castle Healthcare',
    'JAG': 'Jag Healthcare',
}

ER_THRESHOLDS = [9, 13, 20, 40]
RP_THRESHOLDS = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def normalize_address(addr):
    s = addr.upper().strip().rstrip('.')
    s = re.sub(r'\bSTREET\b', 'ST', s); s = re.sub(r'\bDRIVE\b', 'DR', s)
    s = re.sub(r'\bROAD\b', 'RD', s); s = re.sub(r'\bAVENUE\b', 'AVE', s)
    s = re.sub(r'\bBOULEVARD\b', 'BLVD', s); s = re.sub(r'\bLANE\b', 'LN', s)
    s = re.sub(r'\bCOURT\b', 'CT', s); s = re.sub(r'\bCIRCLE\b', 'CIR', s)
    s = re.sub(r'\bPLACE\b', 'PL', s); s = re.sub(r'\bTERRACE\b', 'TER', s)
    s = re.sub(r'\bPARKWAY\b', 'PKWY', s); s = re.sub(r'\bHIGHWAY\b', 'HWY', s)
    s = re.sub(r'\bNORTH\b', 'N', s); s = re.sub(r'\bSOUTH\b', 'S', s)
    s = re.sub(r'\bEAST\b', 'E', s); s = re.sub(r'\bWEST\b', 'W', s)
    s = re.sub(r'\.', '', s); s = re.sub(r'\s+', ' ', s).strip()
    return s

def extract_street_number(addr):
    m = re.match(r'^(\d+)', addr.strip())
    return m.group(1) if m else None

# Load
print("Loading data...")
wb_t = openpyxl.load_workbook(TIERING_PATH, data_only=True)
v20_scores = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws_t = wb_t[sn]
    for r in range(2, ws_t.max_row + 1):
        name = ws_t.cell(r, 1).value
        if name:
            v20_scores[name] = {'ai': ws_t.cell(r, 14).value or 0}

wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']
finance_names = {}
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in T5_EXCLUDE:
        finance_names[str(name).strip()] = ws_a.cell(r, 18).value

all_db_names = set()
db_to_finance = {}
for fname, db_names in FINANCE_TO_DB.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_finance[dn] = fname

entity_rows = defaultdict(list)
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue
    fname = db_to_finance[corp]
    if fname not in finance_names:
        continue

    beds = ws_db.cell(r, headers['Total_Beds']).value or 0
    beds = beds if isinstance(beds, (int, float)) else 0
    addr_raw = str(ws_db.cell(r, headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, headers['City']).value or '').strip().upper()
    tot_rev = ws_db.cell(r, headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0
    served = str(ws_db.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws_db.cell(r, headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws_db.cell(r, headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws_db.cell(r, headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    street_num = extract_street_number(addr_raw)

    entity_rows[fname].append({
        'raw_key': (addr_raw.upper(), city),
        'norm_key': (normalize_address(addr_raw), city),
        'num_key': (street_num, city) if street_num else (normalize_address(addr_raw), city),
        'beds': beds, 'tot_rev': tot_rev,
        'served': served, 'mh': mh, 'pcp': pcp, 'integ': integ,
    })

def compute_full(fname, rows, campus_key_field):
    kept = [r for r in rows if r['beds'] > BED_MIN]
    if not kept:
        return {'tier': 'T4', 'n_camp': 0, 'total_score': 0, 'er': 0, 'ir': 0, 'si': 0, 'rp': 0, 'rs': 0, 'ai': 0, 'total_rev': 0}

    campuses = set(r[campus_key_field] for r in kept)
    served_camps = set(r[campus_key_field] for r in kept if r['served'])
    mh_camps = set(r[campus_key_field] for r in kept if r['mh'])
    pcp_camps = set(r[campus_key_field] for r in kept if r['pcp'])
    integ_camps = set(r[campus_key_field] for r in kept if r['integ'])
    total_rev = sum(r['tot_rev'] for r in kept)
    n_camp = len(campuses)
    n_srv = len(served_camps)

    if n_camp < 7:
        return {'tier': 'T4', 'n_camp': n_camp, 'total_score': 0, 'er': 0, 'ir': 0, 'si': 0, 'rp': 0, 'rs': 0, 'ai': 0, 'total_rev': total_rev}

    er = score_dim(n_camp, ER_THRESHOLDS)
    rp = score_dim(total_rev, RP_THRESHOLDS)
    rs = TOM_RS_SCORES.get(fname, 1)
    si = DATA_SI_SCORES.get(fname, 2)
    v20_name = FINANCE_TO_V20.get(fname)
    v20 = v20_scores.get(v20_name) if v20_name else None
    ai = v20['ai'] if v20 else 0

    n_integ = len(integ_camps); n_mh = len(mh_camps); n_pcp = len(pcp_camps)
    dual = len(mh_camps & pcp_camps)
    int_pct = n_integ / n_camp if n_camp > 0 else 0

    if n_srv == 0: ir = 1
    elif int_pct >= 0.5: ir = 5
    elif int_pct >= 0.25: ir = 4
    elif dual > 0: ir = 4
    elif n_mh > 0 and n_pcp > 0: ir = 3
    elif n_mh > 0 or n_pcp > 0: ir = 2
    else: ir = 1

    total_score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = 'T1' if total_score >= 55 else 'T2' if total_score >= 35 else 'T3'

    return {'tier': tier, 'n_camp': n_camp, 'total_score': total_score,
            'er': er, 'ir': ir, 'si': si, 'rp': rp, 'rs': rs, 'ai': ai, 'total_rev': total_rev}

# Run both options
for label, key_field in [("OPTION A: Standard Normalization + 15-Bed Rule", "norm_key"),
                          ("OPTION B: Street-Number Matching + 15-Bed Rule", "num_key")]:
    print(f"\n{'='*140}")
    print(f"{label}")
    print(f"{'='*140}")
    print(f"{'Entity':<35} {'Camp':>5} {'ER':>3} {'IR':>3} {'SI':>3} {'RP':>3} {'RS':>3} {'AI':>3}  {'Score':>5}  {'Tier':>4}")
    print(f"{'-'*140}")

    tiers = Counter()
    results = []
    for fname in sorted(entity_rows.keys()):
        r = compute_full(fname, entity_rows[fname], key_field)
        tiers[r['tier']] += 1
        results.append((fname, r))

    # Print by tier
    for tier_label in ['T1', 'T2', 'T3', 'T4']:
        tier_ents = [(f, r) for f, r in results if r['tier'] == tier_label]
        if not tier_ents:
            continue
        for fname, r in sorted(tier_ents, key=lambda x: (-x[1]['total_score'], x[0])):
            if r['tier'] == 'T4':
                print(f"{fname:<35} {r['n_camp']:>5}   -   -   -   -   -   -      -    T4")
            else:
                print(f"{fname:<35} {r['n_camp']:>5} {r['er']:>3} {r['ir']:>3} {r['si']:>3} {r['rp']:>3} {r['rs']:>3} {r['ai']:>3}  {r['total_score']:>5}    {r['tier']}")
        print()

    print(f"  T1={tiers.get('T1',0)}  T2={tiers.get('T2',0)}  T3={tiers.get('T3',0)}  T4={tiers.get('T4',0)}  T5={len(T5_EXCLUDE)}  Total={sum(tiers.values())+len(T5_EXCLUDE)}")

# Diff
print(f"\n{'='*140}")
print(f"DIFFERENCES BETWEEN OPTIONS")
print(f"{'='*140}")
print(f"{'Entity':<35} {'OptA Tier':>8} {'OptA Score':>10} {'OptA Camp':>9}  {'OptB Tier':>8} {'OptB Score':>10} {'OptB Camp':>9}")
print(f"{'-'*140}")
for fname in sorted(entity_rows.keys()):
    a = compute_full(fname, entity_rows[fname], 'norm_key')
    b = compute_full(fname, entity_rows[fname], 'num_key')
    if a['tier'] != b['tier'] or a['total_score'] != b['total_score']:
        print(f"{fname:<35} {a['tier']:>8} {a['total_score']:>10} {a['n_camp']:>9}  {b['tier']:>8} {b['total_score']:>10} {b['n_camp']:>9}")
