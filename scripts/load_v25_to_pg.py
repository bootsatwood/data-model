"""
Load V25 facility database into PostgreSQL (truncate-and-reload).

Tables:
  bd.market_intel_corporate_entities  (parent — load first)
  bd.market_intel_facilities          (child — FK to corporate entities)
  bd.market_intel_metros              (untouched — already populated)

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25.xlsx

Usage:
  python scripts/load_v25_to_pg.py [--dry-run]
"""

import os
import sys
import psycopg2
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
VAULT = os.path.expanduser(
    "~/OneDrive - Eventus WholeHealth/Vault"
)
V25_PATH = os.path.join(
    VAULT, "02_Data_Model", "Current", "1_Combined_Database_FINAL_V25.xlsx"
)

# PostgreSQL connection
PG_HOST = "keystone-platform-postgres.postgres.database.azure.com"
PG_PORT = 5432
PG_DATABASE = "postgres"
PG_USER = "ratwood"
PG_PASSWORD = "Ch0EeU3yz7ep5bEWTKT1UoC^NMPCZXzN"


# ---------------------------------------------------------------------------
# Connection
# ---------------------------------------------------------------------------
def get_connection():
    return psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        database=PG_DATABASE,
        user=PG_USER,
        password=PG_PASSWORD,
        sslmode="require",
    )


# ---------------------------------------------------------------------------
# Step 1: Read V25 facilities from Excel
# ---------------------------------------------------------------------------
def read_v25_facilities():
    """Returns a list of dicts, one per facility row."""
    wb = openpyxl.load_workbook(V25_PATH, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    facilities = []
    for row in ws.iter_rows(min_row=2):
        vals = {headers[i]: cell.value for i, cell in enumerate(row)}
        facilities.append({
            "facility_name": clean_str(vals.get("Facility_Name")) or "",
            "corporate_name_raw": clean_str(vals.get("Corporate_Name")),
            "source_type": clean_str(vals.get("Source_Type")) or "",
            "address": clean_str(vals.get("Address")),
            "city": clean_str(vals.get("City")),
            "state": clean_str(vals.get("State")),
            "zip": clean_str(vals.get("ZIP")),
            "county": clean_str(vals.get("County")),
            "ownership_type": clean_str(vals.get("Ownership_Type")),
            "total_beds": safe_int(vals.get("Total_Beds")),
            "census": safe_decimal(vals.get("Census")),
            "do_we_serve": to_bool(vals.get("Do_We_Serve")),
            "integrated_flag": to_bool(vals.get("Integrated_Flag")),
            "pcp_flag": to_bool(vals.get("PCP_Flag")),
            "mh_flag": to_bool(vals.get("MH_Flag")),
            "barrier": clean_str(vals.get("Barrier")),
            "latitude": safe_decimal(vals.get("Latitude")),
            "longitude": safe_decimal(vals.get("Longitude")),
            "contract_status": clean_str(vals.get("Contract_Status")),
            "geographic_tier": clean_str(vals.get("Geographic_Tier")),
            "original_geographic_tier": clean_str(vals.get("Original_Geographic_Tier")),
            "metro_assignment": clean_str(vals.get("Metro_Assignment")),
            "distance_to_metro_center": safe_decimal(vals.get("Distance_to_Metro_Center")),
            "metro_center_used": clean_str(vals.get("Metro_Center_Used")),
            "corp_attribution_source": clean_str(vals.get("Corp_Attribution_Source")),
            "data_quality_flag": clean_str(vals.get("Data_Quality_Flag")),
        })

    wb.close()
    print(f"  Read {len(facilities)} facility rows from V25")
    return facilities


# ---------------------------------------------------------------------------
# Step 2: Build corporate entities from distinct names
# ---------------------------------------------------------------------------
def build_corporate_entities(facilities):
    """Build entity list from distinct Corporate_Name values in facilities."""
    entities = {}
    for f in facilities:
        name = f["corporate_name_raw"]
        if not name:
            continue
        key = name.upper()
        if key not in entities:
            entities[key] = {
                "corporate_name": name,
                "bd_tier": None,
                "total_score": None,
                "enterprise_reach": None,
                "integration_ready": None,
                "strategic_influence": None,
                "revenue_potential_score": None,
                "relationship_strength": None,
                "ai_tech_adoption": None,
                "barrier_type": None,
                "ownership_type": "Corporate",
            }

    print(f"  Built {len(entities)} corporate entities from distinct names")
    return list(entities.values())


# ---------------------------------------------------------------------------
# Step 3: Truncate and reload
# ---------------------------------------------------------------------------
def truncate_tables(cursor):
    """Truncate facilities (child) first, then corporate entities (parent)."""
    cursor.execute("TRUNCATE TABLE bd.market_intel_facilities CASCADE")
    cursor.execute("TRUNCATE TABLE bd.market_intel_corporate_entities CASCADE")
    print("  Truncated bd.market_intel_facilities and bd.market_intel_corporate_entities")


def insert_corporate_entities(cursor, entities):
    """Insert corporate entities and return a name->id mapping."""
    sql = """
        INSERT INTO bd.market_intel_corporate_entities
            (corporate_name, bd_tier, total_score, enterprise_reach,
             integration_ready, strategic_influence, revenue_potential_score,
             relationship_strength, ai_tech_adoption, barrier_type, ownership_type)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    count = 0
    for e in entities:
        cursor.execute(sql, (
            e["corporate_name"],
            e["bd_tier"],
            e["total_score"],
            e["enterprise_reach"],
            e["integration_ready"],
            e["strategic_influence"],
            e["revenue_potential_score"],
            e["relationship_strength"],
            e["ai_tech_adoption"],
            e["barrier_type"],
            e["ownership_type"],
        ))
        count += 1

    print(f"  Inserted {count} corporate entities")

    # Build name -> id mapping
    cursor.execute("SELECT id, corporate_name FROM bd.market_intel_corporate_entities")
    name_to_id = {}
    for row in cursor.fetchall():
        name_to_id[row[1].upper()] = row[0]
    print(f"  Built name->id mapping ({len(name_to_id)} entries)")
    return name_to_id


def insert_facilities(cursor, facilities, corp_name_to_id):
    """Insert facilities with FK references to corporate entities."""
    sql = """
        INSERT INTO bd.market_intel_facilities
            (facility_name, corporate_entity_id, corporate_name_raw, source_type,
             address, city, state, zip, county, ownership_type,
             total_beds, census, do_we_serve, integrated_flag, pcp_flag, mh_flag,
             barrier, latitude, longitude, contract_status,
             geographic_tier, original_geographic_tier, metro_assignment,
             distance_to_metro_center, metro_center_used, corp_attribution_source,
             data_quality_flag)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    count = 0
    fk_matched = 0
    fk_null = 0

    for f in facilities:
        corp_id = None
        raw_name = f["corporate_name_raw"]
        if raw_name:
            corp_id = corp_name_to_id.get(raw_name.upper())
            if corp_id:
                fk_matched += 1
            else:
                fk_null += 1
        else:
            fk_null += 1

        cursor.execute(sql, (
            f["facility_name"],
            corp_id,
            f["corporate_name_raw"],
            f["source_type"],
            f["address"],
            f["city"],
            f["state"],
            f["zip"],
            f["county"],
            f["ownership_type"],
            f["total_beds"],
            f["census"],
            f["do_we_serve"],
            f["integrated_flag"],
            f["pcp_flag"],
            f["mh_flag"],
            f["barrier"],
            f["latitude"],
            f["longitude"],
            f["contract_status"],
            f["geographic_tier"],
            f["original_geographic_tier"],
            f["metro_assignment"],
            f["distance_to_metro_center"],
            f["metro_center_used"],
            f["corp_attribution_source"],
            f["data_quality_flag"],
        ))
        count += 1
        if count % 5000 == 0:
            print(f"    ... {count} facilities inserted")

    print(f"  Inserted {count} facilities ({fk_matched} with corporate FK, {fk_null} with NULL FK)")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_decimal(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_bool(val):
    """Convert Yes/No to Python bool for PostgreSQL BOOLEAN columns."""
    if val is None:
        return False
    return str(val).strip().lower() in ("yes", "true", "1")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print(f"V25 -> PostgreSQL Loader {'(DRY RUN)' if dry_run else ''}")
    print("=" * 60)

    if not os.path.exists(V25_PATH):
        print(f"ERROR: V25 file not found: {V25_PATH}")
        sys.exit(1)
    print(f"  Source: {V25_PATH}")

    print()
    print("Step 1: Reading V25 facilities...")
    facilities = read_v25_facilities()

    print()
    print("Step 2: Building corporate entities from distinct names...")
    entities = build_corporate_entities(facilities)

    print()
    print("Step 3: Connecting to PostgreSQL...")
    conn = get_connection()
    cursor = conn.cursor()
    print("  Connected.")

    if dry_run:
        print()
        print("DRY RUN — no changes will be made.")
        print(f"  Would truncate and reload:")
        print(f"    {len(entities)} corporate entities")
        print(f"    {len(facilities)} facilities")
        cursor.close()
        conn.close()
        return

    print()
    print("Step 4: Truncating existing data...")
    truncate_tables(cursor)

    print()
    print("Step 5: Inserting corporate entities...")
    corp_name_to_id = insert_corporate_entities(cursor, entities)

    print()
    print("Step 6: Inserting facilities...")
    insert_facilities(cursor, facilities, corp_name_to_id)

    print()
    print("Step 7: Committing transaction...")
    conn.commit()
    print("  Committed.")

    # Verification
    print()
    print("Step 8: Verification...")
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_corporate_entities")
    ce_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_facilities")
    f_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_facilities WHERE do_we_serve = true")
    served_count = cursor.fetchone()[0]
    cursor.execute("""
        SELECT ce.corporate_name, COUNT(f.id) AS facilities
        FROM bd.market_intel_corporate_entities ce
        JOIN bd.market_intel_facilities f ON f.corporate_entity_id = ce.id
        GROUP BY ce.corporate_name
        ORDER BY COUNT(f.id) DESC
        LIMIT 10
    """)
    top10 = cursor.fetchall()

    print(f"  Corporate entities in DB: {ce_count}")
    print(f"  Facilities in DB: {f_count}")
    print(f"  Served facilities: {served_count}")
    print(f"  Top 10 corporate entities by facility count:")
    for row in top10:
        print(f"    {row[0]}: {row[1]}")

    cursor.close()
    conn.close()
    print()
    print("Done. V25 loaded successfully.")


if __name__ == "__main__":
    main()
