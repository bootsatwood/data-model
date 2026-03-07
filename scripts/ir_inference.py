"""Infer IR scores for the 16 entities missing Brooke's V20 qualitative scores.
Shows service flag proportions so we can make informed judgments.

IR rules (from Brooke's clarification):
  - Not served → IR=1
  - Served, single service line (MH or PCP only) → IR=2 (conservative; 3 requires behavioral judgment)
  - Served, dual service (MH+PCP) → IR=4
  - Fully integrated → IR=5
  - Service config sets the FLOOR; behavioral engagement can raise within band
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}

MISSING_16 = {
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'AMERICAN HEALTHCARE LLC': {'AMERICAN HEALTHCARE, LLC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'STORYPOINT': {'STORYPOINT'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'CARDON & ASSOCIATES': {'CARDON & ASSOCIATES'},
}

wb = openpyxl.load_workbook(DB_PATH, data_only=True)
ws = wb.active
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

all_db_names = set()
name_to_finance = {}
for fname, db_names in MISSING_16.items():
    for dn in db_names:
        all_db_names.add(dn)
        name_to_finance[dn] = fname

entity_data = {}
for fname in MISSING_16:
    entity_data[fname] = {
        'campuses': set(),
        'served_camps': set(),
        'mh_camps': set(),
        'pcp_camps': set(),
        'integ_camps': set(),
        'rows': 0,
    }

for r in range(2, ws.max_row + 1):
    corp = str(ws.cell(r, headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws.cell(r, headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue

    fname = name_to_finance[corp]
    d = entity_data[fname]
    d['rows'] += 1

    addr = str(ws.cell(r, headers['Address']).value or '').strip().upper()
    city = str(ws.cell(r, headers['City']).value or '').strip().upper()
    campus = (addr, city)
    d['campuses'].add(campus)

    served = str(ws.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws.cell(r, headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws.cell(r, headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws.cell(r, headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')

    if served:
        d['served_camps'].add(campus)
    if mh:
        d['mh_camps'].add(campus)
    if pcp:
        d['pcp_camps'].add(campus)
    if integ:
        d['integ_camps'].add(campus)

# Output
print(f"{'Entity':<30} {'Camp':>4} {'Srv':>4} {'MH':>4} {'PCP':>4} {'Int':>4}  {'Srv%':>5} {'MH%':>5} {'PCP%':>5} {'Int%':>5}  {'Inferred':>15} {'IR':>3}")
print("-" * 115)

for fname in sorted(MISSING_16.keys()):
    d = entity_data[fname]
    camps = len(d['campuses'])
    srv = len(d['served_camps'])
    mh = len(d['mh_camps'])
    pcp = len(d['pcp_camps'])
    integ = len(d['integ_camps'])

    srv_pct = (srv / camps * 100) if camps > 0 else 0
    mh_pct = (mh / camps * 100) if camps > 0 else 0
    pcp_pct = (pcp / camps * 100) if camps > 0 else 0
    int_pct = (integ / camps * 100) if camps > 0 else 0

    # Dual = campuses that have BOTH MH and PCP
    dual_camps = d['mh_camps'] & d['pcp_camps']
    dual = len(dual_camps)

    # IR inference based on PREDOMINANT service config across served campuses
    if srv == 0:
        ir = 1
        config = "Not served"
    elif integ > 0 and int_pct >= 50:
        ir = 5
        config = "Integrated (maj)"
    elif integ > 0 and int_pct >= 25:
        ir = 4
        config = f"Integrated ({integ}/{camps})"
    elif dual > 0:
        ir = 4
        config = f"Dual MH+PCP ({dual})"
    elif mh > 0 and pcp > 0:
        # Have both services but at different campuses
        ir = 3
        config = f"MH({mh})+PCP({pcp}) sep"
    elif mh > 0:
        ir = 2
        config = f"MH only ({mh}/{camps})"
    elif pcp > 0:
        ir = 2
        config = f"PCP only ({pcp}/{camps})"
    else:
        # Served but no MH/PCP/Integrated flags
        ir = 1
        config = f"Served, no svc flags"

    gate = " *GATE" if camps < 7 else ""
    print(f"{fname:<30} {camps:>4} {srv:>4} {mh:>4} {pcp:>4} {integ:>4}  {srv_pct:>4.0f}% {mh_pct:>4.0f}% {pcp_pct:>4.0f}% {int_pct:>4.0f}%  {config:>15} {ir:>3}{gate}")

print()
print("Key: Camp=total campuses, Srv=served campuses, MH/PCP/Int=campuses with that flag")
print("     *GATE = fails MUO gate (<7 campuses in footprint)")
print()
print("IR Logic:")
print("  1 = Not served")
print("  2 = Single service only (MH or PCP, not both)")
print("  3 = Both MH and PCP present but at separate campuses")
print("  4 = Dual (MH+PCP at same campus) or partial Integrated (25-49%)")
print("  5 = Majority of campuses have Integrated flag (>=50%)")
