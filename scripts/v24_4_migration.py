"""
V24.4 Migration — Punchlist #4 Corporate Name Consolidation

Applies 68 SAME pair decisions from Corporate_Name_Dedup_Review_ra_v2.xlsx.
Renames non-canonical Corporate_Name values to their canonical form.

Input:  Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_3.xlsx
Output: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V24_4.xlsx

No deletes. No structural changes. Corporate_Name renames only.
"""

import openpyxl
import os
from datetime import datetime

# Paths
VAULT = os.path.expanduser("~/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current")
INPUT_FILE = os.path.join(VAULT, "1_Combined_Database_FINAL_V24_3.xlsx")
OUTPUT_FILE = os.path.join(VAULT, "1_Combined_Database_FINAL_V24_4.xlsx")

# Corporate_Name rename map — from Punchlist #4 fuzzy match review
# Only includes renames for pairs resolved in the 2026-03-21 session
# (V22.9 renames were already applied and should not be re-run)
RENAME_MAP = {
    # ProMedica cluster (rows 39, 41, 44)
    "HCR ManorCare": "PROMEDICA",
    "HCR Manor Care/ Promedica": "PROMEDICA",
    "Promedica Health Systems": "PROMEDICA",
    "PROMEDICA SENIOR CARE": "PROMEDICA",

    # Dominion cluster (rows 78, 79)
    "Dominion Group": "DOMINION SENIOR LIVING",
    "DOMINION": "DOMINION SENIOR LIVING",

    # Traditions (row 57)
    "Traditions Management": "TRADITIONS",

    # Silver Birch (row 58)
    "SILVER BIRCH": "SILVER BIRCH LIVING",

    # English Meadows (row 71)
    "English Meadows Senior Living Communities": "ENGLISH MEADOWS",

    # Century Care (row 92)
    "Century Care": "CENTURY CARE MANAGEMENT",

    # Providence (row 101)
    "PROVIDENCE, LLC": "PROVIDENCE HEALTH GROUP",

    # Five Star (row 107)
    "FIVE STAR SENIOR LIVING": "FIVE STAR RESIDENCES",

    # Lutheran National (row 134)
    "LUTHERAN NATIONAL COMMUNITIES AND SERVICES": "NATIONAL LUTHERAN COMMUNITIES & SERVICES",

    # Kaplan (row 138)
    "Senior Management LLC (Part of Kaplan Development Group)": "KAPLAN DEVELOPMENT",

    # Davco (row 141)
    "DAVCO HOMES, INC": "DAVCO",

    # Trinity Health (row 45)
    "TRINITY HEALTH SENIOR COMMUNITIES": "TRINITY HEALTH",

    # Continuing Healthcare (row 54)
    "CONTINUING HEALTHCARE": "CONTINUING HEALTHCARE SOLUTIONS",

    # Southern Healthcare (row 146)
    "SOUTHERN HEALTHCARE": "SOUTHERN HEALTHCARE MANAGEMENT, LLC",

    # Blanchard Valley (row 133)
    "BLANCHARD VALLEY CONTINUING CARE SERVICES": "BLANCHARD VALLEY HEALTH SYSTEM",

    # Pinnacle (rows 66, 115) — already partially handled in V24.2
    "PINNACLE": "Pinnacle Living",
    "Pinnacle Senior Living": "Pinnacle Living",

    # Bickford (row 34) — may already be done from V22.9
    "BICKFORD SENIOR LIVING": "BICKFORD",

    # Good Samaritan (rows 6, 8) — may already be done from V22.9
    "Evangelical Lutheran Good Samaritan Society/ Sandford": "GOOD SAMARITAN SOCIETY",
    "THE EVANGELICAL LUTHERAN GOOD SAMARITAN SOCIETY": "GOOD SAMARITAN SOCIETY",
}


def main():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Find Corporate_Name column
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    corp_col = headers.index("Corporate_Name") + 1
    print(f"Corporate_Name column: {corp_col}")

    total_rows = ws.max_row - 1
    print(f"Total data rows: {total_rows}")

    # Apply renames
    renames_applied = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, corp_col).value
        if val in RENAME_MAP:
            new_val = RENAME_MAP[val]
            ws.cell(r, corp_col).value = new_val
            key = f"{val} -> {new_val}"
            renames_applied[key] = renames_applied.get(key, 0) + 1

    # Report
    total_renamed = sum(renames_applied.values())
    print(f"\n{'='*60}")
    print(f"V24.4 Migration Summary")
    print(f"{'='*60}")
    print(f"Total renames applied: {total_renamed}")
    print(f"Unique rename rules matched: {len(renames_applied)}")
    print()

    if renames_applied:
        print(f"{'Old Name':<55} -> {'New Name':<40} | Count")
        print("-" * 110)
        for key, count in sorted(renames_applied.items()):
            old, new = key.split(" -> ")
            print(f"{old:<55} -> {new:<40} | {count}")
    else:
        print("No renames found — all names may already be canonical.")

    # Check for rename rules that didn't match anything
    unmatched = set(RENAME_MAP.keys()) - {k.split(" -> ")[0] for k in renames_applied}
    if unmatched:
        print(f"\nRename rules with 0 matches (already applied or name not in DB):")
        for name in sorted(unmatched):
            print(f"  {name} -> {RENAME_MAP[name]}")

    # Save
    print(f"\nSaving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print(f"Done. V24.4 saved ({total_rows} rows, {total_renamed} renames).")


if __name__ == "__main__":
    main()
