#!/usr/bin/env python3
"""Build Liberty QC Review workbook — all proposed changes before apply."""

import sys
sys.path.insert(0, ".")
from utils import load_db, norm_addr, norm, safe, ensure_report_dir, VAULT, FOOTPRINT
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB = VAULT / "Current" / "1_Combined_Database_FINAL_V22_1.xlsx"
headers, rows = load_db(DB)
print(f"V22.1 loaded: {len(rows):,} rows")

# ── Styles ──
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

cols = [
    "Action", "Change Type", "Excel Row",
    "Facility_Name", "Source_Type", "New Source_Type",
    "Corporate_Name (Current)", "Corporate_Name (New)",
    "Address", "City", "State", "ZIP",
    "Total_Beds", "Census", "Do_We_Serve",
    "Website Match", "Notes",
]


def write_header(ws):
    for c, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"


def write_row(ws, row_num, vals, fill=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def autofit(ws, max_row):
    for c in range(1, len(cols) + 1):
        max_len = len(str(ws.cell(row=1, column=c).value or ""))
        for r in range(2, min(max_row + 1, 200)):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 45))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max_len + 3


# ── Build DB address index ──
db_by_addr = {}
for r in rows:
    akey = norm_addr(safe(r.get("Address", ""))) + "|" + norm(safe(r.get("City", ""))) + "|" + norm(safe(r.get("State", "")))
    db_by_addr.setdefault(akey, []).append(r)

rows_by_excel = {r["_excel_row"]: r for r in rows}


def addr_lookup(address, city, state):
    akey = norm_addr(address) + "|" + norm(city) + "|" + norm(state)
    return db_by_addr.get(akey, [])


# ── Confirmed LHR addresses from PDF directory ──
lhr_addresses = [
    ("Liberty Commons N&R of Alamance County", "791 Boone Station Drive", "Burlington", "NC"),
    ("Three Rivers Health & Rehab Center", "1403 Connor Ave.", "Windsor", "NC"),
    ("Elizabethtown Healthcare & Rehab Center", "208 Mercer Road", "Elizabethtown", "NC"),
    ("Southport Health & Rehab Center", "630 N Fodale Ave", "Southport", "NC"),
    ("Pisgah Manor", "104 Holcombe Cove Road", "Candler", "NC"),
    ("Liberty Commons N&R of Columbus County", "1402 Pinckney Street", "Whiteville", "NC"),
    ("Shoreland Health Care & Retirement Center", "200 Flowers-Pridgen Drive", "Whiteville", "NC"),
    ("Golden Years Nursing Home", "7348 North West Street", "Falcon", "NC"),
    ("Woodlands Nursing & Rehab Center", "400 Pelt Drive", "Fayetteville", "NC"),
    ("Highland House Rehab & Healthcare", "1700 Pamalee Dr.", "Fayetteville", "NC"),
    ("Bermuda Commons N&R Center", "316 NC Highway 801 South", "Advance", "NC"),
    ("Briar Creek at The Barclay", "6041 Piedmont Row Drive", "Charlotte", "NC"),
    ("Oak Forest Health & Rehab Center", "5680 Windy Hill Drive", "Winston-Salem", "NC"),
    ("The Oaks", "901 Bethesda Road", "Winston-Salem", "NC"),
    ("Summerstone Health & Rehab Center", "485 Veteran's Way", "Kernersville", "NC"),
    ("Louisburg Healthcare & Rehab Center", "202 Smoketree Way", "Louisburg", "NC"),
    ("Liberty Commons N&R of Halifax County", "101 Caroline Avenue", "Weldon", "NC"),
    ("Silver Bluff Village", "100 Silver Bluff Dr.", "Canton", "NC"),
    ("Liberty Commons N&R of Johnston County", "2315 Highway 242 North", "Benson", "NC"),
    ("Liberty Commons N&R of Lee County", "310 Commerce Drive", "Sanford", "NC"),
    ("Westfield Rehab and Health Center", "3100 Tramway Road", "Sanford", "NC"),
    ("Royal Park Rehab and Health Center", "2700 Royal Commons Lane", "Matthews", "NC"),
    ("The Pavilion Health Center at Brightmore", "10011 Providence Road West", "Charlotte", "NC"),
    ("Pinehurst Healthcare & Rehab Center", "300 Blake Road", "Pinehurst", "NC"),
    ("The Inn at Quail Haven Village", "155 Blake Road", "Pinehurst", "NC"),
    ("Bradley Creek at Carolina Bay", "740 Diamond Shoals Road", "Wilmington", "NC"),
    ("Liberty Commons Rehab Center", "121 Racine Drive", "Wilmington", "NC"),
    ("Parkview Health & Rehab Center", "1718 Legion Road", "Chapel Hill", "NC"),
    ("Roxboro Healthcare & Rehab Center", "901 Ridge Road", "Roxboro", "NC"),
    ("WoodHaven Nursing Center", "1150 Pine Run Dr.", "Lumberton", "NC"),
    ("Liberty Commons N&R of Rowan County", "4412 South Main Street", "Salisbury", "NC"),
    ("Mary Gran Nursing Center", "120 Southwood Drive", "Clinton", "NC"),
    ("Southwood Nursing & Rehab Center", "180 Southwood Drive", "Clinton", "NC"),
    ("Capital Nursing & Rehab Center", "3000 Holston Lane", "Raleigh", "NC"),
    ("Bloomsbury at Hayes Barton Place", "2750 Oberlin Road", "Raleigh", "NC"),
    ("Swift Creek at The Templeton", "221 Brightmore Drive", "Cary", "NC"),
    ("Warren Hills Rehab & Nursing Center", "864 US Hwy. 158 Business West", "Warrenton", "NC"),
    ("The Foley Center at Chestnut Ridge", "621 Chestnut Ridge Parkway", "Blowing Rock", "NC"),
    ("Yadkin Nursing Care Center", "903 W. Main Street", "Yadkinville", "NC"),
    ("Kempton of Charleston (SNF)", "194 Spring Street", "Charleston", "SC"),
    ("Preserve at Fairfield Glade", "100 Samaritan Way", "Crossville", "TN"),
]

# ── Confirmed LSL addresses from screenshots ──
lsl_addresses = [
    ("Brightmore of South Charlotte", "10225 Old Ardrey Kell Road", "Charlotte", "NC"),
    ("Brightmore of Wilmington", "2124 41st Street", "Wilmington", "NC"),
    ("Carlisle Palm Beach", "450 East Ocean Avenue", "Lantana", "FL"),
    ("Carolina Bay at Autumn Hall", "620 Carolina Bay Drive", "Wilmington", "NC"),
    ("Hayes Barton Place", "2600 Yettington Drive", "Raleigh", "NC"),
    ("Inspire Briar Chapel", "152 Market Chapel Road", "Pittsboro", "NC"),
    ("Inspire Brunswick Forest", "6146 Liberty Hall Drive", "Leland", "NC"),
    ("Inspire Royal Park", "4101 Glenloch Circle", "Matthews", "NC"),
    ("Inspire Sandhill", "440 Town Center Place", "Columbia", "SC"),
    ("Kempton of Charleston (AL)", "194 Spring Street", "Charleston", "SC"),
    ("Kempton of Jacksonville", "3045 Henderson Drive Extension", "Jacksonville", "NC"),
    ("Kempton of Rockhill", "1611 Constitution Boulevard", "Rock Hill", "SC"),
    ("Oakleaf Village of Lexington", "800 N Lake Drive", "Lexington", "SC"),
    ("Pisgah Valley", "95 Holcombe Cove Road", "Candler", "NC"),
    ("Quail Haven Village", "155 Blake Boulevard", "Pinehurst", "NC"),
    ("South Bay at Mount Pleasant", "1400 Liberty Midtown Drive", "Mount Pleasant", "SC"),
    ("The Barclay at SouthPark", "4801 Barclay Downs Drive", "Charlotte", "NC"),
    ("The Peninsula of Charleston", "625 King Street", "Charleston", "SC"),
    ("The Templeton of Cary", "125 Brightmore Drive", "Cary", "NC"),
    ("Wellington Bay", "10430 Stable Lane", "Wellington", "FL"),
]

# ── Source_Type corrections from Facility DB review ──
SOURCE_TYPE_FIXES = {
    8774: "SNF",   # Briar Creek Health Center SNF — currently ALF, should be SNF
    9983: "SNF",   # Warren Hills Nursing Center SNF — currently ALF, should be SNF
}

# ── Bare "LIBERTY" rows confirmed as Liberty Health ──
BARE_LIBERTY_CONFIRMED = {8773, 8774, 9291, 9487, 9641, 9982, 9983}

# ── Liberty Village — separate company ──
LIBERTY_VILLAGE_CORPS = {
    "LIBERTY VILLAGE",
    "Liberty Village of Freeport", "Liberty Village of Geneseo",
    "Liberty Village of Jerseyville", "Liberty Village of Peoria",
    "Liberty Village of Peru", "Liberty Village of Rochelle",
    "Liberty Village of Streator",
}

# ── Not Liberty Health at all ──
NOT_LIBERTY_HEALTH = {
    "SEACOAST AT LIBERTY RIDGE",
    "LIBERTY RIDGE SENIOR LIVING INC",
    "LIBERTY RETIREMENT PROPERTIES OF LIMA LTD",
    "WEST LIBERTY GARDEN APARTMENTS",
}

# ════════════════════════════════════════════════════════
# BUILD CHANGES
# ════════════════════════════════════════════════════════

wb = Workbook()
all_changes = []

# -- Tab 1: Corporate_Name -> LIBERTY --
# Collect all DB rows at confirmed Liberty addresses that aren't already "LIBERTY"
confirmed_liberty_rows = set()

# LHR address matches
for (web_name, addr, city, state) in lhr_addresses:
    for m in addr_lookup(addr, city, state):
        confirmed_liberty_rows.add(m["_excel_row"])

# LSL address matches
for (web_name, addr, city, state) in lsl_addresses:
    for m in addr_lookup(addr, city, state):
        confirmed_liberty_rows.add(m["_excel_row"])

# Bare LIBERTY confirmed
for row_num in BARE_LIBERTY_CONFIRMED:
    confirmed_liberty_rows.add(row_num)

# Also include current "LIBERTY SENIOR LIVING" rows — they're all Liberty Health
for r in rows:
    if safe(r.get("Corporate_Name", "")).upper() == "LIBERTY SENIOR LIVING":
        confirmed_liberty_rows.add(r["_excel_row"])

# Also include LIBERTY HEALTHCARE PROPERTIES rows that matched LHR addresses
for r in rows:
    corp = safe(r.get("Corporate_Name", "")).upper()
    if "LIBERTY HEALTHCARE PROPERTIES" in corp:
        # Check if at an LHR address
        akey = norm_addr(safe(r.get("Address", ""))) + "|" + norm(safe(r.get("City", ""))) + "|" + norm(safe(r.get("State", "")))
        for (web_name, addr, city, state) in lhr_addresses:
            lhr_key = norm_addr(addr) + "|" + norm(city) + "|" + norm(state)
            if akey == lhr_key:
                confirmed_liberty_rows.add(r["_excel_row"])
                break

ws1 = wb.active
ws1.title = "Rename to LIBERTY"
write_header(ws1)

rn = 2
rename_count = 0
for row_num in sorted(confirmed_liberty_rows):
    r = rows_by_excel.get(row_num)
    if not r:
        continue
    corp = safe(r.get("Corporate_Name", ""))
    corp_upper = corp.upper()

    # Skip if already "LIBERTY"
    if corp == "LIBERTY":
        # But check if it needs a Source_Type fix
        if row_num in SOURCE_TYPE_FIXES:
            new_stype = SOURCE_TYPE_FIXES[row_num]
            vals = [
                "FIX SOURCE_TYPE", "Source_Type correction", row_num,
                safe(r.get("Facility_Name", "")),
                safe(r.get("Source_Type", "")), new_stype,
                corp, corp,
                safe(r.get("Address", "")), safe(r.get("City", "")),
                safe(r.get("State", "")), safe(r.get("ZIP", "")),
                r.get("Total_Beds", ""), r.get("Census", ""),
                safe(r.get("Do_We_Serve", "")),
                "", "Facility DB confirms this is SNF, not ALF",
            ]
            write_row(ws1, rn, vals, yellow_fill)
            rn += 1
            rename_count += 1
        continue

    # Skip if it's a not-Liberty-Health company
    if corp in NOT_LIBERTY_HEALTH:
        continue
    # Skip Liberty Village
    if corp in LIBERTY_VILLAGE_CORPS:
        continue

    # Determine website match name
    web_match = ""
    r_addr = safe(r.get("Address", ""))
    r_city = safe(r.get("City", ""))
    r_state = safe(r.get("State", ""))
    for (wn, wa, wc, ws_) in lhr_addresses:
        if (norm_addr(wa) + "|" + norm(wc) + "|" + norm(ws_)
                == norm_addr(r_addr) + "|" + norm(r_city) + "|" + norm(r_state)):
            web_match = f"LHR: {wn}"
            break
    if not web_match:
        for (wn, wa, wc, ws_) in lsl_addresses:
            if (norm_addr(wa) + "|" + norm(wc) + "|" + norm(ws_)
                    == norm_addr(r_addr) + "|" + norm(r_city) + "|" + norm(r_state)):
                web_match = f"LSL: {wn}"
                break
    if not web_match and row_num in BARE_LIBERTY_CONFIRMED:
        web_match = "Facility DB confirmed"

    new_stype = SOURCE_TYPE_FIXES.get(row_num, "")
    change_type = "Corporate_Name rename"
    if new_stype:
        change_type = "Corporate_Name + Source_Type"

    notes = ""
    if corp == "":
        notes = "Blank corporate name"
    elif corp_upper == "LIBERTY SENIOR LIVING":
        notes = "Already Liberty but wrong canonical name"
    elif "PROPCO" in corp_upper or "LLC" in corp_upper:
        notes = "PROPCO/holding company name"
    elif corp_upper == "INDEPENDENT":
        notes = "Was classified independent"
    elif corp_upper == "NOT AVAIL FROM COUNTY":
        notes = "Missing data"

    vals = [
        "RENAME", change_type, row_num,
        safe(r.get("Facility_Name", "")),
        safe(r.get("Source_Type", "")), new_stype if new_stype else "",
        corp, "LIBERTY",
        safe(r.get("Address", "")), safe(r.get("City", "")),
        safe(r.get("State", "")), safe(r.get("ZIP", "")),
        r.get("Total_Beds", ""), r.get("Census", ""),
        safe(r.get("Do_We_Serve", "")),
        web_match, notes,
    ]

    served = safe(r.get("Do_We_Serve", "")).upper() in ("YES", "Y", "TRUE", "1")
    fill = green_fill if served else blue_fill
    write_row(ws1, rn, vals, fill)
    rn += 1
    rename_count += 1

autofit(ws1, ws1.max_row)
print(f"\nTab 1 - Rename to LIBERTY: {rename_count} rows")

# -- Tab 2: Liberty Village — separate company rename --
ws2 = wb.create_sheet("Liberty Village (1c)")
write_header(ws2)

lib_village = [r for r in rows if safe(r.get("Corporate_Name", "")) in LIBERTY_VILLAGE_CORPS]
rn = 2
for r in sorted(lib_village, key=lambda x: safe(x.get("Corporate_Name", ""))):
    corp = safe(r.get("Corporate_Name", ""))
    vals = [
        "RENAME", "Corporate_Name consolidation", r["_excel_row"],
        safe(r.get("Facility_Name", "")),
        safe(r.get("Source_Type", "")), "",
        corp, "LIBERTY VILLAGE",
        safe(r.get("Address", "")), safe(r.get("City", "")),
        safe(r.get("State", "")), safe(r.get("ZIP", "")),
        r.get("Total_Beds", ""), r.get("Census", ""),
        safe(r.get("Do_We_Serve", "")),
        "", "Separate IL-based company, not Liberty Health Management",
    ]
    write_row(ws2, rn, vals, blue_fill)
    rn += 1

autofit(ws2, ws2.max_row)
print(f"Tab 2 - Liberty Village: {len(lib_village)} rows")

# -- Tab 3: Leave Alone --
ws3 = wb.create_sheet("Leave Alone")
write_header(ws3)

not_lhm = [r for r in rows if safe(r.get("Corporate_Name", "")) in NOT_LIBERTY_HEALTH]
rn = 2
for r in not_lhm:
    vals = [
        "NO ACTION", "Not Liberty Health", r["_excel_row"],
        safe(r.get("Facility_Name", "")),
        safe(r.get("Source_Type", "")), "",
        safe(r.get("Corporate_Name", "")), safe(r.get("Corporate_Name", "")),
        safe(r.get("Address", "")), safe(r.get("City", "")),
        safe(r.get("State", "")), safe(r.get("ZIP", "")),
        r.get("Total_Beds", ""), r.get("Census", ""),
        safe(r.get("Do_We_Serve", "")),
        "", "Different company — not part of Liberty Health Management",
    ]
    write_row(ws3, rn, vals, gray_fill)
    rn += 1

autofit(ws3, ws3.max_row)
print(f"Tab 3 - Leave Alone: {len(not_lhm)} rows")

outpath = ensure_report_dir() / "Liberty_QC_Review.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
