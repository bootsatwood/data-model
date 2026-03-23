"""
Load V25.2 facility database into PostgreSQL (facilities-only reload).

PRESERVES bd.market_intel_corporate_entities (key_insight, scores, tiers, etc.)
Only truncates and reloads bd.market_intel_facilities.
Maps corporate_name_raw -> existing entity IDs.
Creates new entity entries only for names that don't already exist.
Adds Campus_ID column if missing from PG schema.

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_2.xlsx

Usage:
  python scripts/load_v25_2_to_pg.py [--dry-run]
"""
import os
import sys
import psycopg2
import openpyxl

VAULT = os.path.expanduser("~/OneDrive - Eventus WholeHealth/Vault")
V25_2_PATH = os.path.join(
    VAULT, "02_Data_Model", "Current", "1_Combined_Database_FINAL_V25_2.xlsx"
)

PG_HOST = "keystone-platform-postgres.postgres.database.azure.com"
PG_PORT = 5432
PG_DATABASE = "postgres"
PG_USER = "ratwood"
PG_PASSWORD = "Ch0EeU3yz7ep5bEWTKT1UoC^NMPCZXzN"


def get_connection():
    return psycopg2.connect(
        host=PG_HOST, port=PG_PORT, database=PG_DATABASE,
        user=PG_USER, password=PG_PASSWORD, sslmode="require",
    )


def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_decimal(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_bool(val):
    if val is None:
        return False
    return str(val).strip().lower() in ("yes", "true", "1")


def read_facilities():
    """Read V25.2 facilities from Excel."""
    wb = openpyxl.load_workbook(V25_2_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    facilities = []
    for row in ws.iter_rows(min_row=2):
        vals = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        facilities.append({
            "facility_name": clean_str(vals.get("Facility_Name")) or "",
            "corporate_name_raw": clean_str(vals.get("Corporate_Name")),
            "source_type": clean_str(vals.get("Source_Type")) or "",
            "address": clean_str(vals.get("Address")),
            "city": clean_str(vals.get("City")),
            "state": clean_str(vals.get("State")),
            "zip": clean_str(vals.get("ZIP")),
            "county": clean_str(vals.get("County")),
            "ownership_type": clean_str(vals.get("Ownership_Type")),
            "total_beds": safe_int(vals.get("Total_Beds")),
            "census": safe_decimal(vals.get("Census")),
            "do_we_serve": to_bool(vals.get("Do_We_Serve")),
            "integrated_flag": to_bool(vals.get("Integrated_Flag")),
            "pcp_flag": to_bool(vals.get("PCP_Flag")),
            "mh_flag": to_bool(vals.get("MH_Flag")),
            "barrier": clean_str(vals.get("Barrier")),
            "latitude": safe_decimal(vals.get("Latitude")),
            "longitude": safe_decimal(vals.get("Longitude")),
            "contract_status": clean_str(vals.get("Contract_Status")),
            "geographic_tier": clean_str(vals.get("Geographic_Tier")),
            "original_geographic_tier": clean_str(vals.get("Original_Geographic_Tier")),
            "metro_assignment": clean_str(vals.get("Metro_Assignment")),
            "distance_to_metro_center": safe_decimal(vals.get("Distance_to_Metro_Center")),
            "metro_center_used": clean_str(vals.get("Metro_Center_Used")),
            "corp_attribution_source": clean_str(vals.get("Corp_Attribution_Source")),
            "data_quality_flag": clean_str(vals.get("Data_Quality_Flag")),
            "campus_id": safe_int(vals.get("Campus_ID")),
        })

    wb.close()
    print(f"  Read {len(facilities)} facility rows from V25.2")
    return facilities


def ensure_campus_id_column(cursor):
    """Add campus_id column to facilities table if it doesn't exist."""
    cursor.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'bd' AND table_name = 'market_intel_facilities'
          AND column_name = 'campus_id'
    """)
    if not cursor.fetchone():
        cursor.execute("""
            ALTER TABLE bd.market_intel_facilities
            ADD COLUMN campus_id INTEGER
        """)
        print("  Added campus_id column to bd.market_intel_facilities")
    else:
        print("  campus_id column already exists")


def get_existing_entity_map(cursor):
    """Get name -> id mapping from existing entity table."""
    cursor.execute("SELECT id, corporate_name FROM bd.market_intel_corporate_entities")
    name_to_id = {}
    for row in cursor.fetchall():
        name_to_id[row[1].upper()] = row[0]
    print(f"  Existing entity map: {len(name_to_id)} entries")
    return name_to_id


def create_missing_entities(cursor, facilities, name_to_id):
    """Create entity entries for corporate names not already in the table."""
    needed = set()
    for f in facilities:
        name = f["corporate_name_raw"]
        if name and name.upper() not in name_to_id:
            needed.add(name)

    if not needed:
        print("  No new entities needed — all corporate names have existing entries")
        return name_to_id

    print(f"  Creating {len(needed)} new entity entries:")
    sql = """
        INSERT INTO bd.market_intel_corporate_entities (corporate_name, ownership_type)
        VALUES (%s, %s)
        RETURNING id
    """
    for name in sorted(needed):
        cursor.execute(sql, (name, "Corporate"))
        new_id = cursor.fetchone()[0]
        name_to_id[name.upper()] = new_id
        print(f"    [{new_id}] {name}")

    return name_to_id


def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print(f"V25.2 -> PostgreSQL Loader {'(DRY RUN)' if dry_run else ''}")
    print("=" * 60)

    if not os.path.exists(V25_2_PATH):
        print(f"ERROR: V25.2 file not found: {V25_2_PATH}")
        sys.exit(1)
    print(f"  Source: {V25_2_PATH}")

    print("\nStep 1: Reading V25.2 facilities...")
    facilities = read_facilities()

    print("\nStep 2: Connecting to PostgreSQL...")
    conn = get_connection()
    cursor = conn.cursor()
    print("  Connected.")

    print("\nStep 3: Loading existing entity map (PRESERVED — not truncating)...")
    name_to_id = get_existing_entity_map(cursor)

    # Collect distinct corp names from V25.2
    v25_2_names = set()
    for f in facilities:
        if f["corporate_name_raw"]:
            v25_2_names.add(f["corporate_name_raw"])
    print(f"  Distinct corporate names in V25.2: {len(v25_2_names)}")

    missing = [n for n in v25_2_names if n.upper() not in name_to_id]
    print(f"  Names needing new entity entries: {len(missing)}")
    if missing:
        for m in sorted(missing)[:20]:
            print(f"    {m}")
        if len(missing) > 20:
            print(f"    ... and {len(missing) - 20} more")

    if dry_run:
        print(f"\nDRY RUN — no changes will be made.")
        print(f"  Would reload {len(facilities)} facilities")
        print(f"  Would create {len(missing)} new entity entries")
        cursor.close()
        conn.close()
        return

    print("\nStep 4: Ensuring campus_id column exists...")
    ensure_campus_id_column(cursor)

    print("\nStep 5: Creating missing entity entries...")
    name_to_id = create_missing_entities(cursor, facilities, name_to_id)

    print("\nStep 6: Truncating facilities table ONLY...")
    cursor.execute("TRUNCATE TABLE bd.market_intel_facilities CASCADE")
    print("  Truncated bd.market_intel_facilities")

    print("\nStep 7: Inserting facilities...")
    sql = """
        INSERT INTO bd.market_intel_facilities
            (facility_name, corporate_entity_id, corporate_name_raw, source_type,
             address, city, state, zip, county, ownership_type,
             total_beds, census, do_we_serve, integrated_flag, pcp_flag, mh_flag,
             barrier, latitude, longitude, contract_status,
             geographic_tier, original_geographic_tier, metro_assignment,
             distance_to_metro_center, metro_center_used, corp_attribution_source,
             data_quality_flag, campus_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """
    count = 0
    fk_matched = 0
    fk_null = 0
    for f in facilities:
        corp_id = None
        raw_name = f["corporate_name_raw"]
        if raw_name:
            corp_id = name_to_id.get(raw_name.upper())
            if corp_id:
                fk_matched += 1
            else:
                fk_null += 1
        else:
            fk_null += 1

        cursor.execute(sql, (
            f["facility_name"], corp_id, f["corporate_name_raw"], f["source_type"],
            f["address"], f["city"], f["state"], f["zip"], f["county"],
            f["ownership_type"], f["total_beds"], f["census"],
            f["do_we_serve"], f["integrated_flag"], f["pcp_flag"], f["mh_flag"],
            f["barrier"], f["latitude"], f["longitude"], f["contract_status"],
            f["geographic_tier"], f["original_geographic_tier"],
            f["metro_assignment"], f["distance_to_metro_center"],
            f["metro_center_used"], f["corp_attribution_source"],
            f["data_quality_flag"], f["campus_id"],
        ))
        count += 1
        if count % 5000 == 0:
            print(f"    ... {count} inserted")

    print(f"  Inserted {count} facilities ({fk_matched} FK matched, {fk_null} FK null)")

    print("\nStep 8: Committing...")
    conn.commit()
    print("  Committed.")

    print("\nStep 9: Verification...")
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_corporate_entities")
    ce_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_facilities")
    f_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_facilities WHERE do_we_serve = true")
    served = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bd.market_intel_facilities WHERE campus_id IS NOT NULL")
    campus = cursor.fetchone()[0]
    cursor.execute("""
        SELECT ce.corporate_name, COUNT(f.id)
        FROM bd.market_intel_corporate_entities ce
        JOIN bd.market_intel_facilities f ON f.corporate_entity_id = ce.id
        GROUP BY ce.corporate_name ORDER BY COUNT(f.id) DESC LIMIT 10
    """)
    top10 = cursor.fetchall()

    print(f"  Corporate entities: {ce_count}")
    print(f"  Facilities: {f_count}")
    print(f"  Served: {served}")
    print(f"  With Campus_ID: {campus}")
    print(f"  Top 10 by facility count:")
    for row in top10:
        print(f"    {row[0]}: {row[1]}")

    cursor.close()
    conn.close()
    print("\nDone. V25.2 loaded successfully.")


if __name__ == "__main__":
    main()
