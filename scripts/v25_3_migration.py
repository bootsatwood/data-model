"""
V25.3 Migration — V25.2 YAD Correction + Missing Facility Inserts

Source: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_2.xlsx
Output: Vault/02_Data_Model/Current/1_Combined_Database_FINAL_V25_3.xlsx

Changes:
  A: REVERT 10 YAD rows incorrectly recoded to Choice Health in V25.2
     (Fletcher, Ramseur, Laurel Park, Windsor, Seven Oaks/Brian Center)
     All 5 are Tzvi Alter-owned via holding LLCs, not Donald Beaver/Choice Health.
     Fletcher + Ramseur were CHOWed from Beaver to Alter in June 2024.
  B: Recode Litchford Falls from LIFEWORKS REHAB to CHOICE HEALTH MANAGEMENT
     (confirmed Donald Beaver ownership via CMS)
  C: Insert 3 missing facilities (Aspen Alcove x2, Vitality West End Richmond)
"""
import openpyxl
import re
import os
from collections import defaultdict

VAULT = os.path.expanduser(
    "~/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current"
)
V25_2_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_2.xlsx")
V25_3_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_3.xlsx")

COL = {}

def detect_columns(ws):
    global COL
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            COL[val] = c

def norm(s):
    if not s: return ''
    return ' '.join(str(s).upper().strip().split())

def find_row(ws, name_frag, state, city=None, corp=None):
    matches = []
    for r in range(2, ws.max_row + 1):
        rn = norm(ws.cell(r, COL['Facility_Name']).value)
        rs = norm(ws.cell(r, COL['State']).value)
        if name_frag.upper() not in rn or state.upper() != rs:
            continue
        if city and city.upper() != norm(ws.cell(r, COL['City']).value):
            continue
        if corp and corp.upper() != norm(ws.cell(r, COL['Corporate_Name']).value):
            continue
        matches.append(r)
    if len(matches) == 1: return matches[0]
    elif len(matches) == 0: return None
    else: return matches

# ============================================================
# PART A: Revert 10 YAD rows incorrectly coded as Choice Health
# ============================================================
YAD_REVERTS = [
    # (name_frag, city, state) — all currently CHOICE HEALTH MANAGEMENT, revert to YAD HEALTHCARE
    ("FLETCHER REHABILITATION AND HEALTHCARE CENTER", "FLETCHER", "NC"),
    ("FLETCHER REHABILITATION AND HEATHCARE CENTER", "Fletcher", "NC"),
    ("RAMSEUR REHABILITATION AND HEALTH CARE CENTER", "Ramseur", "NC"),
    ("UNIVERSAL HEALTHCARE/RAMSEUR", "RAMSEUR", "NC"),
    ("LAUREL PARK REHABILITATION AND HEALTHCARE CENTER", "Elizabeth City", "NC"),
    ("LAUREL PARK REHABILITATION & HEALTHCARE CENTER", "Elizabeth City", "NC"),
    ("WINDSOR REHABILITATION AND HEALTHCARE CENTER", "Windsor", "NC"),
    ("WINDSOR REHAB & HEALTHCARE CENTER", "Windsor", "NC"),
    ("SEVEN OAKS REHABILITATION & HEALTHCARE CENTER", "Columbia", "SC"),
    ("BRIAN CENTER NURSING CARE - ST ANDREWS", "COLUMBIA", "SC"),
]

# ============================================================
# PART B: Litchford Falls -> Choice Health Management
# ============================================================
LITCHFORD = ("LITCHFORD FALLS", "RALEIGH", "NC", "LIFEWORKS REHAB", "CHOICE HEALTH MANAGEMENT")

# ============================================================
# PART C: New facility inserts
# ============================================================
# Column order matches V25.2 headers
INSERTS = [
    {
        "Source_Type": "ALF",
        "Facility_Name": "Aspen Alcove at Bardstown",
        "Corporate_Name": "Triple Crown Senior Living",
        "Address": "107 Thruway Dr",
        "City": "Bardstown",
        "State": "KY",
        "ZIP": "40004",
        "County": "Nelson",
        "Total_Beds": 41,
        "Do_We_Serve": "No",
        "Corp_Attribution_Source": "WEB",
        "Data_Quality_Flag": "V25.3_Insert",
    },
    {
        "Source_Type": "ALF",
        "Facility_Name": "Aspen Alcove at Elizabethtown",
        "Corporate_Name": "Triple Crown Senior Living",
        "Address": "1108 Regency Way",
        "City": "Elizabethtown",
        "State": "KY",
        "ZIP": "42701",
        "County": "Hardin",
        "Total_Beds": 41,
        "Do_We_Serve": "No",
        "Corp_Attribution_Source": "WEB",
        "Data_Quality_Flag": "V25.3_Insert",
    },
    {
        "Source_Type": "ALF",
        "Facility_Name": "Vitality Living West End Richmond",
        "Corporate_Name": "VITALITY LIVING",
        "Address": "1800 Gaskins Rd",
        "City": "Henrico",
        "State": "VA",
        "ZIP": "23238",
        "County": "Henrico",
        "Total_Beds": 136,
        "Do_We_Serve": "No",
        "Corp_Attribution_Source": "WEB",
        "Data_Quality_Flag": "V25.3_Insert",
    },
]

# ============================================================
# PROCESS
# ============================================================

print("Loading V25.2...")
wb = openpyxl.load_workbook(V25_2_PATH)
ws = wb.active
detect_columns(ws)
total_rows = ws.max_row - 1
print(f"V25.2 rows: {total_rows}")

errors = []
changes = defaultdict(list)

# --- PART A: Revert YAD rows ---
print("\n--- Part A: Revert 10 YAD rows (V25.2 correction) ---")
reverted = 0
for name_frag, city, state in YAD_REVERTS:
    row = find_row(ws, name_frag, state, city=city, corp="CHOICE HEALTH MANAGEMENT")
    if row is None:
        errors.append(f"YAD REVERT: NOT FOUND — '{name_frag}' in {city}, {state} under CHOICE HEALTH MANAGEMENT")
    elif isinstance(row, list):
        errors.append(f"YAD REVERT: MULTIPLE ({len(row)}) — '{name_frag}'")
    else:
        old = ws.cell(row, COL['Corporate_Name']).value
        ws.cell(row, COL['Corporate_Name']).value = "YAD HEALTHCARE"
        fac = ws.cell(row, COL['Facility_Name']).value
        changes['yad_revert'].append(
            f"Row {row}: {fac} | {city}, {state} | '{old}' -> 'YAD HEALTHCARE' (Tzvi Alter ownership confirmed)")
        reverted += 1
print(f"  Reverted: {reverted} rows")

# --- PART B: Litchford Falls ---
print("--- Part B: Litchford Falls -> Choice Health ---")
nf, city, state, old_corp, new_corp = LITCHFORD
row = find_row(ws, nf, state, city=city, corp=old_corp)
if row is None:
    errors.append(f"LITCHFORD: NOT FOUND — '{nf}' under '{old_corp}'")
elif isinstance(row, list):
    errors.append(f"LITCHFORD: MULTIPLE ({len(row)})")
else:
    ws.cell(row, COL['Corporate_Name']).value = new_corp
    fac = ws.cell(row, COL['Facility_Name']).value
    changes['litchford'].append(
        f"Row {row}: {fac} | '{old_corp}' -> '{new_corp}' (Donald Beaver ownership confirmed via CMS)")

# --- PART C: New facility inserts ---
print("--- Part C: Insert 3 missing facilities ---")
for ins in INSERTS:
    new_row = ws.max_row + 1
    for col_name, val in ins.items():
        if col_name in COL:
            ws.cell(new_row, COL[col_name]).value = val
    changes['inserts'].append(
        f"Row {new_row}: {ins['Facility_Name']} | {ins['City']}, {ins['State']} | {ins['Corporate_Name']} | {ins['Total_Beds']} beds")

# ============================================================
# REPORT
# ============================================================
print(f"\n{'=' * 70}")
print(f"V25.3 Migration Report")
print(f"{'=' * 70}")

labels = {
    'yad_revert': 'PART A: YAD Revert (V25.2 correction)',
    'litchford': 'PART B: Litchford Falls -> Choice Health',
    'inserts': 'PART C: New Facility Inserts',
}
total_changes = 0
for key, label in labels.items():
    items = changes.get(key, [])
    total_changes += len(items)
    print(f"{label}: {len(items)}")

print(f"{'=' * 70}")
print(f"TOTAL CHANGES: {total_changes}")
print(f"Final row count: {ws.max_row - 1}")
print()

for key, label in labels.items():
    items = changes.get(key, [])
    if items:
        print(f"\n{label}:")
        for item in items:
            print(f"  {item}")

if errors:
    print(f"\n!! {len(errors)} WARNINGS !!")
    for e in errors:
        print(f"  {e}")

# ============================================================
# SAVE
# ============================================================
print(f"\nSaving to {V25_3_PATH}...")
wb.save(V25_3_PATH)
print("Done.")
wb.close()
