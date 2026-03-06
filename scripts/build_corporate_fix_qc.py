#!/usr/bin/env python3
"""Build QC Review Workbook for corporate_fix.py (V22.2 -> V22.3).

Runs the same compute_all_changes() logic as corporate_fix.py, then
writes a multi-tab Excel workbook organized by risk level so every
proposed change can be reviewed before apply.

Tabs:
  1. HIGH — Large Chain to Small    (old corp has 50+ rows, new corp has <5)
  2. HIGH — Served Facility Renames  (Do_We_Serve=Yes, Corporate_Name changing)
  3. MED-HIGH — Served Ownership Flip (served + Ownership_Type changing)
  4. MED — INDEPENDENT to Operator   (currently blank/INDEPENDENT, gaining a name)
  5. SUMMARY — All Phase 2 renames grouped by new_corp with counts
  6. FLAGGED — All 681 flagged rows (NIC operator=unknown)
  7. REMOVAL — Phase 1d row removal detail

Usage:
  python build_corporate_fix_qc.py
"""

import sys
sys.path.insert(0, ".")

from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from corporate_fix import compute_all_changes
from utils import safe, load_db, ensure_report_dir, VAULT

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_headers(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = width + 2


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------

print("Computing all changes (same logic as corporate_fix.py preview)...")
data = compute_all_changes()

# Also load the full DB rows for lookup
DB_SOURCE = VAULT / "Current" / "1_Combined_Database_FINAL_V22_2.xlsx"
_, db_rows = load_db(DB_SOURCE)
db_by_row = {r['_excel_row']: r for r in db_rows}

# Count facilities per old Corporate_Name (V22.2 baseline)
old_corp_counts = Counter()
for r in db_rows:
    corp = safe(r.get('Corporate_Name', ''))
    if corp and corp.upper() != 'INDEPENDENT':
        old_corp_counts[corp] += 1

# Count facilities per new Corporate_Name (post Phase 2)
new_corp_counts = data['corp_counts']

# Build lookup: excel_row -> ownership change
own_change_map = {}
for c in data['ownership_changes']:
    own_change_map[c['excel_row']] = c

# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------

wb = Workbook()

# ===================================================================
# Tab 1: HIGH — Large Chain to Small
# ===================================================================

ws1 = wb.active
ws1.title = "HIGH - Chain to Small"

headers1 = [
    "Excel_Row", "Facility_Name", "City", "State", "Source_Type",
    "Do_We_Serve", "Old_Corporate_Name", "Old_Corp_Count",
    "New_Corporate_Name", "New_Corp_Count", "Match_Type",
    "Count_Ratio", "Risk_Note"
]
ws1.append(headers1)
style_headers(ws1, len(headers1))

chain_to_small = []
for c in data['phase2_auto']:
    er = c['excel_row']
    r = db_by_row.get(er, {})
    old_count = old_corp_counts.get(c['old_corp'], 0)
    new_count = new_corp_counts.get(c['new_corp'], 0)

    # Flag: old corp has 20+ rows AND new corp has fewer than old/10
    if old_count >= 20 and new_count < max(old_count // 10, 5):
        risk = "LARGE CHAIN -> TINY OPERATOR"
    elif old_count >= 50 and new_count < old_count // 3:
        risk = "LARGE CHAIN -> SMALL OPERATOR"
    else:
        continue

    chain_to_small.append({
        'excel_row': er,
        'facility': safe(r.get('Facility_Name', '')),
        'city': safe(r.get('City', '')),
        'state': safe(r.get('State', '')),
        'source_type': safe(r.get('Source_Type', '')),
        'served': safe(r.get('Do_We_Serve', '')),
        'old_corp': c['old_corp'],
        'old_count': old_count,
        'new_corp': c['new_corp'],
        'new_count': new_count,
        'match_type': c['match_type'],
        'ratio': f"{old_count}:{new_count}",
        'risk': risk,
    })

# Sort by old_count descending
chain_to_small.sort(key=lambda x: -x['old_count'])

for item in chain_to_small:
    row_data = [
        item['excel_row'], item['facility'], item['city'], item['state'],
        item['source_type'], item['served'], item['old_corp'], item['old_count'],
        item['new_corp'], item['new_count'], item['match_type'],
        item['ratio'], item['risk']
    ]
    ws1.append(row_data)
    row_num = ws1.max_row
    if item['served'] == 'Yes':
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = RED_FILL
    else:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = ORANGE_FILL

auto_width(ws1)
print(f"  Tab 1: {len(chain_to_small)} large-chain-to-small renames")


# ===================================================================
# Tab 2: HIGH — Served Facility Renames
# ===================================================================

ws2 = wb.create_sheet("HIGH - Served Renames")

headers2 = [
    "Excel_Row", "Facility_Name", "City", "State", "Source_Type",
    "Old_Corporate_Name", "Old_Corp_Count",
    "New_Corporate_Name", "New_Corp_Count", "Match_Type",
    "NIC_Operator", "Ownership_Change"
]
ws2.append(headers2)
style_headers(ws2, len(headers2))

served_renames = []
for c in data['phase2_auto']:
    er = c['excel_row']
    r = db_by_row.get(er, {})
    if safe(r.get('Do_We_Serve', '')) != 'Yes':
        continue

    old_count = old_corp_counts.get(c['old_corp'], 0)
    new_count = new_corp_counts.get(c['new_corp'], 0)
    own_chg = own_change_map.get(er, {})
    own_str = ""
    if own_chg:
        own_str = f"{own_chg['old_own']} -> {own_chg['new_own']}"

    served_renames.append({
        'excel_row': er,
        'facility': safe(r.get('Facility_Name', '')),
        'city': safe(r.get('City', '')),
        'state': safe(r.get('State', '')),
        'source_type': safe(r.get('Source_Type', '')),
        'old_corp': c['old_corp'],
        'old_count': old_count,
        'new_corp': c['new_corp'],
        'new_count': new_count,
        'match_type': c['match_type'],
        'nic_op': c['nic_operator'],
        'own_change': own_str,
    })

served_renames.sort(key=lambda x: (x['state'], x['city']))

for item in served_renames:
    row_data = [
        item['excel_row'], item['facility'], item['city'], item['state'],
        item['source_type'], item['old_corp'], item['old_count'],
        item['new_corp'], item['new_count'], item['match_type'],
        item['nic_op'], item['own_change']
    ]
    ws2.append(row_data)
    row_num = ws2.max_row
    # Color by risk: red if also losing corporate, orange if operator-to-different-operator
    if "Corporate -> Independent" in item['own_change']:
        fill = RED_FILL
    elif item['old_corp'] and item['old_corp'].upper() != 'INDEPENDENT' and item['old_corp'] != 'unknown':
        fill = ORANGE_FILL
    else:
        fill = YELLOW_FILL
    for col in range(1, len(headers2) + 1):
        ws2.cell(row=row_num, column=col).fill = fill

auto_width(ws2)
print(f"  Tab 2: {len(served_renames)} served facility renames")


# ===================================================================
# Tab 3: MED-HIGH — Served Ownership Flip
# ===================================================================

ws3 = wb.create_sheet("MED-HIGH - Served Own Flip")

headers3 = [
    "Excel_Row", "Facility_Name", "City", "State", "Source_Type",
    "Corporate_Name", "Corp_Count", "Old_Ownership", "New_Ownership",
    "Direction"
]
ws3.append(headers3)
style_headers(ws3, len(headers3))

served_own_flips = []
for c in data['ownership_changes']:
    er = c['excel_row']
    r = db_by_row.get(er, {})
    if safe(r.get('Do_We_Serve', '')) != 'Yes':
        continue

    direction = f"{c['old_own']} -> {c['new_own']}"
    served_own_flips.append({
        'excel_row': er,
        'facility': safe(r.get('Facility_Name', '')),
        'city': safe(r.get('City', '')),
        'state': safe(r.get('State', '')),
        'source_type': safe(r.get('Source_Type', '')),
        'corp_name': c['corp_name'],
        'corp_count': c['corp_count'],
        'old_own': c['old_own'],
        'new_own': c['new_own'],
        'direction': direction,
    })

served_own_flips.sort(key=lambda x: (x['direction'], x['state'], x['city']))

for item in served_own_flips:
    row_data = [
        item['excel_row'], item['facility'], item['city'], item['state'],
        item['source_type'], item['corp_name'], item['corp_count'],
        item['old_own'], item['new_own'], item['direction']
    ]
    ws3.append(row_data)
    row_num = ws3.max_row
    if item['direction'].startswith("Corporate -> Independent"):
        fill = RED_FILL
    else:
        fill = GREEN_FILL
    for col in range(1, len(headers3) + 1):
        ws3.cell(row=row_num, column=col).fill = fill

auto_width(ws3)
i2c_served = sum(1 for x in served_own_flips if x['old_own'] == 'Independent')
c2i_served = sum(1 for x in served_own_flips if x['old_own'] == 'Corporate')
print(f"  Tab 3: {len(served_own_flips)} served ownership flips "
      f"({i2c_served} Ind->Corp, {c2i_served} Corp->Ind)")


# ===================================================================
# Tab 4: MED — INDEPENDENT/Blank to Operator
# ===================================================================

ws4 = wb.create_sheet("MED - Ind to Operator")

headers4 = [
    "Excel_Row", "Facility_Name", "City", "State", "Source_Type",
    "Do_We_Serve", "Old_Corporate_Name",
    "New_Corporate_Name", "New_Corp_Count", "Match_Type",
    "NIC_Operator"
]
ws4.append(headers4)
style_headers(ws4, len(headers4))

ind_to_op = []
for c in data['phase2_auto']:
    old = c['old_corp']
    if old and old.upper() != 'INDEPENDENT' and old.lower() != 'unknown':
        continue

    er = c['excel_row']
    r = db_by_row.get(er, {})
    new_count = new_corp_counts.get(c['new_corp'], 0)
    ind_to_op.append({
        'excel_row': er,
        'facility': safe(r.get('Facility_Name', '')),
        'city': safe(r.get('City', '')),
        'state': safe(r.get('State', '')),
        'source_type': safe(r.get('Source_Type', '')),
        'served': safe(r.get('Do_We_Serve', '')),
        'old_corp': old,
        'new_corp': c['new_corp'],
        'new_count': new_count,
        'match_type': c['match_type'],
        'nic_op': c['nic_operator'],
    })

ind_to_op.sort(key=lambda x: (-1 if x['served'] == 'Yes' else 0, x['state'], x['city']))

for item in ind_to_op:
    row_data = [
        item['excel_row'], item['facility'], item['city'], item['state'],
        item['source_type'], item['served'], item['old_corp'],
        item['new_corp'], item['new_count'], item['match_type'],
        item['nic_op']
    ]
    ws4.append(row_data)
    row_num = ws4.max_row
    if item['served'] == 'Yes':
        fill = ORANGE_FILL
    elif item['new_count'] == 1:
        fill = YELLOW_FILL
    else:
        fill = GREEN_FILL
    for col in range(1, len(headers4) + 1):
        ws4.cell(row=row_num, column=col).fill = fill

auto_width(ws4)
served_ind = sum(1 for x in ind_to_op if x['served'] == 'Yes')
print(f"  Tab 4: {len(ind_to_op)} INDEPENDENT/blank -> operator "
      f"({served_ind} served)")


# ===================================================================
# Tab 5: SUMMARY — All Phase 2 renames by new_corp
# ===================================================================

ws5 = wb.create_sheet("SUMMARY - By New Corp")

headers5 = [
    "New_Corporate_Name", "Rename_Count", "Match_Type",
    "New_Corp_Total_Count", "Served_Count",
    "Top_Old_Corp_Names", "States"
]
ws5.append(headers5)
style_headers(ws5, len(headers5))

# Group by new_corp
by_new_corp = {}
for c in data['phase2_auto']:
    key = c['new_corp']
    if key not in by_new_corp:
        by_new_corp[key] = []
    by_new_corp[key].append(c)

summary_rows = []
for new_corp, changes in sorted(by_new_corp.items(), key=lambda x: -len(x[1])):
    rename_ct = len(changes)
    match_types = set(c['match_type'] for c in changes)
    match_str = "/".join(sorted(match_types))
    total_count = new_corp_counts.get(new_corp, 0)

    served_ct = 0
    old_corps = Counter()
    states = set()
    for c in changes:
        r = db_by_row.get(c['excel_row'], {})
        if safe(r.get('Do_We_Serve', '')) == 'Yes':
            served_ct += 1
        old_corps[c['old_corp']] += 1
        states.add(c['state'])

    top_old = "; ".join(f"{corp} ({ct})" for corp, ct in old_corps.most_common(3))
    state_str = ", ".join(sorted(states))

    summary_rows.append([
        new_corp, rename_ct, match_str, total_count, served_ct,
        top_old, state_str
    ])

for row_data in summary_rows:
    ws5.append(row_data)
    row_num = ws5.max_row
    if row_data[4] > 0:  # served_count > 0
        for col in range(1, len(headers5) + 1):
            ws5.cell(row=row_num, column=col).fill = ORANGE_FILL

auto_width(ws5)
print(f"  Tab 5: {len(summary_rows)} distinct new corporate names")


# ===================================================================
# Tab 6: FLAGGED — NIC operator unknown
# ===================================================================

ws6 = wb.create_sheet("FLAGGED - Unknown Op")

headers6 = [
    "Excel_Row", "Facility_Name", "City", "State", "Source_Type",
    "Do_We_Serve", "Current_Corporate_Name", "NIC_Owner",
    "Reason"
]
ws6.append(headers6)
style_headers(ws6, len(headers6))

for f in sorted(data['phase2_flagged'], key=lambda x: (x['state'], x['city'])):
    er = f['excel_row']
    r = db_by_row.get(er, {})
    row_data = [
        er, f['facility'], f['city'], f['state'],
        safe(r.get('Source_Type', '')),
        safe(r.get('Do_We_Serve', '')),
        f['old_corp'], f['nic_owner'], f['reason']
    ]
    ws6.append(row_data)
    row_num = ws6.max_row
    if safe(r.get('Do_We_Serve', '')) == 'Yes':
        for col in range(1, len(headers6) + 1):
            ws6.cell(row=row_num, column=col).fill = ORANGE_FILL

auto_width(ws6)
flagged_served = sum(1 for f in data['phase2_flagged']
                     if safe(db_by_row.get(f['excel_row'], {}).get('Do_We_Serve', '')) == 'Yes')
print(f"  Tab 6: {len(data['phase2_flagged'])} flagged rows "
      f"({flagged_served} served)")


# ===================================================================
# Tab 7: REMOVAL — Phase 1d detail
# ===================================================================

ws7 = wb.create_sheet("REMOVAL - Phase 1d")

headers7 = [
    "Excel_Row", "Facility_Name", "City", "State",
    "Corporate_Name", "Total_Beds", "Do_We_Serve",
    "Real_SNF_Row", "Real_SNF_Name"
]
ws7.append(headers7)
style_headers(ws7, len(headers7))

for d in data['phase1d']:
    row_data = [
        d['excel_row'], d['facility'], d['city'], d['state'],
        d['corp'], d['beds'], d['served'],
        d['real_row'], d['real_snf']
    ]
    ws7.append(row_data)
    row_num = ws7.max_row
    if d['served'] == 'Yes':
        for col in range(1, len(headers7) + 1):
            ws7.cell(row=row_num, column=col).fill = RED_FILL

auto_width(ws7)
print(f"  Tab 7: {len(data['phase1d'])} rows to remove")


# ===================================================================
# Save
# ===================================================================

output_path = ensure_report_dir() / "Corporate_Fix_QC_Review.xlsx"
wb.save(output_path)
print(f"\nSaved: {output_path}")

# ===================================================================
# Console summary
# ===================================================================

print()
print("=" * 60)
print("QC REVIEW WORKBOOK SUMMARY")
print("=" * 60)
print(f"  Tab 1: HIGH - Chain to Small:       {len(chain_to_small)} rows")
print(f"  Tab 2: HIGH - Served Renames:       {len(served_renames)} rows")
print(f"  Tab 3: MED-HIGH - Served Own Flip:  {len(served_own_flips)} rows")
print(f"    Independent -> Corporate:         {i2c_served}")
print(f"    Corporate -> Independent:         {c2i_served}")
print(f"  Tab 4: MED - Ind to Operator:       {len(ind_to_op)} rows ({served_ind} served)")
print(f"  Tab 5: SUMMARY - By New Corp:       {len(summary_rows)} distinct names")
print(f"  Tab 6: FLAGGED - Unknown Op:        {len(data['phase2_flagged'])} rows ({flagged_served} served)")
print(f"  Tab 7: REMOVAL - Phase 1d:          {len(data['phase1d'])} rows")
print()
print("Color key:")
print("  RED    = served + high risk (losing Corporate status or chain mismatch)")
print("  ORANGE = served or medium risk")
print("  YELLOW = low-medium risk (single-facility new name)")
print("  GREEN  = low risk (gaining Corporate status or clean match)")
print()
print("Review priority:")
print("  1. Tab 1 — are any of these false address matches?")
print("  2. Tab 2 — spot-check served renames, especially red/orange rows")
print("  3. Tab 3 — Corp->Ind served rows (red) need confirmation")
print("  4. Tab 5 — scan new corporate names for person-names or facility-names")
