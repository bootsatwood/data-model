"""V20 → V23 Movers Analysis: Entity-by-entity tier changes with dimension-level explanations.
Outputs both console summary and HTML report for stakeholder review."""

import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

# ============================================================
# FILE PATHS
# ============================================================
V20_PATH = "C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx"
V23_BD_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_Final_MUO_Tiering_V23.xlsx"
V23_FIN_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_MUO_Scoring_Workbook_V23_v8.xlsx"
HTML_OUT = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/V20_V23_Movers_Analysis.html"

# ============================================================
# V20 NAME -> V23 DISPLAY NAME MAPPING
# ============================================================
V20_TO_V23 = {
    'ALG': 'ALG',
    'American Senior Communities': 'American Senior Communities',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'Cch Healthcare': 'CCH Healthcare',
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': 'Avardis',
    'Infinity Healthcare Consulting': 'Infinity Healthcare Consulting',
    'Liberty Senior Living': 'Liberty',
    'Life Care Centers Of America': 'Lifecare',
    'Majestic Care': 'Majestic Care',
    'Navion': 'Navion',
    'Otterbein Seniorlife': 'Otterbein Senior Life',
    'Principle Long Term Care': 'Principle',
    'Pruitthealth': 'Pruitt Health',
    'Saber Healthcare Group': 'Saber Healthcare Group',
    'Terra Bella': 'TerraBella Senior Living',
    'Tlc Management': 'TLC Management',
    'Trilogy Health Services': 'Trilogy',
    'Arbors At Ohio': 'Arbors',
    'BHI Senior Living': 'BHI Senior Living',
    'Castle Healthcare': 'Castle Healthcare',
    'Heritage Hall': 'American Healthcare LLC',
    'Jag Healthcare': 'JAG',
    'Kissito': 'Kissito Healthcare',
    'Lionstone Care': 'Lionstone Care',
    'Lutheran Services Carolina': 'Lutheran Services Carolinas',
    'Lutheran Services Carolinas': 'Lutheran Services Carolinas',
    'Momentous Health': 'Momentus Health',
    'Morning Pointe Senior Living': 'Morning Pointe Senior Living',
    'Peak Resources, Inc.': 'Peak Resources',
    'Priority Life Care': 'Priority',
    'Sanstone Health & Rehabilitation': 'Sanstone',
    'Topaz': 'Topaz Healthcare',
    'Yad Healthcare': 'YAD',
    'Caring Place Healthcare': 'Caring Place Healthcare',
    'Clearview': 'Clearview',
    'Pavilion Healthcare': 'Pavilion Healthcare',
    'Eldercare Partners': 'Eldercare Partners',
    'Fundamental Ltc': 'Fundamental LTC',
    'Southern Healthcare Management, LLC': 'Southern Healthcare Mgmt',
    'Sonida Senior Living': 'Sonida Senior Living',
    'Kisco Senior Living': 'Kisco Senior Living',
    'Spring Arbor Management': 'Spring Arbor Management',
    'Cedarhurst Senior Living': 'Cedarhurst Senior Living',
    'Senior Lifestyle': 'Senior Lifestyle',
    'StoryPoint': 'StoryPoint',
    'Runk & Pratt': 'Runk & Pratt',
    'Greencroft': 'Greencroft',
    'LifeSpire of Virginia': 'LifeSpire of Virginia',
    'Brighton': 'Brighton',
    'Sunnyside Communities': 'Sunnyside Communities',
    'Lutheran Life Villages': 'Lutheran Life Villages',
    'Triple Crown Senior Living': 'Triple Crown Senior Living',
    'Warm Hearth Village': 'Warm Hearth Village',
    'Atrium Health': 'Atrium Health',
    # Gap entities (V20 scored, not in Finance 60)
    'Altercare': 'Altercare',
    'Aom Healthcare': 'AOM Healthcare',
    'Aperion Care': 'Aperion Care',
    'Brickyard Healthcare': 'Brickyard Healthcare',
    'Carecore Health': 'Carecore Health',
    'Carespring': 'Carespring',
    'Carrolton': 'Carrolton',
    'Certus Healthcare': 'Certus Healthcare',
    'Choice Health Management': 'Choice Health Management',
    'Ciena Healthcare/Laurel Health Care': 'Ciena Healthcare/Laurel Health Care',
    'Continuing Healthcare Solutions': 'Continuing Healthcare Solutions',
    'Crown Healthcare Group': 'Crown Healthcare Group',
    'Divine Healthcare Management': 'Divine Healthcare Management',
    'Envive Healthcare': 'Envive Healthcare',
    'Gardant Management Solutions, Inc': 'Gardant Management Solutions',
    'HCF Management': 'HCF Management',
    'Health Care Management Group': 'Health Care Management Group',
    'Hillstone Healthcare': 'Hillstone Healthcare',
    'Legacy Health Services': 'Legacy Health Services',
    "Miller's Merry Manor": "Miller's Merry Manor",
    'National Healthcare Corporation': 'National Healthcare Corporation',
    'Ohio Living Communities': 'Ohio Living Communities',
    'Optalis Health & Rehabilitation': 'Optalis Health & Rehabilitation',
    'PACS Group': 'PACS Group',
    'Phoenix Senior Living': 'Phoenix Senior Living',
    'Progressive Quality Care': 'Progressive Quality Care',
    'Seky Holding Co.': 'SEKY Holding Co.',
    'Sprenger Health Care Systems': 'Sprenger Health Care Systems',
    'Sunrise': 'Sunrise Senior Living',
    'Trio Healthcare': 'Trio Healthcare',
    'Vancrest Health Care Centers': 'Vancrest Health Care Centers',
    'White Oak Management': 'White Oak Management',
    'Windsor House, Inc.': 'Windsor House',
    # Absorbed
    'Southern Assisted Living, LLC': None,  # absorbed by Brookdale
    'Sovereign Healthcare Holdings': None,  # = Southern Healthcare Mgmt
    # V20 T5 barriers
    'Bluegrass': 'Bluegrass/Encore',
    'Encore': 'Bluegrass/Encore',
    'Signature': 'SIGNATURE HEALTH',
    'MFA': 'MFA',
    'Communicare': 'COMMUNICARE',
    'Hill Valley': 'Hill Valley',
    'Singh': 'SINGH',
    'Cardon': 'CARDON & ASSOCIATES',
    'Eastern': 'Eastern Healthcare Group',
    'Embassy': 'Embassy',
    'Exceptional Living': 'Exceptional Living',
    'Portopiccolo': 'Portopiccolo',
    'Venza': 'Venza',
    'Aventura': 'Aventura',
    'Journey': 'Journey',
}

# V23 T5 entities (won't appear in scored tabs)
V23_T5_NAMES = {
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare',
    'Embassy', 'Exceptional Living', 'Portopiccolo', 'Venza', 'Aventura', 'Journey',
}

TIER_RANK = {'T1': 1, 'T2': 2, 'T3': 3, 'T4': 4, 'T5': 5}
DIMS = ['ER', 'IR', 'SI', 'RP', 'RS', 'AI']
DIM_WEIGHTS = {'ER': 4, 'IR': 3, 'SI': 3, 'RP': 4, 'RS': 3, 'AI': 3}
DIM_LABELS = {
    'ER': 'Enterprise Reach',
    'IR': 'Integration Readiness',
    'SI': 'Strategic Influence',
    'RP': 'Revenue Potential',
    'RS': 'Relationship Strength',
    'AI': 'AI/Tech Adoption',
}

# Methodology change descriptions for explaining dimension shifts
METHOD_CHANGES = {
    'ER': 'V23 counts street-number-collapsed campuses in 6-state footprint (was raw facility rows nationally)',
    'IR': 'V23 uses DB service flags MH/PCP/Integrated (was Brooke qualitative)',
    'SI': 'V23 uses Tom Notes + web research on REIT/ISNP/ACO (was Brooke qualitative)',
    'RP': 'V23 uses S2 TAM revenue in footprint with $1M/$2.5M/$5M/$10M brackets (was consent-based $250K/$500K/$1M/$2M)',
    'RS': 'V23 uses Tom Shared Savings Notes (was Brooke qualitative)',
    'AI': 'Carried from V20 where available, default 0 for new entities',
}

# ============================================================
# READ V20 WORKBOOK
# ============================================================
print("Loading V20 workbook...")
wb_v20 = openpyxl.load_workbook(V20_PATH, data_only=True)

v20_entities = {}

for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_v20[sn]
    tier_label = sn.split()[0]  # T1, T2, T3
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip().upper() == 'TOTAL':
            continue
        name = str(name).strip()
        score = ws.cell(r, 3).value or 0
        er = ws.cell(r, 9).value or 0
        ir = ws.cell(r, 10).value or 0
        si = ws.cell(r, 11).value or 0
        rp = ws.cell(r, 12).value or 0
        rs = ws.cell(r, 13).value or 0
        ai = ws.cell(r, 14).value or 0
        v20_entities[name] = {
            'tier': tier_label,
            'score': int(score) if score else 0,
            'ER': int(er) if er else 0,
            'IR': int(ir) if ir else 0,
            'SI': int(si) if si else 0,
            'RP': int(rp) if rp else 0,
            'RS': int(rs) if rs else 0,
            'AI': int(ai) if ai else 0,
        }

# V20 T5
v20_t5 = set()
try:
    ws5 = wb_v20['T5- Hard Barriers ']
    for r in range(2, ws5.max_row + 1):
        name = ws5.cell(r, 1).value
        if name and str(name).strip().upper() != 'TOTAL':
            v20_t5.add(str(name).strip())
except KeyError:
    # Try alternate tab names
    for sn in wb_v20.sheetnames:
        if 'T5' in sn or 'Barrier' in sn:
            ws5 = wb_v20[sn]
            for r in range(2, ws5.max_row + 1):
                name = ws5.cell(r, 1).value
                if name and str(name).strip().upper() != 'TOTAL':
                    v20_t5.add(str(name).strip())

print(f"  V20: {len(v20_entities)} scored entities + {len(v20_t5)} T5 barriers")

# ============================================================
# READ V23 BD WORKBOOK
# ============================================================
print("Loading V23 BD workbook...")
wb_v23 = openpyxl.load_workbook(V23_BD_PATH, data_only=True)

v23_entities = {}

for sn in wb_v23.sheetnames:
    ws = wb_v23[sn]
    if 'T1' in sn or 'T2' in sn or 'T3' in sn:
        tier_label = sn.split()[0]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip().upper() == 'TOTAL':
                continue
            name = str(name).strip()
            score = ws.cell(r, 3).value or 0
            er = ws.cell(r, 9).value or 0
            ir = ws.cell(r, 10).value or 0
            si = ws.cell(r, 11).value or 0
            rp = ws.cell(r, 12).value or 0
            rs = ws.cell(r, 13).value or 0
            ai = ws.cell(r, 14).value or 0
            v23_entities[name] = {
                'tier': tier_label,
                'score': int(score) if score else 0,
                'ER': int(er) if er else 0,
                'IR': int(ir) if ir else 0,
                'SI': int(si) if si else 0,
                'RP': int(rp) if rp else 0,
                'RS': int(rs) if rs else 0,
                'AI': int(ai) if ai else 0,
            }
    elif 'T4' in sn:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() in ('TOTAL', 'FACILITY DETAIL', ''):
                continue
            name = str(name).strip()
            camps = ws.cell(r, 2).value or 0
            v23_entities[name] = {
                'tier': 'T4',
                'score': None,
                'ER': None, 'IR': None, 'SI': None,
                'RP': None, 'RS': None, 'AI': None,
                'campuses': int(camps) if camps else 0,
            }

# Add T5 entities
for t5_name in V23_T5_NAMES:
    if t5_name not in v23_entities:
        v23_entities[t5_name] = {
            'tier': 'T5',
            'score': None,
            'ER': None, 'IR': None, 'SI': None,
            'RP': None, 'RS': None, 'AI': None,
        }

print(f"  V23: {len(v23_entities)} total entities")

# ============================================================
# PAIR ENTITIES
# ============================================================
print("\nPairing V20 → V23 entities...")

# Also build case-insensitive lookup for V23
v23_upper_lookup = {}
for name in v23_entities:
    v23_upper_lookup[name.upper()] = name

paired = []
unmatched = []

for v20_name, v20_data in v20_entities.items():
    v23_display = V20_TO_V23.get(v20_name)

    if v23_display is None:
        # Absorbed entity
        paired.append({
            'v20_name': v20_name,
            'v23_name': '(absorbed)',
            'v20': v20_data,
            'v23': None,
            'status': 'absorbed',
        })
        continue

    # Look up in V23 entities
    v23_data = None
    if v23_display in v23_entities:
        v23_data = v23_entities[v23_display]
    elif v23_display.upper() in v23_upper_lookup:
        actual_name = v23_upper_lookup[v23_display.upper()]
        v23_data = v23_entities[actual_name]
        v23_display = actual_name
    else:
        # Try case-insensitive on original name
        if v20_name.upper() in v23_upper_lookup:
            actual_name = v23_upper_lookup[v20_name.upper()]
            v23_data = v23_entities[actual_name]
            v23_display = actual_name

    if v23_data:
        paired.append({
            'v20_name': v20_name,
            'v23_name': v23_display,
            'v20': v20_data,
            'v23': v23_data,
            'status': 'matched',
        })
    else:
        unmatched.append(v20_name)

# V20 T5 entities
for v20_t5_name in v20_t5:
    v23_display = V20_TO_V23.get(v20_t5_name, v20_t5_name)
    if v23_display:
        paired.append({
            'v20_name': v20_t5_name,
            'v23_name': v23_display,
            'v20': {'tier': 'T5', 'score': None, 'ER': None, 'IR': None, 'SI': None, 'RP': None, 'RS': None, 'AI': None},
            'v23': v23_entities.get(v23_display, v23_entities.get(v23_display.upper() if v23_display.upper() in v23_upper_lookup else '', None)),
            'status': 'matched',
        })

print(f"  Paired: {len(paired)}, Unmatched: {len(unmatched)}")
if unmatched:
    print(f"  Unmatched V20 names: {unmatched}")

# ============================================================
# ANALYZE MOVERS
# ============================================================
movers_up = []    # tier improved (lower number)
movers_down = []  # tier worsened (higher number)
same_tier = []    # same tier
absorbed = []     # no longer exist
new_to_v23 = []   # in V23 but not V20

for p in paired:
    if p['status'] == 'absorbed':
        absorbed.append(p)
        continue

    v20_tier = p['v20']['tier']
    v23_tier = p['v23']['tier'] if p['v23'] else '?'

    v20_rank = TIER_RANK.get(v20_tier, 99)
    v23_rank = TIER_RANK.get(v23_tier, 99)

    p['v20_tier'] = v20_tier
    p['v23_tier'] = v23_tier
    p['tier_change'] = v20_rank - v23_rank  # positive = upgraded, negative = downgraded

    # Build dimension change explanations
    dim_changes = []
    if p['v20']['score'] is not None and p['v23'] and p['v23']['score'] is not None:
        p['score_delta'] = p['v23']['score'] - p['v20']['score']
        for dim in DIMS:
            v20_val = p['v20'].get(dim, 0) or 0
            v23_val = p['v23'].get(dim, 0) or 0
            delta = v23_val - v20_val
            weighted_delta = delta * DIM_WEIGHTS[dim]
            if delta != 0:
                dim_changes.append({
                    'dim': dim,
                    'label': DIM_LABELS[dim],
                    'v20': v20_val,
                    'v23': v23_val,
                    'delta': delta,
                    'weighted_delta': weighted_delta,
                    'method_note': METHOD_CHANGES[dim],
                })
    else:
        p['score_delta'] = None

    p['dim_changes'] = dim_changes

    if v20_rank > v23_rank:
        movers_up.append(p)
    elif v20_rank < v23_rank:
        movers_down.append(p)
    else:
        same_tier.append(p)

# Check for V23-only entities (new entities not in V20)
v20_v23_names = set(p['v23_name'] for p in paired if p['v23_name'] != '(absorbed)')
for v23_name, v23_data in v23_entities.items():
    if v23_name not in v20_v23_names and v23_name.upper() not in {n.upper() for n in v20_v23_names}:
        new_to_v23.append({'v23_name': v23_name, 'v23': v23_data})

# Sort
movers_up.sort(key=lambda x: (-abs(x['tier_change']), x['v23_name']))
movers_down.sort(key=lambda x: (-abs(x['tier_change']), x['v23_name']))
same_tier.sort(key=lambda x: x['v23_name'])

# ============================================================
# CONSOLE OUTPUT
# ============================================================
print(f"\n{'='*120}")
print(f"V20 → V23 MOVERS ANALYSIS")
print(f"{'='*120}")

print(f"\n  UPGRADED:    {len(movers_up)} entities moved to a better tier")
print(f"  DOWNGRADED:  {len(movers_down)} entities moved to a worse tier")
print(f"  SAME TIER:   {len(same_tier)} entities stayed in same tier")
print(f"  ABSORBED:    {len(absorbed)} entities no longer exist")
print(f"  NEW TO V23:  {len(new_to_v23)} entities added in V23")

if movers_up:
    print(f"\n{'='*120}")
    print(f"UPGRADED ({len(movers_up)})")
    print(f"{'='*120}")
    for p in movers_up:
        print(f"\n  {p['v23_name']}")
        print(f"    {p['v20_tier']} → {p['v23_tier']}  (score: {p['v20']['score'] or 'n/a'} → {p['v23']['score'] or 'n/a'})")
        if p['dim_changes']:
            for dc in sorted(p['dim_changes'], key=lambda x: -abs(x['weighted_delta'])):
                arrow = '↑' if dc['delta'] > 0 else '↓'
                print(f"    {dc['dim']:>2} {dc['v20']:>1}→{dc['v23']:>1} ({arrow}{abs(dc['delta'])}, wt {dc['weighted_delta']:+d})  {dc['method_note']}")

if movers_down:
    print(f"\n{'='*120}")
    print(f"DOWNGRADED ({len(movers_down)})")
    print(f"{'='*120}")
    for p in movers_down:
        print(f"\n  {p['v23_name']}")
        v23_score_str = str(p['v23']['score']) if p['v23'] and p['v23']['score'] is not None else 'n/a'
        v20_score_str = str(p['v20']['score']) if p['v20']['score'] is not None else 'n/a'
        tier_reason = ''
        if p['v23_tier'] == 'T4':
            camps = p['v23'].get('campuses', '?') if p['v23'] else '?'
            tier_reason = f"  [T4: {camps} campuses, gate requires 7]"
        elif p['v23_tier'] == 'T5':
            tier_reason = '  [T5: barrier identified]'
        print(f"    {p['v20_tier']} → {p['v23_tier']}  (score: {v20_score_str} → {v23_score_str}){tier_reason}")
        if p['dim_changes']:
            for dc in sorted(p['dim_changes'], key=lambda x: -abs(x['weighted_delta'])):
                arrow = '↑' if dc['delta'] > 0 else '↓'
                print(f"    {dc['dim']:>2} {dc['v20']:>1}→{dc['v23']:>1} ({arrow}{abs(dc['delta'])}, wt {dc['weighted_delta']:+d})  {dc['method_note']}")
        elif p['v23_tier'] in ('T4', 'T5'):
            if p['v23_tier'] == 'T4':
                print(f"    Not scored — below MUO gate (V23 requires 7+ campuses in 6-state footprint)")
            else:
                print(f"    Not scored — structural barrier prevents growth")

# ============================================================
# BUILD HTML REPORT
# ============================================================
print(f"\nBuilding HTML report...")

def tier_color(tier):
    return {
        'T1': '#C6EFCE', 'T2': '#FFEB9C', 'T3': '#FFC7CE',
        'T4': '#D6DCE4', 'T5': '#F2DCDB',
    }.get(tier, '#FFFFFF')

def tier_badge(tier):
    return f'<span style="background-color: {tier_color(tier)}; padding: 2px 8px; border-radius: 3px; font-weight: bold;">{tier}</span>'

def dim_change_row(dc):
    color = '#2E7D32' if dc['delta'] > 0 else '#C62828'
    arrow = '&#9650;' if dc['delta'] > 0 else '&#9660;'
    return f'''<tr>
        <td style="padding: 3px 10px;">{dc['dim']} — {dc['label']}</td>
        <td style="text-align: center; padding: 3px 10px;">{dc['v20']}</td>
        <td style="text-align: center; padding: 3px 10px;">{dc['v23']}</td>
        <td style="text-align: center; padding: 3px 10px; color: {color}; font-weight: bold;">{arrow} {abs(dc['delta'])} (wt {dc['weighted_delta']:+d})</td>
        <td style="padding: 3px 10px; font-size: 10pt; color: #666;">{dc['method_note']}</td>
    </tr>'''

html_parts = []
html_parts.append(f'''<html>
<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #333; max-width: 1200px; margin: 0 auto; padding: 20px;">

<h1 style="color: #1F4E79; border-bottom: 3px solid #4472C4; padding-bottom: 10px;">V20 → V23 Movers Analysis</h1>

<p>Entity-by-entity comparison of tier changes between the original V20 scoring (Brooke's qualitative model, December 2025) and V23 (data-driven methodology, March 2026).</p>

<table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; font-size: 11pt; margin: 20px 0;">
<tr style="background-color: #4472C4; color: white; font-weight: bold; text-align: center;">
<td>Category</td><td>Count</td><td>Description</td>
</tr>
<tr><td>Upgraded</td><td style="text-align: center; font-weight: bold; color: #2E7D32;">{len(movers_up)}</td><td>Moved to a better (lower-numbered) tier</td></tr>
<tr style="background-color: #F2F2F2;"><td>Downgraded</td><td style="text-align: center; font-weight: bold; color: #C62828;">{len(movers_down)}</td><td>Moved to a worse (higher-numbered) tier</td></tr>
<tr><td>Same Tier</td><td style="text-align: center;">{len(same_tier)}</td><td>Tier unchanged (score may differ)</td></tr>
<tr style="background-color: #F2F2F2;"><td>Absorbed</td><td style="text-align: center;">{len(absorbed)}</td><td>Entity no longer exists (merged into another)</td></tr>
</table>
''')

# --- UPGRADED ---
if movers_up:
    html_parts.append(f'''
<h2 style="color: #2E7D32; margin-top: 40px;">&#9650; Upgraded ({len(movers_up)})</h2>
<p>These entities moved to a better tier under V23 methodology.</p>
''')
    for p in movers_up:
        v20s = p['v20']['score'] if p['v20']['score'] is not None else 'n/a'
        v23s = p['v23']['score'] if p['v23'] and p['v23']['score'] is not None else 'n/a'
        html_parts.append(f'''
<div style="border: 1px solid #ccc; border-left: 4px solid #2E7D32; margin: 15px 0; padding: 12px 16px; border-radius: 4px;">
<h3 style="margin: 0 0 8px 0; color: #1F4E79;">{p['v23_name']}</h3>
<p style="margin: 4px 0;">{tier_badge(p['v20_tier'])} &rarr; {tier_badge(p['v23_tier'])} &nbsp;&nbsp; Score: {v20s} &rarr; {v23s}</p>
''')
        if p['dim_changes']:
            html_parts.append('''<table border="1" cellpadding="4" cellspacing="0" style="border-collapse: collapse; font-size: 10pt; margin-top: 8px; width: 100%;">
<tr style="background-color: #E8F5E9; font-weight: bold;"><td>Dimension</td><td style="text-align: center;">V20</td><td style="text-align: center;">V23</td><td style="text-align: center;">Change</td><td>Why</td></tr>''')
            for dc in sorted(p['dim_changes'], key=lambda x: -abs(x['weighted_delta'])):
                html_parts.append(dim_change_row(dc))
            html_parts.append('</table>')
        html_parts.append('</div>')

# --- DOWNGRADED ---
if movers_down:
    html_parts.append(f'''
<h2 style="color: #C62828; margin-top: 40px;">&#9660; Downgraded ({len(movers_down)})</h2>
<p>These entities moved to a worse tier under V23 methodology.</p>
''')
    for p in movers_down:
        v20s = p['v20']['score'] if p['v20']['score'] is not None else 'n/a'
        v23s = p['v23']['score'] if p['v23'] and p['v23']['score'] is not None else 'n/a'
        extra = ''
        if p['v23_tier'] == 'T4':
            camps = p['v23'].get('campuses', '?') if p['v23'] else '?'
            extra = f'<br><span style="color: #666; font-size: 10pt;">Below MUO gate: {camps} qualifying campuses in 6-state footprint (requires 7)</span>'
        elif p['v23_tier'] == 'T5':
            extra = '<br><span style="color: #666; font-size: 10pt;">Structural barrier identified — prevents growth</span>'

        html_parts.append(f'''
<div style="border: 1px solid #ccc; border-left: 4px solid #C62828; margin: 15px 0; padding: 12px 16px; border-radius: 4px;">
<h3 style="margin: 0 0 8px 0; color: #1F4E79;">{p['v23_name']}</h3>
<p style="margin: 4px 0;">{tier_badge(p['v20_tier'])} &rarr; {tier_badge(p['v23_tier'])} &nbsp;&nbsp; Score: {v20s} &rarr; {v23s}{extra}</p>
''')
        if p['dim_changes']:
            html_parts.append('''<table border="1" cellpadding="4" cellspacing="0" style="border-collapse: collapse; font-size: 10pt; margin-top: 8px; width: 100%;">
<tr style="background-color: #FFEBEE; font-weight: bold;"><td>Dimension</td><td style="text-align: center;">V20</td><td style="text-align: center;">V23</td><td style="text-align: center;">Change</td><td>Why</td></tr>''')
            for dc in sorted(p['dim_changes'], key=lambda x: -abs(x['weighted_delta'])):
                html_parts.append(dim_change_row(dc))
            html_parts.append('</table>')
        html_parts.append('</div>')

# --- SAME TIER ---
html_parts.append(f'''
<h2 style="color: #1F4E79; margin-top: 40px;">&#8596; Same Tier ({len(same_tier)})</h2>
<p>These entities stayed in the same tier. Score changes still occurred due to methodology differences.</p>

<table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-size: 11pt;">
<tr style="background-color: #4472C4; color: white; font-weight: bold; text-align: center;">
<td style="text-align: left;">Entity</td><td>Tier</td><td>V20 Score</td><td>V23 Score</td><td>Delta</td><td>Top Dimension Changes</td>
</tr>
''')

for i, p in enumerate(same_tier):
    bg = ' style="background-color: #F2F2F2;"' if i % 2 == 1 else ''
    v20s = p['v20']['score'] if p['v20']['score'] is not None else '—'
    v23s = p['v23']['score'] if p['v23'] and p['v23']['score'] is not None else '—'
    delta = ''
    if p['score_delta'] is not None:
        delta = f'{p["score_delta"]:+d}' if p['score_delta'] != 0 else '0'
        delta_color = '#2E7D32' if p['score_delta'] > 0 else '#C62828' if p['score_delta'] < 0 else '#666'
        delta = f'<span style="color: {delta_color};">{delta}</span>'

    # Top 2 dimension changes
    top_dims = sorted(p['dim_changes'], key=lambda x: -abs(x['weighted_delta']))[:2]
    dim_str = ', '.join([f"{dc['dim']} {dc['v20']}→{dc['v23']}" for dc in top_dims]) if top_dims else '(no change)'

    html_parts.append(f'''<tr{bg}>
<td style="text-align: left; padding: 4px 10px;">{p['v23_name']}</td>
<td style="text-align: center;">{tier_badge(p['v23_tier'])}</td>
<td style="text-align: center;">{v20s}</td>
<td style="text-align: center;">{v23s}</td>
<td style="text-align: center; font-weight: bold;">{delta}</td>
<td style="padding: 4px 10px; font-size: 10pt;">{dim_str}</td>
</tr>''')

html_parts.append('</table>')

# --- ABSORBED ---
if absorbed:
    html_parts.append(f'''
<h2 style="color: #666; margin-top: 40px;">Absorbed ({len(absorbed)})</h2>
<table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-size: 11pt;">
<tr style="background-color: #4472C4; color: white; font-weight: bold;">
<td>V20 Entity</td><td>V20 Tier</td><td>Resolution</td>
</tr>
''')
    for p in absorbed:
        html_parts.append(f'''<tr>
<td>{p['v20_name']}</td><td style="text-align: center;">{p['v20']['tier']}</td>
<td>Entity no longer exists — facilities counted under parent entity in V23</td>
</tr>''')
    html_parts.append('</table>')

html_parts.append(f'''
<hr style="border: 1px solid #ccc; margin: 40px 0;">
<p style="font-size: 10pt; color: #999;">Generated {os.popen('date /t').read().strip()} | V20: Final_MUO_Tiering_V20.xlsx | V23: Final_MUO_Tiering_V23.xlsx + MUO_Scoring_Workbook_V23_v8.xlsx</p>
</body></html>''')

html_content = '\n'.join(html_parts)

with open(HTML_OUT, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"\nHTML report saved to: {HTML_OUT}")
print("Done!")
