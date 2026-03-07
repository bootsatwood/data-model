"""Build the Final MUO Tiering workbook — V20 universe rescored with V23 methodology.
Output format matches the V20 'Final MUO Tiering.xlsx' layout (T1/T2/T3/T4/T5 tabs).
Audience: Business Development team."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
V20_PATH = "C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx"
OUT_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V23.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15

# ============================================================
# T5 BARRIERS — union of V20 + V23 identified barriers
# ============================================================
T5_BARRIERS = {
    # V20 original
    'Bluegrass/Encore': {'db_names': {'BLUEGRASS/ENCORE', 'BLUEGRASS', 'ENCORE SENIOR LIVING'}, 'barrier': 'Alliance', 'v20_name': ['Bluegrass', 'Encore']},
    'SIGNATURE HEALTH': {'db_names': {'SIGNATURE HEALTHCARE', 'SIGNATURE HEALTH'}, 'barrier': 'Own provider group', 'v20_name': ['Signature']},
    'MFA': {'db_names': {'MFA', 'MFA MARYVILLE RE LLC'}, 'barrier': 'Alliance (VA)', 'v20_name': ['MFA']},
    'COMMUNICARE': {'db_names': {'COMMUNICARE'}, 'barrier': 'Own provider group', 'v20_name': ['Communicare']},
    'Hill Valley': {'db_names': {'HILL VALLEY'}, 'barrier': 'Own provider group', 'v20_name': ['Hill Valley']},
    'SINGH': {'db_names': {'SINGH'}, 'barrier': 'Own provider group', 'v20_name': ['Singh']},
    'Embassy': {'db_names': {'EMBASSY'}, 'barrier': 'Own provider group', 'v20_name': ['Embassy']},
    'Exceptional Living': {'db_names': {'EXCEPTIONAL LIVING CENTERS'}, 'barrier': 'Own provider group', 'v20_name': ['Exceptional Living']},
    'Portopiccolo': {'db_names': {'PORTOPICCOLO'}, 'barrier': 'Alliance (Telos)', 'v20_name': ['Portopiccolo']},
    'Venza': {'db_names': {'VENZA CARE MANAGEMENT'}, 'barrier': 'Alliance', 'v20_name': ['Venza']},
    'Aventura': {'db_names': {'AVENTURA'}, 'barrier': 'Recent LOB', 'v20_name': ['Aventura']},
    'Journey': {'db_names': {'JOURNEY'}, 'barrier': 'Explicit rejection', 'v20_name': ['Journey']},
    # V23 additions (identified from DB Barrier flag)
    'CARDON & ASSOCIATES': {'db_names': {'CARDON & ASSOCIATES'}, 'barrier': 'Own provider group (91%)', 'v20_name': ['Cardon']},
    'Eastern Healthcare Group': {'db_names': {'EASTERN HEALTHCARE GROUP'}, 'barrier': 'Termination Risk (100%)', 'v20_name': ['Eastern']},
    'CLEARVIEW': {'db_names': {'CLEARVIEW'}, 'barrier': 'Alliance + Own Provider Group (38%)', 'v20_name': ['Clearview']},
    'Pavilion Healthcare': {'db_names': {'PAVILION HEALTHCARE'}, 'barrier': 'Alliance + Own Provider Group (29%)', 'v20_name': ['Pavilion Healthcare']},
}

# All T5 DB names for filtering
T5_DB_NAMES = set()
for info in T5_BARRIERS.values():
    T5_DB_NAMES.update(info['db_names'])

# ============================================================
# ENTITY UNIVERSE — V20's 70 scored entities + their DB name mappings
# ============================================================
# Format: 'Display Name': {'db_names': set(), 'v20_name': str}
ENTITY_UNIVERSE = {
    # --- Already in Finance 60 (35 overlap) ---
    'ALG': {'ALG', 'ALG SENIOR'},
    'American Senior Communities': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'Saber Healthcare Group': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'Infinity Healthcare Consulting': {'INFINITY HEALTHCARE CONSULTING'},
    'Navion': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'Pruitt Health': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Trilogy': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Topaz Healthcare': {'TOPAZ HEALTHCARE'},
    'Morning Pointe Senior Living': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'Principle': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TerraBella Senior Living': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'Sanstone': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'Lionstone Care': {'LIONSTONE CARE', 'LIONSTONE'},
    'Eldercare Partners': {'ELDERCARE PARTNERS'},
    'Otterbein Senior Life': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'Peak Resources': {'PEAK RESOURCES'},
    'Arbors': {'ARBORS', 'ARBORS AT OHIO'},
    'American Healthcare LLC': {'AMERICAN HEALTHCARE, LLC'},
    'CCH Healthcare': {'CCH HEALTHCARE'},
    'Priority': {'PRIORITY LIFE CARE'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'Caring Place Healthcare': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'Fundamental LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Momentus Health': {'MOMENTUS HEALTH'},
    # --- 33 Gap entities (in V20 but not in Finance 60) ---
    'Altercare': {'ALTERCARE'},
    'AOM Healthcare': {'AOM HEALTHCARE'},
    'Aperion Care': {'APERION CARE'},
    'Brickyard Healthcare': {'BRICKYARD HEALTHCARE'},
    'Carecore Health': {'CARECORE HEALTH'},
    'Carespring': {'CARESPRING'},
    'Carrolton': {'CARROLTON FACILTY MANAGEMENT', 'CARROLTON NURSING HOMES'},
    'Certus Healthcare': {'CERTUS HEALTHCARE'},
    'Choice Health Management': {'CHOICE HEALTH MANAGEMENT'},
    'Ciena Healthcare/Laurel Health Care': {'CIENA HEALTHCARE/LAUREL HEALTH CARE'},
    'Continuing Healthcare Solutions': {'CONTINUING HEALTHCARE SOLUTIONS'},
    'Crown Healthcare Group': {'CROWN HEALTHCARE GROUP'},
    'Divine Healthcare Management': {'DIVINE HEALTHCARE MANAGEMENT'},
    'Envive Healthcare': {'ENVIVE HEALTHCARE'},
    'Gardant Management Solutions': {'GARDANT MANAGEMENT SOLUTIONS, INC'},
    'HCF Management': {'HCF MANAGEMENT'},
    'Health Care Management Group': {'HEALTH CARE MANAGEMENT GROUP'},
    'Hillstone Healthcare': {'HILLSTONE HEALTHCARE'},
    'Legacy Health Services': {'LEGACY HEALTH SERVICES'},
    "Miller's Merry Manor": {'MILLERS MERRY MANOR', 'MILLERS MERRY MANOR INDIANAPOLIS EAST LLC'},
    'National Healthcare Corporation': {'NATIONAL HEALTHCARE CORPORATION'},
    'Ohio Living Communities': {'OHIO LIVING COMMUNITIES'},
    'Optalis Health & Rehabilitation': {'OPTALIS HEALTH & REHABILITATION'},
    'PACS Group': {'PACS GROUP'},
    'Phoenix Senior Living': {'PHOENIX SENIOR LIVING'},
    'Progressive Quality Care': {'PROGRESSIVE QUALITY CARE'},
    'SEKY Holding Co.': {'SEKY HOLDING CO.'},
    'Sprenger Health Care Systems': {'SPRENGER HEALTH CARE SYSTEMS'},
    'Sunrise Senior Living': {'SUNRISE SENIOR LIVING', 'SUNRISE'},
    'Trio Healthcare': {'TRIO HEALTHCARE'},
    'Vancrest Health Care Centers': {'VANCREST HEALTH CARE CENTERS'},
    'White Oak Management': {'WHITE OAK MANAGEMENT'},
    'Windsor House': {'WINDSOR HOUSE, INC.'},
}

# V20 name -> our display name (for AI score lookup)
V20_TO_DISPLAY = {
    'ALG': 'ALG',
    'American Senior Communities': 'American Senior Communities',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'Saber Healthcare Group': 'Saber Healthcare Group',
    'Infinity Healthcare Consulting': 'Infinity Healthcare Consulting',
    'Navion': 'Navion',
    'Majestic Care': 'Majestic Care',
    'Pruitthealth': 'Pruitt Health',
    'Trilogy Health Services': 'Trilogy',
    'Topaz': 'Topaz Healthcare',
    'Morning Pointe Senior Living': 'Morning Pointe Senior Living',
    'Principle Long Term Care': 'Principle',
    'Liberty Senior Living': 'Liberty',
    'Terra Bella': 'TerraBella Senior Living',
    'Sanstone Health & Rehabilitation': 'Sanstone',
    'Lionstone Care': 'Lionstone Care',
    'Otterbein Seniorlife': 'Otterbein Senior Life',
    'Tlc Management': 'TLC Management',
    'BHI Senior Living': 'BHI Senior Living',
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': 'Avardis',
    'Peak Resources, Inc.': 'Peak Resources',
    'Arbors At Ohio': 'Arbors',
    'Heritage Hall': 'American Healthcare LLC',
    'Cch Healthcare': 'CCH Healthcare',
    'Priority Life Care': 'Priority',
    'Life Care Centers Of America': 'Lifecare',
    'Kissito': 'Kissito Healthcare',
    'Yad Healthcare': 'YAD',
    'Caring Place Healthcare': 'Caring Place Healthcare',
    'Castle Healthcare': 'Castle Healthcare',
    'Jag Healthcare': 'JAG',
    'Momentous Health': 'Momentus Health',
    # Gap entities (V20 name = display name or close)
    'Altercare': 'Altercare',
    'Aom Healthcare': 'AOM Healthcare',
    'Aperion Care': 'Aperion Care',
    'Brickyard Healthcare': 'Brickyard Healthcare',
    'Carecore Health': 'Carecore Health',
    'Carespring': 'Carespring',
    'Carrolton': 'Carrolton',
    'Certus Healthcare': 'Certus Healthcare',
    'Choice Health Management': 'Choice Health Management',
    'Ciena Healthcare/Laurel Health Care': 'Ciena Healthcare/Laurel Health Care',
    'Continuing Healthcare Solutions': 'Continuing Healthcare Solutions',
    'Crown Healthcare Group': 'Crown Healthcare Group',
    'Divine Healthcare Management': 'Divine Healthcare Management',
    'Envive Healthcare': 'Envive Healthcare',
    'Gardant Management Solutions, Inc': 'Gardant Management Solutions',
    'HCF Management': 'HCF Management',
    'Health Care Management Group': 'Health Care Management Group',
    'Hillstone Healthcare': 'Hillstone Healthcare',
    'Legacy Health Services': 'Legacy Health Services',
    "Miller's Merry Manor": "Miller's Merry Manor",
    'National Healthcare Corporation': 'National Healthcare Corporation',
    'Ohio Living Communities': 'Ohio Living Communities',
    'Optalis Health & Rehabilitation': 'Optalis Health & Rehabilitation',
    'PACS Group': 'PACS Group',
    'Phoenix Senior Living': 'Phoenix Senior Living',
    'Progressive Quality Care': 'Progressive Quality Care',
    'Seky Holding Co.': 'SEKY Holding Co.',
    'Sprenger Health Care Systems': 'Sprenger Health Care Systems',
    'Sunrise': 'Sunrise Senior Living',
    'Trio Healthcare': 'Trio Healthcare',
    'Vancrest Health Care Centers': 'Vancrest Health Care Centers',
    'White Oak Management': 'White Oak Management',
    'Windsor House, Inc.': 'Windsor House',
    # These were in V20 scored but are now T5 in V23
    'Clearview': 'CLEARVIEW',
    'Pavilion Healthcare': 'Pavilion Healthcare',
    'Southern Assisted Living, LLC': None,  # absorbed by Brookdale
    'Sovereign Healthcare Holdings': None,  # = Southern Healthcare Mgmt
}

# ============================================================
# SCORING PARAMETERS (same as V8)
# ============================================================
ER_THRESHOLDS = [9, 13, 20, 40]
RP_THRESHOLDS = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def normalize_address(addr):
    s = addr.upper().strip().rstrip('.')
    for full, abbr in [('STREET', 'ST'), ('DRIVE', 'DR'), ('ROAD', 'RD'), ('AVENUE', 'AVE'),
                       ('BOULEVARD', 'BLVD'), ('LANE', 'LN'), ('COURT', 'CT'), ('CIRCLE', 'CIR'),
                       ('PLACE', 'PL'), ('TERRACE', 'TER'), ('PARKWAY', 'PKWY'), ('HIGHWAY', 'HWY'),
                       ('NORTH', 'N'), ('SOUTH', 'S'), ('EAST', 'E'), ('WEST', 'W')]:
        s = re.sub(r'\b' + full + r'\b', abbr, s)
    s = re.sub(r'\.', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def extract_street_number(addr):
    m = re.match(r'^(\d+)', addr.strip())
    return m.group(1) if m else None

# RS scores — from Tom's Notes (Finance 60 entities only), default 1 for gap entities
TOM_RS_SCORES = {
    'ALG': 2, 'American Senior Communities': 5, 'Brookdale Senior Living': 5,
    'Saber Healthcare Group': 2, 'Infinity Healthcare Consulting': 2,
    'Navion': 3, 'Majestic Care': 3, 'Pruitt Health': 3,
    'Trilogy': 4, 'Topaz Healthcare': 1,
    'Morning Pointe Senior Living': 2, 'Principle': 3, 'Liberty': 4,
    'TerraBella Senior Living': 2, 'Sanstone': 2, 'Lionstone Care': 3,
    'Eldercare Partners': 3, 'Otterbein Senior Life': 1, 'TLC Management': 3,
    'BHI Senior Living': 1, 'Avardis': 2, 'Peak Resources': 2,
    'Arbors': 1, 'American Healthcare LLC': 3, 'CCH Healthcare': 2,
    'Priority': 1, 'Lifecare': 1, 'Kissito Healthcare': 2,
    'YAD': 2, 'Caring Place Healthcare': 2, 'Castle Healthcare': 3,
    'JAG': 2, 'Fundamental LTC': 1, 'Southern Healthcare Mgmt': 1,
    'Momentus Health': 1,
}

# SI scores — from Tom's Notes + web research (Finance 60 entities), default 2 for gap entities
DATA_SI_SCORES = {
    'ALG': 5, 'Brookdale Senior Living': 5, 'Saber Healthcare Group': 5,
    'Trilogy': 5, 'Liberty': 5, 'Pruitt Health': 5, 'Avardis': 5,
    'Lifecare': 5,
    'American Senior Communities': 4, 'Infinity Healthcare Consulting': 4,
    'Majestic Care': 4, 'Navion': 4, 'TLC Management': 4,
    'Sanstone': 3, 'Morning Pointe Senior Living': 3, 'Otterbein Senior Life': 3,
    'Principle': 3, 'Lionstone Care': 3, 'Kissito Healthcare': 3,
    'CCH Healthcare': 3, 'TerraBella Senior Living': 3, 'Peak Resources': 3,
    'BHI Senior Living': 3,
    # Gap entities with known SI from V20 or obvious research
    'Ciena Healthcare/Laurel Health Care': 3,  # Multi-state, 85 total facilities
    'National Healthcare Corporation': 3,  # Multi-state, 82 total facilities
    'PACS Group': 3,  # NYSE: PACS, large operator
    'Sunrise Senior Living': 4,  # Major national operator, 100+ facilities
}

# ============================================================
# LOAD DATA
# ============================================================
print("Loading V23 DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
db_headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

print("Loading V20 Tiering workbook (for AI scores)...")
wb_v20 = openpyxl.load_workbook(V20_PATH, data_only=True)
v20_scores = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws_t = wb_v20[sn]
    for r in range(2, ws_t.max_row + 1):
        name = ws_t.cell(r, 1).value
        if not name:
            continue
        v20_scores[name.strip()] = {
            'ai': ws_t.cell(r, 14).value or 0,
        }

# Build reverse lookup: DB corp name -> display name
all_db_names = set()
db_to_display = {}
for display_name, db_names in ENTITY_UNIVERSE.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_display[dn] = display_name

# ============================================================
# COLLECT FACILITY ROWS
# ============================================================
print("Scanning DB for facility rows...")
entity_facilities = defaultdict(list)

for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, db_headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, db_headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue
    # Skip T5 entities
    if corp in T5_DB_NAMES:
        continue

    display_name = db_to_display[corp]
    fac_name = str(ws_db.cell(r, db_headers['Facility_Name']).value or '').strip()
    addr = str(ws_db.cell(r, db_headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, db_headers['City']).value or '').strip()
    served = str(ws_db.cell(r, db_headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws_db.cell(r, db_headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws_db.cell(r, db_headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws_db.cell(r, db_headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    tot_beds = ws_db.cell(r, db_headers['Total_Beds']).value or 0
    tot_beds = tot_beds if isinstance(tot_beds, (int, float)) else 0
    tot_rev = ws_db.cell(r, db_headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0
    cur_rev = ws_db.cell(r, db_headers['Current_Revenue']).value or 0
    cur_rev = cur_rev if isinstance(cur_rev, (int, float)) else 0

    norm_addr = normalize_address(addr)
    street_num = extract_street_number(addr)
    num_key = (street_num, city.upper()) if street_num else (norm_addr, city.upper())

    entity_facilities[display_name].append({
        'fac_name': fac_name, 'addr': addr, 'city': city, 'state': state,
        'campus_key': num_key, 'served': served, 'mh': mh, 'pcp': pcp, 'integ': integ,
        'tot_beds': tot_beds, 'tot_rev': tot_rev, 'cur_rev': cur_rev,
    })

# ============================================================
# COMPUTE ENTITY ROLLUPS
# ============================================================
print("Computing entity rollups...")
scored_entities = {}
t4_entities = {}

for display_name in sorted(ENTITY_UNIVERSE.keys()):
    facs = entity_facilities.get(display_name, [])
    if not facs:
        t4_entities[display_name] = {
            'n_camp': 0, 'n_srv': 0, 'total_rev': 0, 'cur_rev': 0,
            'states': '(no DB rows in FP)', 'facs': [],
        }
        continue

    # Apply 15-bed minimum
    qualifying = [f for f in facs if f['tot_beds'] > BED_MIN]
    campuses = set(f['campus_key'] for f in qualifying)
    served_camps = set(f['campus_key'] for f in qualifying if f['served'])
    mh_camps = set(f['campus_key'] for f in qualifying if f['mh'])
    pcp_camps = set(f['campus_key'] for f in qualifying if f['pcp'])
    integ_camps = set(f['campus_key'] for f in qualifying if f['integ'])
    total_rev = sum(f['tot_rev'] for f in qualifying)
    cur_rev_total = sum(f['cur_rev'] for f in qualifying)

    n_camp = len(campuses)
    n_srv = len(served_camps)
    states = sorted(set(f['state'] for f in qualifying))

    if n_camp < 7:
        t4_entities[display_name] = {
            'n_camp': n_camp, 'n_srv': n_srv, 'total_rev': total_rev, 'cur_rev': cur_rev_total,
            'states': ', '.join(states) if states else '(no qualifying rows)',
            'facs': facs,
        }
        continue

    # ER
    er = score_dim(n_camp, ER_THRESHOLDS)

    # RP
    rp = score_dim(total_rev, RP_THRESHOLDS)

    # RS
    rs = TOM_RS_SCORES.get(display_name, 1)

    # SI
    si = DATA_SI_SCORES.get(display_name, 2)

    # AI — from V20 where available
    ai = 0
    for v20_name, disp in V20_TO_DISPLAY.items():
        if disp == display_name and v20_name in v20_scores:
            ai = v20_scores[v20_name]['ai']
            break

    # IR — formulaic from service flags
    n_integ = len(integ_camps)
    int_pct = n_integ / n_camp if n_camp > 0 else 0
    dual = len(mh_camps & pcp_camps)

    if n_srv == 0:
        ir = 1
    elif int_pct >= 0.5:
        ir = 5
    elif int_pct >= 0.25:
        ir = 4
    elif dual > 0:
        ir = 4
    elif len(mh_camps) > 0 and len(pcp_camps) > 0:
        ir = 3
    elif len(mh_camps) > 0 or len(pcp_camps) > 0:
        ir = 2
    else:
        ir = 1

    total_score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = 'T1' if total_score >= 55 else 'T2' if total_score >= 35 else 'T3'

    scored_entities[display_name] = {
        'n_camp': n_camp, 'n_srv': n_srv, 'n_states': len(states),
        'states': ', '.join(states), 'total_rev': total_rev, 'cur_rev': cur_rev_total,
        'er': er, 'ir': ir, 'si': si, 'rp': rp, 'rs': rs, 'ai': ai,
        'total_score': total_score, 'tier': tier,
        'total_facs': len(facs),
    }

# ============================================================
# BUILD WORKBOOK (V20 format)
# ============================================================
print("Building workbook...")
wb_out = openpyxl.Workbook()

# Styles
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
DATA_FONT = Font(size=10)
T1_FILL = PatternFill('solid', fgColor='C6EFCE')
T2_FILL = PatternFill('solid', fgColor='FFEB9C')
T3_FILL = PatternFill('solid', fgColor='FFC7CE')
T4_FILL = PatternFill('solid', fgColor='D6DCE4')
T5_FILL = PatternFill('solid', fgColor='F2DCDB')
MONEY_FMT = '$#,##0'
MONEY_M_FMT = '$#,##0.0,,"M"'

def write_scored_tab(wb, tab_name, entities, tier_fill):
    """Write a T1/T2/T3 tab in V20 format."""
    ws = wb.create_sheet(tab_name) if tab_name != wb.sheetnames[0] else wb.active
    if tab_name == wb.sheetnames[0]:
        ws.title = tab_name

    headers = ['MUO Name', 'BD Tier', 'Total Score', 'States', '# States',
               'Campuses', 'We Serve', 'Total Opportunity',
               'Enterprise Reach', 'Integration Ready', 'Strategic Influence',
               'Revenue Potential', 'Relationship Strength', 'AI/Tech Adoption']
    ws.append(headers)

    # Style header
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Column widths (match V20)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 16
    for letter in 'IJKLMN':
        ws.column_dimensions[letter].width = 13

    # Sort by total score descending
    sorted_names = sorted(entities, key=lambda x: (-entities[x]['total_score'], x))

    row_num = 2
    for name in sorted_names:
        e = entities[name]
        rev_display = e['total_rev']

        ws.cell(row_num, 1, name)
        ws.cell(row_num, 2, e['tier'])
        ws.cell(row_num, 3, e['total_score'])
        ws.cell(row_num, 4, e['states'])
        ws.cell(row_num, 5, e['n_states'])
        ws.cell(row_num, 6, e['n_camp'])
        ws.cell(row_num, 7, e['n_srv'])
        c = ws.cell(row_num, 8, rev_display); c.number_format = MONEY_FMT
        ws.cell(row_num, 9, e['er'])
        ws.cell(row_num, 10, e['ir'])
        ws.cell(row_num, 11, e['si'])
        ws.cell(row_num, 12, e['rp'])
        ws.cell(row_num, 13, e['rs'])
        ws.cell(row_num, 14, e['ai'])

        for c in range(1, len(headers) + 1):
            ws.cell(row_num, c).font = DATA_FONT
            ws.cell(row_num, c).alignment = Alignment(horizontal='center')
        ws.cell(row_num, 1).alignment = Alignment(horizontal='left')
        ws.cell(row_num, 4).alignment = Alignment(horizontal='left')

        row_num += 1

    return ws

# Split scored entities by tier
t1 = {k: v for k, v in scored_entities.items() if v['tier'] == 'T1'}
t2 = {k: v for k, v in scored_entities.items() if v['tier'] == 'T2'}
t3 = {k: v for k, v in scored_entities.items() if v['tier'] == 'T3'}

# T1 tab (use default first sheet)
ws = wb_out.active
ws.title = 'T1 MUO'
write_scored_tab.__code__  # just to define it above
# Actually write the tabs
for ws_ref in list(wb_out.sheetnames):
    if ws_ref != 'T1 MUO':
        del wb_out[ws_ref]

# Rewrite T1 on active sheet
ws1 = wb_out.active
headers = ['MUO Name', 'BD Tier', 'Total Score', 'States', '# States',
           'Campuses', 'We Serve', 'Total Opportunity',
           'Enterprise Reach', 'Integration Ready', 'Strategic Influence',
           'Revenue Potential', 'Relationship Strength', 'AI/Tech Adoption']
ws1.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws1.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 25
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 11
ws1.column_dimensions['G'].width = 11
ws1.column_dimensions['H'].width = 16
for letter in 'IJKLMN':
    ws1.column_dimensions[letter].width = 13

def write_entities_to_sheet(ws, entities_dict, start_row=2):
    sorted_names = sorted(entities_dict, key=lambda x: (-entities_dict[x]['total_score'], x))
    row_num = start_row
    for name in sorted_names:
        e = entities_dict[name]
        ws.cell(row_num, 1, name)
        ws.cell(row_num, 2, e['tier'])
        ws.cell(row_num, 3, e['total_score'])
        ws.cell(row_num, 4, e['states'])
        ws.cell(row_num, 5, e['n_states'])
        ws.cell(row_num, 6, e['n_camp'])
        ws.cell(row_num, 7, e['n_srv'])
        c = ws.cell(row_num, 8, e['total_rev']); c.number_format = MONEY_FMT
        ws.cell(row_num, 9, e['er'])
        ws.cell(row_num, 10, e['ir'])
        ws.cell(row_num, 11, e['si'])
        ws.cell(row_num, 12, e['rp'])
        ws.cell(row_num, 13, e['rs'])
        ws.cell(row_num, 14, e['ai'])
        for c_idx in range(1, 15):
            ws.cell(row_num, c_idx).font = DATA_FONT
            ws.cell(row_num, c_idx).alignment = Alignment(horizontal='center')
        ws.cell(row_num, 1).alignment = Alignment(horizontal='left')
        ws.cell(row_num, 4).alignment = Alignment(horizontal='left')
        row_num += 1
    return row_num

write_entities_to_sheet(ws1, t1)

# T2 tab
ws2 = wb_out.create_sheet('T2 MUO')
ws2.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws2.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 25
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 11
ws2.column_dimensions['G'].width = 11
ws2.column_dimensions['H'].width = 16
for letter in 'IJKLMN':
    ws2.column_dimensions[letter].width = 13
write_entities_to_sheet(ws2, t2)

# T3 tab
ws3 = wb_out.create_sheet('T3 MUO')
ws3.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws3.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 11
ws3.column_dimensions['G'].width = 11
ws3.column_dimensions['H'].width = 16
for letter in 'IJKLMN':
    ws3.column_dimensions[letter].width = 13
write_entities_to_sheet(ws3, t3)

# ============================================================
# T4 — Independents
# ============================================================
ws4 = wb_out.create_sheet('T4 - Independents')
t4_headers = ['MUO Name', 'Qualifying Campuses', 'We Serve', 'States', 'Total Opportunity', 'Reason']
ws4.append(t4_headers)
for c in range(1, len(t4_headers) + 1):
    cell = ws4.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 11
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 40

sorted_t4 = sorted(t4_entities.keys(), key=lambda x: (-t4_entities[x]['n_camp'], x))
row_num = 2
for name in sorted_t4:
    t4d = t4_entities[name]
    reason = f"{t4d['n_camp']} qualifying campuses (gate requires 7)" if t4d['n_camp'] > 0 else 'No DB rows in footprint'
    ws4.cell(row_num, 1, name)
    ws4.cell(row_num, 2, t4d['n_camp'])
    ws4.cell(row_num, 3, t4d['n_srv'])
    ws4.cell(row_num, 4, t4d['states'])
    c = ws4.cell(row_num, 5, t4d['total_rev']); c.number_format = MONEY_FMT
    ws4.cell(row_num, 6, reason)
    for c_idx in range(1, len(t4_headers) + 1):
        ws4.cell(row_num, c_idx).font = DATA_FONT
        ws4.cell(row_num, c_idx).fill = T4_FILL
    row_num += 1

# ============================================================
# T5 — Hard Barriers
# ============================================================
ws5 = wb_out.create_sheet('T5 - Hard Barriers')
t5_headers = ['Corporate Name', 'Barrier Type', 'Facilities', 'We Serve', 'States']
ws5.append(t5_headers)
for c in range(1, len(t5_headers) + 1):
    cell = ws5.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 11
ws5.column_dimensions['D'].width = 11
ws5.column_dimensions['E'].width = 20

# Collect T5 facility counts from DB
t5_data = {}
for display_name, info in T5_BARRIERS.items():
    fac_count = 0
    served_count = 0
    states = set()
    for r in range(2, ws_db.max_row + 1):
        corp = str(ws_db.cell(r, db_headers['Corporate_Name']).value or '').strip().upper()
        state = str(ws_db.cell(r, db_headers['State']).value or '').strip()
        if corp in info['db_names'] and state in FOOTPRINT:
            fac_count += 1
            states.add(state)
            served = str(ws_db.cell(r, db_headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
            if served:
                served_count += 1
    t5_data[display_name] = {
        'barrier': info['barrier'], 'facs': fac_count, 'served': served_count,
        'states': ', '.join(sorted(states)) if states else '',
    }

row_num = 2
for name in sorted(T5_BARRIERS.keys()):
    d = t5_data[name]
    ws5.cell(row_num, 1, name)
    ws5.cell(row_num, 2, d['barrier'])
    ws5.cell(row_num, 3, d['facs'])
    ws5.cell(row_num, 4, d['served'])
    ws5.cell(row_num, 5, d['states'])
    for c_idx in range(1, len(t5_headers) + 1):
        ws5.cell(row_num, c_idx).font = DATA_FONT
        ws5.cell(row_num, c_idx).fill = T5_FILL
    row_num += 1

# ============================================================
# SAVE
# ============================================================
print(f"\nSaving to {OUT_PATH}...")
wb_out.save(OUT_PATH)

# Summary
from collections import Counter
print(f"\n{'='*80}")
print(f"FINAL MUO TIERING V23 — SUMMARY")
print(f"{'='*80}")
print(f"  T1: {len(t1)} entities")
for name in sorted(t1, key=lambda x: -t1[x]['total_score']):
    print(f"    {name:<45} {t1[name]['total_score']:>3}")
print(f"  T2: {len(t2)} entities")
for name in sorted(t2, key=lambda x: -t2[x]['total_score']):
    print(f"    {name:<45} {t2[name]['total_score']:>3}")
print(f"  T3: {len(t3)} entities")
for name in sorted(t3, key=lambda x: -t3[x]['total_score']):
    print(f"    {name:<45} {t3[name]['total_score']:>3}")
print(f"  T4: {len(t4_entities)} entities")
for name in sorted(t4_entities, key=lambda x: -t4_entities[x]['n_camp']):
    print(f"    {name:<45} {t4_entities[name]['n_camp']} campuses")
print(f"  T5: {len(T5_BARRIERS)} entities")
print(f"\n  Total: {len(t1) + len(t2) + len(t3) + len(t4_entities) + len(T5_BARRIERS)}")
print("Done!")
