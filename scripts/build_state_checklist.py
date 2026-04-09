"""Build New State Entry Checklist workbook — one tab per expansion state."""
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

conn = psycopg2.connect(
    host='keystone-platform-postgres.postgres.database.azure.com',
    port=5432, database='postgres', user='ratwood',
    password='Ch0EeU3yz7ep5bEWTKT1UoC^NMPCZXzN',
    sslmode='require'
)
cur = conn.cursor()

cur.execute("""
SELECT abbr, name, strategy, market_class, total_som, total_facilities,
       practice_act_status, physician_requirement, is_con,
       muo_revenue, muo_subtitle, muo_note,
       bluestone_present, alliance_level, alliance_count, alliance_note,
       team_hospitals, team_note
FROM bd.expansion_states WHERE strategy != 'Footprint'
ORDER BY CASE strategy WHEN 'S1' THEN 1 WHEN 'S2' THEN 2 WHEN 'Outlier' THEN 3 WHEN 'Watch' THEN 4 END, name
""")
states = cur.fetchall()

cur.execute("SELECT state_abbr, metro, facilities, s2_revenue, muos FROM bd.expansion_state_metros ORDER BY state_abbr, facilities DESC")
metros_all = {}
for r in cur.fetchall():
    metros_all.setdefault(r[0], []).append(r)

cur.execute("SELECT state_abbr, chain_name, facility_count FROM bd.expansion_state_alliance_chains ORDER BY state_abbr, facility_count DESC")
chains_all = {}
for r in cur.fetchall():
    chains_all.setdefault(r[0], []).append(r)
conn.close()

# --- Styles ---
navy = "193241"
purple = "877BE9"
light_blue = "91D9E8"
pewter = "507E8E"
grey_hex = "E1E2E0"

title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
subtitle_font = Font(name="Calibri", size=11, color=light_blue)
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
section_font = Font(name="Calibri", size=12, bold=True, color=navy)
item_font = Font(name="Calibri", size=11, color="333333")
value_font = Font(name="Calibri", size=11, color="333333")
total_font = Font(name="Calibri", size=11, bold=True, color=navy)
legend_font = Font(name="Calibri", size=9, color="666666", italic=True)
detail_font = Font(name="Calibri", size=10, color=pewter, italic=True)

navy_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
section_fill = PatternFill(start_color="F0EEF8", end_color="F0EEF8", fill_type="solid")
total_fill = PatternFill(start_color=grey_hex, end_color=grey_hex, fill_type="solid")
stat_fill = PatternFill(start_color="F8F8FA", end_color="F8F8FA", fill_type="solid")
pewter_fill = PatternFill(start_color=pewter, end_color=pewter, fill_type="solid")
thin_border = Border(bottom=Side(style="thin", color=grey_hex))
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

NL = "\n"

wb = Workbook()
wb.remove(wb.active)

for s in states:
    abbr, name, strategy, mclass = s[0], s[1], s[2], s[3]
    som, fac_count = s[4], s[5]
    pract_act, phys_req, is_con = s[6], s[7], s[8]
    muo_rev, muo_detail, muo_note = s[9], s[10], s[11]
    bluestone = s[12]
    alliance_lvl, alliance_cnt, alliance_note = s[13], s[14], s[15]
    team_hosp, team_note = s[16], s[17]
    metros = metros_all.get(abbr, [])
    chains = chains_all.get(abbr, [])

    tab = f"{abbr} - {name}"[:31]
    ws = wb.create_sheet(title=tab)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 22

    # Title rows
    for r in (1, 2):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    c1 = ws["A1"]
    c1.value = f"New State Entry Checklist — {name} ({abbr})"
    c1.font = title_font
    c1.alignment = Alignment(vertical="center", indent=1)
    c2 = ws["A2"]
    c2.value = f"{mclass}  |  Strategy: {strategy}  |  {datetime.now().strftime('%B %d, %Y')}"
    c2.font = subtitle_font
    c2.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22

    # Stats bar row 3
    ws.row_dimensions[3].height = 40
    stat_data = [
        (f"SOM: {som}", "Total Market"),
        (str(fac_count), "Facilities"),
        (str(muo_rev), "MUO Revenue"),
        str(pract_act),
        ("Yes" if is_con else "No"),
    ]
    stat_labels = ["Total Market", "Facilities", "MUO Revenue", "Practice Act", "CON State"]
    for i in range(5):
        val = stat_data[i] if isinstance(stat_data[i], str) else stat_data[i][0]
        label = stat_labels[i]
        if isinstance(stat_data[i], tuple):
            val = stat_data[i][0]
        c = ws.cell(row=3, column=i + 2)
        c.value = f"{val}{NL}{label}"
        c.font = Font(name="Calibri", size=10, bold=True, color=purple)
        c.fill = stat_fill
        c.alignment = center_wrap

    # Column headers row 5
    for col in range(2, 7):
        ws.cell(row=5, column=col).fill = pewter_fill
        ws.cell(row=5, column=col).font = header_font
        ws.cell(row=5, column=col).alignment = center
    ws["B5"] = "Key Process Step"
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["D5"] = "Timeline"
    ws["F5"] = "Incremental Cost"

    # Helpers
    def write_item(r, text, timeline="", cost="", height=18):
        ws.cell(row=r, column=2, value=text).font = item_font
        ws.cell(row=r, column=2).alignment = wrap
        if timeline:
            ws.cell(row=r, column=4, value=timeline).font = value_font
            ws.cell(row=r, column=4).alignment = center
        if cost:
            ws.cell(row=r, column=6, value=cost).font = value_font
            ws.cell(row=r, column=6).alignment = center
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = thin_border
        ws.row_dimensions[r].height = height

    def write_section(r, text):
        ws.cell(row=r, column=2, value=text).font = section_font
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = section_fill

    # ===== LEGAL / REGULATORY =====
    write_section(7, "Legal / Regulatory")

    write_item(9, "Establish Business License", "[2-3 months]", "Low")
    write_item(10, "Set up TIN", "[2-3 months]", "Low")

    muo_line = "Sign contract with existing MUO partners"
    if muo_detail:
        muo_line += f"{NL}({muo_detail})"
    write_item(11, muo_line, "[2-3 months]", "Low", 32 if muo_detail else 18)

    cred_line = "Credentialing / payor requirements"
    if phys_req:
        cred_line += f"{NL}{phys_req}"
    if pract_act == "Full":
        cred_time, cred_cost = "[1-2 months]", "Low"
    elif pract_act == "Restricted":
        cred_time, cred_cost = "[3-4 months]", "Medium"
    else:
        cred_time, cred_cost = "[2-3 months]", "Medium"
    write_item(12, cred_line, cred_time, cred_cost, 36 if phys_req else 18)

    con_line = "ACO / CON requirements"
    if is_con:
        con_line += f"{NL}CON state — certificate of need applies"
    else:
        con_line += f"{NL}Non-CON state — no certificate of need barrier"
    write_item(13, con_line, "", "", 32)

    # ===== CLINICAL =====
    write_section(15, "Clinical")

    if pract_act == "Full":
        cl_text = f"Hire clinicians{NL}Full practice authority — NPs practice independently, no physician oversight required"
        cl_time, cl_cost = "[2-3 months]", "Medium"
    elif pract_act == "Reduced":
        cl_text = f"Hire clinicians{NL}Reduced practice — collaborative agreement required"
        if phys_req:
            cl_text += f". {phys_req}"
        cl_time, cl_cost = "[3-4 months]", "Medium"
    else:
        cl_text = f"Hire clinicians{NL}Restricted practice — physician supervision required"
        if phys_req:
            cl_text += f". {phys_req}"
        cl_time, cl_cost = "[4-5 months]", "High"
    write_item(17, cl_text, cl_time, cl_cost, 48)

    # ===== OPERATIONAL =====
    write_section(19, "Operational")

    ops_line = "Determine regional oversight"
    if metros:
        metro_parts = [f"{m[1]} ({m[2]} fac, {m[3]})" for m in metros[:5]]
        ops_line += f"{NL}Key metros: {'; '.join(metro_parts)}"
    write_item(21, ops_line, "[1-2 months]", "Low", 48 if metros else 18)

    write_item(22, "Designate a rep to territory OR", "[1-2 months]", "Low")
    write_item(23, "Hire a new rep", "[4-5 months]", "High")
    write_item(24, "EHR needs?", "", "Low")

    # ===== TOTAL =====
    if pract_act == "Full":
        gl_time, gl_cost = "[3-4 months]", "[$50-75k]"
    elif pract_act == "Reduced":
        gl_time, gl_cost = "[4-5 months]", "[$75-100k]"
    else:
        gl_time, gl_cost = "[5-6 months]", "[$100-150k]"

    ws.cell(row=26, column=2, value='Total Estimated To "Go Live"').font = total_font
    ws.cell(row=26, column=4, value=gl_time).font = total_font
    ws.cell(row=26, column=4).alignment = center
    ws.cell(row=26, column=6, value=gl_cost).font = total_font
    ws.cell(row=26, column=6).alignment = center
    for c in range(2, 7):
        ws.cell(row=26, column=c).fill = total_fill

    ws.cell(row=28, column=2, value="Cost Legend: Low = <$25k; Medium = $25-100k; High = >$100k").font = legend_font

    # ===== COMPETITIVE LANDSCAPE (if relevant) =====
    if chains or (alliance_note and alliance_lvl in ("MODERATE", "HIGH")):
        write_section(30, "Competitive Landscape")
        r = 32
        if chains:
            chain_str = "; ".join([f"{c[1]} ({c[2]} fac)" for c in chains])
            ws.cell(row=r, column=2, value=f"Alliance chains: {chain_str}").font = detail_font
            ws.cell(row=r, column=2).alignment = wrap
            ws.row_dimensions[r].height = 28
            r += 1
        if alliance_note:
            ws.cell(row=r, column=2, value=alliance_note).font = detail_font
            ws.cell(row=r, column=2).alignment = wrap
            ws.row_dimensions[r].height = 36

out = "C:/Users/ratwood/Downloads/Eventus_New_State_Entry_Checklist_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Tabs: {wb.sheetnames}")
