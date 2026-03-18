"""Build a single-page HTML BD view of the complete 100-entity MUO universe.
Stacks all 5 tiers on one page. Pulls scored data from Finance workbook where available,
falls back to BD workbook for V20-gap entities."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

# ============================================================
# PATHS
# ============================================================
DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
V20_PATH = "C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx"
V23_FIN_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_MUO_Scoring_Workbook_V23_v8.xlsx"
V23_BD_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/2026-03-07_Final_MUO_Tiering_V23.xlsx"
HTML_OUT = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/BD_Complete_MUO_Universe_V23.html"

BREAKOUT_OUT = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/BD_MUO_Breakouts.html"

FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15
ER_THRESHOLDS = [9, 13, 20, 40]
RP_THRESHOLDS = [1_000_000, 2_500_000, 5_000_000, 10_000_000]

def slugify(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

# Bracket labels for breakout detail
ER_BRACKETS = {5: '40+ campuses', 4: '20-39', 3: '13-19', 2: '9-12', 1: '<9'}
IR_BRACKETS = {5: 'Majority integrated (>=50%)', 4: 'Dual MH+PCP same campus, or 25-49%', 3: 'MH + PCP at separate campuses', 2: 'Single service only (MH or PCP)', 1: 'Not served'}
SI_BRACKETS = {5: 'REIT/ISNP/Enterprise', 4: 'Payer/ACO connected', 3: 'Regional reputation, expanding', 2: 'Minimal local presence', 1: 'No affiliations found'}
RP_BRACKETS = {5: '>$10M', 4: '$5M-$10M', 3: '$2.5M-$5M', 2: '$1M-$2.5M', 1: '<$1M'}
RS_BRACKETS = {5: 'Active SS on Eventus paper', 4: 'Active engagement, verbal commits', 3: 'Have relationship, stalled', 2: 'Tried, hit resistance', 1: 'No relationship'}
AI_BRACKETS = {5: 'Advanced AI/analytics adoption', 4: 'Active tech investment', 3: 'Moderate EHR/data capability', 2: 'Basic digital infrastructure', 1: 'Minimal tech presence', 0: 'Not scored'}

def extract_street_number(addr):
    m = re.match(r'^(\d+)', addr.strip())
    return m.group(1) if m else None

def normalize_address(addr):
    s = addr.upper().strip().rstrip('.')
    for full, abbr in [('STREET', 'ST'), ('DRIVE', 'DR'), ('ROAD', 'RD'), ('AVENUE', 'AVE'),
                       ('BOULEVARD', 'BLVD'), ('LANE', 'LN'), ('COURT', 'CT'), ('CIRCLE', 'CIR'),
                       ('PLACE', 'PL'), ('TERRACE', 'TER'), ('PARKWAY', 'PKWY'), ('HIGHWAY', 'HWY'),
                       ('NORTH', 'N'), ('SOUTH', 'S'), ('EAST', 'E'), ('WEST', 'W')]:
        s = re.sub(r'\b' + full + r'\b', abbr, s)
    s = re.sub(r'\.', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ============================================================
# CANONICAL NAME RESOLUTION
# ============================================================
ALIASES = {
    'consulate health care/independence living centers/nspire healthcare/raydiant health care': 'Avardis',
    'liberty senior living': 'Liberty', 'life care centers of america': 'Lifecare',
    'principle long term care': 'Principle', 'pruitthealth': 'Pruitt Health',
    'terra bella': 'TerraBella Senior Living', 'tlc management': 'TLC Management',
    'trilogy health services': 'Trilogy', 'arbors at ohio': 'Arbors',
    'heritage hall': 'American Healthcare LLC', 'jag healthcare': 'JAG',
    'kissito': 'Kissito Healthcare', 'momentous health': 'Momentus Health',
    'peak resources, inc.': 'Peak Resources', 'priority life care': 'Priority',
    'sanstone health & rehabilitation': 'Sanstone', 'topaz': 'Topaz Healthcare',
    'yad healthcare': 'YAD', 'cch healthcare': 'CCH Healthcare',
    'otterbein seniorlife': 'Otterbein Senior Life',
    'lutheran services carolina': 'Lutheran Services Carolinas',
    'gardant management solutions, inc': 'Gardant Management Solutions',
    'seky holding co.': 'SEKY Holding Co.', 'sunrise': 'Sunrise Senior Living',
    'windsor house, inc.': 'Windsor House', 'aom healthcare': 'AOM Healthcare',
    'fundamental ltc': 'Fundamental LTC',
    'southern healthcare management, llc': 'Southern Healthcare Mgmt',
    'southern assisted living, llc': '(ABSORBED)',
    'sovereign healthcare holdings': '(ABSORBED)',
    'cedarhurst senor living': 'Cedarhurst Senior Living',
    'cedarhurst senior living': 'Cedarhurst Senior Living',
}

def canonicalize(name):
    name = name.strip()
    low = name.lower()
    if low in ALIASES:
        return ALIASES[low]
    # Case-insensitive match against known canonical names from ENTITY_DB_NAMES
    if not hasattr(canonicalize, '_lower_lookup'):
        canonicalize._lower_lookup = {}
    return canonicalize._lower_lookup.get(low, name)

def _build_canonical_lookup():
    """Call after ENTITY_DB_NAMES is defined to populate the case-insensitive lookup."""
    lookup = {}
    for display_name in ENTITY_DB_NAMES:
        lookup[display_name.lower()] = display_name
        for db_name in ENTITY_DB_NAMES[display_name]:
            lookup[db_name.lower()] = display_name
    # Also add T5 names
    for display_name in T5_BARRIERS:
        lookup[display_name.lower()] = display_name
        for db_name in T5_DB_NAMES.get(display_name, set()):
            lookup[db_name.lower()] = display_name
    canonicalize._lower_lookup = lookup

# ============================================================
# T5 BARRIERS
# ============================================================
T5_BARRIERS = {
    'Bluegrass/Encore': 'Alliance, MH Only Opportunity',
    'Signature Health': 'Own provider group, MH Only',
    'MFA': 'Alliance (VA)',
    'CommuniCare': 'Own provider group — kicking us out everywhere',
    'Hill Valley': 'Own provider group (MH Only 85%)',
    'Singh': 'Own provider group',
    'Cardon & Associates': 'Own provider group (91%), Hoosier Alliance',
    'Eastern Healthcare Group': 'Termination Risk (100%)',
    'Clearview': 'Alliance + Own Provider Group (38%), Telos partner',
    'Pavilion Healthcare': 'Alliance + Own Provider Group (29%)',
    'Embassy': 'Own provider group',
    'Exceptional Living': 'Own provider group',
    'Portopiccolo': 'Alliance (Telos)',
    'Venza': 'Alliance',
    'Aventura': 'Recent LOB',
    'Journey': 'Explicit rejection',
}

T5_DB_NAMES = {
    'Bluegrass/Encore': {'BLUEGRASS/ENCORE', 'BLUEGRASS', 'ENCORE SENIOR LIVING'},
    'Signature Health': {'SIGNATURE HEALTHCARE', 'SIGNATURE HEALTH'},
    'MFA': {'MFA', 'MFA MARYVILLE RE LLC'},
    'CommuniCare': {'COMMUNICARE'},
    'Hill Valley': {'HILL VALLEY'},
    'Singh': {'SINGH'},
    'Cardon & Associates': {'CARDON & ASSOCIATES'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'Clearview': {'CLEARVIEW'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'Embassy': {'EMBASSY'},
    'Exceptional Living': {'EXCEPTIONAL LIVING CENTERS'},
    'Portopiccolo': {'PORTOPICCOLO'},
    'Venza': {'VENZA CARE MANAGEMENT'},
    'Aventura': {'AVENTURA'},
    'Journey': {'JOURNEY'},
}

ALL_T5_DB = set()
for names in T5_DB_NAMES.values():
    ALL_T5_DB.update(names)

# ============================================================
# DB NAME -> DISPLAY NAME for all non-T5 entities
# ============================================================
# Build from Finance + BD workbooks, plus known DB names
ENTITY_DB_NAMES = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'American Senior Communities': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'Saber Healthcare Group': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'Infinity Healthcare Consulting': {'INFINITY HEALTHCARE CONSULTING'},
    'Navion': {'NAVION', 'NAVION SENIOR SOLUTIONS'},
    'Majestic Care': {'MAJESTIC CARE'},
    'Pruitt Health': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Trilogy': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Topaz Healthcare': {'TOPAZ HEALTHCARE'},
    'Morning Pointe Senior Living': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'Principle': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TerraBella Senior Living': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'Sanstone': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'Lionstone Care': {'LIONSTONE CARE', 'LIONSTONE'},
    'Eldercare Partners': {'ELDERCARE PARTNERS'},
    'Otterbein Senior Life': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'Peak Resources': {'PEAK RESOURCES'},
    'Arbors': {'ARBORS', 'ARBORS AT OHIO'},
    'American Healthcare LLC': {'AMERICAN HEALTHCARE, LLC'},
    'CCH Healthcare': {'CCH HEALTHCARE'},
    'Priority': {'PRIORITY LIFE CARE'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'Caring Place Healthcare': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'Fundamental LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Momentus Health': {'MOMENTUS HEALTH'},
    'Altercare': {'ALTERCARE'},
    'AOM Healthcare': {'AOM HEALTHCARE'},
    'Aperion Care': {'APERION CARE'},
    'Brickyard Healthcare': {'BRICKYARD HEALTHCARE'},
    'Carecore Health': {'CARECORE HEALTH'},
    'Carespring': {'CARESPRING'},
    'Carrolton': {'CARROLTON FACILTY MANAGEMENT', 'CARROLTON NURSING HOMES'},
    'Certus Healthcare': {'CERTUS HEALTHCARE'},
    'Choice Health Management': {'CHOICE HEALTH MANAGEMENT'},
    'Ciena Healthcare/Laurel Health Care': {'CIENA HEALTHCARE/LAUREL HEALTH CARE'},
    'Continuing Healthcare Solutions': {'CONTINUING HEALTHCARE SOLUTIONS'},
    'Crown Healthcare Group': {'CROWN HEALTHCARE GROUP'},
    'Divine Healthcare Management': {'DIVINE HEALTHCARE MANAGEMENT'},
    'Envive Healthcare': {'ENVIVE HEALTHCARE'},
    'Gardant Management Solutions': {'GARDANT MANAGEMENT SOLUTIONS, INC'},
    'HCF Management': {'HCF MANAGEMENT'},
    'Health Care Management Group': {'HEALTH CARE MANAGEMENT GROUP'},
    'Hillstone Healthcare': {'HILLSTONE HEALTHCARE'},
    'Legacy Health Services': {'LEGACY HEALTH SERVICES'},
    "Miller's Merry Manor": {'MILLERS MERRY MANOR', 'MILLERS MERRY MANOR INDIANAPOLIS EAST LLC'},
    'National Healthcare Corporation': {'NATIONAL HEALTHCARE CORPORATION'},
    'Ohio Living Communities': {'OHIO LIVING COMMUNITIES'},
    'Optalis Health & Rehabilitation': {'OPTALIS HEALTH & REHABILITATION'},
    'PACS Group': {'PACS GROUP'},
    'Phoenix Senior Living': {'PHOENIX SENIOR LIVING'},
    'Progressive Quality Care': {'PROGRESSIVE QUALITY CARE'},
    'SEKY Holding Co.': {'SEKY HOLDING CO.'},
    'Sprenger Health Care Systems': {'SPRENGER HEALTH CARE SYSTEMS'},
    'Sunrise Senior Living': {'SUNRISE SENIOR LIVING', 'SUNRISE'},
    'Trio Healthcare': {'TRIO HEALTHCARE'},
    'Vancrest Health Care Centers': {'VANCREST HEALTH CARE CENTERS'},
    'White Oak Management': {'WHITE OAK MANAGEMENT'},
    'Windsor House': {'WINDSOR HOUSE, INC.'},
    # Finance-only entities that need DB name mappings
    'Sonida Senior Living': {'SONIDA SENIOR LIVING'},
    'StoryPoint': {'STORYPOINT', 'STORYPOINT SENIOR LIVING'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'Senior Lifestyle': {'SENIOR LIFESTYLE', 'SENIOR LIFESTYLE CORPORATION'},
    'Greencroft': {'GREENCROFT', 'GREENCROFT COMMUNITIES'},
    'Spring Arbor Management': {'SPRING ARBOR MANAGEMENT'},
    'Cedarhurst Senior Living': {'CEDARHURST SENIOR LIVING'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES'},
    'Atrium Health': {'ATRIUM HEALTH'},
    'Brighton': {'BRIGHTON'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE'},
    'Triple Crown Senior Living': {'TRIPLE CROWN SENIOR LIVING'},
}

_build_canonical_lookup()

db_to_display = {}
all_db_names = set()
for display_name, db_names in ENTITY_DB_NAMES.items():
    for dn in db_names:
        db_to_display[dn] = display_name
        all_db_names.add(dn)

# SI scores
DATA_SI_SCORES = {
    'ALG': 5, 'Brookdale Senior Living': 5, 'Saber Healthcare Group': 5,
    'Trilogy': 5, 'Liberty': 5, 'Pruitt Health': 5, 'Avardis': 5, 'Lifecare': 5,
    'Sonida Senior Living': 5,
    'American Senior Communities': 4, 'Infinity Healthcare Consulting': 4,
    'Majestic Care': 4, 'Navion': 4, 'TLC Management': 4,
    'Kisco Senior Living': 4, 'Lutheran Life Villages': 4, 'Lutheran Services Carolinas': 4,
    'Sanstone': 3, 'Morning Pointe Senior Living': 3, 'Otterbein Senior Life': 3,
    'Principle': 3, 'Lionstone Care': 3, 'Kissito Healthcare': 3,
    'CCH Healthcare': 3, 'TerraBella Senior Living': 3, 'Peak Resources': 3,
    'BHI Senior Living': 3,
    'Brickyard Healthcare': 5, 'Ciena Healthcare/Laurel Health Care': 5,
    'National Healthcare Corporation': 5, 'PACS Group': 5, 'Ohio Living Communities': 5,
    'Sunrise Senior Living': 4, 'Carespring': 4,
    'Altercare': 3, 'AOM Healthcare': 3, 'Aperion Care': 3, 'Certus Healthcare': 3,
    'Continuing Healthcare Solutions': 3, 'Crown Healthcare Group': 3,
    'Gardant Management Solutions': 3, 'Hillstone Healthcare': 3,
    "Miller's Merry Manor": 3, 'Optalis Health & Rehabilitation': 3,
    'Phoenix Senior Living': 3, 'White Oak Management': 3,
    'StoryPoint': 3, 'Greencroft': 2, 'Senior Lifestyle': 2,
    'Spring Arbor Management': 2, 'Cedarhurst Senior Living': 2, 'MCAP': 2,
}

# RS scores
TOM_RS_SCORES = {
    'ALG': 2, 'American Senior Communities': 5, 'Brookdale Senior Living': 5,
    'Saber Healthcare Group': 2, 'Infinity Healthcare Consulting': 2,
    'Navion': 3, 'Majestic Care': 3, 'Pruitt Health': 3,
    'Trilogy': 4, 'Topaz Healthcare': 1, 'Morning Pointe Senior Living': 2,
    'Principle': 3, 'Liberty': 4, 'TerraBella Senior Living': 2,
    'Sanstone': 2, 'Lionstone Care': 3, 'Eldercare Partners': 3,
    'Otterbein Senior Life': 1, 'TLC Management': 3, 'BHI Senior Living': 1,
    'Avardis': 2, 'Peak Resources': 2, 'Arbors': 1,
    'American Healthcare LLC': 3, 'CCH Healthcare': 2, 'Priority': 1,
    'Lifecare': 1, 'Kissito Healthcare': 2, 'YAD': 2,
    'Caring Place Healthcare': 2, 'Castle Healthcare': 3, 'JAG': 2,
    'Fundamental LTC': 1, 'Southern Healthcare Mgmt': 1, 'Momentus Health': 1,
    'StoryPoint': 3, 'Greencroft': 1, 'Sonida Senior Living': 1,
    'Kisco Senior Living': 1, 'Senior Lifestyle': 1, 'MCAP': 1,
    'Spring Arbor Management': 1, 'Cedarhurst Senior Living': 1,
    'Lutheran Services Carolinas': 4, 'Lutheran Life Villages': 4,
    'Atrium Health': 1, 'LifeSpire of Virginia': 1,
}

# V20 AI scores
print("Loading V20 for AI scores...")
wb_v20 = openpyxl.load_workbook(V20_PATH, data_only=True)
V20_AI = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_v20[sn]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        canon = canonicalize(str(name))
        if canon == '(ABSORBED)':
            continue
        ai = ws.cell(r, 14).value or 0
        V20_AI[canon] = int(ai) if ai else 0

# ============================================================
# LOAD DB AND COMPUTE ALL ENTITY ROLLUPS
# ============================================================
print("Loading V23 DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
db_headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

entity_facilities = defaultdict(list)
for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, db_headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, db_headers['State']).value or '').strip()
    if state not in FOOTPRINT:
        continue
    if corp in ALL_T5_DB:
        continue
    if corp not in all_db_names:
        continue

    display_name = db_to_display[corp]
    fac_name = str(ws_db.cell(r, db_headers['Facility_Name']).value or '').strip()
    addr = str(ws_db.cell(r, db_headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, db_headers['City']).value or '').strip()
    src_type = str(ws_db.cell(r, db_headers['Source_Type']).value or '').strip()
    geo_tier = str(ws_db.cell(r, db_headers.get('Geographic_Tier', 0)).value or '').strip() if db_headers.get('Geographic_Tier') else ''
    metro = str(ws_db.cell(r, db_headers.get('Metro_Assignment', 0)).value or '').strip() if db_headers.get('Metro_Assignment') else ''
    served = str(ws_db.cell(r, db_headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws_db.cell(r, db_headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws_db.cell(r, db_headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws_db.cell(r, db_headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    tot_beds = ws_db.cell(r, db_headers['Total_Beds']).value or 0
    tot_beds = tot_beds if isinstance(tot_beds, (int, float)) else 0
    census = ws_db.cell(r, db_headers.get('Census', 0)).value or 0 if db_headers.get('Census') else 0
    census = census if isinstance(census, (int, float)) else 0
    tot_rev = ws_db.cell(r, db_headers['Total_Potential_Revenue']).value or 0
    tot_rev = tot_rev if isinstance(tot_rev, (int, float)) else 0
    int_rev = ws_db.cell(r, db_headers['Integration_Revenue']).value or 0
    int_rev = int_rev if isinstance(int_rev, (int, float)) else 0
    new_rev = ws_db.cell(r, db_headers['New_Business_Revenue']).value or 0
    new_rev = new_rev if isinstance(new_rev, (int, float)) else 0
    cur_rev = ws_db.cell(r, db_headers.get('Current_Revenue', 0)).value or 0 if db_headers.get('Current_Revenue') else 0
    cur_rev = cur_rev if isinstance(cur_rev, (int, float)) else 0

    street_num = extract_street_number(addr)
    norm_addr = normalize_address(addr)
    num_key = (street_num, city.upper()) if street_num else (norm_addr, city.upper())

    entity_facilities[display_name].append({
        'fac_name': fac_name, 'addr': addr, 'city': city,
        'state': state, 'src_type': src_type, 'geo_tier': geo_tier, 'metro': metro,
        'campus_key': num_key, 'served': served,
        'mh': mh, 'pcp': pcp, 'integ': integ,
        'tot_beds': tot_beds, 'census': census, 'tot_rev': tot_rev,
        'int_rev': int_rev, 'new_rev': new_rev, 'cur_rev': cur_rev,
    })

print("Computing entity scores...")
all_entities = {}

for display_name in sorted(ENTITY_DB_NAMES.keys()):
    facs = entity_facilities.get(display_name, [])
    if not facs:
        all_entities[display_name] = {
            'tier': 'T4', 'score': None, 'n_camp': 0, 'n_srv': 0,
            'total_rev': 0, 'int_rev': 0, 'new_rev': 0, 'states': '', 'n_states': 0,
            'ER': None, 'IR': None, 'SI': None, 'RP': None, 'RS': None, 'AI': None,
            'reason': 'No DB rows in footprint',
        }
        continue

    qualifying = [f for f in facs if f['tot_beds'] > BED_MIN]
    campuses = set(f['campus_key'] for f in qualifying)
    served_camps = set(f['campus_key'] for f in qualifying if f['served'])
    mh_camps = set(f['campus_key'] for f in qualifying if f['mh'])
    pcp_camps = set(f['campus_key'] for f in qualifying if f['pcp'])
    integ_camps = set(f['campus_key'] for f in qualifying if f['integ'])
    total_rev = sum(f['tot_rev'] for f in qualifying)
    int_rev = sum(f['int_rev'] for f in qualifying)
    new_rev = sum(f['new_rev'] for f in qualifying)
    n_camp = len(campuses)
    n_srv = len(served_camps)
    states = sorted(set(f['state'] for f in qualifying))

    if n_camp < 7:
        all_entities[display_name] = {
            'tier': 'T4', 'score': None, 'n_camp': n_camp, 'n_srv': n_srv,
            'total_rev': total_rev, 'int_rev': int_rev, 'new_rev': new_rev,
            'states': ', '.join(states), 'n_states': len(states),
            'ER': None, 'IR': None, 'SI': None, 'RP': None, 'RS': None, 'AI': None,
            'reason': f'{n_camp} qualifying campuses (gate requires 7)',
        }
        continue

    er = score_dim(n_camp, ER_THRESHOLDS)
    rp = score_dim(total_rev, RP_THRESHOLDS)
    rs = TOM_RS_SCORES.get(display_name, 1)
    si = DATA_SI_SCORES.get(display_name, 2)
    ai = V20_AI.get(display_name, 0)

    n_integ = len(integ_camps)
    int_pct = n_integ / n_camp if n_camp > 0 else 0
    dual = len(mh_camps & pcp_camps)
    if n_srv == 0:
        ir = 1
    elif int_pct >= 0.5:
        ir = 5
    elif int_pct >= 0.25 or dual > 0:
        ir = 4
    elif len(mh_camps) > 0 and len(pcp_camps) > 0:
        ir = 3
    elif len(mh_camps) > 0 or len(pcp_camps) > 0:
        ir = 2
    else:
        ir = 1

    total_score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = 'T1' if total_score >= 55 else 'T2' if total_score >= 35 else 'T3'

    all_entities[display_name] = {
        'tier': tier, 'score': total_score, 'n_camp': n_camp, 'n_srv': n_srv,
        'total_rev': total_rev, 'int_rev': int_rev, 'new_rev': new_rev,
        'states': ', '.join(states), 'n_states': len(states),
        'ER': er, 'IR': ir, 'SI': si, 'RP': rp, 'RS': rs, 'AI': ai,
    }

# Also read Finance workbook to capture any entities we missed or to cross-check
print("Reading Finance workbook for cross-check...")
wb_fin = openpyxl.load_workbook(V23_FIN_PATH, data_only=True)
ws_sum = wb_fin['Summary']
for r in range(2, ws_sum.max_row + 1):
    name = ws_sum.cell(r, 1).value
    if not name:
        continue
    canon = canonicalize(str(name))
    if canon not in all_entities or all_entities[canon]['score'] is None:
        score = ws_sum.cell(r, 15).value
        tier = ws_sum.cell(r, 16).value
        camps = ws_sum.cell(r, 17).value
        served = ws_sum.cell(r, 18).value
        rev = ws_sum.cell(r, 19).value
        er = ws_sum.cell(r, 3).value or 0
        ir = ws_sum.cell(r, 5).value or 0
        si = ws_sum.cell(r, 7).value or 0
        rp = ws_sum.cell(r, 9).value or 0
        rs = ws_sum.cell(r, 11).value or 0
        ai = ws_sum.cell(r, 13).value or 0
        if score:
            existing = all_entities.get(canon, {})
            all_entities[canon] = {
                'tier': str(tier).strip(), 'score': int(score),
                'n_camp': int(camps) if camps else 0,
                'n_srv': int(served) if served else 0,
                'total_rev': float(rev) if rev else 0,
                'int_rev': existing.get('int_rev', 0),
                'new_rev': existing.get('new_rev', 0),
                'states': existing.get('states', ''),
                'n_states': existing.get('n_states', 0),
                'ER': int(er), 'IR': int(ir), 'SI': int(si),
                'RP': int(rp), 'RS': int(rs), 'AI': int(ai),
            }

# ============================================================
# IDENTIFY SOURCE COVERAGE
# ============================================================
# Read BD workbook to tag which entities are in it
wb_bd = openpyxl.load_workbook(V23_BD_PATH, data_only=True)
bd_names = set()
for sn in wb_bd.sheetnames:
    ws = wb_bd[sn]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name and str(name).strip() not in ('', 'TOTAL', 'FACILITY DETAIL'):
            bd_names.add(canonicalize(str(name)))

# Read Finance scored names
fin_names = set()
for r in range(2, ws_sum.max_row + 1):
    name = ws_sum.cell(r, 1).value
    if name:
        fin_names.add(canonicalize(str(name)))
ws_t4f = wb_fin['T4 - Independents']
for r in range(2, ws_t4f.max_row + 1):
    name = ws_t4f.cell(r, 1).value
    if name and str(name).strip() not in ('', 'TOTAL', 'FACILITY DETAIL', 'Finance Tier'):
        fin_names.add(canonicalize(str(name)))
for n in T5_BARRIERS:
    fin_names.add(n)

# ============================================================
# BUILD HTML
# ============================================================
print("Building HTML...")

tier_order = {'T1': 1, 'T2': 2, 'T3': 3, 'T4': 4, 'T5': 5}
tier_names = {
    'T1': 'T1 — Strategic Enterprise',
    'T2': 'T2 — Growth & Expansion',
    'T3': 'T3 — Retention / Watch',
    'T4': 'T4 — Independents',
    'T5': 'T5 — Hard Barriers',
}
tier_colors = {
    'T1': ('#C6EFCE', '#193241'), 'T2': ('#FFEB9C', '#193241'),
    'T3': ('#FFC7CE', '#193241'), 'T4': ('#E1E2E0', '#193241'),
    'T5': ('#F2DCDB', '#193241'),
}
tier_descriptions = {
    'T1': 'Score &ge;55. C-suite access, dedicated CAM, quarterly exec reviews.',
    'T2': 'Score 35&ndash;54. Regional leadership access, CAM-led with rotational clinical, regular QBRs.',
    'T3': 'Score &lt;35. Facility-level engagement, field-managed, annual touchpoints.',
    'T4': '&lt;7 qualifying campuses in 6-state footprint. Facility-level engagement, no corporate account management.',
    'T5': 'Structural barriers prevent growth. Defend current revenue, monitor barrier status annually.',
}

# Group entities by tier
by_tier = defaultdict(list)
for name, data in all_entities.items():
    by_tier[data['tier']].append((name, data))

# Add T5 barriers
for name, barrier in T5_BARRIERS.items():
    if name not in all_entities:
        by_tier['T5'].append((name, {'tier': 'T5', 'barrier': barrier}))

html = []
html.append(f'''<html>
<head>
<style>
html {{ scroll-behavior: smooth; }}
body {{ font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #333; max-width: 1300px; margin: 0 auto; padding: 20px; }}
h1 {{ color: #193241; border-bottom: 3px solid #877BE9; padding-bottom: 10px; }}
h2 {{ margin-top: 40px; padding: 10px 16px; border-radius: 4px; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0 30px 0; font-size: 10.5pt; }}
th {{ background-color: #193241; color: white; padding: 8px 10px; text-align: center; font-weight: bold; }}
th:first-child {{ text-align: left; }}
td {{ padding: 5px 10px; border: 1px solid #ddd; }}
tr:nth-child(even) {{ background-color: #F8F8F8; }}
.score {{ text-align: center; font-weight: bold; }}
.dim {{ text-align: center; }}
.money {{ text-align: right; }}
.num {{ text-align: center; }}
.new-badge {{ background-color: #E3F2FD; color: #293A60; font-size: 9pt; padding: 1px 6px; border-radius: 3px; font-weight: bold; margin-left: 6px; }}
.entity-name {{ font-weight: 500; }}
.summary-box {{ display: inline-block; background: #F5F5F5; border: 1px solid #E1E2E0; border-radius: 6px; padding: 12px 18px; margin: 3px; text-align: center; }}
.summary-count {{ font-size: 26pt; font-weight: bold; color: #193241; }}
.summary-label {{ font-size: 10pt; color: #507E8E; }}
.dim-box {{ flex: 1; border: 1px solid #E1E2E0; border-radius: 6px; padding: 14px 6px; margin: 3px; text-align: center; background: #FAFAFA; display: flex; flex-direction: column; justify-content: center; border-top: 3px solid #877BE9; }}
.dim-abbr {{ font-size: 18pt; font-weight: bold; line-height: 1.1; color: #193241; }}
.dim-name {{ font-size: 8.5pt; color: #507E8E; margin-top: 3px; white-space: nowrap; }}
.dim-weight {{ font-size: 8.5pt; color: #999; margin-top: 2px; }}
.chevron-toggle {{ cursor: pointer; user-select: none; display: inline-flex; align-items: center; gap: 6px; color: #507E8E; font-size: 10.5pt; font-weight: 500; margin: 2px 0; padding: 6px 12px; border: 1px solid #E1E2E0; border-radius: 4px; background: #F5F7F8; }}
.chevron-toggle:hover {{ background: #E8ECEE; }}
.chevron-toggle .arrow {{ transition: transform 0.2s; display: inline-block; }}
.chevron-content {{ display: none; }}
</style>
</head>
<body>

<h1 style="margin-bottom: 0;">Multi-Unit Operators Scored by Tiers</h1>
<div style="font-size: 13pt; color: #507E8E; font-weight: 500; margin-bottom: 12px;">Multi-Attribute Analysis &mdash; utilizing source data from EWH&rsquo;s Market Intelligence Database (V23)</div>
<p style="font-size: 10.5pt; color: #507E8E; line-height: 1.6; margin: 0 0 10px 0;">Each multi-unit operator is scored across six dimensions &mdash; enterprise reach, integration readiness, strategic influence, revenue potential, relationship strength, and technology adoption &mdash; then placed into one of five tiers based on composite score. Results reflect 100 corporate entities across our six-state operational footprint (NC, SC, VA, KY, OH, IN).</p>

<!-- ============================================================ -->
<!-- RESULTS + METHODOLOGY — side by side -->
<!-- ============================================================ -->
<div style="display: flex; flex-wrap: wrap; gap: 20px; margin: 25px 0 10px 0; align-items: stretch;">

<!-- LEFT: Results -->
<div style="flex: 0 0 auto;">
<div style="font-size: 13pt; font-weight: bold; color: #193241; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px;">Results</div>
<div style="display: flex; flex-wrap: nowrap;">
''')

total = 0
for tier in ['T1', 'T2', 'T3', 'T4', 'T5']:
    count = len(by_tier.get(tier, []))
    total += count
    bg, fg = tier_colors[tier]
    html.append(f'<a href="#section-{tier}" style="text-decoration: none;"><div class="summary-box" style="cursor: pointer;"><div class="summary-count" style="color: {fg};">{count}</div><div class="summary-label" style="background: {bg}; padding: 2px 10px; border-radius: 3px;">{tier}</div></div></a>')

html.append(f'<div class="summary-box"><div class="summary-count">{total}</div><div class="summary-label">Total</div></div>')
html.append('''</div>
<div style="margin-top: 4px; display: flex; align-items: center; gap: 10px; flex-wrap: wrap;">
<div style="display: inline-flex; align-items: center; gap: 6px; color: #507E8E; font-size: 10.5pt; font-weight: 500; padding: 8px 12px; border: 1px solid #E1E2E0; border-radius: 4px; background: #F5F7F8;">
<span>&#9650;</span> Click on tier to see full results
</div>
<span style="font-size: 9pt; color: #aaa;">T1 &ge;55 &nbsp;|&nbsp; T2 &ge;35 &nbsp;|&nbsp; T3 &lt;35 &nbsp;|&nbsp; T4 &lt;7 campuses &nbsp;|&nbsp; T5 barrier</span>
</div>
</div>

<!-- RIGHT: Methodology -->
<div style="flex: 1; min-width: 0;">
<div style="font-size: 13pt; font-weight: bold; color: #193241; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px;">Methodology</div>
<div style="display: flex; flex-wrap: nowrap; align-items: stretch;">

<div class="dim-box">
<div class="dim-abbr">ER</div>
<div class="dim-name">Enterprise Reach</div>
<div class="dim-weight">Weight &times;4</div>
</div>

<div class="dim-box">
<div class="dim-abbr">IR</div>
<div class="dim-name">Integration Readiness</div>
<div class="dim-weight">Weight &times;3</div>
</div>

<div class="dim-box">
<div class="dim-abbr">SI</div>
<div class="dim-name">Strategic Influence</div>
<div class="dim-weight">Weight &times;3</div>
</div>

<div class="dim-box">
<div class="dim-abbr">RP</div>
<div class="dim-name">Revenue Potential</div>
<div class="dim-weight">Weight &times;4</div>
</div>

<div class="dim-box">
<div class="dim-abbr">RS</div>
<div class="dim-name">Relationship Strength</div>
<div class="dim-weight">Weight &times;3</div>
</div>

<div class="dim-box">
<div class="dim-abbr">AI</div>
<div class="dim-name">Tech Adoption</div>
<div class="dim-weight">Weight &times;3</div>
</div>

</div>
<div style="margin-top: 4px; display: flex; align-items: center; justify-content: flex-end; flex-wrap: wrap; gap: 10px;">
<span style="font-size: 9pt; color: #aaa;">(ER&times;4) + (IR&times;3) + (SI&times;3) + (RP&times;4) + (RS&times;3) + (AI&times;3) &nbsp;|&nbsp; Max 100</span>
<div class="chevron-toggle" style="display: inline-flex; padding: 8px 12px;" onclick="var c=document.getElementById('dim-detail');var a=this.querySelector('.arrow');if(c.style.display==='none'||c.style.display===''){c.style.display='block';a.style.transform='rotate(90deg)';}else{c.style.display='none';a.style.transform='rotate(0deg)';}">
<span class="arrow">&#9654;</span> Scoring Detail &mdash; click to expand
</div>
</div>
</div>

</div>

<div id="dim-detail" class="chevron-content">

<div style="display: flex; flex-wrap: wrap; gap: 12px; margin: 15px 0 10px 0;">

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">ER &mdash; Enterprise Reach</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;4 (max 20)</b><br>
Unique campuses in 6-state footprint<br>
(street-number collapsed, &gt;15 beds)
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">40+ campuses</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">20&ndash;39</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">13&ndash;19</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">9&ndash;12</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">&lt;9</td></tr>
</table>
</div></div>

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">IR &mdash; Integration Readiness</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;3 (max 15)</b><br>
Service configuration from DB flags<br>
(MH_Flag, PCP_Flag, Integrated_Flag)
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">Majority integrated (&ge;50%)</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">Dual MH+PCP same campus, or 25&ndash;49%</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">MH + PCP at separate campuses</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">Single service only (MH or PCP)</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">Not served</td></tr>
</table>
</div></div>

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">SI &mdash; Strategic Influence</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;3 (max 15)</b><br>
Tom&rsquo;s Notes + web research<br>
(REIT, ISNP, ACO, public company)
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">REIT/ISNP/Enterprise (NYSE, I-SNP)</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">Payer/ACO connected</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">Regional reputation, expanding</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">Minimal local presence</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">No affiliations found</td></tr>
</table>
</div></div>

</div>
<div style="display: flex; flex-wrap: wrap; gap: 12px; margin: 0 0 20px 0;">

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">RP &mdash; Revenue Potential</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;4 (max 20)</b><br>
S2 Total Potential Revenue (TAM)<br>
in footprint, &gt;15 bed facilities only
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">&gt;$10M</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">$5M&ndash;$10M</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">$2.5M&ndash;$5M</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">$1M&ndash;$2.5M</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">&lt;$1M</td></tr>
</table>
</div></div>

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">RS &mdash; Relationship Strength</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;3 (max 15)</b><br>
Tom&rsquo;s Shared Savings Notes<br>
(actual sales engagement status)
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">Active SS on Eventus paper</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">Active engagement, verbal commits</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">Have relationship, stalled</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">Tried, hit resistance</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">No relationship</td></tr>
</table>
</div></div>

<div style="flex: 1; min-width: 190px; border: 1px solid #E1E2E0; border-top: 4px solid #877BE9; border-radius: 4px; padding: 12px 14px; background: #FAFAFA;">
<div style="font-weight: bold; color: #193241; font-size: 12pt; margin-bottom: 6px;">AI &mdash; Tech Adoption</div>
<div style="font-size: 9.5pt; color: #507E8E; line-height: 1.5;">
<b>Weight: &times;3 (max 15)</b><br>
V20 Brooke scores where available<br>
(default 0 &mdash; needs data-driven method)
<table style="width: 100%; margin-top: 6px; font-size: 9pt; border-collapse: collapse;">
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">5</td><td style="padding: 2px 4px; border: none;">Advanced AI/analytics adoption</td></tr>
<tr><td style="padding: 2px 4px; border: none;">4</td><td style="padding: 2px 4px; border: none;">Active tech investment</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">3</td><td style="padding: 2px 4px; border: none;">Moderate EHR/data capability</td></tr>
<tr><td style="padding: 2px 4px; border: none;">2</td><td style="padding: 2px 4px; border: none;">Basic digital infrastructure</td></tr>
<tr style="background: #F3F1FC;"><td style="padding: 2px 4px; border: none;">1</td><td style="padding: 2px 4px; border: none;">Minimal tech presence</td></tr>
</table>
</div></div>

</div>
</div>
''')

# Scored tiers: T1, T2, T3
for tier in ['T1', 'T2', 'T3']:
    entities = by_tier.get(tier, [])
    entities.sort(key=lambda x: (-(x[1].get('score') or 0), x[0]))
    bg, fg = tier_colors[tier]

    html.append(f'<h2 id="section-{tier}" style="background-color: {bg}; color: {fg};">{tier_names[tier]} ({len(entities)})</h2>')
    html.append(f'<p style="margin: 0 0 10px 0; font-size: 10pt; color: #666;">{tier_descriptions[tier]}</p>')
    html.append('''<table>
<tr><th style="width: 220px;">MUO Name</th><th style="width: 40px;">Score</th><th>States</th>
<th style="width: 55px;">Campuses</th><th style="width: 45px;">Serve</th>
<th style="width: 95px;">Integration</th><th style="width: 95px;">New Business</th><th style="width: 95px;">Total Opp</th>
<th style="width: 30px;">ER</th><th style="width: 30px;">IR</th><th style="width: 30px;">SI</th>
<th style="width: 30px;">RP</th><th style="width: 30px;">RS</th><th style="width: 30px;">AI</th></tr>''')

    for name, data in entities:
        score = data.get('score', 0) or 0
        states = data.get('states', '')
        camps = data.get('n_camp', 0)
        srv = data.get('n_srv', 0)
        i_rev = data.get('int_rev', 0)
        n_rev = data.get('new_rev', 0)
        t_rev = data.get('total_rev', 0)
        i_str = f'${i_rev:,.0f}' if i_rev else '-'
        n_str = f'${n_rev:,.0f}' if n_rev else '-'
        t_str = f'${t_rev:,.0f}' if t_rev else '-'

        badge = ''
        if name not in bd_names:
            badge = '<span class="new-badge">NEW</span>'

        slug = slugify(name)
        html.append(f'''<tr>
<td class="entity-name"><a href="BD_MUO_Breakouts.html#{slug}" style="color: #193241; text-decoration: none; border-bottom: 1px dotted #877BE9;">{name}</a>{badge}</td>
<td class="score">{score}</td><td>{states}</td>
<td class="num">{camps}</td><td class="num">{srv}</td>
<td class="money">{i_str}</td><td class="money">{n_str}</td><td class="money">{t_str}</td>
<td class="dim">{data.get("ER", "-")}</td><td class="dim">{data.get("IR", "-")}</td>
<td class="dim">{data.get("SI", "-")}</td><td class="dim">{data.get("RP", "-")}</td>
<td class="dim">{data.get("RS", "-")}</td><td class="dim">{data.get("AI", "-")}</td></tr>''')

    html.append('</table>')

# T4
entities_t4 = by_tier.get('T4', [])
entities_t4.sort(key=lambda x: (-x[1].get('n_camp', 0), x[0]))
bg, fg = tier_colors['T4']
html.append(f'<h2 id="section-T4" style="background-color: {bg}; color: {fg};">{tier_names["T4"]} ({len(entities_t4)})</h2>')
html.append(f'<p style="margin: 0 0 10px 0; font-size: 10pt; color: #666;">{tier_descriptions["T4"]}</p>')
html.append('''<table>
<tr><th style="width: 220px;">MUO Name</th><th style="width: 60px;">Campuses</th><th style="width: 45px;">Serve</th>
<th>States</th><th style="width: 95px;">Integration</th><th style="width: 95px;">New Business</th><th style="width: 95px;">Total Opp</th><th>Reason</th></tr>''')

for name, data in entities_t4:
    camps = data.get('n_camp', 0)
    srv = data.get('n_srv', 0)
    states = data.get('states', '')
    i_rev = data.get('int_rev', 0)
    n_rev = data.get('new_rev', 0)
    t_rev = data.get('total_rev', 0)
    i_str = f'${i_rev:,.0f}' if i_rev else '-'
    n_str = f'${n_rev:,.0f}' if n_rev else '-'
    t_str = f'${t_rev:,.0f}' if t_rev else '-'
    reason = data.get('reason', f'{camps} qualifying campuses (gate requires 7)')
    badge = ''
    if name not in bd_names:
        badge = '<span class="new-badge">NEW</span>'

    slug = slugify(name)
    html.append(f'''<tr>
<td class="entity-name"><a href="BD_MUO_Breakouts.html#{slug}" style="color: #193241; text-decoration: none; border-bottom: 1px dotted #877BE9;">{name}</a>{badge}</td>
<td class="num">{camps}</td><td class="num">{srv}</td>
<td>{states}</td><td class="money">{i_str}</td><td class="money">{n_str}</td><td class="money">{t_str}</td><td style="font-size: 10pt; color: #666;">{reason}</td></tr>''')

html.append('</table>')

# T5
entities_t5 = by_tier.get('T5', [])
# Deduplicate
t5_seen = set()
t5_deduped = []
for name, data in entities_t5:
    if name not in t5_seen:
        t5_seen.add(name)
        t5_deduped.append((name, data))
t5_deduped.sort(key=lambda x: x[0])

bg, fg = tier_colors['T5']
html.append(f'<h2 id="section-T5" style="background-color: {bg}; color: {fg};">{tier_names["T5"]} ({len(t5_deduped)})</h2>')
html.append(f'<p style="margin: 0 0 10px 0; font-size: 10pt; color: #666;">{tier_descriptions["T5"]}</p>')
html.append('''<table>
<tr><th style="width: 260px;">Corporate Name</th><th>Barrier</th></tr>''')

for name, data in t5_deduped:
    barrier = data.get('barrier', T5_BARRIERS.get(name, ''))
    slug = slugify(name)
    html.append(f'<tr><td class="entity-name"><a href="BD_MUO_Breakouts.html#{slug}" style="color: #193241; text-decoration: none; border-bottom: 1px dotted #877BE9;">{name}</a></td><td style="font-size: 10pt;">{barrier}</td></tr>')

html.append('</table>')

# Footer
html.append(f'''
<hr style="border: 1px solid #E1E2E0; margin: 40px 0;">
<p style="font-size: 10pt; color: #aaa;">
<b>Sources:</b> V23 Database (4_Economic_Model_Scenario_2_Combined_V23.xlsx), Finance Workbook (MUO_Scoring_Workbook_V23_v8.xlsx), BD Workbook (Final_MUO_Tiering_V23.xlsx), V20 Tiering (Final_MUO_Tiering_V20.xlsx)<br>
<b>Methodology:</b> 6.3_Corporate_Scoring_Methodology_V23.md &nbsp;|&nbsp; <span class="new-badge">NEW</span> = entity added from Finance workbook (not in prior BD view)<br>
Generated March 2026
</p>
</body></html>''')

html_content = '\n'.join(html)
with open(HTML_OUT, 'w', encoding='utf-8') as f:
    f.write(html_content)

# ============================================================
# GENERATE BREAKOUT DETAIL FILE
# ============================================================
print("Building breakout detail file...")

# Build MUO profile lookup
import os
PROFILE_DIR = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/MUO_Profiles/"
profile_files = {}
if os.path.isdir(PROFILE_DIR):
    for fn in os.listdir(PROFILE_DIR):
        if fn.endswith('_MUO_Profile.md'):
            key = fn.replace('_MUO_Profile.md', '').replace('_', ' ').lower()
            profile_files[key] = fn

def find_profile(name):
    """Try to match entity name to an MUO profile file."""
    key = name.lower()
    if key in profile_files:
        return profile_files[key]
    for pkey, pfile in profile_files.items():
        if pkey in key or key in pkey:
            return pfile
    return None

def parse_profile(filename):
    """Parse an MUO profile markdown to extract key sections."""
    path = os.path.join(PROFILE_DIR, filename)
    if not os.path.isfile(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()

    result = {}

    # Extract executive summary metrics table
    metrics = {}
    m = re.search(r'## 1\. Executive Summary\s*\n(.*?)(?=###|\n---)', content, re.DOTALL)
    if m:
        for row in re.findall(r'\|\s*\*\*(.+?)\*\*\s*\|\s*(.+?)\s*\|', m.group(1)):
            metrics[row[0].strip()] = row[1].strip()
    result['metrics'] = metrics

    # Extract Key Insight
    m = re.search(r'### Key Insight\s*\n\s*(.+?)(?=\n---|\n##)', content, re.DOTALL)
    if m:
        result['key_insight'] = m.group(1).strip()

    # Extract Leadership table
    leaders = []
    m = re.search(r'### Executive Leadership\s*\n\s*\|.*?\n\s*\|[-| ]+\|\s*\n(.*?)(?=\n###|\n---|\n##)', content, re.DOTALL)
    if m:
        for row in re.findall(r'\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.*?)\s*\|', m.group(1)):
            name_val = row[0].strip()
            if name_val and not name_val.startswith('-'):
                leaders.append({'name': name_val, 'title': row[1].strip(), 'notable': row[2].strip()})
    result['leaders'] = leaders

    # Extract Ownership info
    m = re.search(r'### Ownership[:\s]*(.+?)\s*\n', content)
    if m:
        result['ownership'] = m.group(1).strip()

    # Extract Recommendation
    m = re.search(r'## 8\. Recommendation\s*\n\s*(.*?)(?=\n---|\n## 9)', content, re.DOTALL)
    if m:
        result['recommendation'] = m.group(1).strip()

    # Extract HQ, CEO, Founded, Website from metrics
    result['hq'] = metrics.get('Headquarters', '')
    result['ceo'] = metrics.get('CEO', '')
    result['founded'] = metrics.get('Founded', '')
    result['website'] = metrics.get('Website', '')
    result['ownership_type'] = metrics.get('Ownership Type', '')
    result['revenue'] = metrics.get('Annual Revenue (Company)', '')
    result['employees'] = metrics.get('Employees', '')

    return result

# Parse all available profiles
parsed_profiles = {}
for pkey, pfile in profile_files.items():
    parsed_profiles[pfile] = parse_profile(pfile)

dim_meta = [
    ('ER', 'Enterprise Reach', 4, ER_BRACKETS),
    ('IR', 'Integration Readiness', 3, IR_BRACKETS),
    ('SI', 'Strategic Influence', 3, SI_BRACKETS),
    ('RP', 'Revenue Potential', 4, RP_BRACKETS),
    ('RS', 'Relationship Strength', 3, RS_BRACKETS),
    ('AI', 'Tech Adoption', 3, AI_BRACKETS),
]

bo = []
bo.append('''<html>
<head>
<style>
html { scroll-behavior: smooth; }
body { font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #333; max-width: 1200px; margin: 0 auto; padding: 20px; }
h1 { color: #193241; border-bottom: 3px solid #877BE9; padding-bottom: 10px; }
h3 { color: #193241; margin: 20px 0 8px 0; font-size: 13pt; }
.entity-section { margin-bottom: 50px; padding-bottom: 30px; border-bottom: 2px solid #E1E2E0; }
.entity-header { display: flex; align-items: center; gap: 14px; margin-bottom: 12px; flex-wrap: wrap; }
.entity-title { font-size: 18pt; font-weight: bold; color: #193241; }
.tier-badge { padding: 4px 14px; border-radius: 4px; font-weight: bold; font-size: 11pt; }
.score-badge { font-size: 14pt; font-weight: bold; color: #877BE9; }
.back-link { font-size: 10pt; color: #507E8E; text-decoration: none; border-bottom: 1px dotted #877BE9; }
.back-link:hover { color: #193241; }
.detail-grid { display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0; }
.detail-card { flex: 1; min-width: 130px; border: 1px solid #E1E2E0; border-radius: 6px; padding: 12px; background: #FAFAFA; }
.detail-card-label { font-size: 9pt; color: #507E8E; text-transform: uppercase; letter-spacing: 0.5px; }
.detail-card-value { font-size: 16pt; font-weight: bold; color: #193241; margin-top: 2px; }
.detail-card-sub { font-size: 9pt; color: #999; margin-top: 2px; }
table.dim-table { border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 10.5pt; }
table.dim-table th { background: #193241; color: white; padding: 6px 10px; text-align: left; font-weight: bold; }
table.dim-table td { padding: 5px 10px; border: 1px solid #E1E2E0; }
table.dim-table tr.active { background: #F3F1FC; font-weight: 500; }
table.fac-table { border-collapse: collapse; width: 100%; margin: 8px 0; font-size: 9.5pt; }
table.fac-table th { background: #193241; color: white; padding: 5px 8px; text-align: left; font-size: 9pt; }
table.fac-table td { padding: 4px 8px; border: 1px solid #E1E2E0; }
table.fac-table tr:nth-child(even) { background: #F8F8F8; }
.svc-bar { display: flex; height: 14px; border-radius: 3px; overflow: hidden; margin: 4px 0; }
.svc-seg { display: inline-block; height: 100%; }
.profile-link { display: inline-block; margin-top: 8px; padding: 5px 12px; border: 1px solid #877BE9; border-radius: 4px; color: #877BE9; font-size: 10pt; text-decoration: none; font-weight: 500; }
.profile-link:hover { background: #F3F1FC; }
</style>
</head>
<body>

<h1>MUO Breakout Detail</h1>
<p style="color: #507E8E; margin-bottom: 30px;">Detailed scoring and facility-level breakdown for each multi-unit operator.
<a href="BD_Complete_MUO_Universe_V23.html" class="back-link" style="margin-left: 12px;">&larr; Back to Tier View</a></p>
''')

# Build sorted list: T1, T2, T3 by score desc; T4 by camps desc; T5 alpha
all_ordered = []
for tier in ['T1', 'T2', 'T3']:
    ents = sorted(by_tier.get(tier, []), key=lambda x: (-(x[1].get('score') or 0), x[0]))
    all_ordered.extend(ents)
ents_t4 = sorted(by_tier.get('T4', []), key=lambda x: (-x[1].get('n_camp', 0), x[0]))
all_ordered.extend(ents_t4)
all_ordered.extend(sorted(t5_deduped, key=lambda x: x[0]))

for name, data in all_ordered:
    slug = slugify(name)
    tier = data.get('tier', '')
    score = data.get('score')
    camps = data.get('n_camp', 0)
    srv = data.get('n_srv', 0)
    states = data.get('states', '')
    i_rev = data.get('int_rev', 0)
    n_rev = data.get('new_rev', 0)
    t_rev = data.get('total_rev', 0)
    barrier = data.get('barrier', T5_BARRIERS.get(name, ''))
    facs = entity_facilities.get(name, [])
    qualifying = [f for f in facs if f['tot_beds'] > BED_MIN]

    # Tier badge color
    badge_colors = {
        'T1': ('background: #C6EFCE; color: #193241;', 'Strategic Enterprise'),
        'T2': ('background: #FFEB9C; color: #193241;', 'Growth & Expansion'),
        'T3': ('background: #FFC7CE; color: #193241;', 'Retention / Watch'),
        'T4': ('background: #E1E2E0; color: #193241;', 'Independent'),
        'T5': ('background: #F2DCDB; color: #193241;', 'Hard Barrier'),
    }
    badge_style, tier_label = badge_colors.get(tier, ('background: #eee;', ''))

    bo.append(f'<div class="entity-section" id="{slug}">')
    bo.append(f'<div class="entity-header">')
    bo.append(f'<div class="entity-title">{name}</div>')
    bo.append(f'<span class="tier-badge" style="{badge_style}">{tier} &mdash; {tier_label}</span>')
    if score is not None:
        bo.append(f'<span class="score-badge">Score: {score}</span>')
    bo.append(f'</div>')

    # Back link + profile link
    bo.append(f'<a href="BD_Complete_MUO_Universe_V23.html#section-{tier}" class="back-link">&larr; Back to {tier} table</a>')
    profile = find_profile(name)
    if profile:
        bo.append(f' &nbsp; <span class="profile-link" title="Full MUO profile available">Full Profile: {profile.replace("_MUO_Profile.md", "").replace("_", " ")}</span>')

    # ---- BUSINESS DESCRIPTION ----
    prof = parsed_profiles.get(profile, {}) if profile else {}
    if prof.get('key_insight'):
        bo.append(f'<div style="margin: 16px 0; padding: 14px 18px; background: #F5F7F8; border-left: 4px solid #877BE9; border-radius: 4px; color: #333; font-size: 10.5pt; line-height: 1.6;">')
        # Metrics row
        meta_parts = []
        if prof.get('hq'):
            meta_parts.append(f'<b>HQ:</b> {prof["hq"]}')
        if prof.get('founded'):
            meta_parts.append(f'<b>Founded:</b> {prof["founded"]}')
        if prof.get('ceo'):
            meta_parts.append(f'<b>CEO:</b> {prof["ceo"]}')
        if prof.get('ownership_type'):
            meta_parts.append(f'<b>Ownership:</b> {prof["ownership_type"]}')
        if prof.get('revenue'):
            meta_parts.append(f'<b>Revenue:</b> {prof["revenue"]}')
        if meta_parts:
            bo.append(f'<div style="font-size: 9.5pt; color: #507E8E; margin-bottom: 8px;">{" &nbsp;|&nbsp; ".join(meta_parts)}</div>')
        bo.append(f'{prof["key_insight"]}')
        bo.append(f'</div>')
    elif qualifying:
        # Auto-generate from DB data
        n_q = len(set(f['campus_key'] for f in qualifying))
        st_list = sorted(set(f['state'] for f in qualifying))
        types = sorted(set(f.get('src_type', '') for f in qualifying) - {''})
        type_str = ' and '.join(types) if types else 'senior care'
        bo.append(f'<div style="margin: 16px 0; padding: 14px 18px; background: #F5F7F8; border-left: 4px solid #E1E2E0; border-radius: 4px; color: #507E8E; font-size: 10.5pt; line-height: 1.6;">')
        bo.append(f'{name} operates {n_q} qualifying campuses ({type_str}) across {", ".join(st_list)} within the EWH operational footprint.')
        bo.append(f'</div>')

    # ---- LEADERSHIP ----
    if prof.get('leaders'):
        bo.append('<h3>Leadership</h3>')
        bo.append('<table class="dim-table">')
        bo.append('<tr><th>Name</th><th>Title</th><th>Notable</th></tr>')
        for ldr in prof['leaders']:
            bo.append(f'<tr class="active"><td style="font-weight: 500;">{ldr["name"]}</td><td>{ldr["title"]}</td><td style="color: #507E8E; font-size: 10pt;">{ldr["notable"]}</td></tr>')
        bo.append('</table>')

    if tier == 'T5':
        bo.append(f'<div style="margin: 16px 0; padding: 12px 16px; background: #FDE8E8; border-left: 4px solid #C62828; border-radius: 4px;">')
        bo.append(f'<b>Barrier:</b> {barrier}')
        bo.append(f'</div>')
        bo.append(f'<p style="color: #507E8E; font-size: 10.5pt;">Structural barriers prevent corporate-level growth. Guidance: defend current revenue, monitor barrier status annually.</p>')
        bo.append(f'</div>')
        continue

    # ---- SUMMARY CARDS ----
    i_str = f'${i_rev:,.0f}' if i_rev else '-'
    n_str = f'${n_rev:,.0f}' if n_rev else '-'
    t_str = f'${t_rev:,.0f}' if t_rev else '-'
    cur = sum(f.get('cur_rev', 0) for f in qualifying)
    cur_str = f'${cur:,.0f}' if cur else '-'

    bo.append('<div class="detail-grid">')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">Campuses</div><div class="detail-card-value">{camps}</div><div class="detail-card-sub">in 6-state footprint</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">Served</div><div class="detail-card-value">{srv}</div><div class="detail-card-sub">campuses with Eventus</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">States</div><div class="detail-card-value">{data.get("n_states", 0)}</div><div class="detail-card-sub">{states}</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">Current Revenue</div><div class="detail-card-value" style="font-size: 13pt;">{cur_str}</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">Integration Opp</div><div class="detail-card-value" style="font-size: 13pt;">{i_str}</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">New Business Opp</div><div class="detail-card-value" style="font-size: 13pt;">{n_str}</div></div>')
    bo.append(f'<div class="detail-card"><div class="detail-card-label">Total Opportunity</div><div class="detail-card-value" style="font-size: 13pt; color: #877BE9;">{t_str}</div></div>')
    bo.append('</div>')

    if tier == 'T4':
        reason = data.get('reason', f'{camps} qualifying campuses (gate requires 7)')
        bo.append(f'<div style="margin: 12px 0; padding: 10px 16px; background: #F5F5F5; border-left: 4px solid #E1E2E0; border-radius: 4px; color: #507E8E; font-size: 10.5pt;">')
        bo.append(f'<b>Gate:</b> {reason}. Entities need 7+ qualifying campuses to enter the scored tiers (T1-T3).')
        bo.append(f'</div>')

    # ---- SERVICE PENETRATION ----
    if qualifying:
        n_q = len(set(f['campus_key'] for f in qualifying))
        n_served = len(set(f['campus_key'] for f in qualifying if f['served']))
        n_not = n_q - n_served
        pct_served = (n_served / n_q * 100) if n_q > 0 else 0
        n_integ = len(set(f['campus_key'] for f in qualifying if f['integ']))
        n_pcp_only = len(set(f['campus_key'] for f in qualifying if f['pcp'] and not f['mh'] and not f['integ']))
        n_mh_only = len(set(f['campus_key'] for f in qualifying if f['mh'] and not f['pcp'] and not f['integ']))
        n_none = n_q - n_integ - n_pcp_only - n_mh_only

        bo.append('<h3>Service Penetration</h3>')
        bo.append('<div class="detail-grid">')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">We Serve</div><div class="detail-card-value" style="color: #2E7D32;">{n_served}</div><div class="detail-card-sub">{pct_served:.0f}% of campuses</div></div>')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">Don\'t Serve</div><div class="detail-card-value" style="color: #C62828;">{n_not}</div><div class="detail-card-sub">{100 - pct_served:.0f}% of campuses</div></div>')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">Integrated</div><div class="detail-card-value">{n_integ}</div><div class="detail-card-sub">PCP + MH</div></div>')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">PCP Only</div><div class="detail-card-value">{n_pcp_only}</div></div>')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">MH Only</div><div class="detail-card-value">{n_mh_only}</div></div>')
        bo.append(f'<div class="detail-card"><div class="detail-card-label">Not Served</div><div class="detail-card-value">{n_none}</div></div>')
        bo.append('</div>')

        # Service penetration bar
        seg_colors = [('#2E7D32', n_integ, 'Integrated'), ('#877BE9', n_pcp_only, 'PCP'), ('#91D9E8', n_mh_only, 'MH'), ('#E1E2E0', n_none, 'None')]
        bo.append('<div style="display: flex; align-items: center; gap: 10px; margin: 4px 0 12px 0;">')
        bo.append('<div class="svc-bar" style="flex: 1; max-width: 500px;">')
        for color, count, label in seg_colors:
            pct = (count / n_q * 100) if n_q > 0 else 0
            if pct > 0:
                bo.append(f'<div class="svc-seg" style="width: {pct:.1f}%; background: {color};" title="{label}: {count}"></div>')
        bo.append('</div>')
        bo.append('<span style="font-size: 9pt; color: #999;">')
        for color, count, label in seg_colors:
            if count > 0:
                bo.append(f'<span style="color: {color};">&#9632;</span> {label} ({count}) &nbsp;')
        bo.append('</span></div>')

    # ---- RELATIONSHIP NOTES ----
    rs_val = data.get('RS', 0) or 0
    rs_label = RS_BRACKETS.get(rs_val, '-')
    bo.append('<h3>Relationship Notes</h3>')
    bo.append(f'<div style="padding: 12px 18px; background: #F5F7F8; border-left: 4px solid #91D9E8; border-radius: 4px; margin-bottom: 12px;">')
    bo.append(f'<div style="font-size: 9.5pt; color: #507E8E; margin-bottom: 6px;"><b>RS Score:</b> {rs_val} &mdash; {rs_label}</div>')
    if prof.get('recommendation'):
        # Pull first paragraph of recommendation as relationship context
        rec_lines = prof['recommendation'].split('\n\n')
        rec_first = rec_lines[0].replace('**', '<b>').replace('**', '</b>') if rec_lines else ''
        bo.append(f'<div style="font-size: 10.5pt; color: #333; line-height: 1.5;">{rec_first}</div>')
    else:
        bo.append(f'<div style="font-size: 10pt; color: #aaa; font-style: italic;">No relationship notes on file. Update via MUO Profile or field intel.</div>')
    bo.append('</div>')

    # ---- LAST KNOWN QBR ----
    bo.append('<h3>Last Known QBR</h3>')
    bo.append(f'<div style="padding: 12px 18px; background: #F5F7F8; border-left: 4px solid #E1E2E0; border-radius: 4px;">')
    bo.append(f'<div style="font-size: 10pt; color: #aaa; font-style: italic;">No QBR data on file. To be updated following next quarterly business review.</div>')
    bo.append('</div>')

    # ---- SCORING DETAIL (scored tiers only) ----
    if tier in ('T1', 'T2', 'T3'):
        bo.append('<h3>Scoring Assessment</h3>')
        bo.append('<table class="dim-table">')
        bo.append('<tr><th style="width: 40px;">Dim</th><th style="width: 170px;">Dimension</th><th style="width: 50px;">Weight</th><th style="width: 50px;">Score</th><th style="width: 60px;">Weighted</th><th>Bracket</th></tr>')
        for abbr, full_name, weight, brackets in dim_meta:
            raw = data.get(abbr, 0) or 0
            weighted = raw * weight
            bracket_label = brackets.get(raw, '-')
            bo.append(f'<tr class="active"><td style="font-weight: bold;">{abbr}</td><td>{full_name}</td><td style="text-align: center;">&times;{weight}</td><td style="text-align: center;">{raw}</td><td style="text-align: center;">{weighted}</td><td style="color: #507E8E;">{bracket_label}</td></tr>')
        bo.append(f'<tr style="background: #193241; color: white; font-weight: bold;"><td colspan="4" style="text-align: right; border: none;">Total Score</td><td style="text-align: center; border: none;">{score}</td><td style="border: none;"></td></tr>')
        bo.append('</table>')

    # ---- STATE DISTRIBUTION ----
    if qualifying:
        state_data = defaultdict(lambda: {'camps': set(), 'served': set(), 'beds': 0, 'rev': 0, 'types': set()})
        for f in qualifying:
            sd = state_data[f['state']]
            sd['camps'].add(f['campus_key'])
            if f['served']:
                sd['served'].add(f['campus_key'])
            sd['beds'] += f['tot_beds']
            sd['rev'] += f['tot_rev']
            sd['types'].add(f.get('src_type', ''))

        bo.append('<h3>Geographic Distribution</h3>')
        bo.append('<table class="dim-table">')
        bo.append('<tr><th>State</th><th style="text-align: center;">Campuses</th><th style="text-align: center;">Served</th><th style="text-align: center;">Beds</th><th style="text-align: center;">Types</th><th style="text-align: right;">Opportunity</th></tr>')
        for st in sorted(state_data.keys()):
            sd = state_data[st]
            n_c = len(sd['camps'])
            n_s = len(sd['served'])
            types = ', '.join(sorted(sd['types'] - {''})) or '-'
            bo.append(f'<tr class="active"><td>{st}</td><td style="text-align: center;">{n_c}</td><td style="text-align: center;">{n_s}</td><td style="text-align: center;">{sd["beds"]:.0f}</td><td>{types}</td><td style="text-align: right;">${sd["rev"]:,.0f}</td></tr>')
        bo.append('</table>')

    # ---- FACILITY LIST ----
    if qualifying:
        bo.append('<h3>Facility List</h3>')
        bo.append('<table class="fac-table">')
        bo.append('<tr><th>Facility</th><th>City</th><th>ST</th><th style="text-align: center;">Beds</th><th style="text-align: center;">Serve</th><th>Config</th><th>Geo</th><th style="text-align: right;">Revenue</th></tr>')
        fac_sorted = sorted(qualifying, key=lambda f: (f['state'], f['city'], f['fac_name']))
        for f in fac_sorted:
            svc = 'Yes' if f['served'] else '-'
            svc_color = 'color: #2E7D32;' if f['served'] else 'color: #ccc;'
            if f['integ']:
                config = 'Integrated'
            elif f['pcp'] and f['mh']:
                config = 'PCP + MH'
            elif f['pcp']:
                config = 'PCP Only'
            elif f['mh']:
                config = 'MH Only'
            else:
                config = '-'
            geo = f.get('geo_tier', '') or '-'
            rev = f['tot_rev']
            rev_str = f'${rev:,.0f}' if rev else '-'
            bo.append(f'<tr><td>{f["fac_name"]}</td><td>{f["city"]}</td><td>{f["state"]}</td><td style="text-align: center;">{f["tot_beds"]:.0f}</td><td style="text-align: center; {svc_color} font-weight: bold;">{svc}</td><td>{config}</td><td>{geo}</td><td style="text-align: right;">{rev_str}</td></tr>')
        bo.append('</table>')

    bo.append('</div>')

bo.append('''
<hr style="border: 1px solid #E1E2E0; margin: 40px 0;">
<p style="font-size: 10pt; color: #aaa;">
<a href="BD_Complete_MUO_Universe_V23.html" class="back-link">&larr; Back to Tier View</a><br>
Generated March 2026
</p>
</body></html>''')

bo_content = '\n'.join(bo)
with open(BREAKOUT_OUT, 'w', encoding='utf-8') as f:
    f.write(bo_content)

print(f"Breakout saved to: {BREAKOUT_OUT}")

# Console summary
print(f"\n{'='*80}")
print(f"MUO SCORED BY TIERS — A MULTI-ATTRIBUTE ANALYSIS")
print(f"{'='*80}")
for tier in ['T1', 'T2', 'T3', 'T4', 'T5']:
    ents = by_tier.get(tier, [])
    if tier == 'T5':
        ents = t5_deduped
    print(f"  {tier}: {len(ents)}")
print(f"  Total: {total}")

new_count = sum(1 for name in all_entities if name not in bd_names)
new_t5 = sum(1 for name in t5_seen if name not in bd_names)
print(f"\n  NEW (added from Finance, not in prior BD): {new_count + new_t5}")
print(f"\nHTML saved to: {HTML_OUT}")
print("Done!")
