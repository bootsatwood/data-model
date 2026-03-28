"""
v25_5_migration.py
------------------
Applies all confirmed but unapplied dedup changes to V25.4, producing V25.5.

Changes:
  - 17 DELETES (14 from 50-dup exercise + 3 from sibling pair review)
  - 4 FIELD UPDATES (Legacy beds+served, Robin Run ALF beds, Robin Run SNF beds, Robin Run Village retype)
  - 1 INSERT (Robin Run IL)

Source of truth:
  - dedup_decisions_log.csv (pairs 5-8)
  - dedup_progress memory (50-dup confirmed deletes)
  - data_remediation_progress memory (field fixes)
"""

import openpyxl
import os
from copy import copy

# --- Paths ---
VAULT = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - Eventus WholeHealth", "Vault", "02_Data_Model", "Current"
)
INPUT_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_4.xlsx")
OUTPUT_PATH = os.path.join(VAULT, "1_Combined_Database_FINAL_V25_5.xlsx")

# --- DELETES: match by (Facility_Name, Address, City, State) ---
DELETES = [
    # 14 from 50-dup exercise (still in V25.4)
    ("MEADOWVIEW TERRACE OF WADESBORO", "ALG SENIOR"),
    ("THE GARDENS OF HENDERSONVILLE", "ALG SENIOR"),
    ("THE LANDINGS OF OAK ISLAND", "ALG SENIOR"),
    ("CLEVELAND HOUSE", "ALG SENIOR"),
    ("AHOSKIE HOUSE", "ALG SENIOR"),
    ("WINDSOR HOUSE", "ALG SENIOR"),
    ("MOCKSVILLE SENIOR LIVING AND MEMORY CARE", "ALG SENIOR"),
    ("BROOKDALE CONCORD PARKWAY", "BROOKDALE"),
    ("BROOKDALE UNION PARK", "BROOKDALE"),
    ("BROOKDALE REIDSVILLE", "BROOKDALE"),
    ("BROOKDALE BRISTOL", "BROOKDALE"),
    ("BROOKDALE WAKE FOREST", "BROOKDALE"),
    ("BROOKDALE WOOSTER", "BROOKDALE"),
    ("SILVER OAKS HEALTH CAMPUS", "TRILOGY"),
    # 3 from sibling pair review (pairs 5-7)
    ("WILDWOOD HEALTHCARE CENTER SNF", "WILDWOOD HEALTHCARE"),
    ("HELLENIC SENIOR LIVING OF MISHAWAKA", "HELLENIC SENIOR LIVING"),
    ("PAOLI HEALTH AND LIVING COMMUNITY", "CARDON & ASSOCIATES"),
]

# --- FIELD UPDATES: match by Facility_Name, apply changes ---
FIELD_UPDATES = [
    {
        "match_name": "THE LEGACY AT NORTH AUGUSTA, INC.",
        "changes": {"Total_Beds": 71, "Do_We_Serve": "Yes"},
    },
    {
        "match_name": "ROBIN RUN ALF",
        "changes": {"Total_Beds": 94},
    },
    {
        "match_name": "ROBIN RUN SNF",
        "changes": {"Total_Beds": 84},
    },
    {
        "match_name": "ROBIN RUN VILLAGE",
        "changes": {"Source_Type": "ILF"},
    },
]

# --- INSERT: Robin Run IL ---
# Will be appended with values matching the column order of V25.4
ROBIN_RUN_IL = {
    "Source_Type": "ILF",
    "Facility_Name": "ROBIN RUN IL",
    "Corporate_Name": "OAKDALE SENIOR ALLIANCE",
    "Address": "5354 W 62nd St",
    "City": "Indianapolis",
    "State": "IN",
    "ZIP": "46268",
    "County": "Marion",
    "Ownership_Type": "Corporate",
    "Total_Beds": 150,
    "Census": 118,
    "Do_We_Serve": "No",
    "Integrated_Flag": None,
    "PCP_Flag": None,
    "MH_Flag": "No",
    "Barrier": None,
    "Latitude": None,
    "Longitude": None,
    "Data_Quality_Flag": None,
    "Contract_Status": None,
    "Geographic_Tier": "A_Metro",
    "Original_Geographic_Tier": "A_Metro",
    "Metro_Assignment": "Indianapolis",
    "Distance_to_Metro_Center": None,
    "Metro_Center_Used": None,
    "Corp_Attribution_Source": "GLR",
    "Campus_ID": None,
}


def main():
    print(f"Reading: {INPUT_PATH}")
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active

    # Build header index
    header = [cell.value for cell in ws[1]]
    col_idx = {name: i for i, name in enumerate(header)}
    name_col = col_idx["Facility_Name"] + 1  # 1-indexed for openpyxl
    corp_col = col_idx["Corporate_Name"] + 1

    total_before = ws.max_row - 1
    print(f"Rows before: {total_before}")

    # --- DELETES ---
    delete_names = {d[0] for d in DELETES}
    delete_lookup = {d[0]: d[1] for d in DELETES}
    rows_to_delete = []

    for row_num in range(2, ws.max_row + 1):
        fname = ws.cell(row=row_num, column=name_col).value
        corp = ws.cell(row=row_num, column=corp_col).value
        if fname and fname.strip() in delete_names:
            fname_clean = fname.strip()
            expected_corp = delete_lookup[fname_clean]
            corp_clean = (corp or "").strip().upper()
            if expected_corp.upper() in corp_clean or corp_clean in expected_corp.upper():
                rows_to_delete.append((row_num, fname_clean))

    # Delete from bottom up to preserve row numbers
    rows_to_delete.sort(key=lambda x: x[0], reverse=True)
    deleted_count = 0
    for row_num, fname in rows_to_delete:
        ws.delete_rows(row_num)
        deleted_count += 1
        print(f"  DELETE row {row_num}: {fname}")

    print(f"Total deleted: {deleted_count}")

    if deleted_count != 17:
        print(f"  WARNING: Expected 17 deletes, got {deleted_count}. Verify manually.")

    # --- FIELD UPDATES ---
    update_count = 0
    for row_num in range(2, ws.max_row + 1):
        fname = ws.cell(row=row_num, column=name_col).value
        if not fname:
            continue
        fname_clean = fname.strip()
        for update in FIELD_UPDATES:
            if fname_clean == update["match_name"]:
                for field, new_val in update["changes"].items():
                    col = col_idx[field] + 1
                    old_val = ws.cell(row=row_num, column=col).value
                    ws.cell(row=row_num, column=col).value = new_val
                    print(f"  UPDATE {fname_clean}: {field} {old_val} -> {new_val}")
                update_count += 1

    print(f"Total updated: {update_count}")

    # --- INSERT ---
    new_row = ws.max_row + 1
    for col_name, value in ROBIN_RUN_IL.items():
        col = col_idx[col_name] + 1
        ws.cell(row=new_row, column=col).value = value
    print(f"  INSERT row {new_row}: ROBIN RUN IL (ILF, 150 beds)")

    total_after = ws.max_row - 1
    print(f"\nRows after: {total_after}")
    print(f"Delta: {total_after - total_before} (expected: -17 deletes + 1 insert = -16)")

    # --- SAVE ---
    print(f"\nSaving: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
