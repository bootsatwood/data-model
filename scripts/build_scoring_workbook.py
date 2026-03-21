"""Build the MUO Scoring Workbook — one tab per scored dimension.
Each tab: MUO entity header row (bold) with rollup score,
then facility detail rows showing how we arrive at the number.
Final tab: Summary with all 6 scores and tier."""

import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DB_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx"
MUO_PATH = "C:/Users/ratwood/Downloads/MUO Data (1).xlsx"
TIERING_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/Final_MUO_Tiering_V20.xlsx"
OUT_PATH = "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/MUO_Scoring_Workbook_V23_v8.xlsx"
FOOTPRINT = {'NC', 'SC', 'VA', 'KY', 'OH', 'IN'}
BED_MIN = 15  # V8: Exclude facilities with <=15 beds from ER campus counting and RP revenue

# --- Exclusions ---
# T5 barriers are excluded from scoring entirely
T5_EXCLUDE = {
    # T5 - Hard Barriers (original)
    'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
    # T5 - Barriers from DB flags (V4)
    'CARDON & ASSOCIATES',          # 91% Permanent - Cardon
    'Eastern Healthcare Group',     # 100% Termination Risk
    'CLEARVIEW',                    # 38% Alliance + Own Provider Group, partnering with Telos
    'Pavilion Healthcare',          # 29% Alliance + Own Provider Group
}
# T4 entities pass through the pipeline — captured by <7 campus gate, listed in T4 tab

# --- Finance -> DB name mapping ---
FINANCE_TO_DB = {
    'ALG': {'ALG', 'ALG SENIOR'},
    'AMERICAN SENIOR COMMUNITIES': {'AMERICAN SENIOR COMMUNITIES'},
    'Brookdale Senior Living': {'BROOKDALE SENIOR LIVING'},
    'SABER HEALTHCARE GROUP': {'SABER HEALTHCARE GROUP', 'SABER HEALTHCARE'},
    'INFINITY HEALTHCARE CONSULTING': {'INFINITY HEALTHCARE CONSULTING'},
    'NAVION': {'NAVION', 'NAVION SENIOR SOLUTIONS', 'NAVION SENIOR LIVING', 'Navion Senior Solutions'},
    'Majestic Care': {'MAJESTIC CARE'},
    'PRUITT HEALTH': {'PRUITT HEALTH', 'PRUITTHEALTH'},
    'Kisco Senior Living': {'KISCO SENIOR LIVING'},
    'TRILOGY': {'TRILOGY', 'TRILOGY HEALTH SERVICES'},
    'Pavilion Healthcare': {'PAVILION HEALTHCARE'},
    'TOPAZ HEALTHCARE': {'TOPAZ HEALTHCARE'},
    'MORNING POINTE SENIOR LIVING': {'MORNING POINTE SENIOR LIVING', 'MORNING POINTE'},
    'PRINCIPLE': {'PRINCIPLE', 'PRINCIPLE LONG TERM CARE'},
    'Liberty': {'LIBERTY'},
    'TERRABELLA SENIOR LIVING': {'TERRABELLA SENIOR LIVING', 'TERRABELLA'},
    'SANSTONE': {'SANSTONE', 'SANSTONE HEALTH & REHABILITATION'},
    'LIONSTONE CARE': {'LIONSTONE CARE', 'LIONSTONE'},
    'ELDERCARE PARTNERS': {'ELDERCARE PARTNERS'},
    'OTTERBEIN SENIOR LIFE': {'OTTERBEIN SENIOR LIFE', 'OTTERBEIN'},
    'TLC Management': {'TLC MANAGEMENT'},
    'BHI Senior Living': {'BHI SENIOR LIVING'},
    'Avardis': {'AVARDIS', 'CONSULATE HEALTH CARE/INDEPENDENCE LIVING CENTERS/NSPIRE HEALTHCARE/RAYDIANT HEALTH CARE'},
    'PEAK RESOURCES': {'PEAK RESOURCES'},
    'ARBORS': {'ARBORS', 'ARBORS AT OHIO'},
    'HERITAGE HALL': {'HERITAGE HALL', 'AMERICAN HEALTHCARE, LLC', 'AHC'},
    'Runk & Pratt': {'RUNK & PRATT'},
    'MCAP': {'MCAP'},
    'Lutheran Services Carolinas': {'LUTHERAN SERVICES CAROLINAS', 'LUTHERAN SERVICES CAROLINA'},
    'Lutheran Life Villages': {'LUTHERAN LIFE VILLAGES', 'LUTHERAN LIFE COMMUNITIES'},
    'Greencroft': {'GREENCROFT'},
    'CCH HEALTHCARE': {'CCH HEALTHCARE'},
    'LifeSpire of Virginia': {'LIFESPIRE OF VIRGINIA'},
    'SENIOR LIFESTYLE': {'SENIOR LIFESTYLE'},
    'CLEARVIEW': {'CLEARVIEW'},
    'PRIORITY': {'PRIORITY LIFE CARE'},
    'CEDARHURST SENOR LIVING': {'CEDARHURST SENIOR LIVING'},
    'Lifecare': {'LIFE CARE CENTERS OF AMERICA'},
    'Kissito Healthcare': {'KISSITO HEALTHCARE', 'KISSITO'},
    'SPRING ARBOR MANAGEMENT': {'SPRING ARBOR MANAGEMENT'},
    'YAD': {'YAD', 'YAD HEALTHCARE'},
    'STORYPOINT': {'STORYPOINT'},
    'CARING PLACE HEALTHCARE': {'CARING PLACE HEALTHCARE', 'CARING PLACE'},
    'Eastern Healthcare Group': {'EASTERN HEALTHCARE GROUP'},
    'SONIDA SENIOR LIVING': {'SONIDA SENIOR LIVING'},
    'Castle Healthcare': {'CASTLE HEALTHCARE'},
    'JAG': {'JAG HEALTHCARE'},
    'FUNDAMENTAL LTC': {'FUNDAMENTAL HEALTHCARE', 'FUNDAMENTAL LTC'},
    'CARDON & ASSOCIATES': {'CARDON & ASSOCIATES'},
    'Southern Healthcare Mgmt': {'SOUTHERN HEALTHCARE MANAGEMENT, LLC'},
    'Triple Crown': {'TRIPLE CROWN', 'TRIPLE CROWN SENIOR LIVING', 'VITALITY SENIOR SERVICES'},
    # T4 entities (below MUO gate but tracked by Finance)
    'Sunnyside Communities': {'SUNNYSIDE COMMUNITIES', 'SUNNYSIDE'},
    'ATRIUM HEALTH': {'ATRIUM HEALTH'},
    'Warm Hearth Village': {'WARM HEARTH VILLAGE', 'WARM HEARTH'},
    'Brighton': {'BRIGHTON'},
    'Momentus Health': {'MOMENTUS HEALTH'},
}

# Finance -> V20 workbook name
FINANCE_TO_V20 = {
    'ALG': 'ALG',
    'AMERICAN SENIOR COMMUNITIES': 'American Senior Communities',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'SABER HEALTHCARE GROUP': 'Saber Healthcare Group',
    'INFINITY HEALTHCARE CONSULTING': 'Infinity Healthcare Consulting',
    'NAVION': 'Navion',
    'Majestic Care': 'Majestic Care',
    'PRUITT HEALTH': 'Pruitthealth',
    'TRILOGY': 'Trilogy Health Services',
    'Pavilion Healthcare': 'Pavilion Healthcare',
    'TOPAZ HEALTHCARE': 'Topaz',
    'MORNING POINTE SENIOR LIVING': 'Morning Pointe Senior Living',
    'PRINCIPLE': 'Principle Long Term Care',
    'Liberty': 'Liberty Senior Living',
    'TERRABELLA SENIOR LIVING': 'Terra Bella',
    'SANSTONE': 'Sanstone Health & Rehabilitation',
    'LIONSTONE CARE': 'Lionstone Care',
    'OTTERBEIN SENIOR LIFE': 'Otterbein Seniorlife',
    'TLC Management': 'Tlc Management',
    'BHI Senior Living': 'BHI Senior Living',
    'Avardis': 'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care',
    'PEAK RESOURCES': 'Peak Resources, Inc.',
    'ARBORS': 'Arbors At Ohio',
    'Lutheran Services Carolinas': 'Lutheran Services Carolina',
    'CCH HEALTHCARE': 'Cch Healthcare',
    'CLEARVIEW': 'Clearview',
    'PRIORITY': 'Priority Life Care',
    'Lifecare': 'Life Care Centers Of America',
    'Kissito Healthcare': 'Kissito',
    'YAD': 'Yad Healthcare',
    'CARING PLACE HEALTHCARE': 'Caring Place Healthcare',
    'Castle Healthcare': 'Castle Healthcare',
    'JAG': 'Jag Healthcare',
}

# --- Bracket definitions ---
ER_THRESHOLDS = [9, 13, 20, 40]  # Option C
RP_THRESHOLDS = [1_000_000, 2_500_000, 5_000_000, 10_000_000]  # Option B

def score_dim(val, thresholds):
    for i, t in enumerate(thresholds):
        if val < t:
            return i + 1
    return 5

def rs_score_served(srv, total):
    """V1/V2 approach — served campus ratio. Kept for reference."""
    if total == 0: return 1
    pct = srv / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if srv > 0: return 2
    return 1

# RS from Tom's Shared Savings Notes (Col W, MUO Data Analysis sheet)
# 5 = Active SS agreement on Eventus paper, strong partnership
# 4 = Active engagement, proposals/meetings in play, verbal commitments
# 3 = Have relationship but stalled, mixed signals, early stage
# 2 = Tried and hit resistance, rejected proposals, defensive posture
# 1 = "Don't really know well" or no meaningful relationship
TOM_RS_SCORES = {
    'ALG':                              2,  # Presented SS, Charlie not interested
    'AMERICAN SENIOR COMMUNITIES':      5,  # Have it on our paper, they're happy
    'Brookdale Senior Living':          5,  # Have on our paper
    'SABER HEALTHCARE GROUP':           2,  # Not very focused, not working
    'INFINITY HEALTHCARE CONSULTING':   2,  # Getting nowhere
    'NAVION':                           3,  # Good to start talking, growing
    'Majestic Care':                    3,  # Expressed interest but hasn't gone anywhere
    'PRUITT HEALTH':                    3,  # Getting more PC, good idea to approach, but focused on own ISNP
    'Kisco Senior Living':             1,  # Don't really know well
    'TRILOGY':                          4,  # Already covered, speaking with Dr. McNamara
    'Pavilion Healthcare':             1,  # Don't really know well
    'TOPAZ HEALTHCARE':                1,  # Don't really know well
    'MORNING POINTE SENIOR LIVING':     2,  # Not a great fit, don't recommend doing anything
    'PRINCIPLE':                        3,  # Meeting scheduled on the 17th
    'Liberty':                          4,  # Gave them proposal in RFP, focused on Liberty Advantage ISNP
    'TERRABELLA SENIOR LIVING':         2,  # Would have to drive growth, shouldn't just reward
    'SANSTONE':                         2,  # Presented SS, canceled meetings, not ready to commit
    'LIONSTONE CARE':                   3,  # Tom needs to talk to Kim B (action pending)
    'ELDERCARE PARTNERS':              3,  # Good relationship with them, out of KY
    'OTTERBEIN SENIOR LIFE':           1,  # Don't really know well
    'TLC Management':                   3,  # Part of Hoosier Alliance, not happy with ISNP, active contact
    'BHI Senior Living':               1,  # All psych
    'Avardis':                          2,  # Don't really know well, they do not want SS
    'PEAK RESOURCES':                   2,  # Not responsive or interested in SS
    'ARBORS':                          1,  # Don't really know well
    'HERITAGE HALL':          3,  # Probably need to do something, Tom to talk to Colvin
    'Runk & Pratt':                    1,  # Don't really know well
    'MCAP':                            1,  # Don't really know well
    'Lutheran Services Carolinas':      4,  # Verbal commitment for SS, part of how we won business
    'Lutheran Life Villages':           4,  # Cathy worked with Mike to put them in SS
    'Greencroft':                      1,  # Don't really know well
    'CCH HEALTHCARE':                   2,  # Talk to them defensively
    'LifeSpire of Virginia':           1,  # Don't really know well
    'SENIOR LIFESTYLE':                1,  # Don't really know well
    'CLEARVIEW':                        3,  # In process of partnering with Telos
    'PRIORITY':                        1,  # Don't really know well
    'CEDARHURST SENOR LIVING':         1,  # Don't really know well
    'Lifecare':                        1,  # Don't really know well
    'Kissito Healthcare':               2,  # COO not interested in SS, just stay close
    'SPRING ARBOR MANAGEMENT':         1,  # Don't really know well
    'YAD':                              2,  # Tried but never gotten an audience, defensive
    'STORYPOINT':                       3,  # Part of a growth plan
    'CARING PLACE HEALTHCARE':          2,  # Psych only, won't do PC, have their own
    'Eastern Healthcare Group':        1,  # Don't really know well
    'SONIDA SENIOR LIVING':            1,  # Growth opp, not doing much
    'Castle Healthcare':                3,  # Brooke and Ian have been more involved
    'JAG':                             2,  # Not much growth opp
    'FUNDAMENTAL LTC':                 1,  # Psych only
    'CARDON & ASSOCIATES':             2,  # Creating own group, Hoosier Alliance, not going to get anything
    'Southern Healthcare Mgmt':         1,  # Not in MUO Data notes
    'Triple Crown':                     1,  # Not in MUO Data notes
}

# SI from Tom's Notes + Web Research (market position, payer/ACO/REIT/ISNP connections)
# 5 = REIT-backed, runs ISNP, publicly traded, enterprise-scale
# 4 = Payer/ACO connected, multi-state enterprise, developing ISNP
# 3 = Regional reputation, established brand, expanding
# 2 = Minimal local presence, small/niche operator
# 1 = No affiliations, no market presence
DATA_SI_SCORES = {
    # SI=5 — REIT/ISNP/Enterprise
    'ALG':                              5,  # LTC Properties REIT relationship
    'Brookdale Senior Living':          5,  # NYSE: BKD, Welltower/DHC REIT partnerships
    'SABER HEALTHCARE GROUP':           5,  # Omega Healthcare 49% JV
    'TRILOGY':                          5,  # 100% owned by American Healthcare REIT
    'Liberty':                          5,  # Runs Liberty Advantage ISNP
    'PRUITT HEALTH':                    5,  # Runs PruittHealth Premier ISNP
    'Avardis':                          5,  # Omega Healthcare REIT lease ($3.1M/mo)
    'SONIDA SENIOR LIVING':             5,  # NYSE: SNDA, $1.8B merger with CNL Healthcare
    'Lifecare':                         5,  # Largest privately held, 200+ facilities

    # SI=4 — Payer/ACO connected, multi-state enterprise
    'AMERICAN SENIOR COMMUNITIES':      4,  # SS agreement on Eventus paper, ACO participation
    'INFINITY HEALTHCARE CONSULTING':   4,  # JV with Longevity Health for I-SNP expansion
    'Majestic Care':                    4,  # Developing own ISNP
    'NAVION':                           4,  # NHI REIT relationship + Blackstone partnership
    'Kisco Senior Living':              4,  # Private, 32 communities in 10 states + DC
    'TLC Management':                   4,  # LTC ACO agreement (Hoosier Alliance — NOTE: barrier concern)
    'Lutheran Life Villages':           4,  # LTC ACO agreement
    'Lutheran Services Carolinas':      4,  # ACO REACH with PPHP, verbal SS commitment

    # SI=3 — Regional reputation
    'SANSTONE':                         3,  # 20 locations across NC, growing, building new
    'MORNING POINTE SENIOR LIVING':     3,  # Regional brand TN/KY
    'OTTERBEIN SENIOR LIFE':            3,  # Ohio nonprofit, 11 communities, S&P A rating
    'PRINCIPLE':                        3,  # Regional presence
    'LIONSTONE CARE':                   3,  # Regional presence
    'Kissito Healthcare':               3,  # Acquired, regional presence
    'CCH HEALTHCARE':                   3,  # Regional operator
    'TERRABELLA SENIOR LIVING':         3,  # Regional NC
    'PEAK RESOURCES':                   3,  # NC regional
    'BHI Senior Living':                3,  # Faith-based nonprofit, 12 communities, BBB+ Fitch

    # SI=2 — Minimal local
    'TOPAZ HEALTHCARE':                 2,
    'ELDERCARE PARTNERS':               2,  # KY based
    'ARBORS':                           2,  # Ohio focused
    'HERITAGE HALL':          2,  # Heritage Hall buildings, VA
    'Runk & Pratt':                     2,  # VA
    'MCAP':                             2,
    'Greencroft':                       2,  # Indiana nonprofit
    'LifeSpire of Virginia':            2,  # VA
    'SENIOR LIFESTYLE':                 2,
    'PRIORITY':                         2,
    'CEDARHURST SENOR LIVING':          2,
    'SPRING ARBOR MANAGEMENT':          2,
    'STORYPOINT':                       2,  # Growth plan per Tom
    'CARING PLACE HEALTHCARE':          2,  # Psych only, own primary care
    'Castle Healthcare':                2,
    'JAG':                              2,  # Ohio, small
    'FUNDAMENTAL LTC':                  2,  # Psych only
    'YAD':                              2,  # Defensive posture
    'Southern Healthcare Mgmt':         2,
    'Triple Crown':                     2,
}

def normalize_address(addr):
    """Normalize common street abbreviations for address collapsing."""
    s = addr.upper().strip().rstrip('.')
    s = re.sub(r'\bSTREET\b', 'ST', s)
    s = re.sub(r'\bDRIVE\b', 'DR', s)
    s = re.sub(r'\bROAD\b', 'RD', s)
    s = re.sub(r'\bAVENUE\b', 'AVE', s)
    s = re.sub(r'\bBOULEVARD\b', 'BLVD', s)
    s = re.sub(r'\bLANE\b', 'LN', s)
    s = re.sub(r'\bCOURT\b', 'CT', s)
    s = re.sub(r'\bCIRCLE\b', 'CIR', s)
    s = re.sub(r'\bPLACE\b', 'PL', s)
    s = re.sub(r'\bTERRACE\b', 'TER', s)
    s = re.sub(r'\bPARKWAY\b', 'PKWY', s)
    s = re.sub(r'\bHIGHWAY\b', 'HWY', s)
    s = re.sub(r'\bNORTH\b', 'N', s)
    s = re.sub(r'\bSOUTH\b', 'S', s)
    s = re.sub(r'\bEAST\b', 'E', s)
    s = re.sub(r'\bWEST\b', 'W', s)
    s = re.sub(r'\.', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def extract_street_number(addr):
    """Pull leading street number from address."""
    m = re.match(r'^(\d+)', addr.strip())
    return m.group(1) if m else None

def er_bracket_label(score):
    labels = {1: '<9', 2: '9-12', 3: '13-19', 4: '20-39', 5: '40+'}
    return labels.get(score, '?')

def rp_bracket_label(score):
    labels = {1: '<$1M', 2: '$1M-$2.5M', 3: '$2.5M-$5M', 4: '$5M-$10M', 5: '>$10M'}
    return labels.get(score, '?')

def rs_bracket_label(score):
    labels = {1: '0%', 2: '1-24%', 3: '25-49%', 4: '50-79%', 5: '80-100%'}
    return labels.get(score, '?')

# --- Styles ---
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT_W = Font(bold=True, size=11, color='FFFFFF')
ENTITY_FILL = PatternFill('solid', fgColor='D9E2F3')
ENTITY_FONT = Font(bold=True, size=10)
DETAIL_FONT = Font(size=9, color='404040')
SCORE_FONT = Font(bold=True, size=11, color='1F4E79')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='B4C6E7')
)
MONEY_FMT = '#,##0'
PCT_FMT = '0%'

def style_header(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER_FONT_W
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_entity_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row, c)
        cell.fill = ENTITY_FILL
        cell.font = ENTITY_FONT

def style_detail_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row, c)
        cell.font = DETAIL_FONT

# ============================================================
# LOAD DATA
# ============================================================
print("Loading V23 DB...")
wb_db = openpyxl.load_workbook(DB_PATH, data_only=True)
ws_db = wb_db.active
db_headers = {ws_db.cell(1, c).value: c for c in range(1, ws_db.max_column + 1) if ws_db.cell(1, c).value}

print("Loading V20 Tiering workbook...")
wb_t = openpyxl.load_workbook(TIERING_PATH, data_only=True)
v20_scores = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws_t = wb_t[sn]
    for r in range(2, ws_t.max_row + 1):
        name = ws_t.cell(r, 1).value
        if not name:
            continue
        v20_scores[name] = {
            'er': ws_t.cell(r, 9).value or 0,
            'ir': ws_t.cell(r, 10).value or 0,
            'si': ws_t.cell(r, 11).value or 0,
            'rp': ws_t.cell(r, 12).value or 0,
            'rs': ws_t.cell(r, 13).value or 0,
            'ai': ws_t.cell(r, 14).value or 0,
            'total': ws_t.cell(r, 3).value or 0,
        }

print("Loading Finance entity list...")
wb_muo = openpyxl.load_workbook(MUO_PATH, data_only=True)
ws_a = wb_muo['Analysis']
finance_entities = []
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    tier = ws_a.cell(r, 18).value
    if name and str(name).strip() not in ('TOTAL TOP 60',) and str(name).strip() not in T5_EXCLUDE:
        finance_entities.append({'name': str(name).strip(), 'fin_tier': tier})

# --- Build reverse lookup: DB corp name -> finance name ---
all_db_names = set()
db_to_finance = {}
for fname, db_names in FINANCE_TO_DB.items():
    for dn in db_names:
        all_db_names.add(dn)
        db_to_finance[dn] = fname

# --- Collect ALL facility rows per entity (footprint only) ---
print("Scanning DB for facility rows...")
entity_facilities = defaultdict(list)

for r in range(2, ws_db.max_row + 1):
    corp = str(ws_db.cell(r, db_headers['Corporate_Name']).value or '').strip().upper()
    state = str(ws_db.cell(r, db_headers['State']).value or '').strip()
    if corp not in all_db_names or state not in FOOTPRINT:
        continue

    fname = db_to_finance[corp]
    # Only include if this finance entity is in our scoreable list
    if fname not in [fe['name'] for fe in finance_entities]:
        continue

    fac_name = str(ws_db.cell(r, db_headers['Facility_Name']).value or '').strip()
    addr = str(ws_db.cell(r, db_headers['Address']).value or '').strip()
    city = str(ws_db.cell(r, db_headers['City']).value or '').strip()
    served = str(ws_db.cell(r, db_headers['Do_We_Serve']).value or '').strip().upper() == 'YES'
    mh = str(ws_db.cell(r, db_headers['MH_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    pcp = str(ws_db.cell(r, db_headers['PCP_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    integ = str(ws_db.cell(r, db_headers['Integrated_Flag']).value or '').strip().upper() in ('YES', 'TRUE', '1')
    census = ws_db.cell(r, db_headers['Census']).value or 0
    tot_beds = ws_db.cell(r, db_headers['Total_Beds']).value or 0
    cur_rev = ws_db.cell(r, db_headers['Current_Revenue']).value or 0
    int_rev = ws_db.cell(r, db_headers['Integration_Revenue']).value or 0
    nb_rev = ws_db.cell(r, db_headers['New_Business_Revenue']).value or 0
    tot_rev = ws_db.cell(r, db_headers['Total_Potential_Revenue']).value or 0
    source = str(ws_db.cell(r, db_headers['Source_Type']).value or '').strip()

    norm_addr = normalize_address(addr)
    street_num = extract_street_number(addr)
    beds_val = tot_beds if isinstance(tot_beds, (int, float)) else 0
    # V8: street-number matching — (street_number, city) as campus key
    # Falls back to (normalized_address, city) if no leading number
    num_key = (street_num, city.upper()) if street_num else (norm_addr, city.upper())

    entity_facilities[fname].append({
        'fac_name': fac_name,
        'addr': addr,
        'city': city,
        'state': state,
        'campus_key': num_key,
        'norm_addr': norm_addr,
        'street_num': street_num,
        'served': served,
        'mh': mh,
        'pcp': pcp,
        'integ': integ,
        'census': census if isinstance(census, (int, float)) else 0,
        'tot_beds': beds_val,
        'cur_rev': cur_rev if isinstance(cur_rev, (int, float)) else 0,
        'int_rev': int_rev if isinstance(int_rev, (int, float)) else 0,
        'nb_rev': nb_rev if isinstance(nb_rev, (int, float)) else 0,
        'tot_rev': tot_rev if isinstance(tot_rev, (int, float)) else 0,
        'source': source,
        'db_corp': corp,
    })

# --- Compute entity-level rollups ---
print("Computing entity rollups...")
entity_rollups = {}
t4_rollups = {}
for fe in finance_entities:
    fname = fe['name']
    facs = entity_facilities.get(fname, [])
    if not facs:
        # No DB rows — record as T4 with zero data
        t4_rollups[fname] = {
            'fin_tier': fe['fin_tier'], 'n_camp': 0, 'n_srv': 0,
            'total_rev': 0, 'states': '(no DB match)', 'facs': [],
        }
        continue

    # V8: Apply 15-bed minimum for ER campus counting and RP revenue
    qualifying = [f for f in facs if f['tot_beds'] > BED_MIN]
    campuses = set(f['campus_key'] for f in qualifying)
    served_camps = set(f['campus_key'] for f in qualifying if f['served'])
    mh_camps = set(f['campus_key'] for f in qualifying if f['mh'])
    pcp_camps = set(f['campus_key'] for f in qualifying if f['pcp'])
    integ_camps = set(f['campus_key'] for f in qualifying if f['integ'])
    total_rev = sum(f['tot_rev'] for f in qualifying)

    n_camp = len(campuses)
    n_srv = len(served_camps)

    if n_camp < 7:
        # T4 - Independent (fails MUO gate)
        states = sorted(set(f['state'] for f in qualifying))
        t4_rollups[fname] = {
            'fin_tier': fe['fin_tier'], 'n_camp': n_camp, 'n_srv': n_srv,
            'total_rev': total_rev, 'states': ', '.join(states), 'facs': facs,
            'excluded_beds': [f for f in facs if f['tot_beds'] <= BED_MIN],
        }
        continue

    # ER
    er = score_dim(n_camp, ER_THRESHOLDS)

    # RP
    rp = score_dim(total_rev, RP_THRESHOLDS)

    # RS — from Tom's Shared Savings Notes
    rs_served = rs_score_served(n_srv, n_camp)  # keep for comparison
    rs = TOM_RS_SCORES.get(fname, 1)
    rs_source = f"Tom's Notes (served ratio was {rs_served})"

    # V20 qualitative scores for AI only; SI now data-driven for all
    v20_name = FINANCE_TO_V20.get(fname)
    v20 = v20_scores.get(v20_name) if v20_name else None

    # SI — ALL entities use data-driven scores (Tom's Notes + web research)
    si = DATA_SI_SCORES.get(fname, 2)
    v20_si = v20['si'] if v20 else None
    si_source = f"Data-driven (was V20={v20_si})" if v20_si is not None else "Data-driven"

    # AI — still from V20 where available, default=0 otherwise
    if v20:
        ai = v20['ai']
        ai_source = 'V20 Brooke'
        v20_ir = v20['ir']
    else:
        ai = 0
        ai_source = 'Default (no V20 score)'
        v20_ir = None

    # IR — ALL entities use formulaic approach from service flags
    n_integ = len(integ_camps)
    n_mh = len(mh_camps)
    n_pcp = len(pcp_camps)
    dual = len(mh_camps & pcp_camps)
    int_pct = n_integ / n_camp if n_camp > 0 else 0

    if n_srv == 0:
        ir = 1
        ir_source = 'Data: not served'
    elif int_pct >= 0.5:
        ir = 5
        ir_source = f'Data: {n_integ}/{n_camp} integrated'
    elif int_pct >= 0.25:
        ir = 4
        ir_source = f'Data: {n_integ}/{n_camp} integrated'
    elif dual > 0:
        ir = 4
        ir_source = f'Data: {dual} dual MH+PCP'
    elif n_mh > 0 and n_pcp > 0:
        ir = 3
        ir_source = f'Data: MH({n_mh})+PCP({n_pcp}) separate'
    elif n_mh > 0 or n_pcp > 0:
        ir = 2
        ir_source = f'Data: single svc ({n_mh}MH/{n_pcp}PCP)'
    else:
        ir = 1
        ir_source = 'Data: served, no svc flags'

    # Track V20 comparison
    if v20_ir is not None:
        ir_source += f' (was V20={v20_ir})'

    total_score = (er * 4) + (ir * 3) + (si * 3) + (rp * 4) + (rs * 3) + (ai * 3)
    tier = 'T1' if total_score >= 55 else 'T2' if total_score >= 35 else 'T3'

    entity_rollups[fname] = {
        'fin_tier': fe['fin_tier'],
        'campuses': campuses, 'served_camps': served_camps,
        'mh_camps': mh_camps, 'pcp_camps': pcp_camps, 'integ_camps': integ_camps,
        'total_rev': total_rev,
        'er': er, 'ir': ir, 'si': si, 'rp': rp, 'rs': rs, 'ai': ai,
        'rs_source': rs_source, 'rs_served': rs_served,
        'ir_source': ir_source, 'si_source': si_source, 'ai_source': ai_source,
        'total_score': total_score, 'tier': tier,
        'n_camp': n_camp, 'n_srv': n_srv,
    }

# Sort entities by score descending
sorted_entities = sorted(entity_rollups.keys(), key=lambda x: (-entity_rollups[x]['total_score'], x))

print(f"Scoreable entities: {len(sorted_entities)}")

# ============================================================
# BUILD WORKBOOK
# ============================================================
print("Building workbook...")
wb_out = openpyxl.Workbook()

# ============================================================
# TAB 1: ER — Enterprise Reach
# ============================================================
ws_er = wb_out.active
ws_er.title = 'ER - Enterprise Reach'

er_cols = ['', 'Facility Name', 'Address', 'City', 'State', 'Source Type', 'Total Beds', 'Campus Key', 'Status',
           '', 'Qualifying Campuses', 'ER Bracket', 'ER Score (×4)', 'Weighted']
ws_er.append(er_cols)
style_header(ws_er, 1, len(er_cols))

ws_er.column_dimensions['A'].width = 30   # Entity/indent
ws_er.column_dimensions['B'].width = 35
ws_er.column_dimensions['C'].width = 30
ws_er.column_dimensions['D'].width = 15
ws_er.column_dimensions['E'].width = 6
ws_er.column_dimensions['F'].width = 12
ws_er.column_dimensions['G'].width = 10
ws_er.column_dimensions['H'].width = 30
ws_er.column_dimensions['I'].width = 16
ws_er.column_dimensions['K'].width = 18
ws_er.column_dimensions['L'].width = 12
ws_er.column_dimensions['M'].width = 14
ws_er.column_dimensions['N'].width = 10

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    facs = entity_facilities.get(fname, [])

    # Entity header row
    ws_er.cell(row_num, 1, fname)
    ws_er.cell(row_num, 11, ru['n_camp'])
    ws_er.cell(row_num, 12, er_bracket_label(ru['er']))
    ws_er.cell(row_num, 13, ru['er'])
    ws_er.cell(row_num, 14, ru['er'] * 4)
    style_entity_row(ws_er, row_num, len(er_cols))
    row_num += 1

    # Track which campuses we've already listed (to mark duplicates)
    seen_campuses = set()
    # Sort facilities by campus key for grouping
    facs_sorted = sorted(facs, key=lambda f: (f['campus_key'], f['fac_name']))

    for f in facs_sorted:
        excluded_bed = f['tot_beds'] <= BED_MIN
        is_unique = not excluded_bed and f['campus_key'] not in seen_campuses
        if not excluded_bed:
            seen_campuses.add(f['campus_key'])

        status = 'EXCLUDED (≤15 beds)' if excluded_bed else ('YES' if is_unique else 'co-located')

        ws_er.cell(row_num, 1, '')  # indent
        ws_er.cell(row_num, 2, f['fac_name'])
        ws_er.cell(row_num, 3, f['addr'])
        ws_er.cell(row_num, 4, f['city'])
        ws_er.cell(row_num, 5, f['state'])
        ws_er.cell(row_num, 6, f['source'])
        ws_er.cell(row_num, 7, f['tot_beds'])
        ws_er.cell(row_num, 8, f"{f['campus_key'][0]}, {f['campus_key'][1]}")
        ws_er.cell(row_num, 9, status)
        style_detail_row(ws_er, row_num, len(er_cols))
        if excluded_bed:
            ws_er.cell(row_num, 9).font = Font(size=9, color='CC0000', italic=True)
        elif not is_unique:
            ws_er.cell(row_num, 9).font = Font(size=9, color='999999', italic=True)
        row_num += 1

    row_num += 1  # blank row between entities

# ============================================================
# TAB 2: IR — Integration Readiness
# ============================================================
ws_ir = wb_out.create_sheet('IR - Integration Readiness')

ir_cols = ['', 'Facility Name', 'Address', 'City', 'State', 'Do We Serve', 'MH Flag', 'PCP Flag', 'Integrated Flag',
           '', 'Served Camps', 'MH Camps', 'PCP Camps', 'Integ Camps', 'Service Config', 'IR Score (×3)', 'Weighted', 'Source']
ws_ir.append(ir_cols)
style_header(ws_ir, 1, len(ir_cols))

ws_ir.column_dimensions['A'].width = 30
ws_ir.column_dimensions['B'].width = 35
ws_ir.column_dimensions['C'].width = 30
ws_ir.column_dimensions['D'].width = 15
ws_ir.column_dimensions['E'].width = 6
ws_ir.column_dimensions['F'].width = 12
ws_ir.column_dimensions['G'].width = 9
ws_ir.column_dimensions['H'].width = 9
ws_ir.column_dimensions['I'].width = 13
ws_ir.column_dimensions['K'].width = 12
ws_ir.column_dimensions['L'].width = 10
ws_ir.column_dimensions['M'].width = 10
ws_ir.column_dimensions['N'].width = 12
ws_ir.column_dimensions['O'].width = 22
ws_ir.column_dimensions['P'].width = 14
ws_ir.column_dimensions['Q'].width = 10
ws_ir.column_dimensions['R'].width = 30

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    facs = entity_facilities.get(fname, [])

    # Determine service config label
    n_integ = len(ru['integ_camps'])
    n_mh = len(ru['mh_camps'])
    n_pcp = len(ru['pcp_camps'])
    if n_integ > 0:
        config = f"Integrated ({n_integ})"
    elif n_mh > 0 and n_pcp > 0:
        config = f"Dual MH+PCP"
    elif n_mh > 0:
        config = f"MH only ({n_mh})"
    elif n_pcp > 0:
        config = f"PCP only ({n_pcp})"
    elif ru['n_srv'] > 0:
        config = "Served, no svc flags"
    else:
        config = "Not served"

    ws_ir.cell(row_num, 1, fname)
    ws_ir.cell(row_num, 11, ru['n_srv'])
    ws_ir.cell(row_num, 12, n_mh)
    ws_ir.cell(row_num, 13, n_pcp)
    ws_ir.cell(row_num, 14, n_integ)
    ws_ir.cell(row_num, 15, config)
    ws_ir.cell(row_num, 16, ru['ir'])
    ws_ir.cell(row_num, 17, ru['ir'] * 3)
    ws_ir.cell(row_num, 18, ru['ir_source'])
    style_entity_row(ws_ir, row_num, len(ir_cols))
    row_num += 1

    facs_sorted = sorted(facs, key=lambda f: (not f['served'], f['campus_key'], f['fac_name']))
    for f in facs_sorted:
        ws_ir.cell(row_num, 1, '')
        ws_ir.cell(row_num, 2, f['fac_name'])
        ws_ir.cell(row_num, 3, f['addr'])
        ws_ir.cell(row_num, 4, f['city'])
        ws_ir.cell(row_num, 5, f['state'])
        ws_ir.cell(row_num, 6, 'Yes' if f['served'] else 'No')
        ws_ir.cell(row_num, 7, 'Yes' if f['mh'] else '')
        ws_ir.cell(row_num, 8, 'Yes' if f['pcp'] else '')
        ws_ir.cell(row_num, 9, 'Yes' if f['integ'] else '')
        style_detail_row(ws_ir, row_num, len(ir_cols))
        row_num += 1

    row_num += 1

# ============================================================
# TAB 3: SI — Strategic Influence
# ============================================================
ws_si = wb_out.create_sheet('SI - Strategic Influence')

si_cols = ['MUO Entity', 'SI Score (×3)', 'Weighted', 'Source', 'Research Basis']
ws_si.append(si_cols)
style_header(ws_si, 1, len(si_cols))

ws_si.column_dimensions['A'].width = 35
ws_si.column_dimensions['B'].width = 14
ws_si.column_dimensions['C'].width = 10
ws_si.column_dimensions['D'].width = 30
ws_si.column_dimensions['E'].width = 70

SI_RESEARCH_NOTES = {
    'ALG':                              'LTC Properties REIT transitioned 12 communities to ALG (2022)',
    'Brookdale Senior Living':          'NYSE: BKD. Largest US senior living co. REIT partnerships (Welltower, DHC)',
    'SABER HEALTHCARE GROUP':           'Omega Healthcare Investors 49% JV ownership stake',
    'TRILOGY':                          '100% owned by American Healthcare REIT ($258M deal, Sept 2024)',
    'Liberty':                          'Runs Liberty Advantage ISNP (HMO I-SNP) in NC',
    'PRUITT HEALTH':                    'Runs PruittHealth Premier ISNP (HMO I-SNP) in NC/SC/GA',
    'Avardis':                          'Omega Healthcare REIT lease ($3.1M/mo, through 2037). Formerly Consulate/LaVie',
    'SONIDA SENIOR LIVING':             'NYSE: SNDA. $1.8B merger with CNL Healthcare Properties (Nov 2025)',
    'Lifecare':                         'Largest privately held senior care co. 200+ facilities. Founded 1970',
    'AMERICAN SENIOR COMMUNITIES':      'SS agreement on Eventus paper. ACO participation',
    'INFINITY HEALTHCARE CONSULTING':   'JV with Longevity Health for I-SNP expansion (Oct 2024)',
    'Majestic Care':                    'Developing own ISNP. Expanding into OH (6 facilities, Jan 2025)',
    'NAVION':                           'NHI REIT relationship ($13.2M). Blackstone partnership. New CEO Jan 2025',
    'Kisco Senior Living':              'Private, 32 communities in 10 states + DC. Family-owned since 1990',
    'TLC Management':                   'LTC ACO agreement. Part of Hoosier Alliance (NOTE: alliance may be barrier)',
    'Lutheran Life Villages':           'LTC ACO agreement. Cathy/Mike put them in SS',
    'Lutheran Services Carolinas':      'ACO REACH with PPHP. Verbal SS commitment',
    'SANSTONE':                         'Independent NC operator, 20 locations. Growing — broke ground on new Pinehurst facility',
    'MORNING POINTE SENIOR LIVING':     'Regional brand TN/KY/GA. Small buildings',
    'OTTERBEIN SENIOR LIFE':            'Ohio nonprofit, 11 communities. S&P A rating. Acquiring Ohio Living Lake Vista',
    'PRINCIPLE':                        'Regional presence. Meeting scheduled (active engagement)',
    'LIONSTONE CARE':                   'Regional presence',
    'Kissito Healthcare':               'Acquired. Regional presence. COO not interested in SS',
    'CCH HEALTHCARE':                   'Regional operator. Defensive posture',
    'TERRABELLA SENIOR LIVING':         'Regional NC operator',
    'PEAK RESOURCES':                   'NC regional. Administrator-driven decisions',
    'BHI Senior Living':                'Faith-based nonprofit. 12 communities Midwest. BBB+ Fitch rating',
}

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    ws_si.cell(row_num, 1, fname)
    ws_si.cell(row_num, 2, ru['si'])
    ws_si.cell(row_num, 3, ru['si'] * 3)
    ws_si.cell(row_num, 4, ru['si_source'])
    ws_si.cell(row_num, 5, SI_RESEARCH_NOTES.get(fname, 'Minimal local presence — no significant affiliations found'))
    style_entity_row(ws_si, row_num, len(si_cols))
    row_num += 1

# ============================================================
# TAB 4: RP — Revenue Potential
# ============================================================
ws_rp = wb_out.create_sheet('RP - Revenue Potential')

rp_cols = ['', 'Facility Name', 'Address', 'City', 'State', 'Census', 'Total Beds',
           'Current Rev', 'Integration Rev', 'New Biz Rev', 'Total Potential Rev',
           '', 'Entity Total Rev', 'RP Bracket', 'RP Score (×4)', 'Weighted']
ws_rp.append(rp_cols)
style_header(ws_rp, 1, len(rp_cols))

ws_rp.column_dimensions['A'].width = 30
ws_rp.column_dimensions['B'].width = 35
ws_rp.column_dimensions['C'].width = 30
ws_rp.column_dimensions['D'].width = 15
ws_rp.column_dimensions['E'].width = 6
ws_rp.column_dimensions['F'].width = 8
ws_rp.column_dimensions['G'].width = 10
ws_rp.column_dimensions['H'].width = 14
ws_rp.column_dimensions['I'].width = 14
ws_rp.column_dimensions['J'].width = 14
ws_rp.column_dimensions['K'].width = 16
ws_rp.column_dimensions['M'].width = 16
ws_rp.column_dimensions['N'].width = 12
ws_rp.column_dimensions['O'].width = 14
ws_rp.column_dimensions['P'].width = 10

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    facs = entity_facilities.get(fname, [])

    ws_rp.cell(row_num, 1, fname)
    ws_rp.cell(row_num, 13, ru['total_rev'])
    ws_rp.cell(row_num, 13).number_format = MONEY_FMT
    ws_rp.cell(row_num, 14, rp_bracket_label(ru['rp']))
    ws_rp.cell(row_num, 15, ru['rp'])
    ws_rp.cell(row_num, 16, ru['rp'] * 4)
    style_entity_row(ws_rp, row_num, len(rp_cols))
    row_num += 1

    facs_sorted = sorted(facs, key=lambda f: (-f['tot_rev'], f['campus_key']))
    for f in facs_sorted:
        excluded_bed = f['tot_beds'] <= BED_MIN
        ws_rp.cell(row_num, 1, '')
        ws_rp.cell(row_num, 2, f['fac_name'])
        ws_rp.cell(row_num, 3, f['addr'])
        ws_rp.cell(row_num, 4, f['city'])
        ws_rp.cell(row_num, 5, f['state'])
        ws_rp.cell(row_num, 6, f['census'])
        ws_rp.cell(row_num, 7, f['tot_beds'])
        c = ws_rp.cell(row_num, 8, f['cur_rev']); c.number_format = MONEY_FMT
        c = ws_rp.cell(row_num, 9, f['int_rev']); c.number_format = MONEY_FMT
        c = ws_rp.cell(row_num, 10, f['nb_rev']); c.number_format = MONEY_FMT
        c = ws_rp.cell(row_num, 11, f['tot_rev']); c.number_format = MONEY_FMT
        style_detail_row(ws_rp, row_num, len(rp_cols))
        if excluded_bed:
            for col in [2, 7, 11]:
                ws_rp.cell(row_num, col).font = Font(size=9, color='CC0000', italic=True)
        row_num += 1

    row_num += 1

# ============================================================
# TAB 5: RS — Relationship Strength
# ============================================================
ws_rs = wb_out.create_sheet('RS - Relationship Strength')

rs_cols = ['', 'Facility Name', 'Address', 'City', 'State', 'Do We Serve', 'Campus Key',
           '', 'Served Camps', 'Total Camps', 'Served %', 'RS Score (×3)', 'Weighted',
           'Old Served-Ratio RS', 'Source', 'Tom Notes Summary']
ws_rs.append(rs_cols)
style_header(ws_rs, 1, len(rs_cols))

ws_rs.column_dimensions['A'].width = 30
ws_rs.column_dimensions['B'].width = 35
ws_rs.column_dimensions['C'].width = 30
ws_rs.column_dimensions['D'].width = 15
ws_rs.column_dimensions['E'].width = 6
ws_rs.column_dimensions['F'].width = 12
ws_rs.column_dimensions['G'].width = 30
ws_rs.column_dimensions['I'].width = 12
ws_rs.column_dimensions['J'].width = 12
ws_rs.column_dimensions['K'].width = 10
ws_rs.column_dimensions['L'].width = 14
ws_rs.column_dimensions['M'].width = 10
ws_rs.column_dimensions['N'].width = 18
ws_rs.column_dimensions['O'].width = 30
ws_rs.column_dimensions['P'].width = 60

# Tom's notes text for display
TOM_NOTES_SUMMARY = {
    'ALG': 'Presented SS, Charlie not interested, wants clinical support instead',
    'AMERICAN SENIOR COMMUNITIES': 'Have it on our paper, they\'re happy',
    'Brookdale Senior Living': 'Have on our paper, main contacts Laura Rivers and Dixon Allen',
    'SABER HEALTHCARE GROUP': 'Dr. Istenes not focused, leveraging not working, need Eventus paper',
    'INFINITY HEALTHCARE CONSULTING': 'Strong engagement but getting nowhere',
    'NAVION': 'High acuity ALF, growing, good to start talking about SS',
    'Majestic Care': 'Expressed interest in own medical group, JV discussed but hasn\'t moved',
    'PRUITT HEALTH': 'Getting more PC, good to approach, but focused on own ISNP',
    'Kisco Senior Living': 'Don\'t really know well',
    'TRILOGY': 'Already covered, speaking with Dr. McNamara',
    'Pavilion Healthcare': 'Don\'t really know well',
    'TOPAZ HEALTHCARE': 'Don\'t really know well',
    'MORNING POINTE SENIOR LIVING': 'Small buildings, not a great fit, don\'t recommend doing anything',
    'PRINCIPLE': 'Meeting scheduled on the 17th',
    'Liberty': 'Gave them proposal in RFP, focused on Liberty Advantage ISNP',
    'TERRABELLA SENIOR LIVING': 'Would have to drive growth, shouldn\'t just reward current',
    'SANSTONE': 'Presented SS, canceled meetings, not ready to commit',
    'LIONSTONE CARE': 'Tom needs to talk to Kim B',
    'ELDERCARE PARTNERS': 'Good relationship, out of KY',
    'OTTERBEIN SENIOR LIFE': 'Don\'t really know well',
    'TLC Management': 'Hoosier Alliance, not happy with ISNP, active contact',
    'BHI Senior Living': 'All psych',
    'Avardis': 'Don\'t really know well, they do not want SS',
    'PEAK RESOURCES': 'Not responsive or interested in SS, lost buildings',
    'ARBORS': 'Don\'t really know well',
    'HERITAGE HALL': 'Heritage Hall buildings, need to do something, Tom to talk to Colvin',
    'Runk & Pratt': 'Don\'t really know well',
    'MCAP': 'Don\'t really know well',
    'Lutheran Services Carolinas': 'Verbal commitment for SS, part of how we won business',
    'Lutheran Life Villages': 'Cathy worked with Mike, put them in SS',
    'Greencroft': 'Don\'t really know well',
    'CCH HEALTHCARE': 'Talk to them defensively',
    'LifeSpire of Virginia': 'Don\'t really know well',
    'SENIOR LIFESTYLE': 'Don\'t really know well',
    'CLEARVIEW': 'In process of partnering with Telos',
    'PRIORITY': 'Don\'t really know well',
    'CEDARHURST SENOR LIVING': 'Don\'t really know well',
    'Lifecare': 'Don\'t really know well',
    'Kissito Healthcare': 'Acquired, COO not interested in SS, just stay close',
    'SPRING ARBOR MANAGEMENT': 'Don\'t really know well',
    'YAD': 'Tried but never gotten audience, defensive',
    'STORYPOINT': 'Part of a growth plan',
    'CARING PLACE HEALTHCARE': 'Psych only, won\'t do PC, have their own',
    'Eastern Healthcare Group': 'Don\'t really know well',
    'SONIDA SENIOR LIVING': 'Growth opp, not doing much',
    'Castle Healthcare': 'Brooke and Ian have been more involved',
    'JAG': 'Ohio, not much growth opp',
    'FUNDAMENTAL LTC': 'Psych only',
    'CARDON & ASSOCIATES': 'Creating own group, Hoosier Alliance, not getting anything',
    'Southern Healthcare Mgmt': '(Not in MUO Data notes)',
    'Triple Crown': '(Not in MUO Data notes)',
}

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    facs = entity_facilities.get(fname, [])
    srv_pct = ru['n_srv'] / ru['n_camp'] if ru['n_camp'] > 0 else 0

    ws_rs.cell(row_num, 1, fname)
    ws_rs.cell(row_num, 9, ru['n_srv'])
    ws_rs.cell(row_num, 10, ru['n_camp'])
    c = ws_rs.cell(row_num, 11, srv_pct); c.number_format = PCT_FMT
    ws_rs.cell(row_num, 12, ru['rs'])
    ws_rs.cell(row_num, 13, ru['rs'] * 3)
    ws_rs.cell(row_num, 14, ru['rs_served'])
    ws_rs.cell(row_num, 15, ru['rs_source'])
    ws_rs.cell(row_num, 16, TOM_NOTES_SUMMARY.get(fname, ''))
    style_entity_row(ws_rs, row_num, len(rs_cols))
    row_num += 1

    seen_campuses = {}
    facs_sorted = sorted(facs, key=lambda f: (not f['served'], f['campus_key'], f['fac_name']))
    for f in facs_sorted:
        ck = f['campus_key']
        is_first = ck not in seen_campuses
        seen_campuses[ck] = True

        ws_rs.cell(row_num, 1, '')
        ws_rs.cell(row_num, 2, f['fac_name'])
        ws_rs.cell(row_num, 3, f['addr'])
        ws_rs.cell(row_num, 4, f['city'])
        ws_rs.cell(row_num, 5, f['state'])
        ws_rs.cell(row_num, 6, 'Yes' if f['served'] else 'No')
        ws_rs.cell(row_num, 7, f"{ck[0]}, {ck[1]}" if is_first else '(co-located)')
        style_detail_row(ws_rs, row_num, len(rs_cols))
        if not is_first:
            ws_rs.cell(row_num, 7).font = Font(size=9, color='999999', italic=True)
        row_num += 1

    row_num += 1

# ============================================================
# TAB 6: AI — AI/Tech Adoption
# ============================================================
ws_ai = wb_out.create_sheet('AI - Tech Adoption')

ai_cols = ['MUO Entity', 'AI Score (×3)', 'Weighted', 'Source', 'Notes']
ws_ai.append(ai_cols)
style_header(ws_ai, 1, len(ai_cols))

ws_ai.column_dimensions['A'].width = 35
ws_ai.column_dimensions['B'].width = 14
ws_ai.column_dimensions['C'].width = 10
ws_ai.column_dimensions['D'].width = 30
ws_ai.column_dimensions['E'].width = 50

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    ws_ai.cell(row_num, 1, fname)
    ws_ai.cell(row_num, 2, ru['ai'])
    ws_ai.cell(row_num, 3, ru['ai'] * 3)
    ws_ai.cell(row_num, 4, ru['ai_source'])
    ws_ai.cell(row_num, 5, 'Qualitative — no facility-level data drives this dimension')
    style_entity_row(ws_ai, row_num, len(ai_cols))
    row_num += 1

# ============================================================
# TAB 7: Summary — All Scores & Tiers
# ============================================================
ws_sum = wb_out.create_sheet('Summary')

sum_cols = ['MUO Entity', 'Finance Tier',
            'ER', 'ER Wt', 'IR', 'IR Wt', 'SI', 'SI Wt', 'RP', 'RP Wt', 'RS', 'RS Wt', 'AI', 'AI Wt',
            'Total Score', 'V23 Tier',
            'Campuses', 'Served', 'FP Revenue']
ws_sum.append(sum_cols)
style_header(ws_sum, 1, len(sum_cols))

ws_sum.column_dimensions['A'].width = 35
ws_sum.column_dimensions['B'].width = 12
for col_letter in ['C','D','E','F','G','H','I','J','K','L','M','N']:
    ws_sum.column_dimensions[col_letter].width = 7
ws_sum.column_dimensions['O'].width = 12
ws_sum.column_dimensions['P'].width = 10
ws_sum.column_dimensions['Q'].width = 10
ws_sum.column_dimensions['R'].width = 8
ws_sum.column_dimensions['S'].width = 16

# Tier color fills
T1_FILL = PatternFill('solid', fgColor='C6EFCE')
T2_FILL = PatternFill('solid', fgColor='FFEB9C')
T3_FILL = PatternFill('solid', fgColor='FFC7CE')

row_num = 2
for fname in sorted_entities:
    ru = entity_rollups[fname]
    ft = str(ru['fin_tier'] or '?').strip()
    if 'not on list' in ft.lower():
        ft = 'NEW'
    elif 'pulled' in ft.lower():
        ft = 'T1*'
    elif ft in ('1', '2', '3'):
        ft = f'T{ft}'

    ws_sum.cell(row_num, 1, fname)
    ws_sum.cell(row_num, 2, ft)
    ws_sum.cell(row_num, 3, ru['er'])
    ws_sum.cell(row_num, 4, ru['er'] * 4)
    ws_sum.cell(row_num, 5, ru['ir'])
    ws_sum.cell(row_num, 6, ru['ir'] * 3)
    ws_sum.cell(row_num, 7, ru['si'])
    ws_sum.cell(row_num, 8, ru['si'] * 3)
    ws_sum.cell(row_num, 9, ru['rp'])
    ws_sum.cell(row_num, 10, ru['rp'] * 4)
    ws_sum.cell(row_num, 11, ru['rs'])
    ws_sum.cell(row_num, 12, ru['rs'] * 3)
    ws_sum.cell(row_num, 13, ru['ai'])
    ws_sum.cell(row_num, 14, ru['ai'] * 3)
    ws_sum.cell(row_num, 15, ru['total_score'])
    ws_sum.cell(row_num, 16, ru['tier'])
    ws_sum.cell(row_num, 17, ru['n_camp'])
    ws_sum.cell(row_num, 18, ru['n_srv'])
    c = ws_sum.cell(row_num, 19, ru['total_rev']); c.number_format = MONEY_FMT

    # Color the tier cell
    tier_cell = ws_sum.cell(row_num, 16)
    if ru['tier'] == 'T1':
        tier_cell.fill = T1_FILL
    elif ru['tier'] == 'T2':
        tier_cell.fill = T2_FILL
    else:
        tier_cell.fill = T3_FILL

    tier_cell.font = Font(bold=True)
    row_num += 1

# ============================================================
# TAB 8: T4 — Independents (<7 campuses in footprint)
# ============================================================
ws_t4 = wb_out.create_sheet('T4 - Independents')

t4_cols = ['MUO Entity', 'Finance Tier', 'FP Campuses', 'Served', 'States', 'FP Revenue', 'Reason']
ws_t4.append(t4_cols)
style_header(ws_t4, 1, len(t4_cols))

ws_t4.column_dimensions['A'].width = 35
ws_t4.column_dimensions['B'].width = 12
ws_t4.column_dimensions['C'].width = 14
ws_t4.column_dimensions['D'].width = 8
ws_t4.column_dimensions['E'].width = 20
ws_t4.column_dimensions['F'].width = 16
ws_t4.column_dimensions['G'].width = 40

T4_FILL = PatternFill('solid', fgColor='D6DCE4')

sorted_t4 = sorted(t4_rollups.keys(), key=lambda x: (-t4_rollups[x]['n_camp'], x))
row_num = 2
for fname in sorted_t4:
    t4d = t4_rollups[fname]
    ft = str(t4d['fin_tier'] or '?').strip()
    if 'not on list' in ft.lower():
        ft = 'NEW'
    elif ft in ('1', '2', '3'):
        ft = f'T{ft}'

    reason = f"{t4d['n_camp']} qualifying campuses in FP (gate requires 7)" if t4d['n_camp'] > 0 else 'No DB rows in footprint'

    ws_t4.cell(row_num, 1, fname)
    ws_t4.cell(row_num, 2, ft)
    ws_t4.cell(row_num, 3, t4d['n_camp'])
    ws_t4.cell(row_num, 4, t4d['n_srv'])
    ws_t4.cell(row_num, 5, t4d['states'])
    c = ws_t4.cell(row_num, 6, t4d['total_rev']); c.number_format = MONEY_FMT
    ws_t4.cell(row_num, 7, reason)

    for col in range(1, len(t4_cols) + 1):
        ws_t4.cell(row_num, col).fill = T4_FILL
    row_num += 1

# Facility detail rows for T4 entities
row_num += 1
ws_t4.cell(row_num, 1, 'FACILITY DETAIL')
ws_t4.cell(row_num, 1).font = Font(bold=True, size=11)
row_num += 1

detail_cols = ['', 'Facility Name', 'Address', 'City', 'State', 'Do We Serve', 'Source Type']
for j, val in enumerate(detail_cols, 1):
    ws_t4.cell(row_num, j, val)
style_header(ws_t4, row_num, len(detail_cols))
row_num += 1

for fname in sorted_t4:
    t4d = t4_rollups[fname]
    if not t4d['facs']:
        continue
    ws_t4.cell(row_num, 1, fname)
    ws_t4.cell(row_num, 3, f"{t4d['n_camp']} campuses")
    style_entity_row(ws_t4, row_num, len(detail_cols))
    row_num += 1

    facs_sorted = sorted(t4d['facs'], key=lambda f: (f['campus_key'], f['fac_name']))
    for f in facs_sorted:
        ws_t4.cell(row_num, 2, f['fac_name'])
        ws_t4.cell(row_num, 3, f['addr'])
        ws_t4.cell(row_num, 4, f['city'])
        ws_t4.cell(row_num, 5, f['state'])
        ws_t4.cell(row_num, 6, 'Yes' if f['served'] else 'No')
        ws_t4.cell(row_num, 7, f['source'])
        style_detail_row(ws_t4, row_num, len(detail_cols))
        row_num += 1
    row_num += 1

print(f"T4 entities: {len(t4_rollups)}")

# ============================================================
# TAB 9: Bracket Reference
# ============================================================
ws_ref = wb_out.create_sheet('Bracket Reference')

ws_ref.column_dimensions['A'].width = 20
ws_ref.column_dimensions['B'].width = 15
ws_ref.column_dimensions['C'].width = 20
ws_ref.column_dimensions['D'].width = 8
ws_ref.column_dimensions['E'].width = 35

ws_ref.column_dimensions['E'].width = 80

ref_data = [
    ['MUO SCORING WORKBOOK V23 — VERSION 8 METHODOLOGY', '', '', '', ''],
    ['Generated: March 7, 2026', '', '', '', ''],
    ['', '', '', '', ''],
    ['SCORING FORMULA', '', '', '', ''],
    ['Score = (ER x4) + (IR x3) + (SI x3) + (RP x4) + (RS x3) + (AI x3)', '', '', '', ''],
    ['Maximum possible score = 100', '', '', '', ''],
    ['', '', '', '', ''],
    ['TIER THRESHOLDS (V6 — revised from V20)', '', '', '', ''],
    ['Tier', 'Score Range', 'V20 Range', '', 'Rationale for change'],
    ['T1 - Strategic Enterprise', '55 - 100', '50 - 100', '', 'V6 data-driven methodology produces higher scores on average; raising threshold prevents T1 inflation'],
    ['T2 - Growth & Expansion', '35 - 54', '25 - 49', '', 'Raised to create meaningful T3 population; V20 had 12 T3 entities, V5 had only 1 at old thresholds'],
    ['T3 - Retention/Watch', '0 - 34', '0 - 24', '', 'Entities below 35 have weak relationships + limited integration + small footprint or revenue'],
    ['T4 - Independents', 'Not scored', 'Not scored', '', '<7 qualifying campuses in footprint — engaged at facility level only'],
    ['T5 - Hard Barriers', 'Not scored', 'Not scored', '', 'Structural barriers prevent growth — defend current revenue only'],
    ['', '', '', '', ''],
    ['MUO GATE RULE', '', '', '', ''],
    ['>=7 qualifying campuses in operational footprint = scored (T1/T2/T3)', '', '', '', ''],
    ['<=6 qualifying campuses = T4 Independent (not scored)', '', '', '', ''],
    ['', '', '', '', ''],
    ['CAMPUS DEFINITION (V8 — Street-Number Matching)', '', '', '', ''],
    ['Campus = unique (street_number, city) pair within the same entity.', '', '', '', ''],
    ['If an address has no leading street number, falls back to (normalized_address, city).', '', '', '', ''],
    ['Co-located SNF+ALF at same street number and city = 1 campus.', '', '', '', ''],
    ['This catches abbreviation differences (Rd/Road, Dr/Drive), name variations, and data entry inconsistencies.', '', '', '', ''],
    ['', '', '', '', ''],
    ['15-BED MINIMUM RULE (V8)', '', '', '', ''],
    ['Facilities with Total_Beds <= 15 are excluded from ER campus counting and RP revenue summation.', '', '', '', ''],
    ['Rationale: Catches IL, MC-exclusive, and small satellite facilities that inflate campus counts.', '', '', '', ''],
    ['These facilities still appear in detail rows (marked red) but do not count toward campuses or revenue.', '', '', '', ''],
    ['', '', '', '', ''],
    ['OPERATIONAL FOOTPRINT', '', '', '', ''],
    ['NC, SC, VA, KY, OH, IN (6 states)', '', '', '', ''],
    ['All dimensions (ER, RP, RS, IR) are scoped to these states only.', '', '', '', ''],
    ['', '', '', '', ''],
    # --- ER ---
    ['DIMENSION 1: ENTERPRISE REACH (ER) — Weight x4', '', '', '', ''],
    ['Method: Data-driven. Count qualifying campuses (street-number collapsed, >15 beds) in operational footprint.', '', '', '', ''],
    ['Brackets (Option C):', 'Score', 'Campus Count', '', ''],
    ['', 1, '<9 campuses', '', ''],
    ['', 2, '9-12 campuses', '', ''],
    ['', 3, '13-19 campuses', '', ''],
    ['', 4, '20-39 campuses', '', ''],
    ['', 5, '40+ campuses', '', ''],
    ['', '', '', '', ''],
    # --- IR ---
    ['DIMENSION 2: INTEGRATION READINESS (IR) — Weight x3', '', '', '', ''],
    ['Method: Data-driven for ALL entities. Service flags from V23 DB (MH_Flag, PCP_Flag, Integrated_Flag).', '', '', '', ''],
    ['V20 used qualitative behavioral scores from Brooke. V6 replaces with formulaic approach.', '', '', '', ''],
    ['Service config sets the floor:', 'Score', '', '', ''],
    ['  Not served (no Do_We_Serve=YES rows)', '', '', 1, ''],
    ['  Single service (MH only or PCP only)', '', '', 2, '(conservative floor; 3 requires behavioral judgment)'],
    ['  Both MH+PCP present at separate campuses', '', '', 3, ''],
    ['  Dual MH+PCP at same campus, or 25-49% integrated', '', '', 4, ''],
    ['  Majority of campuses (>=50%) have Integrated_Flag', '', '', 5, ''],
    ['', '', '', '', ''],
    # --- SI ---
    ['DIMENSION 3: STRATEGIC INFLUENCE (SI) — Weight x3', '', '', '', ''],
    ['Method: Data-driven for ALL entities. Scored from Tom Shared Savings Notes (MUO Data workbook Col W)', '', '', '', ''],
    ['+ web research on REIT ownership, ISNP programs, payer/ACO connections, public company status.', '', '', '', ''],
    ['', 'Score', 'Definition', '', 'Examples from research'],
    ['', 1, 'No affiliations', '', 'No payer, ACO, REIT, or enterprise connections found'],
    ['', 2, 'Minimal local', '', 'Small/niche local operators (Runk & Pratt, MCAP, Greencroft)'],
    ['', 3, 'Regional reputation', '', 'Sanstone (20 NC locations), Otterbein (S&P A rating), Morning Pointe'],
    ['', 4, 'Payer/ACO connected', '', 'Infinity (JV with Longevity for I-SNP), Majestic (developing ISNP), Navion (NHI REIT + Blackstone)'],
    ['', 5, 'REIT/ISNP/Enterprise', '', 'Brookdale (NYSE:BKD), Trilogy (American Healthcare REIT), Liberty (ISNP), Pruitt (ISNP)'],
    ['', '', '', '', ''],
    ['SI RESEARCH SOURCES:', '', '', '', ''],
    ['  ALG: LTC Properties REIT transitioned 12 communities (Senior Housing News, Jul 2022)', '', '', '', ''],
    ['  Avardis: Omega Healthcare REIT lease $3.1M/mo through 2037 (McKnights Senior Living)', '', '', '', ''],
    ['  Brookdale: NYSE:BKD, partnerships with Welltower and DHC REITs (Yahoo Finance, 2025)', '', '', '', ''],
    ['  Infinity: JV with Longevity Health for I-SNP expansion (Skilled Nursing News, Oct 2024)', '', '', '', ''],
    ['  Liberty: Runs Liberty Advantage ISNP (HMO I-SNP) in NC (libertyadvantageplan.com)', '', '', '', ''],
    ['  Majestic Care: Developing own ISNP, expanding into OH with 6 facilities (Skilled Nursing News, Jan 2025)', '', '', '', ''],
    ['  Navion: NHI REIT relationship $13.2M, Blackstone partnership (Senior Housing News)', '', '', '', ''],
    ['  Pruitt Health: Runs PruittHealth Premier ISNP (HMO I-SNP) in NC/SC (Medicare.gov, CMS)', '', '', '', ''],
    ['  Saber: Omega Healthcare 49% JV ownership stake (McKnights)', '', '', '', ''],
    ['  Sonida: NYSE:SNDA, $1.8B merger with CNL Healthcare Properties Nov 2025 (BusinessWire)', '', '', '', ''],
    ['  Trilogy: 100% owned by American Healthcare REIT, $258M deal Sept 2024 (McKnights Senior Living)', '', '', '', ''],
    ['  Life Care Centers: Largest privately held senior care co, 200+ facilities (Wikipedia, PitchBook)', '', '', '', ''],
    ['  Kisco: Private, 32 communities in 10 states + DC, family-owned since 1990 (kiscoseniorliving.com)', '', '', '', ''],
    ['  Otterbein: Ohio nonprofit, 11 communities, S&P A rating 2025, acquiring Ohio Living Lake Vista', '', '', '', ''],
    ['  BHI Senior Living: Faith-based nonprofit, 12 communities Midwest, BBB+ Fitch rating', '', '', '', ''],
    ['  Sanstone: Independent NC operator, 20 locations, broke ground on new Pinehurst facility (sanstonehealth.com)', '', '', '', ''],
    ['', '', '', '', ''],
    # --- RP ---
    ['DIMENSION 4: REVENUE POTENTIAL (RP) — Weight x4', '', '', '', ''],
    ['Method: Data-driven. Sum of Total_Potential_Revenue from V23 S2 Economic Model, footprint only, >15 beds.', '', '', '', ''],
    ['S2 revenue = Census x Adjusted Fee per bed per year (SNF $4,583.50, ALF $3,699.50).', '', '', '', ''],
    ['This is TAM (theoretical addressable market), not near-term closeable revenue.', '', '', '', ''],
    ['Brackets (Option B):', 'Score', 'S2 Total Potential Rev', '', ''],
    ['', 1, '<$1M', '', ''],
    ['', 2, '$1M - $2.5M', '', ''],
    ['', 3, '$2.5M - $5M', '', ''],
    ['', 4, '$5M - $10M', '', ''],
    ['', 5, '>$10M', '', ''],
    ['', '', '', '', ''],
    # --- RS ---
    ['DIMENSION 5: RELATIONSHIP STRENGTH (RS) — Weight x3', '', '', '', ''],
    ['Method: Data-driven for ALL entities. Scored from Tom Shared Savings Notes (MUO Data workbook Col W).', '', '', '', ''],
    ['V20 used qualitative scores from Brooke. V6 replaces with Tom relationship assessment.', '', '', '', ''],
    ['', 'Score', 'Definition', '', 'Signal from Tom Notes'],
    ['', 1, 'No relationship', '', '"Don\'t really know well" or no meaningful contact'],
    ['', 2, 'Tried, hit resistance', '', 'Presented SS and rejected, defensive posture, psych-only/won\'t do PC'],
    ['', 3, 'Have relationship, stalled', '', 'Active contact but no commitment, meetings scheduled, early stage'],
    ['', 4, 'Active engagement', '', 'Verbal commitments, proposals in play, speaking with leadership'],
    ['', 5, 'Active SS on Eventus paper', '', 'Signed shared savings agreement on Eventus paper, happy relationship'],
    ['', '', '', '', ''],
    # --- AI ---
    ['DIMENSION 6: AI/TECH ADOPTION (AI) — Weight x3', '', '', '', ''],
    ['Method: V20 Brooke scores where available (31 entities), default=0 for remainder (14 entities).', '', '', '', ''],
    ['58 of 70 V20 entities scored AI=0. Only 12 had nonzero scores. No surrogate applied per V20 methodology.', '', '', '', ''],
    ['', '', '', '', ''],
    # --- T5 Barriers ---
    ['T5 BARRIER ENTITIES (excluded from scoring)', '', '', '', ''],
    ['Entity', '', 'Barrier Source', '', 'Barrier Detail'],
    ['Bluegrass/Encore', '', 'V20 original', '', 'Alliance, MH Only Opportunity (all rows)'],
    ['Signature Health', '', 'V20 original', '', 'MH Only Opportunity (all rows), own primary care and MH'],
    ['MFA', '', 'V20 original', '', 'Alliance (all rows)'],
    ['CommuniCare', '', 'V20 original', '', 'Own Provider Group (all rows), kicking us out everywhere'],
    ['Hill Valley', '', 'V20 original', '', 'MH Only Opportunity (28 of 33 rows)'],
    ['Singh', '', 'V20 original', '', 'Previously designated barrier'],
    ['Cardon & Associates', '', 'V6 — DB Barrier flag', '', 'Permanent - Cardon (32 of 35 rows = 91%). Creating own group, Hoosier Alliance'],
    ['Eastern Healthcare Group', '', 'V6 — DB Barrier flag', '', 'Termination Risk (19 of 19 rows = 100%)'],
    ['Clearview', '', 'V6 — DB Barrier flag', '', 'Alliance, Own Provider Group (3 of 8 rows = 38%). Partnering with Telos'],
    ['Pavilion Healthcare', '', 'V6 — DB Barrier flag', '', 'Alliance, Own Provider Group (6 of 21 rows = 29%)'],
    ['', '', '', '', ''],
    # --- T4 ---
    ['T4 INDEPENDENT ENTITIES (<7 campuses in footprint — see T4 tab for full list)', '', '', '', ''],
    ['Entities failing the 7-campus MUO gate are listed in the T4 - Independents tab with facility detail.', '', '', '', ''],
    ['', '', '', '', ''],
    # --- Version history ---
    ['VERSION HISTORY', '', '', '', ''],
    ['V1', '', 'Brooke V20 IR (33) + inferred (16). Served-ratio RS. V20/default SI. Thresholds 50/25.', '', 'T1=31 T2=18 T3=0'],
    ['V2', '', 'Formulaic IR for all 49 from DB service flags.', '', 'T1=34 T2=15 T3=0'],
    ['V3', '', 'Tom Shared Savings Notes RS for all 49.', '', 'T1=25 T2=23 T3=1'],
    ['V4', '', '4 barrier entities moved to T5 (Cardon, Eastern HC, Clearview, Pavilion). 45 scoreable.', '', 'T1=25 T2=19 T3=1'],
    ['V5', '', 'Data-driven SI for all 45 from Tom Notes + web research (REIT/ISNP/ACO).', '', 'T1=26 T2=18 T3=1'],
    ['V6', '', 'Tier thresholds raised to 55/35 (from 50/25). All 6 dimensions documented.', '', 'T1=19 T2=21 T3=5'],
    ['V7', '', 'Added T4 - Independents tab with facility detail for entities failing 7-campus gate.', '', ''],
    ['V8', '', 'Street-number matching + 15-bed minimum rule. Campus key = (street_number, city).', '', 'T1=19 T2=17 T3=2 T4=12 T5=10'],
    ['', '', '  Facilities with <=15 beds excluded from ER campus counting and RP revenue.', '', ''],
    ['', '', '  Street-number matching collapses address abbreviation variants (Rd/Road, Dr/Drive, etc.).', '', ''],
    ['', '', '', '', ''],
    ['DATA SOURCES', '', '', '', ''],
    ['DB:', '', 'V23 (4_Economic_Model_Scenario_2_Combined_V23.xlsx)', '', ''],
    ['Universe:', '', 'Finance MUO Data (60 entities). T5 excluded (10), T4 in own tab (12), 38 scored.', '', ''],
    ['IR:', '', 'V23 DB service flags (MH_Flag, PCP_Flag, Integrated_Flag) — all scored entities', '', ''],
    ['SI:', '', 'Tom Shared Savings Notes (Col W) + web research — all scored entities', '', ''],
    ['RS:', '', 'Tom Shared Savings Notes (Col W) — all scored entities', '', ''],
    ['AI:', '', 'V20 Brooke scores where available + default=0 for remainder', '', ''],
    ['ER/RP:', '', 'V23 DB campus counts (street-number collapsed, >15 beds) and S2 Total Potential Revenue', '', ''],
]

# Bold the section headers in Bracket Reference
# Re-derive from content rather than hardcoding line numbers
bold_keywords = {'MUO SCORING', 'SCORING FORMULA', 'TIER THRESHOLDS', 'MUO GATE',
                 'CAMPUS DEFINITION', '15-BED MINIMUM', 'OPERATIONAL FOOTPRINT',
                 'DIMENSION', 'T5 BARRIER', 'T4 INDEPENDENT', 'VERSION HISTORY',
                 'DATA SOURCES', 'SI RESEARCH'}
bold_rows = set()
for idx, rd in enumerate(ref_data, 1):
    if rd[0] and any(kw in rd[0] for kw in bold_keywords):
        bold_rows.add(idx)
for i, row_data in enumerate(ref_data, 1):
    for j, val in enumerate(row_data, 1):
        ws_ref.cell(i, j, val)
    if i in bold_rows:
        ws_ref.cell(i, 1).font = Font(bold=True, size=11)

# ============================================================
# SAVE
# ============================================================
print(f"Saving to {OUT_PATH}...")
wb_out.save(OUT_PATH)
print("Done!")

# Quick summary
from collections import Counter
tiers = Counter(entity_rollups[f]['tier'] for f in sorted_entities)
print(f"\nTier distribution: T1={tiers.get('T1',0)}  T2={tiers.get('T2',0)}  T3={tiers.get('T3',0)}  T4={len(t4_rollups)}")

v20_count = sum(1 for f in sorted_entities if 'V20' in entity_rollups[f]['ir_source'])
inf_count = len(sorted_entities) - v20_count
print(f"Score sources: {v20_count} from V20 Brooke, {inf_count} inferred")
