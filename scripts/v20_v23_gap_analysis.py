"""Map V20 (Final MUO Tiering.xlsx — 70 scored entities) against V23 V8
to find the gap of entities that need scoring."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

# V20 scored entities
wb_v20 = openpyxl.load_workbook('C:/Users/ratwood/Downloads/Final MUO Tiering.xlsx', data_only=True)
v20_scored = {}
for sn in ['T1 MUO', 'T2 MUO', 'T3 MUO']:
    ws = wb_v20[sn]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name:
            score = ws.cell(r, 3).value
            v20_scored[name.strip()] = {'tier': sn.split()[0], 'score': score}

# V20 T5
v20_t5 = set()
ws5 = wb_v20['T5- Hard Barriers ']
for r in range(2, ws5.max_row + 1):
    name = ws5.cell(r, 1).value
    if name and name.strip() != 'TOTAL':
        v20_t5.add(name.strip())

# V23 V8 workbook
wb_v23 = openpyxl.load_workbook(
    'C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/03_Corporate_Accounts/Tiering/MUO_Scoring_Workbook_V23_v8.xlsx',
    data_only=True)

# V23 scored (Summary tab)
ws_sum = wb_v23['Summary']
v23_scored = {}
for r in range(2, ws_sum.max_row + 1):
    name = ws_sum.cell(r, 1).value
    tier = ws_sum.cell(r, 16).value
    score = ws_sum.cell(r, 15).value
    if name:
        v23_scored[name.strip()] = {'tier': tier, 'score': score}

# V23 T4
ws_t4 = wb_v23['T4 - Independents']
v23_t4 = {}
for r in range(2, ws_t4.max_row + 1):
    name = ws_t4.cell(r, 1).value
    camps = ws_t4.cell(r, 3).value
    if name and name != 'FACILITY DETAIL':
        v23_t4[name.strip()] = {'tier': 'T4', 'camps': camps}

# V23 T5
v23_t5 = {'Bluegrass/Encore', 'SIGNATURE HEALTH', 'MFA', 'COMMUNICARE', 'Hill Valley', 'SINGH',
           'CARDON & ASSOCIATES', 'Eastern Healthcare Group', 'CLEARVIEW', 'Pavilion Healthcare'}

# Known name mappings: V20 name -> V23 name
V20_TO_V23 = {
    'ALG': 'ALG',
    'American Senior Communities': 'AMERICAN SENIOR COMMUNITIES',
    'Brookdale Senior Living': 'Brookdale Senior Living',
    'Cch Healthcare': 'CCH HEALTHCARE',
    'Consulate Health Care/Independence Living Centers/Nspire Healthcare/Raydiant Health Care': 'Avardis',
    'Infinity Healthcare Consulting': 'INFINITY HEALTHCARE CONSULTING',
    'Liberty Senior Living': 'Liberty',
    'Life Care Centers Of America': 'Lifecare',
    'Majestic Care': 'Majestic Care',
    'Navion': 'NAVION',
    'Otterbein Seniorlife': 'OTTERBEIN SENIOR LIFE',
    'Principle Long Term Care': 'PRINCIPLE',
    'Pruitthealth': 'PRUITT HEALTH',
    'Saber Healthcare Group': 'SABER HEALTHCARE GROUP',
    'Terra Bella': 'TERRABELLA SENIOR LIVING',
    'Tlc Management': 'TLC Management',
    'Trilogy Health Services': 'TRILOGY',
    'Arbors At Ohio': 'ARBORS',
    'BHI Senior Living': 'BHI Senior Living',
    'Castle Healthcare': 'Castle Healthcare',
    'Heritage Hall': 'AMERICAN HEALTHCARE LLC',
    'Jag Healthcare': 'JAG',
    'Kissito': 'Kissito Healthcare',
    'Lionstone Care': 'LIONSTONE CARE',
    'Lutheran Services Carolina': 'Lutheran Services Carolinas',
    'Momentous Health': 'Momentus Health',
    'Morning Pointe Senior Living': 'MORNING POINTE SENIOR LIVING',
    'Peak Resources, Inc.': 'PEAK RESOURCES',
    'Priority Life Care': 'PRIORITY',
    'Sanstone Health & Rehabilitation': 'SANSTONE',
    'Topaz': 'TOPAZ HEALTHCARE',
    'Yad Healthcare': 'YAD',
    'Caring Place Healthcare': 'CARING PLACE HEALTHCARE',
    'Clearview': 'CLEARVIEW',
    'Pavilion Healthcare': 'Pavilion Healthcare',
}

# Also try case-insensitive matching for anything not in the map
v23_all_names = {}
for name in v23_scored:
    v23_all_names[name.upper()] = ('scored', name)
for name in v23_t4:
    v23_all_names[name.upper()] = ('T4', name)
for name in v23_t5:
    v23_all_names[name.upper()] = ('T5', name)

# Map
matched = []
unmatched_v20 = []

for v20_name, v20_data in v20_scored.items():
    v23_name = V20_TO_V23.get(v20_name)
    if v23_name:
        if v23_name in v23_scored:
            v23_tier = v23_scored[v23_name]['tier']
            v23_score = v23_scored[v23_name]['score']
        elif v23_name in v23_t4:
            v23_tier = 'T4'
            v23_score = None
        elif v23_name in v23_t5:
            v23_tier = 'T5'
            v23_score = None
        else:
            v23_tier = '?'
            v23_score = None
        matched.append((v20_name, v20_data['tier'], v20_data['score'], v23_name, v23_tier, v23_score))
    else:
        # Try case-insensitive
        key = v20_name.upper()
        if key in v23_all_names:
            kind, v23_name = v23_all_names[key]
            if kind == 'scored':
                v23_tier = v23_scored[v23_name]['tier']
                v23_score = v23_scored[v23_name]['score']
            elif kind == 'T4':
                v23_tier = 'T4'
                v23_score = None
            else:
                v23_tier = 'T5'
                v23_score = None
            matched.append((v20_name, v20_data['tier'], v20_data['score'], v23_name, v23_tier, v23_score))
        else:
            unmatched_v20.append((v20_name, v20_data['tier'], v20_data['score']))

# V23 entities not in V20
v23_names_in_v20 = set(v[3] for v in matched)
v23_only = []
for name, data in v23_scored.items():
    if name not in v23_names_in_v20:
        v23_only.append((name, data['tier'], data['score']))
for name, data in v23_t4.items():
    if name not in v23_names_in_v20:
        v23_only.append((name, 'T4', None))

# Print results
print('=' * 120)
print(f'V20 vs V23 GAP ANALYSIS')
print(f'V20: {len(v20_scored)} scored + {len(v20_t5)} T5 = {len(v20_scored) + len(v20_t5)} total')
print(f'V23: {len(v23_scored)} scored + {len(v23_t4)} T4 + {len(v23_t5)} T5 = {len(v23_scored) + len(v23_t4) + len(v23_t5)} total')
print('=' * 120)

print(f'\n--- MATCHED: {len(matched)} V20 entities found in V23 ---')
print(f'{"V20 Name":<65} {"V20":>3} {"V20sc":>5}  {"V23 Name":<35} {"V23":>3} {"V23sc":>5}')
print('-' * 125)
for v20_name, v20_tier, v20_score, v23_name, v23_tier, v23_score in sorted(matched, key=lambda x: x[0]):
    v23s = str(v23_score) if v23_score is not None else '-'
    print(f'{v20_name:<65} {v20_tier:>3} {v20_score:>5}  {v23_name:<35} {v23_tier:>3} {v23s:>5}')

print(f'\n--- GAP: {len(unmatched_v20)} V20 entities NOT in V23 (need scoring) ---')
print(f'{"V20 Name":<65} {"V20 Tier":>8} {"V20 Score":>9}')
print('-' * 85)
for v20_name, v20_tier, v20_score in sorted(unmatched_v20, key=lambda x: x[0]):
    print(f'{v20_name:<65} {v20_tier:>8} {v20_score:>9}')

print(f'\n--- V23-ONLY: {len(v23_only)} entities in V23 but NOT in V20 ---')
print(f'{"V23 Name":<45} {"V23 Tier":>8} {"V23 Score":>9}')
print('-' * 65)
for name, tier, score in sorted(v23_only, key=lambda x: x[0]):
    s = str(score) if score is not None else '-'
    print(f'{name:<45} {tier:>8} {s:>9}')

# V20 T5 mapping
print(f'\n--- V20 T5 BARRIERS ({len(v20_t5)}) vs V23 T5 ({len(v23_t5)}) ---')
for name in sorted(v20_t5):
    in_v23 = 'YES' if any(name.upper() in v.upper() for v in v23_t5) else 'NO'
    print(f'  {name:<35} in V23 T5: {in_v23}')

# Summary
print(f'\n{"=" * 120}')
print(f'SUMMARY')
print(f'{"=" * 120}')
print(f'  Matched (already scored in V23): {len(matched)}')
print(f'  Gap (V20 entities needing V23 scoring): {len(unmatched_v20)}')
print(f'  V23-only (new to Finance 60, not in V20): {len(v23_only)}')
print(f'  V20 T5 barriers: {len(v20_t5)}')
