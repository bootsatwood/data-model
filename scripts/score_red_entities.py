"""Score the 9 red-flagged 'not on list' entities from MUO Data workbook.
Uses campus-collapsed (unique address) Enterprise Reach."""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

# Get the red-flagged "not on list" entities from Analysis sheet
wb_muo = openpyxl.load_workbook("C:/Users/ratwood/Downloads/MUO Data (1).xlsx", data_only=True)
ws_a = wb_muo['Analysis']

red_entities = []
for r in range(18, 78):
    name = ws_a.cell(r, 2).value
    tier = ws_a.cell(r, 18).value
    glr_facs = ws_a.cell(r, 4).value
    core_rev = ws_a.cell(r, 14).value
    if name and tier and 'not on list' in str(tier).lower():
        red_entities.append({
            'name': str(name).strip(),
            'name_upper': str(name).strip().upper(),
            'glr_facs': int(glr_facs) if glr_facs else 0,
            'glr_core_rev': core_rev or 0,
        })

# Load V23 S2 data
wb_s2 = openpyxl.load_workbook(
    "C:/Users/ratwood/OneDrive - Eventus WholeHealth/Vault/02_Data_Model/Current/4_Economic_Model_Scenario_2_Combined_V23.xlsx",
    data_only=True
)
ws2 = wb_s2.active
headers = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1) if ws2.cell(1, c).value}

# Build facility-level data
facility_data = []
for r in range(2, ws2.max_row + 1):
    corp = ws2.cell(r, headers['Corporate_Name']).value
    if not corp:
        continue
    facility_data.append({
        'corp': str(corp).strip().upper(),
        'addr': str(ws2.cell(r, headers['Address']).value or '').strip().upper(),
        'city': str(ws2.cell(r, headers['City']).value or '').strip().upper(),
        'state': str(ws2.cell(r, headers['State']).value or '').strip(),
        'facility': str(ws2.cell(r, headers['Facility_Name']).value or '').strip(),
        'type': str(ws2.cell(r, headers['Source_Type']).value or '').strip(),
        'census': ws2.cell(r, headers['Census']).value or 0,
        'served': str(ws2.cell(r, headers['Do_We_Serve']).value or '').strip().upper() == 'YES',
        'integrated': str(ws2.cell(r, headers['Integrated_Flag']).value or '').strip().upper() == 'YES',
        'pcp': str(ws2.cell(r, headers['PCP_Flag']).value or '').strip().upper() == 'YES',
        'mh': str(ws2.cell(r, headers['MH_Flag']).value or '').strip().upper() == 'YES',
        'cur_rev': ws2.cell(r, headers['Current_Revenue']).value or 0,
        'int_rev': ws2.cell(r, headers['Integration_Revenue']).value or 0,
        'nb_rev': ws2.cell(r, headers['New_Business_Revenue']).value or 0,
        'tot_pot': ws2.cell(r, headers['Total_Potential_Revenue']).value or 0,
        'metro': str(ws2.cell(r, headers['Metro_Assignment']).value or '').strip(),
    })


def match_corp(red_name, db_corp):
    if red_name == db_corp:
        return True
    if red_name in db_corp or db_corp in red_name:
        return True
    mappings = {
        'AMERICAN HEALTHCARE LLC': ['AMERICAN HEALTHCARE'],
        'MCAP': ['MCAP'],
        'GREENCROFT': ['GREENCROFT'],
        'LIFESPIRE OF VIRGINIA': ['LIFESPIRE'],
        'SENIOR LIFESTYLE': ['SENIOR LIFESTYLE'],
        'CEDARHURST SENOR LIVING': ['CEDARHURST'],
        'SPRING ARBOR MANAGEMENT': ['SPRING ARBOR'],
        'STORYPOINT': ['STORYPOINT'],
        'SONIDA SENIOR LIVING': ['SONIDA'],
    }
    for key, vals in mappings.items():
        if red_name == key:
            for v in vals:
                if v in db_corp or db_corp in v:
                    return True
    return False


# Scoring rubrics
def er_score(f):
    if f >= 40: return 5
    if f >= 20: return 4
    if f >= 10: return 3
    if f >= 5: return 2
    return 1

def rp_score(rev):
    if rev >= 2000000: return 5
    if rev >= 1000000: return 4
    if rev >= 500000: return 3
    if rev >= 250000: return 2
    return 1

def ir_score(served, integrated, pcp_only, mh_only):
    if served == 0:
        return 1
    int_pct = integrated / served
    if int_pct >= 0.8: return 5
    if int_pct >= 0.5: return 4
    if integrated > 0 or (pcp_only > 0 and mh_only > 0): return 3
    if pcp_only > 0 or mh_only > 0: return 2
    return 1

def rs_score(served, total):
    if total == 0:
        return 0
    pct = served / total
    if pct >= 0.8: return 5
    if pct >= 0.5: return 4
    if pct >= 0.25: return 3
    if served > 0: return 2
    return 1

def er_bracket_label(n):
    if n < 5: return '1-4'
    if n < 10: return '5-9'
    if n < 20: return '10-19'
    if n < 40: return '20-39'
    return '40+'

def rp_bracket_label(rev):
    if rev < 250000: return '<$250K'
    if rev < 500000: return '$250-499K'
    if rev < 1000000: return '$500K-$999K'
    if rev < 2000000: return '$1.0-1.99M'
    return '>$2M'


# Process each red entity
for ent in red_entities:
    red_name = ent['name_upper']

    # Collect all matching rows
    rows = [f for f in facility_data if match_corp(red_name, f['corp'])]

    if not rows:
        print("=" * 100)
        print(f"  {ent['name']}  --  NO V23 DB MATCH")
        print(f"  GLR shows {ent['glr_facs']} facilities, ${ent['glr_core_rev']:,.0f} core revenue")
        print(f"  Likely a DB name mismatch -- needs reconciliation")
        print("=" * 100)
        print()
        continue

    # Collapse by address -> campuses
    campuses = defaultdict(lambda: {
        'facilities': [], 'census': 0, 'served': False,
        'integrated': False, 'pcp': False, 'mh': False,
        'cur_rev': 0, 'int_rev': 0, 'nb_rev': 0, 'tot_pot': 0,
        'state': '', 'city': '', 'metro': '', 'types': set()
    })

    for f in rows:
        key = f['addr'] + '|' + f['city']
        c = campuses[key]
        c['facilities'].append(f['facility'])
        c['census'] += f['census'] if isinstance(f['census'], (int, float)) else 0
        c['state'] = f['state']
        c['city'] = f['city']
        c['metro'] = f['metro']
        c['types'].add(f['type'])
        if f['served']:
            c['served'] = True
        if f['integrated']:
            c['integrated'] = True
        if f['pcp']:
            c['pcp'] = True
        if f['mh']:
            c['mh'] = True
        for k in ('cur_rev', 'int_rev', 'nb_rev', 'tot_pot'):
            val = f[k]
            if isinstance(val, (int, float)):
                c[k] += val

    # Aggregate campus-level metrics
    n_rows = len(rows)
    n_campuses = len(campuses)
    n_served = sum(1 for c in campuses.values() if c['served'])
    n_integrated = sum(1 for c in campuses.values() if c['integrated'])
    n_pcp_only = sum(1 for c in campuses.values() if c['pcp'] and not c['integrated'] and not c['mh'])
    n_mh_only = sum(1 for c in campuses.values() if c['mh'] and not c['integrated'] and not c['pcp'])
    total_cur = sum(c['cur_rev'] for c in campuses.values())
    total_int = sum(c['int_rev'] for c in campuses.values())
    total_nb = sum(c['nb_rev'] for c in campuses.values())
    total_pot = sum(c['tot_pot'] for c in campuses.values())
    states = sorted(set(c['state'] for c in campuses.values() if c['state']))
    metro_count = sum(1 for c in campuses.values() if c['metro'] and c['metro'] not in ('None', '', 'nan'))

    # Score
    er = er_score(n_campuses)
    ir = ir_score(n_served, n_integrated, n_pcp_only, n_mh_only)
    si = 2  # Default -- needs manual assessment
    rp = rp_score(total_pot)
    rs = rs_score(n_served, n_campuses)
    ai = 0  # No data
    score = er * 4 + ir * 3 + si * 3 + rp * 4 + rs * 3 + ai * 3
    tier = 'T1' if score >= 50 else 'T2' if score >= 25 else 'T3'
    meets_muo = n_campuses >= 7

    # Print
    print("=" * 100)
    print(f"  {ent['name']}")
    print("=" * 100)
    print(f"  DB Rows: {n_rows}  |  Unique Campuses: {n_campuses}  |  MUO Threshold (7+): ", end='')
    if meets_muo:
        print("MEETS")
    else:
        print(f"FAILS ({n_campuses} < 7)")
    print(f"  States: {', '.join(states)}  |  Metro Campuses: {metro_count}")
    srv_pct = n_served / n_campuses * 100 if n_campuses > 0 else 0
    print(f"  Served: {n_served}/{n_campuses} ({srv_pct:.0f}%)  |  INT: {n_integrated}  PCP-only: {n_pcp_only}  MH-only: {n_mh_only}")
    print()
    print(f"  Revenue (S2):")
    print(f"    Current:      ${total_cur:>14,.0f}")
    print(f"    Integration:  ${total_int:>14,.0f}")
    print(f"    New Business: ${total_nb:>14,.0f}")
    print(f"    Total Pot:    ${total_pot:>14,.0f}")
    print()
    print(f"  Scoring (campus-based ER):")
    print(f"    Enterprise Reach:      {er}  (x4 = {er * 4:>2})   {n_campuses} campuses -> {er_bracket_label(n_campuses)} bracket")
    print(f"    Integration Readiness: {ir}  (x3 = {ir * 3:>2})   {n_served} served, {n_integrated} integrated")
    print(f"    Strategic Influence:    {si}  (x3 = {si * 3:>2})   Default (needs manual assessment)")
    print(f"    Revenue Potential:      {rp}  (x4 = {rp * 4:>2})   ${total_pot:,.0f} -> {rp_bracket_label(total_pot)}")
    print(f"    Relationship Strength: {rs}  (x3 = {rs * 3:>2})   {n_served}/{n_campuses} = {srv_pct:.0f}%")
    print(f"    AI/Tech Adoption:      {ai}  (x3 = {ai * 3:>2})   No data")
    print(f"    -----------------------------------------")
    print(f"    TOTAL SCORE:          {score:>2}   ->  {tier}")
    if not meets_muo:
        print(f"    *** DOES NOT MEET MUO 7-FACILITY THRESHOLD ***")
    print()
