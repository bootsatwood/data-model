#!/usr/bin/env python3
"""
QC Validation Auto-Populator V16.0
Automatically populates the QC Validation Workbook with actual values from your model files.
Reduces manual data entry and speeds up validation process.
"""

import pandas as pd
from openpyxl import load_workbook
import os
import sys
import json
from datetime import datetime

class QCValidator:
    def __init__(self):
        self.results = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'files_checked': [],
            'validation_results': {},
            'errors': []
        }
        
    def check_file_exists(self, filepath, description):
        """Check if a required file exists"""
        if os.path.exists(filepath):
            self.results['files_checked'].append(f"✓ {description}")
            return True
        else:
            self.results['errors'].append(f"✗ Missing: {description} at {filepath}")
            return False
    
    def validate_database(self, filepath):
        """Validate Combined Database structure and counts"""
        if not self.check_file_exists(filepath, "Combined Database"):
            return None
            
        try:
            df = pd.read_excel(filepath)
            
            validation = {
                'total_facilities': len(df),
                'snf_count': len(df[df['Source_Type'] == 'SNF']) if 'Source_Type' in df.columns else 0,
                'alf_count': len(df[df['Source_Type'] == 'ALF']) if 'Source_Type' in df.columns else 0,
                'served_count': len(df[df['Do_We_Serve'] == 'Yes']) if 'Do_We_Serve' in df.columns else 0,
                'columns_present': list(df.columns),
                'column_count': len(df.columns)
            }
            
            # Check for barriers
            if 'Has_Barrier' in df.columns:
                validation['total_barriers'] = df['Has_Barrier'].sum()
            
            # Check CCH Healthcare specific
            if 'Parent_Company' in df.columns and 'State' in df.columns:
                cch_df = df[df['Parent_Company'] == 'CCH HEALTHCARE']
                validation['cch_ohio_barriers'] = len(cch_df[(cch_df['State'] == 'OH') & (cch_df.get('Has_Barrier', 0) == 1)])
                validation['cch_nc_barriers'] = len(cch_df[(cch_df['State'] == 'NC') & (cch_df.get('Has_Barrier', 0) == 1)])
            
            self.results['validation_results']['database'] = validation
            print(f"✓ Database validated: {validation['total_facilities']} facilities")
            return validation
            
        except Exception as e:
            self.results['errors'].append(f"Error reading database: {str(e)}")
            return None
    
    def validate_scenario(self, filepath, scenario_num):
        """Validate Economic Model Scenario file"""
        if not self.check_file_exists(filepath, f"Economic Model Scenario {scenario_num}"):
            return None
            
        try:
            # Read all sheets
            xl = pd.ExcelFile(filepath)
            sheets = xl.sheet_names
            
            validation = {
                'sheet_count': len(sheets),
                'sheet_names': sheets,
                'revenue_totals': {}
            }
            
            # Try to read summary data if available
            if 'Summary' in sheets or 'Model' in sheets:
                sheet_name = 'Summary' if 'Summary' in sheets else 'Model'
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                
                # Look for revenue columns
                if 'Current_Revenue' in df.columns:
                    validation['revenue_totals']['current'] = df['Current_Revenue'].sum()
                if 'Integration_Revenue' in df.columns:
                    validation['revenue_totals']['integration'] = df['Integration_Revenue'].sum()
                if 'New_Business_Revenue' in df.columns:
                    validation['revenue_totals']['new_business'] = df['New_Business_Revenue'].sum()
            
            self.results['validation_results'][f'scenario_{scenario_num}'] = validation
            print(f"✓ Scenario {scenario_num} validated")
            return validation
            
        except Exception as e:
            self.results['errors'].append(f"Error reading Scenario {scenario_num}: {str(e)}")
            return None
    
    def validate_comprehensive_report(self, filepath):
        """Validate Comprehensive Report structure"""
        if not self.check_file_exists(filepath, "Comprehensive Report"):
            return None
            
        try:
            xl = pd.ExcelFile(filepath)
            sheets = xl.sheet_names
            
            validation = {
                'sheet_count': len(sheets),
                'sheet_names': sheets,
                'structure_checks': {}
            }
            
            # Check for expected sheets
            expected_sheets = [
                'TAM SAM SOM Facilities',
                'TAM SAM SOM Revenue', 
                'Fee Structure SOM',
                'Top Corporate Rankings',
                'State Analysis'
            ]
            
            for expected in expected_sheets:
                validation['structure_checks'][expected] = any(expected.lower() in sheet.lower() for sheet in sheets)
            
            # Check Table 25 for INDEPENDENT
            if 'Top Corporate Rankings' in sheets or any('corporate' in s.lower() for s in sheets):
                corp_sheet = next((s for s in sheets if 'corporate' in s.lower()), None)
                if corp_sheet:
                    df = pd.read_excel(filepath, sheet_name=corp_sheet)
                    # Check if INDEPENDENT appears in top rankings
                    if 'Corporate_Name' in df.columns or 'Corporate Name' in df.columns:
                        col_name = 'Corporate_Name' if 'Corporate_Name' in df.columns else 'Corporate Name'
                        validation['independent_in_rankings'] = 'INDEPENDENT' in df[col_name].values
                    
            self.results['validation_results']['comprehensive_report'] = validation
            print(f"✓ Comprehensive Report validated: {len(sheets)} sheets")
            return validation
            
        except Exception as e:
            self.results['errors'].append(f"Error reading Comprehensive Report: {str(e)}")
            return None
    
    def validate_fee_schedule(self, filepath):
        """Validate Fee Schedule Reference"""
        if not self.check_file_exists(filepath, "Fee Schedule Reference"):
            return None
            
        try:
            df = pd.read_excel(filepath)
            
            validation = {
                'rows': len(df),
                'columns': len(df.columns),
                'fees_found': {}
            }
            
            # Try to extract specific fee values
            # This would need to be customized based on actual fee schedule structure
            
            self.results['validation_results']['fee_schedule'] = validation
            print(f"✓ Fee Schedule validated")
            return validation
            
        except Exception as e:
            self.results['errors'].append(f"Error reading Fee Schedule: {str(e)}")
            return None
    
    def update_qc_workbook(self, qc_filepath, validation_data):
        """Update the QC Validation Workbook with actual values"""
        try:
            wb = load_workbook(qc_filepath)
            
            # Update Financial Reconciliation sheet
            if 'Financial Reconciliation' in wb.sheetnames:
                ws = wb['Financial Reconciliation']
                if 'database' in validation_data:
                    db = validation_data['database']
                    ws['C3'] = str(db.get('total_facilities', ''))  # Total Facilities
                    ws['C4'] = str(db.get('served_count', ''))      # Total Served
                    ws['C5'] = str(db.get('total_barriers', ''))    # Total Barriers
                    ws['C8'] = str(db.get('snf_count', ''))         # SNF Facilities
                    ws['C9'] = str(db.get('alf_count', ''))         # ALF Facilities
            
            # Update Structure Validation sheet
            if 'Structure Validation' in wb.sheetnames:
                ws = wb['Structure Validation']
                if 'comprehensive_report' in validation_data:
                    cr = validation_data['comprehensive_report']
                    ws['D5'] = str(cr.get('sheet_count', ''))  # Sheet count
                    
                    # Check sheet names
                    for i, sheet_name in enumerate(cr.get('sheet_names', [])[:6], 6):
                        ws[f'D{i}'] = sheet_name
            
            # Update Rule Implementation sheet
            if 'Rule Implementation' in wb.sheetnames:
                ws = wb['Rule Implementation']
                if 'comprehensive_report' in validation_data:
                    cr = validation_data['comprehensive_report']
                    if 'independent_in_rankings' in cr:
                        ws['D5'] = 'Present' if cr['independent_in_rankings'] else 'Not Present'
                        ws['D6'] = 'Present' if cr['independent_in_rankings'] else 'Not Present'
                
                if 'database' in validation_data:
                    db = validation_data['database']
                    ws['D9'] = str(db.get('cch_ohio_barriers', ''))  # CCH Ohio
                    ws['D10'] = str(db.get('cch_nc_barriers', ''))   # CCH NC
            
            # Save updated workbook
            output_path = qc_filepath.replace('.xlsx', '_populated.xlsx')
            wb.save(output_path)
            print(f"\n✓ QC Workbook updated and saved to: {output_path}")
            return True
            
        except Exception as e:
            self.results['errors'].append(f"Error updating QC Workbook: {str(e)}")
            return False
    
    def run_validation(self, file_paths):
        """Run complete validation suite"""
        print("="*60)
        print("QC VALIDATION AUTO-POPULATOR V16.0")
        print("="*60)
        print(f"\nStarting validation at {self.results['timestamp']}\n")
        
        # Validate each file type
        if 'database' in file_paths:
            self.validate_database(file_paths['database'])
        
        if 'scenario_1' in file_paths:
            self.validate_scenario(file_paths['scenario_1'], 1)
        
        if 'scenario_2' in file_paths:
            self.validate_scenario(file_paths['scenario_2'], 2)
            
        if 'scenario_3' in file_paths:
            self.validate_scenario(file_paths['scenario_3'], 3)
        
        if 'comprehensive_report' in file_paths:
            self.validate_comprehensive_report(file_paths['comprehensive_report'])
        
        if 'fee_schedule' in file_paths:
            self.validate_fee_schedule(file_paths['fee_schedule'])
        
        # Update QC Workbook if specified
        if 'qc_workbook' in file_paths and self.results['validation_results']:
            self.update_qc_workbook(file_paths['qc_workbook'], 
                                   self.results['validation_results'])
        
        # Print summary
        print("\n" + "="*60)
        print("VALIDATION SUMMARY")
        print("="*60)
        
        print(f"\nFiles Checked: {len(self.results['files_checked'])}")
        for file in self.results['files_checked']:
            print(f"  {file}")
        
        if self.results['errors']:
            print(f"\nErrors Found: {len(self.results['errors'])}")
            for error in self.results['errors']:
                print(f"  {error}")
        else:
            print("\n✓ All validations completed successfully!")
        
        # Save results to JSON
        with open('qc_validation_results.json', 'w') as f:
            json.dump(self.results, f, indent=2, default=str)
        print(f"\nDetailed results saved to: qc_validation_results.json")
        
        return self.results

# Example usage
if __name__ == "__main__":
    # Define file paths - adjust these to match your actual file locations
    file_paths = {
        'database': 'Combined_Database_FINAL_V15.xlsx',
        'scenario_1': 'Economic_Model_Scenario_1_Combined_V15.xlsx',
        'scenario_2': 'Economic_Model_Scenario_2_Combined_V15.xlsx',
        'scenario_3': 'Economic_Model_Scenario_3_Combined_V15.xlsx',
        'comprehensive_report': 'Comprehensive_Report_Workbook_V15.xlsx',
        'fee_schedule': 'Fee_Schedule_Reference_V15.xlsx',
        'qc_workbook': 'QC_Validation_Workbook_V16.xlsx'
    }
    
    # Run validation
    validator = QCValidator()
    results = validator.run_validation(file_paths)
    
    print("\n✓ Validation complete!")
